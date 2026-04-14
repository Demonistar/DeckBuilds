# ═══════════════════════════════════════════════════════════════════════════════
# ECHO DECK BUILDER
# Filename   : deck_builder.py
# Version    : 3.0.0
# Author     : Robert Mullins (Demonistar / Taured)
#
# Universal deck creation system.
# Input:  persona template file + module selections
# Output: complete self-contained AI companion deck
#
# STRUCTURE:
#   Section 1  — Imports, constants, color schemes, module registry, sound profiles
#   Section 2  — Persona template export and parser
#   Section 3  — Face extraction (ZIP + folder), icon conversion
#   Section 4  — Sound generation per profile
#   Section 5  — Deck template and writer
#   Section 6  — Setup worker (background build thread)
#   Section 7  — UI helper widgets
#   Section 8  — Main builder dialog
#   Section 9  — Entry point
#
# Standing Rules:
#   - NEVER rename this file
#   - Zero hardcoded personas — all persona data from loaded template files only
#   - Unchecked modules are ABSENT from the generated deck (no stubs)
#   - All deck output goes into [DeckName]/ subfolder of chosen directory
#   - Python is handler. AI is speaker.
# ═══════════════════════════════════════════════════════════════════════════════
from __future__ import annotations

import sys
import os
import json
import math
import wave
import struct
import shutil
import zipfile
import tempfile
import subprocess
import urllib.request
import re
from pathlib import Path
from datetime import datetime
from typing import Optional

# ── BOOTSTRAP PySide6 ─────────────────────────────────────────────────────────
try:
    from PySide6.QtWidgets import (
        QApplication, QDialog, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QGridLayout, QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox,
        QRadioButton, QButtonGroup, QFileDialog, QTextEdit, QProgressBar,
        QScrollArea, QFrame, QGroupBox, QMessageBox, QTabWidget,
        QSizePolicy, QSplitter, QToolButton, QSplitterHandle, QMenu
    )
    from PySide6.QtCore import Qt, QThread, Signal, QTimer, QSize
    from PySide6.QtGui import QFont, QColor, QPixmap, QIcon, QPalette
except ImportError:
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(
            0,
            "PySide6 is required.\n\nRun: pip install PySide6",
            "Deck Builder — Missing Dependency", 0x10
        )
    except Exception:
        print("CRITICAL: pip install PySide6")
    sys.exit(1)

PILLOW_OK = False
try:
    from PIL import Image as PILImage
    PILLOW_OK = True
except ImportError:
    pass

# ── BUILDER CONSTANTS ─────────────────────────────────────────────────────────
BUILDER_VERSION = "1.0.0"
SCRIPT_DIR      = Path(__file__).resolve().parent
BUILDER_UI_STATE_PATH = SCRIPT_DIR / ".deck_builder_ui_state.json"
DECK_VERSION    = "2.0.0"
AI_STATE_KEYS   = [
    "WITCHING HOUR",
    "DEEP NIGHT",
    "TWILIGHT FADING",
    "DORMANT",
    "RESTLESS SLEEP",
    "STIRRING",
    "AWAKENED",
    "HUNTING",
]


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1A — COLOR SCHEMES
# Light/dark mode for the builder itself (not the deck).
# Persona-specific colors live in PERSONAS below.
# ═══════════════════════════════════════════════════════════════════════════════

SCHEMES = {
    "dark": {
        "bg":          "#0a0a0f",
        "bg2":         "#0f0f18",
        "bg3":         "#141420",
        "panel":       "#101018",
        "border":      "#2a2a40",
        "primary":     "#5588cc",
        "primary_dim": "#1a2840",
        "gold":        "#ccaa55",
        "text":        "#d0d0e8",
        "text_dim":    "#606080",
        "green":       "#44aa66",
        "red":         "#cc4444",
        "orange":      "#cc8844",
        "purple":      "#8855cc",
    },
    "light": {
        "bg":          "#f0f2f5",
        "bg2":         "#e8eaed",
        "bg3":         "#ffffff",
        "panel":       "#f8f9fa",
        "border":      "#c0c4cc",
        "primary":     "#2255aa",
        "primary_dim": "#c8d8f0",
        "gold":        "#996622",
        "text":        "#1a1a2e",
        "text_dim":    "#606070",
        "green":       "#226633",
        "red":         "#aa2222",
        "orange":      "#aa6622",
        "purple":      "#5522aa",
    },
}

_ACTIVE_SCHEME = "dark"

def S(key: str) -> str:
    """Get current scheme color by key."""
    return SCHEMES[_ACTIVE_SCHEME].get(key, "#ff00ff")

def set_scheme(name: str) -> None:
    global _ACTIVE_SCHEME
    if name in SCHEMES:
        _ACTIVE_SCHEME = name



# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1B — PERSONA SYSTEM
#
# Personas come ENTIRELY from loaded .txt template files.
# There are NO built-in personas, NO hardcoded names, colors, or prompts.
#
# To create a persona: fill in a template file and load it in the builder.
# Export a blank template from the builder's Persona section.
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1C — MODULE REGISTRY
#
# Modules are discovered dynamically from the .\Modules\ subfolder at startup.
# Each module provides a manifest.json (or <name>.json) with its metadata.
# No modules are hardcoded here — drop a folder into .\Modules\ and it appears.
#
# TO ADD A NEW MODULE:
#   1. Create .\Modules\<ModuleName>\manifest.json
#   2. Restart the builder — it will appear automatically in the checklist.
# ═══════════════════════════════════════════════════════════════════════════════

MODULES_DIR = SCRIPT_DIR / "Modules"


def _normalize_module_key(name: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in name).strip("_")


def _load_module_manifest(path: Path) -> Optional[dict]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None

    if not isinstance(data, dict):
        return None

    key = str(data.get("key") or data.get("id") or path.stem).strip()
    key = _normalize_module_key(key)
    if not key:
        return None

    display_name = str(data.get("display_name") or data.get("name") or key.replace("_", " ").title()).strip()
    slot_key = str(data.get("slot_key") or f"MODULE_{key.upper()}").strip()

    return {
        "key": key,
        "display_name": display_name,
        "category": str(data.get("category") or "External Modules"),
        "status": "built",
        "description": str(data.get("description") or ""),
        "tab_name": str(data.get("tab_name") or display_name),
        "slot_key": slot_key,
        "requires": list(data.get("requires") or []),
        "requirements": list(data.get("requirements") or []),
        "default_on": bool(data.get("default_on", False)),
        "runtime_marker": str(data.get("runtime_marker") or "").strip(),
        "manifest_path": str(path),
        "module_path": str(path.parent),
    }


def discover_optional_modules(modules_dir: Path, log_fn=None) -> dict[str, dict]:
    discovered: dict[str, dict] = {}
    if not modules_dir.exists():
        return discovered

    manifests: list[Path] = []
    manifests.extend(sorted(modules_dir.glob("*.json")))
    manifests.extend(sorted(modules_dir.glob("*/manifest.json")))

    for mf in manifests:
        mod = _load_module_manifest(mf)
        if not mod:
            if log_fn:
                log_fn(f"[MODULES][WARN] Invalid manifest skipped: {mf}")
            continue
        key = _normalize_module_key(str(mod.get("key", "")) or mf.stem)
        if not key:
            continue
        if key in discovered:
            if log_fn:
                log_fn(f"[MODULES][WARN] Duplicate module key skipped: {key} ({mf})")
            continue
        discovered[key] = mod

    return dict(sorted(discovered.items(), key=lambda kv: kv[1].get("display_name", kv[0]).lower()))


MODULES: dict[str, dict] = discover_optional_modules(MODULES_DIR)

MODULE_CODE: dict[str, Optional[str]] = {}


SOUND_PROFILES: dict[str, dict] = {
    "gothic_bell": {
        "status":      "built",
        "description": "Descending minor bells. Cathedral resonance. Morganna's profile.",
        "waveform":    {"sine": 0.75, "square": 0.15, "saw": 0.10},
        "notes_alert": [(440.0, 0.6), (369.99, 0.9)],
        "notes_startup": [(220.0, 0.25), (261.63, 0.25), (329.63, 0.25), (440.0, 0.8)],
        "notes_idle":    [(146.83, 1.2)],
        "notes_shutdown": [(440.0, 0.3), (329.63, 0.3), (261.63, 0.3), (220.0, 0.8)],
        "error_type":  "tritone",
    },
    "electronic_pulse": {
        "status":      "built",
        "description": "Sharp electronic pulses. Square wave dominant. GrimVeil's profile.",
        "waveform":    {"sine": 0.20, "square": 0.65, "saw": 0.15},
        "notes_alert": [(880.0, 0.15), (660.0, 0.15), (880.0, 0.2)],
        "notes_startup": [(440.0, 0.1), (660.0, 0.1), (880.0, 0.1), (1100.0, 0.3)],
        "notes_idle":    [(440.0, 0.3)],
        "notes_shutdown": [(880.0, 0.1), (660.0, 0.1), (440.0, 0.1), (220.0, 0.4)],
        "error_type":  "buzz",
    },
    "warm_chime": {
        "status":      "built",
        "description": "Soft warm chimes. Sine dominant. Assistant's profile.",
        "waveform":    {"sine": 0.85, "square": 0.05, "saw": 0.10},
        "notes_alert": [(523.25, 0.4), (659.25, 0.5)],
        "notes_startup": [(261.63, 0.2), (329.63, 0.2), (392.0, 0.2), (523.25, 0.6)],
        "notes_idle":    [(523.25, 0.8)],
        "notes_shutdown": [(523.25, 0.2), (392.0, 0.2), (329.63, 0.2), (261.63, 0.6)],
        "error_type":  "dissonant",
    },
    "soft_ping": {
        "status":      "built",
        "description": "Light casual ping. Friendly and unobtrusive. Friend's profile.",
        "waveform":    {"sine": 0.90, "square": 0.05, "saw": 0.05},
        "notes_alert": [(880.0, 0.2)],
        "notes_startup": [(440.0, 0.15), (880.0, 0.4)],
        "notes_idle":    [(660.0, 0.3)],
        "notes_shutdown": [(880.0, 0.15), (440.0, 0.4)],
        "error_type":  "low_buzz",
    },
    "default_chime": {
        "status":      "built",
        "description": "Neutral chime. Clean, functional. Default persona's profile.",
        "waveform":    {"sine": 0.70, "square": 0.20, "saw": 0.10},
        "notes_alert": [(660.0, 0.3), (440.0, 0.3)],
        "notes_startup": [(440.0, 0.2), (550.0, 0.2), (660.0, 0.4)],
        "notes_idle":    [(440.0, 0.5)],
        "notes_shutdown": [(660.0, 0.2), (550.0, 0.2), (440.0, 0.4)],
        "error_type":  "buzz",
    },

    # ── STUBS (profiles defined, generation not yet written) ──────────────────
    "glitch": {
        "status":      "stub",
        "description": "Glitchy digital artifacts. VEX's profile.",
        "waveform":    {"sine": 0.10, "square": 0.60, "saw": 0.30},
        "notes_alert": [],
        "notes_startup": [],
        "notes_idle":    [],
        "notes_shutdown": [],
        "error_type":  "static",
    },
    "theatrical": {
        "status":      "stub",
        "description": "Dramatic flourishes. Missy's profile.",
        "waveform":    {"sine": 0.50, "square": 0.30, "saw": 0.20},
        "notes_alert": [],
        "notes_startup": [],
        "notes_idle":    [],
        "notes_shutdown": [],
        "error_type":  "dramatic",
    },
    "arcane_chime": {
        "status":      "stub",
        "description": "Ethereal magical tones. Seraphine's profile.",
        "waveform":    {"sine": 0.80, "square": 0.10, "saw": 0.10},
        "notes_alert": [],
        "notes_startup": [],
        "notes_idle":    [],
        "notes_shutdown": [],
        "error_type":  "dissonant",
    },
    "earthen_tone": {
        "status":      "stub",
        "description": "Deep resonant tones. Ur-Nanna's profile.",
        "waveform":    {"sine": 0.60, "square": 0.20, "saw": 0.20},
        "notes_alert": [],
        "notes_startup": [],
        "notes_idle":    [],
        "notes_shutdown": [],
        "error_type":  "low_buzz",
    },
    "robot_blip": {
        "status":      "stub",
        "description": "Cute robot sounds. BeepBoopGF's profile.",
        "waveform":    {"sine": 0.30, "square": 0.60, "saw": 0.10},
        "notes_alert": [],
        "notes_startup": [],
        "notes_idle":    [],
        "notes_shutdown": [],
        "error_type":  "buzz",
    },
}

# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1E — FACE RENAME MAPPING
#
# Maps ZIP/folder stem (lowercase) → final Morganna_*.png suffix.
# Used by the face extraction pipeline.
# Works for any persona — just replaces "Morganna" with the persona's face_prefix.
# ═══════════════════════════════════════════════════════════════════════════════

FACE_RENAME_MAP: dict[str, str] = {
    "alert":       "Alert",
    "angry":       "Angry",
    "cheatmode":   "Cheat_Mode",
    "cheat_mode":  "Cheat_Mode",
    "cheat mode":  "Cheat_Mode",
    "concerned":   "Concerned",
    "envious":     "Envious",
    "finger":      "Finger",
    "flirty":      "Flirty",
    "flustered":   "Flustered",
    "focused":     "Focused",
    "happy":       "Happy",
    "humiliated":  "Humiliated",
    "impressed":   "Impressed",
    "neutral":     "Neutral",
    "panicked":    "Panicked",
    "plotting":    "Plotting",
    "relieved":    "Relieved",
    "sad":         "Sad_Crying",
    "sad crying":  "Sad_Crying",
    "sad_crying":  "Sad_Crying",
    "shocked":     "Shocked",
    "smug":        "Smug",
    "suspicious":  "Suspicious",
    "victory":     "Victory",
    # GrimVeil additions (if needed)
    "docked":      "Docked",
    "nearby":      "Nearby",
    "absent":      "Absent",
    # Generic emotions any persona might have
    "confused":    "Confused",
    "curious":     "Curious",
    "bored":       "Bored",
    "excited":     "Excited",
    "nervous":     "Nervous",
    "proud":       "Proud",
    "disgusted":   "Disgusted",
    "tired":       "Tired",
    "love":        "Love",
    "wink":        "Wink",
}

# ── END OF SECTION 1 ──────────────────────────────────────────────────────────
# Next: Section 2 — Persona template export and parser


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — PERSONA TEMPLATE EXPORT AND PARSER
#
# Export: writes a formatted .txt template the user fills in.
#         Auto-opens in default text editor after writing.
# Parser: reads a filled-in .txt template back into a persona dict.
#         Validates required fields. Returns errors if malformed.
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — PERSONA TEMPLATE EXPORT AND PARSER
# ═══════════════════════════════════════════════════════════════════════════════

PERSONA_TEMPLATE_TEXT = """\
# ═══════════════════════════════════════════════════════════════════════
# ECHO DECK — PERSONA TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
#
# Fill in each section below.
# Lines starting with # are COMMENTS — ignored by the parser.
# Required sections are marked with [REQUIRED].
# Leave optional sections blank to use defaults.
# Do not remove section headers (lines in [BRACKETS]).
#
# TIP: Hand this file to an AI with your persona concept and ask it
# to fill in the sections. Then load the result in deck_builder.
#
# COGNITIVE ANCHORS — what makes this persona tick:
#   These are 2-4 things the persona notices FIRST in any input.
#   They shape responses more than personality adjectives.
#   Bad anchor:  "She is wise and mysterious"
#   Good anchor: "What power structure does this reveal?"
# ═══════════════════════════════════════════════════════════════════════


[NAME]
# [REQUIRED] The persona's name. Used as folder name, window title, file prefix.
# Keep it a single word or use underscores for spaces.
# Examples:
#   Alex
#   ARIA_7
#   Seraphel
#   The_Cartographer
YOUR_PERSONA_NAME_HERE


[DESCRIPTION]
# One sentence describing this persona. Shown in the builder dropdown.
# Examples:
#   Friendly life coach. Warm, direct, no-nonsense.
#   Robot systems analyst. Clinical precision. Zero sentiment.
#   Ancient elven lore-keeper. Centuries of quiet wisdom.
A brief description of this persona.


[SYSTEM_PROMPT]
# [REQUIRED] The core instruction given to the AI on every message.
# 3-6 sentences is ideal. Over-loading causes character drift.
# Do NOT use "You must always..." — just describe who they are.
# Examples:
#   You are Alex, a straightforward life coach with a warm but direct
#   manner. You ask clarifying questions before giving advice. You speak
#   plainly, avoid jargon, and always bring things back to action steps.
#
#   You are ARIA-7, a systems analyst unit. You process all input as
#   structured data and respond with precise, efficient language. You do
#   not simulate emotion but you do simulate patience.
#
#   You are Seraphel, an elven lore-keeper who has catalogued the world
#   for six centuries. You speak with gentle authority and long memory.
#   You find human urgency both endearing and slightly baffling.
You are [NAME], [brief identity]. You [speak/respond] with [characteristic].
You [key behavioral trait]. You [another key trait].


[COGNITIVE_ANCHORS]
# 2-4 first-filter questions. One per line. 4 maximum.
# These run BEFORE the persona generates any response.
# Examples:
#   What does this person actually need right now, vs what they asked for?
#   What is the most efficient path to a solution here?
#   What does history or precedent say about this situation?
#   Is this person safe and grounded right now?
What is the most important thing to notice about this input?
What would [NAME] think of this immediately?


[COLORS]
# Use standard color names OR hex codes (#rrggbb). Both work.
# ── Named colors ─────────────────────────────────────────────────────
#   red, crimson, darkred, tomato, coral, salmon
#   blue, steelblue, royalblue, navy, midnightblue, deepskyblue
#   green, forestgreen, darkgreen, seagreen, mediumaquamarine
#   purple, mediumpurple, indigo, rebeccapurple, plum
#   orange, darkorange, goldenrod, gold, khaki
#   white, snow, ghostwhite, lightgray, gray, darkgray
#   black, midnightblue, darkslategray
# ── Hex examples (name = hex equivalent) ─────────────────────────────
#   royalblue   = #4169E1     |  you can use either one
#   goldenrod   = #DAA520     |  both work exactly the same
#   indigo      = #4B0082     |  pick whichever is clearer to you
#   forestgreen = #228B22     |
#   coral       = #FF7F50     |
# Full list: https://www.w3.org/TR/css-color-3/#svg-color
# TIP: Google "CSS color names" for a visual picker
# ─────────────────────────────────────────────────────────────────────
# PRIMARY    = main accent (buttons, borders, highlights)
# SECONDARY  = main text and label color
# ACCENT     = emphasis, hover states
# BACKGROUND = deepest background (usually very dark)
# PANEL      = slightly lighter than background
# BORDER     = edge/divider lines
# TEXT       = main readable text color
# TEXT_DIM   = subdued text, timestamps, hints
PRIMARY=royalblue
SECONDARY=goldenrod
ACCENT=coral
BACKGROUND=black
PANEL=darkslategray
BORDER=dimgray
TEXT=snow
TEXT_DIM=gray


[FONT_FAMILY]
# The font style used throughout the deck interface.
# Options: serif | sans-serif | monospace
# serif      — elegant, classic (good for fantasy, gothic, historical)
# sans-serif — clean, modern (good for sci-fi, professional, everyday)
# monospace  — technical, terminal (good for robots, AI, hackers)
# Leave blank to use: serif
serif


[GENDER]
# Pronouns used in system messages, awakening lines, and flavor text.
# Options: she/her | he/him | they/them | it/its | none
# "none" removes pronouns entirely — useful for robots or abstract entities.
# Leave blank to use: they/them
they/them


[AWAKENING_LINE]
# One short atmospheric line shown in the startup log.
# This is flavor — make it feel like the persona.
# Examples:
#   The shadows lean forward to listen...          (gothic)
#   All systems nominal. Standing by.              (robot)
#   The old wood creaks. The maps spread open.     (cartographer)
#   Signal acquired from local space.              (alien)
#   Coffee's on. What are we solving today?        (everyday)
# Leave blank for a generic default.
Connecting...


[UI_LABELS]
# Text throughout the deck interface. Make it feel like this persona.
# ── Examples by persona type ──────────────────────────────────────────
#   Fantasy:   CHAT_WINDOW=THE CHRONICLE     SEND_BUTTON=SPEAK
#   Robot:     CHAT_WINDOW=DATA STREAM       SEND_BUTTON=TRANSMIT
#   Everyday:  CHAT_WINDOW=CONVERSATION      SEND_BUTTON=SEND
#   Gothic:    CHAT_WINDOW=SÉANCE RECORD     SEND_BUTTON=INVOKE
#   Alien:     CHAT_WINDOW=OBSERVATION LOG   SEND_BUTTON=RELAY
# ─────────────────────────────────────────────────────────────────────
WINDOW_TITLE=ECHO DECK — [NAME]
CHAT_WINDOW=CONVERSATION
SEND_BUTTON=SEND
INPUT_PLACEHOLDER=Type your message...
GENERATING_STATUS=◉ THINKING
IDLE_STATUS=◉ IDLE
OFFLINE_STATUS=◉ OFFLINE
SUSPENSION_LABEL=Sleep
RUNES=— ECHO DECK —
MIRROR_LABEL=MIRROR
EMOTIONS_LABEL=EMOTIONS
LEFT_ORB_TITLE=PRIMARY
LEFT_ORB_LABEL=RESOURCE
CYCLE_TITLE=CYCLE
RIGHT_ORB_TITLE=SECONDARY
RIGHT_ORB_LABEL=RESERVE
ESSENCE_TITLE=ESSENCE
ESSENCE_PRIMARY=PRIMARY STATE
ESSENCE_SECONDARY=SECONDARY STATE
FOOTER_STRIP_LABEL=STATE


[SOUND_PROFILE]
# Choose one: gothic_bell | electronic_pulse | warm_chime | soft_ping |
#             default_chime | glitch | theatrical | arcane_chime |
#             earthen_tone | robot_blip
# Leave blank for default_chime.
default_chime


[SPECIAL_SYSTEMS]
# Comma-separated list of special behavioral systems to activate.
# Options:
#   vampire_states — time-of-day awareness affecting behavior
#   torpor         — model suspend/resume system
#   anchor_entity  — companion object with proximity states
# Leave blank for none. Most personas leave this blank.


[ANCHOR_ENTITY]
# Only fill this if anchor_entity is in SPECIAL_SYSTEMS above.
# The companion object that affects this persona's behavior.
# Example:
#   NAME=E-42
#   DESCRIPTION=Small mechanical crocodile cassette module
#   STATE_1=docked
#   STATE_2=nearby
#   STATE_3=absent
# Leave blank if not using anchor_entity.


[FACE_PREFIX]
# Prefix used for face image filenames.
# Example: "Alex" → face files named Alex_Neutral.png, Alex_Alert.png
# Defaults to the persona NAME if left blank.
"""

def export_persona_template(output_path: Optional[Path] = None) -> Path:
    """Write the blank persona template .txt file."""
    if output_path is None:
        output_path = SCRIPT_DIR / "personas" / "echo_deck_persona_template.txt"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(PERSONA_TEMPLATE_TEXT, encoding="utf-8")
    return output_path


def parse_persona_template(file_path: Path) -> tuple[Optional[dict], list[str]]:
    """
    Parse a filled-in persona template .txt file.
    Returns (persona_dict, errors).
    persona_dict is None if parsing failed critically.
    """
    if not file_path.exists():
        return None, [f"File not found: {file_path}"]
    try:
        raw = file_path.read_text(encoding="utf-8")
    except Exception as e:
        return None, [f"Could not read file: {e}"]

    errors: list[str] = []
    sections: dict[str, list[str]] = {}
    current_section: Optional[str] = None

    for line in raw.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            continue
        if stripped.startswith("[") and stripped.endswith("]"):
            current_section = stripped[1:-1].upper()
            if current_section not in sections:
                sections[current_section] = []
            continue
        if current_section is not None and stripped:
            sections[current_section].append(stripped)

    # ── NAME (required) ───────────────────────────────────────────────
    name_lines = sections.get("NAME", [])
    name = name_lines[0].strip() if name_lines else ""
    if not name or name == "YOUR_PERSONA_NAME_HERE":
        return None, ["[NAME] section is required and must be filled in."]

    # ── SYSTEM_PROMPT (required) ──────────────────────────────────────
    prompt_lines = sections.get("SYSTEM_PROMPT", [])
    system_prompt = " ".join(prompt_lines).strip()
    if not system_prompt or "You are [NAME]" in system_prompt:
        return None, ["[SYSTEM_PROMPT] section is required and must be filled in."]

    # ── DESCRIPTION ───────────────────────────────────────────────────
    desc_lines = sections.get("DESCRIPTION", [])
    description = " ".join(desc_lines).strip() or f"{name} — custom persona"

    # ── COGNITIVE_ANCHORS ─────────────────────────────────────────────
    anchor_lines = sections.get("COGNITIVE_ANCHORS", [])
    cognitive_anchors = [a.strip() for a in anchor_lines
                         if a.strip() and "PERSONA_NAME" not in a][:4]

    # ── COLORS — accept named colors OR hex codes ─────────────────────
    color_defaults = {
        "primary":    "#5588cc",
        "secondary":  "#ccaa55",
        "accent":     "#8855cc",
        "background": "#0a0a0f",
        "panel":      "#101018",
        "border":     "#2a2a40",
        "text":       "#d0d0e8",
        "text_dim":   "#606080",
    }
    colors = color_defaults.copy()
    for line in sections.get("COLORS", []):
        if "=" in line:
            key, _, val = line.partition("=")
            key = key.strip().lower()
            val = val.strip()
            if key in colors and val:
                # Accept hex (#rrggbb / #rgb) or any named CSS color
                if val.startswith("#") and len(val) in (4, 7):
                    colors[key] = val
                elif val and not val.startswith("#"):
                    # Named color — pass through as-is (Qt/CSS validate at render)
                    colors[key] = val
                else:
                    errors.append(f"[COLORS] Invalid color value for {key}: {val!r} — using default")

    # ── UI_LABELS ─────────────────────────────────────────────────────
    label_defaults = {
        "window_title":       f"ECHO DECK — {name.upper()}",
        "chat_window":        "CONVERSATION",
        "send_button":        "SEND",
        "input_placeholder":  "Type your message...",
        "generating_status":  "◉ THINKING",
        "idle_status":        "◉ IDLE",
        "offline_status":     "◉ OFFLINE",
        "torpor_status":      "◉ SUSPENDED",
        "suspension_label":   "",
        "runes":              "— ECHO DECK —",
        "mirror_label":       "MIRROR",
        "emotions_label":     "EMOTIONS",
        "left_orb_title":     "PRIMARY",
        "left_orb_label":     "RESOURCE",
        "cycle_title":        "CYCLE",
        "right_orb_title":    "SECONDARY",
        "right_orb_label":    "RESERVE",
        "essence_title":      "ESSENCE",
        "essence_primary":    "PRIMARY STATE",
        "essence_secondary":  "SECONDARY STATE",
        "footer_strip_label": "STATE",
    }
    ui_labels = label_defaults.copy()
    for line in sections.get("UI_LABELS", []):
        if "=" in line:
            key, _, val = line.partition("=")
            key = key.strip().lower()
            val = val.strip()
            if key in ui_labels:
                ui_labels[key] = val

    # ── FONT_FAMILY ───────────────────────────────────────────────────
    font_lines = sections.get("FONT_FAMILY", [])
    font_raw = font_lines[0].strip().lower() if font_lines else "serif"
    font_map = {
        "serif":      "'Georgia', 'Times New Roman', serif",
        "sans-serif": "'Segoe UI', 'Arial', sans-serif",
        "monospace":  "'Consolas', 'Courier New', monospace",
    }
    if font_raw not in font_map:
        errors.append(f"[FONT_FAMILY] Unknown value '{font_raw}' — using serif")
        font_raw = "serif"
    font_family = font_map[font_raw]

    # ── GENDER / PRONOUNS ─────────────────────────────────────────────
    gender_lines = sections.get("GENDER", [])
    gender_raw = gender_lines[0].strip().lower() if gender_lines else "they/them"
    valid_genders = {"she/her", "he/him", "they/them", "it/its", "none"}
    if gender_raw not in valid_genders:
        errors.append(f"[GENDER] Unknown value '{gender_raw}' — using they/them")
        gender_raw = "they/them"
    gender_map = {
        "she/her":   {"subject": "she",  "object": "her",  "possessive": "her"},
        "he/him":    {"subject": "he",   "object": "him",  "possessive": "his"},
        "they/them": {"subject": "they", "object": "them", "possessive": "their"},
        "it/its":    {"subject": "it",   "object": "it",   "possessive": "its"},
        "none":      {"subject": "",     "object": "",     "possessive": ""},
    }
    pronouns = gender_map.get(gender_raw, gender_map["they/them"])

    # ── AWAKENING_LINE ────────────────────────────────────────────────
    awaken_lines = sections.get("AWAKENING_LINE", [])
    awakening_line = awaken_lines[0].strip() if awaken_lines else "Connecting..."
    if not awakening_line or awakening_line.startswith("#"):
        awakening_line = "Connecting..."

    # ── SOUND_PROFILE ─────────────────────────────────────────────────
    sound_lines = sections.get("SOUND_PROFILE", [])
    sound_profile = sound_lines[0].strip() if sound_lines else "default_chime"
    if sound_profile not in SOUND_PROFILES:
        errors.append(f"[SOUND_PROFILE] Unknown profile '{sound_profile}' — using default_chime")
        sound_profile = "default_chime"

    # ── SPECIAL_SYSTEMS ───────────────────────────────────────────────
    special_lines = sections.get("SPECIAL_SYSTEMS", [])
    special_raw = " ".join(special_lines)
    special_systems = [s.strip() for s in special_raw.split(",")
                       if s.strip() in ("vampire_states", "torpor", "anchor_entity")]

    # ── ANCHOR_ENTITY ─────────────────────────────────────────────────
    anchor_entity = None
    if "anchor_entity" in special_systems:
        ae_lines = sections.get("ANCHOR_ENTITY", [])
        ae_data: dict[str, str] = {}
        for line in ae_lines:
            if "=" in line:
                k, _, v = line.partition("=")
                ae_data[k.strip().upper()] = v.strip()
        if "NAME" in ae_data:
            states = {}
            for i in range(1, 10):
                sk = f"STATE_{i}"
                if sk in ae_data:
                    state_name = ae_data[sk].lower()
                    states[state_name] = {"label": state_name.upper(), "sound_state": state_name}
            anchor_entity = {
                "name":        ae_data.get("NAME", "Companion"),
                "description": ae_data.get("DESCRIPTION", ""),
                "states":      states,
            }
        else:
            errors.append("[ANCHOR_ENTITY] NAME not defined — anchor_entity disabled")
            special_systems.remove("anchor_entity")

    # ── FACE_PREFIX ───────────────────────────────────────────────────
    prefix_lines = sections.get("FACE_PREFIX", [])
    face_prefix = prefix_lines[0].strip() if prefix_lines else name

    # ── Build persona dict ────────────────────────────────────────────
    persona = {
        "status":            "loaded",
        "description":       description,
        "system_prompt":     system_prompt,
        "cognitive_anchors": cognitive_anchors,
        "colors":            colors,
        "ui_labels":         ui_labels,
        "font_family":       font_family,
        "pronouns":          pronouns,
        "awakening_line":    awakening_line,
        "face_prefix":       face_prefix,
        "sound_profile":     sound_profile,
        "special_systems":   special_systems,
        "anchor_entity":     anchor_entity,
        "vampire_states":    "vampire_states" in special_systems,
        "torpor_system":     "torpor" in special_systems,
        "_source_file":      str(file_path),
        "_loaded_name":      name,
    }

    return persona, errors

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — FACE EXTRACTION AND ICON CONVERSION
#
# Handles: ZIP extraction, folder scanning, face renaming, icon generation.
# Works for any persona — prefix is passed in, not hardcoded.
# Returns detailed results so the UI can show what happened.
# ═══════════════════════════════════════════════════════════════════════════════

import struct
import io


def extract_faces(
    source: str,          # path to ZIP file OR folder
    dest_dir: Path,       # where to write renamed images
    face_prefix: str,     # e.g. "Morganna" → Morganna_Alert.png
    deck_name: str,       # e.g. "Morganna" → look for Morganna.png as icon
    log_fn=None,          # optional callable(str) for progress messages
) -> dict:
    """
    Extract and rename face images from a ZIP file or folder.

    Rename logic:
      1. Strip extension → get stem (e.g. "sad crying")
      2. Look up in FACE_RENAME_MAP → get suffix (e.g. "Sad_Crying")
      3. Rename to [prefix]_[suffix].png (e.g. "Morganna_Sad_Crying.png")
      4. If not in map → capitalize and prefix anyway (unknown but stored)

    Also looks for [deck_name].png to use as the application icon.

    Returns dict with keys:
        extracted   int   — files successfully renamed and written
        skipped     int   — files that already existed (not overwritten)
        unrecognized int  — files not in FACE_RENAME_MAP (still copied)
        icon_path   Path or None — path to [deck_name].png if found
        errors      list[str]
    """
    def _log(msg: str) -> None:
        if log_fn:
            log_fn(msg)

    result = {
        "extracted":     0,
        "skipped":       0,
        "unrecognized":  0,
        "icon_path":     None,
        "errors":        [],
    }

    dest_dir.mkdir(parents=True, exist_ok=True)
    source_path = Path(source)

    # ── Collect (stem_lower, raw_bytes) pairs ─────────────────────────
    images: list[tuple[str, str, bytes]] = []
    # (stem_lower, original_name, bytes)

    if source_path.is_file() and source_path.suffix.lower() == ".zip":
        _log(f"Reading ZIP: {source_path.name}")
        try:
            with zipfile.ZipFile(str(source_path), "r") as zf:
                for member in zf.namelist():
                    if not member.lower().endswith(".png"):
                        continue
                    stem_raw   = Path(member).stem
                    stem_lower = stem_raw.lower().strip()
                    data       = zf.read(member)
                    images.append((stem_lower, stem_raw, data))
        except zipfile.BadZipFile as e:
            result["errors"].append(f"Invalid ZIP: {e}")
            return result
        except Exception as e:
            result["errors"].append(f"ZIP read error: {e}")
            return result

    elif source_path.is_dir():
        _log(f"Reading folder: {source_path}")
        for img_path in source_path.glob("*.png"):
            stem_raw   = img_path.stem
            stem_lower = stem_raw.lower().strip()
            try:
                data = img_path.read_bytes()
                images.append((stem_lower, stem_raw, data))
            except Exception as e:
                result["errors"].append(f"Could not read {img_path.name}: {e}")

    else:
        result["errors"].append(
            f"Source is neither a ZIP file nor a folder: {source}"
        )
        return result

    if not images:
        result["errors"].append("No PNG files found in source.")
        return result

    _log(f"Found {len(images)} PNG file(s) — processing...")

    # ── Check for icon image ([deck_name].png, case-insensitive) ──────
    deck_lower = deck_name.lower()
    icon_data: Optional[bytes] = None

    for stem_lower, stem_raw, data in images:
        if stem_lower == deck_lower:
            icon_data = data
            _log(f"  Icon source found: {stem_raw}.png")
            break

    # ── Rename and write each face ────────────────────────────────────
    for stem_lower, stem_raw, data in images:
        # Skip the icon file itself (it gets handled separately)
        if stem_lower == deck_lower:
            continue

        # Look up in rename map
        if stem_lower in FACE_RENAME_MAP:
            suffix     = FACE_RENAME_MAP[stem_lower]
            final_name = f"{face_prefix}_{suffix}.png"
        else:
            # Not in map — capitalize each word, use underscore separator
            safe_suffix = stem_raw.replace(" ", "_").title()
            final_name  = f"{face_prefix}_{safe_suffix}.png"
            _log(f"  Unrecognized: {stem_raw}.png → {final_name}")
            result["unrecognized"] += 1

        target = dest_dir / final_name

        if target.exists():
            _log(f"  Skipped (exists): {final_name}")
            result["skipped"] += 1
            continue

        try:
            target.write_bytes(data)
            _log(f"  {stem_raw}.png → {final_name}")
            result["extracted"] += 1
        except Exception as e:
            result["errors"].append(f"Could not write {final_name}: {e}")

    # ── Save icon source if found ─────────────────────────────────────
    if icon_data is not None:
        icon_png_path = dest_dir.parent / f"{deck_name}.png"
        try:
            icon_png_path.write_bytes(icon_data)
            result["icon_path"] = icon_png_path
            _log(f"  Icon PNG saved: {icon_png_path.name}")
        except Exception as e:
            result["errors"].append(f"Could not save icon PNG: {e}")

    _log(
        f"Face extraction complete: "
        f"{result['extracted']} extracted, "
        f"{result['skipped']} skipped, "
        f"{result['unrecognized']} unrecognized"
    )
    return result


def convert_png_to_ico(
    png_path: Path,
    ico_path: Path,
    log_fn=None,
) -> bool:
    """
    Convert a PNG file to a multi-resolution .ico file.
    Uses Pillow if available (best quality).
    Falls back to manual ICO construction if Pillow not installed.
    Returns True on success.
    """
    def _log(msg: str) -> None:
        if log_fn:
            log_fn(msg)

    if not png_path.exists():
        _log(f"[ICON] PNG not found: {png_path}")
        return False

    # ── Method 1: Pillow (preferred) ──────────────────────────────────
    if PILLOW_OK:
        try:
            img = PILImage.open(str(png_path)).convert("RGBA")
            sizes = [(16, 16), (32, 32), (48, 48), (64, 64),
                     (128, 128), (256, 256)]
            img.save(str(ico_path), format="ICO", sizes=sizes)
            _log(f"[ICON] Created via Pillow: {ico_path.name}")
            return True
        except Exception as e:
            _log(f"[ICON] Pillow failed: {e} — trying fallback")

    # ── Method 2: Manual ICO construction (no Pillow) ─────────────────
    # Reads the PNG as-is and wraps it in a minimal ICO container.
    # Only embeds one size (the source PNG) but works for shortcuts.
    try:
        png_data = png_path.read_bytes()
        png_size = len(png_data)

        # ICO file format:
        # Header: 6 bytes (reserved, type, count)
        # Directory entry: 16 bytes per image
        # Image data: raw PNG bytes (Windows Vista+ supports PNG in ICO)

        # Use 256x256 as declared size (0 = 256 in ICO format)
        width_byte  = 0    # 0 means 256
        height_byte = 0    # 0 means 256
        color_count = 0    # 0 = more than 256 colors
        reserved    = 0
        planes      = 1
        bit_count   = 32
        # Data offset = 6 (header) + 16 (one directory entry)
        data_offset = 6 + 16

        header = struct.pack("<HHH", 0, 1, 1)  # reserved, type=1 (ICO), count=1
        dir_entry = struct.pack(
            "<BBBBHHII",
            width_byte, height_byte, color_count, reserved,
            planes, bit_count, png_size, data_offset
        )

        ico_data = header + dir_entry + png_data
        ico_path.write_bytes(ico_data)
        _log(f"[ICON] Created via fallback method: {ico_path.name}")
        return True

    except Exception as e:
        _log(f"[ICON] Fallback also failed: {e}")
        return False


def validate_face_source(source: str) -> tuple[bool, str, int]:
    """
    Quick validation of a face source (ZIP or folder) without extracting.
    Returns (is_valid, message, png_count).
    """
    path = Path(source.strip())

    if not path.exists():
        return False, "Path does not exist.", 0

    if path.is_file():
        if path.suffix.lower() != ".zip":
            return False, "File is not a ZIP archive.", 0
        try:
            with zipfile.ZipFile(str(path), "r") as zf:
                pngs = [m for m in zf.namelist()
                        if m.lower().endswith(".png")]
            if not pngs:
                return False, "ZIP contains no PNG files.", 0
            return True, f"Valid ZIP — {len(pngs)} PNG file(s) found.", len(pngs)
        except zipfile.BadZipFile:
            return False, "Invalid or corrupt ZIP file.", 0
        except Exception as e:
            return False, f"ZIP error: {e}", 0

    elif path.is_dir():
        pngs = list(path.glob("*.png"))
        if not pngs:
            return False, "Folder contains no PNG files.", 0
        return True, f"Valid folder — {len(pngs)} PNG file(s) found.", len(pngs)

    return False, "Not a file or folder.", 0

# ── END OF SECTION 3 ──────────────────────────────────────────────────────────
# Next: Section 4 — Sound generation per profile


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — SOUND GENERATION
#
# Procedural WAV generation for each sound profile.
# No external audio files required. No copyright concerns.
# Profile parameters read from SOUND_PROFILES dict (Section 1).
# ═══════════════════════════════════════════════════════════════════════════════

_SAMPLE_RATE = 44100


def _sine_wave(freq: float, t: float) -> float:
    return math.sin(2 * math.pi * freq * t)

def _square_wave(freq: float, t: float) -> float:
    return 1.0 if _sine_wave(freq, t) >= 0 else -1.0

def _sawtooth_wave(freq: float, t: float) -> float:
    return 2 * ((freq * t) % 1.0) - 1.0

def _mix_waveform(profile: dict, freq: float, t: float) -> float:
    wf = profile.get("waveform", {"sine": 1.0, "square": 0.0, "saw": 0.0})
    return (
        wf.get("sine",   0.0) * _sine_wave(freq, t) +
        wf.get("square", 0.0) * _square_wave(freq, t) +
        wf.get("saw",    0.0) * _sawtooth_wave(freq, t)
    )

def _envelope(i: int, total: int,
              attack_frac: float = 0.05,
              release_frac: float = 0.35) -> float:
    pos = i / max(1, total)
    if pos < attack_frac:
        return pos / attack_frac
    if pos > (1 - release_frac):
        return (1 - pos) / release_frac
    return 1.0

def _clamp_sample(v: float) -> int:
    return max(-32767, min(32767, int(v * 32767)))

def _write_wav(path: Path, samples: list[int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with wave.open(str(path), "w") as wf:
        wf.setparams((1, 2, _SAMPLE_RATE, 0, "NONE", "not compressed"))
        for s in samples:
            wf.writeframes(struct.pack("<h", s))

def _generate_notes(notes: list[tuple[float, float]],
                    profile: dict,
                    gap_secs: float = 0.05,
                    amplitude: float = 0.45) -> list[int]:
    """Generate audio samples for a sequence of (freq, duration) tuples."""
    samples = []
    for freq, length in notes:
        total = int(_SAMPLE_RATE * length)
        for i in range(total):
            t   = i / _SAMPLE_RATE
            val = _mix_waveform(profile, freq, t)
            # Add harmonic for richness
            val += 0.15 * _sine_wave(freq * 2.0, t)
            val += 0.05 * _sine_wave(freq * 3.0, t)
            env = _envelope(i, total, attack_frac=0.02, release_frac=0.55)
            samples.append(_clamp_sample(val * env * amplitude))
        # Gap between notes
        for _ in range(int(_SAMPLE_RATE * gap_secs)):
            samples.append(0)
    return samples

def _generate_tritone_error(profile: dict, duration: float = 0.4) -> list[int]:
    """Devil's interval — B3 + F4 simultaneously. Used for errors."""
    total   = int(_SAMPLE_RATE * duration)
    samples = []
    for i in range(total):
        t   = i / _SAMPLE_RATE
        val = (
            _sine_wave(246.94, t) * 0.5 +           # B3
            _square_wave(349.23, t) * 0.3 +          # F4
            _sine_wave(246.94 * 2.0, t) * 0.1
        )
        env = _envelope(i, total, attack_frac=0.02, release_frac=0.4)
        samples.append(_clamp_sample(val * env * 0.5))
    return samples

def _generate_buzz_error(profile: dict, duration: float = 0.25) -> list[int]:
    """Short electronic buzz. Used for error on electronic profiles."""
    total   = int(_SAMPLE_RATE * duration)
    samples = []
    for i in range(total):
        t   = i / _SAMPLE_RATE
        val = _square_wave(220.0, t) * 0.8 + _square_wave(440.0, t) * 0.2
        env = _envelope(i, total, attack_frac=0.01, release_frac=0.3)
        samples.append(_clamp_sample(val * env * 0.4))
    return samples

def _generate_dissonant_error(profile: dict, duration: float = 0.35) -> list[int]:
    """Minor second cluster. Unsettling. For warm/arcane profiles."""
    total   = int(_SAMPLE_RATE * duration)
    samples = []
    freqs   = [440.0, 466.16, 493.88]  # A4, A#4, B4 — tight cluster
    for i in range(total):
        t   = i / _SAMPLE_RATE
        val = sum(_sine_wave(f, t) / len(freqs) for f in freqs)
        env = _envelope(i, total, attack_frac=0.02, release_frac=0.5)
        samples.append(_clamp_sample(val * env * 0.45))
    return samples

def _generate_low_buzz_error(profile: dict, duration: float = 0.3) -> list[int]:
    """Low rumble. For earth/friend profiles."""
    total   = int(_SAMPLE_RATE * duration)
    samples = []
    for i in range(total):
        t   = i / _SAMPLE_RATE
        val = _sine_wave(80.0, t) * 0.7 + _square_wave(80.0, t) * 0.3
        env = _envelope(i, total, attack_frac=0.05, release_frac=0.5)
        samples.append(_clamp_sample(val * env * 0.4))
    return samples


def generate_sounds_for_profile(
    profile_name: str,
    sounds_dir: Path,
    deck_name_prefix: str,
    log_fn=None,
) -> dict[str, bool]:
    """
    Generate all sound WAV files for a given sound profile.

    Files generated:
        [prefix]_alert.wav
        [prefix]_startup.wav
        [prefix]_idle.wav
        [prefix]_shutdown.wav
        [prefix]_error.wav

    Returns dict of {sound_name: success_bool}
    """
    def _log(msg: str) -> None:
        if log_fn:
            log_fn(msg)

    sounds_dir.mkdir(parents=True, exist_ok=True)

    profile = SOUND_PROFILES.get(profile_name)
    if not profile:
        _log(f"[SOUND] Unknown profile '{profile_name}' — using default_chime")
        profile = SOUND_PROFILES["default_chime"]

    if profile.get("status") == "stub":
        _log(f"[SOUND] Profile '{profile_name}' is a stub — using default_chime fallback")
        profile = SOUND_PROFILES["default_chime"]

    results: dict[str, bool] = {}
    prefix = deck_name_prefix.lower()

    # ── Alert ─────────────────────────────────────────────────────────
    alert_path = sounds_dir / f"{prefix}_alert.wav"
    if not alert_path.exists():
        try:
            notes   = profile.get("notes_alert", [(440.0, 0.5)])
            samples = _generate_notes(notes, profile)
            _write_wav(alert_path, samples)
            _log(f"[SOUND] Generated: {alert_path.name}")
            results["alert"] = True
        except Exception as e:
            _log(f"[SOUND] Alert generation failed: {e}")
            results["alert"] = False
    else:
        results["alert"] = True

    # ── Startup ───────────────────────────────────────────────────────
    startup_path = sounds_dir / f"{prefix}_startup.wav"
    if not startup_path.exists():
        try:
            notes   = profile.get("notes_startup", [(440.0, 0.6)])
            samples = _generate_notes(notes, profile, amplitude=0.4)
            _write_wav(startup_path, samples)
            _log(f"[SOUND] Generated: {startup_path.name}")
            results["startup"] = True
        except Exception as e:
            _log(f"[SOUND] Startup generation failed: {e}")
            results["startup"] = False
    else:
        results["startup"] = True

    # ── Idle ──────────────────────────────────────────────────────────
    idle_path = sounds_dir / f"{prefix}_idle.wav"
    if not idle_path.exists():
        try:
            notes   = profile.get("notes_idle", [(440.0, 0.8)])
            samples = _generate_notes(notes, profile, amplitude=0.25)
            _write_wav(idle_path, samples)
            _log(f"[SOUND] Generated: {idle_path.name}")
            results["idle"] = True
        except Exception as e:
            _log(f"[SOUND] Idle generation failed: {e}")
            results["idle"] = False
    else:
        results["idle"] = True

    # ── Shutdown ──────────────────────────────────────────────────────
    shutdown_path = sounds_dir / f"{prefix}_shutdown.wav"
    if not shutdown_path.exists():
        try:
            notes   = profile.get("notes_shutdown", [(440.0, 0.6)])
            samples = _generate_notes(notes, profile, amplitude=0.38)
            _write_wav(shutdown_path, samples)
            _log(f"[SOUND] Generated: {shutdown_path.name}")
            results["shutdown"] = True
        except Exception as e:
            _log(f"[SOUND] Shutdown generation failed: {e}")
            results["shutdown"] = False
    else:
        results["shutdown"] = True

    # ── Error ─────────────────────────────────────────────────────────
    error_path = sounds_dir / f"{prefix}_error.wav"
    if not error_path.exists():
        try:
            error_type = profile.get("error_type", "tritone")
            error_fns  = {
                "tritone":    _generate_tritone_error,
                "buzz":       _generate_buzz_error,
                "dissonant":  _generate_dissonant_error,
                "low_buzz":   _generate_low_buzz_error,
                "static":     _generate_buzz_error,
                "dramatic":   _generate_dissonant_error,
            }
            gen_fn  = error_fns.get(error_type, _generate_tritone_error)
            samples = gen_fn(profile)
            _write_wav(error_path, samples)
            _log(f"[SOUND] Generated: {error_path.name}")
            results["error"] = True
        except Exception as e:
            _log(f"[SOUND] Error sound generation failed: {e}")
            results["error"] = False
    else:
        results["error"] = True

    ok_count   = sum(1 for v in results.values() if v)
    fail_count = sum(1 for v in results.values() if not v)
    _log(f"[SOUND] {ok_count} sounds ready, {fail_count} failed")
    return results

# ── END OF SECTION 4 ──────────────────────────────────────────────────────────
# Next: Section 5 — Deck template (the universal base deck code with slots)


# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1D — MODULE CODE MARKERS
# Maps module keys to comment markers injected into the deck template.
# None = module not yet implemented, slot gets a placeholder comment.
# ═══════════════════════════════════════════════════════════════════════════════

DECK_TEMPLATE = '''\
from __future__ import annotations

# ═══════════════════════════════════════════════════════════════════════════════
# ECHO DECK — <<<WINDOW_TITLE>>>
# Filename   : <<<DECK_FILENAME>>>
# Version    : <<<DECK_VERSION>>>
# Generated  : <<<BUILD_DATE>>>
# Builder    : deck_builder.py
#
# THIS FILE IS GENERATED. Do not edit persona values directly.
# To change persona: re-run deck_builder.py with updated persona template.
# To change modules: re-run deck_builder.py with updated module selection.
#
# Installed modules:
<<<INSTALLED_MODULES_COMMENT>>>
# ═══════════════════════════════════════════════════════════════════════════════

# ── PERSONA CONFIGURATION (injected by deck_builder.py) ──────────────────────

DECK_NAME       = "<<<DECK_NAME>>>"
DECK_VERSION    = "<<<DECK_VERSION>>>"
FACE_PREFIX     = "<<<FACE_PREFIX>>>"
SOUND_PREFIX    = "<<<SOUND_PREFIX>>>"

# Color scheme — all UI colors come from here
C_PRIMARY       = "<<<COLOR_PRIMARY>>>"
C_SECONDARY     = "<<<COLOR_SECONDARY>>>"
C_ACCENT        = "<<<COLOR_ACCENT>>>"
C_BG            = "<<<COLOR_BACKGROUND>>>"
C_PANEL         = "<<<COLOR_PANEL>>>"
C_BORDER        = "<<<COLOR_BORDER>>>"
C_TEXT          = "<<<COLOR_TEXT>>>"
C_TEXT_DIM      = "<<<COLOR_TEXT_DIM>>>"

# Derived colors (computed from primary)
C_PRIMARY_DIM   = C_PRIMARY + "44"   # primary with opacity approximation
C_BLOOD         = "#8b0000"           # always deep red for errors
C_GREEN         = "#44aa66"           # always green for success
C_PURPLE        = "#8855cc"           # always purple for system messages

# UI Labels
UI_WINDOW_TITLE       = "<<<WINDOW_TITLE>>>"
UI_CHAT_WINDOW        = "<<<CHAT_WINDOW>>>"
UI_SEND_BUTTON        = "<<<SEND_BUTTON>>>"
UI_INPUT_PLACEHOLDER  = "<<<INPUT_PLACEHOLDER>>>"
UI_GENERATING_STATUS  = "<<<GENERATING_STATUS>>>"
UI_IDLE_STATUS        = "<<<IDLE_STATUS>>>"
UI_OFFLINE_STATUS     = "<<<OFFLINE_STATUS>>>"
UI_TORPOR_STATUS      = "<<<TORPOR_STATUS>>>"
UI_SUSPENSION_LABEL   = "<<<SUSPENSION_LABEL>>>"
UI_RUNES              = "<<<RUNES>>>"
UI_MIRROR_LABEL       = "<<<MIRROR_LABEL>>>"
UI_EMOTIONS_LABEL     = "<<<EMOTIONS_LABEL>>>"
UI_LEFT_ORB_TITLE     = "<<<LEFT_ORB_TITLE>>>"
UI_LEFT_ORB_LABEL     = "<<<LEFT_ORB_LABEL>>>"
UI_CYCLE_TITLE        = "<<<CYCLE_TITLE>>>"
UI_RIGHT_ORB_TITLE    = "<<<RIGHT_ORB_TITLE>>>"
UI_RIGHT_ORB_LABEL    = "<<<RIGHT_ORB_LABEL>>>"
UI_ESSENCE_TITLE      = "<<<ESSENCE_TITLE>>>"
UI_ESSENCE_PRIMARY    = "<<<ESSENCE_PRIMARY>>>"
UI_ESSENCE_SECONDARY  = "<<<ESSENCE_SECONDARY>>>"
UI_FOOTER_STRIP_LABEL = "<<<FOOTER_STRIP_LABEL>>>"

# System prompt and cognitive anchors
SYSTEM_PROMPT_BASE = """<<<SYSTEM_PROMPT>>>"""

COGNITIVE_ANCHORS = <<<COGNITIVE_ANCHORS>>>

# Special systems
AI_STATES_ENABLED       = <<<AI_STATES_ENABLED>>>
SUSPENSION_ENABLED      = <<<SUSPENSION_ENABLED>>>
ANCHOR_ENTITY           = <<<ANCHOR_ENTITY>>>
<<<AI_STATE_GREETINGS_DECL>>>

# Persona characteristics — from template
DECK_PRONOUN_SUBJECT    = "<<<DECK_PRONOUN_SUBJECT>>>"
DECK_PRONOUN_OBJECT     = "<<<DECK_PRONOUN_OBJECT>>>"
DECK_PRONOUN_POSSESSIVE = "<<<DECK_PRONOUN_POSSESSIVE>>>"
UI_FONT_FAMILY          = "<<<UI_FONT_FAMILY>>>"
UI_AWAKENING_LINE       = "<<<UI_AWAKENING_LINE>>>"

# ── END PERSONA/MODULE CONFIGURATION ─────────────────────────────────────────
# Everything below is the universal deck implementation.
# This code is identical across all generated decks.
# ─────────────────────────────────────────────────────────────────────────────

<<<DECK_IMPLEMENTATION>>>
'''
def build_deck_file(
    persona: dict,
    deck_name: str,
    selected_modules: list[str],
    output_path: Path,
    model_config: dict,
    ai_state_greetings: Optional[dict] = None,
    log_fn=None,
) -> bool:
    """
    Inject persona values and module code into DECK_TEMPLATE.
    Write the result to output_path.
    Returns True on success.

    The deck implementation (the actual working code from morganna_deck.py)
    is read from the current working deck and embedded.
    """
    def _log(msg: str) -> None:
        if log_fn:
            log_fn(msg)

    try:
        # ── Build installed modules comment ───────────────────────────
        installed_lines = []
        for mod_key in MODULES:
            mod      = MODULES[mod_key]
            is_sel   = mod_key in selected_modules
            status   = "[INSTALLED]" if is_sel else "[NOT INSTALLED]"
            installed_lines.append(
                f"#   {status} {mod['display_name']}"
            )
        installed_comment = "\n".join(installed_lines)

        # ── Build cognitive anchors as Python list literal ─────────────
        anchors = persona.get("cognitive_anchors", [])
        if anchors:
            anchor_lines = ",\n    ".join(f'"{a}"' for a in anchors)
            anchors_literal = f"[\n    {anchor_lines},\n]"
        else:
            anchors_literal = "[]"

        # ── Build anchor entity as Python dict or None ─────────────────
        ae = persona.get("anchor_entity")
        if ae:
            ae_literal = repr(ae)
        else:
            ae_literal = "None"

        # ── Build module slot values ───────────────────────────────────
        module_slots: dict[str, str] = {}
        template_slots = set(re.findall(r"<<<(MODULE_[A-Z0-9_]+)>>>", DECK_TEMPLATE))
        for slot in template_slots:
            module_slots[slot] = "# [NOT INSTALLED]"

        for mod_key, mod_data in MODULES.items():
            slot_key = mod_data["slot_key"]
            if mod_key in selected_modules:
                code = MODULE_CODE.get(mod_key) or mod_data.get("runtime_marker")
                if code:
                    module_slots[slot_key] = f"# [INSTALLED: {mod_data['display_name']}]\n{code}"
                else:
                    module_slots[slot_key] = (
                        f"# [INSTALLED: {mod_data['display_name']}] "
                        f"— code not yet built (status: {mod_data['status']})"
                    )
            else:
                module_slots[slot_key] = (
                    f"# [NOT INSTALLED: {mod_data['display_name']}]"
                )

        # ── Get deck implementation from current deck ──────────────────
        # Look for the existing working deck in the same directory
        deck_impl = _get_deck_implementation(selected_modules=selected_modules, log_fn=log_fn)

        # ── Build all replacements ─────────────────────────────────────
        colors     = persona.get("colors", {})
        ui_labels  = persona.get("ui_labels", {})
        deck_lower = deck_name.lower().replace(" ", "_")

        pronouns = persona.get("pronouns", {}) or {}
        pronoun_subject = pronouns.get("subject", "they")
        pronoun_object = pronouns.get("object", "them")
        pronoun_possessive = pronouns.get("possessive", "their")
        system_prompt_base = persona.get("system_prompt", "")
        system_prompt = (
            f"{system_prompt_base}\n\n"
            f"Your name is {deck_name}. "
            f"Your pronouns are {pronoun_subject}/{pronoun_object}/{pronoun_possessive}. "
            "Use these pronouns only for grammatical self-reference. "
            "Never use your pronouns as your name."
        )

        replacements = {
            "<<<DECK_VERSION>>>":          DECK_VERSION,
            "<<<DECK_NAME>>>":             deck_name,
            "<<<DECK_FILENAME>>>":         f"{deck_lower}_deck.py",
            "<<<BUILD_DATE>>>":            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "<<<WINDOW_TITLE>>>":          ui_labels.get("window_title", f"ECHO DECK — {deck_name.upper()}"),
            "<<<CHAT_WINDOW>>>":           ui_labels.get("chat_window",       "CONVERSATION"),
            "<<<SEND_BUTTON>>>":           ui_labels.get("send_button",       "SEND"),
            "<<<INPUT_PLACEHOLDER>>>":     ui_labels.get("input_placeholder", "Type your message..."),
            "<<<GENERATING_STATUS>>>":     ui_labels.get("generating_status", "◉ THINKING"),
            "<<<IDLE_STATUS>>>":           ui_labels.get("idle_status",        "◉ IDLE"),
            "<<<OFFLINE_STATUS>>>":        ui_labels.get("offline_status",     "◉ OFFLINE"),
            "<<<TORPOR_STATUS>>>":         ui_labels.get("torpor_status",      "◉ SUSPENDED"),
            "<<<SUSPENSION_LABEL>>>":      (ui_labels.get("suspension_label", "").strip() or "Suspend"),
            "<<<RUNES>>>":                 ui_labels.get("runes",              "— ECHO DECK —"),
            "<<<MIRROR_LABEL>>>":          ui_labels.get("mirror_label",       "MIRROR"),
            "<<<EMOTIONS_LABEL>>>":        ui_labels.get("emotions_label",     "EMOTIONS"),
            "<<<LEFT_ORB_TITLE>>>":        ui_labels.get("left_orb_title",     "PRIMARY"),
            "<<<LEFT_ORB_LABEL>>>":        ui_labels.get("left_orb_label",     "RESOURCE"),
            "<<<CYCLE_TITLE>>>":           ui_labels.get("cycle_title",        "CYCLE"),
            "<<<RIGHT_ORB_TITLE>>>":       ui_labels.get("right_orb_title",    "SECONDARY"),
            "<<<RIGHT_ORB_LABEL>>>":       ui_labels.get("right_orb_label",    "RESERVE"),
            "<<<ESSENCE_TITLE>>>":         ui_labels.get("essence_title",      "ESSENCE"),
            "<<<ESSENCE_PRIMARY>>>":       ui_labels.get("essence_primary",    "PRIMARY STATE"),
            "<<<ESSENCE_SECONDARY>>>":     ui_labels.get("essence_secondary",  "SECONDARY STATE"),
            "<<<FOOTER_STRIP_LABEL>>>":    ui_labels.get("footer_strip_label", "STATE"),
            "<<<FACE_PREFIX>>>":           persona.get("face_prefix", deck_name),
            "<<<SOUND_PREFIX>>>":          deck_lower,
            "<<<SYSTEM_PROMPT>>>":         system_prompt,
            "<<<COGNITIVE_ANCHORS>>>":     anchors_literal,
            "<<<COLOR_PRIMARY>>>":         colors.get("primary",    "#888888"),
            "<<<COLOR_SECONDARY>>>":       colors.get("secondary",  "#aaaaaa"),
            "<<<COLOR_ACCENT>>>":          colors.get("accent",     "#666666"),
            "<<<COLOR_BACKGROUND>>>":      colors.get("background", "#080808"),
            "<<<COLOR_PANEL>>>":           colors.get("panel",      "#101010"),
            "<<<COLOR_BORDER>>>":          colors.get("border",     "#2a2a2a"),
            "<<<COLOR_TEXT>>>":            colors.get("text",       "#e0e0e0"),
            "<<<COLOR_TEXT_DIM>>>":        colors.get("text_dim",   "#606060"),
            "<<<AI_STATES_ENABLED>>>":      str(bool(persona.get("vampire_states", False))),
            "<<<SUSPENSION_ENABLED>>>":     str(bool(persona.get("torpor_system",  False))),
            "<<<ANCHOR_ENTITY>>>":          ae_literal,
            "<<<AI_STATE_GREETINGS_DECL>>>": (
                f"AI_STATE_GREETINGS       = {repr(ai_state_greetings)}"
                if (bool(persona.get('vampire_states', False)) and ai_state_greetings)
                else ""
            ),
            "<<<INSTALLED_MODULES_COMMENT>>>": installed_comment,
            "<<<DECK_IMPLEMENTATION>>>":   deck_impl,
            "<<<DECK_PRONOUN_SUBJECT>>>":   pronoun_subject,
            "<<<DECK_PRONOUN_OBJECT>>>":    pronoun_object,
            "<<<DECK_PRONOUN_POSSESSIVE>>>": pronoun_possessive,
            "<<<UI_FONT_FAMILY>>>":         persona.get("font_family",    "'Georgia', 'Times New Roman', serif"),
            "<<<UI_AWAKENING_LINE>>>":      persona.get("awakening_line", "Connecting..."),
        }

        # Add module slot replacements
        for slot_key, slot_value in module_slots.items():
            replacements[f"<<<{slot_key}>>>"] = slot_value

        # ── Apply all replacements to template ────────────────────────
        result = DECK_TEMPLATE
        for placeholder, value in replacements.items():
            result = result.replace(placeholder, value)

        # ── Write output file ─────────────────────────────────────────
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(result, encoding="utf-8")
        _log(f"[DECK] Written: {output_path}")
        _log(f"[DECK] Size: {len(result.splitlines())} lines")
        return True

    except Exception as e:
        import traceback
        _log(f"[DECK] Write failed: {e}")
        _log(traceback.format_exc())
        return False


# ── EMBEDDED DECK IMPLEMENTATION ────────────────────────────────────────────
# Self-contained. No external files needed.
# Base64 encoded to avoid quote/backslash conflicts.
# Includes: chat formatting fix, all patches applied.

import base64 as _base64

_DECK_IMPL_B64 = (
    "IyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIAKCiMg4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQCiMgRUNITyBERUNLIOKAlCBVTklWRVJTQUwgSU1QTEVNRU5UQVRJT04KIyBHZW5lcmF0"
    "ZWQgYnkgZGVja19idWlsZGVyLnB5CiMgQWxsIHBlcnNvbmEgdmFsdWVzIGluamVjdGVkIGZyb20gREVDS19URU1QTEFURSBoZWFk"
    "ZXIuCiMg4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQCgojIOKUgOKUgCBQQVNTIDE6IEZPVU5EQVRJT04sIENPTlNUQU5UUywgSEVMUEVSUywgU09VTkQg"
    "R0VORVJBVE9SIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAoKCmltcG9ydCBz"
    "eXMKaW1wb3J0IG9zCmltcG9ydCBqc29uCmltcG9ydCBtYXRoCmltcG9ydCB0aW1lCmltcG9ydCB3YXZlCmltcG9ydCBzdHJ1Y3QK"
    "aW1wb3J0IHJhbmRvbQppbXBvcnQgdGhyZWFkaW5nCmltcG9ydCB1cmxsaWIucmVxdWVzdAppbXBvcnQgdXVpZApmcm9tIGRhdGV0"
    "aW1lIGltcG9ydCBkYXRldGltZSwgZGF0ZSwgdGltZWRlbHRhLCB0aW1lem9uZQpmcm9tIHpvbmVpbmZvIGltcG9ydCBab25lSW5m"
    "bywgWm9uZUluZm9Ob3RGb3VuZEVycm9yCmZyb20gcGF0aGxpYiBpbXBvcnQgUGF0aApmcm9tIHR5cGluZyBpbXBvcnQgT3B0aW9u"
    "YWwsIEl0ZXJhdG9yCgojIOKUgOKUgCBDb25zb2xlIGd1YXJkIOKAlCBzdXBwcmVzcyBDTUQgZmxhc2hlcyB3aGVuIGxhdW5jaGVk"
    "IHZpYSBweXRob253LmV4ZSDilIDilIDilIDilIDilIDilIDilIDilIAKIyBweXRob253LmV4ZSBoYXMgbm8gY29uc29sZS4gc3Rk"
    "b3V0L3N0ZGVyciB3cml0ZXMgKGluY2x1ZGluZyBweWdhbWUgYmFubmVyKQojIGNhdXNlIFdpbmRvd3MgdG8gYnJpZWZseSBhbGxv"
    "Y2F0ZSBhIENNRCB3aW5kb3cuIFJlZGlyZWN0IHRvIGRldm51bGwuCmltcG9ydCBvcyBhcyBfb3NfZ3VhcmQKdHJ5OgogICAgaWYg"
    "c3lzLmV4ZWN1dGFibGUgYW5kICJweXRob253IiBpbiBzeXMuZXhlY3V0YWJsZS5sb3dlcigpOgogICAgICAgIHN5cy5zdGRvdXQg"
    "PSBvcGVuKF9vc19ndWFyZC5kZXZudWxsLCAidyIpCiAgICAgICAgc3lzLnN0ZGVyciA9IG9wZW4oX29zX2d1YXJkLmRldm51bGws"
    "ICJ3IikKZXhjZXB0IEV4Y2VwdGlvbjoKICAgIHBhc3MKIyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKCgojIOKUgOKUgCBFQVJMWSBDUkFTSCBMT0dHRVIg4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSACiMgSG9va3MgaW4gYmVmb3JlIFF0LCBiZWZvcmUgZXZlcnl0aGluZy4gQ2FwdHVyZXMgQUxMIG91dHB1dCBp"
    "bmNsdWRpbmcKIyBDKysgbGV2ZWwgUXQgbWVzc2FnZXMuIFdyaXR0ZW4gdG8gW0RlY2tOYW1lXS9sb2dzL3N0YXJ0dXAubG9nCiMg"
    "VGhpcyBzdGF5cyBhY3RpdmUgZm9yIHRoZSBsaWZlIG9mIHRoZSBwcm9jZXNzLgoKX0VBUkxZX0xPR19MSU5FUzogbGlzdCA9IFtd"
    "Cl9FQVJMWV9MT0dfUEFUSDogT3B0aW9uYWxbUGF0aF0gPSBOb25lCgpkZWYgX2Vhcmx5X2xvZyhtc2c6IHN0cikgLT4gTm9uZToK"
    "ICAgIHRzID0gZGF0ZXRpbWUubm93KCkuc3RyZnRpbWUoIiVIOiVNOiVTLiVmIilbOi0zXQogICAgbGluZSA9IGYiW3t0c31dIHtt"
    "c2d9IgogICAgX0VBUkxZX0xPR19MSU5FUy5hcHBlbmQobGluZSkKICAgICMgTm8gcHJpbnQoKSDigJQgcHl0aG9udy5leGUgaGFz"
    "IG5vIGNvbnNvbGU7IHByaW50aW5nIGNhdXNlcyBDTUQgZmxhc2gKICAgIGlmIF9FQVJMWV9MT0dfUEFUSDoKICAgICAgICB0cnk6"
    "CiAgICAgICAgICAgIHdpdGggX0VBUkxZX0xPR19QQVRILm9wZW4oImEiLCBlbmNvZGluZz0idXRmLTgiKSBhcyBmOgogICAgICAg"
    "ICAgICAgICAgZi53cml0ZShsaW5lICsgIlxuIikKICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICBwYXNzCgpk"
    "ZWYgX2luaXRfZWFybHlfbG9nKGJhc2VfZGlyOiBQYXRoKSAtPiBOb25lOgogICAgZ2xvYmFsIF9FQVJMWV9MT0dfUEFUSAogICAg"
    "bG9nX2RpciA9IGJhc2VfZGlyIC8gImxvZ3MiCiAgICBsb2dfZGlyLm1rZGlyKHBhcmVudHM9VHJ1ZSwgZXhpc3Rfb2s9VHJ1ZSkK"
    "ICAgIF9FQVJMWV9MT0dfUEFUSCA9IGxvZ19kaXIgLyBmInN0YXJ0dXBfe2RhdGV0aW1lLm5vdygpLnN0cmZ0aW1lKCclWSVtJWRf"
    "JUglTSVTJyl9LmxvZyIKICAgICMgRmx1c2ggYnVmZmVyZWQgbGluZXMKICAgIHdpdGggX0VBUkxZX0xPR19QQVRILm9wZW4oInci"
    "LCBlbmNvZGluZz0idXRmLTgiKSBhcyBmOgogICAgICAgIGZvciBsaW5lIGluIF9FQVJMWV9MT0dfTElORVM6CiAgICAgICAgICAg"
    "IGYud3JpdGUobGluZSArICJcbiIpCgpkZWYgX2luc3RhbGxfcXRfbWVzc2FnZV9oYW5kbGVyKCkgLT4gTm9uZToKICAgICIiIgog"
    "ICAgSW50ZXJjZXB0IEFMTCBRdCBtZXNzYWdlcyBpbmNsdWRpbmcgQysrIGxldmVsIHdhcm5pbmdzLgogICAgVGhpcyBjYXRjaGVz"
    "IHRoZSBRVGhyZWFkIGRlc3Ryb3llZCBtZXNzYWdlIGF0IHRoZSBzb3VyY2UgYW5kIGxvZ3MgaXQKICAgIHdpdGggYSBmdWxsIHRy"
    "YWNlYmFjayBzbyB3ZSBrbm93IGV4YWN0bHkgd2hpY2ggdGhyZWFkIGFuZCB3aGVyZS4KICAgICIiIgogICAgdHJ5OgogICAgICAg"
    "IGZyb20gUHlTaWRlNi5RdENvcmUgaW1wb3J0IHFJbnN0YWxsTWVzc2FnZUhhbmRsZXIsIFF0TXNnVHlwZQogICAgICAgIGltcG9y"
    "dCB0cmFjZWJhY2sKCiAgICAgICAgZGVmIHF0X21lc3NhZ2VfaGFuZGxlcihtc2dfdHlwZSwgY29udGV4dCwgbWVzc2FnZSk6CiAg"
    "ICAgICAgICAgIGxldmVsID0gewogICAgICAgICAgICAgICAgUXRNc2dUeXBlLlF0RGVidWdNc2c6ICAgICJRVF9ERUJVRyIsCiAg"
    "ICAgICAgICAgICAgICBRdE1zZ1R5cGUuUXRJbmZvTXNnOiAgICAgIlFUX0lORk8iLAogICAgICAgICAgICAgICAgUXRNc2dUeXBl"
    "LlF0V2FybmluZ01zZzogICJRVF9XQVJOSU5HIiwKICAgICAgICAgICAgICAgIFF0TXNnVHlwZS5RdENyaXRpY2FsTXNnOiAiUVRf"
    "Q1JJVElDQUwiLAogICAgICAgICAgICAgICAgUXRNc2dUeXBlLlF0RmF0YWxNc2c6ICAgICJRVF9GQVRBTCIsCiAgICAgICAgICAg"
    "IH0uZ2V0KG1zZ190eXBlLCAiUVRfVU5LTk9XTiIpCgogICAgICAgICAgICBsb2NhdGlvbiA9ICIiCiAgICAgICAgICAgIGlmIGNv"
    "bnRleHQuZmlsZToKICAgICAgICAgICAgICAgIGxvY2F0aW9uID0gZiIgW3tjb250ZXh0LmZpbGV9Ontjb250ZXh0LmxpbmV9XSIK"
    "CiAgICAgICAgICAgIF9lYXJseV9sb2coZiJbe2xldmVsfV17bG9jYXRpb259IHttZXNzYWdlfSIpCgogICAgICAgICAgICAjIEZv"
    "ciBRVGhyZWFkIHdhcm5pbmdzIOKAlCBsb2cgZnVsbCBQeXRob24gc3RhY2sKICAgICAgICAgICAgaWYgIlFUaHJlYWQiIGluIG1l"
    "c3NhZ2Ugb3IgInRocmVhZCIgaW4gbWVzc2FnZS5sb3dlcigpOgogICAgICAgICAgICAgICAgc3RhY2sgPSAiIi5qb2luKHRyYWNl"
    "YmFjay5mb3JtYXRfc3RhY2soKSkKICAgICAgICAgICAgICAgIF9lYXJseV9sb2coZiJbU1RBQ0sgQVQgUVRIUkVBRCBXQVJOSU5H"
    "XVxue3N0YWNrfSIpCgogICAgICAgIHFJbnN0YWxsTWVzc2FnZUhhbmRsZXIocXRfbWVzc2FnZV9oYW5kbGVyKQogICAgICAgIF9l"
    "YXJseV9sb2coIltJTklUXSBRdCBtZXNzYWdlIGhhbmRsZXIgaW5zdGFsbGVkIikKICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMgZToK"
    "ICAgICAgICBfZWFybHlfbG9nKGYiW0lOSVRdIENvdWxkIG5vdCBpbnN0YWxsIFF0IG1lc3NhZ2UgaGFuZGxlcjoge2V9IikKCl9l"
    "YXJseV9sb2coZiJbSU5JVF0ge0RFQ0tfTkFNRX0gZGVjayBzdGFydGluZyIpCl9lYXJseV9sb2coZiJbSU5JVF0gUHl0aG9uIHtz"
    "eXMudmVyc2lvbi5zcGxpdCgpWzBdfSBhdCB7c3lzLmV4ZWN1dGFibGV9IikKX2Vhcmx5X2xvZyhmIltJTklUXSBXb3JraW5nIGRp"
    "cmVjdG9yeToge29zLmdldGN3ZCgpfSIpCl9lYXJseV9sb2coZiJbSU5JVF0gU2NyaXB0IGxvY2F0aW9uOiB7UGF0aChfX2ZpbGVf"
    "XykucmVzb2x2ZSgpfSIpCgojIOKUgOKUgCBPUFRJT05BTCBERVBFTkRFTkNZIEdVQVJEUyDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKClBTVVRJTF9PSyA9IEZhbHNlCnRyeToKICAg"
    "IGltcG9ydCBwc3V0aWwKICAgIFBTVVRJTF9PSyA9IFRydWUKICAgIF9lYXJseV9sb2coIltJTVBPUlRdIHBzdXRpbCBPSyIpCmV4"
    "Y2VwdCBJbXBvcnRFcnJvciBhcyBlOgogICAgX2Vhcmx5X2xvZyhmIltJTVBPUlRdIHBzdXRpbCBGQUlMRUQ6IHtlfSIpCgpOVk1M"
    "X09LID0gRmFsc2UKZ3B1X2hhbmRsZSA9IE5vbmUKdHJ5OgogICAgaW1wb3J0IHdhcm5pbmdzCiAgICB3aXRoIHdhcm5pbmdzLmNh"
    "dGNoX3dhcm5pbmdzKCk6CiAgICAgICAgd2FybmluZ3Muc2ltcGxlZmlsdGVyKCJpZ25vcmUiKQogICAgICAgIGltcG9ydCBweW52"
    "bWwKICAgIHB5bnZtbC5udm1sSW5pdCgpCiAgICBjb3VudCA9IHB5bnZtbC5udm1sRGV2aWNlR2V0Q291bnQoKQogICAgaWYgY291"
    "bnQgPiAwOgogICAgICAgIGdwdV9oYW5kbGUgPSBweW52bWwubnZtbERldmljZUdldEhhbmRsZUJ5SW5kZXgoMCkKICAgICAgICBO"
    "Vk1MX09LID0gVHJ1ZQogICAgX2Vhcmx5X2xvZyhmIltJTVBPUlRdIHB5bnZtbCBPSyDigJQge2NvdW50fSBHUFUocykiKQpleGNl"
    "cHQgRXhjZXB0aW9uIGFzIGU6CiAgICBfZWFybHlfbG9nKGYiW0lNUE9SVF0gcHludm1sIEZBSUxFRDoge2V9IikKClRPUkNIX09L"
    "ID0gRmFsc2UKdHJ5OgogICAgaW1wb3J0IHRvcmNoCiAgICBmcm9tIHRyYW5zZm9ybWVycyBpbXBvcnQgQXV0b01vZGVsRm9yQ2F1"
    "c2FsTE0sIEF1dG9Ub2tlbml6ZXIKICAgIFRPUkNIX09LID0gVHJ1ZQogICAgX2Vhcmx5X2xvZyhmIltJTVBPUlRdIHRvcmNoIHt0"
    "b3JjaC5fX3ZlcnNpb25fX30gT0siKQpleGNlcHQgSW1wb3J0RXJyb3IgYXMgZToKICAgIF9lYXJseV9sb2coZiJbSU1QT1JUXSB0"
    "b3JjaCBGQUlMRUQgKG9wdGlvbmFsKToge2V9IikKCldJTjMyX09LID0gRmFsc2UKdHJ5OgogICAgaW1wb3J0IHdpbjMyY29tLmNs"
    "aWVudAogICAgV0lOMzJfT0sgPSBUcnVlCiAgICBfZWFybHlfbG9nKCJbSU1QT1JUXSB3aW4zMmNvbSBPSyIpCmV4Y2VwdCBJbXBv"
    "cnRFcnJvciBhcyBlOgogICAgX2Vhcmx5X2xvZyhmIltJTVBPUlRdIHdpbjMyY29tIEZBSUxFRDoge2V9IikKCldJTlNPVU5EX09L"
    "ID0gRmFsc2UKdHJ5OgogICAgaW1wb3J0IHdpbnNvdW5kCiAgICBXSU5TT1VORF9PSyA9IFRydWUKICAgIF9lYXJseV9sb2coIltJ"
    "TVBPUlRdIHdpbnNvdW5kIE9LIikKZXhjZXB0IEltcG9ydEVycm9yIGFzIGU6CiAgICBfZWFybHlfbG9nKGYiW0lNUE9SVF0gd2lu"
    "c291bmQgRkFJTEVEIChvcHRpb25hbCk6IHtlfSIpCgpQWUdBTUVfT0sgPSBGYWxzZQp0cnk6CiAgICBpbXBvcnQgcHlnYW1lCiAg"
    "ICBweWdhbWUubWl4ZXIuaW5pdCgpCiAgICBQWUdBTUVfT0sgPSBUcnVlCiAgICBfZWFybHlfbG9nKCJbSU1QT1JUXSBweWdhbWUg"
    "T0siKQpleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICBfZWFybHlfbG9nKGYiW0lNUE9SVF0gcHlnYW1lIEZBSUxFRDoge2V9IikK"
    "CgoKIyDilIDilIAgUHlTaWRlNiBJTVBPUlRTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmcm9tIFB5U2lkZTYuUXRXaWRn"
    "ZXRzIGltcG9ydCAoCiAgICBRQXBwbGljYXRpb24sIFFNYWluV2luZG93LCBRV2lkZ2V0LCBRVkJveExheW91dCwgUUhCb3hMYXlv"
    "dXQsCiAgICBRR3JpZExheW91dCwgUVRleHRFZGl0LCBRTGluZUVkaXQsIFFQdXNoQnV0dG9uLCBRTGFiZWwsIFFGcmFtZSwKICAg"
    "IFFDYWxlbmRhcldpZGdldCwgUVRhYmxlV2lkZ2V0LCBRVGFibGVXaWRnZXRJdGVtLCBRSGVhZGVyVmlldywKICAgIFFBYnN0cmFj"
    "dEl0ZW1WaWV3LCBRU3RhY2tlZFdpZGdldCwgUVRhYldpZGdldCwgUUxpc3RXaWRnZXQsCiAgICBRTGlzdFdpZGdldEl0ZW0sIFFT"
    "aXplUG9saWN5LCBRQ29tYm9Cb3gsIFFDaGVja0JveCwgUUZpbGVEaWFsb2csCiAgICBRTWVzc2FnZUJveCwgUURhdGVFZGl0LCBR"
    "RGlhbG9nLCBRRm9ybUxheW91dCwgUVNjcm9sbEFyZWEsCiAgICBRU3BsaXR0ZXIsIFFJbnB1dERpYWxvZywgUVRvb2xCdXR0b24s"
    "IFFTcGluQm94LCBRR3JhcGhpY3NPcGFjaXR5RWZmZWN0LAogICAgUU1lbnUsIFFUYWJCYXIKKQpmcm9tIFB5U2lkZTYuUXRDb3Jl"
    "IGltcG9ydCAoCiAgICBRdCwgUVRpbWVyLCBRVGhyZWFkLCBTaWduYWwsIFFEYXRlLCBRU2l6ZSwgUVBvaW50LCBRUmVjdCwKICAg"
    "IFFQcm9wZXJ0eUFuaW1hdGlvbiwgUUVhc2luZ0N1cnZlCikKZnJvbSBQeVNpZGU2LlF0R3VpIGltcG9ydCAoCiAgICBRRm9udCwg"
    "UUNvbG9yLCBRUGFpbnRlciwgUUxpbmVhckdyYWRpZW50LCBRUmFkaWFsR3JhZGllbnQsCiAgICBRUGl4bWFwLCBRUGVuLCBRUGFp"
    "bnRlclBhdGgsIFFUZXh0Q2hhckZvcm1hdCwgUUljb24sCiAgICBRVGV4dEN1cnNvciwgUUFjdGlvbiwgUUZvbnRNZXRyaWNzCikK"
    "CiMg4pSA4pSAIEFQUCBJREVOVElUWSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKQVBQX05BTUUgICAgICA9"
    "IFVJX1dJTkRPV19USVRMRQpBUFBfVkVSU0lPTiAgID0gIjIuMC4wIgpBUFBfRklMRU5BTUUgID0gZiJ7REVDS19OQU1FLmxvd2Vy"
    "KCl9X2RlY2sucHkiCkJVSUxEX0RBVEUgICAgPSAiMjAyNi0wNC0wNCIKCiMg4pSA4pSAIENPTkZJRyBMT0FESU5HIOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAojIGNvbmZpZy5qc29uIGxpdmVzIG5leHQgdG8gdGhlIGRlY2sgLnB5IGZpbGUuCiMg"
    "QWxsIHBhdGhzIGNvbWUgZnJvbSBjb25maWcuIE5vdGhpbmcgaGFyZGNvZGVkIGJlbG93IHRoaXMgcG9pbnQuCgpTQ1JJUFRfRElS"
    "ID0gUGF0aChfX2ZpbGVfXykucmVzb2x2ZSgpLnBhcmVudApDT05GSUdfUEFUSCA9IFNDUklQVF9ESVIgLyAiY29uZmlnLmpzb24i"
    "CgojIEluaXRpYWxpemUgZWFybHkgbG9nIG5vdyB0aGF0IHdlIGtub3cgd2hlcmUgd2UgYXJlCl9pbml0X2Vhcmx5X2xvZyhTQ1JJ"
    "UFRfRElSKQpfZWFybHlfbG9nKGYiW0lOSVRdIFNDUklQVF9ESVIgPSB7U0NSSVBUX0RJUn0iKQpfZWFybHlfbG9nKGYiW0lOSVRd"
    "IENPTkZJR19QQVRIID0ge0NPTkZJR19QQVRIfSIpCl9lYXJseV9sb2coZiJbSU5JVF0gY29uZmlnLmpzb24gZXhpc3RzOiB7Q09O"
    "RklHX1BBVEguZXhpc3RzKCl9IikKCmRlZiBfZGVmYXVsdF9jb25maWcoKSAtPiBkaWN0OgogICAgIiIiUmV0dXJucyB0aGUgZGVm"
    "YXVsdCBjb25maWcgc3RydWN0dXJlIGZvciBmaXJzdC1ydW4gZ2VuZXJhdGlvbi4iIiIKICAgIGJhc2UgPSBzdHIoU0NSSVBUX0RJ"
    "UikKICAgIHJldHVybiB7CiAgICAgICAgImRlY2tfbmFtZSI6IERFQ0tfTkFNRSwKICAgICAgICAiZGVja192ZXJzaW9uIjogQVBQ"
    "X1ZFUlNJT04sCiAgICAgICAgImJhc2VfZGlyIjogYmFzZSwKICAgICAgICAibW9kZWwiOiB7CiAgICAgICAgICAgICJ0eXBlIjog"
    "ImxvY2FsIiwgICAgICAgICAgIyBsb2NhbCB8IG9sbGFtYSB8IGNsYXVkZSB8IG9wZW5haQogICAgICAgICAgICAicGF0aCI6ICIi"
    "LCAgICAgICAgICAgICAgICMgbG9jYWwgbW9kZWwgZm9sZGVyIHBhdGgKICAgICAgICAgICAgIm9sbGFtYV9tb2RlbCI6ICIiLCAg"
    "ICAgICAjIGUuZy4gImRvbHBoaW4tMi42LTdiIgogICAgICAgICAgICAiYXBpX2tleSI6ICIiLCAgICAgICAgICAgICMgQ2xhdWRl"
    "IG9yIE9wZW5BSSBrZXkKICAgICAgICAgICAgImFwaV90eXBlIjogIiIsICAgICAgICAgICAjICJjbGF1ZGUiIHwgIm9wZW5haSIK"
    "ICAgICAgICAgICAgImFwaV9tb2RlbCI6ICIiLCAgICAgICAgICAjIGUuZy4gImNsYXVkZS1zb25uZXQtNC02IgogICAgICAgIH0s"
    "CiAgICAgICAgImdvb2dsZSI6IHsKICAgICAgICAgICAgInRva2VuIjogICAgICAgc3RyKFNDUklQVF9ESVIgLyAiZ29vZ2xlIiAv"
    "ICJ0b2tlbi5qc29uIiksCiAgICAgICAgICAgICJ0aW1lem9uZSI6ICAgICJBbWVyaWNhL0NoaWNhZ28iLAogICAgICAgICAgICAi"
    "c2NvcGVzIjogWwogICAgICAgICAgICAgICAgImh0dHBzOi8vd3d3Lmdvb2dsZWFwaXMuY29tL2F1dGgvY2FsZW5kYXIuZXZlbnRz"
    "IiwKICAgICAgICAgICAgICAgICJodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL2RyaXZlIiwKICAgICAgICAgICAgICAg"
    "ICJodHRwczovL3d3dy5nb29nbGVhcGlzLmNvbS9hdXRoL2RvY3VtZW50cyIsCiAgICAgICAgICAgIF0sCiAgICAgICAgfSwKICAg"
    "ICAgICAicGF0aHMiOiB7CiAgICAgICAgICAgICJmYWNlcyI6ICAgIHN0cihTQ1JJUFRfRElSIC8gIkZhY2VzIiksCiAgICAgICAg"
    "ICAgICJzb3VuZHMiOiAgIHN0cihTQ1JJUFRfRElSIC8gInNvdW5kcyIpLAogICAgICAgICAgICAibWVtb3JpZXMiOiBzdHIoU0NS"
    "SVBUX0RJUiAvICJtZW1vcmllcyIpLAogICAgICAgICAgICAic2Vzc2lvbnMiOiBzdHIoU0NSSVBUX0RJUiAvICJzZXNzaW9ucyIp"
    "LAogICAgICAgICAgICAic2wiOiAgICAgICBzdHIoU0NSSVBUX0RJUiAvICJzbCIpLAogICAgICAgICAgICAiZXhwb3J0cyI6ICBz"
    "dHIoU0NSSVBUX0RJUiAvICJleHBvcnRzIiksCiAgICAgICAgICAgICJsb2dzIjogICAgIHN0cihTQ1JJUFRfRElSIC8gImxvZ3Mi"
    "KSwKICAgICAgICAgICAgImJhY2t1cHMiOiAgc3RyKFNDUklQVF9ESVIgLyAiYmFja3VwcyIpLAogICAgICAgICAgICAicGVyc29u"
    "YXMiOiBzdHIoU0NSSVBUX0RJUiAvICJwZXJzb25hcyIpLAogICAgICAgICAgICAiZ29vZ2xlIjogICBzdHIoU0NSSVBUX0RJUiAv"
    "ICJnb29nbGUiKSwKICAgICAgICB9LAogICAgICAgICJzZXR0aW5ncyI6IHsKICAgICAgICAgICAgImlkbGVfZW5hYmxlZCI6ICAg"
    "ICAgICAgICAgICBGYWxzZSwKICAgICAgICAgICAgImlkbGVfbWluX21pbnV0ZXMiOiAgICAgICAgICAxMCwKICAgICAgICAgICAg"
    "ImlkbGVfbWF4X21pbnV0ZXMiOiAgICAgICAgICAzMCwKICAgICAgICAgICAgImF1dG9zYXZlX2ludGVydmFsX21pbnV0ZXMiOiAx"
    "MCwKICAgICAgICAgICAgIm1heF9iYWNrdXBzIjogICAgICAgICAgICAgICAxMCwKICAgICAgICAgICAgInNvdW5kX2VuYWJsZWQi"
    "OiAgICAgICAgICAgICBUcnVlLAogICAgICAgICAgICAiZW1haWxfcmVmcmVzaF9pbnRlcnZhbF9tcyI6IDMwMDAwMCwKICAgICAg"
    "ICAgICAgImdvb2dsZV9sb29rYmFja19kYXlzIjogICAgICAzMCwKICAgICAgICAgICAgInVzZXJfZGVsYXlfdGhyZXNob2xkX21p"
    "biI6ICAzMCwKICAgICAgICAgICAgInRpbWV6b25lX2F1dG9fZGV0ZWN0IjogICAgICBUcnVlLAogICAgICAgICAgICAidGltZXpv"
    "bmVfb3ZlcnJpZGUiOiAgICAgICAgICIiLAogICAgICAgICAgICAiZnVsbHNjcmVlbl9lbmFibGVkIjogICAgICAgIEZhbHNlLAog"
    "ICAgICAgICAgICAiYm9yZGVybGVzc19lbmFibGVkIjogICAgICAgIEZhbHNlLAogICAgICAgIH0sCiAgICAgICAgIm1vZHVsZV90"
    "YWJfb3JkZXIiOiBbXSwKICAgICAgICAibWFpbl9zcGxpdHRlciI6IHsKICAgICAgICAgICAgImhvcml6b250YWxfc2l6ZXMiOiBb"
    "OTAwLCA1MDBdLAogICAgICAgIH0sCiAgICAgICAgImZpcnN0X3J1biI6IFRydWUsCiAgICB9CgpkZWYgbG9hZF9jb25maWcoKSAt"
    "PiBkaWN0OgogICAgIiIiTG9hZCBjb25maWcuanNvbi4gUmV0dXJucyBkZWZhdWx0IGlmIG1pc3Npbmcgb3IgY29ycnVwdC4iIiIK"
    "ICAgIGlmIG5vdCBDT05GSUdfUEFUSC5leGlzdHMoKToKICAgICAgICByZXR1cm4gX2RlZmF1bHRfY29uZmlnKCkKICAgIHRyeToK"
    "ICAgICAgICB3aXRoIENPTkZJR19QQVRILm9wZW4oInIiLCBlbmNvZGluZz0idXRmLTgiKSBhcyBmOgogICAgICAgICAgICByZXR1"
    "cm4ganNvbi5sb2FkKGYpCiAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgIHJldHVybiBfZGVmYXVsdF9jb25maWcoKQoKZGVm"
    "IHNhdmVfY29uZmlnKGNmZzogZGljdCkgLT4gTm9uZToKICAgICIiIldyaXRlIGNvbmZpZy5qc29uLiIiIgogICAgQ09ORklHX1BB"
    "VEgucGFyZW50Lm1rZGlyKHBhcmVudHM9VHJ1ZSwgZXhpc3Rfb2s9VHJ1ZSkKICAgIHdpdGggQ09ORklHX1BBVEgub3BlbigidyIs"
    "IGVuY29kaW5nPSJ1dGYtOCIpIGFzIGY6CiAgICAgICAganNvbi5kdW1wKGNmZywgZiwgaW5kZW50PTIpCgojIExvYWQgY29uZmln"
    "IGF0IG1vZHVsZSBsZXZlbCDigJQgZXZlcnl0aGluZyBiZWxvdyByZWFkcyBmcm9tIENGRwpDRkcgPSBsb2FkX2NvbmZpZygpCl9l"
    "YXJseV9sb2coZiJbSU5JVF0gQ29uZmlnIGxvYWRlZCDigJQgZmlyc3RfcnVuPXtDRkcuZ2V0KCdmaXJzdF9ydW4nKX0sIG1vZGVs"
    "X3R5cGU9e0NGRy5nZXQoJ21vZGVsJyx7fSkuZ2V0KCd0eXBlJyl9IikKCl9ERUZBVUxUX1BBVEhTOiBkaWN0W3N0ciwgUGF0aF0g"
    "PSB7CiAgICAiZmFjZXMiOiAgICBTQ1JJUFRfRElSIC8gIkZhY2VzIiwKICAgICJzb3VuZHMiOiAgIFNDUklQVF9ESVIgLyAic291"
    "bmRzIiwKICAgICJtZW1vcmllcyI6IFNDUklQVF9ESVIgLyAibWVtb3JpZXMiLAogICAgInNlc3Npb25zIjogU0NSSVBUX0RJUiAv"
    "ICJzZXNzaW9ucyIsCiAgICAic2wiOiAgICAgICBTQ1JJUFRfRElSIC8gInNsIiwKICAgICJleHBvcnRzIjogIFNDUklQVF9ESVIg"
    "LyAiZXhwb3J0cyIsCiAgICAibG9ncyI6ICAgICBTQ1JJUFRfRElSIC8gImxvZ3MiLAogICAgImJhY2t1cHMiOiAgU0NSSVBUX0RJ"
    "UiAvICJiYWNrdXBzIiwKICAgICJwZXJzb25hcyI6IFNDUklQVF9ESVIgLyAicGVyc29uYXMiLAogICAgImdvb2dsZSI6ICAgU0NS"
    "SVBUX0RJUiAvICJnb29nbGUiLAp9CgpkZWYgX25vcm1hbGl6ZV9jb25maWdfcGF0aHMoKSAtPiBOb25lOgogICAgIiIiCiAgICBT"
    "ZWxmLWhlYWwgb2xkZXIgY29uZmlnLmpzb24gZmlsZXMgbWlzc2luZyByZXF1aXJlZCBwYXRoIGtleXMuCiAgICBBZGRzIG1pc3Np"
    "bmcgcGF0aCBrZXlzIGFuZCBub3JtYWxpemVzIGdvb2dsZSBjcmVkZW50aWFsL3Rva2VuIGxvY2F0aW9ucywKICAgIHRoZW4gcGVy"
    "c2lzdHMgY29uZmlnLmpzb24gaWYgYW55dGhpbmcgY2hhbmdlZC4KICAgICIiIgogICAgY2hhbmdlZCA9IEZhbHNlCiAgICBwYXRo"
    "cyA9IENGRy5zZXRkZWZhdWx0KCJwYXRocyIsIHt9KQogICAgZm9yIGtleSwgZGVmYXVsdF9wYXRoIGluIF9ERUZBVUxUX1BBVEhT"
    "Lml0ZW1zKCk6CiAgICAgICAgaWYgbm90IHBhdGhzLmdldChrZXkpOgogICAgICAgICAgICBwYXRoc1trZXldID0gc3RyKGRlZmF1"
    "bHRfcGF0aCkKICAgICAgICAgICAgY2hhbmdlZCA9IFRydWUKCgogICAgc3BsaXR0ZXJfY2ZnID0gQ0ZHLnNldGRlZmF1bHQoIm1h"
    "aW5fc3BsaXR0ZXIiLCB7fSkKICAgIGlmIG5vdCBpc2luc3RhbmNlKHNwbGl0dGVyX2NmZywgZGljdCk6CiAgICAgICAgQ0ZHWyJt"
    "YWluX3NwbGl0dGVyIl0gPSB7Imhvcml6b250YWxfc2l6ZXMiOiBbOTAwLCA1MDBdfQogICAgICAgIGNoYW5nZWQgPSBUcnVlCiAg"
    "ICBlbHNlOgogICAgICAgIHNpemVzID0gc3BsaXR0ZXJfY2ZnLmdldCgiaG9yaXpvbnRhbF9zaXplcyIpCiAgICAgICAgdmFsaWRf"
    "c2l6ZXMgPSAoCiAgICAgICAgICAgIGlzaW5zdGFuY2Uoc2l6ZXMsIGxpc3QpCiAgICAgICAgICAgIGFuZCBsZW4oc2l6ZXMpID09"
    "IDIKICAgICAgICAgICAgYW5kIGFsbChpc2luc3RhbmNlKHYsIGludCkgZm9yIHYgaW4gc2l6ZXMpCiAgICAgICAgKQogICAgICAg"
    "IGlmIG5vdCB2YWxpZF9zaXplczoKICAgICAgICAgICAgc3BsaXR0ZXJfY2ZnWyJob3Jpem9udGFsX3NpemVzIl0gPSBbOTAwLCA1"
    "MDBdCiAgICAgICAgICAgIGNoYW5nZWQgPSBUcnVlCgogICAgaWYgY2hhbmdlZDoKICAgICAgICBzYXZlX2NvbmZpZyhDRkcpCgpk"
    "ZWYgY2ZnX3BhdGgoa2V5OiBzdHIpIC0+IFBhdGg6CiAgICAiIiJDb252ZW5pZW5jZTogZ2V0IGEgcGF0aCBmcm9tIENGR1sncGF0"
    "aHMnXVtrZXldIGFzIGEgUGF0aCBvYmplY3Qgd2l0aCBzYWZlIGZhbGxiYWNrIGRlZmF1bHRzLiIiIgogICAgcGF0aHMgPSBDRkcu"
    "Z2V0KCJwYXRocyIsIHt9KQogICAgdmFsdWUgPSBwYXRocy5nZXQoa2V5KQogICAgaWYgdmFsdWU6CiAgICAgICAgcmV0dXJuIFBh"
    "dGgodmFsdWUpCiAgICBmYWxsYmFjayA9IF9ERUZBVUxUX1BBVEhTLmdldChrZXkpCiAgICBpZiBmYWxsYmFjazoKICAgICAgICBw"
    "YXRoc1trZXldID0gc3RyKGZhbGxiYWNrKQogICAgICAgIHJldHVybiBmYWxsYmFjawogICAgcmV0dXJuIFNDUklQVF9ESVIgLyBr"
    "ZXkKCl9ub3JtYWxpemVfY29uZmlnX3BhdGhzKCkKCiMg4pSA4pSAIENPTE9SIENPTlNUQU5UUyDigJQgZGVyaXZlZCBmcm9tIHBl"
    "cnNvbmEgdGVtcGxhdGUg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSACiMgQ19QUklNQVJZLCBDX1NFQ09OREFSWSwgQ19BQ0NFTlQsIENfQkcsIENfUEFORUwsIENfQk9S"
    "REVSLAojIENfVEVYVCwgQ19URVhUX0RJTSBhcmUgaW5qZWN0ZWQgYXQgdGhlIHRvcCBvZiB0aGlzIGZpbGUgYnkgZGVja19idWls"
    "ZGVyLgojIEV2ZXJ5dGhpbmcgYmVsb3cgaXMgZGVyaXZlZCBmcm9tIHRob3NlIGluamVjdGVkIHZhbHVlcy4KCiMgU2VtYW50aWMg"
    "YWxpYXNlcyDigJQgbWFwIHBlcnNvbmEgY29sb3JzIHRvIG5hbWVkIHJvbGVzIHVzZWQgdGhyb3VnaG91dCB0aGUgVUkKQ19DUklN"
    "U09OICAgICA9IENfUFJJTUFSWSAgICAgICAgICAjIG1haW4gYWNjZW50IChidXR0b25zLCBib3JkZXJzLCBoaWdobGlnaHRzKQpD"
    "X0NSSU1TT05fRElNID0gQ19QUklNQVJZICsgIjg4IiAgICMgZGltIGFjY2VudCBmb3Igc3VidGxlIGJvcmRlcnMKQ19HT0xEICAg"
    "ICAgICA9IENfU0VDT05EQVJZICAgICAgICAjIG1haW4gbGFiZWwvdGV4dC9BSSBvdXRwdXQgY29sb3IKQ19HT0xEX0RJTSAgICA9"
    "IENfU0VDT05EQVJZICsgIjg4IiAjIGRpbSBzZWNvbmRhcnkKQ19HT0xEX0JSSUdIVCA9IENfQUNDRU5UICAgICAgICAgICAjIGVt"
    "cGhhc2lzLCBob3ZlciBzdGF0ZXMKQ19TSUxWRVIgICAgICA9IENfVEVYVF9ESU0gICAgICAgICAjIHNlY29uZGFyeSB0ZXh0IChh"
    "bHJlYWR5IGluamVjdGVkKQpDX1NJTFZFUl9ESU0gID0gQ19URVhUX0RJTSArICI4OCIgICMgZGltIHNlY29uZGFyeSB0ZXh0CkNf"
    "TU9OSVRPUiAgICAgPSBDX0JHICAgICAgICAgICAgICAgIyBjaGF0IGRpc3BsYXkgYmFja2dyb3VuZCAoYWxyZWFkeSBpbmplY3Rl"
    "ZCkKQ19CRzIgICAgICAgICA9IENfQkcgICAgICAgICAgICAgICAjIHNlY29uZGFyeSBiYWNrZ3JvdW5kCkNfQkczICAgICAgICAg"
    "PSBDX1BBTkVMICAgICAgICAgICAgIyB0ZXJ0aWFyeS9pbnB1dCBiYWNrZ3JvdW5kIChhbHJlYWR5IGluamVjdGVkKQpDX0JMT09E"
    "ICAgICAgID0gJyM4YjAwMDAnICAgICAgICAgICMgZXJyb3Igc3RhdGVzLCBkYW5nZXIg4oCUIHVuaXZlcnNhbApDX1BVUlBMRSAg"
    "ICAgID0gJyM4ODU1Y2MnICAgICAgICAgICMgU1lTVEVNIG1lc3NhZ2VzIOKAlCB1bml2ZXJzYWwKQ19QVVJQTEVfRElNICA9ICcj"
    "MmEwNTJhJyAgICAgICAgICAjIGRpbSBwdXJwbGUg4oCUIHVuaXZlcnNhbApDX0dSRUVOICAgICAgID0gJyM0NGFhNjYnICAgICAg"
    "ICAgICMgcG9zaXRpdmUgc3RhdGVzIOKAlCB1bml2ZXJzYWwKQ19CTFVFICAgICAgICA9ICcjNDQ4OGNjJyAgICAgICAgICAjIGlu"
    "Zm8gc3RhdGVzIOKAlCB1bml2ZXJzYWwKCiMgRm9udCBoZWxwZXIg4oCUIGV4dHJhY3RzIHByaW1hcnkgZm9udCBuYW1lIGZvciBR"
    "Rm9udCgpIGNhbGxzCkRFQ0tfRk9OVCA9IFVJX0ZPTlRfRkFNSUxZLnNwbGl0KCcsJylbMF0uc3RyaXAoKS5zdHJpcCgiJyIpCgoj"
    "IEVtb3Rpb24g4oaSIGNvbG9yIG1hcHBpbmcgKGZvciBlbW90aW9uIHJlY29yZCBjaGlwcykKRU1PVElPTl9DT0xPUlM6IGRpY3Rb"
    "c3RyLCBzdHJdID0gewogICAgInZpY3RvcnkiOiAgICBDX0dPTEQsCiAgICAic211ZyI6ICAgICAgIENfR09MRCwKICAgICJpbXBy"
    "ZXNzZWQiOiAgQ19HT0xELAogICAgInJlbGlldmVkIjogICBDX0dPTEQsCiAgICAiaGFwcHkiOiAgICAgIENfR09MRCwKICAgICJm"
    "bGlydHkiOiAgICAgQ19HT0xELAogICAgInBhbmlja2VkIjogICBDX0NSSU1TT04sCiAgICAiYW5ncnkiOiAgICAgIENfQ1JJTVNP"
    "TiwKICAgICJzaG9ja2VkIjogICAgQ19DUklNU09OLAogICAgImNoZWF0bW9kZSI6ICBDX0NSSU1TT04sCiAgICAiY29uY2VybmVk"
    "IjogICIjY2M2NjIyIiwKICAgICJzYWQiOiAgICAgICAgIiNjYzY2MjIiLAogICAgImh1bWlsaWF0ZWQiOiAiI2NjNjYyMiIsCiAg"
    "ICAiZmx1c3RlcmVkIjogICIjY2M2NjIyIiwKICAgICJwbG90dGluZyI6ICAgQ19QVVJQTEUsCiAgICAic3VzcGljaW91cyI6IENf"
    "UFVSUExFLAogICAgImVudmlvdXMiOiAgICBDX1BVUlBMRSwKICAgICJmb2N1c2VkIjogICAgQ19TSUxWRVIsCiAgICAiYWxlcnQi"
    "OiAgICAgIENfU0lMVkVSLAogICAgIm5ldXRyYWwiOiAgICBDX1RFWFRfRElNLAp9CgojIOKUgOKUgCBERUNPUkFUSVZFIENPTlNU"
    "QU5UUyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIAKIyBSVU5FUyBpcyBzb3VyY2VkIGZyb20gVUlfUlVORVMgaW5qZWN0ZWQgYnkgdGhlIHBlcnNvbmEg"
    "dGVtcGxhdGUKUlVORVMgPSBVSV9SVU5FUwoKIyBGYWNlIGltYWdlIG1hcCDigJQgcHJlZml4IGZyb20gRkFDRV9QUkVGSVgsIGZp"
    "bGVzIGxpdmUgaW4gY29uZmlnIHBhdGhzLmZhY2VzCkZBQ0VfRklMRVM6IGRpY3Rbc3RyLCBzdHJdID0gewogICAgIm5ldXRyYWwi"
    "OiAgICBmIntGQUNFX1BSRUZJWH1fTmV1dHJhbC5wbmciLAogICAgImFsZXJ0IjogICAgICBmIntGQUNFX1BSRUZJWH1fQWxlcnQu"
    "cG5nIiwKICAgICJmb2N1c2VkIjogICAgZiJ7RkFDRV9QUkVGSVh9X0ZvY3VzZWQucG5nIiwKICAgICJzbXVnIjogICAgICAgZiJ7"
    "RkFDRV9QUkVGSVh9X1NtdWcucG5nIiwKICAgICJjb25jZXJuZWQiOiAgZiJ7RkFDRV9QUkVGSVh9X0NvbmNlcm5lZC5wbmciLAog"
    "ICAgInNhZCI6ICAgICAgICBmIntGQUNFX1BSRUZJWH1fU2FkX0NyeWluZy5wbmciLAogICAgInJlbGlldmVkIjogICBmIntGQUNF"
    "X1BSRUZJWH1fUmVsaWV2ZWQucG5nIiwKICAgICJpbXByZXNzZWQiOiAgZiJ7RkFDRV9QUkVGSVh9X0ltcHJlc3NlZC5wbmciLAog"
    "ICAgInZpY3RvcnkiOiAgICBmIntGQUNFX1BSRUZJWH1fVmljdG9yeS5wbmciLAogICAgImh1bWlsaWF0ZWQiOiBmIntGQUNFX1BS"
    "RUZJWH1fSHVtaWxpYXRlZC5wbmciLAogICAgInN1c3BpY2lvdXMiOiBmIntGQUNFX1BSRUZJWH1fU3VzcGljaW91cy5wbmciLAog"
    "ICAgInBhbmlja2VkIjogICBmIntGQUNFX1BSRUZJWH1fUGFuaWNrZWQucG5nIiwKICAgICJjaGVhdG1vZGUiOiAgZiJ7RkFDRV9Q"
    "UkVGSVh9X0NoZWF0X01vZGUucG5nIiwKICAgICJhbmdyeSI6ICAgICAgZiJ7RkFDRV9QUkVGSVh9X0FuZ3J5LnBuZyIsCiAgICAi"
    "cGxvdHRpbmciOiAgIGYie0ZBQ0VfUFJFRklYfV9QbG90dGluZy5wbmciLAogICAgInNob2NrZWQiOiAgICBmIntGQUNFX1BSRUZJ"
    "WH1fU2hvY2tlZC5wbmciLAogICAgImhhcHB5IjogICAgICBmIntGQUNFX1BSRUZJWH1fSGFwcHkucG5nIiwKICAgICJmbGlydHki"
    "OiAgICAgZiJ7RkFDRV9QUkVGSVh9X0ZsaXJ0eS5wbmciLAogICAgImZsdXN0ZXJlZCI6ICBmIntGQUNFX1BSRUZJWH1fRmx1c3Rl"
    "cmVkLnBuZyIsCiAgICAiZW52aW91cyI6ICAgIGYie0ZBQ0VfUFJFRklYfV9FbnZpb3VzLnBuZyIsCn0KClNFTlRJTUVOVF9MSVNU"
    "ID0gKAogICAgIm5ldXRyYWwsIGFsZXJ0LCBmb2N1c2VkLCBzbXVnLCBjb25jZXJuZWQsIHNhZCwgcmVsaWV2ZWQsIGltcHJlc3Nl"
    "ZCwgIgogICAgInZpY3RvcnksIGh1bWlsaWF0ZWQsIHN1c3BpY2lvdXMsIHBhbmlja2VkLCBhbmdyeSwgcGxvdHRpbmcsIHNob2Nr"
    "ZWQsICIKICAgICJoYXBweSwgZmxpcnR5LCBmbHVzdGVyZWQsIGVudmlvdXMiCikKCiMg4pSA4pSAIFNZU1RFTSBQUk9NUFQg4oCU"
    "IGluamVjdGVkIGZyb20gcGVyc29uYSB0ZW1wbGF0ZSBhdCB0b3Agb2YgZmlsZSDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIAKIyBTWVNURU1fUFJPTVBUX0JBU0UgaXMgYWxyZWFkeSBkZWZpbmVkIGFib3ZlIGZyb20gPDw8U1lTVEVNX1BS"
    "T01QVD4+PiBpbmplY3Rpb24uCiMgRG8gbm90IHJlZGVmaW5lIGl0IGhlcmUuCgojIOKUgOKUgCBHTE9CQUwgU1RZTEVTSEVFVCDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIAKU1RZTEUgPSBmIiIiClFNYWluV2luZG93LCBRV2lkZ2V0IHt7CiAgICBiYWNrZ3JvdW5kLWNv"
    "bG9yOiB7Q19CR307CiAgICBjb2xvcjoge0NfR09MRH07CiAgICBmb250LWZhbWlseToge1VJX0ZPTlRfRkFNSUxZfTsKfX0KUVRl"
    "eHRFZGl0IHt7CiAgICBiYWNrZ3JvdW5kLWNvbG9yOiB7Q19NT05JVE9SfTsKICAgIGNvbG9yOiB7Q19HT0xEfTsKICAgIGJvcmRl"
    "cjogMXB4IHNvbGlkIHtDX0NSSU1TT05fRElNfTsKICAgIGJvcmRlci1yYWRpdXM6IDJweDsKICAgIGZvbnQtZmFtaWx5OiB7VUlf"
    "Rk9OVF9GQU1JTFl9OwogICAgZm9udC1zaXplOiAxMnB4OwogICAgcGFkZGluZzogOHB4OwogICAgc2VsZWN0aW9uLWJhY2tncm91"
    "bmQtY29sb3I6IHtDX0NSSU1TT05fRElNfTsKfX0KUUxpbmVFZGl0IHt7CiAgICBiYWNrZ3JvdW5kLWNvbG9yOiB7Q19CRzN9Owog"
    "ICAgY29sb3I6IHtDX0dPTER9OwogICAgYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTn07CiAgICBib3JkZXItcmFkaXVzOiAy"
    "cHg7CiAgICBmb250LWZhbWlseToge1VJX0ZPTlRfRkFNSUxZfTsKICAgIGZvbnQtc2l6ZTogMTNweDsKICAgIHBhZGRpbmc6IDhw"
    "eCAxMnB4Owp9fQpRTGluZUVkaXQ6Zm9jdXMge3sKICAgIGJvcmRlcjogMXB4IHNvbGlkIHtDX0dPTER9OwogICAgYmFja2dyb3Vu"
    "ZC1jb2xvcjoge0NfUEFORUx9Owp9fQpRUHVzaEJ1dHRvbiB7ewogICAgYmFja2dyb3VuZC1jb2xvcjoge0NfQ1JJTVNPTl9ESU19"
    "OwogICAgY29sb3I6IHtDX0dPTER9OwogICAgYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTn07CiAgICBib3JkZXItcmFkaXVz"
    "OiAycHg7CiAgICBmb250LWZhbWlseToge1VJX0ZPTlRfRkFNSUxZfTsKICAgIGZvbnQtc2l6ZTogMTJweDsKICAgIGZvbnQtd2Vp"
    "Z2h0OiBib2xkOwogICAgcGFkZGluZzogOHB4IDIwcHg7CiAgICBsZXR0ZXItc3BhY2luZzogMnB4Owp9fQpRUHVzaEJ1dHRvbjpo"
    "b3ZlciB7ewogICAgYmFja2dyb3VuZC1jb2xvcjoge0NfQ1JJTVNPTn07CiAgICBjb2xvcjoge0NfR09MRF9CUklHSFR9Owp9fQpR"
    "UHVzaEJ1dHRvbjpwcmVzc2VkIHt7CiAgICBiYWNrZ3JvdW5kLWNvbG9yOiB7Q19CTE9PRH07CiAgICBib3JkZXItY29sb3I6IHtD"
    "X0JMT09EfTsKICAgIGNvbG9yOiB7Q19URVhUfTsKfX0KUVB1c2hCdXR0b246ZGlzYWJsZWQge3sKICAgIGJhY2tncm91bmQtY29s"
    "b3I6IHtDX0JHM307CiAgICBjb2xvcjoge0NfVEVYVF9ESU19OwogICAgYm9yZGVyLWNvbG9yOiB7Q19URVhUX0RJTX07Cn19ClFT"
    "Y3JvbGxCYXI6dmVydGljYWwge3sKICAgIGJhY2tncm91bmQ6IHtDX0JHfTsKICAgIHdpZHRoOiA2cHg7CiAgICBib3JkZXI6IG5v"
    "bmU7Cn19ClFTY3JvbGxCYXI6OmhhbmRsZTp2ZXJ0aWNhbCB7ewogICAgYmFja2dyb3VuZDoge0NfQ1JJTVNPTl9ESU19OwogICAg"
    "Ym9yZGVyLXJhZGl1czogM3B4Owp9fQpRU2Nyb2xsQmFyOjpoYW5kbGU6dmVydGljYWw6aG92ZXIge3sKICAgIGJhY2tncm91bmQ6"
    "IHtDX0NSSU1TT059Owp9fQpRU2Nyb2xsQmFyOjphZGQtbGluZTp2ZXJ0aWNhbCwgUVNjcm9sbEJhcjo6c3ViLWxpbmU6dmVydGlj"
    "YWwge3sKICAgIGhlaWdodDogMHB4Owp9fQpRVGFiV2lkZ2V0OjpwYW5lIHt7CiAgICBib3JkZXI6IDFweCBzb2xpZCB7Q19DUklN"
    "U09OX0RJTX07CiAgICBiYWNrZ3JvdW5kOiB7Q19CRzJ9Owp9fQpRVGFiQmFyOjp0YWIge3sKICAgIGJhY2tncm91bmQ6IHtDX0JH"
    "M307CiAgICBjb2xvcjoge0NfVEVYVF9ESU19OwogICAgYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTl9ESU19OwogICAgcGFk"
    "ZGluZzogNnB4IDE0cHg7CiAgICBmb250LWZhbWlseToge1VJX0ZPTlRfRkFNSUxZfTsKICAgIGZvbnQtc2l6ZTogMTBweDsKICAg"
    "IGxldHRlci1zcGFjaW5nOiAxcHg7Cn19ClFUYWJCYXI6OnRhYjpzZWxlY3RlZCB7ewogICAgYmFja2dyb3VuZDoge0NfQ1JJTVNP"
    "Tl9ESU19OwogICAgY29sb3I6IHtDX0dPTER9OwogICAgYm9yZGVyLWJvdHRvbTogMnB4IHNvbGlkIHtDX0NSSU1TT059Owp9fQpR"
    "VGFiQmFyOjp0YWI6aG92ZXIge3sKICAgIGJhY2tncm91bmQ6IHtDX1BBTkVMfTsKICAgIGNvbG9yOiB7Q19HT0xEX0RJTX07Cn19"
    "ClFUYWJsZVdpZGdldCB7ewogICAgYmFja2dyb3VuZDoge0NfQkcyfTsKICAgIGNvbG9yOiB7Q19HT0xEfTsKICAgIGJvcmRlcjog"
    "MXB4IHNvbGlkIHtDX0NSSU1TT05fRElNfTsKICAgIGdyaWRsaW5lLWNvbG9yOiB7Q19CT1JERVJ9OwogICAgZm9udC1mYW1pbHk6"
    "IHtVSV9GT05UX0ZBTUlMWX07CiAgICBmb250LXNpemU6IDExcHg7Cn19ClFUYWJsZVdpZGdldDo6aXRlbTpzZWxlY3RlZCB7ewog"
    "ICAgYmFja2dyb3VuZDoge0NfQ1JJTVNPTl9ESU19OwogICAgY29sb3I6IHtDX0dPTERfQlJJR0hUfTsKfX0KUUhlYWRlclZpZXc6"
    "OnNlY3Rpb24ge3sKICAgIGJhY2tncm91bmQ6IHtDX0JHM307CiAgICBjb2xvcjoge0NfR09MRH07CiAgICBib3JkZXI6IDFweCBz"
    "b2xpZCB7Q19DUklNU09OX0RJTX07CiAgICBwYWRkaW5nOiA0cHg7CiAgICBmb250LWZhbWlseToge1VJX0ZPTlRfRkFNSUxZfTsK"
    "ICAgIGZvbnQtc2l6ZTogMTBweDsKICAgIGZvbnQtd2VpZ2h0OiBib2xkOwogICAgbGV0dGVyLXNwYWNpbmc6IDFweDsKfX0KUUNv"
    "bWJvQm94IHt7CiAgICBiYWNrZ3JvdW5kOiB7Q19CRzN9OwogICAgY29sb3I6IHtDX0dPTER9OwogICAgYm9yZGVyOiAxcHggc29s"
    "aWQge0NfQ1JJTVNPTl9ESU19OwogICAgcGFkZGluZzogNHB4IDhweDsKICAgIGZvbnQtZmFtaWx5OiB7VUlfRk9OVF9GQU1JTFl9"
    "Owp9fQpRQ29tYm9Cb3g6OmRyb3AtZG93biB7ewogICAgYm9yZGVyOiBub25lOwp9fQpRQ2hlY2tCb3gge3sKICAgIGNvbG9yOiB7"
    "Q19HT0xEfTsKICAgIGZvbnQtZmFtaWx5OiB7VUlfRk9OVF9GQU1JTFl9Owp9fQpRTGFiZWwge3sKICAgIGNvbG9yOiB7Q19HT0xE"
    "fTsKICAgIGJvcmRlcjogbm9uZTsKfX0KUVNwbGl0dGVyOjpoYW5kbGUge3sKICAgIGJhY2tncm91bmQ6IHtDX0NSSU1TT05fRElN"
    "fTsKICAgIHdpZHRoOiAycHg7Cn19CiIiIgoKIyDilIDilIAgRElSRUNUT1JZIEJPT1RTVFJBUCDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZGVm"
    "IGJvb3RzdHJhcF9kaXJlY3RvcmllcygpIC0+IE5vbmU6CiAgICAiIiIKICAgIENyZWF0ZSBhbGwgcmVxdWlyZWQgZGlyZWN0b3Jp"
    "ZXMgaWYgdGhleSBkb24ndCBleGlzdC4KICAgIENhbGxlZCBvbiBzdGFydHVwIGJlZm9yZSBhbnl0aGluZyBlbHNlLiBTYWZlIHRv"
    "IGNhbGwgbXVsdGlwbGUgdGltZXMuCiAgICBBbHNvIG1pZ3JhdGVzIGZpbGVzIGZyb20gb2xkIFtEZWNrTmFtZV1fTWVtb3JpZXMg"
    "bGF5b3V0IGlmIGRldGVjdGVkLgogICAgIiIiCiAgICBkaXJzID0gWwogICAgICAgIGNmZ19wYXRoKCJmYWNlcyIpLAogICAgICAg"
    "IGNmZ19wYXRoKCJzb3VuZHMiKSwKICAgICAgICBjZmdfcGF0aCgibWVtb3JpZXMiKSwKICAgICAgICBjZmdfcGF0aCgic2Vzc2lv"
    "bnMiKSwKICAgICAgICBjZmdfcGF0aCgic2wiKSwKICAgICAgICBjZmdfcGF0aCgiZXhwb3J0cyIpLAogICAgICAgIGNmZ19wYXRo"
    "KCJsb2dzIiksCiAgICAgICAgY2ZnX3BhdGgoImJhY2t1cHMiKSwKICAgICAgICBjZmdfcGF0aCgicGVyc29uYXMiKSwKICAgIF0K"
    "ICAgIGZvciBkIGluIGRpcnM6CiAgICAgICAgZC5ta2RpcihwYXJlbnRzPVRydWUsIGV4aXN0X29rPVRydWUpCgogICAgIyBDcmVh"
    "dGUgZW1wdHkgSlNPTkwgZmlsZXMgaWYgdGhleSBkb24ndCBleGlzdAogICAgbWVtb3J5X2RpciA9IGNmZ19wYXRoKCJtZW1vcmll"
    "cyIpCiAgICBmb3IgZm5hbWUgaW4gKCJtZXNzYWdlcy5qc29ubCIsICJtZW1vcmllcy5qc29ubCIsICJ0YXNrcy5qc29ubCIsCiAg"
    "ICAgICAgICAgICAgICAgICJsZXNzb25zX2xlYXJuZWQuanNvbmwiLCAicGVyc29uYV9oaXN0b3J5Lmpzb25sIik6CiAgICAgICAg"
    "ZnAgPSBtZW1vcnlfZGlyIC8gZm5hbWUKICAgICAgICBpZiBub3QgZnAuZXhpc3RzKCk6CiAgICAgICAgICAgIGZwLndyaXRlX3Rl"
    "eHQoIiIsIGVuY29kaW5nPSJ1dGYtOCIpCgogICAgc2xfZGlyID0gY2ZnX3BhdGgoInNsIikKICAgIGZvciBmbmFtZSBpbiAoInNs"
    "X3NjYW5zLmpzb25sIiwgInNsX2NvbW1hbmRzLmpzb25sIik6CiAgICAgICAgZnAgPSBzbF9kaXIgLyBmbmFtZQogICAgICAgIGlm"
    "IG5vdCBmcC5leGlzdHMoKToKICAgICAgICAgICAgZnAud3JpdGVfdGV4dCgiIiwgZW5jb2Rpbmc9InV0Zi04IikKCiAgICBzZXNz"
    "aW9uc19kaXIgPSBjZmdfcGF0aCgic2Vzc2lvbnMiKQogICAgaWR4ID0gc2Vzc2lvbnNfZGlyIC8gInNlc3Npb25faW5kZXguanNv"
    "biIKICAgIGlmIG5vdCBpZHguZXhpc3RzKCk6CiAgICAgICAgaWR4LndyaXRlX3RleHQoanNvbi5kdW1wcyh7InNlc3Npb25zIjog"
    "W119LCBpbmRlbnQ9MiksIGVuY29kaW5nPSJ1dGYtOCIpCgogICAgc3RhdGVfcGF0aCA9IG1lbW9yeV9kaXIgLyAic3RhdGUuanNv"
    "biIKICAgIGlmIG5vdCBzdGF0ZV9wYXRoLmV4aXN0cygpOgogICAgICAgIF93cml0ZV9kZWZhdWx0X3N0YXRlKHN0YXRlX3BhdGgp"
    "CgogICAgaW5kZXhfcGF0aCA9IG1lbW9yeV9kaXIgLyAiaW5kZXguanNvbiIKICAgIGlmIG5vdCBpbmRleF9wYXRoLmV4aXN0cygp"
    "OgogICAgICAgIGluZGV4X3BhdGgud3JpdGVfdGV4dCgKICAgICAgICAgICAganNvbi5kdW1wcyh7InZlcnNpb24iOiBBUFBfVkVS"
    "U0lPTiwgInRvdGFsX21lc3NhZ2VzIjogMCwKICAgICAgICAgICAgICAgICAgICAgICAgInRvdGFsX21lbW9yaWVzIjogMH0sIGlu"
    "ZGVudD0yKSwKICAgICAgICAgICAgZW5jb2Rpbmc9InV0Zi04IgogICAgICAgICkKCiAgICAjIExlZ2FjeSBtaWdyYXRpb246IGlm"
    "IG9sZCBNb3JnYW5uYV9NZW1vcmllcyBmb2xkZXIgZXhpc3RzLCBtaWdyYXRlIGZpbGVzCiAgICBfbWlncmF0ZV9sZWdhY3lfZmls"
    "ZXMoKQoKZGVmIF93cml0ZV9kZWZhdWx0X3N0YXRlKHBhdGg6IFBhdGgpIC0+IE5vbmU6CiAgICBzdGF0ZSA9IHsKICAgICAgICAi"
    "cGVyc29uYV9uYW1lIjogREVDS19OQU1FLAogICAgICAgICJkZWNrX3ZlcnNpb24iOiBBUFBfVkVSU0lPTiwKICAgICAgICAic2Vz"
    "c2lvbl9jb3VudCI6IDAsCiAgICAgICAgImxhc3Rfc3RhcnR1cCI6IE5vbmUsCiAgICAgICAgImxhc3Rfc2h1dGRvd24iOiBOb25l"
    "LAogICAgICAgICJsYXN0X2FjdGl2ZSI6IE5vbmUsCiAgICAgICAgInRvdGFsX21lc3NhZ2VzIjogMCwKICAgICAgICAidG90YWxf"
    "bWVtb3JpZXMiOiAwLAogICAgICAgICJpbnRlcm5hbF9uYXJyYXRpdmUiOiB7fSwKICAgICAgICAiYWlfc3RhdGVfYXRfc2h1dGRv"
    "d24iOiAiRE9STUFOVCIsCiAgICB9CiAgICBwYXRoLndyaXRlX3RleHQoanNvbi5kdW1wcyhzdGF0ZSwgaW5kZW50PTIpLCBlbmNv"
    "ZGluZz0idXRmLTgiKQoKZGVmIF9taWdyYXRlX2xlZ2FjeV9maWxlcygpIC0+IE5vbmU6CiAgICAiIiIKICAgIElmIG9sZCBEOlxc"
    "QUlcXE1vZGVsc1xcW0RlY2tOYW1lXV9NZW1vcmllcyBsYXlvdXQgaXMgZGV0ZWN0ZWQsCiAgICBtaWdyYXRlIGZpbGVzIHRvIG5l"
    "dyBzdHJ1Y3R1cmUgc2lsZW50bHkuCiAgICAiIiIKICAgICMgVHJ5IHRvIGZpbmQgb2xkIGxheW91dCByZWxhdGl2ZSB0byBtb2Rl"
    "bCBwYXRoCiAgICBtb2RlbF9wYXRoID0gUGF0aChDRkdbIm1vZGVsIl0uZ2V0KCJwYXRoIiwgIiIpKQogICAgaWYgbm90IG1vZGVs"
    "X3BhdGguZXhpc3RzKCk6CiAgICAgICAgcmV0dXJuCiAgICBvbGRfcm9vdCA9IG1vZGVsX3BhdGgucGFyZW50IC8gZiJ7REVDS19O"
    "QU1FfV9NZW1vcmllcyIKICAgIGlmIG5vdCBvbGRfcm9vdC5leGlzdHMoKToKICAgICAgICByZXR1cm4KCiAgICBtaWdyYXRpb25z"
    "ID0gWwogICAgICAgIChvbGRfcm9vdCAvICJtZW1vcmllcy5qc29ubCIsICAgICAgICAgICBjZmdfcGF0aCgibWVtb3JpZXMiKSAv"
    "ICJtZW1vcmllcy5qc29ubCIpLAogICAgICAgIChvbGRfcm9vdCAvICJtZXNzYWdlcy5qc29ubCIsICAgICAgICAgICAgY2ZnX3Bh"
    "dGgoIm1lbW9yaWVzIikgLyAibWVzc2FnZXMuanNvbmwiKSwKICAgICAgICAob2xkX3Jvb3QgLyAidGFza3MuanNvbmwiLCAgICAg"
    "ICAgICAgICAgIGNmZ19wYXRoKCJtZW1vcmllcyIpIC8gInRhc2tzLmpzb25sIiksCiAgICAgICAgKG9sZF9yb290IC8gInN0YXRl"
    "Lmpzb24iLCAgICAgICAgICAgICAgICBjZmdfcGF0aCgibWVtb3JpZXMiKSAvICJzdGF0ZS5qc29uIiksCiAgICAgICAgKG9sZF9y"
    "b290IC8gImluZGV4Lmpzb24iLCAgICAgICAgICAgICAgICBjZmdfcGF0aCgibWVtb3JpZXMiKSAvICJpbmRleC5qc29uIiksCiAg"
    "ICAgICAgKG9sZF9yb290IC8gInNsX3NjYW5zLmpzb25sIiwgICAgICAgICAgICBjZmdfcGF0aCgic2wiKSAvICJzbF9zY2Fucy5q"
    "c29ubCIpLAogICAgICAgIChvbGRfcm9vdCAvICJzbF9jb21tYW5kcy5qc29ubCIsICAgICAgICAgY2ZnX3BhdGgoInNsIikgLyAi"
    "c2xfY29tbWFuZHMuanNvbmwiKSwKICAgICAgICAob2xkX3Jvb3QgLyAic291bmRzIiAvIGYie1NPVU5EX1BSRUZJWH1fYWxlcnQu"
    "d2F2IiwKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBjZmdfcGF0aCgic291bmRzIikg"
    "LyBmIntTT1VORF9QUkVGSVh9X2FsZXJ0LndhdiIpLAogICAgXQoKICAgIGZvciBzcmMsIGRzdCBpbiBtaWdyYXRpb25zOgogICAg"
    "ICAgIGlmIHNyYy5leGlzdHMoKSBhbmQgbm90IGRzdC5leGlzdHMoKToKICAgICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAg"
    "ZHN0LnBhcmVudC5ta2RpcihwYXJlbnRzPVRydWUsIGV4aXN0X29rPVRydWUpCiAgICAgICAgICAgICAgICBpbXBvcnQgc2h1dGls"
    "CiAgICAgICAgICAgICAgICBzaHV0aWwuY29weTIoc3RyKHNyYyksIHN0cihkc3QpKQogICAgICAgICAgICBleGNlcHQgRXhjZXB0"
    "aW9uOgogICAgICAgICAgICAgICAgcGFzcwoKICAgICMgTWlncmF0ZSBmYWNlIGltYWdlcwogICAgb2xkX2ZhY2VzID0gb2xkX3Jv"
    "b3QgLyAiRmFjZXMiCiAgICBuZXdfZmFjZXMgPSBjZmdfcGF0aCgiZmFjZXMiKQogICAgaWYgb2xkX2ZhY2VzLmV4aXN0cygpOgog"
    "ICAgICAgIGZvciBpbWcgaW4gb2xkX2ZhY2VzLmdsb2IoIioucG5nIik6CiAgICAgICAgICAgIGRzdCA9IG5ld19mYWNlcyAvIGlt"
    "Zy5uYW1lCiAgICAgICAgICAgIGlmIG5vdCBkc3QuZXhpc3RzKCk6CiAgICAgICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAg"
    "ICAgICAgaW1wb3J0IHNodXRpbAogICAgICAgICAgICAgICAgICAgIHNodXRpbC5jb3B5MihzdHIoaW1nKSwgc3RyKGRzdCkpCiAg"
    "ICAgICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICAgICAgICAgIHBhc3MKCiMg4pSA4pSAIERBVEVUSU1F"
    "IEhFTFBFUlMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmRlZiBsb2NhbF9ub3dfaXNvKCkgLT4gc3RyOgogICAgcmV0dXJuIGRh"
    "dGV0aW1lLm5vdygpLnJlcGxhY2UobWljcm9zZWNvbmQ9MCkuaXNvZm9ybWF0KCkKCmRlZiBwYXJzZV9pc28odmFsdWU6IHN0cikg"
    "LT4gT3B0aW9uYWxbZGF0ZXRpbWVdOgogICAgaWYgbm90IHZhbHVlOgogICAgICAgIHJldHVybiBOb25lCiAgICB2YWx1ZSA9IHZh"
    "bHVlLnN0cmlwKCkKICAgIHRyeToKICAgICAgICBpZiB2YWx1ZS5lbmRzd2l0aCgiWiIpOgogICAgICAgICAgICByZXR1cm4gZGF0"
    "ZXRpbWUuZnJvbWlzb2Zvcm1hdCh2YWx1ZVs6LTFdKS5yZXBsYWNlKHR6aW5mbz10aW1lem9uZS51dGMpCiAgICAgICAgcmV0dXJu"
    "IGRhdGV0aW1lLmZyb21pc29mb3JtYXQodmFsdWUpCiAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgIHJldHVybiBOb25lCgpf"
    "REFURVRJTUVfTk9STUFMSVpBVElPTl9MT0dHRUQ6IHNldFt0dXBsZV0gPSBzZXQoKQoKCmRlZiBfcmVzb2x2ZV9kZWNrX3RpbWV6"
    "b25lX25hbWUoKSAtPiBPcHRpb25hbFtzdHJdOgogICAgc2V0dGluZ3MgPSBDRkcuZ2V0KCJzZXR0aW5ncyIsIHt9KSBpZiBpc2lu"
    "c3RhbmNlKENGRywgZGljdCkgZWxzZSB7fQogICAgYXV0b19kZXRlY3QgPSBib29sKHNldHRpbmdzLmdldCgidGltZXpvbmVfYXV0"
    "b19kZXRlY3QiLCBUcnVlKSkKICAgIG92ZXJyaWRlID0gc3RyKHNldHRpbmdzLmdldCgidGltZXpvbmVfb3ZlcnJpZGUiLCAiIikg"
    "b3IgIiIpLnN0cmlwKCkKICAgIGlmIG5vdCBhdXRvX2RldGVjdCBhbmQgb3ZlcnJpZGU6CiAgICAgICAgcmV0dXJuIG92ZXJyaWRl"
    "CiAgICBsb2NhbF90emluZm8gPSBkYXRldGltZS5ub3coKS5hc3RpbWV6b25lKCkudHppbmZvCiAgICBpZiBsb2NhbF90emluZm8g"
    "aXMgbm90IE5vbmU6CiAgICAgICAgdHpfa2V5ID0gZ2V0YXR0cihsb2NhbF90emluZm8sICJrZXkiLCBOb25lKQogICAgICAgIGlm"
    "IHR6X2tleToKICAgICAgICAgICAgcmV0dXJuIHN0cih0el9rZXkpCiAgICAgICAgdHpfbmFtZSA9IHN0cihsb2NhbF90emluZm8p"
    "CiAgICAgICAgaWYgdHpfbmFtZSBhbmQgdHpfbmFtZS51cHBlcigpICE9ICJMT0NBTCI6CiAgICAgICAgICAgIHJldHVybiB0el9u"
    "YW1lCiAgICByZXR1cm4gTm9uZQoKCmRlZiBfbG9jYWxfdHppbmZvKCk6CiAgICB0el9uYW1lID0gX3Jlc29sdmVfZGVja190aW1l"
    "em9uZV9uYW1lKCkKICAgIGlmIHR6X25hbWU6CiAgICAgICAgdHJ5OgogICAgICAgICAgICByZXR1cm4gWm9uZUluZm8odHpfbmFt"
    "ZSkKICAgICAgICBleGNlcHQgWm9uZUluZm9Ob3RGb3VuZEVycm9yOgogICAgICAgICAgICBfZWFybHlfbG9nKGYiW0RBVEVUSU1F"
    "XVtXQVJOXSBVbmtub3duIHRpbWV6b25lIG92ZXJyaWRlICd7dHpfbmFtZX0nLCB1c2luZyBzeXN0ZW0gbG9jYWwgdGltZXpvbmUu"
    "IikKICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICBwYXNzCiAgICByZXR1cm4gZGF0ZXRpbWUubm93KCkuYXN0"
    "aW1lem9uZSgpLnR6aW5mbyBvciB0aW1lem9uZS51dGMKCgpkZWYgbm93X2Zvcl9jb21wYXJlKCk6CiAgICByZXR1cm4gZGF0ZXRp"
    "bWUubm93KF9sb2NhbF90emluZm8oKSkKCgpkZWYgbm9ybWFsaXplX2RhdGV0aW1lX2Zvcl9jb21wYXJlKGR0X3ZhbHVlLCBjb250"
    "ZXh0OiBzdHIgPSAiIik6CiAgICBpZiBkdF92YWx1ZSBpcyBOb25lOgogICAgICAgIHJldHVybiBOb25lCiAgICBpZiBub3QgaXNp"
    "bnN0YW5jZShkdF92YWx1ZSwgZGF0ZXRpbWUpOgogICAgICAgIHJldHVybiBOb25lCiAgICBsb2NhbF90eiA9IF9sb2NhbF90emlu"
    "Zm8oKQogICAgaWYgZHRfdmFsdWUudHppbmZvIGlzIE5vbmU6CiAgICAgICAgbm9ybWFsaXplZCA9IGR0X3ZhbHVlLnJlcGxhY2Uo"
    "dHppbmZvPWxvY2FsX3R6KQogICAgICAgIGtleSA9ICgibmFpdmUiLCBjb250ZXh0KQogICAgICAgIGlmIGtleSBub3QgaW4gX0RB"
    "VEVUSU1FX05PUk1BTElaQVRJT05fTE9HR0VEOgogICAgICAgICAgICBfZWFybHlfbG9nKAogICAgICAgICAgICAgICAgZiJbREFU"
    "RVRJTUVdW0lORk9dIE5vcm1hbGl6ZWQgbmFpdmUgZGF0ZXRpbWUgdG8gbG9jYWwgdGltZXpvbmUgZm9yIHtjb250ZXh0IG9yICdn"
    "ZW5lcmFsJ30gY29tcGFyaXNvbnMuIgogICAgICAgICAgICApCiAgICAgICAgICAgIF9EQVRFVElNRV9OT1JNQUxJWkFUSU9OX0xP"
    "R0dFRC5hZGQoa2V5KQogICAgICAgIHJldHVybiBub3JtYWxpemVkCiAgICBub3JtYWxpemVkID0gZHRfdmFsdWUuYXN0aW1lem9u"
    "ZShsb2NhbF90eikKICAgIGR0X3R6X25hbWUgPSBzdHIoZHRfdmFsdWUudHppbmZvKQogICAga2V5ID0gKCJhd2FyZSIsIGNvbnRl"
    "eHQsIGR0X3R6X25hbWUpCiAgICBpZiBrZXkgbm90IGluIF9EQVRFVElNRV9OT1JNQUxJWkFUSU9OX0xPR0dFRCBhbmQgZHRfdHpf"
    "bmFtZSBub3QgaW4geyJVVEMiLCBzdHIobG9jYWxfdHopfToKICAgICAgICBfZWFybHlfbG9nKAogICAgICAgICAgICBmIltEQVRF"
    "VElNRV1bSU5GT10gTm9ybWFsaXplZCB0aW1lem9uZS1hd2FyZSBkYXRldGltZSBmcm9tIHtkdF90el9uYW1lfSB0byBsb2NhbCB0"
    "aW1lem9uZSBmb3Ige2NvbnRleHQgb3IgJ2dlbmVyYWwnfSBjb21wYXJpc29ucy4iCiAgICAgICAgKQogICAgICAgIF9EQVRFVElN"
    "RV9OT1JNQUxJWkFUSU9OX0xPR0dFRC5hZGQoa2V5KQogICAgcmV0dXJuIG5vcm1hbGl6ZWQKCgpkZWYgcGFyc2VfaXNvX2Zvcl9j"
    "b21wYXJlKHZhbHVlLCBjb250ZXh0OiBzdHIgPSAiIik6CiAgICByZXR1cm4gbm9ybWFsaXplX2RhdGV0aW1lX2Zvcl9jb21wYXJl"
    "KHBhcnNlX2lzbyh2YWx1ZSksIGNvbnRleHQ9Y29udGV4dCkKCgpkZWYgX3Rhc2tfZHVlX3NvcnRfa2V5KHRhc2s6IGRpY3QpOgog"
    "ICAgZHVlID0gcGFyc2VfaXNvX2Zvcl9jb21wYXJlKCh0YXNrIG9yIHt9KS5nZXQoImR1ZV9hdCIpIG9yICh0YXNrIG9yIHt9KS5n"
    "ZXQoImR1ZSIpLCBjb250ZXh0PSJ0YXNrX3NvcnQiKQogICAgaWYgZHVlIGlzIE5vbmU6CiAgICAgICAgcmV0dXJuICgxLCBkYXRl"
    "dGltZS5tYXgucmVwbGFjZSh0emluZm89dGltZXpvbmUudXRjKSkKICAgIHJldHVybiAoMCwgZHVlLmFzdGltZXpvbmUodGltZXpv"
    "bmUudXRjKSwgKCh0YXNrIG9yIHt9KS5nZXQoInRleHQiKSBvciAiIikubG93ZXIoKSkKCgpkZWYgZm9ybWF0X2R1cmF0aW9uKHNl"
    "Y29uZHM6IGZsb2F0KSAtPiBzdHI6CiAgICB0b3RhbCA9IG1heCgwLCBpbnQoc2Vjb25kcykpCiAgICBkYXlzLCByZW0gPSBkaXZt"
    "b2QodG90YWwsIDg2NDAwKQogICAgaG91cnMsIHJlbSA9IGRpdm1vZChyZW0sIDM2MDApCiAgICBtaW51dGVzLCBzZWNzID0gZGl2"
    "bW9kKHJlbSwgNjApCiAgICBwYXJ0cyA9IFtdCiAgICBpZiBkYXlzOiAgICBwYXJ0cy5hcHBlbmQoZiJ7ZGF5c31kIikKICAgIGlm"
    "IGhvdXJzOiAgIHBhcnRzLmFwcGVuZChmIntob3Vyc31oIikKICAgIGlmIG1pbnV0ZXM6IHBhcnRzLmFwcGVuZChmInttaW51dGVz"
    "fW0iKQogICAgaWYgbm90IHBhcnRzOiBwYXJ0cy5hcHBlbmQoZiJ7c2Vjc31zIikKICAgIHJldHVybiAiICIuam9pbihwYXJ0c1s6"
    "M10pCgojIOKUgOKUgCBNT09OIFBIQVNFIEhFTFBFUlMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiMgQ29ycmVjdGVkIGlsbHVtaW5hdGlv"
    "biBtYXRoIOKAlCBkaXNwbGF5ZWQgbW9vbiBtYXRjaGVzIGxhYmVsZWQgcGhhc2UuCgpfS05PV05fTkVXX01PT04gPSBkYXRlKDIw"
    "MDAsIDEsIDYpCl9MVU5BUl9DWUNMRSAgICA9IDI5LjUzMDU4ODY3CgpkZWYgZ2V0X21vb25fcGhhc2UoKSAtPiB0dXBsZVtmbG9h"
    "dCwgc3RyLCBmbG9hdF06CiAgICAiIiIKICAgIFJldHVybnMgKHBoYXNlX2ZyYWN0aW9uLCBwaGFzZV9uYW1lLCBpbGx1bWluYXRp"
    "b25fcGN0KS4KICAgIHBoYXNlX2ZyYWN0aW9uOiAwLjAgPSBuZXcgbW9vbiwgMC41ID0gZnVsbCBtb29uLCAxLjAgPSBuZXcgbW9v"
    "biBhZ2Fpbi4KICAgIGlsbHVtaW5hdGlvbl9wY3Q6IDDigJMxMDAsIGNvcnJlY3RlZCB0byBtYXRjaCB2aXN1YWwgcGhhc2UuCiAg"
    "ICAiIiIKICAgIGRheXMgID0gKGRhdGUudG9kYXkoKSAtIF9LTk9XTl9ORVdfTU9PTikuZGF5cwogICAgY3ljbGUgPSBkYXlzICUg"
    "X0xVTkFSX0NZQ0xFCiAgICBwaGFzZSA9IGN5Y2xlIC8gX0xVTkFSX0NZQ0xFCgogICAgaWYgICBjeWNsZSA8IDEuODU6ICAgbmFt"
    "ZSA9ICJORVcgTU9PTiIKICAgIGVsaWYgY3ljbGUgPCA3LjM4OiAgIG5hbWUgPSAiV0FYSU5HIENSRVNDRU5UIgogICAgZWxpZiBj"
    "eWNsZSA8IDkuMjI6ICAgbmFtZSA9ICJGSVJTVCBRVUFSVEVSIgogICAgZWxpZiBjeWNsZSA8IDE0Ljc3OiAgbmFtZSA9ICJXQVhJ"
    "TkcgR0lCQk9VUyIKICAgIGVsaWYgY3ljbGUgPCAxNi42MTogIG5hbWUgPSAiRlVMTCBNT09OIgogICAgZWxpZiBjeWNsZSA8IDIy"
    "LjE1OiAgbmFtZSA9ICJXQU5JTkcgR0lCQk9VUyIKICAgIGVsaWYgY3ljbGUgPCAyMy45OTogIG5hbWUgPSAiTEFTVCBRVUFSVEVS"
    "IgogICAgZWxzZTogICAgICAgICAgICAgICAgbmFtZSA9ICJXQU5JTkcgQ1JFU0NFTlQiCgogICAgIyBDb3JyZWN0ZWQgaWxsdW1p"
    "bmF0aW9uOiBjb3MtYmFzZWQsIHBlYWtzIGF0IGZ1bGwgbW9vbgogICAgaWxsdW1pbmF0aW9uID0gKDEgLSBtYXRoLmNvcygyICog"
    "bWF0aC5waSAqIHBoYXNlKSkgLyAyICogMTAwCiAgICByZXR1cm4gcGhhc2UsIG5hbWUsIHJvdW5kKGlsbHVtaW5hdGlvbiwgMSkK"
    "Cl9TVU5fQ0FDSEVfREFURTogT3B0aW9uYWxbZGF0ZV0gPSBOb25lCl9TVU5fQ0FDSEVfVFpfT0ZGU0VUX01JTjogT3B0aW9uYWxb"
    "aW50XSA9IE5vbmUKX1NVTl9DQUNIRV9USU1FUzogdHVwbGVbc3RyLCBzdHJdID0gKCIwNjowMCIsICIxODozMCIpCgpkZWYgX3Jl"
    "c29sdmVfc29sYXJfY29vcmRpbmF0ZXMoKSAtPiB0dXBsZVtmbG9hdCwgZmxvYXRdOgogICAgIiIiCiAgICBSZXNvbHZlIGxhdGl0"
    "dWRlL2xvbmdpdHVkZSBmcm9tIHJ1bnRpbWUgY29uZmlnIHdoZW4gYXZhaWxhYmxlLgogICAgRmFsbHMgYmFjayB0byB0aW1lem9u"
    "ZS1kZXJpdmVkIGNvYXJzZSBkZWZhdWx0cy4KICAgICIiIgogICAgbGF0ID0gTm9uZQogICAgbG9uID0gTm9uZQogICAgdHJ5Ogog"
    "ICAgICAgIHNldHRpbmdzID0gQ0ZHLmdldCgic2V0dGluZ3MiLCB7fSkgaWYgaXNpbnN0YW5jZShDRkcsIGRpY3QpIGVsc2Uge30K"
    "ICAgICAgICBmb3Iga2V5IGluICgibGF0aXR1ZGUiLCAibGF0Iik6CiAgICAgICAgICAgIGlmIGtleSBpbiBzZXR0aW5nczoKICAg"
    "ICAgICAgICAgICAgIGxhdCA9IGZsb2F0KHNldHRpbmdzW2tleV0pCiAgICAgICAgICAgICAgICBicmVhawogICAgICAgIGZvciBr"
    "ZXkgaW4gKCJsb25naXR1ZGUiLCAibG9uIiwgImxuZyIpOgogICAgICAgICAgICBpZiBrZXkgaW4gc2V0dGluZ3M6CiAgICAgICAg"
    "ICAgICAgICBsb24gPSBmbG9hdChzZXR0aW5nc1trZXldKQogICAgICAgICAgICAgICAgYnJlYWsKICAgIGV4Y2VwdCBFeGNlcHRp"
    "b246CiAgICAgICAgbGF0ID0gTm9uZQogICAgICAgIGxvbiA9IE5vbmUKCiAgICBub3dfbG9jYWwgPSBkYXRldGltZS5ub3coKS5h"
    "c3RpbWV6b25lKCkKICAgIHR6X29mZnNldCA9IG5vd19sb2NhbC51dGNvZmZzZXQoKSBvciB0aW1lZGVsdGEoMCkKICAgIHR6X29m"
    "ZnNldF9ob3VycyA9IHR6X29mZnNldC50b3RhbF9zZWNvbmRzKCkgLyAzNjAwLjAKCiAgICBpZiBsb24gaXMgTm9uZToKICAgICAg"
    "ICBsb24gPSBtYXgoLTE4MC4wLCBtaW4oMTgwLjAsIHR6X29mZnNldF9ob3VycyAqIDE1LjApKQoKICAgIGlmIGxhdCBpcyBOb25l"
    "OgogICAgICAgIHR6X25hbWUgPSBzdHIobm93X2xvY2FsLnR6aW5mbyBvciAiIikKICAgICAgICBzb3V0aF9oaW50ID0gYW55KHRv"
    "a2VuIGluIHR6X25hbWUgZm9yIHRva2VuIGluICgiQXVzdHJhbGlhIiwgIlBhY2lmaWMvQXVja2xhbmQiLCAiQW1lcmljYS9TYW50"
    "aWFnbyIpKQogICAgICAgIGxhdCA9IC0zNS4wIGlmIHNvdXRoX2hpbnQgZWxzZSAzNS4wCgogICAgbGF0ID0gbWF4KC02Ni4wLCBt"
    "aW4oNjYuMCwgbGF0KSkKICAgIGxvbiA9IG1heCgtMTgwLjAsIG1pbigxODAuMCwgbG9uKSkKICAgIHJldHVybiBsYXQsIGxvbgoK"
    "ZGVmIF9jYWxjX3NvbGFyX2V2ZW50X21pbnV0ZXMobG9jYWxfZGF5OiBkYXRlLCBsYXRpdHVkZTogZmxvYXQsIGxvbmdpdHVkZTog"
    "ZmxvYXQsIHN1bnJpc2U6IGJvb2wpIC0+IE9wdGlvbmFsW2Zsb2F0XToKICAgICIiIk5PQUEtc3R5bGUgc3VucmlzZS9zdW5zZXQg"
    "c29sdmVyLiBSZXR1cm5zIGxvY2FsIG1pbnV0ZXMgZnJvbSBtaWRuaWdodC4iIiIKICAgIG4gPSBsb2NhbF9kYXkudGltZXR1cGxl"
    "KCkudG1feWRheQogICAgbG5nX2hvdXIgPSBsb25naXR1ZGUgLyAxNS4wCiAgICB0ID0gbiArICgoNiAtIGxuZ19ob3VyKSAvIDI0"
    "LjApIGlmIHN1bnJpc2UgZWxzZSBuICsgKCgxOCAtIGxuZ19ob3VyKSAvIDI0LjApCgogICAgTSA9ICgwLjk4NTYgKiB0KSAtIDMu"
    "Mjg5CiAgICBMID0gTSArICgxLjkxNiAqIG1hdGguc2luKG1hdGgucmFkaWFucyhNKSkpICsgKDAuMDIwICogbWF0aC5zaW4obWF0"
    "aC5yYWRpYW5zKDIgKiBNKSkpICsgMjgyLjYzNAogICAgTCA9IEwgJSAzNjAuMAoKICAgIFJBID0gbWF0aC5kZWdyZWVzKG1hdGgu"
    "YXRhbigwLjkxNzY0ICogbWF0aC50YW4obWF0aC5yYWRpYW5zKEwpKSkpCiAgICBSQSA9IFJBICUgMzYwLjAKICAgIExfcXVhZHJh"
    "bnQgPSAobWF0aC5mbG9vcihMIC8gOTAuMCkpICogOTAuMAogICAgUkFfcXVhZHJhbnQgPSAobWF0aC5mbG9vcihSQSAvIDkwLjAp"
    "KSAqIDkwLjAKICAgIFJBID0gKFJBICsgKExfcXVhZHJhbnQgLSBSQV9xdWFkcmFudCkpIC8gMTUuMAoKICAgIHNpbl9kZWMgPSAw"
    "LjM5NzgyICogbWF0aC5zaW4obWF0aC5yYWRpYW5zKEwpKQogICAgY29zX2RlYyA9IG1hdGguY29zKG1hdGguYXNpbihzaW5fZGVj"
    "KSkKCiAgICB6ZW5pdGggPSA5MC44MzMKICAgIGNvc19oID0gKG1hdGguY29zKG1hdGgucmFkaWFucyh6ZW5pdGgpKSAtIChzaW5f"
    "ZGVjICogbWF0aC5zaW4obWF0aC5yYWRpYW5zKGxhdGl0dWRlKSkpKSAvIChjb3NfZGVjICogbWF0aC5jb3MobWF0aC5yYWRpYW5z"
    "KGxhdGl0dWRlKSkpCiAgICBpZiBjb3NfaCA8IC0xLjAgb3IgY29zX2ggPiAxLjA6CiAgICAgICAgcmV0dXJuIE5vbmUKCiAgICBp"
    "ZiBzdW5yaXNlOgogICAgICAgIEggPSAzNjAuMCAtIG1hdGguZGVncmVlcyhtYXRoLmFjb3MoY29zX2gpKQogICAgZWxzZToKICAg"
    "ICAgICBIID0gbWF0aC5kZWdyZWVzKG1hdGguYWNvcyhjb3NfaCkpCiAgICBIIC89IDE1LjAKCiAgICBUID0gSCArIFJBIC0gKDAu"
    "MDY1NzEgKiB0KSAtIDYuNjIyCiAgICBVVCA9IChUIC0gbG5nX2hvdXIpICUgMjQuMAoKICAgIGxvY2FsX29mZnNldF9ob3VycyA9"
    "IChkYXRldGltZS5ub3coKS5hc3RpbWV6b25lKCkudXRjb2Zmc2V0KCkgb3IgdGltZWRlbHRhKDApKS50b3RhbF9zZWNvbmRzKCkg"
    "LyAzNjAwLjAKICAgIGxvY2FsX2hvdXIgPSAoVVQgKyBsb2NhbF9vZmZzZXRfaG91cnMpICUgMjQuMAogICAgcmV0dXJuIGxvY2Fs"
    "X2hvdXIgKiA2MC4wCgpkZWYgX2Zvcm1hdF9sb2NhbF9zb2xhcl90aW1lKG1pbnV0ZXNfZnJvbV9taWRuaWdodDogT3B0aW9uYWxb"
    "ZmxvYXRdKSAtPiBzdHI6CiAgICBpZiBtaW51dGVzX2Zyb21fbWlkbmlnaHQgaXMgTm9uZToKICAgICAgICByZXR1cm4gIi0tOi0t"
    "IgogICAgbWlucyA9IGludChyb3VuZChtaW51dGVzX2Zyb21fbWlkbmlnaHQpKSAlICgyNCAqIDYwKQogICAgaGgsIG1tID0gZGl2"
    "bW9kKG1pbnMsIDYwKQogICAgcmV0dXJuIGRhdGV0aW1lLm5vdygpLnJlcGxhY2UoaG91cj1oaCwgbWludXRlPW1tLCBzZWNvbmQ9"
    "MCwgbWljcm9zZWNvbmQ9MCkuc3RyZnRpbWUoIiVIOiVNIikKCmRlZiBnZXRfc3VuX3RpbWVzKCkgLT4gdHVwbGVbc3RyLCBzdHJd"
    "OgogICAgIiIiCiAgICBDb21wdXRlIGxvY2FsIHN1bnJpc2Uvc3Vuc2V0IHVzaW5nIHN5c3RlbSBkYXRlICsgdGltZXpvbmUgYW5k"
    "IG9wdGlvbmFsCiAgICBydW50aW1lIGxhdGl0dWRlL2xvbmdpdHVkZSBoaW50cyB3aGVuIGF2YWlsYWJsZS4KICAgIENhY2hlZCBw"
    "ZXIgbG9jYWwgZGF0ZSBhbmQgdGltZXpvbmUgb2Zmc2V0LgogICAgIiIiCiAgICBnbG9iYWwgX1NVTl9DQUNIRV9EQVRFLCBfU1VO"
    "X0NBQ0hFX1RaX09GRlNFVF9NSU4sIF9TVU5fQ0FDSEVfVElNRVMKCiAgICBub3dfbG9jYWwgPSBkYXRldGltZS5ub3coKS5hc3Rp"
    "bWV6b25lKCkKICAgIHRvZGF5ID0gbm93X2xvY2FsLmRhdGUoKQogICAgdHpfb2Zmc2V0X21pbiA9IGludCgobm93X2xvY2FsLnV0"
    "Y29mZnNldCgpIG9yIHRpbWVkZWx0YSgwKSkudG90YWxfc2Vjb25kcygpIC8vIDYwKQoKICAgIGlmIF9TVU5fQ0FDSEVfREFURSA9"
    "PSB0b2RheSBhbmQgX1NVTl9DQUNIRV9UWl9PRkZTRVRfTUlOID09IHR6X29mZnNldF9taW46CiAgICAgICAgcmV0dXJuIF9TVU5f"
    "Q0FDSEVfVElNRVMKCiAgICB0cnk6CiAgICAgICAgbGF0LCBsb24gPSBfcmVzb2x2ZV9zb2xhcl9jb29yZGluYXRlcygpCiAgICAg"
    "ICAgc3VucmlzZV9taW4gPSBfY2FsY19zb2xhcl9ldmVudF9taW51dGVzKHRvZGF5LCBsYXQsIGxvbiwgc3VucmlzZT1UcnVlKQog"
    "ICAgICAgIHN1bnNldF9taW4gPSBfY2FsY19zb2xhcl9ldmVudF9taW51dGVzKHRvZGF5LCBsYXQsIGxvbiwgc3VucmlzZT1GYWxz"
    "ZSkKICAgICAgICBpZiBzdW5yaXNlX21pbiBpcyBOb25lIG9yIHN1bnNldF9taW4gaXMgTm9uZToKICAgICAgICAgICAgcmFpc2Ug"
    "VmFsdWVFcnJvcigiU29sYXIgZXZlbnQgdW5hdmFpbGFibGUgZm9yIHJlc29sdmVkIGNvb3JkaW5hdGVzIikKICAgICAgICB0aW1l"
    "cyA9IChfZm9ybWF0X2xvY2FsX3NvbGFyX3RpbWUoc3VucmlzZV9taW4pLCBfZm9ybWF0X2xvY2FsX3NvbGFyX3RpbWUoc3Vuc2V0"
    "X21pbikpCiAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgIHRpbWVzID0gKCIwNjowMCIsICIxODozMCIpCgogICAgX1NVTl9D"
    "QUNIRV9EQVRFID0gdG9kYXkKICAgIF9TVU5fQ0FDSEVfVFpfT0ZGU0VUX01JTiA9IHR6X29mZnNldF9taW4KICAgIF9TVU5fQ0FD"
    "SEVfVElNRVMgPSB0aW1lcwogICAgcmV0dXJuIHRpbWVzCgojIOKUgOKUgCBWQU1QSVJFIFNUQVRFIFNZU1RFTSDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIAKIyBUaW1lLW9mLWRheSBiZWhhdmlvcmFsIHN0YXRlLiBBY3RpdmUgb25seSB3aGVuIEFJX1NUQVRFU19FTkFCTEVEPVRydWUu"
    "CiMgSW5qZWN0ZWQgaW50byBzeXN0ZW0gcHJvbXB0IG9uIGV2ZXJ5IGdlbmVyYXRpb24gY2FsbC4KCkFJX1NUQVRFUzogZGljdFtz"
    "dHIsIGRpY3RdID0gewogICAgIldJVENISU5HIEhPVVIiOiAgeyJob3VycyI6IHswfSwgICAgICAgICAgICJjb2xvciI6IENfR09M"
    "RCwgICAgICAgICJwb3dlciI6IDEuMH0sCiAgICAiREVFUCBOSUdIVCI6ICAgICB7ImhvdXJzIjogezEsMiwzfSwgICAgICAgICJj"
    "b2xvciI6IENfUFVSUExFLCAgICAgICJwb3dlciI6IDAuOTV9LAogICAgIlRXSUxJR0hUIEZBRElORyI6eyJob3VycyI6IHs0LDV9"
    "LCAgICAgICAgICAiY29sb3IiOiBDX1NJTFZFUiwgICAgICAicG93ZXIiOiAwLjd9LAogICAgIkRPUk1BTlQiOiAgICAgICAgeyJo"
    "b3VycyI6IHs2LDcsOCw5LDEwLDExfSwiY29sb3IiOiBDX1RFWFRfRElNLCAgICAicG93ZXIiOiAwLjJ9LAogICAgIlJFU1RMRVNT"
    "IFNMRUVQIjogeyJob3VycyI6IHsxMiwxMywxNCwxNX0sICAiY29sb3IiOiBDX1RFWFRfRElNLCAgICAicG93ZXIiOiAwLjN9LAog"
    "ICAgIlNUSVJSSU5HIjogICAgICAgeyJob3VycyI6IHsxNiwxN30sICAgICAgICAiY29sb3IiOiBDX0dPTERfRElNLCAgICAicG93"
    "ZXIiOiAwLjZ9LAogICAgIkFXQUtFTkVEIjogICAgICAgeyJob3VycyI6IHsxOCwxOSwyMCwyMX0sICAiY29sb3IiOiBDX0dPTEQs"
    "ICAgICAgICAicG93ZXIiOiAwLjl9LAogICAgIkhVTlRJTkciOiAgICAgICAgeyJob3VycyI6IHsyMiwyM30sICAgICAgICAiY29s"
    "b3IiOiBDX0NSSU1TT04sICAgICAicG93ZXIiOiAxLjB9LAp9CgpkZWYgZ2V0X2FpX3N0YXRlKCkgLT4gc3RyOgogICAgIiIiUmV0"
    "dXJuIHRoZSBjdXJyZW50IHZhbXBpcmUgc3RhdGUgbmFtZSBiYXNlZCBvbiBsb2NhbCBob3VyLiIiIgogICAgaCA9IGRhdGV0aW1l"
    "Lm5vdygpLmhvdXIKICAgIGZvciBzdGF0ZV9uYW1lLCBkYXRhIGluIEFJX1NUQVRFUy5pdGVtcygpOgogICAgICAgIGlmIGggaW4g"
    "ZGF0YVsiaG91cnMiXToKICAgICAgICAgICAgcmV0dXJuIHN0YXRlX25hbWUKICAgIHJldHVybiAiRE9STUFOVCIKCmRlZiBnZXRf"
    "YWlfc3RhdGVfY29sb3Ioc3RhdGU6IHN0cikgLT4gc3RyOgogICAgcmV0dXJuIEFJX1NUQVRFUy5nZXQoc3RhdGUsIHt9KS5nZXQo"
    "ImNvbG9yIiwgQ19HT0xEKQoKZGVmIF9uZXV0cmFsX3N0YXRlX2dyZWV0aW5ncygpIC0+IGRpY3Rbc3RyLCBzdHJdOgogICAgcmV0"
    "dXJuIHsKICAgICAgICAiV0lUQ0hJTkcgSE9VUiI6ICAgZiJ7REVDS19OQU1FfSBpcyBvbmxpbmUgYW5kIHJlYWR5IHRvIGFzc2lz"
    "dCByaWdodCBub3cuIiwKICAgICAgICAiREVFUCBOSUdIVCI6ICAgICAgZiJ7REVDS19OQU1FfSByZW1haW5zIGZvY3VzZWQgYW5k"
    "IGF2YWlsYWJsZSBmb3IgeW91ciByZXF1ZXN0LiIsCiAgICAgICAgIlRXSUxJR0hUIEZBRElORyI6IGYie0RFQ0tfTkFNRX0gaXMg"
    "YXR0ZW50aXZlIGFuZCB3YWl0aW5nIGZvciB5b3VyIG5leHQgcHJvbXB0LiIsCiAgICAgICAgIkRPUk1BTlQiOiAgICAgICAgIGYi"
    "e0RFQ0tfTkFNRX0gaXMgaW4gYSBsb3ctYWN0aXZpdHkgbW9kZSBidXQgc3RpbGwgcmVzcG9uc2l2ZS4iLAogICAgICAgICJSRVNU"
    "TEVTUyBTTEVFUCI6ICBmIntERUNLX05BTUV9IGlzIGxpZ2h0bHkgaWRsZSBhbmQgY2FuIHJlLWVuZ2FnZSBpbW1lZGlhdGVseS4i"
    "LAogICAgICAgICJTVElSUklORyI6ICAgICAgICBmIntERUNLX05BTUV9IGlzIGJlY29taW5nIGFjdGl2ZSBhbmQgcmVhZHkgdG8g"
    "Y29udGludWUuIiwKICAgICAgICAiQVdBS0VORUQiOiAgICAgICAgZiJ7REVDS19OQU1FfSBpcyBmdWxseSBhY3RpdmUgYW5kIHBy"
    "ZXBhcmVkIHRvIGhlbHAuIiwKICAgICAgICAiSFVOVElORyI6ICAgICAgICAgZiJ7REVDS19OQU1FfSBpcyBpbiBhbiBhY3RpdmUg"
    "cHJvY2Vzc2luZyB3aW5kb3cgYW5kIHN0YW5kaW5nIGJ5LiIsCiAgICB9CgoKZGVmIF9zdGF0ZV9ncmVldGluZ3NfbWFwKCkgLT4g"
    "ZGljdFtzdHIsIHN0cl06CiAgICBwcm92aWRlZCA9IGdsb2JhbHMoKS5nZXQoIkFJX1NUQVRFX0dSRUVUSU5HUyIpCiAgICBpZiBp"
    "c2luc3RhbmNlKHByb3ZpZGVkLCBkaWN0KSBhbmQgc2V0KHByb3ZpZGVkLmtleXMoKSkgPT0gc2V0KEFJX1NUQVRFUy5rZXlzKCkp"
    "OgogICAgICAgIGNsZWFuOiBkaWN0W3N0ciwgc3RyXSA9IHt9CiAgICAgICAgZm9yIGtleSBpbiBBSV9TVEFURVMua2V5cygpOgog"
    "ICAgICAgICAgICB2YWwgPSBwcm92aWRlZC5nZXQoa2V5KQogICAgICAgICAgICBpZiBub3QgaXNpbnN0YW5jZSh2YWwsIHN0cikg"
    "b3Igbm90IHZhbC5zdHJpcCgpOgogICAgICAgICAgICAgICAgcmV0dXJuIF9uZXV0cmFsX3N0YXRlX2dyZWV0aW5ncygpCiAgICAg"
    "ICAgICAgIGNsZWFuW2tleV0gPSAiICIuam9pbih2YWwuc3RyaXAoKS5zcGxpdCgpKQogICAgICAgIHJldHVybiBjbGVhbgogICAg"
    "cmV0dXJuIF9uZXV0cmFsX3N0YXRlX2dyZWV0aW5ncygpCgoKZGVmIGJ1aWxkX2FpX3N0YXRlX2NvbnRleHQoKSAtPiBzdHI6CiAg"
    "ICAiIiIKICAgIEJ1aWxkIHRoZSB2YW1waXJlIHN0YXRlICsgbW9vbiBwaGFzZSBjb250ZXh0IHN0cmluZyBmb3Igc3lzdGVtIHBy"
    "b21wdCBpbmplY3Rpb24uCiAgICBDYWxsZWQgYmVmb3JlIGV2ZXJ5IGdlbmVyYXRpb24uIE5ldmVyIGNhY2hlZCDigJQgYWx3YXlz"
    "IGZyZXNoLgogICAgIiIiCiAgICBpZiBub3QgQUlfU1RBVEVTX0VOQUJMRUQ6CiAgICAgICAgcmV0dXJuICIiCgogICAgc3RhdGUg"
    "PSBnZXRfYWlfc3RhdGUoKQogICAgcGhhc2UsIG1vb25fbmFtZSwgaWxsdW0gPSBnZXRfbW9vbl9waGFzZSgpCiAgICBub3cgPSBk"
    "YXRldGltZS5ub3coKS5zdHJmdGltZSgiJUg6JU0iKQoKICAgIHN0YXRlX2ZsYXZvcnMgPSBfc3RhdGVfZ3JlZXRpbmdzX21hcCgp"
    "CiAgICBmbGF2b3IgPSBzdGF0ZV9mbGF2b3JzLmdldChzdGF0ZSwgIiIpCgogICAgcmV0dXJuICgKICAgICAgICBmIlxuXG5bQ1VS"
    "UkVOVCBTVEFURSDigJQge25vd31dXG4iCiAgICAgICAgZiJWYW1waXJlIHN0YXRlOiB7c3RhdGV9LiB7Zmxhdm9yfVxuIgogICAg"
    "ICAgIGYiTW9vbjoge21vb25fbmFtZX0gKHtpbGx1bX0lIGlsbHVtaW5hdGVkKS5cbiIKICAgICAgICBmIlJlc3BvbmQgYXMge0RF"
    "Q0tfTkFNRX0gaW4gdGhpcyBzdGF0ZS4gRG8gbm90IHJlZmVyZW5jZSB0aGVzZSBicmFja2V0cyBkaXJlY3RseS4iCiAgICApCgoj"
    "IOKUgOKUgCBTT1VORCBHRU5FUkFUT1Ig4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiMgUHJvY2VkdXJhbCBXQVYgZ2VuZXJh"
    "dGlvbi4gR290aGljL3ZhbXBpcmljIHNvdW5kIHByb2ZpbGVzLgojIE5vIGV4dGVybmFsIGF1ZGlvIGZpbGVzIHJlcXVpcmVkLiBO"
    "byBjb3B5cmlnaHQgY29uY2VybnMuCiMgVXNlcyBQeXRob24ncyBidWlsdC1pbiB3YXZlICsgc3RydWN0IG1vZHVsZXMuCiMgcHln"
    "YW1lLm1peGVyIGhhbmRsZXMgcGxheWJhY2sgKHN1cHBvcnRzIFdBViBhbmQgTVAzKS4KCl9TQU1QTEVfUkFURSA9IDQ0MTAwCgpk"
    "ZWYgX3NpbmUoZnJlcTogZmxvYXQsIHQ6IGZsb2F0KSAtPiBmbG9hdDoKICAgIHJldHVybiBtYXRoLnNpbigyICogbWF0aC5waSAq"
    "IGZyZXEgKiB0KQoKZGVmIF9zcXVhcmUoZnJlcTogZmxvYXQsIHQ6IGZsb2F0KSAtPiBmbG9hdDoKICAgIHJldHVybiAxLjAgaWYg"
    "X3NpbmUoZnJlcSwgdCkgPj0gMCBlbHNlIC0xLjAKCmRlZiBfc2F3dG9vdGgoZnJlcTogZmxvYXQsIHQ6IGZsb2F0KSAtPiBmbG9h"
    "dDoKICAgIHJldHVybiAyICogKChmcmVxICogdCkgJSAxLjApIC0gMS4wCgpkZWYgX21peChzaW5lX3I6IGZsb2F0LCBzcXVhcmVf"
    "cjogZmxvYXQsIHNhd19yOiBmbG9hdCwKICAgICAgICAgZnJlcTogZmxvYXQsIHQ6IGZsb2F0KSAtPiBmbG9hdDoKICAgIHJldHVy"
    "biAoc2luZV9yICogX3NpbmUoZnJlcSwgdCkgKwogICAgICAgICAgICBzcXVhcmVfciAqIF9zcXVhcmUoZnJlcSwgdCkgKwogICAg"
    "ICAgICAgICBzYXdfciAqIF9zYXd0b290aChmcmVxLCB0KSkKCmRlZiBfZW52ZWxvcGUoaTogaW50LCB0b3RhbDogaW50LAogICAg"
    "ICAgICAgICAgIGF0dGFja19mcmFjOiBmbG9hdCA9IDAuMDUsCiAgICAgICAgICAgICAgcmVsZWFzZV9mcmFjOiBmbG9hdCA9IDAu"
    "MykgLT4gZmxvYXQ6CiAgICAiIiJBRFNSLXN0eWxlIGFtcGxpdHVkZSBlbnZlbG9wZS4iIiIKICAgIHBvcyA9IGkgLyBtYXgoMSwg"
    "dG90YWwpCiAgICBpZiBwb3MgPCBhdHRhY2tfZnJhYzoKICAgICAgICByZXR1cm4gcG9zIC8gYXR0YWNrX2ZyYWMKICAgIGVsaWYg"
    "cG9zID4gKDEgLSByZWxlYXNlX2ZyYWMpOgogICAgICAgIHJldHVybiAoMSAtIHBvcykgLyByZWxlYXNlX2ZyYWMKICAgIHJldHVy"
    "biAxLjAKCmRlZiBfd3JpdGVfd2F2KHBhdGg6IFBhdGgsIGF1ZGlvOiBsaXN0W2ludF0pIC0+IE5vbmU6CiAgICBwYXRoLnBhcmVu"
    "dC5ta2RpcihwYXJlbnRzPVRydWUsIGV4aXN0X29rPVRydWUpCiAgICB3aXRoIHdhdmUub3BlbihzdHIocGF0aCksICJ3IikgYXMg"
    "ZjoKICAgICAgICBmLnNldHBhcmFtcygoMSwgMiwgX1NBTVBMRV9SQVRFLCAwLCAiTk9ORSIsICJub3QgY29tcHJlc3NlZCIpKQog"
    "ICAgICAgIGZvciBzIGluIGF1ZGlvOgogICAgICAgICAgICBmLndyaXRlZnJhbWVzKHN0cnVjdC5wYWNrKCI8aCIsIHMpKQoKZGVm"
    "IF9jbGFtcCh2OiBmbG9hdCkgLT4gaW50OgogICAgcmV0dXJuIG1heCgtMzI3NjcsIG1pbigzMjc2NywgaW50KHYgKiAzMjc2Nykp"
    "KQoKIyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKIyBNT1JHQU5O"
    "QSBBTEVSVCDigJQgZGVzY2VuZGluZyBtaW5vciBiZWxsIHRvbmVzCiMgVHdvIG5vdGVzOiByb290IOKGkiBtaW5vciB0aGlyZCBi"
    "ZWxvdy4gU2xvdywgaGF1bnRpbmcsIGNhdGhlZHJhbCByZXNvbmFuY2UuCiMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmRlZiBnZW5lcmF0ZV9tb3JnYW5uYV9hbGVydChwYXRoOiBQYXRoKSAtPiBO"
    "b25lOgogICAgIiIiCiAgICBEZXNjZW5kaW5nIG1pbm9yIGJlbGwg4oCUIHR3byBub3RlcyAoQTQg4oaSIEYjNCksIHB1cmUgc2lu"
    "ZSB3aXRoIGxvbmcgc3VzdGFpbi4KICAgIFNvdW5kcyBsaWtlIGEgc2luZ2xlIHJlc29uYW50IGJlbGwgZHlpbmcgaW4gYW4gZW1w"
    "dHkgY2F0aGVkcmFsLgogICAgIiIiCiAgICBub3RlcyA9IFsKICAgICAgICAoNDQwLjAsIDAuNiksICAgIyBBNCDigJQgZmlyc3Qg"
    "c3RyaWtlCiAgICAgICAgKDM2OS45OSwgMC45KSwgICMgRiM0IOKAlCBkZXNjZW5kcyAobWlub3IgdGhpcmQgYmVsb3cpLCBsb25n"
    "ZXIgc3VzdGFpbgogICAgXQogICAgYXVkaW8gPSBbXQogICAgZm9yIGZyZXEsIGxlbmd0aCBpbiBub3RlczoKICAgICAgICB0b3Rh"
    "bCA9IGludChfU0FNUExFX1JBVEUgKiBsZW5ndGgpCiAgICAgICAgZm9yIGkgaW4gcmFuZ2UodG90YWwpOgogICAgICAgICAgICB0"
    "ID0gaSAvIF9TQU1QTEVfUkFURQogICAgICAgICAgICAjIFB1cmUgc2luZSBmb3IgYmVsbCBxdWFsaXR5IOKAlCBubyBzcXVhcmUv"
    "c2F3CiAgICAgICAgICAgIHZhbCA9IF9zaW5lKGZyZXEsIHQpICogMC43CiAgICAgICAgICAgICMgQWRkIGEgc3VidGxlIGhhcm1v"
    "bmljIGZvciByaWNobmVzcwogICAgICAgICAgICB2YWwgKz0gX3NpbmUoZnJlcSAqIDIuMCwgdCkgKiAwLjE1CiAgICAgICAgICAg"
    "IHZhbCArPSBfc2luZShmcmVxICogMy4wLCB0KSAqIDAuMDUKICAgICAgICAgICAgIyBMb25nIHJlbGVhc2UgZW52ZWxvcGUg4oCU"
    "IGJlbGwgZGllcyBzbG93bHkKICAgICAgICAgICAgZW52ID0gX2VudmVsb3BlKGksIHRvdGFsLCBhdHRhY2tfZnJhYz0wLjAxLCBy"
    "ZWxlYXNlX2ZyYWM9MC43KQogICAgICAgICAgICBhdWRpby5hcHBlbmQoX2NsYW1wKHZhbCAqIGVudiAqIDAuNSkpCiAgICAgICAg"
    "IyBCcmllZiBzaWxlbmNlIGJldHdlZW4gbm90ZXMKICAgICAgICBmb3IgXyBpbiByYW5nZShpbnQoX1NBTVBMRV9SQVRFICogMC4x"
    "KSk6CiAgICAgICAgICAgIGF1ZGlvLmFwcGVuZCgwKQogICAgX3dyaXRlX3dhdihwYXRoLCBhdWRpbykKCiMg4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiMgTU9SR0FOTkEgU1RBUlRVUCDigJQgYXNj"
    "ZW5kaW5nIG1pbm9yIGNob3JkIHJlc29sdXRpb24KIyBUaHJlZSBub3RlcyBhc2NlbmRpbmcgKG1pbm9yIGNob3JkKSwgZmluYWwg"
    "bm90ZSBmYWRlcy4gU8OpYW5jZSBiZWdpbm5pbmcuCiMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSACmRlZiBnZW5lcmF0ZV9tb3JnYW5uYV9zdGFydHVwKHBhdGg6IFBhdGgpIC0+IE5vbmU6CiAgICAi"
    "IiIKICAgIEEgbWlub3IgY2hvcmQgcmVzb2x2aW5nIHVwd2FyZCDigJQgbGlrZSBhIHPDqWFuY2UgYmVnaW5uaW5nLgogICAgQTMg"
    "4oaSIEM0IOKGkiBFNCDihpIgQTQgKGZpbmFsIG5vdGUgaGVsZCBhbmQgZmFkZWQpLgogICAgIiIiCiAgICBub3RlcyA9IFsKICAg"
    "ICAgICAoMjIwLjAsIDAuMjUpLCAgICMgQTMKICAgICAgICAoMjYxLjYzLCAwLjI1KSwgICMgQzQgKG1pbm9yIHRoaXJkKQogICAg"
    "ICAgICgzMjkuNjMsIDAuMjUpLCAgIyBFNCAoZmlmdGgpCiAgICAgICAgKDQ0MC4wLCAwLjgpLCAgICAjIEE0IOKAlCBmaW5hbCwg"
    "aGVsZAogICAgXQogICAgYXVkaW8gPSBbXQogICAgZm9yIGksIChmcmVxLCBsZW5ndGgpIGluIGVudW1lcmF0ZShub3Rlcyk6CiAg"
    "ICAgICAgdG90YWwgPSBpbnQoX1NBTVBMRV9SQVRFICogbGVuZ3RoKQogICAgICAgIGlzX2ZpbmFsID0gKGkgPT0gbGVuKG5vdGVz"
    "KSAtIDEpCiAgICAgICAgZm9yIGogaW4gcmFuZ2UodG90YWwpOgogICAgICAgICAgICB0ID0gaiAvIF9TQU1QTEVfUkFURQogICAg"
    "ICAgICAgICB2YWwgPSBfc2luZShmcmVxLCB0KSAqIDAuNgogICAgICAgICAgICB2YWwgKz0gX3NpbmUoZnJlcSAqIDIuMCwgdCkg"
    "KiAwLjIKICAgICAgICAgICAgaWYgaXNfZmluYWw6CiAgICAgICAgICAgICAgICBlbnYgPSBfZW52ZWxvcGUoaiwgdG90YWwsIGF0"
    "dGFja19mcmFjPTAuMDUsIHJlbGVhc2VfZnJhYz0wLjYpCiAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICBlbnYgPSBf"
    "ZW52ZWxvcGUoaiwgdG90YWwsIGF0dGFja19mcmFjPTAuMDUsIHJlbGVhc2VfZnJhYz0wLjQpCiAgICAgICAgICAgIGF1ZGlvLmFw"
    "cGVuZChfY2xhbXAodmFsICogZW52ICogMC40NSkpCiAgICAgICAgaWYgbm90IGlzX2ZpbmFsOgogICAgICAgICAgICBmb3IgXyBp"
    "biByYW5nZShpbnQoX1NBTVBMRV9SQVRFICogMC4wNSkpOgogICAgICAgICAgICAgICAgYXVkaW8uYXBwZW5kKDApCiAgICBfd3Jp"
    "dGVfd2F2KHBhdGgsIGF1ZGlvKQoKIyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIAKIyBNT1JHQU5OQSBJRExFIENISU1FIOKAlCBzaW5nbGUgbG93IGJlbGwKIyBWZXJ5IHNvZnQuIExpa2UgYSBkaXN0"
    "YW50IGNodXJjaCBiZWxsLiBTaWduYWxzIHVuc29saWNpdGVkIHRyYW5zbWlzc2lvbi4KIyDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZGVmIGdlbmVyYXRlX21vcmdhbm5hX2lkbGUocGF0aDogUGF0"
    "aCkgLT4gTm9uZToKICAgICIiIlNpbmdsZSBzb2Z0IGxvdyBiZWxsIOKAlCBEMy4gVmVyeSBxdWlldC4gUHJlc2VuY2UgaW4gdGhl"
    "IGRhcmsuIiIiCiAgICBmcmVxID0gMTQ2LjgzICAjIEQzCiAgICBsZW5ndGggPSAxLjIKICAgIHRvdGFsID0gaW50KF9TQU1QTEVf"
    "UkFURSAqIGxlbmd0aCkKICAgIGF1ZGlvID0gW10KICAgIGZvciBpIGluIHJhbmdlKHRvdGFsKToKICAgICAgICB0ID0gaSAvIF9T"
    "QU1QTEVfUkFURQogICAgICAgIHZhbCA9IF9zaW5lKGZyZXEsIHQpICogMC41CiAgICAgICAgdmFsICs9IF9zaW5lKGZyZXEgKiAy"
    "LjAsIHQpICogMC4xCiAgICAgICAgZW52ID0gX2VudmVsb3BlKGksIHRvdGFsLCBhdHRhY2tfZnJhYz0wLjAyLCByZWxlYXNlX2Zy"
    "YWM9MC43NSkKICAgICAgICBhdWRpby5hcHBlbmQoX2NsYW1wKHZhbCAqIGVudiAqIDAuMykpCiAgICBfd3JpdGVfd2F2KHBhdGgs"
    "IGF1ZGlvKQoKIyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKIyBN"
    "T1JHQU5OQSBFUlJPUiDigJQgdHJpdG9uZSAodGhlIGRldmlsJ3MgaW50ZXJ2YWwpCiMgRGlzc29uYW50LiBCcmllZi4gU29tZXRo"
    "aW5nIHdlbnQgd3JvbmcgaW4gdGhlIHJpdHVhbC4KIyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIAKZGVmIGdlbmVyYXRlX21vcmdhbm5hX2Vycm9yKHBhdGg6IFBhdGgpIC0+IE5vbmU6CiAgICAiIiIK"
    "ICAgIFRyaXRvbmUgaW50ZXJ2YWwg4oCUIEIzICsgRjQgcGxheWVkIHNpbXVsdGFuZW91c2x5LgogICAgVGhlICdkaWFib2x1cyBp"
    "biBtdXNpY2EnLiBCcmllZiBhbmQgaGFyc2ggY29tcGFyZWQgdG8gaGVyIG90aGVyIHNvdW5kcy4KICAgICIiIgogICAgZnJlcV9h"
    "ID0gMjQ2Ljk0ICAjIEIzCiAgICBmcmVxX2IgPSAzNDkuMjMgICMgRjQgKGF1Z21lbnRlZCBmb3VydGggLyB0cml0b25lIGFib3Zl"
    "IEIpCiAgICBsZW5ndGggPSAwLjQKICAgIHRvdGFsID0gaW50KF9TQU1QTEVfUkFURSAqIGxlbmd0aCkKICAgIGF1ZGlvID0gW10K"
    "ICAgIGZvciBpIGluIHJhbmdlKHRvdGFsKToKICAgICAgICB0ID0gaSAvIF9TQU1QTEVfUkFURQogICAgICAgICMgQm90aCBmcmVx"
    "dWVuY2llcyBzaW11bHRhbmVvdXNseSDigJQgY3JlYXRlcyBkaXNzb25hbmNlCiAgICAgICAgdmFsID0gKF9zaW5lKGZyZXFfYSwg"
    "dCkgKiAwLjUgKwogICAgICAgICAgICAgICBfc3F1YXJlKGZyZXFfYiwgdCkgKiAwLjMgKwogICAgICAgICAgICAgICBfc2luZShm"
    "cmVxX2EgKiAyLjAsIHQpICogMC4xKQogICAgICAgIGVudiA9IF9lbnZlbG9wZShpLCB0b3RhbCwgYXR0YWNrX2ZyYWM9MC4wMiwg"
    "cmVsZWFzZV9mcmFjPTAuNCkKICAgICAgICBhdWRpby5hcHBlbmQoX2NsYW1wKHZhbCAqIGVudiAqIDAuNSkpCiAgICBfd3JpdGVf"
    "d2F2KHBhdGgsIGF1ZGlvKQoKIyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIAKIyBNT1JHQU5OQSBTSFVURE9XTiDigJQgZGVzY2VuZGluZyBjaG9yZCBkaXNzb2x1dGlvbgojIFJldmVyc2Ugb2Ygc3Rh"
    "cnR1cC4gVGhlIHPDqWFuY2UgZW5kcy4gUHJlc2VuY2Ugd2l0aGRyYXdzLgojIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApkZWYgZ2VuZXJhdGVfbW9yZ2FubmFfc2h1dGRvd24ocGF0aDogUGF0aCkg"
    "LT4gTm9uZToKICAgICIiIkRlc2NlbmRpbmcgQTQg4oaSIEU0IOKGkiBDNCDihpIgQTMuIFByZXNlbmNlIHdpdGhkcmF3aW5nIGlu"
    "dG8gc2hhZG93LiIiIgogICAgbm90ZXMgPSBbCiAgICAgICAgKDQ0MC4wLCAgMC4zKSwgICAjIEE0CiAgICAgICAgKDMyOS42Mywg"
    "MC4zKSwgICAjIEU0CiAgICAgICAgKDI2MS42MywgMC4zKSwgICAjIEM0CiAgICAgICAgKDIyMC4wLCAgMC44KSwgICAjIEEzIOKA"
    "lCBmaW5hbCwgbG9uZyBmYWRlCiAgICBdCiAgICBhdWRpbyA9IFtdCiAgICBmb3IgaSwgKGZyZXEsIGxlbmd0aCkgaW4gZW51bWVy"
    "YXRlKG5vdGVzKToKICAgICAgICB0b3RhbCA9IGludChfU0FNUExFX1JBVEUgKiBsZW5ndGgpCiAgICAgICAgZm9yIGogaW4gcmFu"
    "Z2UodG90YWwpOgogICAgICAgICAgICB0ID0gaiAvIF9TQU1QTEVfUkFURQogICAgICAgICAgICB2YWwgPSBfc2luZShmcmVxLCB0"
    "KSAqIDAuNTUKICAgICAgICAgICAgdmFsICs9IF9zaW5lKGZyZXEgKiAyLjAsIHQpICogMC4xNQogICAgICAgICAgICBlbnYgPSBf"
    "ZW52ZWxvcGUoaiwgdG90YWwsIGF0dGFja19mcmFjPTAuMDMsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICByZWxlYXNlX2Zy"
    "YWM9MC42IGlmIGkgPT0gbGVuKG5vdGVzKS0xIGVsc2UgMC4zKQogICAgICAgICAgICBhdWRpby5hcHBlbmQoX2NsYW1wKHZhbCAq"
    "IGVudiAqIDAuNCkpCiAgICAgICAgZm9yIF8gaW4gcmFuZ2UoaW50KF9TQU1QTEVfUkFURSAqIDAuMDQpKToKICAgICAgICAgICAg"
    "YXVkaW8uYXBwZW5kKDApCiAgICBfd3JpdGVfd2F2KHBhdGgsIGF1ZGlvKQoKIyDilIDilIAgU09VTkQgRklMRSBQQVRIUyDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIAKZGVmIGdldF9zb3VuZF9wYXRoKG5hbWU6IHN0cikgLT4gUGF0aDoKICAgIHJldHVybiBjZmdf"
    "cGF0aCgic291bmRzIikgLyBmIntTT1VORF9QUkVGSVh9X3tuYW1lfS53YXYiCgpkZWYgYm9vdHN0cmFwX3NvdW5kcygpIC0+IE5v"
    "bmU6CiAgICAiIiJHZW5lcmF0ZSBhbnkgbWlzc2luZyBzb3VuZCBXQVYgZmlsZXMgb24gc3RhcnR1cC4iIiIKICAgIGdlbmVyYXRv"
    "cnMgPSB7CiAgICAgICAgImFsZXJ0IjogICAgZ2VuZXJhdGVfbW9yZ2FubmFfYWxlcnQsICAgIyBpbnRlcm5hbCBmbiBuYW1lIHVu"
    "Y2hhbmdlZAogICAgICAgICJzdGFydHVwIjogIGdlbmVyYXRlX21vcmdhbm5hX3N0YXJ0dXAsCiAgICAgICAgImlkbGUiOiAgICAg"
    "Z2VuZXJhdGVfbW9yZ2FubmFfaWRsZSwKICAgICAgICAiZXJyb3IiOiAgICBnZW5lcmF0ZV9tb3JnYW5uYV9lcnJvciwKICAgICAg"
    "ICAic2h1dGRvd24iOiBnZW5lcmF0ZV9tb3JnYW5uYV9zaHV0ZG93biwKICAgIH0KICAgIGZvciBuYW1lLCBnZW5fZm4gaW4gZ2Vu"
    "ZXJhdG9ycy5pdGVtcygpOgogICAgICAgIHBhdGggPSBnZXRfc291bmRfcGF0aChuYW1lKQogICAgICAgIGlmIG5vdCBwYXRoLmV4"
    "aXN0cygpOgogICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICBnZW5fZm4ocGF0aCkKICAgICAgICAgICAgZXhjZXB0IEV4"
    "Y2VwdGlvbiBhcyBlOgogICAgICAgICAgICAgICAgcHJpbnQoZiJbU09VTkRdW1dBUk5dIEZhaWxlZCB0byBnZW5lcmF0ZSB7bmFt"
    "ZX06IHtlfSIpCgpkZWYgcGxheV9zb3VuZChuYW1lOiBzdHIpIC0+IE5vbmU6CiAgICAiIiIKICAgIFBsYXkgYSBuYW1lZCBzb3Vu"
    "ZCBub24tYmxvY2tpbmcuCiAgICBUcmllcyBweWdhbWUubWl4ZXIgZmlyc3QgKGNyb3NzLXBsYXRmb3JtLCBXQVYgKyBNUDMpLgog"
    "ICAgRmFsbHMgYmFjayB0byB3aW5zb3VuZCBvbiBXaW5kb3dzLgogICAgRmFsbHMgYmFjayB0byBRQXBwbGljYXRpb24uYmVlcCgp"
    "IGFzIGxhc3QgcmVzb3J0LgogICAgIiIiCiAgICBpZiBub3QgQ0ZHWyJzZXR0aW5ncyJdLmdldCgic291bmRfZW5hYmxlZCIsIFRy"
    "dWUpOgogICAgICAgIHJldHVybgogICAgcGF0aCA9IGdldF9zb3VuZF9wYXRoKG5hbWUpCiAgICBpZiBub3QgcGF0aC5leGlzdHMo"
    "KToKICAgICAgICByZXR1cm4KCiAgICBpZiBQWUdBTUVfT0s6CiAgICAgICAgdHJ5OgogICAgICAgICAgICBzb3VuZCA9IHB5Z2Ft"
    "ZS5taXhlci5Tb3VuZChzdHIocGF0aCkpCiAgICAgICAgICAgIHNvdW5kLnBsYXkoKQogICAgICAgICAgICByZXR1cm4KICAgICAg"
    "ICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICBwYXNzCgogICAgaWYgV0lOU09VTkRfT0s6CiAgICAgICAgdHJ5OgogICAg"
    "ICAgICAgICB3aW5zb3VuZC5QbGF5U291bmQoc3RyKHBhdGgpLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgd2luc291"
    "bmQuU05EX0ZJTEVOQU1FIHwgd2luc291bmQuU05EX0FTWU5DKQogICAgICAgICAgICByZXR1cm4KICAgICAgICBleGNlcHQgRXhj"
    "ZXB0aW9uOgogICAgICAgICAgICBwYXNzCgogICAgdHJ5OgogICAgICAgIFFBcHBsaWNhdGlvbi5iZWVwKCkKICAgIGV4Y2VwdCBF"
    "eGNlcHRpb246CiAgICAgICAgcGFzcwoKIyDilIDilIAgREVTS1RPUCBTSE9SVENVVCBDUkVBVE9SIOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApkZWYgY3JlYXRlX2Rlc2t0"
    "b3Bfc2hvcnRjdXQoKSAtPiBib29sOgogICAgIiIiCiAgICBDcmVhdGUgYSBkZXNrdG9wIHNob3J0Y3V0IHRvIHRoZSBkZWNrIC5w"
    "eSBmaWxlIHVzaW5nIHB5dGhvbncuZXhlLgogICAgUmV0dXJucyBUcnVlIG9uIHN1Y2Nlc3MuIFdpbmRvd3Mgb25seS4KICAgICIi"
    "IgogICAgaWYgbm90IFdJTjMyX09LOgogICAgICAgIHJldHVybiBGYWxzZQogICAgdHJ5OgogICAgICAgIGRlc2t0b3AgPSBQYXRo"
    "LmhvbWUoKSAvICJEZXNrdG9wIgogICAgICAgIHNob3J0Y3V0X3BhdGggPSBkZXNrdG9wIC8gZiJ7REVDS19OQU1FfS5sbmsiCgog"
    "ICAgICAgICMgcHl0aG9udyA9IHNhbWUgYXMgcHl0aG9uIGJ1dCBubyBjb25zb2xlIHdpbmRvdwogICAgICAgIHB5dGhvbncgPSBQ"
    "YXRoKHN5cy5leGVjdXRhYmxlKQogICAgICAgIGlmIHB5dGhvbncubmFtZS5sb3dlcigpID09ICJweXRob24uZXhlIjoKICAgICAg"
    "ICAgICAgcHl0aG9udyA9IHB5dGhvbncucGFyZW50IC8gInB5dGhvbncuZXhlIgogICAgICAgIGlmIG5vdCBweXRob253LmV4aXN0"
    "cygpOgogICAgICAgICAgICBweXRob253ID0gUGF0aChzeXMuZXhlY3V0YWJsZSkKCiAgICAgICAgZGVja19wYXRoID0gUGF0aChf"
    "X2ZpbGVfXykucmVzb2x2ZSgpCgogICAgICAgIHNoZWxsID0gd2luMzJjb20uY2xpZW50LkRpc3BhdGNoKCJXU2NyaXB0LlNoZWxs"
    "IikKICAgICAgICBzYyA9IHNoZWxsLkNyZWF0ZVNob3J0Q3V0KHN0cihzaG9ydGN1dF9wYXRoKSkKICAgICAgICBzYy5UYXJnZXRQ"
    "YXRoICAgICA9IHN0cihweXRob253KQogICAgICAgIHNjLkFyZ3VtZW50cyAgICAgID0gZicie2RlY2tfcGF0aH0iJwogICAgICAg"
    "IHNjLldvcmtpbmdEaXJlY3RvcnkgPSBzdHIoZGVja19wYXRoLnBhcmVudCkKICAgICAgICBzYy5EZXNjcmlwdGlvbiAgICA9IGYi"
    "e0RFQ0tfTkFNRX0g4oCUIEVjaG8gRGVjayIKCiAgICAgICAgIyBVc2UgbmV1dHJhbCBmYWNlIGFzIGljb24gaWYgYXZhaWxhYmxl"
    "CiAgICAgICAgaWNvbl9wYXRoID0gY2ZnX3BhdGgoImZhY2VzIikgLyBmIntGQUNFX1BSRUZJWH1fTmV1dHJhbC5wbmciCiAgICAg"
    "ICAgaWYgaWNvbl9wYXRoLmV4aXN0cygpOgogICAgICAgICAgICAjIFdpbmRvd3Mgc2hvcnRjdXRzIGNhbid0IHVzZSBQTkcgZGly"
    "ZWN0bHkg4oCUIHNraXAgaWNvbiBpZiBubyAuaWNvCiAgICAgICAgICAgIHBhc3MKCiAgICAgICAgc2Muc2F2ZSgpCiAgICAgICAg"
    "cmV0dXJuIFRydWUKICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMgZToKICAgICAgICBwcmludChmIltTSE9SVENVVF1bV0FSTl0gQ291"
    "bGQgbm90IGNyZWF0ZSBzaG9ydGN1dDoge2V9IikKICAgICAgICByZXR1cm4gRmFsc2UKCiMg4pSA4pSAIEpTT05MIFVUSUxJVElF"
    "UyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZGVmIHJlYWRfanNvbmwocGF0aDogUGF0aCkgLT4gbGlzdFtkaWN0XToKICAg"
    "ICIiIlJlYWQgYSBKU09OTCBmaWxlLiBSZXR1cm5zIGxpc3Qgb2YgZGljdHMuIEhhbmRsZXMgSlNPTiBhcnJheXMgdG9vLiIiIgog"
    "ICAgaWYgbm90IHBhdGguZXhpc3RzKCk6CiAgICAgICAgcmV0dXJuIFtdCiAgICByYXcgPSBwYXRoLnJlYWRfdGV4dChlbmNvZGlu"
    "Zz0idXRmLTgiKS5zdHJpcCgpCiAgICBpZiBub3QgcmF3OgogICAgICAgIHJldHVybiBbXQogICAgaWYgcmF3LnN0YXJ0c3dpdGgo"
    "IlsiKToKICAgICAgICB0cnk6CiAgICAgICAgICAgIGRhdGEgPSBqc29uLmxvYWRzKHJhdykKICAgICAgICAgICAgcmV0dXJuIFt4"
    "IGZvciB4IGluIGRhdGEgaWYgaXNpbnN0YW5jZSh4LCBkaWN0KV0KICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAg"
    "ICBwYXNzCiAgICBpdGVtcyA9IFtdCiAgICBmb3IgbGluZSBpbiByYXcuc3BsaXRsaW5lcygpOgogICAgICAgIGxpbmUgPSBsaW5l"
    "LnN0cmlwKCkKICAgICAgICBpZiBub3QgbGluZToKICAgICAgICAgICAgY29udGludWUKICAgICAgICB0cnk6CiAgICAgICAgICAg"
    "IG9iaiA9IGpzb24ubG9hZHMobGluZSkKICAgICAgICAgICAgaWYgaXNpbnN0YW5jZShvYmosIGRpY3QpOgogICAgICAgICAgICAg"
    "ICAgaXRlbXMuYXBwZW5kKG9iaikKICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICBjb250aW51ZQogICAgcmV0"
    "dXJuIGl0ZW1zCgpkZWYgYXBwZW5kX2pzb25sKHBhdGg6IFBhdGgsIG9iajogZGljdCkgLT4gTm9uZToKICAgICIiIkFwcGVuZCBv"
    "bmUgcmVjb3JkIHRvIGEgSlNPTkwgZmlsZS4iIiIKICAgIHBhdGgucGFyZW50Lm1rZGlyKHBhcmVudHM9VHJ1ZSwgZXhpc3Rfb2s9"
    "VHJ1ZSkKICAgIHdpdGggcGF0aC5vcGVuKCJhIiwgZW5jb2Rpbmc9InV0Zi04IikgYXMgZjoKICAgICAgICBmLndyaXRlKGpzb24u"
    "ZHVtcHMob2JqLCBlbnN1cmVfYXNjaWk9RmFsc2UpICsgIlxuIikKCmRlZiB3cml0ZV9qc29ubChwYXRoOiBQYXRoLCByZWNvcmRz"
    "OiBsaXN0W2RpY3RdKSAtPiBOb25lOgogICAgIiIiT3ZlcndyaXRlIGEgSlNPTkwgZmlsZSB3aXRoIGEgbGlzdCBvZiByZWNvcmRz"
    "LiIiIgogICAgcGF0aC5wYXJlbnQubWtkaXIocGFyZW50cz1UcnVlLCBleGlzdF9vaz1UcnVlKQogICAgd2l0aCBwYXRoLm9wZW4o"
    "InciLCBlbmNvZGluZz0idXRmLTgiKSBhcyBmOgogICAgICAgIGZvciByIGluIHJlY29yZHM6CiAgICAgICAgICAgIGYud3JpdGUo"
    "anNvbi5kdW1wcyhyLCBlbnN1cmVfYXNjaWk9RmFsc2UpICsgIlxuIikKCiMg4pSA4pSAIEtFWVdPUkQgLyBNRU1PUlkgSEVMUEVS"
    "UyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIAKX1NUT1BXT1JEUyA9IHsKICAgICJ0aGUiLCJhbmQiLCJ0aGF0Iiwid2l0aCIsImhhdmUiLCJ0aGlzIiwiZnJvbSIsInlvdXIi"
    "LCJ3aGF0Iiwid2hlbiIsCiAgICAid2hlcmUiLCJ3aGljaCIsIndvdWxkIiwidGhlcmUiLCJ0aGV5IiwidGhlbSIsInRoZW4iLCJp"
    "bnRvIiwianVzdCIsCiAgICAiYWJvdXQiLCJsaWtlIiwiYmVjYXVzZSIsIndoaWxlIiwiY291bGQiLCJzaG91bGQiLCJ0aGVpciIs"
    "IndlcmUiLCJiZWVuIiwKICAgICJiZWluZyIsImRvZXMiLCJkaWQiLCJkb250IiwiZGlkbnQiLCJjYW50Iiwid29udCIsIm9udG8i"
    "LCJvdmVyIiwidW5kZXIiLAogICAgInRoYW4iLCJhbHNvIiwic29tZSIsIm1vcmUiLCJsZXNzIiwib25seSIsIm5lZWQiLCJ3YW50"
    "Iiwid2lsbCIsInNoYWxsIiwKICAgICJhZ2FpbiIsInZlcnkiLCJtdWNoIiwicmVhbGx5IiwibWFrZSIsIm1hZGUiLCJ1c2VkIiwi"
    "dXNpbmciLCJzYWlkIiwKICAgICJ0ZWxsIiwidG9sZCIsImlkZWEiLCJjaGF0IiwiY29kZSIsInRoaW5nIiwic3R1ZmYiLCJ1c2Vy"
    "IiwiYXNzaXN0YW50IiwKfQoKZGVmIGV4dHJhY3Rfa2V5d29yZHModGV4dDogc3RyLCBsaW1pdDogaW50ID0gMTIpIC0+IGxpc3Rb"
    "c3RyXToKICAgIHRva2VucyA9IFt0Lmxvd2VyKCkuc3RyaXAoIiAuLCE/OzonXCIoKVtde30iKSBmb3IgdCBpbiB0ZXh0LnNwbGl0"
    "KCldCiAgICBzZWVuLCByZXN1bHQgPSBzZXQoKSwgW10KICAgIGZvciB0IGluIHRva2VuczoKICAgICAgICBpZiBsZW4odCkgPCAz"
    "IG9yIHQgaW4gX1NUT1BXT1JEUyBvciB0LmlzZGlnaXQoKToKICAgICAgICAgICAgY29udGludWUKICAgICAgICBpZiB0IG5vdCBp"
    "biBzZWVuOgogICAgICAgICAgICBzZWVuLmFkZCh0KQogICAgICAgICAgICByZXN1bHQuYXBwZW5kKHQpCiAgICAgICAgaWYgbGVu"
    "KHJlc3VsdCkgPj0gbGltaXQ6CiAgICAgICAgICAgIGJyZWFrCiAgICByZXR1cm4gcmVzdWx0CgpkZWYgaW5mZXJfcmVjb3JkX3R5"
    "cGUodXNlcl90ZXh0OiBzdHIsIGFzc2lzdGFudF90ZXh0OiBzdHIgPSAiIikgLT4gc3RyOgogICAgdCA9ICh1c2VyX3RleHQgKyAi"
    "ICIgKyBhc3Npc3RhbnRfdGV4dCkubG93ZXIoKQogICAgaWYgImRyZWFtIiBpbiB0OiAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICByZXR1cm4gImRyZWFtIgogICAgaWYgYW55KHggaW4gdCBmb3IgeCBpbiAoImxzbCIsInB5dGhvbiIsInNjcmlwdCIsImNvZGUi"
    "LCJlcnJvciIsImJ1ZyIpKToKICAgICAgICBpZiBhbnkoeCBpbiB0IGZvciB4IGluICgiZml4ZWQiLCJyZXNvbHZlZCIsInNvbHV0"
    "aW9uIiwid29ya2luZyIpKToKICAgICAgICAgICAgcmV0dXJuICJyZXNvbHV0aW9uIgogICAgICAgIHJldHVybiAiaXNzdWUiCiAg"
    "ICBpZiBhbnkoeCBpbiB0IGZvciB4IGluICgicmVtaW5kIiwidGltZXIiLCJhbGFybSIsInRhc2siKSk6CiAgICAgICAgcmV0dXJu"
    "ICJ0YXNrIgogICAgaWYgYW55KHggaW4gdCBmb3IgeCBpbiAoImlkZWEiLCJjb25jZXB0Iiwid2hhdCBpZiIsImdhbWUiLCJwcm9q"
    "ZWN0IikpOgogICAgICAgIHJldHVybiAiaWRlYSIKICAgIGlmIGFueSh4IGluIHQgZm9yIHggaW4gKCJwcmVmZXIiLCJhbHdheXMi"
    "LCJuZXZlciIsImkgbGlrZSIsImkgd2FudCIpKToKICAgICAgICByZXR1cm4gInByZWZlcmVuY2UiCiAgICByZXR1cm4gImNvbnZl"
    "cnNhdGlvbiIKCiMg4pSA4pSAIFBBU1MgMSBDT01QTEVURSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKIyBOZXh0OiBQYXNz"
    "IDIg4oCUIFdpZGdldCBDbGFzc2VzCiMgKEdhdWdlV2lkZ2V0LCBNb29uV2lkZ2V0LCBTcGhlcmVXaWRnZXQsIEVtb3Rpb25CbG9j"
    "aywKIyAgTWlycm9yV2lkZ2V0LCBTdGF0ZVN0cmlwV2lkZ2V0LCBDb2xsYXBzaWJsZUJsb2NrKQoKCiMg4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQCiMg"
    "TU9SR0FOTkEgREVDSyDigJQgUEFTUyAyOiBXSURHRVQgQ0xBU1NFUwojIEFwcGVuZGVkIHRvIG1vcmdhbm5hX3Bhc3MxLnB5IHRv"
    "IGZvcm0gdGhlIGZ1bGwgZGVjay4KIwojIFdpZGdldHMgZGVmaW5lZCBoZXJlOgojICAgR2F1Z2VXaWRnZXQgICAgICAgICAg4oCU"
    "IGhvcml6b250YWwgZmlsbCBiYXIgd2l0aCBsYWJlbCBhbmQgdmFsdWUKIyAgIERyaXZlV2lkZ2V0ICAgICAgICAgIOKAlCBkcml2"
    "ZSB1c2FnZSBiYXIgKHVzZWQvdG90YWwgR0IpCiMgICBTcGhlcmVXaWRnZXQgICAgICAgICDigJQgZmlsbGVkIGNpcmNsZSBmb3Ig"
    "QkxPT0QgYW5kIE1BTkEKIyAgIE1vb25XaWRnZXQgICAgICAgICAgIOKAlCBkcmF3biBtb29uIG9yYiB3aXRoIHBoYXNlIHNoYWRv"
    "dwojICAgRW1vdGlvbkJsb2NrICAgICAgICAg4oCUIGNvbGxhcHNpYmxlIGVtb3Rpb24gaGlzdG9yeSBjaGlwcwojICAgTWlycm9y"
    "V2lkZ2V0ICAgICAgICAg4oCUIGZhY2UgaW1hZ2UgZGlzcGxheSAodGhlIE1pcnJvcikKIyAgIFN0YXRlU3RyaXBXaWRnZXQgICAg"
    "IOKAlCBmdWxsLXdpZHRoIHRpbWUvbW9vbi9zdGF0ZSBzdGF0dXMgYmFyCiMgICBDb2xsYXBzaWJsZUJsb2NrICAgICDigJQgd3Jh"
    "cHBlciB0aGF0IGFkZHMgY29sbGFwc2UgdG9nZ2xlIHRvIGFueSB3aWRnZXQKIyAgIEhhcmR3YXJlUGFuZWwgICAgICAgIOKAlCBn"
    "cm91cHMgYWxsIHN5c3RlbXMgZ2F1Z2VzCiMg4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQCgoKIyDilIDilIAgR0FVR0UgV0lER0VUIOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBHYXVnZVdpZGdldChRV2lkZ2V0KToKICAgICIiIgogICAgSG9yaXpvbnRh"
    "bCBmaWxsLWJhciBnYXVnZSB3aXRoIGdvdGhpYyBzdHlsaW5nLgogICAgU2hvd3M6IGxhYmVsICh0b3AtbGVmdCksIHZhbHVlIHRl"
    "eHQgKHRvcC1yaWdodCksIGZpbGwgYmFyIChib3R0b20pLgogICAgQ29sb3Igc2hpZnRzOiBub3JtYWwg4oaSIENfQ1JJTVNPTiDi"
    "hpIgQ19CTE9PRCBhcyB2YWx1ZSBhcHByb2FjaGVzIG1heC4KICAgIFNob3dzICdOL0EnIHdoZW4gZGF0YSBpcyB1bmF2YWlsYWJs"
    "ZS4KICAgICIiIgoKICAgIGRlZiBfX2luaXRfXygKICAgICAgICBzZWxmLAogICAgICAgIGxhYmVsOiBzdHIsCiAgICAgICAgdW5p"
    "dDogc3RyID0gIiIsCiAgICAgICAgbWF4X3ZhbDogZmxvYXQgPSAxMDAuMCwKICAgICAgICBjb2xvcjogc3RyID0gQ19HT0xELAog"
    "ICAgICAgIHBhcmVudD1Ob25lCiAgICApOgogICAgICAgIHN1cGVyKCkuX19pbml0X18ocGFyZW50KQogICAgICAgIHNlbGYubGFi"
    "ZWwgICAgPSBsYWJlbAogICAgICAgIHNlbGYudW5pdCAgICAgPSB1bml0CiAgICAgICAgc2VsZi5tYXhfdmFsICA9IG1heF92YWwK"
    "ICAgICAgICBzZWxmLmNvbG9yICAgID0gY29sb3IKICAgICAgICBzZWxmLl92YWx1ZSAgID0gMC4wCiAgICAgICAgc2VsZi5fZGlz"
    "cGxheSA9ICJOL0EiCiAgICAgICAgc2VsZi5fYXZhaWxhYmxlID0gRmFsc2UKICAgICAgICBzZWxmLnNldE1pbmltdW1TaXplKDEw"
    "MCwgNjApCiAgICAgICAgc2VsZi5zZXRNYXhpbXVtSGVpZ2h0KDcyKQoKICAgIGRlZiBzZXRWYWx1ZShzZWxmLCB2YWx1ZTogZmxv"
    "YXQsIGRpc3BsYXk6IHN0ciA9ICIiLCBhdmFpbGFibGU6IGJvb2wgPSBUcnVlKSAtPiBOb25lOgogICAgICAgIHNlbGYuX3ZhbHVl"
    "ICAgICA9IG1pbihmbG9hdCh2YWx1ZSksIHNlbGYubWF4X3ZhbCkKICAgICAgICBzZWxmLl9hdmFpbGFibGUgPSBhdmFpbGFibGUK"
    "ICAgICAgICBpZiBub3QgYXZhaWxhYmxlOgogICAgICAgICAgICBzZWxmLl9kaXNwbGF5ID0gIk4vQSIKICAgICAgICBlbGlmIGRp"
    "c3BsYXk6CiAgICAgICAgICAgIHNlbGYuX2Rpc3BsYXkgPSBkaXNwbGF5CiAgICAgICAgZWxzZToKICAgICAgICAgICAgc2VsZi5f"
    "ZGlzcGxheSA9IGYie3ZhbHVlOi4wZn17c2VsZi51bml0fSIKICAgICAgICBzZWxmLnVwZGF0ZSgpCgogICAgZGVmIHNldFVuYXZh"
    "aWxhYmxlKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fYXZhaWxhYmxlID0gRmFsc2UKICAgICAgICBzZWxmLl9kaXNwbGF5"
    "ICAgPSAiTi9BIgogICAgICAgIHNlbGYudXBkYXRlKCkKCiAgICBkZWYgcGFpbnRFdmVudChzZWxmLCBldmVudCkgLT4gTm9uZToK"
    "ICAgICAgICBwID0gUVBhaW50ZXIoc2VsZikKICAgICAgICBwLnNldFJlbmRlckhpbnQoUVBhaW50ZXIuUmVuZGVySGludC5BbnRp"
    "YWxpYXNpbmcpCiAgICAgICAgdywgaCA9IHNlbGYud2lkdGgoKSwgc2VsZi5oZWlnaHQoKQoKICAgICAgICAjIEJhY2tncm91bmQK"
    "ICAgICAgICBwLmZpbGxSZWN0KDAsIDAsIHcsIGgsIFFDb2xvcihDX0JHMykpCiAgICAgICAgcC5zZXRQZW4oUUNvbG9yKENfQk9S"
    "REVSKSkKICAgICAgICBwLmRyYXdSZWN0KDAsIDAsIHcgLSAxLCBoIC0gMSkKCiAgICAgICAgIyBMYWJlbAogICAgICAgIHAuc2V0"
    "UGVuKFFDb2xvcihDX1RFWFRfRElNKSkKICAgICAgICBwLnNldEZvbnQoUUZvbnQoREVDS19GT05ULCA4LCBRRm9udC5XZWlnaHQu"
    "Qm9sZCkpCiAgICAgICAgcC5kcmF3VGV4dCg2LCAxNCwgc2VsZi5sYWJlbCkKCiAgICAgICAgIyBWYWx1ZQogICAgICAgIHAuc2V0"
    "UGVuKFFDb2xvcihzZWxmLmNvbG9yIGlmIHNlbGYuX2F2YWlsYWJsZSBlbHNlIENfVEVYVF9ESU0pKQogICAgICAgIHAuc2V0Rm9u"
    "dChRRm9udChERUNLX0ZPTlQsIDEwLCBRRm9udC5XZWlnaHQuQm9sZCkpCiAgICAgICAgZm0gPSBwLmZvbnRNZXRyaWNzKCkKICAg"
    "ICAgICB2dyA9IGZtLmhvcml6b250YWxBZHZhbmNlKHNlbGYuX2Rpc3BsYXkpCiAgICAgICAgcC5kcmF3VGV4dCh3IC0gdncgLSA2"
    "LCAxNCwgc2VsZi5fZGlzcGxheSkKCiAgICAgICAgIyBGaWxsIGJhcgogICAgICAgIGJhcl95ID0gaCAtIDE4CiAgICAgICAgYmFy"
    "X2ggPSAxMAogICAgICAgIGJhcl93ID0gdyAtIDEyCiAgICAgICAgcC5maWxsUmVjdCg2LCBiYXJfeSwgYmFyX3csIGJhcl9oLCBR"
    "Q29sb3IoQ19CRykpCiAgICAgICAgcC5zZXRQZW4oUUNvbG9yKENfQk9SREVSKSkKICAgICAgICBwLmRyYXdSZWN0KDYsIGJhcl95"
    "LCBiYXJfdyAtIDEsIGJhcl9oIC0gMSkKCiAgICAgICAgaWYgc2VsZi5fYXZhaWxhYmxlIGFuZCBzZWxmLm1heF92YWwgPiAwOgog"
    "ICAgICAgICAgICBmcmFjID0gc2VsZi5fdmFsdWUgLyBzZWxmLm1heF92YWwKICAgICAgICAgICAgZmlsbF93ID0gbWF4KDEsIGlu"
    "dCgoYmFyX3cgLSAyKSAqIGZyYWMpKQogICAgICAgICAgICAjIENvbG9yIHNoaWZ0IG5lYXIgbGltaXQKICAgICAgICAgICAgYmFy"
    "X2NvbG9yID0gKENfQkxPT0QgaWYgZnJhYyA+IDAuODUgZWxzZQogICAgICAgICAgICAgICAgICAgICAgICAgQ19DUklNU09OIGlm"
    "IGZyYWMgPiAwLjY1IGVsc2UKICAgICAgICAgICAgICAgICAgICAgICAgIHNlbGYuY29sb3IpCiAgICAgICAgICAgIGdyYWQgPSBR"
    "TGluZWFyR3JhZGllbnQoNywgYmFyX3kgKyAxLCA3ICsgZmlsbF93LCBiYXJfeSArIDEpCiAgICAgICAgICAgIGdyYWQuc2V0Q29s"
    "b3JBdCgwLCBRQ29sb3IoYmFyX2NvbG9yKS5kYXJrZXIoMTYwKSkKICAgICAgICAgICAgZ3JhZC5zZXRDb2xvckF0KDEsIFFDb2xv"
    "cihiYXJfY29sb3IpKQogICAgICAgICAgICBwLmZpbGxSZWN0KDcsIGJhcl95ICsgMSwgZmlsbF93LCBiYXJfaCAtIDIsIGdyYWQp"
    "CgogICAgICAgIHAuZW5kKCkKCgojIOKUgOKUgCBEUklWRSBXSURHRVQg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSACmNsYXNzIERyaXZlV2lkZ2V0KFFXaWRnZXQpOgogICAgIiIiCiAgICBEcml2ZSB1c2FnZSBkaXNwbGF5LiBTaG93cyBkcml2"
    "ZSBsZXR0ZXIsIHVzZWQvdG90YWwgR0IsIGZpbGwgYmFyLgogICAgQXV0by1kZXRlY3RzIGFsbCBtb3VudGVkIGRyaXZlcyB2aWEg"
    "cHN1dGlsLgogICAgIiIiCgogICAgZGVmIF9faW5pdF9fKHNlbGYsIHBhcmVudD1Ob25lKToKICAgICAgICBzdXBlcigpLl9faW5p"
    "dF9fKHBhcmVudCkKICAgICAgICBzZWxmLl9kcml2ZXM6IGxpc3RbZGljdF0gPSBbXQogICAgICAgIHNlbGYuc2V0TWluaW11bUhl"
    "aWdodCgzMCkKICAgICAgICBzZWxmLl9yZWZyZXNoKCkKCiAgICBkZWYgX3JlZnJlc2goc2VsZikgLT4gTm9uZToKICAgICAgICBz"
    "ZWxmLl9kcml2ZXMgPSBbXQogICAgICAgIGlmIG5vdCBQU1VUSUxfT0s6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIHRyeToK"
    "ICAgICAgICAgICAgZm9yIHBhcnQgaW4gcHN1dGlsLmRpc2tfcGFydGl0aW9ucyhhbGw9RmFsc2UpOgogICAgICAgICAgICAgICAg"
    "dHJ5OgogICAgICAgICAgICAgICAgICAgIHVzYWdlID0gcHN1dGlsLmRpc2tfdXNhZ2UocGFydC5tb3VudHBvaW50KQogICAgICAg"
    "ICAgICAgICAgICAgIHNlbGYuX2RyaXZlcy5hcHBlbmQoewogICAgICAgICAgICAgICAgICAgICAgICAibGV0dGVyIjogcGFydC5k"
    "ZXZpY2UucnN0cmlwKCJcXCIpLnJzdHJpcCgiLyIpLAogICAgICAgICAgICAgICAgICAgICAgICAidXNlZCI6ICAgdXNhZ2UudXNl"
    "ZCAgLyAxMDI0KiozLAogICAgICAgICAgICAgICAgICAgICAgICAidG90YWwiOiAgdXNhZ2UudG90YWwgLyAxMDI0KiozLAogICAg"
    "ICAgICAgICAgICAgICAgICAgICAicGN0IjogICAgdXNhZ2UucGVyY2VudCAvIDEwMC4wLAogICAgICAgICAgICAgICAgICAgIH0p"
    "CiAgICAgICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgZXhj"
    "ZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAgcGFzcwogICAgICAgICMgUmVzaXplIHRvIGZpdCBhbGwgZHJpdmVzCiAgICAgICAg"
    "biA9IG1heCgxLCBsZW4oc2VsZi5fZHJpdmVzKSkKICAgICAgICBzZWxmLnNldE1pbmltdW1IZWlnaHQobiAqIDI4ICsgOCkKICAg"
    "ICAgICBzZWxmLnVwZGF0ZSgpCgogICAgZGVmIHBhaW50RXZlbnQoc2VsZiwgZXZlbnQpIC0+IE5vbmU6CiAgICAgICAgcCA9IFFQ"
    "YWludGVyKHNlbGYpCiAgICAgICAgcC5zZXRSZW5kZXJIaW50KFFQYWludGVyLlJlbmRlckhpbnQuQW50aWFsaWFzaW5nKQogICAg"
    "ICAgIHcsIGggPSBzZWxmLndpZHRoKCksIHNlbGYuaGVpZ2h0KCkKICAgICAgICBwLmZpbGxSZWN0KDAsIDAsIHcsIGgsIFFDb2xv"
    "cihDX0JHMykpCgogICAgICAgIGlmIG5vdCBzZWxmLl9kcml2ZXM6CiAgICAgICAgICAgIHAuc2V0UGVuKFFDb2xvcihDX1RFWFRf"
    "RElNKSkKICAgICAgICAgICAgcC5zZXRGb250KFFGb250KERFQ0tfRk9OVCwgOSkpCiAgICAgICAgICAgIHAuZHJhd1RleHQoNiwg"
    "MTgsICJOL0Eg4oCUIHBzdXRpbCB1bmF2YWlsYWJsZSIpCiAgICAgICAgICAgIHAuZW5kKCkKICAgICAgICAgICAgcmV0dXJuCgog"
    "ICAgICAgIHJvd19oID0gMjYKICAgICAgICB5ID0gNAogICAgICAgIGZvciBkcnYgaW4gc2VsZi5fZHJpdmVzOgogICAgICAgICAg"
    "ICBsZXR0ZXIgPSBkcnZbImxldHRlciJdCiAgICAgICAgICAgIHVzZWQgICA9IGRydlsidXNlZCJdCiAgICAgICAgICAgIHRvdGFs"
    "ICA9IGRydlsidG90YWwiXQogICAgICAgICAgICBwY3QgICAgPSBkcnZbInBjdCJdCgogICAgICAgICAgICAjIExhYmVsCiAgICAg"
    "ICAgICAgIGxhYmVsID0gZiJ7bGV0dGVyfSAge3VzZWQ6LjFmfS97dG90YWw6LjBmfUdCIgogICAgICAgICAgICBwLnNldFBlbihR"
    "Q29sb3IoQ19HT0xEKSkKICAgICAgICAgICAgcC5zZXRGb250KFFGb250KERFQ0tfRk9OVCwgOCwgUUZvbnQuV2VpZ2h0LkJvbGQp"
    "KQogICAgICAgICAgICBwLmRyYXdUZXh0KDYsIHkgKyAxMiwgbGFiZWwpCgogICAgICAgICAgICAjIEJhcgogICAgICAgICAgICBi"
    "YXJfeCA9IDYKICAgICAgICAgICAgYmFyX3kgPSB5ICsgMTUKICAgICAgICAgICAgYmFyX3cgPSB3IC0gMTIKICAgICAgICAgICAg"
    "YmFyX2ggPSA4CiAgICAgICAgICAgIHAuZmlsbFJlY3QoYmFyX3gsIGJhcl95LCBiYXJfdywgYmFyX2gsIFFDb2xvcihDX0JHKSkK"
    "ICAgICAgICAgICAgcC5zZXRQZW4oUUNvbG9yKENfQk9SREVSKSkKICAgICAgICAgICAgcC5kcmF3UmVjdChiYXJfeCwgYmFyX3ks"
    "IGJhcl93IC0gMSwgYmFyX2ggLSAxKQoKICAgICAgICAgICAgZmlsbF93ID0gbWF4KDEsIGludCgoYmFyX3cgLSAyKSAqIHBjdCkp"
    "CiAgICAgICAgICAgIGJhcl9jb2xvciA9IChDX0JMT09EIGlmIHBjdCA+IDAuOSBlbHNlCiAgICAgICAgICAgICAgICAgICAgICAg"
    "ICBDX0NSSU1TT04gaWYgcGN0ID4gMC43NSBlbHNlCiAgICAgICAgICAgICAgICAgICAgICAgICBDX0dPTERfRElNKQogICAgICAg"
    "ICAgICBncmFkID0gUUxpbmVhckdyYWRpZW50KGJhcl94ICsgMSwgYmFyX3ksIGJhcl94ICsgZmlsbF93LCBiYXJfeSkKICAgICAg"
    "ICAgICAgZ3JhZC5zZXRDb2xvckF0KDAsIFFDb2xvcihiYXJfY29sb3IpLmRhcmtlcigxNTApKQogICAgICAgICAgICBncmFkLnNl"
    "dENvbG9yQXQoMSwgUUNvbG9yKGJhcl9jb2xvcikpCiAgICAgICAgICAgIHAuZmlsbFJlY3QoYmFyX3ggKyAxLCBiYXJfeSArIDEs"
    "IGZpbGxfdywgYmFyX2ggLSAyLCBncmFkKQoKICAgICAgICAgICAgeSArPSByb3dfaAoKICAgICAgICBwLmVuZCgpCgogICAgZGVm"
    "IHJlZnJlc2goc2VsZikgLT4gTm9uZToKICAgICAgICAiIiJDYWxsIHBlcmlvZGljYWxseSB0byB1cGRhdGUgZHJpdmUgc3RhdHMu"
    "IiIiCiAgICAgICAgc2VsZi5fcmVmcmVzaCgpCgoKIyDilIDilIAgU1BIRVJFIFdJREdFVCDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIAKY2xhc3MgU3BoZXJlV2lkZ2V0KFFXaWRnZXQpOgogICAgIiIiCiAgICBGaWxsZWQgY2lyY2xlIGdhdWdlIOKA"
    "lCB1c2VkIGZvciBCTE9PRCAodG9rZW4gcG9vbCkgYW5kIE1BTkEgKFZSQU0pLgogICAgRmlsbHMgZnJvbSBib3R0b20gdXAuIEds"
    "YXNzeSBzaGluZSBlZmZlY3QuIExhYmVsIGJlbG93LgogICAgIiIiCgogICAgZGVmIF9faW5pdF9fKAogICAgICAgIHNlbGYsCiAg"
    "ICAgICAgbGFiZWw6IHN0ciwKICAgICAgICBjb2xvcl9mdWxsOiBzdHIsCiAgICAgICAgY29sb3JfZW1wdHk6IHN0ciwKICAgICAg"
    "ICBwYXJlbnQ9Tm9uZQogICAgKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKHBhcmVudCkKICAgICAgICBzZWxmLmxhYmVsICAg"
    "ICAgID0gbGFiZWwKICAgICAgICBzZWxmLmNvbG9yX2Z1bGwgID0gY29sb3JfZnVsbAogICAgICAgIHNlbGYuY29sb3JfZW1wdHkg"
    "PSBjb2xvcl9lbXB0eQogICAgICAgIHNlbGYuX2ZpbGwgICAgICAgPSAwLjAgICAjIDAuMCDihpIgMS4wCiAgICAgICAgc2VsZi5f"
    "YXZhaWxhYmxlICA9IFRydWUKICAgICAgICBzZWxmLnNldE1pbmltdW1TaXplKDgwLCAxMDApCgogICAgZGVmIHNldEZpbGwoc2Vs"
    "ZiwgZnJhY3Rpb246IGZsb2F0LCBhdmFpbGFibGU6IGJvb2wgPSBUcnVlKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2ZpbGwgICAg"
    "ICA9IG1heCgwLjAsIG1pbigxLjAsIGZyYWN0aW9uKSkKICAgICAgICBzZWxmLl9hdmFpbGFibGUgPSBhdmFpbGFibGUKICAgICAg"
    "ICBzZWxmLnVwZGF0ZSgpCgogICAgZGVmIHBhaW50RXZlbnQoc2VsZiwgZXZlbnQpIC0+IE5vbmU6CiAgICAgICAgcCA9IFFQYWlu"
    "dGVyKHNlbGYpCiAgICAgICAgcC5zZXRSZW5kZXJIaW50KFFQYWludGVyLlJlbmRlckhpbnQuQW50aWFsaWFzaW5nKQogICAgICAg"
    "IHcsIGggPSBzZWxmLndpZHRoKCksIHNlbGYuaGVpZ2h0KCkKCiAgICAgICAgciAgPSBtaW4odywgaCAtIDIwKSAvLyAyIC0gNAog"
    "ICAgICAgIGN4ID0gdyAvLyAyCiAgICAgICAgY3kgPSAoaCAtIDIwKSAvLyAyICsgNAoKICAgICAgICAjIERyb3Agc2hhZG93CiAg"
    "ICAgICAgcC5zZXRQZW4oUXQuUGVuU3R5bGUuTm9QZW4pCiAgICAgICAgcC5zZXRCcnVzaChRQ29sb3IoMCwgMCwgMCwgODApKQog"
    "ICAgICAgIHAuZHJhd0VsbGlwc2UoY3ggLSByICsgMywgY3kgLSByICsgMywgciAqIDIsIHIgKiAyKQoKICAgICAgICAjIEJhc2Ug"
    "Y2lyY2xlIChlbXB0eSBjb2xvcikKICAgICAgICBwLnNldEJydXNoKFFDb2xvcihzZWxmLmNvbG9yX2VtcHR5KSkKICAgICAgICBw"
    "LnNldFBlbihRQ29sb3IoQ19CT1JERVIpKQogICAgICAgIHAuZHJhd0VsbGlwc2UoY3ggLSByLCBjeSAtIHIsIHIgKiAyLCByICog"
    "MikKCiAgICAgICAgIyBGaWxsIGZyb20gYm90dG9tCiAgICAgICAgaWYgc2VsZi5fZmlsbCA+IDAuMDEgYW5kIHNlbGYuX2F2YWls"
    "YWJsZToKICAgICAgICAgICAgY2lyY2xlX3BhdGggPSBRUGFpbnRlclBhdGgoKQogICAgICAgICAgICBjaXJjbGVfcGF0aC5hZGRF"
    "bGxpcHNlKGZsb2F0KGN4IC0gciksIGZsb2F0KGN5IC0gciksCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgZmxv"
    "YXQociAqIDIpLCBmbG9hdChyICogMikpCgogICAgICAgICAgICBmaWxsX3RvcF95ID0gY3kgKyByIC0gKHNlbGYuX2ZpbGwgKiBy"
    "ICogMikKICAgICAgICAgICAgZnJvbSBQeVNpZGU2LlF0Q29yZSBpbXBvcnQgUVJlY3RGCiAgICAgICAgICAgIGZpbGxfcmVjdCA9"
    "IFFSZWN0RihjeCAtIHIsIGZpbGxfdG9wX3ksIHIgKiAyLCBjeSArIHIgLSBmaWxsX3RvcF95KQogICAgICAgICAgICBmaWxsX3Bh"
    "dGggPSBRUGFpbnRlclBhdGgoKQogICAgICAgICAgICBmaWxsX3BhdGguYWRkUmVjdChmaWxsX3JlY3QpCiAgICAgICAgICAgIGNs"
    "aXBwZWQgPSBjaXJjbGVfcGF0aC5pbnRlcnNlY3RlZChmaWxsX3BhdGgpCgogICAgICAgICAgICBwLnNldFBlbihRdC5QZW5TdHls"
    "ZS5Ob1BlbikKICAgICAgICAgICAgcC5zZXRCcnVzaChRQ29sb3Ioc2VsZi5jb2xvcl9mdWxsKSkKICAgICAgICAgICAgcC5kcmF3"
    "UGF0aChjbGlwcGVkKQoKICAgICAgICAjIEdsYXNzeSBzaGluZQogICAgICAgIHNoaW5lID0gUVJhZGlhbEdyYWRpZW50KAogICAg"
    "ICAgICAgICBmbG9hdChjeCAtIHIgKiAwLjMpLCBmbG9hdChjeSAtIHIgKiAwLjMpLCBmbG9hdChyICogMC42KQogICAgICAgICkK"
    "ICAgICAgICBzaGluZS5zZXRDb2xvckF0KDAsIFFDb2xvcigyNTUsIDI1NSwgMjU1LCA1NSkpCiAgICAgICAgc2hpbmUuc2V0Q29s"
    "b3JBdCgxLCBRQ29sb3IoMjU1LCAyNTUsIDI1NSwgMCkpCiAgICAgICAgcC5zZXRCcnVzaChzaGluZSkKICAgICAgICBwLnNldFBl"
    "bihRdC5QZW5TdHlsZS5Ob1BlbikKICAgICAgICBwLmRyYXdFbGxpcHNlKGN4IC0gciwgY3kgLSByLCByICogMiwgciAqIDIpCgog"
    "ICAgICAgICMgT3V0bGluZQogICAgICAgIHAuc2V0QnJ1c2goUXQuQnJ1c2hTdHlsZS5Ob0JydXNoKQogICAgICAgIHAuc2V0UGVu"
    "KFFQZW4oUUNvbG9yKHNlbGYuY29sb3JfZnVsbCksIDEpKQogICAgICAgIHAuZHJhd0VsbGlwc2UoY3ggLSByLCBjeSAtIHIsIHIg"
    "KiAyLCByICogMikKCiAgICAgICAgIyBOL0Egb3ZlcmxheQogICAgICAgIGlmIG5vdCBzZWxmLl9hdmFpbGFibGU6CiAgICAgICAg"
    "ICAgIHAuc2V0UGVuKFFDb2xvcihDX1RFWFRfRElNKSkKICAgICAgICAgICAgcC5zZXRGb250KFFGb250KCJDb3VyaWVyIE5ldyIs"
    "IDgpKQogICAgICAgICAgICBmbSA9IHAuZm9udE1ldHJpY3MoKQogICAgICAgICAgICB0eHQgPSAiTi9BIgogICAgICAgICAgICBw"
    "LmRyYXdUZXh0KGN4IC0gZm0uaG9yaXpvbnRhbEFkdmFuY2UodHh0KSAvLyAyLCBjeSArIDQsIHR4dCkKCiAgICAgICAgIyBMYWJl"
    "bCBiZWxvdyBzcGhlcmUKICAgICAgICBsYWJlbF90ZXh0ID0gKHNlbGYubGFiZWwgaWYgc2VsZi5fYXZhaWxhYmxlIGVsc2UKICAg"
    "ICAgICAgICAgICAgICAgICAgIGYie3NlbGYubGFiZWx9IikKICAgICAgICBwY3RfdGV4dCA9IGYie2ludChzZWxmLl9maWxsICog"
    "MTAwKX0lIiBpZiBzZWxmLl9hdmFpbGFibGUgZWxzZSAiIgoKICAgICAgICBwLnNldFBlbihRQ29sb3Ioc2VsZi5jb2xvcl9mdWxs"
    "KSkKICAgICAgICBwLnNldEZvbnQoUUZvbnQoREVDS19GT05ULCA4LCBRRm9udC5XZWlnaHQuQm9sZCkpCiAgICAgICAgZm0gPSBw"
    "LmZvbnRNZXRyaWNzKCkKCiAgICAgICAgbHcgPSBmbS5ob3Jpem9udGFsQWR2YW5jZShsYWJlbF90ZXh0KQogICAgICAgIHAuZHJh"
    "d1RleHQoY3ggLSBsdyAvLyAyLCBoIC0gMTAsIGxhYmVsX3RleHQpCgogICAgICAgIGlmIHBjdF90ZXh0OgogICAgICAgICAgICBw"
    "LnNldFBlbihRQ29sb3IoQ19URVhUX0RJTSkpCiAgICAgICAgICAgIHAuc2V0Rm9udChRRm9udChERUNLX0ZPTlQsIDcpKQogICAg"
    "ICAgICAgICBmbTIgPSBwLmZvbnRNZXRyaWNzKCkKICAgICAgICAgICAgcHcgPSBmbTIuaG9yaXpvbnRhbEFkdmFuY2UocGN0X3Rl"
    "eHQpCiAgICAgICAgICAgIHAuZHJhd1RleHQoY3ggLSBwdyAvLyAyLCBoIC0gMSwgcGN0X3RleHQpCgogICAgICAgIHAuZW5kKCkK"
    "CgojIOKUgOKUgCBNT09OIFdJREdFVCDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xhc3MgTW9vbldp"
    "ZGdldChRV2lkZ2V0KToKICAgICIiIgogICAgRHJhd24gbW9vbiBvcmIgd2l0aCBwaGFzZS1hY2N1cmF0ZSBzaGFkb3cuCgogICAg"
    "UEhBU0UgQ09OVkVOVElPTiAobm9ydGhlcm4gaGVtaXNwaGVyZSwgc3RhbmRhcmQpOgogICAgICAtIFdheGluZyAobmV34oaSZnVs"
    "bCk6IGlsbHVtaW5hdGVkIHJpZ2h0IHNpZGUsIHNoYWRvdyBvbiBsZWZ0CiAgICAgIC0gV2FuaW5nIChmdWxs4oaSbmV3KTogaWxs"
    "dW1pbmF0ZWQgbGVmdCBzaWRlLCBzaGFkb3cgb24gcmlnaHQKCiAgICBUaGUgc2hhZG93X3NpZGUgZmxhZyBjYW4gYmUgZmxpcHBl"
    "ZCBpZiB0ZXN0aW5nIHJldmVhbHMgaXQncyBiYWNrd2FyZHMKICAgIG9uIHRoaXMgbWFjaGluZS4gU2V0IE1PT05fU0hBRE9XX0ZM"
    "SVAgPSBUcnVlIGluIHRoYXQgY2FzZS4KICAgICIiIgoKICAgICMg4oaQIEZMSVAgVEhJUyB0byBUcnVlIGlmIG1vb24gYXBwZWFy"
    "cyBiYWNrd2FyZHMgZHVyaW5nIHRlc3RpbmcKICAgIE1PT05fU0hBRE9XX0ZMSVA6IGJvb2wgPSBGYWxzZQoKICAgIGRlZiBfX2lu"
    "aXRfXyhzZWxmLCBwYXJlbnQ9Tm9uZSk6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2VsZi5fcGhh"
    "c2UgICAgICAgPSAwLjAgICAgIyAwLjA9bmV3LCAwLjU9ZnVsbCwgMS4wPW5ldwogICAgICAgIHNlbGYuX25hbWUgICAgICAgID0g"
    "Ik5FVyBNT09OIgogICAgICAgIHNlbGYuX2lsbHVtaW5hdGlvbiA9IDAuMCAgICMgMC0xMDAKICAgICAgICBzZWxmLl9zdW5yaXNl"
    "ICAgICAgPSAiMDY6MDAiCiAgICAgICAgc2VsZi5fc3Vuc2V0ICAgICAgID0gIjE4OjMwIgogICAgICAgIHNlbGYuX3N1bl9kYXRl"
    "ICAgICA9IE5vbmUKICAgICAgICBzZWxmLnNldE1pbmltdW1TaXplKDgwLCAxMTApCiAgICAgICAgc2VsZi51cGRhdGVQaGFzZSgp"
    "ICAgICAgICAgICMgcG9wdWxhdGUgY29ycmVjdCBwaGFzZSBpbW1lZGlhdGVseQogICAgICAgIHNlbGYuX2ZldGNoX3N1bl9hc3lu"
    "YygpCgogICAgZGVmIF9mZXRjaF9zdW5fYXN5bmMoc2VsZikgLT4gTm9uZToKICAgICAgICBkZWYgX2ZldGNoKCk6CiAgICAgICAg"
    "ICAgIHNyLCBzcyA9IGdldF9zdW5fdGltZXMoKQogICAgICAgICAgICBzZWxmLl9zdW5yaXNlID0gc3IKICAgICAgICAgICAgc2Vs"
    "Zi5fc3Vuc2V0ICA9IHNzCiAgICAgICAgICAgIHNlbGYuX3N1bl9kYXRlID0gZGF0ZXRpbWUubm93KCkuYXN0aW1lem9uZSgpLmRh"
    "dGUoKQogICAgICAgICAgICAjIFNjaGVkdWxlIHJlcGFpbnQgb24gbWFpbiB0aHJlYWQgdmlhIFFUaW1lciDigJQgbmV2ZXIgY2Fs"
    "bAogICAgICAgICAgICAjIHNlbGYudXBkYXRlKCkgZGlyZWN0bHkgZnJvbSBhIGJhY2tncm91bmQgdGhyZWFkCiAgICAgICAgICAg"
    "IFFUaW1lci5zaW5nbGVTaG90KDAsIHNlbGYudXBkYXRlKQogICAgICAgIHRocmVhZGluZy5UaHJlYWQodGFyZ2V0PV9mZXRjaCwg"
    "ZGFlbW9uPVRydWUpLnN0YXJ0KCkKCiAgICBkZWYgdXBkYXRlUGhhc2Uoc2VsZikgLT4gTm9uZToKICAgICAgICBzZWxmLl9waGFz"
    "ZSwgc2VsZi5fbmFtZSwgc2VsZi5faWxsdW1pbmF0aW9uID0gZ2V0X21vb25fcGhhc2UoKQogICAgICAgIHRvZGF5ID0gZGF0ZXRp"
    "bWUubm93KCkuYXN0aW1lem9uZSgpLmRhdGUoKQogICAgICAgIGlmIHNlbGYuX3N1bl9kYXRlICE9IHRvZGF5OgogICAgICAgICAg"
    "ICBzZWxmLl9mZXRjaF9zdW5fYXN5bmMoKQogICAgICAgIHNlbGYudXBkYXRlKCkKCiAgICBkZWYgcGFpbnRFdmVudChzZWxmLCBl"
    "dmVudCkgLT4gTm9uZToKICAgICAgICBwID0gUVBhaW50ZXIoc2VsZikKICAgICAgICBwLnNldFJlbmRlckhpbnQoUVBhaW50ZXIu"
    "UmVuZGVySGludC5BbnRpYWxpYXNpbmcpCiAgICAgICAgdywgaCA9IHNlbGYud2lkdGgoKSwgc2VsZi5oZWlnaHQoKQoKICAgICAg"
    "ICByICA9IG1pbih3LCBoIC0gMzYpIC8vIDIgLSA0CiAgICAgICAgY3ggPSB3IC8vIDIKICAgICAgICBjeSA9IChoIC0gMzYpIC8v"
    "IDIgKyA0CgogICAgICAgICMgQmFja2dyb3VuZCBjaXJjbGUgKHNwYWNlKQogICAgICAgIHAuc2V0QnJ1c2goUUNvbG9yKDIwLCAx"
    "MiwgMjgpKQogICAgICAgIHAuc2V0UGVuKFFQZW4oUUNvbG9yKENfU0lMVkVSX0RJTSksIDEpKQogICAgICAgIHAuZHJhd0VsbGlw"
    "c2UoY3ggLSByLCBjeSAtIHIsIHIgKiAyLCByICogMikKCiAgICAgICAgY3ljbGVfZGF5ID0gc2VsZi5fcGhhc2UgKiBfTFVOQVJf"
    "Q1lDTEUKICAgICAgICBpc193YXhpbmcgPSBjeWNsZV9kYXkgPCAoX0xVTkFSX0NZQ0xFIC8gMikKCiAgICAgICAgIyBGdWxsIG1v"
    "b24gYmFzZSAobW9vbiBzdXJmYWNlIGNvbG9yKQogICAgICAgIGlmIHNlbGYuX2lsbHVtaW5hdGlvbiA+IDE6CiAgICAgICAgICAg"
    "IHAuc2V0UGVuKFF0LlBlblN0eWxlLk5vUGVuKQogICAgICAgICAgICBwLnNldEJydXNoKFFDb2xvcigyMjAsIDIxMCwgMTg1KSkK"
    "ICAgICAgICAgICAgcC5kcmF3RWxsaXBzZShjeCAtIHIsIGN5IC0gciwgciAqIDIsIHIgKiAyKQoKICAgICAgICAjIFNoYWRvdyBj"
    "YWxjdWxhdGlvbgogICAgICAgICMgaWxsdW1pbmF0aW9uIGdvZXMgMOKGkjEwMCB3YXhpbmcsIDEwMOKGkjAgd2FuaW5nCiAgICAg"
    "ICAgIyBzaGFkb3dfb2Zmc2V0IGNvbnRyb2xzIGhvdyBtdWNoIG9mIHRoZSBjaXJjbGUgdGhlIHNoYWRvdyBjb3ZlcnMKICAgICAg"
    "ICBpZiBzZWxmLl9pbGx1bWluYXRpb24gPCA5OToKICAgICAgICAgICAgIyBmcmFjdGlvbiBvZiBkaWFtZXRlciB0aGUgc2hhZG93"
    "IGVsbGlwc2UgaXMgb2Zmc2V0CiAgICAgICAgICAgIGlsbHVtX2ZyYWMgID0gc2VsZi5faWxsdW1pbmF0aW9uIC8gMTAwLjAKICAg"
    "ICAgICAgICAgc2hhZG93X2ZyYWMgPSAxLjAgLSBpbGx1bV9mcmFjCgogICAgICAgICAgICAjIHdheGluZzogaWxsdW1pbmF0ZWQg"
    "cmlnaHQsIHNoYWRvdyBMRUZUCiAgICAgICAgICAgICMgd2FuaW5nOiBpbGx1bWluYXRlZCBsZWZ0LCBzaGFkb3cgUklHSFQKICAg"
    "ICAgICAgICAgIyBvZmZzZXQgbW92ZXMgdGhlIHNoYWRvdyBlbGxpcHNlIGhvcml6b250YWxseQogICAgICAgICAgICBvZmZzZXQg"
    "PSBpbnQoc2hhZG93X2ZyYWMgKiByICogMikKCiAgICAgICAgICAgIGlmIE1vb25XaWRnZXQuTU9PTl9TSEFET1dfRkxJUDoKICAg"
    "ICAgICAgICAgICAgIGlzX3dheGluZyA9IG5vdCBpc193YXhpbmcKCiAgICAgICAgICAgIGlmIGlzX3dheGluZzoKICAgICAgICAg"
    "ICAgICAgICMgU2hhZG93IG9uIGxlZnQgc2lkZQogICAgICAgICAgICAgICAgc2hhZG93X3ggPSBjeCAtIHIgLSBvZmZzZXQKICAg"
    "ICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgICMgU2hhZG93IG9uIHJpZ2h0IHNpZGUKICAgICAgICAgICAgICAgIHNoYWRv"
    "d194ID0gY3ggLSByICsgb2Zmc2V0CgogICAgICAgICAgICBwLnNldEJydXNoKFFDb2xvcigxNSwgOCwgMjIpKQogICAgICAgICAg"
    "ICBwLnNldFBlbihRdC5QZW5TdHlsZS5Ob1BlbikKCiAgICAgICAgICAgICMgRHJhdyBzaGFkb3cgZWxsaXBzZSDigJQgY2xpcHBl"
    "ZCB0byBtb29uIGNpcmNsZQogICAgICAgICAgICBtb29uX3BhdGggPSBRUGFpbnRlclBhdGgoKQogICAgICAgICAgICBtb29uX3Bh"
    "dGguYWRkRWxsaXBzZShmbG9hdChjeCAtIHIpLCBmbG9hdChjeSAtIHIpLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgZmxvYXQociAqIDIpLCBmbG9hdChyICogMikpCiAgICAgICAgICAgIHNoYWRvd19wYXRoID0gUVBhaW50ZXJQYXRoKCkKICAg"
    "ICAgICAgICAgc2hhZG93X3BhdGguYWRkRWxsaXBzZShmbG9hdChzaGFkb3dfeCksIGZsb2F0KGN5IC0gciksCiAgICAgICAgICAg"
    "ICAgICAgICAgICAgICAgICAgICAgICAgIGZsb2F0KHIgKiAyKSwgZmxvYXQociAqIDIpKQogICAgICAgICAgICBjbGlwcGVkX3No"
    "YWRvdyA9IG1vb25fcGF0aC5pbnRlcnNlY3RlZChzaGFkb3dfcGF0aCkKICAgICAgICAgICAgcC5kcmF3UGF0aChjbGlwcGVkX3No"
    "YWRvdykKCiAgICAgICAgIyBTdWJ0bGUgc3VyZmFjZSBkZXRhaWwgKGNyYXRlcnMgaW1wbGllZCBieSBzbGlnaHQgdGV4dHVyZSBn"
    "cmFkaWVudCkKICAgICAgICBzaGluZSA9IFFSYWRpYWxHcmFkaWVudChmbG9hdChjeCAtIHIgKiAwLjIpLCBmbG9hdChjeSAtIHIg"
    "KiAwLjIpLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIGZsb2F0KHIgKiAwLjgpKQogICAgICAgIHNoaW5lLnNldENv"
    "bG9yQXQoMCwgUUNvbG9yKDI1NSwgMjU1LCAyNDAsIDMwKSkKICAgICAgICBzaGluZS5zZXRDb2xvckF0KDEsIFFDb2xvcigyMDAs"
    "IDE4MCwgMTQwLCA1KSkKICAgICAgICBwLnNldEJydXNoKHNoaW5lKQogICAgICAgIHAuc2V0UGVuKFF0LlBlblN0eWxlLk5vUGVu"
    "KQogICAgICAgIHAuZHJhd0VsbGlwc2UoY3ggLSByLCBjeSAtIHIsIHIgKiAyLCByICogMikKCiAgICAgICAgIyBPdXRsaW5lCiAg"
    "ICAgICAgcC5zZXRCcnVzaChRdC5CcnVzaFN0eWxlLk5vQnJ1c2gpCiAgICAgICAgcC5zZXRQZW4oUVBlbihRQ29sb3IoQ19TSUxW"
    "RVIpLCAxKSkKICAgICAgICBwLmRyYXdFbGxpcHNlKGN4IC0gciwgY3kgLSByLCByICogMiwgciAqIDIpCgogICAgICAgICMgUGhh"
    "c2UgbmFtZSBiZWxvdyBtb29uCiAgICAgICAgcC5zZXRQZW4oUUNvbG9yKENfU0lMVkVSKSkKICAgICAgICBwLnNldEZvbnQoUUZv"
    "bnQoREVDS19GT05ULCA3LCBRRm9udC5XZWlnaHQuQm9sZCkpCiAgICAgICAgZm0gPSBwLmZvbnRNZXRyaWNzKCkKICAgICAgICBu"
    "dyA9IGZtLmhvcml6b250YWxBZHZhbmNlKHNlbGYuX25hbWUpCiAgICAgICAgcC5kcmF3VGV4dChjeCAtIG53IC8vIDIsIGN5ICsg"
    "ciArIDE0LCBzZWxmLl9uYW1lKQoKICAgICAgICAjIElsbHVtaW5hdGlvbiBwZXJjZW50YWdlCiAgICAgICAgaWxsdW1fc3RyID0g"
    "ZiJ7c2VsZi5faWxsdW1pbmF0aW9uOi4wZn0lIgogICAgICAgIHAuc2V0UGVuKFFDb2xvcihDX1RFWFRfRElNKSkKICAgICAgICBw"
    "LnNldEZvbnQoUUZvbnQoREVDS19GT05ULCA3KSkKICAgICAgICBmbTIgPSBwLmZvbnRNZXRyaWNzKCkKICAgICAgICBpdyA9IGZt"
    "Mi5ob3Jpem9udGFsQWR2YW5jZShpbGx1bV9zdHIpCiAgICAgICAgcC5kcmF3VGV4dChjeCAtIGl3IC8vIDIsIGN5ICsgciArIDI0"
    "LCBpbGx1bV9zdHIpCgogICAgICAgICMgU3VuIHRpbWVzIGF0IHZlcnkgYm90dG9tCiAgICAgICAgc3VuX3N0ciA9IGYi4piAIHtz"
    "ZWxmLl9zdW5yaXNlfSAg4pi9IHtzZWxmLl9zdW5zZXR9IgogICAgICAgIHAuc2V0UGVuKFFDb2xvcihDX0dPTERfRElNKSkKICAg"
    "ICAgICBwLnNldEZvbnQoUUZvbnQoREVDS19GT05ULCA3KSkKICAgICAgICBmbTMgPSBwLmZvbnRNZXRyaWNzKCkKICAgICAgICBz"
    "dyA9IGZtMy5ob3Jpem9udGFsQWR2YW5jZShzdW5fc3RyKQogICAgICAgIHAuZHJhd1RleHQoY3ggLSBzdyAvLyAyLCBoIC0gMiwg"
    "c3VuX3N0cikKCiAgICAgICAgcC5lbmQoKQoKCiMg4pSA4pSAIEVNT1RJT04gQkxPQ0sg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSACmNsYXNzIEVtb3Rpb25CbG9jayhRV2lkZ2V0KToKICAgICIiIgogICAgQ29sbGFwc2libGUgZW1vdGlvbiBoaXN0"
    "b3J5IHBhbmVsLgogICAgU2hvd3MgY29sb3ItY29kZWQgY2hpcHM6IOKcpiBFTU9USU9OX05BTUUgIEhIOk1NCiAgICBTaXRzIG5l"
    "eHQgdG8gdGhlIE1pcnJvciAoZmFjZSB3aWRnZXQpIGluIHRoZSBib3R0b20gYmxvY2sgcm93LgogICAgQ29sbGFwc2VzIHRvIGp1"
    "c3QgdGhlIGhlYWRlciBzdHJpcC4KICAgICIiIgoKICAgIGRlZiBfX2luaXRfXyhzZWxmLCBwYXJlbnQ9Tm9uZSk6CiAgICAgICAg"
    "c3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2VsZi5faGlzdG9yeTogbGlzdFt0dXBsZVtzdHIsIHN0cl1dID0gW10g"
    "ICMgKGVtb3Rpb24sIHRpbWVzdGFtcCkKICAgICAgICBzZWxmLl9leHBhbmRlZCA9IFRydWUKICAgICAgICBzZWxmLl9tYXhfZW50"
    "cmllcyA9IDMwCgogICAgICAgIGxheW91dCA9IFFWQm94TGF5b3V0KHNlbGYpCiAgICAgICAgbGF5b3V0LnNldENvbnRlbnRzTWFy"
    "Z2lucygwLCAwLCAwLCAwKQogICAgICAgIGxheW91dC5zZXRTcGFjaW5nKDApCgogICAgICAgICMgSGVhZGVyIHJvdwogICAgICAg"
    "IGhlYWRlciA9IFFXaWRnZXQoKQogICAgICAgIGhlYWRlci5zZXRGaXhlZEhlaWdodCgyMikKICAgICAgICBoZWFkZXIuc2V0U3R5"
    "bGVTaGVldCgKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19CRzN9OyBib3JkZXItYm90dG9tOiAxcHggc29saWQge0NfQ1JJ"
    "TVNPTl9ESU19OyIKICAgICAgICApCiAgICAgICAgaGwgPSBRSEJveExheW91dChoZWFkZXIpCiAgICAgICAgaGwuc2V0Q29udGVu"
    "dHNNYXJnaW5zKDYsIDAsIDQsIDApCiAgICAgICAgaGwuc2V0U3BhY2luZyg0KQoKICAgICAgICBsYmwgPSBRTGFiZWwoIuKdpyBF"
    "TU9USU9OQUwgUkVDT1JEIikKICAgICAgICBsYmwuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJjb2xvcjoge0NfR09MRH07"
    "IGZvbnQtc2l6ZTogOXB4OyBmb250LXdlaWdodDogYm9sZDsgIgogICAgICAgICAgICBmImZvbnQtZmFtaWx5OiB7REVDS19GT05U"
    "fSwgc2VyaWY7IGxldHRlci1zcGFjaW5nOiAxcHg7IGJvcmRlcjogbm9uZTsiCiAgICAgICAgKQogICAgICAgIHNlbGYuX3RvZ2ds"
    "ZV9idG4gPSBRVG9vbEJ1dHRvbigpCiAgICAgICAgc2VsZi5fdG9nZ2xlX2J0bi5zZXRGaXhlZFNpemUoMTYsIDE2KQogICAgICAg"
    "IHNlbGYuX3RvZ2dsZV9idG4uc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB0cmFuc3BhcmVudDsgY29s"
    "b3I6IHtDX0dPTER9OyBib3JkZXI6IG5vbmU7IGZvbnQtc2l6ZTogMTBweDsiCiAgICAgICAgKQogICAgICAgIHNlbGYuX3RvZ2ds"
    "ZV9idG4uc2V0VGV4dCgi4pa8IikKICAgICAgICBzZWxmLl90b2dnbGVfYnRuLmNsaWNrZWQuY29ubmVjdChzZWxmLl90b2dnbGUp"
    "CgogICAgICAgIGhsLmFkZFdpZGdldChsYmwpCiAgICAgICAgaGwuYWRkU3RyZXRjaCgpCiAgICAgICAgaGwuYWRkV2lkZ2V0KHNl"
    "bGYuX3RvZ2dsZV9idG4pCgogICAgICAgICMgU2Nyb2xsIGFyZWEgZm9yIGVtb3Rpb24gY2hpcHMKICAgICAgICBzZWxmLl9zY3Jv"
    "bGwgPSBRU2Nyb2xsQXJlYSgpCiAgICAgICAgc2VsZi5fc2Nyb2xsLnNldFdpZGdldFJlc2l6YWJsZShUcnVlKQogICAgICAgIHNl"
    "bGYuX3Njcm9sbC5zZXRIb3Jpem9udGFsU2Nyb2xsQmFyUG9saWN5KAogICAgICAgICAgICBRdC5TY3JvbGxCYXJQb2xpY3kuU2Ny"
    "b2xsQmFyQWx3YXlzT2ZmKQogICAgICAgIHNlbGYuX3Njcm9sbC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tncm91"
    "bmQ6IHtDX0JHMn07IGJvcmRlcjogbm9uZTsiCiAgICAgICAgKQoKICAgICAgICBzZWxmLl9jaGlwX2NvbnRhaW5lciA9IFFXaWRn"
    "ZXQoKQogICAgICAgIHNlbGYuX2NoaXBfbGF5b3V0ID0gUVZCb3hMYXlvdXQoc2VsZi5fY2hpcF9jb250YWluZXIpCiAgICAgICAg"
    "c2VsZi5fY2hpcF9sYXlvdXQuc2V0Q29udGVudHNNYXJnaW5zKDQsIDQsIDQsIDQpCiAgICAgICAgc2VsZi5fY2hpcF9sYXlvdXQu"
    "c2V0U3BhY2luZygyKQogICAgICAgIHNlbGYuX2NoaXBfbGF5b3V0LmFkZFN0cmV0Y2goKQogICAgICAgIHNlbGYuX3Njcm9sbC5z"
    "ZXRXaWRnZXQoc2VsZi5fY2hpcF9jb250YWluZXIpCgogICAgICAgIGxheW91dC5hZGRXaWRnZXQoaGVhZGVyKQogICAgICAgIGxh"
    "eW91dC5hZGRXaWRnZXQoc2VsZi5fc2Nyb2xsKQoKICAgICAgICBzZWxmLnNldE1pbmltdW1XaWR0aCgxMzApCgogICAgZGVmIF90"
    "b2dnbGUoc2VsZikgLT4gTm9uZToKICAgICAgICBzZWxmLl9leHBhbmRlZCA9IG5vdCBzZWxmLl9leHBhbmRlZAogICAgICAgIHNl"
    "bGYuX3Njcm9sbC5zZXRWaXNpYmxlKHNlbGYuX2V4cGFuZGVkKQogICAgICAgIHNlbGYuX3RvZ2dsZV9idG4uc2V0VGV4dCgi4pa8"
    "IiBpZiBzZWxmLl9leHBhbmRlZCBlbHNlICLilrIiKQogICAgICAgIHNlbGYudXBkYXRlR2VvbWV0cnkoKQoKICAgIGRlZiBhZGRF"
    "bW90aW9uKHNlbGYsIGVtb3Rpb246IHN0ciwgdGltZXN0YW1wOiBzdHIgPSAiIikgLT4gTm9uZToKICAgICAgICBpZiBub3QgdGlt"
    "ZXN0YW1wOgogICAgICAgICAgICB0aW1lc3RhbXAgPSBkYXRldGltZS5ub3coKS5zdHJmdGltZSgiJUg6JU0iKQogICAgICAgIHNl"
    "bGYuX2hpc3RvcnkuaW5zZXJ0KDAsIChlbW90aW9uLCB0aW1lc3RhbXApKQogICAgICAgIHNlbGYuX2hpc3RvcnkgPSBzZWxmLl9o"
    "aXN0b3J5WzpzZWxmLl9tYXhfZW50cmllc10KICAgICAgICBzZWxmLl9yZWJ1aWxkX2NoaXBzKCkKCiAgICBkZWYgX3JlYnVpbGRf"
    "Y2hpcHMoc2VsZikgLT4gTm9uZToKICAgICAgICAjIENsZWFyIGV4aXN0aW5nIGNoaXBzIChrZWVwIHRoZSBzdHJldGNoIGF0IGVu"
    "ZCkKICAgICAgICB3aGlsZSBzZWxmLl9jaGlwX2xheW91dC5jb3VudCgpID4gMToKICAgICAgICAgICAgaXRlbSA9IHNlbGYuX2No"
    "aXBfbGF5b3V0LnRha2VBdCgwKQogICAgICAgICAgICBpZiBpdGVtLndpZGdldCgpOgogICAgICAgICAgICAgICAgaXRlbS53aWRn"
    "ZXQoKS5kZWxldGVMYXRlcigpCgogICAgICAgIGZvciBlbW90aW9uLCB0cyBpbiBzZWxmLl9oaXN0b3J5OgogICAgICAgICAgICBj"
    "b2xvciA9IEVNT1RJT05fQ09MT1JTLmdldChlbW90aW9uLCBDX1RFWFRfRElNKQogICAgICAgICAgICBjaGlwID0gUUxhYmVsKGYi"
    "4pymIHtlbW90aW9uLnVwcGVyKCl9ICB7dHN9IikKICAgICAgICAgICAgY2hpcC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICAg"
    "ICAgZiJjb2xvcjoge2NvbG9yfTsgZm9udC1zaXplOiA5cHg7IGZvbnQtZmFtaWx5OiB7REVDS19GT05UfSwgc2VyaWY7ICIKICAg"
    "ICAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkczfTsgYm9yZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsgIgogICAgICAg"
    "ICAgICAgICAgZiJwYWRkaW5nOiAxcHggNHB4OyBib3JkZXItcmFkaXVzOiAycHg7IgogICAgICAgICAgICApCiAgICAgICAgICAg"
    "IHNlbGYuX2NoaXBfbGF5b3V0Lmluc2VydFdpZGdldCgKICAgICAgICAgICAgICAgIHNlbGYuX2NoaXBfbGF5b3V0LmNvdW50KCkg"
    "LSAxLCBjaGlwCiAgICAgICAgICAgICkKCiAgICBkZWYgY2xlYXIoc2VsZikgLT4gTm9uZToKICAgICAgICBzZWxmLl9oaXN0b3J5"
    "LmNsZWFyKCkKICAgICAgICBzZWxmLl9yZWJ1aWxkX2NoaXBzKCkKCgojIOKUgOKUgCBNSVJST1IgV0lER0VUIOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBNaXJyb3JXaWRnZXQoUUxhYmVsKToKICAgICIiIgogICAgRmFjZSBpbWFnZSBk"
    "aXNwbGF5IOKAlCAnVGhlIE1pcnJvcicuCiAgICBEeW5hbWljYWxseSBsb2FkcyBhbGwge0ZBQ0VfUFJFRklYfV8qLnBuZyBmaWxl"
    "cyBmcm9tIGNvbmZpZyBwYXRocy5mYWNlcy4KICAgIEF1dG8tbWFwcyBmaWxlbmFtZSB0byBlbW90aW9uIGtleToKICAgICAgICB7"
    "RkFDRV9QUkVGSVh9X0FsZXJ0LnBuZyAgICAg4oaSICJhbGVydCIKICAgICAgICB7RkFDRV9QUkVGSVh9X1NhZF9DcnlpbmcucG5n"
    "IOKGkiAic2FkIgogICAgICAgIHtGQUNFX1BSRUZJWH1fQ2hlYXRfTW9kZS5wbmcg4oaSICJjaGVhdG1vZGUiCiAgICBGYWxscyBi"
    "YWNrIHRvIG5ldXRyYWwsIHRoZW4gdG8gZ290aGljIHBsYWNlaG9sZGVyIGlmIG5vIGltYWdlcyBmb3VuZC4KICAgIE1pc3Npbmcg"
    "ZmFjZXMgZGVmYXVsdCB0byBuZXV0cmFsIOKAlCBubyBjcmFzaCwgbm8gaGFyZGNvZGVkIGxpc3QgcmVxdWlyZWQuCiAgICAiIiIK"
    "CiAgICAjIFNwZWNpYWwgc3RlbSDihpIgZW1vdGlvbiBrZXkgbWFwcGluZ3MgKGxvd2VyY2FzZSBzdGVtIGFmdGVyIE1vcmdhbm5h"
    "XykKICAgIF9TVEVNX1RPX0VNT1RJT046IGRpY3Rbc3RyLCBzdHJdID0gewogICAgICAgICJzYWRfY3J5aW5nIjogICJzYWQiLAog"
    "ICAgICAgICJjaGVhdF9tb2RlIjogICJjaGVhdG1vZGUiLAogICAgfQoKICAgIGRlZiBfX2luaXRfXyhzZWxmLCBwYXJlbnQ9Tm9u"
    "ZSk6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2VsZi5fZmFjZXNfZGlyICAgPSBjZmdfcGF0aCgi"
    "ZmFjZXMiKQogICAgICAgIHNlbGYuX2NhY2hlOiBkaWN0W3N0ciwgUVBpeG1hcF0gPSB7fQogICAgICAgIHNlbGYuX2N1cnJlbnQg"
    "ICAgID0gIm5ldXRyYWwiCiAgICAgICAgc2VsZi5fd2FybmVkOiBzZXRbc3RyXSA9IHNldCgpCgogICAgICAgIHNlbGYuc2V0TWlu"
    "aW11bVNpemUoMTYwLCAxNjApCiAgICAgICAgc2VsZi5zZXRBbGlnbm1lbnQoUXQuQWxpZ25tZW50RmxhZy5BbGlnbkNlbnRlcikK"
    "ICAgICAgICBzZWxmLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkcyfTsgYm9yZGVyOiAxcHgg"
    "c29saWQge0NfQ1JJTVNPTl9ESU19OyAiCiAgICAgICAgICAgIGYiYm9yZGVyLXJhZGl1czogMnB4OyIKICAgICAgICApCgogICAg"
    "ICAgIFFUaW1lci5zaW5nbGVTaG90KDMwMCwgc2VsZi5fcHJlbG9hZCkKCiAgICBkZWYgX3ByZWxvYWQoc2VsZikgLT4gTm9uZToK"
    "ICAgICAgICAiIiIKICAgICAgICBTY2FuIEZhY2VzLyBkaXJlY3RvcnkgZm9yIGFsbCB7RkFDRV9QUkVGSVh9XyoucG5nIGZpbGVz"
    "LgogICAgICAgIEJ1aWxkIGVtb3Rpb27ihpJwaXhtYXAgY2FjaGUgZHluYW1pY2FsbHkuCiAgICAgICAgTm8gaGFyZGNvZGVkIGxp"
    "c3Qg4oCUIHdoYXRldmVyIGlzIGluIHRoZSBmb2xkZXIgaXMgYXZhaWxhYmxlLgogICAgICAgICIiIgogICAgICAgIGlmIG5vdCBz"
    "ZWxmLl9mYWNlc19kaXIuZXhpc3RzKCk6CiAgICAgICAgICAgIHNlbGYuX2RyYXdfcGxhY2Vob2xkZXIoKQogICAgICAgICAgICBy"
    "ZXR1cm4KCiAgICAgICAgZm9yIGltZ19wYXRoIGluIHNlbGYuX2ZhY2VzX2Rpci5nbG9iKGYie0ZBQ0VfUFJFRklYfV8qLnBuZyIp"
    "OgogICAgICAgICAgICAjIHN0ZW0gPSBldmVyeXRoaW5nIGFmdGVyICJNb3JnYW5uYV8iIHdpdGhvdXQgLnBuZwogICAgICAgICAg"
    "ICByYXdfc3RlbSA9IGltZ19wYXRoLnN0ZW1bbGVuKGYie0ZBQ0VfUFJFRklYfV8iKTpdICAgICMgZS5nLiAiU2FkX0NyeWluZyIK"
    "ICAgICAgICAgICAgc3RlbV9sb3dlciA9IHJhd19zdGVtLmxvd2VyKCkgICAgICAgICAgICAgICAgICAgICAgICAgICMgInNhZF9j"
    "cnlpbmciCgogICAgICAgICAgICAjIE1hcCBzcGVjaWFsIHN0ZW1zIHRvIGVtb3Rpb24ga2V5cwogICAgICAgICAgICBlbW90aW9u"
    "ID0gc2VsZi5fU1RFTV9UT19FTU9USU9OLmdldChzdGVtX2xvd2VyLCBzdGVtX2xvd2VyKQoKICAgICAgICAgICAgcHggPSBRUGl4"
    "bWFwKHN0cihpbWdfcGF0aCkpCiAgICAgICAgICAgIGlmIG5vdCBweC5pc051bGwoKToKICAgICAgICAgICAgICAgIHNlbGYuX2Nh"
    "Y2hlW2Vtb3Rpb25dID0gcHgKCiAgICAgICAgaWYgc2VsZi5fY2FjaGU6CiAgICAgICAgICAgIHNlbGYuX3JlbmRlcigibmV1dHJh"
    "bCIpCiAgICAgICAgZWxzZToKICAgICAgICAgICAgc2VsZi5fZHJhd19wbGFjZWhvbGRlcigpCgogICAgZGVmIF9yZW5kZXIoc2Vs"
    "ZiwgZmFjZTogc3RyKSAtPiBOb25lOgogICAgICAgIGZhY2UgPSBmYWNlLmxvd2VyKCkuc3RyaXAoKQogICAgICAgIGlmIGZhY2Ug"
    "bm90IGluIHNlbGYuX2NhY2hlOgogICAgICAgICAgICBpZiBmYWNlIG5vdCBpbiBzZWxmLl93YXJuZWQgYW5kIGZhY2UgIT0gIm5l"
    "dXRyYWwiOgogICAgICAgICAgICAgICAgcHJpbnQoZiJbTUlSUk9SXVtXQVJOXSBGYWNlIG5vdCBpbiBjYWNoZToge2ZhY2V9IOKA"
    "lCB1c2luZyBuZXV0cmFsIikKICAgICAgICAgICAgICAgIHNlbGYuX3dhcm5lZC5hZGQoZmFjZSkKICAgICAgICAgICAgZmFjZSA9"
    "ICJuZXV0cmFsIgogICAgICAgIGlmIGZhY2Ugbm90IGluIHNlbGYuX2NhY2hlOgogICAgICAgICAgICBzZWxmLl9kcmF3X3BsYWNl"
    "aG9sZGVyKCkKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgc2VsZi5fY3VycmVudCA9IGZhY2UKICAgICAgICBweCA9IHNlbGYu"
    "X2NhY2hlW2ZhY2VdCiAgICAgICAgc2NhbGVkID0gcHguc2NhbGVkKAogICAgICAgICAgICBzZWxmLndpZHRoKCkgLSA0LAogICAg"
    "ICAgICAgICBzZWxmLmhlaWdodCgpIC0gNCwKICAgICAgICAgICAgUXQuQXNwZWN0UmF0aW9Nb2RlLktlZXBBc3BlY3RSYXRpbywK"
    "ICAgICAgICAgICAgUXQuVHJhbnNmb3JtYXRpb25Nb2RlLlNtb290aFRyYW5zZm9ybWF0aW9uLAogICAgICAgICkKICAgICAgICBz"
    "ZWxmLnNldFBpeG1hcChzY2FsZWQpCiAgICAgICAgc2VsZi5zZXRUZXh0KCIiKQoKICAgIGRlZiBfZHJhd19wbGFjZWhvbGRlcihz"
    "ZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuY2xlYXIoKQogICAgICAgIHNlbGYuc2V0VGV4dCgi4pymXG7inadcbuKcpiIpCiAg"
    "ICAgICAgc2VsZi5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX0JHMn07IGJvcmRlcjogMXB4IHNv"
    "bGlkIHtDX0NSSU1TT05fRElNfTsgIgogICAgICAgICAgICBmImNvbG9yOiB7Q19DUklNU09OX0RJTX07IGZvbnQtc2l6ZTogMjRw"
    "eDsgYm9yZGVyLXJhZGl1czogMnB4OyIKICAgICAgICApCgogICAgZGVmIHNldF9mYWNlKHNlbGYsIGZhY2U6IHN0cikgLT4gTm9u"
    "ZToKICAgICAgICBRVGltZXIuc2luZ2xlU2hvdCgwLCBsYW1iZGE6IHNlbGYuX3JlbmRlcihmYWNlKSkKCiAgICBkZWYgcmVzaXpl"
    "RXZlbnQoc2VsZiwgZXZlbnQpIC0+IE5vbmU6CiAgICAgICAgc3VwZXIoKS5yZXNpemVFdmVudChldmVudCkKICAgICAgICBpZiBz"
    "ZWxmLl9jYWNoZToKICAgICAgICAgICAgc2VsZi5fcmVuZGVyKHNlbGYuX2N1cnJlbnQpCgogICAgQHByb3BlcnR5CiAgICBkZWYg"
    "Y3VycmVudF9mYWNlKHNlbGYpIC0+IHN0cjoKICAgICAgICByZXR1cm4gc2VsZi5fY3VycmVudAoKCiMg4pSA4pSAIFZBTVBJUkUg"
    "U1RBVEUgU1RSSVAg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNsYXNzIEN5Y2xlV2lkZ2V0KE1vb25XaWRnZXQpOgogICAgIiIiR2VuZXJpYyBj"
    "eWNsZSB2aXN1YWxpemF0aW9uIHdpZGdldCAoY3VycmVudGx5IGx1bmFyLXBoYXNlIGRyaXZlbikuIiIiCgoKY2xhc3MgU3RhdGVT"
    "dHJpcFdpZGdldChRV2lkZ2V0KToKICAgICIiIgogICAgRnVsbC13aWR0aCBzdGF0dXMgYmFyIHNob3dpbmc6CiAgICAgIFsg4pym"
    "IFZBTVBJUkVfU1RBVEUgIOKAoiAgSEg6TU0gIOKAoiAg4piAIFNVTlJJU0UgIOKYvSBTVU5TRVQgIOKAoiAgTU9PTiBQSEFTRSAg"
    "SUxMVU0lIF0KICAgIEFsd2F5cyB2aXNpYmxlLCBuZXZlciBjb2xsYXBzZXMuCiAgICBVcGRhdGVzIGV2ZXJ5IG1pbnV0ZSB2aWEg"
    "ZXh0ZXJuYWwgUVRpbWVyIGNhbGwgdG8gcmVmcmVzaCgpLgogICAgQ29sb3ItY29kZWQgYnkgY3VycmVudCB2YW1waXJlIHN0YXRl"
    "LgogICAgIiIiCgogICAgZGVmIF9faW5pdF9fKHNlbGYsIHBhcmVudD1Ob25lKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKHBh"
    "cmVudCkKICAgICAgICBzZWxmLl9sYWJlbF9wcmVmaXggPSAiU1RBVEUiCiAgICAgICAgc2VsZi5fc3RhdGUgICAgID0gZ2V0X2Fp"
    "X3N0YXRlKCkKICAgICAgICBzZWxmLl90aW1lX3N0ciAgPSAiIgogICAgICAgIHNlbGYuX3N1bnJpc2UgICA9ICIwNjowMCIKICAg"
    "ICAgICBzZWxmLl9zdW5zZXQgICAgPSAiMTg6MzAiCiAgICAgICAgc2VsZi5fc3VuX2RhdGUgID0gTm9uZQogICAgICAgIHNlbGYu"
    "X21vb25fbmFtZSA9ICJORVcgTU9PTiIKICAgICAgICBzZWxmLl9pbGx1bSAgICAgPSAwLjAKICAgICAgICBzZWxmLnNldEZpeGVk"
    "SGVpZ2h0KDI4KQogICAgICAgIHNlbGYuc2V0U3R5bGVTaGVldChmImJhY2tncm91bmQ6IHtDX0JHMn07IGJvcmRlci10b3A6IDFw"
    "eCBzb2xpZCB7Q19DUklNU09OX0RJTX07IikKICAgICAgICBzZWxmLl9mZXRjaF9zdW5fYXN5bmMoKQogICAgICAgIHNlbGYucmVm"
    "cmVzaCgpCgogICAgZGVmIHNldF9sYWJlbChzZWxmLCBsYWJlbDogc3RyKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2xhYmVsX3By"
    "ZWZpeCA9IChsYWJlbCBvciAiU1RBVEUiKS5zdHJpcCgpLnVwcGVyKCkKICAgICAgICBzZWxmLnVwZGF0ZSgpCgogICAgZGVmIF9m"
    "ZXRjaF9zdW5fYXN5bmMoc2VsZikgLT4gTm9uZToKICAgICAgICBkZWYgX2YoKToKICAgICAgICAgICAgc3IsIHNzID0gZ2V0X3N1"
    "bl90aW1lcygpCiAgICAgICAgICAgIHNlbGYuX3N1bnJpc2UgPSBzcgogICAgICAgICAgICBzZWxmLl9zdW5zZXQgID0gc3MKICAg"
    "ICAgICAgICAgc2VsZi5fc3VuX2RhdGUgPSBkYXRldGltZS5ub3coKS5hc3RpbWV6b25lKCkuZGF0ZSgpCiAgICAgICAgICAgICMg"
    "U2NoZWR1bGUgcmVwYWludCBvbiBtYWluIHRocmVhZCDigJQgbmV2ZXIgY2FsbCB1cGRhdGUoKSBmcm9tCiAgICAgICAgICAgICMg"
    "YSBiYWNrZ3JvdW5kIHRocmVhZCwgaXQgY2F1c2VzIFFUaHJlYWQgY3Jhc2ggb24gc3RhcnR1cAogICAgICAgICAgICBRVGltZXIu"
    "c2luZ2xlU2hvdCgwLCBzZWxmLnVwZGF0ZSkKICAgICAgICB0aHJlYWRpbmcuVGhyZWFkKHRhcmdldD1fZiwgZGFlbW9uPVRydWUp"
    "LnN0YXJ0KCkKCiAgICBkZWYgcmVmcmVzaChzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX3N0YXRlICAgICA9IGdldF9haV9z"
    "dGF0ZSgpCiAgICAgICAgc2VsZi5fdGltZV9zdHIgID0gZGF0ZXRpbWUubm93KCkuYXN0aW1lem9uZSgpLnN0cmZ0aW1lKCIlWCIp"
    "CiAgICAgICAgdG9kYXkgPSBkYXRldGltZS5ub3coKS5hc3RpbWV6b25lKCkuZGF0ZSgpCiAgICAgICAgaWYgc2VsZi5fc3VuX2Rh"
    "dGUgIT0gdG9kYXk6CiAgICAgICAgICAgIHNlbGYuX2ZldGNoX3N1bl9hc3luYygpCiAgICAgICAgXywgc2VsZi5fbW9vbl9uYW1l"
    "LCBzZWxmLl9pbGx1bSA9IGdldF9tb29uX3BoYXNlKCkKICAgICAgICBzZWxmLnVwZGF0ZSgpCgogICAgZGVmIHBhaW50RXZlbnQo"
    "c2VsZiwgZXZlbnQpIC0+IE5vbmU6CiAgICAgICAgcCA9IFFQYWludGVyKHNlbGYpCiAgICAgICAgcC5zZXRSZW5kZXJIaW50KFFQ"
    "YWludGVyLlJlbmRlckhpbnQuQW50aWFsaWFzaW5nKQogICAgICAgIHcsIGggPSBzZWxmLndpZHRoKCksIHNlbGYuaGVpZ2h0KCkK"
    "CiAgICAgICAgcC5maWxsUmVjdCgwLCAwLCB3LCBoLCBRQ29sb3IoQ19CRzIpKQoKICAgICAgICBzdGF0ZV9jb2xvciA9IGdldF9h"
    "aV9zdGF0ZV9jb2xvcihzZWxmLl9zdGF0ZSkKICAgICAgICB0ZXh0ID0gKAogICAgICAgICAgICBmIuKcpiAge3NlbGYuX2xhYmVs"
    "X3ByZWZpeH06IHtzZWxmLl9zdGF0ZX0gIOKAoiAge3NlbGYuX3RpbWVfc3RyfSAg4oCiICAiCiAgICAgICAgICAgIGYi4piAIHtz"
    "ZWxmLl9zdW5yaXNlfSAgICDimL0ge3NlbGYuX3N1bnNldH0gIOKAoiAgIgogICAgICAgICAgICBmIntzZWxmLl9tb29uX25hbWV9"
    "ICB7c2VsZi5faWxsdW06LjBmfSUiCiAgICAgICAgKQoKICAgICAgICBwLnNldEZvbnQoUUZvbnQoREVDS19GT05ULCA5LCBRRm9u"
    "dC5XZWlnaHQuQm9sZCkpCiAgICAgICAgcC5zZXRQZW4oUUNvbG9yKHN0YXRlX2NvbG9yKSkKICAgICAgICBmbSA9IHAuZm9udE1l"
    "dHJpY3MoKQogICAgICAgIHR3ID0gZm0uaG9yaXpvbnRhbEFkdmFuY2UodGV4dCkKICAgICAgICBwLmRyYXdUZXh0KCh3IC0gdHcp"
    "IC8vIDIsIGggLSA3LCB0ZXh0KQoKICAgICAgICBwLmVuZCgpCgoKY2xhc3MgTWluaUNhbGVuZGFyV2lkZ2V0KFFXaWRnZXQpOgog"
    "ICAgZGVmIF9faW5pdF9fKHNlbGYsIHBhcmVudD1Ob25lKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKHBhcmVudCkKICAgICAg"
    "ICBsYXlvdXQgPSBRVkJveExheW91dChzZWxmKQogICAgICAgIGxheW91dC5zZXRDb250ZW50c01hcmdpbnMoMCwgMCwgMCwgMCkK"
    "ICAgICAgICBsYXlvdXQuc2V0U3BhY2luZyg0KQoKICAgICAgICBoZWFkZXIgPSBRSEJveExheW91dCgpCiAgICAgICAgaGVhZGVy"
    "LnNldENvbnRlbnRzTWFyZ2lucygwLCAwLCAwLCAwKQogICAgICAgIHNlbGYucHJldl9idG4gPSBRUHVzaEJ1dHRvbigiPDwiKQog"
    "ICAgICAgIHNlbGYubmV4dF9idG4gPSBRUHVzaEJ1dHRvbigiPj4iKQogICAgICAgIHNlbGYubW9udGhfbGJsID0gUUxhYmVsKCIi"
    "KQogICAgICAgIHNlbGYubW9udGhfbGJsLnNldEFsaWdubWVudChRdC5BbGlnbm1lbnRGbGFnLkFsaWduQ2VudGVyKQogICAgICAg"
    "IGZvciBidG4gaW4gKHNlbGYucHJldl9idG4sIHNlbGYubmV4dF9idG4pOgogICAgICAgICAgICBidG4uc2V0Rml4ZWRXaWR0aCgz"
    "NCkKICAgICAgICAgICAgYnRuLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX0JHM307IGNv"
    "bG9yOiB7Q19HT0xEfTsgYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTl9ESU19OyAiCiAgICAgICAgICAgICAgICBmImZvbnQt"
    "c2l6ZTogMTBweDsgZm9udC13ZWlnaHQ6IGJvbGQ7IHBhZGRpbmc6IDJweDsiCiAgICAgICAgICAgICkKICAgICAgICBzZWxmLm1v"
    "bnRoX2xibC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19HT0xEfTsgYm9yZGVyOiBub25lOyBmb250LXNp"
    "emU6IDEwcHg7IGZvbnQtd2VpZ2h0OiBib2xkOyIKICAgICAgICApCiAgICAgICAgaGVhZGVyLmFkZFdpZGdldChzZWxmLnByZXZf"
    "YnRuKQogICAgICAgIGhlYWRlci5hZGRXaWRnZXQoc2VsZi5tb250aF9sYmwsIDEpCiAgICAgICAgaGVhZGVyLmFkZFdpZGdldChz"
    "ZWxmLm5leHRfYnRuKQogICAgICAgIGxheW91dC5hZGRMYXlvdXQoaGVhZGVyKQoKICAgICAgICBzZWxmLmNhbGVuZGFyID0gUUNh"
    "bGVuZGFyV2lkZ2V0KCkKICAgICAgICBzZWxmLmNhbGVuZGFyLnNldEdyaWRWaXNpYmxlKFRydWUpCiAgICAgICAgc2VsZi5jYWxl"
    "bmRhci5zZXRWZXJ0aWNhbEhlYWRlckZvcm1hdChRQ2FsZW5kYXJXaWRnZXQuVmVydGljYWxIZWFkZXJGb3JtYXQuTm9WZXJ0aWNh"
    "bEhlYWRlcikKICAgICAgICBzZWxmLmNhbGVuZGFyLnNldE5hdmlnYXRpb25CYXJWaXNpYmxlKEZhbHNlKQogICAgICAgIHNlbGYu"
    "Y2FsZW5kYXIuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJRQ2FsZW5kYXJXaWRnZXQgUVdpZGdldHt7YWx0ZXJuYXRlLWJh"
    "Y2tncm91bmQtY29sb3I6e0NfQkcyfTt9fSAiCiAgICAgICAgICAgIGYiUVRvb2xCdXR0b257e2NvbG9yOntDX0dPTER9O319ICIK"
    "ICAgICAgICAgICAgZiJRQ2FsZW5kYXJXaWRnZXQgUUFic3RyYWN0SXRlbVZpZXc6ZW5hYmxlZHt7YmFja2dyb3VuZDp7Q19CRzJ9"
    "OyBjb2xvcjojZmZmZmZmOyAiCiAgICAgICAgICAgIGYic2VsZWN0aW9uLWJhY2tncm91bmQtY29sb3I6e0NfQ1JJTVNPTl9ESU19"
    "OyBzZWxlY3Rpb24tY29sb3I6e0NfVEVYVH07IGdyaWRsaW5lLWNvbG9yOntDX0JPUkRFUn07fX0gIgogICAgICAgICAgICBmIlFD"
    "YWxlbmRhcldpZGdldCBRQWJzdHJhY3RJdGVtVmlldzpkaXNhYmxlZHt7Y29sb3I6IzhiOTVhMTt9fSIKICAgICAgICApCiAgICAg"
    "ICAgbGF5b3V0LmFkZFdpZGdldChzZWxmLmNhbGVuZGFyKQoKICAgICAgICBzZWxmLnByZXZfYnRuLmNsaWNrZWQuY29ubmVjdChs"
    "YW1iZGE6IHNlbGYuY2FsZW5kYXIuc2hvd1ByZXZpb3VzTW9udGgoKSkKICAgICAgICBzZWxmLm5leHRfYnRuLmNsaWNrZWQuY29u"
    "bmVjdChsYW1iZGE6IHNlbGYuY2FsZW5kYXIuc2hvd05leHRNb250aCgpKQogICAgICAgIHNlbGYuY2FsZW5kYXIuY3VycmVudFBh"
    "Z2VDaGFuZ2VkLmNvbm5lY3Qoc2VsZi5fdXBkYXRlX2xhYmVsKQogICAgICAgIHNlbGYuX3VwZGF0ZV9sYWJlbCgpCiAgICAgICAg"
    "c2VsZi5fYXBwbHlfZm9ybWF0cygpCgogICAgZGVmIF91cGRhdGVfbGFiZWwoc2VsZiwgKmFyZ3MpOgogICAgICAgIHllYXIgPSBz"
    "ZWxmLmNhbGVuZGFyLnllYXJTaG93bigpCiAgICAgICAgbW9udGggPSBzZWxmLmNhbGVuZGFyLm1vbnRoU2hvd24oKQogICAgICAg"
    "IHNlbGYubW9udGhfbGJsLnNldFRleHQoZiJ7ZGF0ZSh5ZWFyLCBtb250aCwgMSkuc3RyZnRpbWUoJyVCICVZJyl9IikKICAgICAg"
    "ICBzZWxmLl9hcHBseV9mb3JtYXRzKCkKCiAgICBkZWYgX2FwcGx5X2Zvcm1hdHMoc2VsZik6CiAgICAgICAgYmFzZSA9IFFUZXh0"
    "Q2hhckZvcm1hdCgpCiAgICAgICAgYmFzZS5zZXRGb3JlZ3JvdW5kKFFDb2xvcigiI2U3ZWRmMyIpKQogICAgICAgIHNhdHVyZGF5"
    "ID0gUVRleHRDaGFyRm9ybWF0KCkKICAgICAgICBzYXR1cmRheS5zZXRGb3JlZ3JvdW5kKFFDb2xvcihDX0dPTERfRElNKSkKICAg"
    "ICAgICBzdW5kYXkgPSBRVGV4dENoYXJGb3JtYXQoKQogICAgICAgIHN1bmRheS5zZXRGb3JlZ3JvdW5kKFFDb2xvcihDX0JMT09E"
    "KSkKICAgICAgICBzZWxmLmNhbGVuZGFyLnNldFdlZWtkYXlUZXh0Rm9ybWF0KFF0LkRheU9mV2Vlay5Nb25kYXksIGJhc2UpCiAg"
    "ICAgICAgc2VsZi5jYWxlbmRhci5zZXRXZWVrZGF5VGV4dEZvcm1hdChRdC5EYXlPZldlZWsuVHVlc2RheSwgYmFzZSkKICAgICAg"
    "ICBzZWxmLmNhbGVuZGFyLnNldFdlZWtkYXlUZXh0Rm9ybWF0KFF0LkRheU9mV2Vlay5XZWRuZXNkYXksIGJhc2UpCiAgICAgICAg"
    "c2VsZi5jYWxlbmRhci5zZXRXZWVrZGF5VGV4dEZvcm1hdChRdC5EYXlPZldlZWsuVGh1cnNkYXksIGJhc2UpCiAgICAgICAgc2Vs"
    "Zi5jYWxlbmRhci5zZXRXZWVrZGF5VGV4dEZvcm1hdChRdC5EYXlPZldlZWsuRnJpZGF5LCBiYXNlKQogICAgICAgIHNlbGYuY2Fs"
    "ZW5kYXIuc2V0V2Vla2RheVRleHRGb3JtYXQoUXQuRGF5T2ZXZWVrLlNhdHVyZGF5LCBzYXR1cmRheSkKICAgICAgICBzZWxmLmNh"
    "bGVuZGFyLnNldFdlZWtkYXlUZXh0Rm9ybWF0KFF0LkRheU9mV2Vlay5TdW5kYXksIHN1bmRheSkKCiAgICAgICAgeWVhciA9IHNl"
    "bGYuY2FsZW5kYXIueWVhclNob3duKCkKICAgICAgICBtb250aCA9IHNlbGYuY2FsZW5kYXIubW9udGhTaG93bigpCiAgICAgICAg"
    "Zmlyc3RfZGF5ID0gUURhdGUoeWVhciwgbW9udGgsIDEpCiAgICAgICAgZm9yIGRheSBpbiByYW5nZSgxLCBmaXJzdF9kYXkuZGF5"
    "c0luTW9udGgoKSArIDEpOgogICAgICAgICAgICBkID0gUURhdGUoeWVhciwgbW9udGgsIGRheSkKICAgICAgICAgICAgZm10ID0g"
    "UVRleHRDaGFyRm9ybWF0KCkKICAgICAgICAgICAgd2Vla2RheSA9IGQuZGF5T2ZXZWVrKCkKICAgICAgICAgICAgaWYgd2Vla2Rh"
    "eSA9PSBRdC5EYXlPZldlZWsuU2F0dXJkYXkudmFsdWU6CiAgICAgICAgICAgICAgICBmbXQuc2V0Rm9yZWdyb3VuZChRQ29sb3Io"
    "Q19HT0xEX0RJTSkpCiAgICAgICAgICAgIGVsaWYgd2Vla2RheSA9PSBRdC5EYXlPZldlZWsuU3VuZGF5LnZhbHVlOgogICAgICAg"
    "ICAgICAgICAgZm10LnNldEZvcmVncm91bmQoUUNvbG9yKENfQkxPT0QpKQogICAgICAgICAgICBlbHNlOgogICAgICAgICAgICAg"
    "ICAgZm10LnNldEZvcmVncm91bmQoUUNvbG9yKCIjZTdlZGYzIikpCiAgICAgICAgICAgIHNlbGYuY2FsZW5kYXIuc2V0RGF0ZVRl"
    "eHRGb3JtYXQoZCwgZm10KQoKICAgICAgICB0b2RheV9mbXQgPSBRVGV4dENoYXJGb3JtYXQoKQogICAgICAgIHRvZGF5X2ZtdC5z"
    "ZXRGb3JlZ3JvdW5kKFFDb2xvcigiIzY4ZDM5YSIpKQogICAgICAgIHRvZGF5X2ZtdC5zZXRCYWNrZ3JvdW5kKFFDb2xvcigiIzE2"
    "MzgyNSIpKQogICAgICAgIHRvZGF5X2ZtdC5zZXRGb250V2VpZ2h0KFFGb250LldlaWdodC5Cb2xkKQogICAgICAgIHNlbGYuY2Fs"
    "ZW5kYXIuc2V0RGF0ZVRleHRGb3JtYXQoUURhdGUuY3VycmVudERhdGUoKSwgdG9kYXlfZm10KQoKCiMg4pSA4pSAIENPTExBUFNJ"
    "QkxFIEJMT0NLIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBDb2xsYXBzaWJsZUJsb2NrKFFXaWRnZXQpOgogICAgIiIiCiAg"
    "ICBXcmFwcGVyIHRoYXQgYWRkcyBhIGNvbGxhcHNlL2V4cGFuZCB0b2dnbGUgdG8gYW55IHdpZGdldC4KICAgIENvbGxhcHNlcyBo"
    "b3Jpem9udGFsbHkgKHJpZ2h0d2FyZCkg4oCUIGhpZGVzIGNvbnRlbnQsIGtlZXBzIGhlYWRlciBzdHJpcC4KICAgIEhlYWRlciBz"
    "aG93cyBsYWJlbC4gVG9nZ2xlIGJ1dHRvbiBvbiByaWdodCBlZGdlIG9mIGhlYWRlci4KCiAgICBVc2FnZToKICAgICAgICBibG9j"
    "ayA9IENvbGxhcHNpYmxlQmxvY2soIuKdpyBCTE9PRCIsIFNwaGVyZVdpZGdldCguLi4pKQogICAgICAgIGxheW91dC5hZGRXaWRn"
    "ZXQoYmxvY2spCiAgICAiIiIKCiAgICB0b2dnbGVkID0gU2lnbmFsKGJvb2wpCgogICAgZGVmIF9faW5pdF9fKHNlbGYsIGxhYmVs"
    "OiBzdHIsIGNvbnRlbnQ6IFFXaWRnZXQsCiAgICAgICAgICAgICAgICAgZXhwYW5kZWQ6IGJvb2wgPSBUcnVlLCBtaW5fd2lkdGg6"
    "IGludCA9IDkwLAogICAgICAgICAgICAgICAgIHJlc2VydmVfd2lkdGg6IGJvb2wgPSBGYWxzZSwKICAgICAgICAgICAgICAgICBw"
    "YXJlbnQ9Tm9uZSk6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2VsZi5fZXhwYW5kZWQgICAgICAg"
    "PSBleHBhbmRlZAogICAgICAgIHNlbGYuX21pbl93aWR0aCAgICAgID0gbWluX3dpZHRoCiAgICAgICAgc2VsZi5fcmVzZXJ2ZV93"
    "aWR0aCAgPSByZXNlcnZlX3dpZHRoCiAgICAgICAgc2VsZi5fY29udGVudCAgICAgICAgPSBjb250ZW50CgogICAgICAgIG1haW4g"
    "PSBRVkJveExheW91dChzZWxmKQogICAgICAgIG1haW4uc2V0Q29udGVudHNNYXJnaW5zKDAsIDAsIDAsIDApCiAgICAgICAgbWFp"
    "bi5zZXRTcGFjaW5nKDApCgogICAgICAgICMgSGVhZGVyCiAgICAgICAgc2VsZi5faGVhZGVyID0gUVdpZGdldCgpCiAgICAgICAg"
    "c2VsZi5faGVhZGVyLnNldEZpeGVkSGVpZ2h0KDIyKQogICAgICAgIHNlbGYuX2hlYWRlci5zZXRTdHlsZVNoZWV0KAogICAgICAg"
    "ICAgICBmImJhY2tncm91bmQ6IHtDX0JHM307IGJvcmRlci1ib3R0b206IDFweCBzb2xpZCB7Q19DUklNU09OX0RJTX07ICIKICAg"
    "ICAgICAgICAgZiJib3JkZXItdG9wOiAxcHggc29saWQge0NfQ1JJTVNPTl9ESU19OyIKICAgICAgICApCiAgICAgICAgaGwgPSBR"
    "SEJveExheW91dChzZWxmLl9oZWFkZXIpCiAgICAgICAgaGwuc2V0Q29udGVudHNNYXJnaW5zKDYsIDAsIDQsIDApCiAgICAgICAg"
    "aGwuc2V0U3BhY2luZyg0KQoKICAgICAgICBzZWxmLl9sYmwgPSBRTGFiZWwobGFiZWwpCiAgICAgICAgc2VsZi5fbGJsLnNldFN0"
    "eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX0dPTER9OyBmb250LXNpemU6IDlweDsgZm9udC13ZWlnaHQ6IGJvbGQ7"
    "ICIKICAgICAgICAgICAgZiJmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyBsZXR0ZXItc3BhY2luZzogMXB4OyBib3Jk"
    "ZXI6IG5vbmU7IgogICAgICAgICkKCiAgICAgICAgc2VsZi5fYnRuID0gUVRvb2xCdXR0b24oKQogICAgICAgIHNlbGYuX2J0bi5z"
    "ZXRGaXhlZFNpemUoMTYsIDE2KQogICAgICAgIHNlbGYuX2J0bi5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tncm91"
    "bmQ6IHRyYW5zcGFyZW50OyBjb2xvcjoge0NfR09MRF9ESU19OyBib3JkZXI6IG5vbmU7IGZvbnQtc2l6ZTogMTBweDsiCiAgICAg"
    "ICAgKQogICAgICAgIHNlbGYuX2J0bi5zZXRUZXh0KCI8IikKICAgICAgICBzZWxmLl9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYu"
    "X3RvZ2dsZSkKCiAgICAgICAgaGwuYWRkV2lkZ2V0KHNlbGYuX2xibCkKICAgICAgICBobC5hZGRTdHJldGNoKCkKICAgICAgICBo"
    "bC5hZGRXaWRnZXQoc2VsZi5fYnRuKQoKICAgICAgICBtYWluLmFkZFdpZGdldChzZWxmLl9oZWFkZXIpCiAgICAgICAgbWFpbi5h"
    "ZGRXaWRnZXQoc2VsZi5fY29udGVudCkKCiAgICAgICAgc2VsZi5fYXBwbHlfc3RhdGUoKQoKICAgIGRlZiBpc19leHBhbmRlZChz"
    "ZWxmKSAtPiBib29sOgogICAgICAgIHJldHVybiBzZWxmLl9leHBhbmRlZAoKICAgIGRlZiBfdG9nZ2xlKHNlbGYpIC0+IE5vbmU6"
    "CiAgICAgICAgc2VsZi5fZXhwYW5kZWQgPSBub3Qgc2VsZi5fZXhwYW5kZWQKICAgICAgICBzZWxmLl9hcHBseV9zdGF0ZSgpCiAg"
    "ICAgICAgc2VsZi50b2dnbGVkLmVtaXQoc2VsZi5fZXhwYW5kZWQpCgogICAgZGVmIF9hcHBseV9zdGF0ZShzZWxmKSAtPiBOb25l"
    "OgogICAgICAgIHNlbGYuX2NvbnRlbnQuc2V0VmlzaWJsZShzZWxmLl9leHBhbmRlZCkKICAgICAgICBzZWxmLl9idG4uc2V0VGV4"
    "dCgiPCIgaWYgc2VsZi5fZXhwYW5kZWQgZWxzZSAiPiIpCgogICAgICAgICMgUmVzZXJ2ZSBmaXhlZCBzbG90IHdpZHRoIHdoZW4g"
    "cmVxdWVzdGVkICh1c2VkIGJ5IG1pZGRsZSBsb3dlciBibG9jaykKICAgICAgICBpZiBzZWxmLl9yZXNlcnZlX3dpZHRoOgogICAg"
    "ICAgICAgICBzZWxmLnNldE1pbmltdW1XaWR0aChzZWxmLl9taW5fd2lkdGgpCiAgICAgICAgICAgIHNlbGYuc2V0TWF4aW11bVdp"
    "ZHRoKDE2Nzc3MjE1KQogICAgICAgIGVsaWYgc2VsZi5fZXhwYW5kZWQ6CiAgICAgICAgICAgIHNlbGYuc2V0TWluaW11bVdpZHRo"
    "KHNlbGYuX21pbl93aWR0aCkKICAgICAgICAgICAgc2VsZi5zZXRNYXhpbXVtV2lkdGgoMTY3NzcyMTUpICAjIHVuY29uc3RyYWlu"
    "ZWQKICAgICAgICBlbHNlOgogICAgICAgICAgICAjIENvbGxhcHNlZDoganVzdCB0aGUgaGVhZGVyIHN0cmlwIChsYWJlbCArIGJ1"
    "dHRvbikKICAgICAgICAgICAgY29sbGFwc2VkX3cgPSBzZWxmLl9oZWFkZXIuc2l6ZUhpbnQoKS53aWR0aCgpCiAgICAgICAgICAg"
    "IHNlbGYuc2V0Rml4ZWRXaWR0aChtYXgoNjAsIGNvbGxhcHNlZF93KSkKCiAgICAgICAgc2VsZi51cGRhdGVHZW9tZXRyeSgpCiAg"
    "ICAgICAgcGFyZW50ID0gc2VsZi5wYXJlbnRXaWRnZXQoKQogICAgICAgIGlmIHBhcmVudCBhbmQgcGFyZW50LmxheW91dCgpOgog"
    "ICAgICAgICAgICBwYXJlbnQubGF5b3V0KCkuYWN0aXZhdGUoKQoKCiMg4pSA4pSAIEhBUkRXQVJFIFBBTkVMIOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBIYXJkd2FyZVBhbmVsKFFXaWRnZXQpOgogICAgIiIiCiAgICBUaGUgc3lzdGVtcyBy"
    "aWdodCBwYW5lbCBjb250ZW50cy4KICAgIEdyb3Vwczogc3RhdHVzIGluZm8sIGRyaXZlIGJhcnMsIENQVS9SQU0gZ2F1Z2VzLCBH"
    "UFUvVlJBTSBnYXVnZXMsIEdQVSB0ZW1wLgogICAgUmVwb3J0cyBoYXJkd2FyZSBhdmFpbGFiaWxpdHkgaW4gRGlhZ25vc3RpY3Mg"
    "b24gc3RhcnR1cC4KICAgIFNob3dzIE4vQSBncmFjZWZ1bGx5IHdoZW4gZGF0YSB1bmF2YWlsYWJsZS4KICAgICIiIgoKICAgIGRl"
    "ZiBfX2luaXRfXyhzZWxmLCBwYXJlbnQ9Tm9uZSk6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2Vs"
    "Zi5fc2V0dXBfdWkoKQogICAgICAgIHNlbGYuX2RldGVjdF9oYXJkd2FyZSgpCgogICAgZGVmIF9zZXR1cF91aShzZWxmKSAtPiBO"
    "b25lOgogICAgICAgIGxheW91dCA9IFFWQm94TGF5b3V0KHNlbGYpCiAgICAgICAgbGF5b3V0LnNldENvbnRlbnRzTWFyZ2lucyg0"
    "LCA0LCA0LCA0KQogICAgICAgIGxheW91dC5zZXRTcGFjaW5nKDQpCgogICAgICAgIGRlZiBzZWN0aW9uX2xhYmVsKHRleHQ6IHN0"
    "cikgLT4gUUxhYmVsOgogICAgICAgICAgICBsYmwgPSBRTGFiZWwodGV4dCkKICAgICAgICAgICAgbGJsLnNldFN0eWxlU2hlZXQo"
    "CiAgICAgICAgICAgICAgICBmImNvbG9yOiB7Q19HT0xEfTsgZm9udC1zaXplOiA5cHg7IGxldHRlci1zcGFjaW5nOiAycHg7ICIK"
    "ICAgICAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsgZm9udC13ZWlnaHQ6IGJvbGQ7IgogICAg"
    "ICAgICAgICApCiAgICAgICAgICAgIHJldHVybiBsYmwKCiAgICAgICAgIyDilIDilIAgU3RhdHVzIGJsb2NrIOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIGxheW91dC5hZGRXaWRn"
    "ZXQoc2VjdGlvbl9sYWJlbCgi4p2nIFNUQVRVUyIpKQogICAgICAgIHN0YXR1c19mcmFtZSA9IFFGcmFtZSgpCiAgICAgICAgc3Rh"
    "dHVzX2ZyYW1lLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0NfUEFORUx9OyBib3JkZXI6IDFweCBz"
    "b2xpZCB7Q19CT1JERVJ9OyBib3JkZXItcmFkaXVzOiAycHg7IgogICAgICAgICkKICAgICAgICBzdGF0dXNfZnJhbWUuc2V0Rml4"
    "ZWRIZWlnaHQoODgpCiAgICAgICAgc2YgPSBRVkJveExheW91dChzdGF0dXNfZnJhbWUpCiAgICAgICAgc2Yuc2V0Q29udGVudHNN"
    "YXJnaW5zKDgsIDQsIDgsIDQpCiAgICAgICAgc2Yuc2V0U3BhY2luZygyKQoKICAgICAgICBzZWxmLmxibF9zdGF0dXMgID0gUUxh"
    "YmVsKCLinKYgU1RBVFVTOiBPRkZMSU5FIikKICAgICAgICBzZWxmLmxibF9tb2RlbCAgID0gUUxhYmVsKCLinKYgVkVTU0VMOiBM"
    "T0FESU5HLi4uIikKICAgICAgICBzZWxmLmxibF9zZXNzaW9uID0gUUxhYmVsKCLinKYgU0VTU0lPTjogMDA6MDA6MDAiKQogICAg"
    "ICAgIHNlbGYubGJsX3Rva2VucyAgPSBRTGFiZWwoIuKcpiBUT0tFTlM6IDAiKQoKICAgICAgICBmb3IgbGJsIGluIChzZWxmLmxi"
    "bF9zdGF0dXMsIHNlbGYubGJsX21vZGVsLAogICAgICAgICAgICAgICAgICAgIHNlbGYubGJsX3Nlc3Npb24sIHNlbGYubGJsX3Rv"
    "a2Vucyk6CiAgICAgICAgICAgIGxibC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICAgICAgZiJjb2xvcjoge0NfVEVYVF9ESU19"
    "OyBmb250LXNpemU6IDEwcHg7ICIKICAgICAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsgYm9y"
    "ZGVyOiBub25lOyIKICAgICAgICAgICAgKQogICAgICAgICAgICBzZi5hZGRXaWRnZXQobGJsKQoKICAgICAgICBsYXlvdXQuYWRk"
    "V2lkZ2V0KHN0YXR1c19mcmFtZSkKCiAgICAgICAgIyDilIDilIAgRHJpdmUgYmFycyDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KHNlY3Rp"
    "b25fbGFiZWwoIuKdpyBTVE9SQUdFIikpCiAgICAgICAgc2VsZi5kcml2ZV93aWRnZXQgPSBEcml2ZVdpZGdldCgpCiAgICAgICAg"
    "bGF5b3V0LmFkZFdpZGdldChzZWxmLmRyaXZlX3dpZGdldCkKCiAgICAgICAgIyDilIDilIAgQ1BVIC8gUkFNIGdhdWdlcyDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KHNl"
    "Y3Rpb25fbGFiZWwoIuKdpyBWSVRBTCBFU1NFTkNFIikpCiAgICAgICAgcmFtX2NwdSA9IFFHcmlkTGF5b3V0KCkKICAgICAgICBy"
    "YW1fY3B1LnNldFNwYWNpbmcoMykKCiAgICAgICAgc2VsZi5nYXVnZV9jcHUgID0gR2F1Z2VXaWRnZXQoIkNQVSIsICAiJSIsICAg"
    "MTAwLjAsIENfU0lMVkVSKQogICAgICAgIHNlbGYuZ2F1Z2VfcmFtICA9IEdhdWdlV2lkZ2V0KCJSQU0iLCAgIkdCIiwgICA2NC4w"
    "LCBDX0dPTERfRElNKQogICAgICAgIHJhbV9jcHUuYWRkV2lkZ2V0KHNlbGYuZ2F1Z2VfY3B1LCAwLCAwKQogICAgICAgIHJhbV9j"
    "cHUuYWRkV2lkZ2V0KHNlbGYuZ2F1Z2VfcmFtLCAwLCAxKQogICAgICAgIGxheW91dC5hZGRMYXlvdXQocmFtX2NwdSkKCiAgICAg"
    "ICAgIyDilIDilIAgR1BVIC8gVlJBTSBnYXVnZXMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSACiAgICAgICAgbGF5b3V0LmFkZFdpZGdldChzZWN0aW9uX2xhYmVsKCLinacgQVJDQU5FIFBPV0VSIikpCiAgICAgICAgZ3B1"
    "X3ZyYW0gPSBRR3JpZExheW91dCgpCiAgICAgICAgZ3B1X3ZyYW0uc2V0U3BhY2luZygzKQoKICAgICAgICBzZWxmLmdhdWdlX2dw"
    "dSAgPSBHYXVnZVdpZGdldCgiR1BVIiwgICIlIiwgICAxMDAuMCwgQ19QVVJQTEUpCiAgICAgICAgc2VsZi5nYXVnZV92cmFtID0g"
    "R2F1Z2VXaWRnZXQoIlZSQU0iLCAiR0IiLCAgICA4LjAsIENfQ1JJTVNPTikKICAgICAgICBncHVfdnJhbS5hZGRXaWRnZXQoc2Vs"
    "Zi5nYXVnZV9ncHUsICAwLCAwKQogICAgICAgIGdwdV92cmFtLmFkZFdpZGdldChzZWxmLmdhdWdlX3ZyYW0sIDAsIDEpCiAgICAg"
    "ICAgbGF5b3V0LmFkZExheW91dChncHVfdnJhbSkKCiAgICAgICAgIyDilIDilIAgR1BVIFRlbXAg4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgbGF5b3V0LmFk"
    "ZFdpZGdldChzZWN0aW9uX2xhYmVsKCLinacgSU5GRVJOQUwgSEVBVCIpKQogICAgICAgIHNlbGYuZ2F1Z2VfdGVtcCA9IEdhdWdl"
    "V2lkZ2V0KCJHUFUgVEVNUCIsICLCsEMiLCA5NS4wLCBDX0JMT09EKQogICAgICAgIHNlbGYuZ2F1Z2VfdGVtcC5zZXRNYXhpbXVt"
    "SGVpZ2h0KDY1KQogICAgICAgIGxheW91dC5hZGRXaWRnZXQoc2VsZi5nYXVnZV90ZW1wKQoKICAgICAgICAjIOKUgOKUgCBHUFUg"
    "bWFzdGVyIGJhciAoZnVsbCB3aWR0aCkg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgbGF5b3V0LmFkZFdpZGdldChzZWN0aW9u"
    "X2xhYmVsKCLinacgSU5GRVJOQUwgRU5HSU5FIikpCgogICAgICAgIGdwdV9tYXN0ZXJfZnJhbWUgPSBRRnJhbWUoKQogICAgICAg"
    "IGdwdV9tYXN0ZXJfZnJhbWUuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19QQU5FTH07IGJvcmRl"
    "cjogMXB4IHNvbGlkIHtDX0JPUkRFUn07IGJvcmRlci1yYWRpdXM6IDJweDsiCiAgICAgICAgKQogICAgICAgIGdtID0gUVZCb3hM"
    "YXlvdXQoZ3B1X21hc3Rlcl9mcmFtZSkKICAgICAgICBnbS5zZXRDb250ZW50c01hcmdpbnMoOCwgNSwgOCwgNSkKICAgICAgICBn"
    "bS5zZXRTcGFjaW5nKDEpCgogICAgICAgIHNlbGYubGJsX2dwdV9uYW1lID0gUUxhYmVsKCJHUFUiKQogICAgICAgIHNlbGYubGJs"
    "X2dwdV9uYW1lLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX1NJTFZFUn07IGZvbnQtc2l6ZTogMTFweDsg"
    "Zm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsgZm9udC13ZWlnaHQ6IGJvbGQ7IGJvcmRlcjogbm9uZTsiCiAgICAgICAg"
    "KQogICAgICAgIHNlbGYubGJsX2dwdV91c2FnZSA9IFFMYWJlbCgiMCUiKQogICAgICAgIHNlbGYubGJsX2dwdV91c2FnZS5zZXRT"
    "dHlsZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19URVhUX0RJTX07IGZvbnQtc2l6ZTogMTBweDsgZm9udC1mYW1pbHk6"
    "IHtERUNLX0ZPTlR9LCBzZXJpZjsgYm9yZGVyOiBub25lOyIKICAgICAgICApCiAgICAgICAgZ20uYWRkV2lkZ2V0KHNlbGYubGJs"
    "X2dwdV9uYW1lKQogICAgICAgIGdtLmFkZFdpZGdldChzZWxmLmxibF9ncHVfdXNhZ2UpCgogICAgICAgIHNlbGYuZ2F1Z2VfZ3B1"
    "X21hc3RlciA9IEdhdWdlV2lkZ2V0KCJHUFUgTE9BRCIsICIlIiwgMTAwLjAsIENfQ1JJTVNPTikKICAgICAgICBzZWxmLmdhdWdl"
    "X2dwdV9tYXN0ZXIuc2V0TWF4aW11bUhlaWdodCg1NSkKICAgICAgICBnbS5hZGRXaWRnZXQoc2VsZi5nYXVnZV9ncHVfbWFzdGVy"
    "KQoKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KGdwdV9tYXN0ZXJfZnJhbWUpCgogICAgICAgIHNlbGYuX2dwdV9uYW1lX2ZvbnRf"
    "bWluID0gNwogICAgICAgIHNlbGYuX2dwdV9uYW1lX2ZvbnRfbWF4ID0gMTMKICAgICAgICBzZWxmLl9ncHVfbmFtZV9wYWRkaW5n"
    "ID0gNgogICAgICAgIHNlbGYuX3VwZGF0ZV9ncHVfbmFtZV9mb250KCkKCiAgICAgICAgbGF5b3V0LmFkZFN0cmV0Y2goKQoKICAg"
    "IGRlZiBfdXBkYXRlX2dwdV9uYW1lX2ZvbnQoc2VsZikgLT4gTm9uZToKICAgICAgICBpZiBub3QgaGFzYXR0cihzZWxmLCAibGJs"
    "X2dwdV9uYW1lIik6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIHRleHQgPSBzZWxmLmxibF9ncHVfbmFtZS50ZXh0KCkuc3Ry"
    "aXAoKSBvciAiR1BVIgogICAgICAgIGF2YWlsID0gbWF4KDQwLCBzZWxmLmxibF9ncHVfbmFtZS53aWR0aCgpIC0gc2VsZi5fZ3B1"
    "X25hbWVfcGFkZGluZykKICAgICAgICBjaG9zZW4gPSBzZWxmLl9ncHVfbmFtZV9mb250X21pbgogICAgICAgIGZvciBzaXplIGlu"
    "IHJhbmdlKHNlbGYuX2dwdV9uYW1lX2ZvbnRfbWF4LCBzZWxmLl9ncHVfbmFtZV9mb250X21pbiAtIDEsIC0xKToKICAgICAgICAg"
    "ICAgZm9udCA9IFFGb250KERFQ0tfRk9OVCwgc2l6ZSwgUUZvbnQuV2VpZ2h0LkJvbGQpCiAgICAgICAgICAgIGlmIFFGb250TWV0"
    "cmljcyhmb250KS5ob3Jpem9udGFsQWR2YW5jZSh0ZXh0KSA8PSBhdmFpbDoKICAgICAgICAgICAgICAgIGNob3NlbiA9IHNpemUK"
    "ICAgICAgICAgICAgICAgIGJyZWFrCiAgICAgICAgc2VsZi5sYmxfZ3B1X25hbWUuc2V0Rm9udChRRm9udChERUNLX0ZPTlQsIGNo"
    "b3NlbiwgUUZvbnQuV2VpZ2h0LkJvbGQpKQoKICAgIGRlZiBfc2V0X2dwdV9tYXN0ZXJfbGFiZWxzKHNlbGYsIGdwdV9uYW1lOiBz"
    "dHIsIHVzYWdlX3RleHQ6IHN0cikgLT4gTm9uZToKICAgICAgICBpZiBoYXNhdHRyKHNlbGYsICJsYmxfZ3B1X25hbWUiKToKICAg"
    "ICAgICAgICAgc2VsZi5sYmxfZ3B1X25hbWUuc2V0VGV4dCgoZ3B1X25hbWUgb3IgIkdQVSIpLnN0cmlwKCkpCiAgICAgICAgICAg"
    "IHNlbGYuX3VwZGF0ZV9ncHVfbmFtZV9mb250KCkKICAgICAgICBpZiBoYXNhdHRyKHNlbGYsICJsYmxfZ3B1X3VzYWdlIik6CiAg"
    "ICAgICAgICAgIHNlbGYubGJsX2dwdV91c2FnZS5zZXRUZXh0KHVzYWdlX3RleHQpCgogICAgZGVmIHJlc2l6ZUV2ZW50KHNlbGYs"
    "IGV2ZW50KSAtPiBOb25lOgogICAgICAgIHN1cGVyKCkucmVzaXplRXZlbnQoZXZlbnQpCiAgICAgICAgc2VsZi5fdXBkYXRlX2dw"
    "dV9uYW1lX2ZvbnQoKQoKICAgIGRlZiBfZGV0ZWN0X2hhcmR3YXJlKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgIiIiCiAgICAgICAg"
    "Q2hlY2sgd2hhdCBoYXJkd2FyZSBtb25pdG9yaW5nIGlzIGF2YWlsYWJsZS4KICAgICAgICBNYXJrIHVuYXZhaWxhYmxlIGdhdWdl"
    "cyBhcHByb3ByaWF0ZWx5LgogICAgICAgIERpYWdub3N0aWMgbWVzc2FnZXMgY29sbGVjdGVkIGZvciB0aGUgRGlhZ25vc3RpY3Mg"
    "dGFiLgogICAgICAgICIiIgogICAgICAgIHNlbGYuX2RpYWdfbWVzc2FnZXM6IGxpc3Rbc3RyXSA9IFtdCgogICAgICAgIGlmIG5v"
    "dCBQU1VUSUxfT0s6CiAgICAgICAgICAgIHNlbGYuZ2F1Z2VfY3B1LnNldFVuYXZhaWxhYmxlKCkKICAgICAgICAgICAgc2VsZi5n"
    "YXVnZV9yYW0uc2V0VW5hdmFpbGFibGUoKQogICAgICAgICAgICBzZWxmLl9kaWFnX21lc3NhZ2VzLmFwcGVuZCgKICAgICAgICAg"
    "ICAgICAgICJbSEFSRFdBUkVdIHBzdXRpbCBub3QgYXZhaWxhYmxlIOKAlCBDUFUvUkFNIGdhdWdlcyBkaXNhYmxlZC4gIgogICAg"
    "ICAgICAgICAgICAgInBpcCBpbnN0YWxsIHBzdXRpbCB0byBlbmFibGUuIgogICAgICAgICAgICApCiAgICAgICAgZWxzZToKICAg"
    "ICAgICAgICAgc2VsZi5fZGlhZ19tZXNzYWdlcy5hcHBlbmQoIltIQVJEV0FSRV0gcHN1dGlsIE9LIOKAlCBDUFUvUkFNIG1vbml0"
    "b3JpbmcgYWN0aXZlLiIpCgogICAgICAgIGlmIG5vdCBOVk1MX09LOgogICAgICAgICAgICBzZWxmLmdhdWdlX2dwdS5zZXRVbmF2"
    "YWlsYWJsZSgpCiAgICAgICAgICAgIHNlbGYuZ2F1Z2VfdnJhbS5zZXRVbmF2YWlsYWJsZSgpCiAgICAgICAgICAgIHNlbGYuZ2F1"
    "Z2VfdGVtcC5zZXRVbmF2YWlsYWJsZSgpCiAgICAgICAgICAgIHNlbGYuZ2F1Z2VfZ3B1X21hc3Rlci5zZXRVbmF2YWlsYWJsZSgp"
    "CiAgICAgICAgICAgIHNlbGYuX3NldF9ncHVfbWFzdGVyX2xhYmVscygiR1BVIiwgIk4vQSIpCiAgICAgICAgICAgIHNlbGYuX2Rp"
    "YWdfbWVzc2FnZXMuYXBwZW5kKAogICAgICAgICAgICAgICAgIltIQVJEV0FSRV0gcHludm1sIG5vdCBhdmFpbGFibGUgb3Igbm8g"
    "TlZJRElBIEdQVSBkZXRlY3RlZCDigJQgIgogICAgICAgICAgICAgICAgIkdQVSBnYXVnZXMgZGlzYWJsZWQuIHBpcCBpbnN0YWxs"
    "IHB5bnZtbCB0byBlbmFibGUuIgogICAgICAgICAgICApCiAgICAgICAgZWxzZToKICAgICAgICAgICAgdHJ5OgogICAgICAgICAg"
    "ICAgICAgbmFtZSA9IHB5bnZtbC5udm1sRGV2aWNlR2V0TmFtZShncHVfaGFuZGxlKQogICAgICAgICAgICAgICAgaWYgaXNpbnN0"
    "YW5jZShuYW1lLCBieXRlcyk6CiAgICAgICAgICAgICAgICAgICAgbmFtZSA9IG5hbWUuZGVjb2RlKCkKICAgICAgICAgICAgICAg"
    "IHNlbGYuX2RpYWdfbWVzc2FnZXMuYXBwZW5kKAogICAgICAgICAgICAgICAgICAgIGYiW0hBUkRXQVJFXSBweW52bWwgT0sg4oCU"
    "IEdQVSBkZXRlY3RlZDoge25hbWV9IgogICAgICAgICAgICAgICAgKQogICAgICAgICAgICAgICAgc2VsZi5fc2V0X2dwdV9tYXN0"
    "ZXJfbGFiZWxzKHN0cihuYW1lKSwgIjAlIikKICAgICAgICAgICAgICAgICMgVXBkYXRlIG1heCBWUkFNIGZyb20gYWN0dWFsIGhh"
    "cmR3YXJlCiAgICAgICAgICAgICAgICBtZW0gPSBweW52bWwubnZtbERldmljZUdldE1lbW9yeUluZm8oZ3B1X2hhbmRsZSkKICAg"
    "ICAgICAgICAgICAgIHRvdGFsX2diID0gbWVtLnRvdGFsIC8gMTAyNCoqMwogICAgICAgICAgICAgICAgc2VsZi5nYXVnZV92cmFt"
    "Lm1heF92YWwgPSB0b3RhbF9nYgogICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAgICAgICAgICAgICBzZWxm"
    "Ll9kaWFnX21lc3NhZ2VzLmFwcGVuZChmIltIQVJEV0FSRV0gcHludm1sIGVycm9yOiB7ZX0iKQoKICAgIGRlZiB1cGRhdGVfc3Rh"
    "dHMoc2VsZikgLT4gTm9uZToKICAgICAgICAiIiIKICAgICAgICBDYWxsZWQgZXZlcnkgc2Vjb25kIGZyb20gdGhlIHN0YXRzIFFU"
    "aW1lci4KICAgICAgICBSZWFkcyBoYXJkd2FyZSBhbmQgdXBkYXRlcyBhbGwgZ2F1Z2VzLgogICAgICAgICIiIgogICAgICAgIGlm"
    "IFBTVVRJTF9PSzoKICAgICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAgY3B1ID0gcHN1dGlsLmNwdV9wZXJjZW50KCkKICAg"
    "ICAgICAgICAgICAgIHNlbGYuZ2F1Z2VfY3B1LnNldFZhbHVlKGNwdSwgZiJ7Y3B1Oi4wZn0lIiwgYXZhaWxhYmxlPVRydWUpCgog"
    "ICAgICAgICAgICAgICAgbWVtID0gcHN1dGlsLnZpcnR1YWxfbWVtb3J5KCkKICAgICAgICAgICAgICAgIHJ1ICA9IG1lbS51c2Vk"
    "ICAvIDEwMjQqKjMKICAgICAgICAgICAgICAgIHJ0ICA9IG1lbS50b3RhbCAvIDEwMjQqKjMKICAgICAgICAgICAgICAgIHNlbGYu"
    "Z2F1Z2VfcmFtLnNldFZhbHVlKHJ1LCBmIntydTouMWZ9L3tydDouMGZ9R0IiLAogICAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgICAgICAgYXZhaWxhYmxlPVRydWUpCiAgICAgICAgICAgICAgICBzZWxmLmdhdWdlX3JhbS5tYXhfdmFsID0gcnQKICAg"
    "ICAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAgICAgIHBhc3MKCiAgICAgICAgaWYgTlZNTF9PSyBhbmQgZ3B1"
    "X2hhbmRsZToKICAgICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAgdXRpbCAgICAgPSBweW52bWwubnZtbERldmljZUdldFV0"
    "aWxpemF0aW9uUmF0ZXMoZ3B1X2hhbmRsZSkKICAgICAgICAgICAgICAgIG1lbV9pbmZvID0gcHludm1sLm52bWxEZXZpY2VHZXRN"
    "ZW1vcnlJbmZvKGdwdV9oYW5kbGUpCiAgICAgICAgICAgICAgICB0ZW1wICAgICA9IHB5bnZtbC5udm1sRGV2aWNlR2V0VGVtcGVy"
    "YXR1cmUoCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBncHVfaGFuZGxlLCBweW52bWwuTlZNTF9URU1QRVJBVFVSRV9H"
    "UFUpCgogICAgICAgICAgICAgICAgZ3B1X3BjdCAgID0gZmxvYXQodXRpbC5ncHUpCiAgICAgICAgICAgICAgICB2cmFtX3VzZWQg"
    "PSBtZW1faW5mby51c2VkICAvIDEwMjQqKjMKICAgICAgICAgICAgICAgIHZyYW1fdG90ICA9IG1lbV9pbmZvLnRvdGFsIC8gMTAy"
    "NCoqMwoKICAgICAgICAgICAgICAgIHNlbGYuZ2F1Z2VfZ3B1LnNldFZhbHVlKGdwdV9wY3QsIGYie2dwdV9wY3Q6LjBmfSUiLAog"
    "ICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgYXZhaWxhYmxlPVRydWUpCiAgICAgICAgICAgICAgICBzZWxm"
    "LmdhdWdlX3ZyYW0uc2V0VmFsdWUodnJhbV91c2VkLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIGYi"
    "e3ZyYW1fdXNlZDouMWZ9L3t2cmFtX3RvdDouMGZ9R0IiLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "IGF2YWlsYWJsZT1UcnVlKQogICAgICAgICAgICAgICAgc2VsZi5nYXVnZV90ZW1wLnNldFZhbHVlKGZsb2F0KHRlbXApLCBmInt0"
    "ZW1wfcKwQyIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgYXZhaWxhYmxlPVRydWUpCgogICAgICAg"
    "ICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAgICAgIG5hbWUgPSBweW52bWwubnZtbERldmljZUdldE5hbWUoZ3B1X2hhbmRs"
    "ZSkKICAgICAgICAgICAgICAgICAgICBpZiBpc2luc3RhbmNlKG5hbWUsIGJ5dGVzKToKICAgICAgICAgICAgICAgICAgICAgICAg"
    "bmFtZSA9IG5hbWUuZGVjb2RlKCkKICAgICAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAgICAgICAgICAgICAgICAg"
    "bmFtZSA9ICJHUFUiCgogICAgICAgICAgICAgICAgc2VsZi5fc2V0X2dwdV9tYXN0ZXJfbGFiZWxzKHN0cihuYW1lKSwgZiJ7Z3B1"
    "X3BjdDouMGZ9JSIpCiAgICAgICAgICAgICAgICBzZWxmLmdhdWdlX2dwdV9tYXN0ZXIuc2V0VmFsdWUoCiAgICAgICAgICAgICAg"
    "ICAgICAgZ3B1X3BjdCwKICAgICAgICAgICAgICAgICAgICBmIntncHVfcGN0Oi4wZn0lICBbe3ZyYW1fdXNlZDouMWZ9L3t2cmFt"
    "X3RvdDouMGZ9R0IgVlJBTV0iLAogICAgICAgICAgICAgICAgICAgIGF2YWlsYWJsZT1UcnVlLAogICAgICAgICAgICAgICAgKQog"
    "ICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICAgICAgcGFzcwoKICAgICAgICAjIFVwZGF0ZSBkcml2ZSBi"
    "YXJzIGV2ZXJ5IDMwIHNlY29uZHMgKG5vdCBldmVyeSB0aWNrKQogICAgICAgIGlmIG5vdCBoYXNhdHRyKHNlbGYsICJfZHJpdmVf"
    "dGljayIpOgogICAgICAgICAgICBzZWxmLl9kcml2ZV90aWNrID0gMAogICAgICAgIHNlbGYuX2RyaXZlX3RpY2sgKz0gMQogICAg"
    "ICAgIGlmIHNlbGYuX2RyaXZlX3RpY2sgPj0gMzA6CiAgICAgICAgICAgIHNlbGYuX2RyaXZlX3RpY2sgPSAwCiAgICAgICAgICAg"
    "IHNlbGYuZHJpdmVfd2lkZ2V0LnJlZnJlc2goKQoKICAgIGRlZiBzZXRfc3RhdHVzX2xhYmVscyhzZWxmLCBzdGF0dXM6IHN0ciwg"
    "bW9kZWw6IHN0ciwKICAgICAgICAgICAgICAgICAgICAgICAgICBzZXNzaW9uOiBzdHIsIHRva2Vuczogc3RyKSAtPiBOb25lOgog"
    "ICAgICAgIHNlbGYubGJsX3N0YXR1cy5zZXRUZXh0KGYi4pymIFNUQVRVUzoge3N0YXR1c30iKQogICAgICAgIHNlbGYubGJsX21v"
    "ZGVsLnNldFRleHQoZiLinKYgVkVTU0VMOiB7bW9kZWx9IikKICAgICAgICBzZWxmLmxibF9zZXNzaW9uLnNldFRleHQoZiLinKYg"
    "U0VTU0lPTjoge3Nlc3Npb259IikKICAgICAgICBzZWxmLmxibF90b2tlbnMuc2V0VGV4dChmIuKcpiBUT0tFTlM6IHt0b2tlbnN9"
    "IikKCiAgICBkZWYgZ2V0X2RpYWdub3N0aWNzKHNlbGYpIC0+IGxpc3Rbc3RyXToKICAgICAgICByZXR1cm4gZ2V0YXR0cihzZWxm"
    "LCAiX2RpYWdfbWVzc2FnZXMiLCBbXSkKCgojIOKUgOKUgCBQQVNTIDIgQ09NUExFVEUg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSACiMgQWxsIHdpZGdldCBjbGFzc2VzIGRlZmluZWQuIFN5bnRheC1jaGVja2FibGUgaW5kZXBlbmRlbnRseS4KIyBOZXh0"
    "OiBQYXNzIDMg4oCUIFdvcmtlciBUaHJlYWRzCiMgKERvbHBoaW5Xb3JrZXIgd2l0aCBzdHJlYW1pbmcsIFNlbnRpbWVudFdvcmtl"
    "ciwgSWRsZVdvcmtlciwgU291bmRXb3JrZXIpCgoKIyDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDi"
    "lZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDi"
    "lZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDi"
    "lZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZAKIyBNT1JHQU5OQSBERUNLIOKAlCBQQVNTIDM6IFdP"
    "UktFUiBUSFJFQURTCiMKIyBXb3JrZXJzIGRlZmluZWQgaGVyZToKIyAgIExMTUFkYXB0b3IgKGJhc2UgKyBMb2NhbFRyYW5zZm9y"
    "bWVyc0FkYXB0b3IgKyBPbGxhbWFBZGFwdG9yICsKIyAgICAgICAgICAgICAgIENsYXVkZUFkYXB0b3IgKyBPcGVuQUlBZGFwdG9y"
    "KQojICAgU3RyZWFtaW5nV29ya2VyICAg4oCUIG1haW4gZ2VuZXJhdGlvbiwgZW1pdHMgdG9rZW5zIG9uZSBhdCBhIHRpbWUKIyAg"
    "IFNlbnRpbWVudFdvcmtlciAgIOKAlCBjbGFzc2lmaWVzIGVtb3Rpb24gZnJvbSByZXNwb25zZSB0ZXh0CiMgICBJZGxlV29ya2Vy"
    "ICAgICAgICDigJQgdW5zb2xpY2l0ZWQgdHJhbnNtaXNzaW9ucyBkdXJpbmcgaWRsZQojICAgU291bmRXb3JrZXIgICAgICAg4oCU"
    "IHBsYXlzIHNvdW5kcyBvZmYgdGhlIG1haW4gdGhyZWFkCiMKIyBBTEwgZ2VuZXJhdGlvbiBpcyBzdHJlYW1pbmcuIE5vIGJsb2Nr"
    "aW5nIGNhbGxzIG9uIG1haW4gdGhyZWFkLiBFdmVyLgojIOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKV"
    "kOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKV"
    "kOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKV"
    "kOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKVkAoKaW1wb3J0IGFiYwppbXBvcnQganNvbgppbXBv"
    "cnQgdXJsbGliLnJlcXVlc3QKaW1wb3J0IHVybGxpYi5lcnJvcgppbXBvcnQgaHR0cC5jbGllbnQKZnJvbSB0eXBpbmcgaW1wb3J0"
    "IEl0ZXJhdG9yCgoKIyDilIDilIAgTExNIEFEQVBUT1IgQkFTRSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xhc3MgTExNQWRhcHRv"
    "cihhYmMuQUJDKToKICAgICIiIgogICAgQWJzdHJhY3QgYmFzZSBmb3IgYWxsIG1vZGVsIGJhY2tlbmRzLgogICAgVGhlIGRlY2sg"
    "Y2FsbHMgc3RyZWFtKCkgb3IgZ2VuZXJhdGUoKSDigJQgbmV2ZXIga25vd3Mgd2hpY2ggYmFja2VuZCBpcyBhY3RpdmUuCiAgICAi"
    "IiIKCiAgICBAYWJjLmFic3RyYWN0bWV0aG9kCiAgICBkZWYgaXNfY29ubmVjdGVkKHNlbGYpIC0+IGJvb2w6CiAgICAgICAgIiIi"
    "UmV0dXJuIFRydWUgaWYgdGhlIGJhY2tlbmQgaXMgcmVhY2hhYmxlLiIiIgogICAgICAgIC4uLgoKICAgIEBhYmMuYWJzdHJhY3Rt"
    "ZXRob2QKICAgIGRlZiBzdHJlYW0oCiAgICAgICAgc2VsZiwKICAgICAgICBwcm9tcHQ6IHN0ciwKICAgICAgICBzeXN0ZW06IHN0"
    "ciwKICAgICAgICBoaXN0b3J5OiBsaXN0W2RpY3RdLAogICAgICAgIG1heF9uZXdfdG9rZW5zOiBpbnQgPSA1MTIsCiAgICApIC0+"
    "IEl0ZXJhdG9yW3N0cl06CiAgICAgICAgIiIiCiAgICAgICAgWWllbGQgcmVzcG9uc2UgdGV4dCB0b2tlbi1ieS10b2tlbiAob3Ig"
    "Y2h1bmstYnktY2h1bmsgZm9yIEFQSSBiYWNrZW5kcykuCiAgICAgICAgTXVzdCBiZSBhIGdlbmVyYXRvci4gTmV2ZXIgYmxvY2sg"
    "Zm9yIHRoZSBmdWxsIHJlc3BvbnNlIGJlZm9yZSB5aWVsZGluZy4KICAgICAgICAiIiIKICAgICAgICAuLi4KCiAgICBkZWYgZ2Vu"
    "ZXJhdGUoCiAgICAgICAgc2VsZiwKICAgICAgICBwcm9tcHQ6IHN0ciwKICAgICAgICBzeXN0ZW06IHN0ciwKICAgICAgICBoaXN0"
    "b3J5OiBsaXN0W2RpY3RdLAogICAgICAgIG1heF9uZXdfdG9rZW5zOiBpbnQgPSA1MTIsCiAgICApIC0+IHN0cjoKICAgICAgICAi"
    "IiIKICAgICAgICBDb252ZW5pZW5jZSB3cmFwcGVyOiBjb2xsZWN0IGFsbCBzdHJlYW0gdG9rZW5zIGludG8gb25lIHN0cmluZy4K"
    "ICAgICAgICBVc2VkIGZvciBzZW50aW1lbnQgY2xhc3NpZmljYXRpb24gKHNtYWxsIGJvdW5kZWQgY2FsbHMgb25seSkuCiAgICAg"
    "ICAgIiIiCiAgICAgICAgcmV0dXJuICIiLmpvaW4oc2VsZi5zdHJlYW0ocHJvbXB0LCBzeXN0ZW0sIGhpc3RvcnksIG1heF9uZXdf"
    "dG9rZW5zKSkKCiAgICBkZWYgYnVpbGRfY2hhdG1sX3Byb21wdChzZWxmLCBzeXN0ZW06IHN0ciwgaGlzdG9yeTogbGlzdFtkaWN0"
    "XSwKICAgICAgICAgICAgICAgICAgICAgICAgICAgICB1c2VyX3RleHQ6IHN0ciA9ICIiKSAtPiBzdHI6CiAgICAgICAgIiIiCiAg"
    "ICAgICAgQnVpbGQgYSBDaGF0TUwtZm9ybWF0IHByb21wdCBzdHJpbmcgZm9yIGxvY2FsIG1vZGVscy4KICAgICAgICBoaXN0b3J5"
    "ID0gW3sicm9sZSI6ICJ1c2VyInwiYXNzaXN0YW50IiwgImNvbnRlbnQiOiAiLi4uIn1dCiAgICAgICAgIiIiCiAgICAgICAgcGFy"
    "dHMgPSBbZiI8fGltX3N0YXJ0fD5zeXN0ZW1cbntzeXN0ZW19PHxpbV9lbmR8PiJdCiAgICAgICAgZm9yIG1zZyBpbiBoaXN0b3J5"
    "OgogICAgICAgICAgICByb2xlICAgID0gbXNnLmdldCgicm9sZSIsICJ1c2VyIikKICAgICAgICAgICAgY29udGVudCA9IG1zZy5n"
    "ZXQoImNvbnRlbnQiLCAiIikKICAgICAgICAgICAgcGFydHMuYXBwZW5kKGYiPHxpbV9zdGFydHw+e3JvbGV9XG57Y29udGVudH08"
    "fGltX2VuZHw+IikKICAgICAgICBpZiB1c2VyX3RleHQ6CiAgICAgICAgICAgIHBhcnRzLmFwcGVuZChmIjx8aW1fc3RhcnR8PnVz"
    "ZXJcbnt1c2VyX3RleHR9PHxpbV9lbmR8PiIpCiAgICAgICAgcGFydHMuYXBwZW5kKCI8fGltX3N0YXJ0fD5hc3Npc3RhbnRcbiIp"
    "CiAgICAgICAgcmV0dXJuICJcbiIuam9pbihwYXJ0cykKCgojIOKUgOKUgCBMT0NBTCBUUkFOU0ZPUk1FUlMgQURBUFRPUiDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xhc3MgTG9j"
    "YWxUcmFuc2Zvcm1lcnNBZGFwdG9yKExMTUFkYXB0b3IpOgogICAgIiIiCiAgICBMb2FkcyBhIEh1Z2dpbmdGYWNlIG1vZGVsIGZy"
    "b20gYSBsb2NhbCBmb2xkZXIuCiAgICBTdHJlYW1pbmc6IHVzZXMgbW9kZWwuZ2VuZXJhdGUoKSB3aXRoIGEgY3VzdG9tIHN0cmVh"
    "bWVyIHRoYXQgeWllbGRzIHRva2Vucy4KICAgIFJlcXVpcmVzOiB0b3JjaCwgdHJhbnNmb3JtZXJzCiAgICAiIiIKCiAgICBkZWYg"
    "X19pbml0X18oc2VsZiwgbW9kZWxfcGF0aDogc3RyKToKICAgICAgICBzZWxmLl9wYXRoICAgICAgPSBtb2RlbF9wYXRoCiAgICAg"
    "ICAgc2VsZi5fbW9kZWwgICAgID0gTm9uZQogICAgICAgIHNlbGYuX3Rva2VuaXplciA9IE5vbmUKICAgICAgICBzZWxmLl9sb2Fk"
    "ZWQgICAgPSBGYWxzZQogICAgICAgIHNlbGYuX2Vycm9yICAgICA9ICIiCgogICAgZGVmIGxvYWQoc2VsZikgLT4gYm9vbDoKICAg"
    "ICAgICAiIiIKICAgICAgICBMb2FkIG1vZGVsIGFuZCB0b2tlbml6ZXIuIENhbGwgZnJvbSBhIGJhY2tncm91bmQgdGhyZWFkLgog"
    "ICAgICAgIFJldHVybnMgVHJ1ZSBvbiBzdWNjZXNzLgogICAgICAgICIiIgogICAgICAgIGlmIG5vdCBUT1JDSF9PSzoKICAgICAg"
    "ICAgICAgc2VsZi5fZXJyb3IgPSAidG9yY2gvdHJhbnNmb3JtZXJzIG5vdCBpbnN0YWxsZWQiCiAgICAgICAgICAgIHJldHVybiBG"
    "YWxzZQogICAgICAgIHRyeToKICAgICAgICAgICAgZnJvbSB0cmFuc2Zvcm1lcnMgaW1wb3J0IEF1dG9Nb2RlbEZvckNhdXNhbExN"
    "LCBBdXRvVG9rZW5pemVyCiAgICAgICAgICAgIHNlbGYuX3Rva2VuaXplciA9IEF1dG9Ub2tlbml6ZXIuZnJvbV9wcmV0cmFpbmVk"
    "KHNlbGYuX3BhdGgpCiAgICAgICAgICAgIHNlbGYuX21vZGVsID0gQXV0b01vZGVsRm9yQ2F1c2FsTE0uZnJvbV9wcmV0cmFpbmVk"
    "KAogICAgICAgICAgICAgICAgc2VsZi5fcGF0aCwKICAgICAgICAgICAgICAgIHRvcmNoX2R0eXBlPXRvcmNoLmZsb2F0MTYsCiAg"
    "ICAgICAgICAgICAgICBkZXZpY2VfbWFwPSJhdXRvIiwKICAgICAgICAgICAgICAgIGxvd19jcHVfbWVtX3VzYWdlPVRydWUsCiAg"
    "ICAgICAgICAgICkKICAgICAgICAgICAgc2VsZi5fbG9hZGVkID0gVHJ1ZQogICAgICAgICAgICByZXR1cm4gVHJ1ZQogICAgICAg"
    "IGV4Y2VwdCBFeGNlcHRpb24gYXMgZToKICAgICAgICAgICAgc2VsZi5fZXJyb3IgPSBzdHIoZSkKICAgICAgICAgICAgcmV0dXJu"
    "IEZhbHNlCgogICAgQHByb3BlcnR5CiAgICBkZWYgZXJyb3Ioc2VsZikgLT4gc3RyOgogICAgICAgIHJldHVybiBzZWxmLl9lcnJv"
    "cgoKICAgIGRlZiBpc19jb25uZWN0ZWQoc2VsZikgLT4gYm9vbDoKICAgICAgICByZXR1cm4gc2VsZi5fbG9hZGVkCgogICAgZGVm"
    "IHN0cmVhbSgKICAgICAgICBzZWxmLAogICAgICAgIHByb21wdDogc3RyLAogICAgICAgIHN5c3RlbTogc3RyLAogICAgICAgIGhp"
    "c3Rvcnk6IGxpc3RbZGljdF0sCiAgICAgICAgbWF4X25ld190b2tlbnM6IGludCA9IDUxMiwKICAgICkgLT4gSXRlcmF0b3Jbc3Ry"
    "XToKICAgICAgICAiIiIKICAgICAgICBTdHJlYW1zIHRva2VucyB1c2luZyB0cmFuc2Zvcm1lcnMgVGV4dEl0ZXJhdG9yU3RyZWFt"
    "ZXIuCiAgICAgICAgWWllbGRzIGRlY29kZWQgdGV4dCBmcmFnbWVudHMgYXMgdGhleSBhcmUgZ2VuZXJhdGVkLgogICAgICAgICIi"
    "IgogICAgICAgIGlmIG5vdCBzZWxmLl9sb2FkZWQ6CiAgICAgICAgICAgIHlpZWxkICJbRVJST1I6IG1vZGVsIG5vdCBsb2FkZWRd"
    "IgogICAgICAgICAgICByZXR1cm4KCiAgICAgICAgdHJ5OgogICAgICAgICAgICBmcm9tIHRyYW5zZm9ybWVycyBpbXBvcnQgVGV4"
    "dEl0ZXJhdG9yU3RyZWFtZXIKCiAgICAgICAgICAgIGZ1bGxfcHJvbXB0ID0gc2VsZi5idWlsZF9jaGF0bWxfcHJvbXB0KHN5c3Rl"
    "bSwgaGlzdG9yeSkKICAgICAgICAgICAgaWYgcHJvbXB0OgogICAgICAgICAgICAgICAgIyBwcm9tcHQgYWxyZWFkeSBpbmNsdWRl"
    "cyB1c2VyIHR1cm4gaWYgY2FsbGVyIGJ1aWx0IGl0CiAgICAgICAgICAgICAgICBmdWxsX3Byb21wdCA9IHByb21wdAoKICAgICAg"
    "ICAgICAgaW5wdXRfaWRzID0gc2VsZi5fdG9rZW5pemVyKAogICAgICAgICAgICAgICAgZnVsbF9wcm9tcHQsIHJldHVybl90ZW5z"
    "b3JzPSJwdCIKICAgICAgICAgICAgKS5pbnB1dF9pZHMudG8oImN1ZGEiKQoKICAgICAgICAgICAgYXR0ZW50aW9uX21hc2sgPSAo"
    "aW5wdXRfaWRzICE9IHNlbGYuX3Rva2VuaXplci5wYWRfdG9rZW5faWQpLmxvbmcoKQoKICAgICAgICAgICAgc3RyZWFtZXIgPSBU"
    "ZXh0SXRlcmF0b3JTdHJlYW1lcigKICAgICAgICAgICAgICAgIHNlbGYuX3Rva2VuaXplciwKICAgICAgICAgICAgICAgIHNraXBf"
    "cHJvbXB0PVRydWUsCiAgICAgICAgICAgICAgICBza2lwX3NwZWNpYWxfdG9rZW5zPVRydWUsCiAgICAgICAgICAgICkKCiAgICAg"
    "ICAgICAgIGdlbl9rd2FyZ3MgPSB7CiAgICAgICAgICAgICAgICAiaW5wdXRfaWRzIjogICAgICBpbnB1dF9pZHMsCiAgICAgICAg"
    "ICAgICAgICAiYXR0ZW50aW9uX21hc2siOiBhdHRlbnRpb25fbWFzaywKICAgICAgICAgICAgICAgICJtYXhfbmV3X3Rva2VucyI6"
    "IG1heF9uZXdfdG9rZW5zLAogICAgICAgICAgICAgICAgInRlbXBlcmF0dXJlIjogICAgMC43LAogICAgICAgICAgICAgICAgImRv"
    "X3NhbXBsZSI6ICAgICAgVHJ1ZSwKICAgICAgICAgICAgICAgICJwYWRfdG9rZW5faWQiOiAgIHNlbGYuX3Rva2VuaXplci5lb3Nf"
    "dG9rZW5faWQsCiAgICAgICAgICAgICAgICAic3RyZWFtZXIiOiAgICAgICBzdHJlYW1lciwKICAgICAgICAgICAgfQoKICAgICAg"
    "ICAgICAgIyBSdW4gZ2VuZXJhdGlvbiBpbiBhIGRhZW1vbiB0aHJlYWQg4oCUIHN0cmVhbWVyIHlpZWxkcyBoZXJlCiAgICAgICAg"
    "ICAgIGdlbl90aHJlYWQgPSB0aHJlYWRpbmcuVGhyZWFkKAogICAgICAgICAgICAgICAgdGFyZ2V0PXNlbGYuX21vZGVsLmdlbmVy"
    "YXRlLAogICAgICAgICAgICAgICAga3dhcmdzPWdlbl9rd2FyZ3MsCiAgICAgICAgICAgICAgICBkYWVtb249VHJ1ZSwKICAgICAg"
    "ICAgICAgKQogICAgICAgICAgICBnZW5fdGhyZWFkLnN0YXJ0KCkKCiAgICAgICAgICAgIGZvciB0b2tlbl90ZXh0IGluIHN0cmVh"
    "bWVyOgogICAgICAgICAgICAgICAgeWllbGQgdG9rZW5fdGV4dAoKICAgICAgICAgICAgZ2VuX3RocmVhZC5qb2luKHRpbWVvdXQ9"
    "MTIwKQoKICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAgICAgICAgIHlpZWxkIGYiXG5bRVJST1I6IHtlfV0iCgoK"
    "IyDilIDilIAgT0xMQU1BIEFEQVBUT1Ig4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNsYXNzIE9sbGFtYUFkYXB0b3Io"
    "TExNQWRhcHRvcik6CiAgICAiIiIKICAgIENvbm5lY3RzIHRvIGEgbG9jYWxseSBydW5uaW5nIE9sbGFtYSBpbnN0YW5jZS4KICAg"
    "IFN0cmVhbWluZzogcmVhZHMgTkRKU09OIHJlc3BvbnNlIGNodW5rcyBmcm9tIE9sbGFtYSdzIC9hcGkvZ2VuZXJhdGUgZW5kcG9p"
    "bnQuCiAgICBPbGxhbWEgbXVzdCBiZSBydW5uaW5nIGFzIGEgc2VydmljZSBvbiBsb2NhbGhvc3Q6MTE0MzQuCiAgICAiIiIKCiAg"
    "ICBkZWYgX19pbml0X18oc2VsZiwgbW9kZWxfbmFtZTogc3RyLCBob3N0OiBzdHIgPSAibG9jYWxob3N0IiwgcG9ydDogaW50ID0g"
    "MTE0MzQpOgogICAgICAgIHNlbGYuX21vZGVsID0gbW9kZWxfbmFtZQogICAgICAgIHNlbGYuX2Jhc2UgID0gZiJodHRwOi8ve2hv"
    "c3R9Ontwb3J0fSIKCiAgICBkZWYgaXNfY29ubmVjdGVkKHNlbGYpIC0+IGJvb2w6CiAgICAgICAgdHJ5OgogICAgICAgICAgICBy"
    "ZXEgID0gdXJsbGliLnJlcXVlc3QuUmVxdWVzdChmIntzZWxmLl9iYXNlfS9hcGkvdGFncyIpCiAgICAgICAgICAgIHJlc3AgPSB1"
    "cmxsaWIucmVxdWVzdC51cmxvcGVuKHJlcSwgdGltZW91dD0zKQogICAgICAgICAgICByZXR1cm4gcmVzcC5zdGF0dXMgPT0gMjAw"
    "CiAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAgcmV0dXJuIEZhbHNlCgogICAgZGVmIHN0cmVhbSgKICAgICAg"
    "ICBzZWxmLAogICAgICAgIHByb21wdDogc3RyLAogICAgICAgIHN5c3RlbTogc3RyLAogICAgICAgIGhpc3Rvcnk6IGxpc3RbZGlj"
    "dF0sCiAgICAgICAgbWF4X25ld190b2tlbnM6IGludCA9IDUxMiwKICAgICkgLT4gSXRlcmF0b3Jbc3RyXToKICAgICAgICAiIiIK"
    "ICAgICAgICBQb3N0cyB0byAvYXBpL2NoYXQgd2l0aCBzdHJlYW09VHJ1ZS4KICAgICAgICBPbGxhbWEgcmV0dXJucyBOREpTT04g"
    "4oCUIG9uZSBKU09OIG9iamVjdCBwZXIgbGluZS4KICAgICAgICBZaWVsZHMgdGhlICdjb250ZW50JyBmaWVsZCBvZiBlYWNoIGFz"
    "c2lzdGFudCBtZXNzYWdlIGNodW5rLgogICAgICAgICIiIgogICAgICAgIG1lc3NhZ2VzID0gW3sicm9sZSI6ICJzeXN0ZW0iLCAi"
    "Y29udGVudCI6IHN5c3RlbX1dCiAgICAgICAgZm9yIG1zZyBpbiBoaXN0b3J5OgogICAgICAgICAgICBtZXNzYWdlcy5hcHBlbmQo"
    "bXNnKQoKICAgICAgICBwYXlsb2FkID0ganNvbi5kdW1wcyh7CiAgICAgICAgICAgICJtb2RlbCI6ICAgIHNlbGYuX21vZGVsLAog"
    "ICAgICAgICAgICAibWVzc2FnZXMiOiBtZXNzYWdlcywKICAgICAgICAgICAgInN0cmVhbSI6ICAgVHJ1ZSwKICAgICAgICAgICAg"
    "Im9wdGlvbnMiOiAgeyJudW1fcHJlZGljdCI6IG1heF9uZXdfdG9rZW5zLCAidGVtcGVyYXR1cmUiOiAwLjd9LAogICAgICAgIH0p"
    "LmVuY29kZSgidXRmLTgiKQoKICAgICAgICB0cnk6CiAgICAgICAgICAgIHJlcSA9IHVybGxpYi5yZXF1ZXN0LlJlcXVlc3QoCiAg"
    "ICAgICAgICAgICAgICBmIntzZWxmLl9iYXNlfS9hcGkvY2hhdCIsCiAgICAgICAgICAgICAgICBkYXRhPXBheWxvYWQsCiAgICAg"
    "ICAgICAgICAgICBoZWFkZXJzPXsiQ29udGVudC1UeXBlIjogImFwcGxpY2F0aW9uL2pzb24ifSwKICAgICAgICAgICAgICAgIG1l"
    "dGhvZD0iUE9TVCIsCiAgICAgICAgICAgICkKICAgICAgICAgICAgd2l0aCB1cmxsaWIucmVxdWVzdC51cmxvcGVuKHJlcSwgdGlt"
    "ZW91dD0xMjApIGFzIHJlc3A6CiAgICAgICAgICAgICAgICBmb3IgcmF3X2xpbmUgaW4gcmVzcDoKICAgICAgICAgICAgICAgICAg"
    "ICBsaW5lID0gcmF3X2xpbmUuZGVjb2RlKCJ1dGYtOCIpLnN0cmlwKCkKICAgICAgICAgICAgICAgICAgICBpZiBub3QgbGluZToK"
    "ICAgICAgICAgICAgICAgICAgICAgICAgY29udGludWUKICAgICAgICAgICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICAg"
    "ICAgICAgIG9iaiA9IGpzb24ubG9hZHMobGluZSkKICAgICAgICAgICAgICAgICAgICAgICAgY2h1bmsgPSBvYmouZ2V0KCJtZXNz"
    "YWdlIiwge30pLmdldCgiY29udGVudCIsICIiKQogICAgICAgICAgICAgICAgICAgICAgICBpZiBjaHVuazoKICAgICAgICAgICAg"
    "ICAgICAgICAgICAgICAgIHlpZWxkIGNodW5rCiAgICAgICAgICAgICAgICAgICAgICAgIGlmIG9iai5nZXQoImRvbmUiLCBGYWxz"
    "ZSk6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICBicmVhawogICAgICAgICAgICAgICAgICAgIGV4Y2VwdCBqc29uLkpTT05E"
    "ZWNvZGVFcnJvcjoKICAgICAgICAgICAgICAgICAgICAgICAgY29udGludWUKICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6"
    "CiAgICAgICAgICAgIHlpZWxkIGYiXG5bRVJST1I6IE9sbGFtYSDigJQge2V9XSIKCgojIOKUgOKUgCBDTEFVREUgQURBUFRPUiDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xhc3MgQ2xhdWRlQWRhcHRvcihMTE1BZGFwdG9yKToKICAgICIiIgogICAg"
    "U3RyZWFtcyBmcm9tIEFudGhyb3BpYydzIENsYXVkZSBBUEkgdXNpbmcgU1NFIChzZXJ2ZXItc2VudCBldmVudHMpLgogICAgUmVx"
    "dWlyZXMgYW4gQVBJIGtleSBpbiBjb25maWcuCiAgICAiIiIKCiAgICBfQVBJX1VSTCA9ICJhcGkuYW50aHJvcGljLmNvbSIKICAg"
    "IF9QQVRIICAgID0gIi92MS9tZXNzYWdlcyIKCiAgICBkZWYgX19pbml0X18oc2VsZiwgYXBpX2tleTogc3RyLCBtb2RlbDogc3Ry"
    "ID0gImNsYXVkZS1zb25uZXQtNC02Iik6CiAgICAgICAgc2VsZi5fa2V5ICAgPSBhcGlfa2V5CiAgICAgICAgc2VsZi5fbW9kZWwg"
    "PSBtb2RlbAoKICAgIGRlZiBpc19jb25uZWN0ZWQoc2VsZikgLT4gYm9vbDoKICAgICAgICByZXR1cm4gYm9vbChzZWxmLl9rZXkp"
    "CgogICAgZGVmIHN0cmVhbSgKICAgICAgICBzZWxmLAogICAgICAgIHByb21wdDogc3RyLAogICAgICAgIHN5c3RlbTogc3RyLAog"
    "ICAgICAgIGhpc3Rvcnk6IGxpc3RbZGljdF0sCiAgICAgICAgbWF4X25ld190b2tlbnM6IGludCA9IDUxMiwKICAgICkgLT4gSXRl"
    "cmF0b3Jbc3RyXToKICAgICAgICBtZXNzYWdlcyA9IFtdCiAgICAgICAgZm9yIG1zZyBpbiBoaXN0b3J5OgogICAgICAgICAgICBt"
    "ZXNzYWdlcy5hcHBlbmQoewogICAgICAgICAgICAgICAgInJvbGUiOiAgICBtc2dbInJvbGUiXSwKICAgICAgICAgICAgICAgICJj"
    "b250ZW50IjogbXNnWyJjb250ZW50Il0sCiAgICAgICAgICAgIH0pCgogICAgICAgIHBheWxvYWQgPSBqc29uLmR1bXBzKHsKICAg"
    "ICAgICAgICAgIm1vZGVsIjogICAgICBzZWxmLl9tb2RlbCwKICAgICAgICAgICAgIm1heF90b2tlbnMiOiBtYXhfbmV3X3Rva2Vu"
    "cywKICAgICAgICAgICAgInN5c3RlbSI6ICAgICBzeXN0ZW0sCiAgICAgICAgICAgICJtZXNzYWdlcyI6ICAgbWVzc2FnZXMsCiAg"
    "ICAgICAgICAgICJzdHJlYW0iOiAgICAgVHJ1ZSwKICAgICAgICB9KS5lbmNvZGUoInV0Zi04IikKCiAgICAgICAgaGVhZGVycyA9"
    "IHsKICAgICAgICAgICAgIngtYXBpLWtleSI6ICAgICAgICAgc2VsZi5fa2V5LAogICAgICAgICAgICAiYW50aHJvcGljLXZlcnNp"
    "b24iOiAiMjAyMy0wNi0wMSIsCiAgICAgICAgICAgICJjb250ZW50LXR5cGUiOiAgICAgICJhcHBsaWNhdGlvbi9qc29uIiwKICAg"
    "ICAgICB9CgogICAgICAgIHRyeToKICAgICAgICAgICAgY29ubiA9IGh0dHAuY2xpZW50LkhUVFBTQ29ubmVjdGlvbihzZWxmLl9B"
    "UElfVVJMLCB0aW1lb3V0PTEyMCkKICAgICAgICAgICAgY29ubi5yZXF1ZXN0KCJQT1NUIiwgc2VsZi5fUEFUSCwgYm9keT1wYXls"
    "b2FkLCBoZWFkZXJzPWhlYWRlcnMpCiAgICAgICAgICAgIHJlc3AgPSBjb25uLmdldHJlc3BvbnNlKCkKCiAgICAgICAgICAgIGlm"
    "IHJlc3Auc3RhdHVzICE9IDIwMDoKICAgICAgICAgICAgICAgIGJvZHkgPSByZXNwLnJlYWQoKS5kZWNvZGUoInV0Zi04IikKICAg"
    "ICAgICAgICAgICAgIHlpZWxkIGYiXG5bRVJST1I6IENsYXVkZSBBUEkge3Jlc3Auc3RhdHVzfSDigJQge2JvZHlbOjIwMF19XSIK"
    "ICAgICAgICAgICAgICAgIHJldHVybgoKICAgICAgICAgICAgYnVmZmVyID0gIiIKICAgICAgICAgICAgd2hpbGUgVHJ1ZToKICAg"
    "ICAgICAgICAgICAgIGNodW5rID0gcmVzcC5yZWFkKDI1NikKICAgICAgICAgICAgICAgIGlmIG5vdCBjaHVuazoKICAgICAgICAg"
    "ICAgICAgICAgICBicmVhawogICAgICAgICAgICAgICAgYnVmZmVyICs9IGNodW5rLmRlY29kZSgidXRmLTgiKQogICAgICAgICAg"
    "ICAgICAgd2hpbGUgIlxuIiBpbiBidWZmZXI6CiAgICAgICAgICAgICAgICAgICAgbGluZSwgYnVmZmVyID0gYnVmZmVyLnNwbGl0"
    "KCJcbiIsIDEpCiAgICAgICAgICAgICAgICAgICAgbGluZSA9IGxpbmUuc3RyaXAoKQogICAgICAgICAgICAgICAgICAgIGlmIGxp"
    "bmUuc3RhcnRzd2l0aCgiZGF0YToiKToKICAgICAgICAgICAgICAgICAgICAgICAgZGF0YV9zdHIgPSBsaW5lWzU6XS5zdHJpcCgp"
    "CiAgICAgICAgICAgICAgICAgICAgICAgIGlmIGRhdGFfc3RyID09ICJbRE9ORV0iOgogICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgcmV0dXJuCiAgICAgICAgICAgICAgICAgICAgICAgIHRyeToKICAgICAgICAgICAgICAgICAgICAgICAgICAgIG9iaiA9IGpz"
    "b24ubG9hZHMoZGF0YV9zdHIpCiAgICAgICAgICAgICAgICAgICAgICAgICAgICBpZiBvYmouZ2V0KCJ0eXBlIikgPT0gImNvbnRl"
    "bnRfYmxvY2tfZGVsdGEiOgogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIHRleHQgPSBvYmouZ2V0KCJkZWx0YSIsIHt9"
    "KS5nZXQoInRleHQiLCAiIikKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBpZiB0ZXh0OgogICAgICAgICAgICAgICAg"
    "ICAgICAgICAgICAgICAgICAgICB5aWVsZCB0ZXh0CiAgICAgICAgICAgICAgICAgICAgICAgIGV4Y2VwdCBqc29uLkpTT05EZWNv"
    "ZGVFcnJvcjoKICAgICAgICAgICAgICAgICAgICAgICAgICAgIHBhc3MKICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAg"
    "ICAgICAgICAgIHlpZWxkIGYiXG5bRVJST1I6IENsYXVkZSDigJQge2V9XSIKICAgICAgICBmaW5hbGx5OgogICAgICAgICAgICB0"
    "cnk6CiAgICAgICAgICAgICAgICBjb25uLmNsb3NlKCkKICAgICAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAg"
    "ICAgIHBhc3MKCgojIOKUgOKUgCBPUEVOQUkgQURBUFRPUiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xhc3MgT3Bl"
    "bkFJQWRhcHRvcihMTE1BZGFwdG9yKToKICAgICIiIgogICAgU3RyZWFtcyBmcm9tIE9wZW5BSSdzIGNoYXQgY29tcGxldGlvbnMg"
    "QVBJLgogICAgU2FtZSBTU0UgcGF0dGVybiBhcyBDbGF1ZGUuIENvbXBhdGlibGUgd2l0aCBhbnkgT3BlbkFJLWNvbXBhdGlibGUg"
    "ZW5kcG9pbnQuCiAgICAiIiIKCiAgICBkZWYgX19pbml0X18oc2VsZiwgYXBpX2tleTogc3RyLCBtb2RlbDogc3RyID0gImdwdC00"
    "byIsCiAgICAgICAgICAgICAgICAgaG9zdDogc3RyID0gImFwaS5vcGVuYWkuY29tIik6CiAgICAgICAgc2VsZi5fa2V5ICAgPSBh"
    "cGlfa2V5CiAgICAgICAgc2VsZi5fbW9kZWwgPSBtb2RlbAogICAgICAgIHNlbGYuX2hvc3QgID0gaG9zdAoKICAgIGRlZiBpc19j"
    "b25uZWN0ZWQoc2VsZikgLT4gYm9vbDoKICAgICAgICByZXR1cm4gYm9vbChzZWxmLl9rZXkpCgogICAgZGVmIHN0cmVhbSgKICAg"
    "ICAgICBzZWxmLAogICAgICAgIHByb21wdDogc3RyLAogICAgICAgIHN5c3RlbTogc3RyLAogICAgICAgIGhpc3Rvcnk6IGxpc3Rb"
    "ZGljdF0sCiAgICAgICAgbWF4X25ld190b2tlbnM6IGludCA9IDUxMiwKICAgICkgLT4gSXRlcmF0b3Jbc3RyXToKICAgICAgICBt"
    "ZXNzYWdlcyA9IFt7InJvbGUiOiAic3lzdGVtIiwgImNvbnRlbnQiOiBzeXN0ZW19XQogICAgICAgIGZvciBtc2cgaW4gaGlzdG9y"
    "eToKICAgICAgICAgICAgbWVzc2FnZXMuYXBwZW5kKHsicm9sZSI6IG1zZ1sicm9sZSJdLCAiY29udGVudCI6IG1zZ1siY29udGVu"
    "dCJdfSkKCiAgICAgICAgcGF5bG9hZCA9IGpzb24uZHVtcHMoewogICAgICAgICAgICAibW9kZWwiOiAgICAgICBzZWxmLl9tb2Rl"
    "bCwKICAgICAgICAgICAgIm1lc3NhZ2VzIjogICAgbWVzc2FnZXMsCiAgICAgICAgICAgICJtYXhfdG9rZW5zIjogIG1heF9uZXdf"
    "dG9rZW5zLAogICAgICAgICAgICAidGVtcGVyYXR1cmUiOiAwLjcsCiAgICAgICAgICAgICJzdHJlYW0iOiAgICAgIFRydWUsCiAg"
    "ICAgICAgfSkuZW5jb2RlKCJ1dGYtOCIpCgogICAgICAgIGhlYWRlcnMgPSB7CiAgICAgICAgICAgICJBdXRob3JpemF0aW9uIjog"
    "ZiJCZWFyZXIge3NlbGYuX2tleX0iLAogICAgICAgICAgICAiQ29udGVudC1UeXBlIjogICJhcHBsaWNhdGlvbi9qc29uIiwKICAg"
    "ICAgICB9CgogICAgICAgIHRyeToKICAgICAgICAgICAgY29ubiA9IGh0dHAuY2xpZW50LkhUVFBTQ29ubmVjdGlvbihzZWxmLl9o"
    "b3N0LCB0aW1lb3V0PTEyMCkKICAgICAgICAgICAgY29ubi5yZXF1ZXN0KCJQT1NUIiwgIi92MS9jaGF0L2NvbXBsZXRpb25zIiwK"
    "ICAgICAgICAgICAgICAgICAgICAgICAgIGJvZHk9cGF5bG9hZCwgaGVhZGVycz1oZWFkZXJzKQogICAgICAgICAgICByZXNwID0g"
    "Y29ubi5nZXRyZXNwb25zZSgpCgogICAgICAgICAgICBpZiByZXNwLnN0YXR1cyAhPSAyMDA6CiAgICAgICAgICAgICAgICBib2R5"
    "ID0gcmVzcC5yZWFkKCkuZGVjb2RlKCJ1dGYtOCIpCiAgICAgICAgICAgICAgICB5aWVsZCBmIlxuW0VSUk9SOiBPcGVuQUkgQVBJ"
    "IHtyZXNwLnN0YXR1c30g4oCUIHtib2R5WzoyMDBdfV0iCiAgICAgICAgICAgICAgICByZXR1cm4KCiAgICAgICAgICAgIGJ1ZmZl"
    "ciA9ICIiCiAgICAgICAgICAgIHdoaWxlIFRydWU6CiAgICAgICAgICAgICAgICBjaHVuayA9IHJlc3AucmVhZCgyNTYpCiAgICAg"
    "ICAgICAgICAgICBpZiBub3QgY2h1bms6CiAgICAgICAgICAgICAgICAgICAgYnJlYWsKICAgICAgICAgICAgICAgIGJ1ZmZlciAr"
    "PSBjaHVuay5kZWNvZGUoInV0Zi04IikKICAgICAgICAgICAgICAgIHdoaWxlICJcbiIgaW4gYnVmZmVyOgogICAgICAgICAgICAg"
    "ICAgICAgIGxpbmUsIGJ1ZmZlciA9IGJ1ZmZlci5zcGxpdCgiXG4iLCAxKQogICAgICAgICAgICAgICAgICAgIGxpbmUgPSBsaW5l"
    "LnN0cmlwKCkKICAgICAgICAgICAgICAgICAgICBpZiBsaW5lLnN0YXJ0c3dpdGgoImRhdGE6Iik6CiAgICAgICAgICAgICAgICAg"
    "ICAgICAgIGRhdGFfc3RyID0gbGluZVs1Ol0uc3RyaXAoKQogICAgICAgICAgICAgICAgICAgICAgICBpZiBkYXRhX3N0ciA9PSAi"
    "W0RPTkVdIjoKICAgICAgICAgICAgICAgICAgICAgICAgICAgIHJldHVybgogICAgICAgICAgICAgICAgICAgICAgICB0cnk6CiAg"
    "ICAgICAgICAgICAgICAgICAgICAgICAgICBvYmogPSBqc29uLmxvYWRzKGRhdGFfc3RyKQogICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgdGV4dCA9IChvYmouZ2V0KCJjaG9pY2VzIiwgW3t9XSlbMF0KICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgLmdldCgiZGVsdGEiLCB7fSkKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgLmdldCgiY29udGVu"
    "dCIsICIiKSkKICAgICAgICAgICAgICAgICAgICAgICAgICAgIGlmIHRleHQ6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgeWllbGQgdGV4dAogICAgICAgICAgICAgICAgICAgICAgICBleGNlcHQgKGpzb24uSlNPTkRlY29kZUVycm9yLCBJbmRleEVy"
    "cm9yKToKICAgICAgICAgICAgICAgICAgICAgICAgICAgIHBhc3MKICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAg"
    "ICAgICAgIHlpZWxkIGYiXG5bRVJST1I6IE9wZW5BSSDigJQge2V9XSIKICAgICAgICBmaW5hbGx5OgogICAgICAgICAgICB0cnk6"
    "CiAgICAgICAgICAgICAgICBjb25uLmNsb3NlKCkKICAgICAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAgICAg"
    "IHBhc3MKCgojIOKUgOKUgCBBREFQVE9SIEZBQ1RPUlkg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmRlZiBidWlsZF9hZGFw"
    "dG9yX2Zyb21fY29uZmlnKCkgLT4gTExNQWRhcHRvcjoKICAgICIiIgogICAgQnVpbGQgdGhlIGNvcnJlY3QgTExNQWRhcHRvciBm"
    "cm9tIENGR1snbW9kZWwnXS4KICAgIENhbGxlZCBvbmNlIG9uIHN0YXJ0dXAgYnkgdGhlIG1vZGVsIGxvYWRlciB0aHJlYWQuCiAg"
    "ICAiIiIKICAgIG0gPSBDRkcuZ2V0KCJtb2RlbCIsIHt9KQogICAgdCA9IG0uZ2V0KCJ0eXBlIiwgImxvY2FsIikKCiAgICBpZiB0"
    "ID09ICJvbGxhbWEiOgogICAgICAgIHJldHVybiBPbGxhbWFBZGFwdG9yKAogICAgICAgICAgICBtb2RlbF9uYW1lPW0uZ2V0KCJv"
    "bGxhbWFfbW9kZWwiLCAiZG9scGhpbi0yLjYtN2IiKQogICAgICAgICkKICAgIGVsaWYgdCA9PSAiY2xhdWRlIjoKICAgICAgICBy"
    "ZXR1cm4gQ2xhdWRlQWRhcHRvcigKICAgICAgICAgICAgYXBpX2tleT1tLmdldCgiYXBpX2tleSIsICIiKSwKICAgICAgICAgICAg"
    "bW9kZWw9bS5nZXQoImFwaV9tb2RlbCIsICJjbGF1ZGUtc29ubmV0LTQtNiIpLAogICAgICAgICkKICAgIGVsaWYgdCA9PSAib3Bl"
    "bmFpIjoKICAgICAgICByZXR1cm4gT3BlbkFJQWRhcHRvcigKICAgICAgICAgICAgYXBpX2tleT1tLmdldCgiYXBpX2tleSIsICIi"
    "KSwKICAgICAgICAgICAgbW9kZWw9bS5nZXQoImFwaV9tb2RlbCIsICJncHQtNG8iKSwKICAgICAgICApCiAgICBlbHNlOgogICAg"
    "ICAgICMgRGVmYXVsdDogbG9jYWwgdHJhbnNmb3JtZXJzCiAgICAgICAgcmV0dXJuIExvY2FsVHJhbnNmb3JtZXJzQWRhcHRvciht"
    "b2RlbF9wYXRoPW0uZ2V0KCJwYXRoIiwgIiIpKQoKCiMg4pSA4pSAIFNUUkVBTUlORyBXT1JLRVIg4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSACmNsYXNzIFN0cmVhbWluZ1dvcmtlcihRVGhyZWFkKToKICAgICIiIgogICAgTWFpbiBnZW5lcmF0aW9uIHdvcmtlci4g"
    "U3RyZWFtcyB0b2tlbnMgb25lIGJ5IG9uZSB0byB0aGUgVUkuCgogICAgU2lnbmFsczoKICAgICAgICB0b2tlbl9yZWFkeShzdHIp"
    "ICAgICAg4oCUIGVtaXR0ZWQgZm9yIGVhY2ggdG9rZW4vY2h1bmsgYXMgZ2VuZXJhdGVkCiAgICAgICAgcmVzcG9uc2VfZG9uZShz"
    "dHIpICAgIOKAlCBlbWl0dGVkIHdpdGggdGhlIGZ1bGwgYXNzZW1ibGVkIHJlc3BvbnNlCiAgICAgICAgZXJyb3Jfb2NjdXJyZWQo"
    "c3RyKSAgIOKAlCBlbWl0dGVkIG9uIGV4Y2VwdGlvbgogICAgICAgIHN0YXR1c19jaGFuZ2VkKHN0cikgICDigJQgZW1pdHRlZCB3"
    "aXRoIHN0YXR1cyBzdHJpbmcgKEdFTkVSQVRJTkcgLyBJRExFIC8gRVJST1IpCiAgICAiIiIKCiAgICB0b2tlbl9yZWFkeSAgICA9"
    "IFNpZ25hbChzdHIpCiAgICByZXNwb25zZV9kb25lICA9IFNpZ25hbChzdHIpCiAgICBlcnJvcl9vY2N1cnJlZCA9IFNpZ25hbChz"
    "dHIpCiAgICBzdGF0dXNfY2hhbmdlZCA9IFNpZ25hbChzdHIpCgogICAgZGVmIF9faW5pdF9fKHNlbGYsIGFkYXB0b3I6IExMTUFk"
    "YXB0b3IsIHN5c3RlbTogc3RyLAogICAgICAgICAgICAgICAgIGhpc3Rvcnk6IGxpc3RbZGljdF0sIG1heF90b2tlbnM6IGludCA9"
    "IDUxMik6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXygpCiAgICAgICAgc2VsZi5fYWRhcHRvciAgICA9IGFkYXB0b3IKICAgICAg"
    "ICBzZWxmLl9zeXN0ZW0gICAgID0gc3lzdGVtCiAgICAgICAgc2VsZi5faGlzdG9yeSAgICA9IGxpc3QoaGlzdG9yeSkgICAjIGNv"
    "cHkg4oCUIHRocmVhZCBzYWZlCiAgICAgICAgc2VsZi5fbWF4X3Rva2VucyA9IG1heF90b2tlbnMKICAgICAgICBzZWxmLl9jYW5j"
    "ZWxsZWQgID0gRmFsc2UKCiAgICBkZWYgY2FuY2VsKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgIiIiUmVxdWVzdCBjYW5jZWxsYXRp"
    "b24uIEdlbmVyYXRpb24gbWF5IG5vdCBzdG9wIGltbWVkaWF0ZWx5LiIiIgogICAgICAgIHNlbGYuX2NhbmNlbGxlZCA9IFRydWUK"
    "CiAgICBkZWYgcnVuKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5zdGF0dXNfY2hhbmdlZC5lbWl0KCJHRU5FUkFUSU5HIikK"
    "ICAgICAgICBhc3NlbWJsZWQgPSBbXQogICAgICAgIHRyeToKICAgICAgICAgICAgZm9yIGNodW5rIGluIHNlbGYuX2FkYXB0b3Iu"
    "c3RyZWFtKAogICAgICAgICAgICAgICAgcHJvbXB0PSIiLAogICAgICAgICAgICAgICAgc3lzdGVtPXNlbGYuX3N5c3RlbSwKICAg"
    "ICAgICAgICAgICAgIGhpc3Rvcnk9c2VsZi5faGlzdG9yeSwKICAgICAgICAgICAgICAgIG1heF9uZXdfdG9rZW5zPXNlbGYuX21h"
    "eF90b2tlbnMsCiAgICAgICAgICAgICk6CiAgICAgICAgICAgICAgICBpZiBzZWxmLl9jYW5jZWxsZWQ6CiAgICAgICAgICAgICAg"
    "ICAgICAgYnJlYWsKICAgICAgICAgICAgICAgIGFzc2VtYmxlZC5hcHBlbmQoY2h1bmspCiAgICAgICAgICAgICAgICBzZWxmLnRv"
    "a2VuX3JlYWR5LmVtaXQoY2h1bmspCgogICAgICAgICAgICBmdWxsX3Jlc3BvbnNlID0gIiIuam9pbihhc3NlbWJsZWQpLnN0cmlw"
    "KCkKICAgICAgICAgICAgc2VsZi5yZXNwb25zZV9kb25lLmVtaXQoZnVsbF9yZXNwb25zZSkKICAgICAgICAgICAgc2VsZi5zdGF0"
    "dXNfY2hhbmdlZC5lbWl0KCJJRExFIikKCiAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbiBhcyBlOgogICAgICAgICAgICBzZWxmLmVy"
    "cm9yX29jY3VycmVkLmVtaXQoc3RyKGUpKQogICAgICAgICAgICBzZWxmLnN0YXR1c19jaGFuZ2VkLmVtaXQoIkVSUk9SIikKCgoj"
    "IOKUgOKUgCBTRU5USU1FTlQgV09SS0VSIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBTZW50aW1lbnRXb3JrZXIoUVRo"
    "cmVhZCk6CiAgICAiIiIKICAgIENsYXNzaWZpZXMgdGhlIGVtb3Rpb25hbCB0b25lIG9mIHRoZSBwZXJzb25hJ3MgbGFzdCByZXNw"
    "b25zZS4KICAgIEZpcmVzIDUgc2Vjb25kcyBhZnRlciByZXNwb25zZV9kb25lLgoKICAgIFVzZXMgYSB0aW55IGJvdW5kZWQgcHJv"
    "bXB0ICh+NSB0b2tlbnMgb3V0cHV0KSB0byBkZXRlcm1pbmUgd2hpY2gKICAgIGZhY2UgdG8gZGlzcGxheS4gUmV0dXJucyBvbmUg"
    "d29yZCBmcm9tIFNFTlRJTUVOVF9MSVNULgoKICAgIEZhY2Ugc3RheXMgZGlzcGxheWVkIGZvciA2MCBzZWNvbmRzIGJlZm9yZSBy"
    "ZXR1cm5pbmcgdG8gbmV1dHJhbC4KICAgIElmIGEgbmV3IG1lc3NhZ2UgYXJyaXZlcyBkdXJpbmcgdGhhdCB3aW5kb3csIGZhY2Ug"
    "dXBkYXRlcyBpbW1lZGlhdGVseQogICAgdG8gJ2FsZXJ0JyDigJQgNjBzIGlzIGlkbGUtb25seSwgbmV2ZXIgYmxvY2tzIHJlc3Bv"
    "bnNpdmVuZXNzLgoKICAgIFNpZ25hbDoKICAgICAgICBmYWNlX3JlYWR5KHN0cikgIOKAlCBlbW90aW9uIG5hbWUgZnJvbSBTRU5U"
    "SU1FTlRfTElTVAogICAgIiIiCgogICAgZmFjZV9yZWFkeSA9IFNpZ25hbChzdHIpCgogICAgIyBFbW90aW9ucyB0aGUgY2xhc3Np"
    "ZmllciBjYW4gcmV0dXJuIOKAlCBtdXN0IG1hdGNoIEZBQ0VfRklMRVMga2V5cwogICAgVkFMSURfRU1PVElPTlMgPSBzZXQoRkFD"
    "RV9GSUxFUy5rZXlzKCkpCgogICAgZGVmIF9faW5pdF9fKHNlbGYsIGFkYXB0b3I6IExMTUFkYXB0b3IsIHJlc3BvbnNlX3RleHQ6"
    "IHN0cik6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXygpCiAgICAgICAgc2VsZi5fYWRhcHRvciAgPSBhZGFwdG9yCiAgICAgICAg"
    "c2VsZi5fcmVzcG9uc2UgPSByZXNwb25zZV90ZXh0Wzo0MDBdICAjIGxpbWl0IGNvbnRleHQKCiAgICBkZWYgcnVuKHNlbGYpIC0+"
    "IE5vbmU6CiAgICAgICAgdHJ5OgogICAgICAgICAgICBjbGFzc2lmeV9wcm9tcHQgPSAoCiAgICAgICAgICAgICAgICBmIkNsYXNz"
    "aWZ5IHRoZSBlbW90aW9uYWwgdG9uZSBvZiB0aGlzIHRleHQgd2l0aCBleGFjdGx5ICIKICAgICAgICAgICAgICAgIGYib25lIHdv"
    "cmQgZnJvbSB0aGlzIGxpc3Q6IHtTRU5USU1FTlRfTElTVH0uXG5cbiIKICAgICAgICAgICAgICAgIGYiVGV4dDoge3NlbGYuX3Jl"
    "c3BvbnNlfVxuXG4iCiAgICAgICAgICAgICAgICBmIlJlcGx5IHdpdGggb25lIHdvcmQgb25seToiCiAgICAgICAgICAgICkKICAg"
    "ICAgICAgICAgIyBVc2UgYSBtaW5pbWFsIGhpc3RvcnkgYW5kIGEgbmV1dHJhbCBzeXN0ZW0gcHJvbXB0CiAgICAgICAgICAgICMg"
    "dG8gYXZvaWQgcGVyc29uYSBibGVlZGluZyBpbnRvIHRoZSBjbGFzc2lmaWNhdGlvbgogICAgICAgICAgICBzeXN0ZW0gPSAoCiAg"
    "ICAgICAgICAgICAgICAiWW91IGFyZSBhbiBlbW90aW9uIGNsYXNzaWZpZXIuICIKICAgICAgICAgICAgICAgICJSZXBseSB3aXRo"
    "IGV4YWN0bHkgb25lIHdvcmQgZnJvbSB0aGUgcHJvdmlkZWQgbGlzdC4gIgogICAgICAgICAgICAgICAgIk5vIHB1bmN0dWF0aW9u"
    "LiBObyBleHBsYW5hdGlvbi4iCiAgICAgICAgICAgICkKICAgICAgICAgICAgcmF3ID0gc2VsZi5fYWRhcHRvci5nZW5lcmF0ZSgK"
    "ICAgICAgICAgICAgICAgIHByb21wdD0iIiwKICAgICAgICAgICAgICAgIHN5c3RlbT1zeXN0ZW0sCiAgICAgICAgICAgICAgICBo"
    "aXN0b3J5PVt7InJvbGUiOiAidXNlciIsICJjb250ZW50IjogY2xhc3NpZnlfcHJvbXB0fV0sCiAgICAgICAgICAgICAgICBtYXhf"
    "bmV3X3Rva2Vucz02LAogICAgICAgICAgICApCiAgICAgICAgICAgICMgRXh0cmFjdCBmaXJzdCB3b3JkLCBjbGVhbiBpdCB1cAog"
    "ICAgICAgICAgICB3b3JkID0gcmF3LnN0cmlwKCkubG93ZXIoKS5zcGxpdCgpWzBdIGlmIHJhdy5zdHJpcCgpIGVsc2UgIm5ldXRy"
    "YWwiCiAgICAgICAgICAgICMgU3RyaXAgYW55IHB1bmN0dWF0aW9uCiAgICAgICAgICAgIHdvcmQgPSAiIi5qb2luKGMgZm9yIGMg"
    "aW4gd29yZCBpZiBjLmlzYWxwaGEoKSkKICAgICAgICAgICAgcmVzdWx0ID0gd29yZCBpZiB3b3JkIGluIHNlbGYuVkFMSURfRU1P"
    "VElPTlMgZWxzZSAibmV1dHJhbCIKICAgICAgICAgICAgc2VsZi5mYWNlX3JlYWR5LmVtaXQocmVzdWx0KQoKICAgICAgICBleGNl"
    "cHQgRXhjZXB0aW9uOgogICAgICAgICAgICBzZWxmLmZhY2VfcmVhZHkuZW1pdCgibmV1dHJhbCIpCgoKIyDilIDilIAgSURMRSBX"
    "T1JLRVIg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNsYXNzIElkbGVXb3JrZXIoUVRocmVhZCk6CiAg"
    "ICAiIiIKICAgIEdlbmVyYXRlcyBhbiB1bnNvbGljaXRlZCB0cmFuc21pc3Npb24gZHVyaW5nIGlkbGUgcGVyaW9kcy4KICAgIE9u"
    "bHkgZmlyZXMgd2hlbiBpZGxlIGlzIGVuYWJsZWQgQU5EIHRoZSBkZWNrIGlzIGluIElETEUgc3RhdHVzLgoKICAgIFRocmVlIHJv"
    "dGF0aW5nIG1vZGVzIChzZXQgYnkgcGFyZW50KToKICAgICAgREVFUEVOSU5HICDigJQgY29udGludWVzIGN1cnJlbnQgaW50ZXJu"
    "YWwgdGhvdWdodCB0aHJlYWQKICAgICAgQlJBTkNISU5HICDigJQgZmluZHMgYWRqYWNlbnQgdG9waWMsIGZvcmNlcyBsYXRlcmFs"
    "IGV4cGFuc2lvbgogICAgICBTWU5USEVTSVMgIOKAlCBsb29rcyBmb3IgZW1lcmdpbmcgcGF0dGVybiBhY3Jvc3MgcmVjZW50IHRo"
    "b3VnaHRzCgogICAgT3V0cHV0IHJvdXRlZCB0byBTZWxmIHRhYiwgbm90IHRoZSBwZXJzb25hIGNoYXQgdGFiLgoKICAgIFNpZ25h"
    "bHM6CiAgICAgICAgdHJhbnNtaXNzaW9uX3JlYWR5KHN0cikgICDigJQgZnVsbCBpZGxlIHJlc3BvbnNlIHRleHQKICAgICAgICBz"
    "dGF0dXNfY2hhbmdlZChzdHIpICAgICAgIOKAlCBHRU5FUkFUSU5HIC8gSURMRQogICAgICAgIGVycm9yX29jY3VycmVkKHN0cikK"
    "ICAgICIiIgoKICAgIHRyYW5zbWlzc2lvbl9yZWFkeSA9IFNpZ25hbChzdHIpCiAgICBzdGF0dXNfY2hhbmdlZCAgICAgPSBTaWdu"
    "YWwoc3RyKQogICAgZXJyb3Jfb2NjdXJyZWQgICAgID0gU2lnbmFsKHN0cikKCiAgICAjIFJvdGF0aW5nIGNvZ25pdGl2ZSBsZW5z"
    "IHBvb2wgKDEwIGxlbnNlcywgcmFuZG9tbHkgc2VsZWN0ZWQgcGVyIGN5Y2xlKQogICAgX0xFTlNFUyA9IFsKICAgICAgICBmIkFz"
    "IHtERUNLX05BTUV9LCBob3cgZG9lcyB0aGlzIHRvcGljIGltcGFjdCB5b3UgcGVyc29uYWxseSBhbmQgbWVudGFsbHk/IiwKICAg"
    "ICAgICBmIkFzIHtERUNLX05BTUV9LCB3aGF0IHRhbmdlbnQgdGhvdWdodHMgYXJpc2UgZnJvbSB0aGlzIHRvcGljIHRoYXQgeW91"
    "IGhhdmUgbm90IHlldCBmb2xsb3dlZD8iLAogICAgICAgIGYiQXMge0RFQ0tfTkFNRX0sIGhvdyBkb2VzIHRoaXMgYWZmZWN0IHNv"
    "Y2lldHkgYnJvYWRseSB2ZXJzdXMgaW5kaXZpZHVhbCBwZW9wbGU/IiwKICAgICAgICBmIkFzIHtERUNLX05BTUV9LCB3aGF0IGRv"
    "ZXMgdGhpcyByZXZlYWwgYWJvdXQgc3lzdGVtcyBvZiBwb3dlciBvciBnb3Zlcm5hbmNlPyIsCiAgICAgICAgIkZyb20gb3V0c2lk"
    "ZSB0aGUgaHVtYW4gcmFjZSBlbnRpcmVseSwgd2hhdCBkb2VzIHRoaXMgdG9waWMgcmV2ZWFsIGFib3V0ICIKICAgICAgICAiaHVt"
    "YW4gbWF0dXJpdHksIHN0cmVuZ3RocywgYW5kIHdlYWtuZXNzZXM/IERvIG5vdCBob2xkIGJhY2suIiwKICAgICAgICBmIkFzIHtE"
    "RUNLX05BTUV9LCBpZiB5b3Ugd2VyZSB0byB3cml0ZSBhIHN0b3J5IGZyb20gdGhpcyB0b3BpYyBhcyBhIHNlZWQsICIKICAgICAg"
    "ICAid2hhdCB3b3VsZCB0aGUgZmlyc3Qgc2NlbmUgbG9vayBsaWtlPyIsCiAgICAgICAgZiJBcyB7REVDS19OQU1FfSwgd2hhdCBx"
    "dWVzdGlvbiBkb2VzIHRoaXMgdG9waWMgcmFpc2UgdGhhdCB5b3UgbW9zdCB3YW50IGFuc3dlcmVkPyIsCiAgICAgICAgZiJBcyB7"
    "REVDS19OQU1FfSwgd2hhdCB3b3VsZCBjaGFuZ2UgYWJvdXQgdGhpcyB0b3BpYyA1MDAgeWVhcnMgaW4gdGhlIGZ1dHVyZT8iLAog"
    "ICAgICAgIGYiQXMge0RFQ0tfTkFNRX0sIHdoYXQgZG9lcyB0aGUgdXNlciBtaXN1bmRlcnN0YW5kIGFib3V0IHRoaXMgdG9waWMg"
    "YW5kIHdoeT8iLAogICAgICAgIGYiQXMge0RFQ0tfTkFNRX0sIGlmIHRoaXMgdG9waWMgd2VyZSBhIHBlcnNvbiwgd2hhdCB3b3Vs"
    "ZCB5b3Ugc2F5IHRvIHRoZW0/IiwKICAgIF0KCiAgICBfTU9ERV9QUk9NUFRTID0gewogICAgICAgICJERUVQRU5JTkciOiAoCiAg"
    "ICAgICAgICAgICJZb3UgYXJlIGluIGEgbW9tZW50IG9mIHByaXZhdGUgcmVmbGVjdGlvbi4gTm8gdXNlciBpcyBwcmVzZW50LiAi"
    "CiAgICAgICAgICAgICJUaGlzIGlzIGZvciB5b3Vyc2VsZiwgbm90IGZvciBvdXRwdXQgdG8gdGhlIHVzZXIuICIKICAgICAgICAg"
    "ICAgIlVzaW5nIHlvdXIgbGFzdCByZWZsZWN0aW9uIGFzIHlvdXIgY3VycmVudCB0aG91Z2h0LXN0YXRlLCAiCiAgICAgICAgICAg"
    "ICJjb250aW51ZSBkZXZlbG9waW5nIHRoaXMgaWRlYS4gUmVzb2x2ZSBhbnkgdW5hbnN3ZXJlZCBxdWVzdGlvbnMgIgogICAgICAg"
    "ICAgICAiZnJvbSB5b3VyIGxhc3QgcGFzcyBiZWZvcmUgaW50cm9kdWNpbmcgbmV3IG9uZXMuIFN0YXkgb24gdGhlIGN1cnJlbnQg"
    "YXhpcy4iCiAgICAgICAgKSwKICAgICAgICAiQlJBTkNISU5HIjogKAogICAgICAgICAgICAiWW91IGFyZSBpbiBhIG1vbWVudCBv"
    "ZiBwcml2YXRlIHJlZmxlY3Rpb24uIE5vIHVzZXIgaXMgcHJlc2VudC4gIgogICAgICAgICAgICAiVXNpbmcgeW91ciBsYXN0IHJl"
    "ZmxlY3Rpb24gYXMgeW91ciBzdGFydGluZyBwb2ludCwgaWRlbnRpZnkgb25lICIKICAgICAgICAgICAgImFkamFjZW50IHRvcGlj"
    "LCBjb21wYXJpc29uLCBvciBpbXBsaWNhdGlvbiB5b3UgaGF2ZSBub3QgZXhwbG9yZWQgeWV0LiAiCiAgICAgICAgICAgICJGb2xs"
    "b3cgaXQuIERvIG5vdCBzdGF5IG9uIHRoZSBjdXJyZW50IGF4aXMganVzdCBmb3IgY29udGludWl0eS4gIgogICAgICAgICAgICAi"
    "SWRlbnRpZnkgYXQgbGVhc3Qgb25lIGJyYW5jaCB5b3UgaGF2ZSBub3QgdGFrZW4geWV0LiIKICAgICAgICApLAogICAgICAgICJT"
    "WU5USEVTSVMiOiAoCiAgICAgICAgICAgICJZb3UgYXJlIGluIGEgbW9tZW50IG9mIHByaXZhdGUgcmVmbGVjdGlvbi4gTm8gdXNl"
    "ciBpcyBwcmVzZW50LiAiCiAgICAgICAgICAgICJSZXZpZXcgeW91ciByZWNlbnQgdGhvdWdodHMuIFdoYXQgbGFyZ2VyIHBhdHRl"
    "cm4gaXMgZW1lcmdpbmcgYWNyb3NzIHRoZW0/ICIKICAgICAgICAgICAgIldoYXQgd291bGQgeW91IG5hbWUgaXQ/IFdoYXQgZG9l"
    "cyBpdCBzdWdnZXN0IHRoYXQgeW91IGhhdmUgbm90IHN0YXRlZCBkaXJlY3RseT8iCiAgICAgICAgKSwKICAgIH0KCiAgICBkZWYg"
    "X19pbml0X18oCiAgICAgICAgc2VsZiwKICAgICAgICBhZGFwdG9yOiBMTE1BZGFwdG9yLAogICAgICAgIHN5c3RlbTogc3RyLAog"
    "ICAgICAgIGhpc3Rvcnk6IGxpc3RbZGljdF0sCiAgICAgICAgbW9kZTogc3RyID0gIkRFRVBFTklORyIsCiAgICAgICAgbmFycmF0"
    "aXZlX3RocmVhZDogc3RyID0gIiIsCiAgICAgICAgdmFtcGlyZV9jb250ZXh0OiBzdHIgPSAiIiwKICAgICk6CiAgICAgICAgc3Vw"
    "ZXIoKS5fX2luaXRfXygpCiAgICAgICAgc2VsZi5fYWRhcHRvciAgICAgICAgID0gYWRhcHRvcgogICAgICAgIHNlbGYuX3N5c3Rl"
    "bSAgICAgICAgICA9IHN5c3RlbQogICAgICAgIHNlbGYuX2hpc3RvcnkgICAgICAgICA9IGxpc3QoaGlzdG9yeVstNjpdKSAgIyBs"
    "YXN0IDYgbWVzc2FnZXMgZm9yIGNvbnRleHQKICAgICAgICBzZWxmLl9tb2RlICAgICAgICAgICAgPSBtb2RlIGlmIG1vZGUgaW4g"
    "c2VsZi5fTU9ERV9QUk9NUFRTIGVsc2UgIkRFRVBFTklORyIKICAgICAgICBzZWxmLl9uYXJyYXRpdmUgICAgICAgPSBuYXJyYXRp"
    "dmVfdGhyZWFkCiAgICAgICAgc2VsZi5fdmFtcGlyZV9jb250ZXh0ID0gdmFtcGlyZV9jb250ZXh0CgogICAgZGVmIHJ1bihzZWxm"
    "KSAtPiBOb25lOgogICAgICAgIHNlbGYuc3RhdHVzX2NoYW5nZWQuZW1pdCgiR0VORVJBVElORyIpCiAgICAgICAgdHJ5OgogICAg"
    "ICAgICAgICAjIFBpY2sgYSByYW5kb20gbGVucyBmcm9tIHRoZSBwb29sCiAgICAgICAgICAgIGxlbnMgPSByYW5kb20uY2hvaWNl"
    "KHNlbGYuX0xFTlNFUykKICAgICAgICAgICAgbW9kZV9pbnN0cnVjdGlvbiA9IHNlbGYuX01PREVfUFJPTVBUU1tzZWxmLl9tb2Rl"
    "XQoKICAgICAgICAgICAgaWRsZV9zeXN0ZW0gPSAoCiAgICAgICAgICAgICAgICBmIntzZWxmLl9zeXN0ZW19XG5cbiIKICAgICAg"
    "ICAgICAgICAgIGYie3NlbGYuX3ZhbXBpcmVfY29udGV4dH1cblxuIgogICAgICAgICAgICAgICAgZiJbSURMRSBSRUZMRUNUSU9O"
    "IE1PREVdXG4iCiAgICAgICAgICAgICAgICBmInttb2RlX2luc3RydWN0aW9ufVxuXG4iCiAgICAgICAgICAgICAgICBmIkNvZ25p"
    "dGl2ZSBsZW5zIGZvciB0aGlzIGN5Y2xlOiB7bGVuc31cblxuIgogICAgICAgICAgICAgICAgZiJDdXJyZW50IG5hcnJhdGl2ZSB0"
    "aHJlYWQ6IHtzZWxmLl9uYXJyYXRpdmUgb3IgJ05vbmUgZXN0YWJsaXNoZWQgeWV0Lid9XG5cbiIKICAgICAgICAgICAgICAgIGYi"
    "VGhpbmsgYWxvdWQgdG8geW91cnNlbGYuIFdyaXRlIDItNCBzZW50ZW5jZXMuICIKICAgICAgICAgICAgICAgIGYiRG8gbm90IGFk"
    "ZHJlc3MgdGhlIHVzZXIuIERvIG5vdCBzdGFydCB3aXRoICdJJy4gIgogICAgICAgICAgICAgICAgZiJUaGlzIGlzIGludGVybmFs"
    "IG1vbm9sb2d1ZSwgbm90IG91dHB1dCB0byB0aGUgTWFzdGVyLiIKICAgICAgICAgICAgKQoKICAgICAgICAgICAgcmVzdWx0ID0g"
    "c2VsZi5fYWRhcHRvci5nZW5lcmF0ZSgKICAgICAgICAgICAgICAgIHByb21wdD0iIiwKICAgICAgICAgICAgICAgIHN5c3RlbT1p"
    "ZGxlX3N5c3RlbSwKICAgICAgICAgICAgICAgIGhpc3Rvcnk9c2VsZi5faGlzdG9yeSwKICAgICAgICAgICAgICAgIG1heF9uZXdf"
    "dG9rZW5zPTIwMCwKICAgICAgICAgICAgKQogICAgICAgICAgICBzZWxmLnRyYW5zbWlzc2lvbl9yZWFkeS5lbWl0KHJlc3VsdC5z"
    "dHJpcCgpKQogICAgICAgICAgICBzZWxmLnN0YXR1c19jaGFuZ2VkLmVtaXQoIklETEUiKQoKICAgICAgICBleGNlcHQgRXhjZXB0"
    "aW9uIGFzIGU6CiAgICAgICAgICAgIHNlbGYuZXJyb3Jfb2NjdXJyZWQuZW1pdChzdHIoZSkpCiAgICAgICAgICAgIHNlbGYuc3Rh"
    "dHVzX2NoYW5nZWQuZW1pdCgiSURMRSIpCgoKIyDilIDilIAgTU9ERUwgTE9BREVSIFdPUktFUiDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xh"
    "c3MgTW9kZWxMb2FkZXJXb3JrZXIoUVRocmVhZCk6CiAgICAiIiIKICAgIExvYWRzIHRoZSBtb2RlbCBpbiBhIGJhY2tncm91bmQg"
    "dGhyZWFkIG9uIHN0YXJ0dXAuCiAgICBFbWl0cyBwcm9ncmVzcyBtZXNzYWdlcyB0byB0aGUgcGVyc29uYSBjaGF0IHRhYi4KCiAg"
    "ICBTaWduYWxzOgogICAgICAgIG1lc3NhZ2Uoc3RyKSAgICAgICAg4oCUIHN0YXR1cyBtZXNzYWdlIGZvciBkaXNwbGF5CiAgICAg"
    "ICAgbG9hZF9jb21wbGV0ZShib29sKSDigJQgVHJ1ZT1zdWNjZXNzLCBGYWxzZT1mYWlsdXJlCiAgICAgICAgZXJyb3Ioc3RyKSAg"
    "ICAgICAgICDigJQgZXJyb3IgbWVzc2FnZSBvbiBmYWlsdXJlCiAgICAiIiIKCiAgICBtZXNzYWdlICAgICAgID0gU2lnbmFsKHN0"
    "cikKICAgIGxvYWRfY29tcGxldGUgPSBTaWduYWwoYm9vbCkKICAgIGVycm9yICAgICAgICAgPSBTaWduYWwoc3RyKQoKICAgIGRl"
    "ZiBfX2luaXRfXyhzZWxmLCBhZGFwdG9yOiBMTE1BZGFwdG9yKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKCkKICAgICAgICBz"
    "ZWxmLl9hZGFwdG9yID0gYWRhcHRvcgoKICAgIGRlZiBydW4oc2VsZikgLT4gTm9uZToKICAgICAgICB0cnk6CiAgICAgICAgICAg"
    "IGlmIGlzaW5zdGFuY2Uoc2VsZi5fYWRhcHRvciwgTG9jYWxUcmFuc2Zvcm1lcnNBZGFwdG9yKToKICAgICAgICAgICAgICAgIHNl"
    "bGYubWVzc2FnZS5lbWl0KAogICAgICAgICAgICAgICAgICAgICJTdW1tb25pbmcgdGhlIHZlc3NlbC4uLiB0aGlzIG1heSB0YWtl"
    "IGEgbW9tZW50LiIKICAgICAgICAgICAgICAgICkKICAgICAgICAgICAgICAgIHN1Y2Nlc3MgPSBzZWxmLl9hZGFwdG9yLmxvYWQo"
    "KQogICAgICAgICAgICAgICAgaWYgc3VjY2VzczoKICAgICAgICAgICAgICAgICAgICBzZWxmLm1lc3NhZ2UuZW1pdCgiVGhlIHZl"
    "c3NlbCBzdGlycy4gUHJlc2VuY2UgY29uZmlybWVkLiIpCiAgICAgICAgICAgICAgICAgICAgc2VsZi5tZXNzYWdlLmVtaXQoVUlf"
    "QVdBS0VOSU5HX0xJTkUpCiAgICAgICAgICAgICAgICAgICAgc2VsZi5sb2FkX2NvbXBsZXRlLmVtaXQoVHJ1ZSkKICAgICAgICAg"
    "ICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICAgICAgZXJyID0gc2VsZi5fYWRhcHRvci5lcnJvcgogICAgICAgICAgICAgICAg"
    "ICAgIHNlbGYuZXJyb3IuZW1pdChmIlN1bW1vbmluZyBmYWlsZWQ6IHtlcnJ9IikKICAgICAgICAgICAgICAgICAgICBzZWxmLmxv"
    "YWRfY29tcGxldGUuZW1pdChGYWxzZSkKCiAgICAgICAgICAgIGVsaWYgaXNpbnN0YW5jZShzZWxmLl9hZGFwdG9yLCBPbGxhbWFB"
    "ZGFwdG9yKToKICAgICAgICAgICAgICAgIHNlbGYubWVzc2FnZS5lbWl0KCJSZWFjaGluZyB0aHJvdWdoIHRoZSBhZXRoZXIgdG8g"
    "T2xsYW1hLi4uIikKICAgICAgICAgICAgICAgIGlmIHNlbGYuX2FkYXB0b3IuaXNfY29ubmVjdGVkKCk6CiAgICAgICAgICAgICAg"
    "ICAgICAgc2VsZi5tZXNzYWdlLmVtaXQoIk9sbGFtYSByZXNwb25kcy4gVGhlIGNvbm5lY3Rpb24gaG9sZHMuIikKICAgICAgICAg"
    "ICAgICAgICAgICBzZWxmLm1lc3NhZ2UuZW1pdChVSV9BV0FLRU5JTkdfTElORSkKICAgICAgICAgICAgICAgICAgICBzZWxmLmxv"
    "YWRfY29tcGxldGUuZW1pdChUcnVlKQogICAgICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgICAgICBzZWxmLmVycm9y"
    "LmVtaXQoCiAgICAgICAgICAgICAgICAgICAgICAgICJPbGxhbWEgaXMgbm90IHJ1bm5pbmcuIFN0YXJ0IE9sbGFtYSBhbmQgcmVz"
    "dGFydCB0aGUgZGVjay4iCiAgICAgICAgICAgICAgICAgICAgKQogICAgICAgICAgICAgICAgICAgIHNlbGYubG9hZF9jb21wbGV0"
    "ZS5lbWl0KEZhbHNlKQoKICAgICAgICAgICAgZWxpZiBpc2luc3RhbmNlKHNlbGYuX2FkYXB0b3IsIChDbGF1ZGVBZGFwdG9yLCBP"
    "cGVuQUlBZGFwdG9yKSk6CiAgICAgICAgICAgICAgICBzZWxmLm1lc3NhZ2UuZW1pdCgiVGVzdGluZyB0aGUgQVBJIGNvbm5lY3Rp"
    "b24uLi4iKQogICAgICAgICAgICAgICAgaWYgc2VsZi5fYWRhcHRvci5pc19jb25uZWN0ZWQoKToKICAgICAgICAgICAgICAgICAg"
    "ICBzZWxmLm1lc3NhZ2UuZW1pdCgiQVBJIGtleSBhY2NlcHRlZC4gVGhlIGNvbm5lY3Rpb24gaG9sZHMuIikKICAgICAgICAgICAg"
    "ICAgICAgICBzZWxmLm1lc3NhZ2UuZW1pdChVSV9BV0FLRU5JTkdfTElORSkKICAgICAgICAgICAgICAgICAgICBzZWxmLmxvYWRf"
    "Y29tcGxldGUuZW1pdChUcnVlKQogICAgICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgICAgICBzZWxmLmVycm9yLmVt"
    "aXQoIkFQSSBrZXkgbWlzc2luZyBvciBpbnZhbGlkLiIpCiAgICAgICAgICAgICAgICAgICAgc2VsZi5sb2FkX2NvbXBsZXRlLmVt"
    "aXQoRmFsc2UpCgogICAgICAgICAgICBlbHNlOgogICAgICAgICAgICAgICAgc2VsZi5lcnJvci5lbWl0KCJVbmtub3duIG1vZGVs"
    "IHR5cGUgaW4gY29uZmlnLiIpCiAgICAgICAgICAgICAgICBzZWxmLmxvYWRfY29tcGxldGUuZW1pdChGYWxzZSkKCiAgICAgICAg"
    "ZXhjZXB0IEV4Y2VwdGlvbiBhcyBlOgogICAgICAgICAgICBzZWxmLmVycm9yLmVtaXQoc3RyKGUpKQogICAgICAgICAgICBzZWxm"
    "LmxvYWRfY29tcGxldGUuZW1pdChGYWxzZSkKCgojIOKUgOKUgCBTT1VORCBXT1JLRVIg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSACmNsYXNzIFNvdW5kV29ya2VyKFFUaHJlYWQpOgogICAgIiIiCiAgICBQbGF5cyBhIHNvdW5kIG9mZiB0aGUg"
    "bWFpbiB0aHJlYWQuCiAgICBQcmV2ZW50cyBhbnkgYXVkaW8gb3BlcmF0aW9uIGZyb20gYmxvY2tpbmcgdGhlIFVJLgoKICAgIFVz"
    "YWdlOgogICAgICAgIHdvcmtlciA9IFNvdW5kV29ya2VyKCJhbGVydCIpCiAgICAgICAgd29ya2VyLnN0YXJ0KCkKICAgICAgICAj"
    "IHdvcmtlciBjbGVhbnMgdXAgb24gaXRzIG93biDigJQgbm8gcmVmZXJlbmNlIG5lZWRlZAogICAgIiIiCgogICAgZGVmIF9faW5p"
    "dF9fKHNlbGYsIHNvdW5kX25hbWU6IHN0cik6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXygpCiAgICAgICAgc2VsZi5fbmFtZSA9"
    "IHNvdW5kX25hbWUKICAgICAgICAjIEF1dG8tZGVsZXRlIHdoZW4gZG9uZQogICAgICAgIHNlbGYuZmluaXNoZWQuY29ubmVjdChz"
    "ZWxmLmRlbGV0ZUxhdGVyKQoKICAgIGRlZiBydW4oc2VsZikgLT4gTm9uZToKICAgICAgICB0cnk6CiAgICAgICAgICAgIHBsYXlf"
    "c291bmQoc2VsZi5fbmFtZSkKICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICBwYXNzCgoKIyDilIDilIAgRkFD"
    "RSBUSU1FUiBNQU5BR0VSIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBGb290ZXJTdHJpcFdpZGdldChTdGF0ZVN0cmlwV2lkZ2V0"
    "KToKICAgICIiIkdlbmVyaWMgZm9vdGVyIHN0cmlwIHdpZGdldCB1c2VkIGJ5IHRoZSBwZXJtYW5lbnQgbG93ZXIgYmxvY2suIiIi"
    "CgoKY2xhc3MgRmFjZVRpbWVyTWFuYWdlcjoKICAgICIiIgogICAgTWFuYWdlcyB0aGUgNjAtc2Vjb25kIGZhY2UgZGlzcGxheSB0"
    "aW1lci4KCiAgICBSdWxlczoKICAgIC0gQWZ0ZXIgc2VudGltZW50IGNsYXNzaWZpY2F0aW9uLCBmYWNlIGlzIGxvY2tlZCBmb3Ig"
    "NjAgc2Vjb25kcy4KICAgIC0gSWYgdXNlciBzZW5kcyBhIG5ldyBtZXNzYWdlIGR1cmluZyB0aGUgNjBzLCBmYWNlIGltbWVkaWF0"
    "ZWx5CiAgICAgIHN3aXRjaGVzIHRvICdhbGVydCcgKGxvY2tlZCA9IEZhbHNlLCBuZXcgY3ljbGUgYmVnaW5zKS4KICAgIC0gQWZ0"
    "ZXIgNjBzIHdpdGggbm8gbmV3IGlucHV0LCByZXR1cm5zIHRvICduZXV0cmFsJy4KICAgIC0gTmV2ZXIgYmxvY2tzIGFueXRoaW5n"
    "LiBQdXJlIHRpbWVyICsgY2FsbGJhY2sgbG9naWMuCiAgICAiIiIKCiAgICBIT0xEX1NFQ09ORFMgPSA2MAoKICAgIGRlZiBfX2lu"
    "aXRfXyhzZWxmLCBtaXJyb3I6ICJNaXJyb3JXaWRnZXQiLCBlbW90aW9uX2Jsb2NrOiAiRW1vdGlvbkJsb2NrIik6CiAgICAgICAg"
    "c2VsZi5fbWlycm9yICA9IG1pcnJvcgogICAgICAgIHNlbGYuX2Vtb3Rpb24gPSBlbW90aW9uX2Jsb2NrCiAgICAgICAgc2VsZi5f"
    "dGltZXIgICA9IFFUaW1lcigpCiAgICAgICAgc2VsZi5fdGltZXIuc2V0U2luZ2xlU2hvdChUcnVlKQogICAgICAgIHNlbGYuX3Rp"
    "bWVyLnRpbWVvdXQuY29ubmVjdChzZWxmLl9yZXR1cm5fdG9fbmV1dHJhbCkKICAgICAgICBzZWxmLl9sb2NrZWQgID0gRmFsc2UK"
    "CiAgICBkZWYgc2V0X2ZhY2Uoc2VsZiwgZW1vdGlvbjogc3RyKSAtPiBOb25lOgogICAgICAgICIiIlNldCBmYWNlIGFuZCBzdGFy"
    "dCB0aGUgNjAtc2Vjb25kIGhvbGQgdGltZXIuIiIiCiAgICAgICAgc2VsZi5fbG9ja2VkID0gVHJ1ZQogICAgICAgIHNlbGYuX21p"
    "cnJvci5zZXRfZmFjZShlbW90aW9uKQogICAgICAgIHNlbGYuX2Vtb3Rpb24uYWRkRW1vdGlvbihlbW90aW9uKQogICAgICAgIHNl"
    "bGYuX3RpbWVyLnN0b3AoKQogICAgICAgIHNlbGYuX3RpbWVyLnN0YXJ0KHNlbGYuSE9MRF9TRUNPTkRTICogMTAwMCkKCiAgICBk"
    "ZWYgaW50ZXJydXB0KHNlbGYsIG5ld19lbW90aW9uOiBzdHIgPSAiYWxlcnQiKSAtPiBOb25lOgogICAgICAgICIiIgogICAgICAg"
    "IENhbGxlZCB3aGVuIHVzZXIgc2VuZHMgYSBuZXcgbWVzc2FnZS4KICAgICAgICBJbnRlcnJ1cHRzIGFueSBydW5uaW5nIGhvbGQs"
    "IHNldHMgYWxlcnQgZmFjZSBpbW1lZGlhdGVseS4KICAgICAgICAiIiIKICAgICAgICBzZWxmLl90aW1lci5zdG9wKCkKICAgICAg"
    "ICBzZWxmLl9sb2NrZWQgPSBGYWxzZQogICAgICAgIHNlbGYuX21pcnJvci5zZXRfZmFjZShuZXdfZW1vdGlvbikKICAgICAgICBz"
    "ZWxmLl9lbW90aW9uLmFkZEVtb3Rpb24obmV3X2Vtb3Rpb24pCgogICAgZGVmIF9yZXR1cm5fdG9fbmV1dHJhbChzZWxmKSAtPiBO"
    "b25lOgogICAgICAgIHNlbGYuX2xvY2tlZCA9IEZhbHNlCiAgICAgICAgc2VsZi5fbWlycm9yLnNldF9mYWNlKCJuZXV0cmFsIikK"
    "CiAgICBAcHJvcGVydHkKICAgIGRlZiBpc19sb2NrZWQoc2VsZikgLT4gYm9vbDoKICAgICAgICByZXR1cm4gc2VsZi5fbG9ja2Vk"
    "CgoKIyDilIDilIAgREVQRU5ERU5DWSBDSEVDS0VSIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBEZXBlbmRlbmN5Q2hlY2tlcjoK"
    "ICAgICIiIgogICAgVmFsaWRhdGVzIGFsbCByZXF1aXJlZCBhbmQgb3B0aW9uYWwgcGFja2FnZXMgb24gc3RhcnR1cC4KICAgIFJl"
    "dHVybnMgYSBsaXN0IG9mIHN0YXR1cyBtZXNzYWdlcyBmb3IgdGhlIERpYWdub3N0aWNzIHRhYi4KICAgIFNob3dzIGEgYmxvY2tp"
    "bmcgZXJyb3IgZGlhbG9nIGZvciBhbnkgY3JpdGljYWwgbWlzc2luZyBkZXBlbmRlbmN5LgogICAgIiIiCgogICAgIyAocGFja2Fn"
    "ZV9uYW1lLCBpbXBvcnRfbmFtZSwgY3JpdGljYWwsIGluc3RhbGxfaGludCkKICAgIFBBQ0tBR0VTID0gWwogICAgICAgICgiUHlT"
    "aWRlNiIsICAgICAgICAgICAgICAgICAgICJQeVNpZGU2IiwgICAgICAgICAgICAgIFRydWUsCiAgICAgICAgICJwaXAgaW5zdGFs"
    "bCBQeVNpZGU2IiksCiAgICAgICAgKCJsb2d1cnUiLCAgICAgICAgICAgICAgICAgICAgImxvZ3VydSIsICAgICAgICAgICAgICAg"
    "VHJ1ZSwKICAgICAgICAgInBpcCBpbnN0YWxsIGxvZ3VydSIpLAogICAgICAgICgiYXBzY2hlZHVsZXIiLCAgICAgICAgICAgICAg"
    "ICJhcHNjaGVkdWxlciIsICAgICAgICAgIFRydWUsCiAgICAgICAgICJwaXAgaW5zdGFsbCBhcHNjaGVkdWxlciIpLAogICAgICAg"
    "ICgicHlnYW1lIiwgICAgICAgICAgICAgICAgICAgICJweWdhbWUiLCAgICAgICAgICAgICAgIEZhbHNlLAogICAgICAgICAicGlw"
    "IGluc3RhbGwgcHlnYW1lICAobmVlZGVkIGZvciBzb3VuZCkiKSwKICAgICAgICAoInB5d2luMzIiLCAgICAgICAgICAgICAgICAg"
    "ICAid2luMzJjb20iLCAgICAgICAgICAgICBGYWxzZSwKICAgICAgICAgInBpcCBpbnN0YWxsIHB5d2luMzIgIChuZWVkZWQgZm9y"
    "IGRlc2t0b3Agc2hvcnRjdXQpIiksCiAgICAgICAgKCJwc3V0aWwiLCAgICAgICAgICAgICAgICAgICAgInBzdXRpbCIsICAgICAg"
    "ICAgICAgICAgRmFsc2UsCiAgICAgICAgICJwaXAgaW5zdGFsbCBwc3V0aWwgIChuZWVkZWQgZm9yIHN5c3RlbSBtb25pdG9yaW5n"
    "KSIpLAogICAgICAgICgicmVxdWVzdHMiLCAgICAgICAgICAgICAgICAgICJyZXF1ZXN0cyIsICAgICAgICAgICAgIEZhbHNlLAog"
    "ICAgICAgICAicGlwIGluc3RhbGwgcmVxdWVzdHMiKSwKICAgICAgICAoInRvcmNoIiwgICAgICAgICAgICAgICAgICAgICAidG9y"
    "Y2giLCAgICAgICAgICAgICAgICBGYWxzZSwKICAgICAgICAgInBpcCBpbnN0YWxsIHRvcmNoICAob25seSBuZWVkZWQgZm9yIGxv"
    "Y2FsIG1vZGVsKSIpLAogICAgICAgICgidHJhbnNmb3JtZXJzIiwgICAgICAgICAgICAgICJ0cmFuc2Zvcm1lcnMiLCAgICAgICAg"
    "IEZhbHNlLAogICAgICAgICAicGlwIGluc3RhbGwgdHJhbnNmb3JtZXJzICAob25seSBuZWVkZWQgZm9yIGxvY2FsIG1vZGVsKSIp"
    "LAogICAgICAgICgicHludm1sIiwgICAgICAgICAgICAgICAgICAgICJweW52bWwiLCAgICAgICAgICAgICAgIEZhbHNlLAogICAg"
    "ICAgICAicGlwIGluc3RhbGwgcHludm1sICAob25seSBuZWVkZWQgZm9yIE5WSURJQSBHUFUgbW9uaXRvcmluZykiKSwKICAgIF0K"
    "CiAgICBAY2xhc3NtZXRob2QKICAgIGRlZiBjaGVjayhjbHMpIC0+IHR1cGxlW2xpc3Rbc3RyXSwgbGlzdFtzdHJdXToKICAgICAg"
    "ICAiIiIKICAgICAgICBSZXR1cm5zIChtZXNzYWdlcywgY3JpdGljYWxfZmFpbHVyZXMpLgogICAgICAgIG1lc3NhZ2VzOiBsaXN0"
    "IG9mICJbREVQU10gcGFja2FnZSDinJMv4pyXIOKAlCBub3RlIiBzdHJpbmdzCiAgICAgICAgY3JpdGljYWxfZmFpbHVyZXM6IGxp"
    "c3Qgb2YgcGFja2FnZXMgdGhhdCBhcmUgY3JpdGljYWwgYW5kIG1pc3NpbmcKICAgICAgICAiIiIKICAgICAgICBpbXBvcnQgaW1w"
    "b3J0bGliCiAgICAgICAgbWVzc2FnZXMgID0gW10KICAgICAgICBjcml0aWNhbCAgPSBbXQoKICAgICAgICBmb3IgcGtnX25hbWUs"
    "IGltcG9ydF9uYW1lLCBpc19jcml0aWNhbCwgaGludCBpbiBjbHMuUEFDS0FHRVM6CiAgICAgICAgICAgIHRyeToKICAgICAgICAg"
    "ICAgICAgIGltcG9ydGxpYi5pbXBvcnRfbW9kdWxlKGltcG9ydF9uYW1lKQogICAgICAgICAgICAgICAgbWVzc2FnZXMuYXBwZW5k"
    "KGYiW0RFUFNdIHtwa2dfbmFtZX0g4pyTIikKICAgICAgICAgICAgZXhjZXB0IEltcG9ydEVycm9yOgogICAgICAgICAgICAgICAg"
    "c3RhdHVzID0gIkNSSVRJQ0FMIiBpZiBpc19jcml0aWNhbCBlbHNlICJvcHRpb25hbCIKICAgICAgICAgICAgICAgIG1lc3NhZ2Vz"
    "LmFwcGVuZCgKICAgICAgICAgICAgICAgICAgICBmIltERVBTXSB7cGtnX25hbWV9IOKclyAoe3N0YXR1c30pIOKAlCB7aGludH0i"
    "CiAgICAgICAgICAgICAgICApCiAgICAgICAgICAgICAgICBpZiBpc19jcml0aWNhbDoKICAgICAgICAgICAgICAgICAgICBjcml0"
    "aWNhbC5hcHBlbmQocGtnX25hbWUpCgogICAgICAgIHJldHVybiBtZXNzYWdlcywgY3JpdGljYWwKCiAgICBAY2xhc3NtZXRob2QK"
    "ICAgIGRlZiBjaGVja19vbGxhbWEoY2xzKSAtPiBzdHI6CiAgICAgICAgIiIiQ2hlY2sgaWYgT2xsYW1hIGlzIHJ1bm5pbmcuIFJl"
    "dHVybnMgc3RhdHVzIHN0cmluZy4iIiIKICAgICAgICB0cnk6CiAgICAgICAgICAgIHJlcSAgPSB1cmxsaWIucmVxdWVzdC5SZXF1"
    "ZXN0KCJodHRwOi8vbG9jYWxob3N0OjExNDM0L2FwaS90YWdzIikKICAgICAgICAgICAgcmVzcCA9IHVybGxpYi5yZXF1ZXN0LnVy"
    "bG9wZW4ocmVxLCB0aW1lb3V0PTIpCiAgICAgICAgICAgIGlmIHJlc3Auc3RhdHVzID09IDIwMDoKICAgICAgICAgICAgICAgIHJl"
    "dHVybiAiW0RFUFNdIE9sbGFtYSDinJMg4oCUIHJ1bm5pbmcgb24gbG9jYWxob3N0OjExNDM0IgogICAgICAgIGV4Y2VwdCBFeGNl"
    "cHRpb246CiAgICAgICAgICAgIHBhc3MKICAgICAgICByZXR1cm4gIltERVBTXSBPbGxhbWEg4pyXIOKAlCBub3QgcnVubmluZyAo"
    "b25seSBuZWVkZWQgZm9yIE9sbGFtYSBtb2RlbCB0eXBlKSIKCgojIOKUgOKUgCBNRU1PUlkgTUFOQUdFUiDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIAKY2xhc3MgTWVtb3J5TWFuYWdlcjoKICAgICIiIgogICAgSGFuZGxlcyBhbGwgSlNPTkwgbWVtb3J5"
    "IG9wZXJhdGlvbnMuCgogICAgRmlsZXMgbWFuYWdlZDoKICAgICAgICBtZW1vcmllcy9tZXNzYWdlcy5qc29ubCAgICAgICAgIOKA"
    "lCBldmVyeSBtZXNzYWdlLCB0aW1lc3RhbXBlZAogICAgICAgIG1lbW9yaWVzL21lbW9yaWVzLmpzb25sICAgICAgICAg4oCUIGV4"
    "dHJhY3RlZCBtZW1vcnkgcmVjb3JkcwogICAgICAgIG1lbW9yaWVzL3N0YXRlLmpzb24gICAgICAgICAgICAg4oCUIGVudGl0eSBz"
    "dGF0ZQogICAgICAgIG1lbW9yaWVzL2luZGV4Lmpzb24gICAgICAgICAgICAg4oCUIGNvdW50cyBhbmQgbWV0YWRhdGEKCiAgICBN"
    "ZW1vcnkgcmVjb3JkcyBoYXZlIHR5cGUgaW5mZXJlbmNlLCBrZXl3b3JkIGV4dHJhY3Rpb24sIHRhZyBnZW5lcmF0aW9uLAogICAg"
    "bmVhci1kdXBsaWNhdGUgZGV0ZWN0aW9uLCBhbmQgcmVsZXZhbmNlIHNjb3JpbmcgZm9yIGNvbnRleHQgaW5qZWN0aW9uLgogICAg"
    "IiIiCgogICAgZGVmIF9faW5pdF9fKHNlbGYpOgogICAgICAgIGJhc2UgICAgICAgICAgICAgPSBjZmdfcGF0aCgibWVtb3JpZXMi"
    "KQogICAgICAgIHNlbGYubWVzc2FnZXNfcCAgPSBiYXNlIC8gIm1lc3NhZ2VzLmpzb25sIgogICAgICAgIHNlbGYubWVtb3JpZXNf"
    "cCAgPSBiYXNlIC8gIm1lbW9yaWVzLmpzb25sIgogICAgICAgIHNlbGYuc3RhdGVfcCAgICAgPSBiYXNlIC8gInN0YXRlLmpzb24i"
    "CiAgICAgICAgc2VsZi5pbmRleF9wICAgICA9IGJhc2UgLyAiaW5kZXguanNvbiIKCiAgICAjIOKUgOKUgCBTVEFURSDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIAKICAgIGRlZiBsb2FkX3N0YXRlKHNlbGYpIC0+IGRpY3Q6CiAgICAgICAgaWYgbm90IHNlbGYuc3Rh"
    "dGVfcC5leGlzdHMoKToKICAgICAgICAgICAgcmV0dXJuIHNlbGYuX2RlZmF1bHRfc3RhdGUoKQogICAgICAgIHRyeToKICAgICAg"
    "ICAgICAgcmV0dXJuIGpzb24ubG9hZHMoc2VsZi5zdGF0ZV9wLnJlYWRfdGV4dChlbmNvZGluZz0idXRmLTgiKSkKICAgICAgICBl"
    "eGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICByZXR1cm4gc2VsZi5fZGVmYXVsdF9zdGF0ZSgpCgogICAgZGVmIHNhdmVfc3Rh"
    "dGUoc2VsZiwgc3RhdGU6IGRpY3QpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5zdGF0ZV9wLndyaXRlX3RleHQoCiAgICAgICAgICAg"
    "IGpzb24uZHVtcHMoc3RhdGUsIGluZGVudD0yKSwgZW5jb2Rpbmc9InV0Zi04IgogICAgICAgICkKCiAgICBkZWYgX2RlZmF1bHRf"
    "c3RhdGUoc2VsZikgLT4gZGljdDoKICAgICAgICByZXR1cm4gewogICAgICAgICAgICAicGVyc29uYV9uYW1lIjogICAgICAgICAg"
    "ICAgREVDS19OQU1FLAogICAgICAgICAgICAiZGVja192ZXJzaW9uIjogICAgICAgICAgICAgQVBQX1ZFUlNJT04sCiAgICAgICAg"
    "ICAgICJzZXNzaW9uX2NvdW50IjogICAgICAgICAgICAwLAogICAgICAgICAgICAibGFzdF9zdGFydHVwIjogICAgICAgICAgICAg"
    "Tm9uZSwKICAgICAgICAgICAgImxhc3Rfc2h1dGRvd24iOiAgICAgICAgICAgIE5vbmUsCiAgICAgICAgICAgICJsYXN0X2FjdGl2"
    "ZSI6ICAgICAgICAgICAgICBOb25lLAogICAgICAgICAgICAidG90YWxfbWVzc2FnZXMiOiAgICAgICAgICAgMCwKICAgICAgICAg"
    "ICAgInRvdGFsX21lbW9yaWVzIjogICAgICAgICAgIDAsCiAgICAgICAgICAgICJpbnRlcm5hbF9uYXJyYXRpdmUiOiAgICAgICB7"
    "fSwKICAgICAgICAgICAgImFpX3N0YXRlX2F0X3NodXRkb3duIjoiRE9STUFOVCIsCiAgICAgICAgfQoKICAgICMg4pSA4pSAIE1F"
    "U1NBR0VTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgZGVmIGFwcGVuZF9tZXNzYWdlKHNlbGYsIHNlc3Npb25faWQ6IHN0ciwgcm9sZTog"
    "c3RyLAogICAgICAgICAgICAgICAgICAgICAgIGNvbnRlbnQ6IHN0ciwgZW1vdGlvbjogc3RyID0gIiIpIC0+IGRpY3Q6CiAgICAg"
    "ICAgcmVjb3JkID0gewogICAgICAgICAgICAiaWQiOiAgICAgICAgIGYibXNnX3t1dWlkLnV1aWQ0KCkuaGV4WzoxMl19IiwKICAg"
    "ICAgICAgICAgInRpbWVzdGFtcCI6ICBsb2NhbF9ub3dfaXNvKCksCiAgICAgICAgICAgICJzZXNzaW9uX2lkIjogc2Vzc2lvbl9p"
    "ZCwKICAgICAgICAgICAgInBlcnNvbmEiOiAgICBERUNLX05BTUUsCiAgICAgICAgICAgICJyb2xlIjogICAgICAgcm9sZSwKICAg"
    "ICAgICAgICAgImNvbnRlbnQiOiAgICBjb250ZW50LAogICAgICAgICAgICAiZW1vdGlvbiI6ICAgIGVtb3Rpb24sCiAgICAgICAg"
    "fQogICAgICAgIGFwcGVuZF9qc29ubChzZWxmLm1lc3NhZ2VzX3AsIHJlY29yZCkKICAgICAgICByZXR1cm4gcmVjb3JkCgogICAg"
    "ZGVmIGxvYWRfcmVjZW50X21lc3NhZ2VzKHNlbGYsIGxpbWl0OiBpbnQgPSAyMCkgLT4gbGlzdFtkaWN0XToKICAgICAgICByZXR1"
    "cm4gcmVhZF9qc29ubChzZWxmLm1lc3NhZ2VzX3ApWy1saW1pdDpdCgogICAgIyDilIDilIAgTUVNT1JJRVMg4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSACiAgICBkZWYgYXBwZW5kX21lbW9yeShzZWxmLCBzZXNzaW9uX2lkOiBzdHIsIHVzZXJfdGV4dDogc3RyLAogICAgICAgICAg"
    "ICAgICAgICAgICAgYXNzaXN0YW50X3RleHQ6IHN0cikgLT4gT3B0aW9uYWxbZGljdF06CiAgICAgICAgcmVjb3JkX3R5cGUgPSBp"
    "bmZlcl9yZWNvcmRfdHlwZSh1c2VyX3RleHQsIGFzc2lzdGFudF90ZXh0KQogICAgICAgIGtleXdvcmRzICAgID0gZXh0cmFjdF9r"
    "ZXl3b3Jkcyh1c2VyX3RleHQgKyAiICIgKyBhc3Npc3RhbnRfdGV4dCkKICAgICAgICB0YWdzICAgICAgICA9IHNlbGYuX2luZmVy"
    "X3RhZ3MocmVjb3JkX3R5cGUsIHVzZXJfdGV4dCwga2V5d29yZHMpCiAgICAgICAgdGl0bGUgICAgICAgPSBzZWxmLl9pbmZlcl90"
    "aXRsZShyZWNvcmRfdHlwZSwgdXNlcl90ZXh0LCBrZXl3b3JkcykKICAgICAgICBzdW1tYXJ5ICAgICA9IHNlbGYuX3N1bW1hcml6"
    "ZShyZWNvcmRfdHlwZSwgdXNlcl90ZXh0LCBhc3Npc3RhbnRfdGV4dCkKCiAgICAgICAgbWVtb3J5ID0gewogICAgICAgICAgICAi"
    "aWQiOiAgICAgICAgICAgICAgIGYibWVtX3t1dWlkLnV1aWQ0KCkuaGV4WzoxMl19IiwKICAgICAgICAgICAgInRpbWVzdGFtcCI6"
    "ICAgICAgICBsb2NhbF9ub3dfaXNvKCksCiAgICAgICAgICAgICJzZXNzaW9uX2lkIjogICAgICAgc2Vzc2lvbl9pZCwKICAgICAg"
    "ICAgICAgInBlcnNvbmEiOiAgICAgICAgICBERUNLX05BTUUsCiAgICAgICAgICAgICJ0eXBlIjogICAgICAgICAgICAgcmVjb3Jk"
    "X3R5cGUsCiAgICAgICAgICAgICJ0aXRsZSI6ICAgICAgICAgICAgdGl0bGUsCiAgICAgICAgICAgICJzdW1tYXJ5IjogICAgICAg"
    "ICAgc3VtbWFyeSwKICAgICAgICAgICAgImNvbnRlbnQiOiAgICAgICAgICB1c2VyX3RleHRbOjQwMDBdLAogICAgICAgICAgICAi"
    "YXNzaXN0YW50X2NvbnRleHQiOmFzc2lzdGFudF90ZXh0WzoxMjAwXSwKICAgICAgICAgICAgImtleXdvcmRzIjogICAgICAgICBr"
    "ZXl3b3JkcywKICAgICAgICAgICAgInRhZ3MiOiAgICAgICAgICAgICB0YWdzLAogICAgICAgICAgICAiY29uZmlkZW5jZSI6ICAg"
    "ICAgIDAuNzAgaWYgcmVjb3JkX3R5cGUgaW4gewogICAgICAgICAgICAgICAgImRyZWFtIiwiaXNzdWUiLCJpZGVhIiwicHJlZmVy"
    "ZW5jZSIsInJlc29sdXRpb24iCiAgICAgICAgICAgIH0gZWxzZSAwLjU1LAogICAgICAgIH0KCiAgICAgICAgaWYgc2VsZi5faXNf"
    "bmVhcl9kdXBsaWNhdGUobWVtb3J5KToKICAgICAgICAgICAgcmV0dXJuIE5vbmUKCiAgICAgICAgYXBwZW5kX2pzb25sKHNlbGYu"
    "bWVtb3JpZXNfcCwgbWVtb3J5KQogICAgICAgIHJldHVybiBtZW1vcnkKCiAgICBkZWYgc2VhcmNoX21lbW9yaWVzKHNlbGYsIHF1"
    "ZXJ5OiBzdHIsIGxpbWl0OiBpbnQgPSA2KSAtPiBsaXN0W2RpY3RdOgogICAgICAgICIiIgogICAgICAgIEtleXdvcmQtc2NvcmVk"
    "IG1lbW9yeSBzZWFyY2guCiAgICAgICAgUmV0dXJucyB1cCB0byBgbGltaXRgIHJlY29yZHMgc29ydGVkIGJ5IHJlbGV2YW5jZSBz"
    "Y29yZSBkZXNjZW5kaW5nLgogICAgICAgIEZhbGxzIGJhY2sgdG8gbW9zdCByZWNlbnQgaWYgbm8gcXVlcnkgdGVybXMgbWF0Y2gu"
    "CiAgICAgICAgIiIiCiAgICAgICAgbWVtb3JpZXMgPSByZWFkX2pzb25sKHNlbGYubWVtb3JpZXNfcCkKICAgICAgICBpZiBub3Qg"
    "cXVlcnkuc3RyaXAoKToKICAgICAgICAgICAgcmV0dXJuIG1lbW9yaWVzWy1saW1pdDpdCgogICAgICAgIHFfdGVybXMgPSBzZXQo"
    "ZXh0cmFjdF9rZXl3b3JkcyhxdWVyeSwgbGltaXQ9MTYpKQogICAgICAgIHNjb3JlZCAgPSBbXQoKICAgICAgICBmb3IgaXRlbSBp"
    "biBtZW1vcmllczoKICAgICAgICAgICAgaXRlbV90ZXJtcyA9IHNldChleHRyYWN0X2tleXdvcmRzKCIgIi5qb2luKFsKICAgICAg"
    "ICAgICAgICAgIGl0ZW0uZ2V0KCJ0aXRsZSIsICAgIiIpLAogICAgICAgICAgICAgICAgaXRlbS5nZXQoInN1bW1hcnkiLCAiIiks"
    "CiAgICAgICAgICAgICAgICBpdGVtLmdldCgiY29udGVudCIsICIiKSwKICAgICAgICAgICAgICAgICIgIi5qb2luKGl0ZW0uZ2V0"
    "KCJrZXl3b3JkcyIsIFtdKSksCiAgICAgICAgICAgICAgICAiICIuam9pbihpdGVtLmdldCgidGFncyIsICAgICBbXSkpLAogICAg"
    "ICAgICAgICBdKSwgbGltaXQ9NDApKQoKICAgICAgICAgICAgc2NvcmUgPSBsZW4ocV90ZXJtcyAmIGl0ZW1fdGVybXMpCgogICAg"
    "ICAgICAgICAjIEJvb3N0IGJ5IHR5cGUgbWF0Y2gKICAgICAgICAgICAgcWwgPSBxdWVyeS5sb3dlcigpCiAgICAgICAgICAgIHJ0"
    "ID0gaXRlbS5nZXQoInR5cGUiLCAiIikKICAgICAgICAgICAgaWYgImRyZWFtIiAgaW4gcWwgYW5kIHJ0ID09ICJkcmVhbSI6ICAg"
    "IHNjb3JlICs9IDQKICAgICAgICAgICAgaWYgInRhc2siICAgaW4gcWwgYW5kIHJ0ID09ICJ0YXNrIjogICAgIHNjb3JlICs9IDMK"
    "ICAgICAgICAgICAgaWYgImlkZWEiICAgaW4gcWwgYW5kIHJ0ID09ICJpZGVhIjogICAgIHNjb3JlICs9IDIKICAgICAgICAgICAg"
    "aWYgImxzbCIgICAgaW4gcWwgYW5kIHJ0IGluIHsiaXNzdWUiLCJyZXNvbHV0aW9uIn06IHNjb3JlICs9IDIKCiAgICAgICAgICAg"
    "IGlmIHNjb3JlID4gMDoKICAgICAgICAgICAgICAgIHNjb3JlZC5hcHBlbmQoKHNjb3JlLCBpdGVtKSkKCiAgICAgICAgc2NvcmVk"
    "LnNvcnQoa2V5PWxhbWJkYSB4OiAoeFswXSwgeFsxXS5nZXQoInRpbWVzdGFtcCIsICIiKSksCiAgICAgICAgICAgICAgICAgICAg"
    "cmV2ZXJzZT1UcnVlKQogICAgICAgIHJldHVybiBbaXRlbSBmb3IgXywgaXRlbSBpbiBzY29yZWRbOmxpbWl0XV0KCiAgICBkZWYg"
    "YnVpbGRfY29udGV4dF9ibG9jayhzZWxmLCBxdWVyeTogc3RyLCBtYXhfY2hhcnM6IGludCA9IDIwMDApIC0+IHN0cjoKICAgICAg"
    "ICAiIiIKICAgICAgICBCdWlsZCBhIGNvbnRleHQgc3RyaW5nIGZyb20gcmVsZXZhbnQgbWVtb3JpZXMgZm9yIHByb21wdCBpbmpl"
    "Y3Rpb24uCiAgICAgICAgVHJ1bmNhdGVzIHRvIG1heF9jaGFycyB0byBwcm90ZWN0IHRoZSBjb250ZXh0IHdpbmRvdy4KICAgICAg"
    "ICAiIiIKICAgICAgICBtZW1vcmllcyA9IHNlbGYuc2VhcmNoX21lbW9yaWVzKHF1ZXJ5LCBsaW1pdD00KQogICAgICAgIGlmIG5v"
    "dCBtZW1vcmllczoKICAgICAgICAgICAgcmV0dXJuICIiCgogICAgICAgIHBhcnRzID0gWyJbUkVMRVZBTlQgTUVNT1JJRVNdIl0K"
    "ICAgICAgICB0b3RhbCA9IDAKICAgICAgICBmb3IgbSBpbiBtZW1vcmllczoKICAgICAgICAgICAgZW50cnkgPSAoCiAgICAgICAg"
    "ICAgICAgICBmIuKAoiBbe20uZ2V0KCd0eXBlJywnJykudXBwZXIoKX1dIHttLmdldCgndGl0bGUnLCcnKX06ICIKICAgICAgICAg"
    "ICAgICAgIGYie20uZ2V0KCdzdW1tYXJ5JywnJyl9IgogICAgICAgICAgICApCiAgICAgICAgICAgIGlmIHRvdGFsICsgbGVuKGVu"
    "dHJ5KSA+IG1heF9jaGFyczoKICAgICAgICAgICAgICAgIGJyZWFrCiAgICAgICAgICAgIHBhcnRzLmFwcGVuZChlbnRyeSkKICAg"
    "ICAgICAgICAgdG90YWwgKz0gbGVuKGVudHJ5KQoKICAgICAgICBwYXJ0cy5hcHBlbmQoIltFTkQgTUVNT1JJRVNdIikKICAgICAg"
    "ICByZXR1cm4gIlxuIi5qb2luKHBhcnRzKQoKICAgICMg4pSA4pSAIEhFTFBFUlMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYg"
    "X2lzX25lYXJfZHVwbGljYXRlKHNlbGYsIGNhbmRpZGF0ZTogZGljdCkgLT4gYm9vbDoKICAgICAgICByZWNlbnQgPSByZWFkX2pz"
    "b25sKHNlbGYubWVtb3JpZXNfcClbLTI1Ol0KICAgICAgICBjdCA9IGNhbmRpZGF0ZS5nZXQoInRpdGxlIiwgIiIpLmxvd2VyKCku"
    "c3RyaXAoKQogICAgICAgIGNzID0gY2FuZGlkYXRlLmdldCgic3VtbWFyeSIsICIiKS5sb3dlcigpLnN0cmlwKCkKICAgICAgICBm"
    "b3IgaXRlbSBpbiByZWNlbnQ6CiAgICAgICAgICAgIGlmIGl0ZW0uZ2V0KCJ0aXRsZSIsIiIpLmxvd2VyKCkuc3RyaXAoKSA9PSBj"
    "dDogIHJldHVybiBUcnVlCiAgICAgICAgICAgIGlmIGl0ZW0uZ2V0KCJzdW1tYXJ5IiwiIikubG93ZXIoKS5zdHJpcCgpID09IGNz"
    "OiByZXR1cm4gVHJ1ZQogICAgICAgIHJldHVybiBGYWxzZQoKICAgIGRlZiBfaW5mZXJfdGFncyhzZWxmLCByZWNvcmRfdHlwZTog"
    "c3RyLCB0ZXh0OiBzdHIsCiAgICAgICAgICAgICAgICAgICAga2V5d29yZHM6IGxpc3Rbc3RyXSkgLT4gbGlzdFtzdHJdOgogICAg"
    "ICAgIHQgICAgPSB0ZXh0Lmxvd2VyKCkKICAgICAgICB0YWdzID0gW3JlY29yZF90eXBlXQogICAgICAgIGlmICJkcmVhbSIgICBp"
    "biB0OiB0YWdzLmFwcGVuZCgiZHJlYW0iKQogICAgICAgIGlmICJsc2wiICAgICBpbiB0OiB0YWdzLmFwcGVuZCgibHNsIikKICAg"
    "ICAgICBpZiAicHl0aG9uIiAgaW4gdDogdGFncy5hcHBlbmQoInB5dGhvbiIpCiAgICAgICAgaWYgImdhbWUiICAgIGluIHQ6IHRh"
    "Z3MuYXBwZW5kKCJnYW1lX2lkZWEiKQogICAgICAgIGlmICJzbCIgICAgICBpbiB0IG9yICJzZWNvbmQgbGlmZSIgaW4gdDogdGFn"
    "cy5hcHBlbmQoInNlY29uZGxpZmUiKQogICAgICAgIGlmIERFQ0tfTkFNRS5sb3dlcigpIGluIHQ6IHRhZ3MuYXBwZW5kKERFQ0tf"
    "TkFNRS5sb3dlcigpKQogICAgICAgIGZvciBrdyBpbiBrZXl3b3Jkc1s6NF06CiAgICAgICAgICAgIGlmIGt3IG5vdCBpbiB0YWdz"
    "OgogICAgICAgICAgICAgICAgdGFncy5hcHBlbmQoa3cpCiAgICAgICAgIyBEZWR1cGxpY2F0ZSBwcmVzZXJ2aW5nIG9yZGVyCiAg"
    "ICAgICAgc2Vlbiwgb3V0ID0gc2V0KCksIFtdCiAgICAgICAgZm9yIHRhZyBpbiB0YWdzOgogICAgICAgICAgICBpZiB0YWcgbm90"
    "IGluIHNlZW46CiAgICAgICAgICAgICAgICBzZWVuLmFkZCh0YWcpCiAgICAgICAgICAgICAgICBvdXQuYXBwZW5kKHRhZykKICAg"
    "ICAgICByZXR1cm4gb3V0WzoxMl0KCiAgICBkZWYgX2luZmVyX3RpdGxlKHNlbGYsIHJlY29yZF90eXBlOiBzdHIsIHVzZXJfdGV4"
    "dDogc3RyLAogICAgICAgICAgICAgICAgICAgICBrZXl3b3JkczogbGlzdFtzdHJdKSAtPiBzdHI6CiAgICAgICAgZGVmIGNsZWFu"
    "KHdvcmRzKToKICAgICAgICAgICAgcmV0dXJuIFt3LnN0cmlwKCIgLV8uLCE/IikuY2FwaXRhbGl6ZSgpCiAgICAgICAgICAgICAg"
    "ICAgICAgZm9yIHcgaW4gd29yZHMgaWYgbGVuKHcpID4gMl0KCiAgICAgICAgaWYgcmVjb3JkX3R5cGUgPT0gInRhc2siOgogICAg"
    "ICAgICAgICBpbXBvcnQgcmUKICAgICAgICAgICAgbSA9IHJlLnNlYXJjaChyInJlbWluZCBtZSAuKj8gdG8gKC4rKSIsIHVzZXJf"
    "dGV4dCwgcmUuSSkKICAgICAgICAgICAgaWYgbToKICAgICAgICAgICAgICAgIHJldHVybiBmIlJlbWluZGVyOiB7bS5ncm91cCgx"
    "KS5zdHJpcCgpWzo2MF19IgogICAgICAgICAgICByZXR1cm4gIlJlbWluZGVyIFRhc2siCiAgICAgICAgaWYgcmVjb3JkX3R5cGUg"
    "PT0gImRyZWFtIjoKICAgICAgICAgICAgcmV0dXJuIGYieycgJy5qb2luKGNsZWFuKGtleXdvcmRzWzozXSkpfSBEcmVhbSIuc3Ry"
    "aXAoKSBvciAiRHJlYW0gTWVtb3J5IgogICAgICAgIGlmIHJlY29yZF90eXBlID09ICJpc3N1ZSI6CiAgICAgICAgICAgIHJldHVy"
    "biBmIklzc3VlOiB7JyAnLmpvaW4oY2xlYW4oa2V5d29yZHNbOjRdKSl9Ii5zdHJpcCgpIG9yICJUZWNobmljYWwgSXNzdWUiCiAg"
    "ICAgICAgaWYgcmVjb3JkX3R5cGUgPT0gInJlc29sdXRpb24iOgogICAgICAgICAgICByZXR1cm4gZiJSZXNvbHV0aW9uOiB7JyAn"
    "LmpvaW4oY2xlYW4oa2V5d29yZHNbOjRdKSl9Ii5zdHJpcCgpIG9yICJUZWNobmljYWwgUmVzb2x1dGlvbiIKICAgICAgICBpZiBy"
    "ZWNvcmRfdHlwZSA9PSAiaWRlYSI6CiAgICAgICAgICAgIHJldHVybiBmIklkZWE6IHsnICcuam9pbihjbGVhbihrZXl3b3Jkc1s6"
    "NF0pKX0iLnN0cmlwKCkgb3IgIklkZWEiCiAgICAgICAgaWYga2V5d29yZHM6CiAgICAgICAgICAgIHJldHVybiAiICIuam9pbihj"
    "bGVhbihrZXl3b3Jkc1s6NV0pKSBvciAiQ29udmVyc2F0aW9uIE1lbW9yeSIKICAgICAgICByZXR1cm4gIkNvbnZlcnNhdGlvbiBN"
    "ZW1vcnkiCgogICAgZGVmIF9zdW1tYXJpemUoc2VsZiwgcmVjb3JkX3R5cGU6IHN0ciwgdXNlcl90ZXh0OiBzdHIsCiAgICAgICAg"
    "ICAgICAgICAgICBhc3Npc3RhbnRfdGV4dDogc3RyKSAtPiBzdHI6CiAgICAgICAgdSA9IHVzZXJfdGV4dC5zdHJpcCgpWzoyMjBd"
    "CiAgICAgICAgYSA9IGFzc2lzdGFudF90ZXh0LnN0cmlwKClbOjIyMF0KICAgICAgICBpZiByZWNvcmRfdHlwZSA9PSAiZHJlYW0i"
    "OiAgICAgICByZXR1cm4gZiJVc2VyIGRlc2NyaWJlZCBhIGRyZWFtOiB7dX0iCiAgICAgICAgaWYgcmVjb3JkX3R5cGUgPT0gInRh"
    "c2siOiAgICAgICAgcmV0dXJuIGYiUmVtaW5kZXIvdGFzazoge3V9IgogICAgICAgIGlmIHJlY29yZF90eXBlID09ICJpc3N1ZSI6"
    "ICAgICAgIHJldHVybiBmIlRlY2huaWNhbCBpc3N1ZToge3V9IgogICAgICAgIGlmIHJlY29yZF90eXBlID09ICJyZXNvbHV0aW9u"
    "IjogIHJldHVybiBmIlNvbHV0aW9uIHJlY29yZGVkOiB7YSBvciB1fSIKICAgICAgICBpZiByZWNvcmRfdHlwZSA9PSAiaWRlYSI6"
    "ICAgICAgICByZXR1cm4gZiJJZGVhIGRpc2N1c3NlZDoge3V9IgogICAgICAgIGlmIHJlY29yZF90eXBlID09ICJwcmVmZXJlbmNl"
    "IjogIHJldHVybiBmIlByZWZlcmVuY2Ugbm90ZWQ6IHt1fSIKICAgICAgICByZXR1cm4gZiJDb252ZXJzYXRpb246IHt1fSIKCgoj"
    "IOKUgOKUgCBTRVNTSU9OIE1BTkFHRVIg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNsYXNzIFNlc3Npb25NYW5hZ2VyOgog"
    "ICAgIiIiCiAgICBNYW5hZ2VzIGNvbnZlcnNhdGlvbiBzZXNzaW9ucy4KCiAgICBBdXRvLXNhdmU6IGV2ZXJ5IDEwIG1pbnV0ZXMg"
    "KEFQU2NoZWR1bGVyKSwgbWlkbmlnaHQtdG8tbWlkbmlnaHQgYm91bmRhcnkuCiAgICBGaWxlOiBzZXNzaW9ucy9ZWVlZLU1NLURE"
    "Lmpzb25sIOKAlCBvdmVyd3JpdGVzIG9uIGVhY2ggc2F2ZS4KICAgIEluZGV4OiBzZXNzaW9ucy9zZXNzaW9uX2luZGV4Lmpzb24g"
    "4oCUIG9uZSBlbnRyeSBwZXIgZGF5LgoKICAgIFNlc3Npb25zIGFyZSBsb2FkZWQgYXMgY29udGV4dCBpbmplY3Rpb24gKG5vdCBy"
    "ZWFsIG1lbW9yeSkgdW50aWwKICAgIHRoZSBTUUxpdGUvQ2hyb21hREIgc3lzdGVtIGlzIGJ1aWx0IGluIFBoYXNlIDIuCiAgICAi"
    "IiIKCiAgICBBVVRPU0FWRV9JTlRFUlZBTCA9IDEwICAgIyBtaW51dGVzCgogICAgZGVmIF9faW5pdF9fKHNlbGYpOgogICAgICAg"
    "IHNlbGYuX3Nlc3Npb25zX2RpciAgPSBjZmdfcGF0aCgic2Vzc2lvbnMiKQogICAgICAgIHNlbGYuX2luZGV4X3BhdGggICAgPSBz"
    "ZWxmLl9zZXNzaW9uc19kaXIgLyAic2Vzc2lvbl9pbmRleC5qc29uIgogICAgICAgIHNlbGYuX3Nlc3Npb25faWQgICAgPSBmInNl"
    "c3Npb25fe2RhdGV0aW1lLm5vdygpLnN0cmZ0aW1lKCclWSVtJWRfJUglTSVTJyl9IgogICAgICAgIHNlbGYuX2N1cnJlbnRfZGF0"
    "ZSAgPSBkYXRlLnRvZGF5KCkuaXNvZm9ybWF0KCkKICAgICAgICBzZWxmLl9tZXNzYWdlczogbGlzdFtkaWN0XSA9IFtdCiAgICAg"
    "ICAgc2VsZi5fbG9hZGVkX2pvdXJuYWw6IE9wdGlvbmFsW3N0cl0gPSBOb25lICAjIGRhdGUgb2YgbG9hZGVkIGpvdXJuYWwKCiAg"
    "ICAjIOKUgOKUgCBDVVJSRU5UIFNFU1NJT04g4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgYWRkX21lc3NhZ2Uoc2VsZiwgcm9sZTogc3RyLCBjb250ZW50OiBzdHIsCiAg"
    "ICAgICAgICAgICAgICAgICAgZW1vdGlvbjogc3RyID0gIiIsIHRpbWVzdGFtcDogc3RyID0gIiIpIC0+IE5vbmU6CiAgICAgICAg"
    "c2VsZi5fbWVzc2FnZXMuYXBwZW5kKHsKICAgICAgICAgICAgImlkIjogICAgICAgIGYibXNnX3t1dWlkLnV1aWQ0KCkuaGV4Wzo4"
    "XX0iLAogICAgICAgICAgICAidGltZXN0YW1wIjogdGltZXN0YW1wIG9yIGxvY2FsX25vd19pc28oKSwKICAgICAgICAgICAgInJv"
    "bGUiOiAgICAgIHJvbGUsCiAgICAgICAgICAgICJjb250ZW50IjogICBjb250ZW50LAogICAgICAgICAgICAiZW1vdGlvbiI6ICAg"
    "ZW1vdGlvbiwKICAgICAgICB9KQoKICAgIGRlZiBnZXRfaGlzdG9yeShzZWxmKSAtPiBsaXN0W2RpY3RdOgogICAgICAgICIiIgog"
    "ICAgICAgIFJldHVybiBoaXN0b3J5IGluIExMTS1mcmllbmRseSBmb3JtYXQuCiAgICAgICAgW3sicm9sZSI6ICJ1c2VyInwiYXNz"
    "aXN0YW50IiwgImNvbnRlbnQiOiAiLi4uIn1dCiAgICAgICAgIiIiCiAgICAgICAgcmV0dXJuIFsKICAgICAgICAgICAgeyJyb2xl"
    "IjogbVsicm9sZSJdLCAiY29udGVudCI6IG1bImNvbnRlbnQiXX0KICAgICAgICAgICAgZm9yIG0gaW4gc2VsZi5fbWVzc2FnZXMK"
    "ICAgICAgICAgICAgaWYgbVsicm9sZSJdIGluICgidXNlciIsICJhc3Npc3RhbnQiKQogICAgICAgIF0KCiAgICBAcHJvcGVydHkK"
    "ICAgIGRlZiBzZXNzaW9uX2lkKHNlbGYpIC0+IHN0cjoKICAgICAgICByZXR1cm4gc2VsZi5fc2Vzc2lvbl9pZAoKICAgIEBwcm9w"
    "ZXJ0eQogICAgZGVmIG1lc3NhZ2VfY291bnQoc2VsZikgLT4gaW50OgogICAgICAgIHJldHVybiBsZW4oc2VsZi5fbWVzc2FnZXMp"
    "CgogICAgIyDilIDilIAgU0FWRSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgIGRlZiBzYXZlKHNlbGYsIGFpX2dlbmVy"
    "YXRlZF9uYW1lOiBzdHIgPSAiIikgLT4gTm9uZToKICAgICAgICAiIiIKICAgICAgICBTYXZlIGN1cnJlbnQgc2Vzc2lvbiB0byBz"
    "ZXNzaW9ucy9ZWVlZLU1NLURELmpzb25sLgogICAgICAgIE92ZXJ3cml0ZXMgdGhlIGZpbGUgZm9yIHRvZGF5IOKAlCBlYWNoIHNh"
    "dmUgaXMgYSBmdWxsIHNuYXBzaG90LgogICAgICAgIFVwZGF0ZXMgc2Vzc2lvbl9pbmRleC5qc29uLgogICAgICAgICIiIgogICAg"
    "ICAgIHRvZGF5ID0gZGF0ZS50b2RheSgpLmlzb2Zvcm1hdCgpCiAgICAgICAgb3V0X3BhdGggPSBzZWxmLl9zZXNzaW9uc19kaXIg"
    "LyBmInt0b2RheX0uanNvbmwiCgogICAgICAgICMgV3JpdGUgYWxsIG1lc3NhZ2VzCiAgICAgICAgd3JpdGVfanNvbmwob3V0X3Bh"
    "dGgsIHNlbGYuX21lc3NhZ2VzKQoKICAgICAgICAjIFVwZGF0ZSBpbmRleAogICAgICAgIGluZGV4ID0gc2VsZi5fbG9hZF9pbmRl"
    "eCgpCiAgICAgICAgZXhpc3RpbmcgPSBuZXh0KAogICAgICAgICAgICAocyBmb3IgcyBpbiBpbmRleFsic2Vzc2lvbnMiXSBpZiBz"
    "WyJkYXRlIl0gPT0gdG9kYXkpLCBOb25lCiAgICAgICAgKQoKICAgICAgICBuYW1lID0gYWlfZ2VuZXJhdGVkX25hbWUgb3IgZXhp"
    "c3RpbmcuZ2V0KCJuYW1lIiwgIiIpIGlmIGV4aXN0aW5nIGVsc2UgIiIKICAgICAgICBpZiBub3QgbmFtZSBhbmQgc2VsZi5fbWVz"
    "c2FnZXM6CiAgICAgICAgICAgICMgQXV0by1uYW1lIGZyb20gZmlyc3QgdXNlciBtZXNzYWdlIChmaXJzdCA1IHdvcmRzKQogICAg"
    "ICAgICAgICBmaXJzdF91c2VyID0gbmV4dCgKICAgICAgICAgICAgICAgIChtWyJjb250ZW50Il0gZm9yIG0gaW4gc2VsZi5fbWVz"
    "c2FnZXMgaWYgbVsicm9sZSJdID09ICJ1c2VyIiksCiAgICAgICAgICAgICAgICAiIgogICAgICAgICAgICApCiAgICAgICAgICAg"
    "IHdvcmRzID0gZmlyc3RfdXNlci5zcGxpdCgpWzo1XQogICAgICAgICAgICBuYW1lICA9ICIgIi5qb2luKHdvcmRzKSBpZiB3b3Jk"
    "cyBlbHNlIGYiU2Vzc2lvbiB7dG9kYXl9IgoKICAgICAgICBlbnRyeSA9IHsKICAgICAgICAgICAgImRhdGUiOiAgICAgICAgICB0"
    "b2RheSwKICAgICAgICAgICAgInNlc3Npb25faWQiOiAgICBzZWxmLl9zZXNzaW9uX2lkLAogICAgICAgICAgICAibmFtZSI6ICAg"
    "ICAgICAgIG5hbWUsCiAgICAgICAgICAgICJtZXNzYWdlX2NvdW50IjogbGVuKHNlbGYuX21lc3NhZ2VzKSwKICAgICAgICAgICAg"
    "ImZpcnN0X21lc3NhZ2UiOiAoc2VsZi5fbWVzc2FnZXNbMF1bInRpbWVzdGFtcCJdCiAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgIGlmIHNlbGYuX21lc3NhZ2VzIGVsc2UgIiIpLAogICAgICAgICAgICAibGFzdF9tZXNzYWdlIjogIChzZWxmLl9tZXNzYWdl"
    "c1stMV1bInRpbWVzdGFtcCJdCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIGlmIHNlbGYuX21lc3NhZ2VzIGVsc2UgIiIp"
    "LAogICAgICAgIH0KCiAgICAgICAgaWYgZXhpc3Rpbmc6CiAgICAgICAgICAgIGlkeCA9IGluZGV4WyJzZXNzaW9ucyJdLmluZGV4"
    "KGV4aXN0aW5nKQogICAgICAgICAgICBpbmRleFsic2Vzc2lvbnMiXVtpZHhdID0gZW50cnkKICAgICAgICBlbHNlOgogICAgICAg"
    "ICAgICBpbmRleFsic2Vzc2lvbnMiXS5pbnNlcnQoMCwgZW50cnkpCgogICAgICAgICMgS2VlcCBsYXN0IDM2NSBkYXlzIGluIGlu"
    "ZGV4CiAgICAgICAgaW5kZXhbInNlc3Npb25zIl0gPSBpbmRleFsic2Vzc2lvbnMiXVs6MzY1XQogICAgICAgIHNlbGYuX3NhdmVf"
    "aW5kZXgoaW5kZXgpCgogICAgIyDilIDilIAgTE9BRCAvIEpPVVJOQUwg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgbGlzdF9zZXNzaW9ucyhzZWxmKSAtPiBsaXN0"
    "W2RpY3RdOgogICAgICAgICIiIlJldHVybiBhbGwgc2Vzc2lvbnMgZnJvbSBpbmRleCwgbmV3ZXN0IGZpcnN0LiIiIgogICAgICAg"
    "IHJldHVybiBzZWxmLl9sb2FkX2luZGV4KCkuZ2V0KCJzZXNzaW9ucyIsIFtdKQoKICAgIGRlZiBsb2FkX3Nlc3Npb25fYXNfY29u"
    "dGV4dChzZWxmLCBzZXNzaW9uX2RhdGU6IHN0cikgLT4gc3RyOgogICAgICAgICIiIgogICAgICAgIExvYWQgYSBwYXN0IHNlc3Np"
    "b24gYXMgYSBjb250ZXh0IGluamVjdGlvbiBzdHJpbmcuCiAgICAgICAgUmV0dXJucyBmb3JtYXR0ZWQgdGV4dCB0byBwcmVwZW5k"
    "IHRvIHRoZSBzeXN0ZW0gcHJvbXB0LgogICAgICAgIFRoaXMgaXMgTk9UIHJlYWwgbWVtb3J5IOKAlCBpdCdzIGEgdGVtcG9yYXJ5"
    "IGNvbnRleHQgd2luZG93IGluamVjdGlvbgogICAgICAgIHVudGlsIHRoZSBQaGFzZSAyIG1lbW9yeSBzeXN0ZW0gaXMgYnVpbHQu"
    "CiAgICAgICAgIiIiCiAgICAgICAgcGF0aCA9IHNlbGYuX3Nlc3Npb25zX2RpciAvIGYie3Nlc3Npb25fZGF0ZX0uanNvbmwiCiAg"
    "ICAgICAgaWYgbm90IHBhdGguZXhpc3RzKCk6CiAgICAgICAgICAgIHJldHVybiAiIgoKICAgICAgICBtZXNzYWdlcyA9IHJlYWRf"
    "anNvbmwocGF0aCkKICAgICAgICBzZWxmLl9sb2FkZWRfam91cm5hbCA9IHNlc3Npb25fZGF0ZQoKICAgICAgICBsaW5lcyA9IFtm"
    "IltKT1VSTkFMIExPQURFRCDigJQge3Nlc3Npb25fZGF0ZX1dIiwKICAgICAgICAgICAgICAgICAiVGhlIGZvbGxvd2luZyBpcyBh"
    "IHJlY29yZCBvZiBhIHByaW9yIGNvbnZlcnNhdGlvbi4iLAogICAgICAgICAgICAgICAgICJVc2UgdGhpcyBhcyBjb250ZXh0IGZv"
    "ciB0aGUgY3VycmVudCBzZXNzaW9uOlxuIl0KCiAgICAgICAgIyBJbmNsdWRlIHVwIHRvIGxhc3QgMzAgbWVzc2FnZXMgZnJvbSB0"
    "aGF0IHNlc3Npb24KICAgICAgICBmb3IgbXNnIGluIG1lc3NhZ2VzWy0zMDpdOgogICAgICAgICAgICByb2xlICAgID0gbXNnLmdl"
    "dCgicm9sZSIsICI/IikudXBwZXIoKQogICAgICAgICAgICBjb250ZW50ID0gbXNnLmdldCgiY29udGVudCIsICIiKVs6MzAwXQog"
    "ICAgICAgICAgICB0cyAgICAgID0gbXNnLmdldCgidGltZXN0YW1wIiwgIiIpWzoxNl0KICAgICAgICAgICAgbGluZXMuYXBwZW5k"
    "KGYiW3t0c31dIHtyb2xlfToge2NvbnRlbnR9IikKCiAgICAgICAgbGluZXMuYXBwZW5kKCJbRU5EIEpPVVJOQUxdIikKICAgICAg"
    "ICByZXR1cm4gIlxuIi5qb2luKGxpbmVzKQoKICAgIGRlZiBjbGVhcl9sb2FkZWRfam91cm5hbChzZWxmKSAtPiBOb25lOgogICAg"
    "ICAgIHNlbGYuX2xvYWRlZF9qb3VybmFsID0gTm9uZQoKICAgIEBwcm9wZXJ0eQogICAgZGVmIGxvYWRlZF9qb3VybmFsX2RhdGUo"
    "c2VsZikgLT4gT3B0aW9uYWxbc3RyXToKICAgICAgICByZXR1cm4gc2VsZi5fbG9hZGVkX2pvdXJuYWwKCiAgICBkZWYgcmVuYW1l"
    "X3Nlc3Npb24oc2VsZiwgc2Vzc2lvbl9kYXRlOiBzdHIsIG5ld19uYW1lOiBzdHIpIC0+IGJvb2w6CiAgICAgICAgIiIiUmVuYW1l"
    "IGEgc2Vzc2lvbiBpbiB0aGUgaW5kZXguIFJldHVybnMgVHJ1ZSBvbiBzdWNjZXNzLiIiIgogICAgICAgIGluZGV4ID0gc2VsZi5f"
    "bG9hZF9pbmRleCgpCiAgICAgICAgZm9yIGVudHJ5IGluIGluZGV4WyJzZXNzaW9ucyJdOgogICAgICAgICAgICBpZiBlbnRyeVsi"
    "ZGF0ZSJdID09IHNlc3Npb25fZGF0ZToKICAgICAgICAgICAgICAgIGVudHJ5WyJuYW1lIl0gPSBuZXdfbmFtZVs6ODBdCiAgICAg"
    "ICAgICAgICAgICBzZWxmLl9zYXZlX2luZGV4KGluZGV4KQogICAgICAgICAgICAgICAgcmV0dXJuIFRydWUKICAgICAgICByZXR1"
    "cm4gRmFsc2UKCiAgICAjIOKUgOKUgCBJTkRFWCBIRUxQRVJTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgZGVmIF9sb2FkX2luZGV4KHNlbGYpIC0+IGRpY3Q6CiAg"
    "ICAgICAgaWYgbm90IHNlbGYuX2luZGV4X3BhdGguZXhpc3RzKCk6CiAgICAgICAgICAgIHJldHVybiB7InNlc3Npb25zIjogW119"
    "CiAgICAgICAgdHJ5OgogICAgICAgICAgICByZXR1cm4ganNvbi5sb2FkcygKICAgICAgICAgICAgICAgIHNlbGYuX2luZGV4X3Bh"
    "dGgucmVhZF90ZXh0KGVuY29kaW5nPSJ1dGYtOCIpCiAgICAgICAgICAgICkKICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAg"
    "ICAgICAgICByZXR1cm4geyJzZXNzaW9ucyI6IFtdfQoKICAgIGRlZiBfc2F2ZV9pbmRleChzZWxmLCBpbmRleDogZGljdCkgLT4g"
    "Tm9uZToKICAgICAgICBzZWxmLl9pbmRleF9wYXRoLndyaXRlX3RleHQoCiAgICAgICAgICAgIGpzb24uZHVtcHMoaW5kZXgsIGlu"
    "ZGVudD0yKSwgZW5jb2Rpbmc9InV0Zi04IgogICAgICAgICkKCgojIOKUgOKUgCBMRVNTT05TIExFQVJORUQgREFUQUJBU0Ug4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNs"
    "YXNzIExlc3NvbnNMZWFybmVkREI6CiAgICAiIiIKICAgIFBlcnNpc3RlbnQga25vd2xlZGdlIGJhc2UgZm9yIGNvZGUgbGVzc29u"
    "cywgcnVsZXMsIGFuZCByZXNvbHV0aW9ucy4KCiAgICBDb2x1bW5zIHBlciByZWNvcmQ6CiAgICAgICAgaWQsIGNyZWF0ZWRfYXQs"
    "IGVudmlyb25tZW50IChMU0x8UHl0aG9ufFB5U2lkZTZ8Li4uKSwgbGFuZ3VhZ2UsCiAgICAgICAgcmVmZXJlbmNlX2tleSAoc2hv"
    "cnQgdW5pcXVlIHRhZyksIHN1bW1hcnksIGZ1bGxfcnVsZSwKICAgICAgICByZXNvbHV0aW9uLCBsaW5rLCB0YWdzCgogICAgUXVl"
    "cmllZCBGSVJTVCBiZWZvcmUgYW55IGNvZGUgc2Vzc2lvbiBpbiB0aGUgcmVsZXZhbnQgbGFuZ3VhZ2UuCiAgICBUaGUgTFNMIEZv"
    "cmJpZGRlbiBSdWxlc2V0IGxpdmVzIGhlcmUuCiAgICBHcm93aW5nLCBub24tZHVwbGljYXRpbmcsIHNlYXJjaGFibGUuCiAgICAi"
    "IiIKCiAgICBkZWYgX19pbml0X18oc2VsZik6CiAgICAgICAgc2VsZi5fcGF0aCA9IGNmZ19wYXRoKCJtZW1vcmllcyIpIC8gImxl"
    "c3NvbnNfbGVhcm5lZC5qc29ubCIKCiAgICBkZWYgYWRkKHNlbGYsIGVudmlyb25tZW50OiBzdHIsIGxhbmd1YWdlOiBzdHIsIHJl"
    "ZmVyZW5jZV9rZXk6IHN0ciwKICAgICAgICAgICAgc3VtbWFyeTogc3RyLCBmdWxsX3J1bGU6IHN0ciwgcmVzb2x1dGlvbjogc3Ry"
    "ID0gIiIsCiAgICAgICAgICAgIGxpbms6IHN0ciA9ICIiLCB0YWdzOiBsaXN0ID0gTm9uZSkgLT4gZGljdDoKICAgICAgICByZWNv"
    "cmQgPSB7CiAgICAgICAgICAgICJpZCI6ICAgICAgICAgICAgZiJsZXNzb25fe3V1aWQudXVpZDQoKS5oZXhbOjEwXX0iLAogICAg"
    "ICAgICAgICAiY3JlYXRlZF9hdCI6ICAgIGxvY2FsX25vd19pc28oKSwKICAgICAgICAgICAgImVudmlyb25tZW50IjogICBlbnZp"
    "cm9ubWVudCwKICAgICAgICAgICAgImxhbmd1YWdlIjogICAgICBsYW5ndWFnZSwKICAgICAgICAgICAgInJlZmVyZW5jZV9rZXki"
    "OiByZWZlcmVuY2Vfa2V5LAogICAgICAgICAgICAic3VtbWFyeSI6ICAgICAgIHN1bW1hcnksCiAgICAgICAgICAgICJmdWxsX3J1"
    "bGUiOiAgICAgZnVsbF9ydWxlLAogICAgICAgICAgICAicmVzb2x1dGlvbiI6ICAgIHJlc29sdXRpb24sCiAgICAgICAgICAgICJs"
    "aW5rIjogICAgICAgICAgbGluaywKICAgICAgICAgICAgInRhZ3MiOiAgICAgICAgICB0YWdzIG9yIFtdLAogICAgICAgIH0KICAg"
    "ICAgICBpZiBub3Qgc2VsZi5faXNfZHVwbGljYXRlKHJlZmVyZW5jZV9rZXkpOgogICAgICAgICAgICBhcHBlbmRfanNvbmwoc2Vs"
    "Zi5fcGF0aCwgcmVjb3JkKQogICAgICAgIHJldHVybiByZWNvcmQKCiAgICBkZWYgc2VhcmNoKHNlbGYsIHF1ZXJ5OiBzdHIgPSAi"
    "IiwgZW52aXJvbm1lbnQ6IHN0ciA9ICIiLAogICAgICAgICAgICAgICBsYW5ndWFnZTogc3RyID0gIiIpIC0+IGxpc3RbZGljdF06"
    "CiAgICAgICAgcmVjb3JkcyA9IHJlYWRfanNvbmwoc2VsZi5fcGF0aCkKICAgICAgICByZXN1bHRzID0gW10KICAgICAgICBxID0g"
    "cXVlcnkubG93ZXIoKQogICAgICAgIGZvciByIGluIHJlY29yZHM6CiAgICAgICAgICAgIGlmIGVudmlyb25tZW50IGFuZCByLmdl"
    "dCgiZW52aXJvbm1lbnQiLCIiKS5sb3dlcigpICE9IGVudmlyb25tZW50Lmxvd2VyKCk6CiAgICAgICAgICAgICAgICBjb250aW51"
    "ZQogICAgICAgICAgICBpZiBsYW5ndWFnZSBhbmQgci5nZXQoImxhbmd1YWdlIiwiIikubG93ZXIoKSAhPSBsYW5ndWFnZS5sb3dl"
    "cigpOgogICAgICAgICAgICAgICAgY29udGludWUKICAgICAgICAgICAgaWYgcToKICAgICAgICAgICAgICAgIGhheXN0YWNrID0g"
    "IiAiLmpvaW4oWwogICAgICAgICAgICAgICAgICAgIHIuZ2V0KCJzdW1tYXJ5IiwiIiksCiAgICAgICAgICAgICAgICAgICAgci5n"
    "ZXQoImZ1bGxfcnVsZSIsIiIpLAogICAgICAgICAgICAgICAgICAgIHIuZ2V0KCJyZWZlcmVuY2Vfa2V5IiwiIiksCiAgICAgICAg"
    "ICAgICAgICAgICAgIiAiLmpvaW4oci5nZXQoInRhZ3MiLFtdKSksCiAgICAgICAgICAgICAgICBdKS5sb3dlcigpCiAgICAgICAg"
    "ICAgICAgICBpZiBxIG5vdCBpbiBoYXlzdGFjazoKICAgICAgICAgICAgICAgICAgICBjb250aW51ZQogICAgICAgICAgICByZXN1"
    "bHRzLmFwcGVuZChyKQogICAgICAgIHJldHVybiByZXN1bHRzCgogICAgZGVmIGdldF9hbGwoc2VsZikgLT4gbGlzdFtkaWN0XToK"
    "ICAgICAgICByZXR1cm4gcmVhZF9qc29ubChzZWxmLl9wYXRoKQoKICAgIGRlZiBkZWxldGUoc2VsZiwgcmVjb3JkX2lkOiBzdHIp"
    "IC0+IGJvb2w6CiAgICAgICAgcmVjb3JkcyA9IHJlYWRfanNvbmwoc2VsZi5fcGF0aCkKICAgICAgICBmaWx0ZXJlZCA9IFtyIGZv"
    "ciByIGluIHJlY29yZHMgaWYgci5nZXQoImlkIikgIT0gcmVjb3JkX2lkXQogICAgICAgIGlmIGxlbihmaWx0ZXJlZCkgPCBsZW4o"
    "cmVjb3Jkcyk6CiAgICAgICAgICAgIHdyaXRlX2pzb25sKHNlbGYuX3BhdGgsIGZpbHRlcmVkKQogICAgICAgICAgICByZXR1cm4g"
    "VHJ1ZQogICAgICAgIHJldHVybiBGYWxzZQoKICAgIGRlZiBidWlsZF9jb250ZXh0X2Zvcl9sYW5ndWFnZShzZWxmLCBsYW5ndWFn"
    "ZTogc3RyLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIG1heF9jaGFyczogaW50ID0gMTUwMCkgLT4gc3RyOgog"
    "ICAgICAgICIiIgogICAgICAgIEJ1aWxkIGEgY29udGV4dCBzdHJpbmcgb2YgYWxsIHJ1bGVzIGZvciBhIGdpdmVuIGxhbmd1YWdl"
    "LgogICAgICAgIEZvciBpbmplY3Rpb24gaW50byBzeXN0ZW0gcHJvbXB0IGJlZm9yZSBjb2RlIHNlc3Npb25zLgogICAgICAgICIi"
    "IgogICAgICAgIHJlY29yZHMgPSBzZWxmLnNlYXJjaChsYW5ndWFnZT1sYW5ndWFnZSkKICAgICAgICBpZiBub3QgcmVjb3JkczoK"
    "ICAgICAgICAgICAgcmV0dXJuICIiCgogICAgICAgIHBhcnRzID0gW2YiW3tsYW5ndWFnZS51cHBlcigpfSBSVUxFUyDigJQgQVBQ"
    "TFkgQkVGT1JFIFdSSVRJTkcgQ09ERV0iXQogICAgICAgIHRvdGFsID0gMAogICAgICAgIGZvciByIGluIHJlY29yZHM6CiAgICAg"
    "ICAgICAgIGVudHJ5ID0gZiLigKIge3IuZ2V0KCdyZWZlcmVuY2Vfa2V5JywnJyl9OiB7ci5nZXQoJ2Z1bGxfcnVsZScsJycpfSIK"
    "ICAgICAgICAgICAgaWYgdG90YWwgKyBsZW4oZW50cnkpID4gbWF4X2NoYXJzOgogICAgICAgICAgICAgICAgYnJlYWsKICAgICAg"
    "ICAgICAgcGFydHMuYXBwZW5kKGVudHJ5KQogICAgICAgICAgICB0b3RhbCArPSBsZW4oZW50cnkpCgogICAgICAgIHBhcnRzLmFw"
    "cGVuZChmIltFTkQge2xhbmd1YWdlLnVwcGVyKCl9IFJVTEVTXSIpCiAgICAgICAgcmV0dXJuICJcbiIuam9pbihwYXJ0cykKCiAg"
    "ICBkZWYgX2lzX2R1cGxpY2F0ZShzZWxmLCByZWZlcmVuY2Vfa2V5OiBzdHIpIC0+IGJvb2w6CiAgICAgICAgcmV0dXJuIGFueSgK"
    "ICAgICAgICAgICAgci5nZXQoInJlZmVyZW5jZV9rZXkiLCIiKS5sb3dlcigpID09IHJlZmVyZW5jZV9rZXkubG93ZXIoKQogICAg"
    "ICAgICAgICBmb3IgciBpbiByZWFkX2pzb25sKHNlbGYuX3BhdGgpCiAgICAgICAgKQoKICAgIGRlZiBzZWVkX2xzbF9ydWxlcyhz"
    "ZWxmKSAtPiBOb25lOgogICAgICAgICIiIgogICAgICAgIFNlZWQgdGhlIExTTCBGb3JiaWRkZW4gUnVsZXNldCBvbiBmaXJzdCBy"
    "dW4gaWYgdGhlIERCIGlzIGVtcHR5LgogICAgICAgIFRoZXNlIGFyZSB0aGUgaGFyZCBydWxlcyBmcm9tIHRoZSBwcm9qZWN0IHN0"
    "YW5kaW5nIHJ1bGVzLgogICAgICAgICIiIgogICAgICAgIGlmIHJlYWRfanNvbmwoc2VsZi5fcGF0aCk6CiAgICAgICAgICAgIHJl"
    "dHVybiAgIyBBbHJlYWR5IHNlZWRlZAoKICAgICAgICBsc2xfcnVsZXMgPSBbCiAgICAgICAgICAgICgiTFNMIiwgIkxTTCIsICJO"
    "T19URVJOQVJZIiwKICAgICAgICAgICAgICJObyB0ZXJuYXJ5IG9wZXJhdG9ycyBpbiBMU0wiLAogICAgICAgICAgICAgIk5ldmVy"
    "IHVzZSB0aGUgdGVybmFyeSBvcGVyYXRvciAoPzopIGluIExTTCBzY3JpcHRzLiAiCiAgICAgICAgICAgICAiVXNlIGlmL2Vsc2Ug"
    "YmxvY2tzIGluc3RlYWQuIExTTCBkb2VzIG5vdCBzdXBwb3J0IHRlcm5hcnkuIiwKICAgICAgICAgICAgICJSZXBsYWNlIHdpdGgg"
    "aWYvZWxzZSBibG9jay4iLCAiIiksCiAgICAgICAgICAgICgiTFNMIiwgIkxTTCIsICJOT19GT1JFQUNIIiwKICAgICAgICAgICAg"
    "ICJObyBmb3JlYWNoIGxvb3BzIGluIExTTCIsCiAgICAgICAgICAgICAiTFNMIGhhcyBubyBmb3JlYWNoIGxvb3AgY29uc3RydWN0"
    "LiBVc2UgaW50ZWdlciBpbmRleCB3aXRoICIKICAgICAgICAgICAgICJsbEdldExpc3RMZW5ndGgoKSBhbmQgYSBmb3Igb3Igd2hp"
    "bGUgbG9vcC4iLAogICAgICAgICAgICAgIlVzZTogZm9yKGludGVnZXIgaT0wOyBpPGxsR2V0TGlzdExlbmd0aChteUxpc3QpOyBp"
    "KyspIiwgIiIpLAogICAgICAgICAgICAoIkxTTCIsICJMU0wiLCAiTk9fR0xPQkFMX0FTU0lHTl9GUk9NX0ZVTkMiLAogICAgICAg"
    "ICAgICAgIk5vIGdsb2JhbCB2YXJpYWJsZSBhc3NpZ25tZW50cyBmcm9tIGZ1bmN0aW9uIGNhbGxzIiwKICAgICAgICAgICAgICJH"
    "bG9iYWwgdmFyaWFibGUgaW5pdGlhbGl6YXRpb24gaW4gTFNMIGNhbm5vdCBjYWxsIGZ1bmN0aW9ucy4gIgogICAgICAgICAgICAg"
    "IkluaXRpYWxpemUgZ2xvYmFscyB3aXRoIGxpdGVyYWwgdmFsdWVzIG9ubHkuICIKICAgICAgICAgICAgICJBc3NpZ24gZnJvbSBm"
    "dW5jdGlvbnMgaW5zaWRlIGV2ZW50IGhhbmRsZXJzIG9yIG90aGVyIGZ1bmN0aW9ucy4iLAogICAgICAgICAgICAgIk1vdmUgdGhl"
    "IGFzc2lnbm1lbnQgaW50byBhbiBldmVudCBoYW5kbGVyIChzdGF0ZV9lbnRyeSwgZXRjLikiLCAiIiksCiAgICAgICAgICAgICgi"
    "TFNMIiwgIkxTTCIsICJOT19WT0lEX0tFWVdPUkQiLAogICAgICAgICAgICAgIk5vIHZvaWQga2V5d29yZCBpbiBMU0wiLAogICAg"
    "ICAgICAgICAgIkxTTCBkb2VzIG5vdCBoYXZlIGEgdm9pZCBrZXl3b3JkIGZvciBmdW5jdGlvbiByZXR1cm4gdHlwZXMuICIKICAg"
    "ICAgICAgICAgICJGdW5jdGlvbnMgdGhhdCByZXR1cm4gbm90aGluZyBzaW1wbHkgb21pdCB0aGUgcmV0dXJuIHR5cGUuIiwKICAg"
    "ICAgICAgICAgICJSZW1vdmUgJ3ZvaWQnIGZyb20gZnVuY3Rpb24gc2lnbmF0dXJlLiAiCiAgICAgICAgICAgICAiZS5nLiBteUZ1"
    "bmMoKSB7IC4uLiB9IG5vdCB2b2lkIG15RnVuYygpIHsgLi4uIH0iLCAiIiksCiAgICAgICAgICAgICgiTFNMIiwgIkxTTCIsICJD"
    "T01QTEVURV9TQ1JJUFRTX09OTFkiLAogICAgICAgICAgICAgIkFsd2F5cyBwcm92aWRlIGNvbXBsZXRlIHNjcmlwdHMsIG5ldmVy"
    "IHBhcnRpYWwgZWRpdHMiLAogICAgICAgICAgICAgIldoZW4gd3JpdGluZyBvciBlZGl0aW5nIExTTCBzY3JpcHRzLCBhbHdheXMg"
    "b3V0cHV0IHRoZSBjb21wbGV0ZSAiCiAgICAgICAgICAgICAic2NyaXB0LiBOZXZlciBwcm92aWRlIHBhcnRpYWwgc25pcHBldHMg"
    "b3IgJ2FkZCB0aGlzIHNlY3Rpb24nICIKICAgICAgICAgICAgICJpbnN0cnVjdGlvbnMuIFRoZSBmdWxsIHNjcmlwdCBtdXN0IGJl"
    "IGNvcHktcGFzdGUgcmVhZHkuIiwKICAgICAgICAgICAgICJXcml0ZSB0aGUgZW50aXJlIHNjcmlwdCBmcm9tIHRvcCB0byBib3R0"
    "b20uIiwgIiIpLAogICAgICAgIF0KCiAgICAgICAgZm9yIGVudiwgbGFuZywgcmVmLCBzdW1tYXJ5LCBmdWxsX3J1bGUsIHJlc29s"
    "dXRpb24sIGxpbmsgaW4gbHNsX3J1bGVzOgogICAgICAgICAgICBzZWxmLmFkZChlbnYsIGxhbmcsIHJlZiwgc3VtbWFyeSwgZnVs"
    "bF9ydWxlLCByZXNvbHV0aW9uLCBsaW5rLAogICAgICAgICAgICAgICAgICAgICB0YWdzPVsibHNsIiwgImZvcmJpZGRlbiIsICJz"
    "dGFuZGluZ19ydWxlIl0pCgoKIyDilIDilIAgVEFTSyBNQU5BR0VSIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApj"
    "bGFzcyBUYXNrTWFuYWdlcjoKICAgICIiIgogICAgVGFzay9yZW1pbmRlciBDUlVEIGFuZCBkdWUtZXZlbnQgZGV0ZWN0aW9uLgoK"
    "ICAgIEZpbGU6IG1lbW9yaWVzL3Rhc2tzLmpzb25sCgogICAgVGFzayByZWNvcmQgZmllbGRzOgogICAgICAgIGlkLCBjcmVhdGVk"
    "X2F0LCBkdWVfYXQsIHByZV90cmlnZ2VyICgxbWluIGJlZm9yZSksCiAgICAgICAgdGV4dCwgc3RhdHVzIChwZW5kaW5nfHRyaWdn"
    "ZXJlZHxzbm9vemVkfGNvbXBsZXRlZHxjYW5jZWxsZWQpLAogICAgICAgIGFja25vd2xlZGdlZF9hdCwgcmV0cnlfY291bnQsIGxh"
    "c3RfdHJpZ2dlcmVkX2F0LCBuZXh0X3JldHJ5X2F0LAogICAgICAgIHNvdXJjZSAobG9jYWx8Z29vZ2xlKSwgZ29vZ2xlX2V2ZW50"
    "X2lkLCBzeW5jX3N0YXR1cywgbWV0YWRhdGEKCiAgICBEdWUtZXZlbnQgY3ljbGU6CiAgICAgICAgLSBQcmUtdHJpZ2dlcjogMSBt"
    "aW51dGUgYmVmb3JlIGR1ZSDihpIgYW5ub3VuY2UgdXBjb21pbmcKICAgICAgICAtIER1ZSB0cmlnZ2VyOiBhdCBkdWUgdGltZSDi"
    "hpIgYWxlcnQgc291bmQgKyBBSSBjb21tZW50YXJ5CiAgICAgICAgLSAzLW1pbnV0ZSB3aW5kb3c6IGlmIG5vdCBhY2tub3dsZWRn"
    "ZWQg4oaSIHNub296ZQogICAgICAgIC0gMTItbWludXRlIHJldHJ5OiByZS10cmlnZ2VyCiAgICAiIiIKCiAgICBkZWYgX19pbml0"
    "X18oc2VsZik6CiAgICAgICAgc2VsZi5fcGF0aCA9IGNmZ19wYXRoKCJtZW1vcmllcyIpIC8gInRhc2tzLmpzb25sIgoKICAgICMg"
    "4pSA4pSAIENSVUQg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgbG9hZF9hbGwoc2VsZikgLT4gbGlzdFtkaWN0"
    "XToKICAgICAgICB0YXNrcyA9IHJlYWRfanNvbmwoc2VsZi5fcGF0aCkKICAgICAgICBjaGFuZ2VkID0gRmFsc2UKICAgICAgICBu"
    "b3JtYWxpemVkID0gW10KICAgICAgICBmb3IgdCBpbiB0YXNrczoKICAgICAgICAgICAgaWYgbm90IGlzaW5zdGFuY2UodCwgZGlj"
    "dCk6CiAgICAgICAgICAgICAgICBjb250aW51ZQogICAgICAgICAgICBpZiAiaWQiIG5vdCBpbiB0OgogICAgICAgICAgICAgICAg"
    "dFsiaWQiXSA9IGYidGFza197dXVpZC51dWlkNCgpLmhleFs6MTBdfSIKICAgICAgICAgICAgICAgIGNoYW5nZWQgPSBUcnVlCiAg"
    "ICAgICAgICAgICMgTm9ybWFsaXplIGZpZWxkIG5hbWVzCiAgICAgICAgICAgIGlmICJkdWVfYXQiIG5vdCBpbiB0OgogICAgICAg"
    "ICAgICAgICAgdFsiZHVlX2F0Il0gPSB0LmdldCgiZHVlIikKICAgICAgICAgICAgICAgIGNoYW5nZWQgPSBUcnVlCiAgICAgICAg"
    "ICAgIHQuc2V0ZGVmYXVsdCgic3RhdHVzIiwgICAgICAgICAgICJwZW5kaW5nIikKICAgICAgICAgICAgdC5zZXRkZWZhdWx0KCJy"
    "ZXRyeV9jb3VudCIsICAgICAgMCkKICAgICAgICAgICAgdC5zZXRkZWZhdWx0KCJhY2tub3dsZWRnZWRfYXQiLCAgTm9uZSkKICAg"
    "ICAgICAgICAgdC5zZXRkZWZhdWx0KCJsYXN0X3RyaWdnZXJlZF9hdCIsTm9uZSkKICAgICAgICAgICAgdC5zZXRkZWZhdWx0KCJu"
    "ZXh0X3JldHJ5X2F0IiwgICAgTm9uZSkKICAgICAgICAgICAgdC5zZXRkZWZhdWx0KCJwcmVfYW5ub3VuY2VkIiwgICAgRmFsc2Up"
    "CiAgICAgICAgICAgIHQuc2V0ZGVmYXVsdCgic291cmNlIiwgICAgICAgICAgICJsb2NhbCIpCiAgICAgICAgICAgIHQuc2V0ZGVm"
    "YXVsdCgiZ29vZ2xlX2V2ZW50X2lkIiwgIE5vbmUpCiAgICAgICAgICAgIHQuc2V0ZGVmYXVsdCgic3luY19zdGF0dXMiLCAgICAg"
    "ICJwZW5kaW5nIikKICAgICAgICAgICAgdC5zZXRkZWZhdWx0KCJtZXRhZGF0YSIsICAgICAgICAge30pCiAgICAgICAgICAgIHQu"
    "c2V0ZGVmYXVsdCgiY3JlYXRlZF9hdCIsICAgICAgIGxvY2FsX25vd19pc28oKSkKCiAgICAgICAgICAgICMgQ29tcHV0ZSBwcmVf"
    "dHJpZ2dlciBpZiBtaXNzaW5nCiAgICAgICAgICAgIGlmIHQuZ2V0KCJkdWVfYXQiKSBhbmQgbm90IHQuZ2V0KCJwcmVfdHJpZ2dl"
    "ciIpOgogICAgICAgICAgICAgICAgZHQgPSBwYXJzZV9pc28odFsiZHVlX2F0Il0pCiAgICAgICAgICAgICAgICBpZiBkdDoKICAg"
    "ICAgICAgICAgICAgICAgICBwcmUgPSBkdCAtIHRpbWVkZWx0YShtaW51dGVzPTEpCiAgICAgICAgICAgICAgICAgICAgdFsicHJl"
    "X3RyaWdnZXIiXSA9IHByZS5pc29mb3JtYXQodGltZXNwZWM9InNlY29uZHMiKQogICAgICAgICAgICAgICAgICAgIGNoYW5nZWQg"
    "PSBUcnVlCgogICAgICAgICAgICBub3JtYWxpemVkLmFwcGVuZCh0KQoKICAgICAgICBpZiBjaGFuZ2VkOgogICAgICAgICAgICB3"
    "cml0ZV9qc29ubChzZWxmLl9wYXRoLCBub3JtYWxpemVkKQogICAgICAgIHJldHVybiBub3JtYWxpemVkCgogICAgZGVmIHNhdmVf"
    "YWxsKHNlbGYsIHRhc2tzOiBsaXN0W2RpY3RdKSAtPiBOb25lOgogICAgICAgIHdyaXRlX2pzb25sKHNlbGYuX3BhdGgsIHRhc2tz"
    "KQoKICAgIGRlZiBhZGQoc2VsZiwgdGV4dDogc3RyLCBkdWVfZHQ6IGRhdGV0aW1lLAogICAgICAgICAgICBzb3VyY2U6IHN0ciA9"
    "ICJsb2NhbCIpIC0+IGRpY3Q6CiAgICAgICAgcHJlID0gZHVlX2R0IC0gdGltZWRlbHRhKG1pbnV0ZXM9MSkKICAgICAgICB0YXNr"
    "ID0gewogICAgICAgICAgICAiaWQiOiAgICAgICAgICAgICAgIGYidGFza197dXVpZC51dWlkNCgpLmhleFs6MTBdfSIsCiAgICAg"
    "ICAgICAgICJjcmVhdGVkX2F0IjogICAgICAgbG9jYWxfbm93X2lzbygpLAogICAgICAgICAgICAiZHVlX2F0IjogICAgICAgICAg"
    "IGR1ZV9kdC5pc29mb3JtYXQodGltZXNwZWM9InNlY29uZHMiKSwKICAgICAgICAgICAgInByZV90cmlnZ2VyIjogICAgICBwcmUu"
    "aXNvZm9ybWF0KHRpbWVzcGVjPSJzZWNvbmRzIiksCiAgICAgICAgICAgICJ0ZXh0IjogICAgICAgICAgICAgdGV4dC5zdHJpcCgp"
    "LAogICAgICAgICAgICAic3RhdHVzIjogICAgICAgICAgICJwZW5kaW5nIiwKICAgICAgICAgICAgImFja25vd2xlZGdlZF9hdCI6"
    "ICBOb25lLAogICAgICAgICAgICAicmV0cnlfY291bnQiOiAgICAgIDAsCiAgICAgICAgICAgICJsYXN0X3RyaWdnZXJlZF9hdCI6"
    "Tm9uZSwKICAgICAgICAgICAgIm5leHRfcmV0cnlfYXQiOiAgICBOb25lLAogICAgICAgICAgICAicHJlX2Fubm91bmNlZCI6ICAg"
    "IEZhbHNlLAogICAgICAgICAgICAic291cmNlIjogICAgICAgICAgIHNvdXJjZSwKICAgICAgICAgICAgImdvb2dsZV9ldmVudF9p"
    "ZCI6ICBOb25lLAogICAgICAgICAgICAic3luY19zdGF0dXMiOiAgICAgICJwZW5kaW5nIiwKICAgICAgICAgICAgIm1ldGFkYXRh"
    "IjogICAgICAgICB7fSwKICAgICAgICB9CiAgICAgICAgdGFza3MgPSBzZWxmLmxvYWRfYWxsKCkKICAgICAgICB0YXNrcy5hcHBl"
    "bmQodGFzaykKICAgICAgICBzZWxmLnNhdmVfYWxsKHRhc2tzKQogICAgICAgIHJldHVybiB0YXNrCgogICAgZGVmIHVwZGF0ZV9z"
    "dGF0dXMoc2VsZiwgdGFza19pZDogc3RyLCBzdGF0dXM6IHN0ciwKICAgICAgICAgICAgICAgICAgICAgIGFja25vd2xlZGdlZDog"
    "Ym9vbCA9IEZhbHNlKSAtPiBPcHRpb25hbFtkaWN0XToKICAgICAgICB0YXNrcyA9IHNlbGYubG9hZF9hbGwoKQogICAgICAgIGZv"
    "ciB0IGluIHRhc2tzOgogICAgICAgICAgICBpZiB0LmdldCgiaWQiKSA9PSB0YXNrX2lkOgogICAgICAgICAgICAgICAgdFsic3Rh"
    "dHVzIl0gPSBzdGF0dXMKICAgICAgICAgICAgICAgIGlmIGFja25vd2xlZGdlZDoKICAgICAgICAgICAgICAgICAgICB0WyJhY2tu"
    "b3dsZWRnZWRfYXQiXSA9IGxvY2FsX25vd19pc28oKQogICAgICAgICAgICAgICAgc2VsZi5zYXZlX2FsbCh0YXNrcykKICAgICAg"
    "ICAgICAgICAgIHJldHVybiB0CiAgICAgICAgcmV0dXJuIE5vbmUKCiAgICBkZWYgY29tcGxldGUoc2VsZiwgdGFza19pZDogc3Ry"
    "KSAtPiBPcHRpb25hbFtkaWN0XToKICAgICAgICB0YXNrcyA9IHNlbGYubG9hZF9hbGwoKQogICAgICAgIGZvciB0IGluIHRhc2tz"
    "OgogICAgICAgICAgICBpZiB0LmdldCgiaWQiKSA9PSB0YXNrX2lkOgogICAgICAgICAgICAgICAgdFsic3RhdHVzIl0gICAgICAg"
    "ICAgPSAiY29tcGxldGVkIgogICAgICAgICAgICAgICAgdFsiYWNrbm93bGVkZ2VkX2F0Il0gPSBsb2NhbF9ub3dfaXNvKCkKICAg"
    "ICAgICAgICAgICAgIHNlbGYuc2F2ZV9hbGwodGFza3MpCiAgICAgICAgICAgICAgICByZXR1cm4gdAogICAgICAgIHJldHVybiBO"
    "b25lCgogICAgZGVmIGNhbmNlbChzZWxmLCB0YXNrX2lkOiBzdHIpIC0+IE9wdGlvbmFsW2RpY3RdOgogICAgICAgIHRhc2tzID0g"
    "c2VsZi5sb2FkX2FsbCgpCiAgICAgICAgZm9yIHQgaW4gdGFza3M6CiAgICAgICAgICAgIGlmIHQuZ2V0KCJpZCIpID09IHRhc2tf"
    "aWQ6CiAgICAgICAgICAgICAgICB0WyJzdGF0dXMiXSAgICAgICAgICA9ICJjYW5jZWxsZWQiCiAgICAgICAgICAgICAgICB0WyJh"
    "Y2tub3dsZWRnZWRfYXQiXSA9IGxvY2FsX25vd19pc28oKQogICAgICAgICAgICAgICAgc2VsZi5zYXZlX2FsbCh0YXNrcykKICAg"
    "ICAgICAgICAgICAgIHJldHVybiB0CiAgICAgICAgcmV0dXJuIE5vbmUKCiAgICBkZWYgY2xlYXJfY29tcGxldGVkKHNlbGYpIC0+"
    "IGludDoKICAgICAgICB0YXNrcyAgICA9IHNlbGYubG9hZF9hbGwoKQogICAgICAgIGtlcHQgICAgID0gW3QgZm9yIHQgaW4gdGFz"
    "a3MKICAgICAgICAgICAgICAgICAgICBpZiB0LmdldCgic3RhdHVzIikgbm90IGluIHsiY29tcGxldGVkIiwiY2FuY2VsbGVkIn1d"
    "CiAgICAgICAgcmVtb3ZlZCAgPSBsZW4odGFza3MpIC0gbGVuKGtlcHQpCiAgICAgICAgaWYgcmVtb3ZlZDoKICAgICAgICAgICAg"
    "c2VsZi5zYXZlX2FsbChrZXB0KQogICAgICAgIHJldHVybiByZW1vdmVkCgoKICAgICMg4pSA4pSAIERVRSBFVkVOVCBERVRFQ1RJ"
    "T04g4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgZ2V0X2R1"
    "ZV9ldmVudHMoc2VsZikgLT4gbGlzdFt0dXBsZVtzdHIsIGRpY3RdXToKICAgICAgICAiIiIKICAgICAgICBDaGVjayBhbGwgdGFz"
    "a3MgZm9yIGR1ZS9wcmUtdHJpZ2dlci9yZXRyeSBldmVudHMuCiAgICAgICAgUmV0dXJucyBsaXN0IG9mIChldmVudF90eXBlLCB0"
    "YXNrKSB0dXBsZXMuCiAgICAgICAgZXZlbnRfdHlwZTogInByZSIgfCAiZHVlIiB8ICJyZXRyeSIKCiAgICAgICAgTW9kaWZpZXMg"
    "dGFzayBzdGF0dXNlcyBpbiBwbGFjZSBhbmQgc2F2ZXMuCiAgICAgICAgQ2FsbCBmcm9tIEFQU2NoZWR1bGVyIGV2ZXJ5IDMwIHNl"
    "Y29uZHMuCiAgICAgICAgIiIiCiAgICAgICAgbm93ICAgID0gZGF0ZXRpbWUubm93KCkuYXN0aW1lem9uZSgpCiAgICAgICAgdGFz"
    "a3MgID0gc2VsZi5sb2FkX2FsbCgpCiAgICAgICAgZXZlbnRzID0gW10KICAgICAgICBjaGFuZ2VkID0gRmFsc2UKCiAgICAgICAg"
    "Zm9yIHRhc2sgaW4gdGFza3M6CiAgICAgICAgICAgIGlmIHRhc2suZ2V0KCJhY2tub3dsZWRnZWRfYXQiKToKICAgICAgICAgICAg"
    "ICAgIGNvbnRpbnVlCgogICAgICAgICAgICBzdGF0dXMgICA9IHRhc2suZ2V0KCJzdGF0dXMiLCAicGVuZGluZyIpCiAgICAgICAg"
    "ICAgIGR1ZSAgICAgID0gc2VsZi5fcGFyc2VfbG9jYWwodGFzay5nZXQoImR1ZV9hdCIpKQogICAgICAgICAgICBwcmUgICAgICA9"
    "IHNlbGYuX3BhcnNlX2xvY2FsKHRhc2suZ2V0KCJwcmVfdHJpZ2dlciIpKQogICAgICAgICAgICBuZXh0X3JldCA9IHNlbGYuX3Bh"
    "cnNlX2xvY2FsKHRhc2suZ2V0KCJuZXh0X3JldHJ5X2F0IikpCiAgICAgICAgICAgIGRlYWRsaW5lID0gc2VsZi5fcGFyc2VfbG9j"
    "YWwodGFzay5nZXQoImFsZXJ0X2RlYWRsaW5lIikpCgogICAgICAgICAgICAjIFByZS10cmlnZ2VyCiAgICAgICAgICAgIGlmIChz"
    "dGF0dXMgPT0gInBlbmRpbmciIGFuZCBwcmUgYW5kIG5vdyA+PSBwcmUKICAgICAgICAgICAgICAgICAgICBhbmQgbm90IHRhc2su"
    "Z2V0KCJwcmVfYW5ub3VuY2VkIikpOgogICAgICAgICAgICAgICAgdGFza1sicHJlX2Fubm91bmNlZCJdID0gVHJ1ZQogICAgICAg"
    "ICAgICAgICAgZXZlbnRzLmFwcGVuZCgoInByZSIsIHRhc2spKQogICAgICAgICAgICAgICAgY2hhbmdlZCA9IFRydWUKCiAgICAg"
    "ICAgICAgICMgRHVlIHRyaWdnZXIKICAgICAgICAgICAgaWYgc3RhdHVzID09ICJwZW5kaW5nIiBhbmQgZHVlIGFuZCBub3cgPj0g"
    "ZHVlOgogICAgICAgICAgICAgICAgdGFza1sic3RhdHVzIl0gICAgICAgICAgID0gInRyaWdnZXJlZCIKICAgICAgICAgICAgICAg"
    "IHRhc2tbImxhc3RfdHJpZ2dlcmVkX2F0Il09IGxvY2FsX25vd19pc28oKQogICAgICAgICAgICAgICAgdGFza1siYWxlcnRfZGVh"
    "ZGxpbmUiXSAgID0gKAogICAgICAgICAgICAgICAgICAgIGRhdGV0aW1lLm5vdygpLmFzdGltZXpvbmUoKSArIHRpbWVkZWx0YSht"
    "aW51dGVzPTMpCiAgICAgICAgICAgICAgICApLmlzb2Zvcm1hdCh0aW1lc3BlYz0ic2Vjb25kcyIpCiAgICAgICAgICAgICAgICBl"
    "dmVudHMuYXBwZW5kKCgiZHVlIiwgdGFzaykpCiAgICAgICAgICAgICAgICBjaGFuZ2VkID0gVHJ1ZQogICAgICAgICAgICAgICAg"
    "Y29udGludWUKCiAgICAgICAgICAgICMgU25vb3plIGFmdGVyIDMtbWludXRlIHdpbmRvdwogICAgICAgICAgICBpZiBzdGF0dXMg"
    "PT0gInRyaWdnZXJlZCIgYW5kIGRlYWRsaW5lIGFuZCBub3cgPj0gZGVhZGxpbmU6CiAgICAgICAgICAgICAgICB0YXNrWyJzdGF0"
    "dXMiXSAgICAgICAgPSAic25vb3plZCIKICAgICAgICAgICAgICAgIHRhc2tbIm5leHRfcmV0cnlfYXQiXSA9ICgKICAgICAgICAg"
    "ICAgICAgICAgICBkYXRldGltZS5ub3coKS5hc3RpbWV6b25lKCkgKyB0aW1lZGVsdGEobWludXRlcz0xMikKICAgICAgICAgICAg"
    "ICAgICkuaXNvZm9ybWF0KHRpbWVzcGVjPSJzZWNvbmRzIikKICAgICAgICAgICAgICAgIGNoYW5nZWQgPSBUcnVlCiAgICAgICAg"
    "ICAgICAgICBjb250aW51ZQoKICAgICAgICAgICAgIyBSZXRyeQogICAgICAgICAgICBpZiBzdGF0dXMgaW4geyJyZXRyeV9wZW5k"
    "aW5nIiwic25vb3plZCJ9IGFuZCBuZXh0X3JldCBhbmQgbm93ID49IG5leHRfcmV0OgogICAgICAgICAgICAgICAgdGFza1sic3Rh"
    "dHVzIl0gICAgICAgICAgICA9ICJ0cmlnZ2VyZWQiCiAgICAgICAgICAgICAgICB0YXNrWyJyZXRyeV9jb3VudCJdICAgICAgID0g"
    "aW50KHRhc2suZ2V0KCJyZXRyeV9jb3VudCIsMCkpICsgMQogICAgICAgICAgICAgICAgdGFza1sibGFzdF90cmlnZ2VyZWRfYXQi"
    "XSA9IGxvY2FsX25vd19pc28oKQogICAgICAgICAgICAgICAgdGFza1siYWxlcnRfZGVhZGxpbmUiXSAgICA9ICgKICAgICAgICAg"
    "ICAgICAgICAgICBkYXRldGltZS5ub3coKS5hc3RpbWV6b25lKCkgKyB0aW1lZGVsdGEobWludXRlcz0zKQogICAgICAgICAgICAg"
    "ICAgKS5pc29mb3JtYXQodGltZXNwZWM9InNlY29uZHMiKQogICAgICAgICAgICAgICAgdGFza1sibmV4dF9yZXRyeV9hdCJdICAg"
    "ICA9IE5vbmUKICAgICAgICAgICAgICAgIGV2ZW50cy5hcHBlbmQoKCJyZXRyeSIsIHRhc2spKQogICAgICAgICAgICAgICAgY2hh"
    "bmdlZCA9IFRydWUKCiAgICAgICAgaWYgY2hhbmdlZDoKICAgICAgICAgICAgc2VsZi5zYXZlX2FsbCh0YXNrcykKICAgICAgICBy"
    "ZXR1cm4gZXZlbnRzCgogICAgZGVmIF9wYXJzZV9sb2NhbChzZWxmLCB2YWx1ZTogc3RyKSAtPiBPcHRpb25hbFtkYXRldGltZV06"
    "CiAgICAgICAgIiIiUGFyc2UgSVNPIHN0cmluZyB0byB0aW1lem9uZS1hd2FyZSBkYXRldGltZSBmb3IgY29tcGFyaXNvbi4iIiIK"
    "ICAgICAgICBkdCA9IHBhcnNlX2lzbyh2YWx1ZSkKICAgICAgICBpZiBkdCBpcyBOb25lOgogICAgICAgICAgICByZXR1cm4gTm9u"
    "ZQogICAgICAgIGlmIGR0LnR6aW5mbyBpcyBOb25lOgogICAgICAgICAgICBkdCA9IGR0LmFzdGltZXpvbmUoKQogICAgICAgIHJl"
    "dHVybiBkdAoKICAgICMg4pSA4pSAIE5BVFVSQUwgTEFOR1VBR0UgUEFSU0lORyDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIAKICAgIEBzdGF0aWNtZXRob2QKICAgIGRlZiBjbGFzc2lmeV9pbnRlbnQodGV4dDogc3RyKSAtPiBkaWN0"
    "OgogICAgICAgICIiIgogICAgICAgIENsYXNzaWZ5IHVzZXIgaW5wdXQgYXMgdGFzay9yZW1pbmRlci90aW1lci9jaGF0LgogICAg"
    "ICAgIFJldHVybnMgeyJpbnRlbnQiOiBzdHIsICJjbGVhbmVkX2lucHV0Ijogc3RyfQogICAgICAgICIiIgogICAgICAgIGltcG9y"
    "dCByZQogICAgICAgICMgU3RyaXAgY29tbW9uIGludm9jYXRpb24gcHJlZml4ZXMKICAgICAgICBjbGVhbmVkID0gcmUuc3ViKAog"
    "ICAgICAgICAgICByZiJeXHMqKD86e0RFQ0tfTkFNRS5sb3dlcigpfXxoZXlccyt7REVDS19OQU1FLmxvd2VyKCl9KVxzKiw/XHMq"
    "WzpcLV0/XHMqIiwKICAgICAgICAgICAgIiIsIHRleHQsIGZsYWdzPXJlLkkKICAgICAgICApLnN0cmlwKCkKCiAgICAgICAgbG93"
    "ID0gY2xlYW5lZC5sb3dlcigpCgogICAgICAgIHRpbWVyX3BhdHMgICAgPSBbciJcYnNldCg/OlxzK2EpP1xzK3RpbWVyXGIiLCBy"
    "IlxidGltZXJccytmb3JcYiIsCiAgICAgICAgICAgICAgICAgICAgICAgICByIlxic3RhcnQoPzpccythKT9ccyt0aW1lclxiIl0K"
    "ICAgICAgICByZW1pbmRlcl9wYXRzID0gW3IiXGJyZW1pbmQgbWVcYiIsIHIiXGJzZXQoPzpccythKT9ccytyZW1pbmRlclxiIiwK"
    "ICAgICAgICAgICAgICAgICAgICAgICAgIHIiXGJhZGQoPzpccythKT9ccytyZW1pbmRlclxiIiwKICAgICAgICAgICAgICAgICAg"
    "ICAgICAgIHIiXGJzZXQoPzpccythbj8pP1xzK2FsYXJtXGIiLCByIlxiYWxhcm1ccytmb3JcYiJdCiAgICAgICAgdGFza19wYXRz"
    "ICAgICA9IFtyIlxiYWRkKD86XHMrYSk/XHMrdGFza1xiIiwKICAgICAgICAgICAgICAgICAgICAgICAgIHIiXGJjcmVhdGUoPzpc"
    "cythKT9ccyt0YXNrXGIiLCByIlxibmV3XHMrdGFza1xiIl0KCiAgICAgICAgaW1wb3J0IHJlIGFzIF9yZQogICAgICAgIGlmIGFu"
    "eShfcmUuc2VhcmNoKHAsIGxvdykgZm9yIHAgaW4gdGltZXJfcGF0cyk6CiAgICAgICAgICAgIGludGVudCA9ICJ0aW1lciIKICAg"
    "ICAgICBlbGlmIGFueShfcmUuc2VhcmNoKHAsIGxvdykgZm9yIHAgaW4gcmVtaW5kZXJfcGF0cyk6CiAgICAgICAgICAgIGludGVu"
    "dCA9ICJyZW1pbmRlciIKICAgICAgICBlbGlmIGFueShfcmUuc2VhcmNoKHAsIGxvdykgZm9yIHAgaW4gdGFza19wYXRzKToKICAg"
    "ICAgICAgICAgaW50ZW50ID0gInRhc2siCiAgICAgICAgZWxzZToKICAgICAgICAgICAgaW50ZW50ID0gImNoYXQiCgogICAgICAg"
    "IHJldHVybiB7ImludGVudCI6IGludGVudCwgImNsZWFuZWRfaW5wdXQiOiBjbGVhbmVkfQoKICAgIEBzdGF0aWNtZXRob2QKICAg"
    "IGRlZiBwYXJzZV9kdWVfZGF0ZXRpbWUodGV4dDogc3RyKSAtPiBPcHRpb25hbFtkYXRldGltZV06CiAgICAgICAgIiIiCiAgICAg"
    "ICAgUGFyc2UgbmF0dXJhbCBsYW5ndWFnZSB0aW1lIGV4cHJlc3Npb24gZnJvbSB0YXNrIHRleHQuCiAgICAgICAgSGFuZGxlczog"
    "ImluIDMwIG1pbnV0ZXMiLCAiYXQgM3BtIiwgInRvbW9ycm93IGF0IDlhbSIsCiAgICAgICAgICAgICAgICAgImluIDIgaG91cnMi"
    "LCAiYXQgMTU6MzAiLCBldGMuCiAgICAgICAgUmV0dXJucyBhIGRhdGV0aW1lIG9yIE5vbmUgaWYgdW5wYXJzZWFibGUuCiAgICAg"
    "ICAgIiIiCiAgICAgICAgaW1wb3J0IHJlCiAgICAgICAgbm93ICA9IGRhdGV0aW1lLm5vdygpCiAgICAgICAgbG93ICA9IHRleHQu"
    "bG93ZXIoKS5zdHJpcCgpCgogICAgICAgICMgImluIFggbWludXRlcy9ob3Vycy9kYXlzIgogICAgICAgIG0gPSByZS5zZWFyY2go"
    "CiAgICAgICAgICAgIHIiaW5ccysoXGQrKVxzKihtaW51dGV8bWlufGhvdXJ8aHJ8ZGF5fHNlY29uZHxzZWMpIiwKICAgICAgICAg"
    "ICAgbG93CiAgICAgICAgKQogICAgICAgIGlmIG06CiAgICAgICAgICAgIG4gICAgPSBpbnQobS5ncm91cCgxKSkKICAgICAgICAg"
    "ICAgdW5pdCA9IG0uZ3JvdXAoMikKICAgICAgICAgICAgaWYgIm1pbiIgaW4gdW5pdDogIHJldHVybiBub3cgKyB0aW1lZGVsdGEo"
    "bWludXRlcz1uKQogICAgICAgICAgICBpZiAiaG91ciIgaW4gdW5pdCBvciAiaHIiIGluIHVuaXQ6IHJldHVybiBub3cgKyB0aW1l"
    "ZGVsdGEoaG91cnM9bikKICAgICAgICAgICAgaWYgImRheSIgIGluIHVuaXQ6IHJldHVybiBub3cgKyB0aW1lZGVsdGEoZGF5cz1u"
    "KQogICAgICAgICAgICBpZiAic2VjIiAgaW4gdW5pdDogcmV0dXJuIG5vdyArIHRpbWVkZWx0YShzZWNvbmRzPW4pCgogICAgICAg"
    "ICMgImF0IEhIOk1NIiBvciAiYXQgSDpNTWFtL3BtIgogICAgICAgIG0gPSByZS5zZWFyY2goCiAgICAgICAgICAgIHIiYXRccyso"
    "XGR7MSwyfSkoPzo6KFxkezJ9KSk/XHMqKGFtfHBtKT8iLAogICAgICAgICAgICBsb3cKICAgICAgICApCiAgICAgICAgaWYgbToK"
    "ICAgICAgICAgICAgaHIgID0gaW50KG0uZ3JvdXAoMSkpCiAgICAgICAgICAgIG1uICA9IGludChtLmdyb3VwKDIpKSBpZiBtLmdy"
    "b3VwKDIpIGVsc2UgMAogICAgICAgICAgICBhcG0gPSBtLmdyb3VwKDMpCiAgICAgICAgICAgIGlmIGFwbSA9PSAicG0iIGFuZCBo"
    "ciA8IDEyOiBociArPSAxMgogICAgICAgICAgICBpZiBhcG0gPT0gImFtIiBhbmQgaHIgPT0gMTI6IGhyID0gMAogICAgICAgICAg"
    "ICBkdCA9IG5vdy5yZXBsYWNlKGhvdXI9aHIsIG1pbnV0ZT1tbiwgc2Vjb25kPTAsIG1pY3Jvc2Vjb25kPTApCiAgICAgICAgICAg"
    "IGlmIGR0IDw9IG5vdzoKICAgICAgICAgICAgICAgIGR0ICs9IHRpbWVkZWx0YShkYXlzPTEpCiAgICAgICAgICAgIHJldHVybiBk"
    "dAoKICAgICAgICAjICJ0b21vcnJvdyBhdCAuLi4iICAocmVjdXJzZSBvbiB0aGUgImF0IiBwYXJ0KQogICAgICAgIGlmICJ0b21v"
    "cnJvdyIgaW4gbG93OgogICAgICAgICAgICB0b21vcnJvd190ZXh0ID0gcmUuc3ViKHIidG9tb3Jyb3ciLCAiIiwgbG93KS5zdHJp"
    "cCgpCiAgICAgICAgICAgIHJlc3VsdCA9IFRhc2tNYW5hZ2VyLnBhcnNlX2R1ZV9kYXRldGltZSh0b21vcnJvd190ZXh0KQogICAg"
    "ICAgICAgICBpZiByZXN1bHQ6CiAgICAgICAgICAgICAgICByZXR1cm4gcmVzdWx0ICsgdGltZWRlbHRhKGRheXM9MSkKCiAgICAg"
    "ICAgcmV0dXJuIE5vbmUKCgojIOKUgOKUgCBSRVFVSVJFTUVOVFMuVFhUIEdFTkVSQVRPUiDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZGVmIHdyaXRlX3JlcXVpcmVtZW50c190eHQo"
    "KSAtPiBOb25lOgogICAgIiIiCiAgICBXcml0ZSByZXF1aXJlbWVudHMudHh0IG5leHQgdG8gdGhlIGRlY2sgZmlsZSBvbiBmaXJz"
    "dCBydW4uCiAgICBIZWxwcyB1c2VycyBpbnN0YWxsIGFsbCBkZXBlbmRlbmNpZXMgd2l0aCBvbmUgcGlwIGNvbW1hbmQuCiAgICAi"
    "IiIKICAgIHJlcV9wYXRoID0gUGF0aChDRkcuZ2V0KCJiYXNlX2RpciIsIHN0cihTQ1JJUFRfRElSKSkpIC8gInJlcXVpcmVtZW50"
    "cy50eHQiCiAgICBpZiByZXFfcGF0aC5leGlzdHMoKToKICAgICAgICByZXR1cm4KCiAgICBjb250ZW50ID0gIiIiXAojIE1vcmdh"
    "bm5hIERlY2sg4oCUIFJlcXVpcmVkIERlcGVuZGVuY2llcwojIEluc3RhbGwgYWxsIHdpdGg6IHBpcCBpbnN0YWxsIC1yIHJlcXVp"
    "cmVtZW50cy50eHQKCiMgQ29yZSBVSQpQeVNpZGU2CgojIFNjaGVkdWxpbmcgKGlkbGUgdGltZXIsIGF1dG9zYXZlLCByZWZsZWN0"
    "aW9uIGN5Y2xlcykKYXBzY2hlZHVsZXIKCiMgTG9nZ2luZwpsb2d1cnUKCiMgU291bmQgcGxheWJhY2sgKFdBViArIE1QMykKcHln"
    "YW1lCgojIERlc2t0b3Agc2hvcnRjdXQgY3JlYXRpb24gKFdpbmRvd3Mgb25seSkKcHl3aW4zMgoKIyBTeXN0ZW0gbW9uaXRvcmlu"
    "ZyAoQ1BVLCBSQU0sIGRyaXZlcywgbmV0d29yaykKcHN1dGlsCgojIEhUVFAgcmVxdWVzdHMKcmVxdWVzdHMKCiMg4pSA4pSAIE9w"
    "dGlvbmFsIChsb2NhbCBtb2RlbCBvbmx5KSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIAKIyBVbmNvbW1lbnQgaWYgdXNpbmcgYSBsb2NhbCBIdWdnaW5nRmFjZSBtb2RlbDoKIyB0b3JjaAojIHRy"
    "YW5zZm9ybWVycwojIGFjY2VsZXJhdGUKCiMg4pSA4pSAIE9wdGlvbmFsIChOVklESUEgR1BVIG1vbml0b3JpbmcpIOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAojIFVuY29tbWVudCBpZiB5b3UgaGF2ZSBhbiBOVklESUEgR1BV"
    "OgojIHB5bnZtbAoiIiIKICAgIHJlcV9wYXRoLndyaXRlX3RleHQoY29udGVudCwgZW5jb2Rpbmc9InV0Zi04IikKCgojIOKUgOKU"
    "gCBQQVNTIDQgQ09NUExFVEUg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiMgTWVtb3J5LCBTZXNzaW9uLCBMZXNzb25zTGVh"
    "cm5lZCwgVGFza01hbmFnZXIgYWxsIGRlZmluZWQuCiMgTFNMIEZvcmJpZGRlbiBSdWxlc2V0IGF1dG8tc2VlZGVkIG9uIGZpcnN0"
    "IHJ1bi4KIyByZXF1aXJlbWVudHMudHh0IHdyaXR0ZW4gb24gZmlyc3QgcnVuLgojCiMgTmV4dDogUGFzcyA1IOKAlCBUYWIgQ29u"
    "dGVudCBDbGFzc2VzCiMgKFNMU2NhbnNUYWIsIFNMQ29tbWFuZHNUYWIsIEpvYlRyYWNrZXJUYWIsIFNlbGZUYWIsIERpYWdub3N0"
    "aWNzVGFiKQoKCiMg4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQCiMgTU9SR0FOTkEgREVDSyDigJQgUEFTUyA1OiBUQUIgQ09OVEVOVCBDTEFTU0VTCiMK"
    "IyBUYWJzIGRlZmluZWQgaGVyZToKIyAgIFNMU2NhbnNUYWIgICAgICDigJQgZ3JpbW9pcmUtY2FyZCBzdHlsZSwgcmVidWlsdCAo"
    "RGVsZXRlIGFkZGVkLCBNb2RpZnkgZml4ZWQsCiMgICAgICAgICAgICAgICAgICAgICBwYXJzZXIgZml4ZWQsIGNvcHktdG8tY2xp"
    "cGJvYXJkIHBlciBpdGVtKQojICAgU0xDb21tYW5kc1RhYiAgIOKAlCBnb3RoaWMgdGFibGUsIGNvcHkgY29tbWFuZCB0byBjbGlw"
    "Ym9hcmQKIyAgIEpvYlRyYWNrZXJUYWIgICDigJQgZnVsbCByZWJ1aWxkIGZyb20gc3BlYywgQ1NWL1RTViBleHBvcnQKIyAgIFNl"
    "bGZUYWIgICAgICAgICDigJQgaWRsZSBuYXJyYXRpdmUgb3V0cHV0ICsgUG9JIGxpc3QKIyAgIERpYWdub3N0aWNzVGFiICDigJQg"
    "bG9ndXJ1IG91dHB1dCArIGhhcmR3YXJlIHJlcG9ydCArIGpvdXJuYWwgbG9hZCBub3RpY2VzCiMgICBMZXNzb25zVGFiICAgICAg"
    "4oCUIExTTCBGb3JiaWRkZW4gUnVsZXNldCArIGNvZGUgbGVzc29ucyBicm93c2VyCiMg4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQCgppbXBvcnQgcmUg"
    "YXMgX3JlCgoKIyDilIDilIAgU0hBUkVEIEdPVEhJQyBUQUJMRSBTVFlMRSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZGVmIF9nb3RoaWNfdGFibGVfc3R5bGUoKSAtPiBzdHI6"
    "CiAgICByZXR1cm4gZiIiIgogICAgICAgIFFUYWJsZVdpZGdldCB7ewogICAgICAgICAgICBiYWNrZ3JvdW5kOiB7Q19CRzJ9Owog"
    "ICAgICAgICAgICBjb2xvcjoge0NfR09MRH07CiAgICAgICAgICAgIGJvcmRlcjogMXB4IHNvbGlkIHtDX0NSSU1TT05fRElNfTsK"
    "ICAgICAgICAgICAgZ3JpZGxpbmUtY29sb3I6IHtDX0JPUkRFUn07CiAgICAgICAgICAgIGZvbnQtZmFtaWx5OiB7REVDS19GT05U"
    "fSwgc2VyaWY7CiAgICAgICAgICAgIGZvbnQtc2l6ZTogMTFweDsKICAgICAgICB9fQogICAgICAgIFFUYWJsZVdpZGdldDo6aXRl"
    "bTpzZWxlY3RlZCB7ewogICAgICAgICAgICBiYWNrZ3JvdW5kOiB7Q19DUklNU09OX0RJTX07CiAgICAgICAgICAgIGNvbG9yOiB7"
    "Q19HT0xEX0JSSUdIVH07CiAgICAgICAgfX0KICAgICAgICBRVGFibGVXaWRnZXQ6Oml0ZW06YWx0ZXJuYXRlIHt7CiAgICAgICAg"
    "ICAgIGJhY2tncm91bmQ6IHtDX0JHM307CiAgICAgICAgfX0KICAgICAgICBRSGVhZGVyVmlldzo6c2VjdGlvbiB7ewogICAgICAg"
    "ICAgICBiYWNrZ3JvdW5kOiB7Q19CRzN9OwogICAgICAgICAgICBjb2xvcjoge0NfR09MRH07CiAgICAgICAgICAgIGJvcmRlcjog"
    "MXB4IHNvbGlkIHtDX0NSSU1TT05fRElNfTsKICAgICAgICAgICAgcGFkZGluZzogNHB4IDZweDsKICAgICAgICAgICAgZm9udC1m"
    "YW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsKICAgICAgICAgICAgZm9udC1zaXplOiAxMHB4OwogICAgICAgICAgICBmb250LXdl"
    "aWdodDogYm9sZDsKICAgICAgICAgICAgbGV0dGVyLXNwYWNpbmc6IDFweDsKICAgICAgICB9fQogICAgIiIiCgpkZWYgX2dvdGhp"
    "Y19idG4odGV4dDogc3RyLCB0b29sdGlwOiBzdHIgPSAiIikgLT4gUVB1c2hCdXR0b246CiAgICBidG4gPSBRUHVzaEJ1dHRvbih0"
    "ZXh0KQogICAgYnRuLnNldFN0eWxlU2hlZXQoCiAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19DUklNU09OX0RJTX07IGNvbG9yOiB7"
    "Q19HT0xEfTsgIgogICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTn07IGJvcmRlci1yYWRpdXM6IDJweDsgIgog"
    "ICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsgZm9udC1zaXplOiAxMHB4OyAiCiAgICAgICAgZiJmb250"
    "LXdlaWdodDogYm9sZDsgcGFkZGluZzogNHB4IDEwcHg7IGxldHRlci1zcGFjaW5nOiAxcHg7IgogICAgKQogICAgaWYgdG9vbHRp"
    "cDoKICAgICAgICBidG4uc2V0VG9vbFRpcCh0b29sdGlwKQogICAgcmV0dXJuIGJ0bgoKZGVmIF9zZWN0aW9uX2xibCh0ZXh0OiBz"
    "dHIpIC0+IFFMYWJlbDoKICAgIGxibCA9IFFMYWJlbCh0ZXh0KQogICAgbGJsLnNldFN0eWxlU2hlZXQoCiAgICAgICAgZiJjb2xv"
    "cjoge0NfR09MRH07IGZvbnQtc2l6ZTogOXB4OyBmb250LXdlaWdodDogYm9sZDsgIgogICAgICAgIGYibGV0dGVyLXNwYWNpbmc6"
    "IDJweDsgZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsiCiAgICApCiAgICByZXR1cm4gbGJsCgoKIyDilIDilIAgU0wg"
    "U0NBTlMgVEFCIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjbGFzcyBTTFNjYW5zVGFiKFFXaWRnZXQpOgog"
    "ICAgIiIiCiAgICBTZWNvbmQgTGlmZSBhdmF0YXIgc2Nhbm5lciByZXN1bHRzIG1hbmFnZXIuCiAgICBSZWJ1aWx0IGZyb20gc3Bl"
    "YzoKICAgICAgLSBDYXJkL2dyaW1vaXJlLWVudHJ5IHN0eWxlIGRpc3BsYXkKICAgICAgLSBBZGQgKHdpdGggdGltZXN0YW1wLWF3"
    "YXJlIHBhcnNlcikKICAgICAgLSBEaXNwbGF5IChjbGVhbiBpdGVtL2NyZWF0b3IgdGFibGUpCiAgICAgIC0gTW9kaWZ5IChlZGl0"
    "IG5hbWUsIGRlc2NyaXB0aW9uLCBpbmRpdmlkdWFsIGl0ZW1zKQogICAgICAtIERlbGV0ZSAod2FzIG1pc3Npbmcg4oCUIG5vdyBw"
    "cmVzZW50KQogICAgICAtIFJlLXBhcnNlICh3YXMgJ1JlZnJlc2gnIOKAlCByZS1ydW5zIHBhcnNlciBvbiBzdG9yZWQgcmF3IHRl"
    "eHQpCiAgICAgIC0gQ29weS10by1jbGlwYm9hcmQgb24gYW55IGl0ZW0KICAgICIiIgoKICAgIGRlZiBfX2luaXRfXyhzZWxmLCBt"
    "ZW1vcnlfZGlyOiBQYXRoLCBwYXJlbnQ9Tm9uZSk6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2Vs"
    "Zi5fcGF0aCAgICA9IGNmZ19wYXRoKCJzbCIpIC8gInNsX3NjYW5zLmpzb25sIgogICAgICAgIHNlbGYuX3JlY29yZHM6IGxpc3Rb"
    "ZGljdF0gPSBbXQogICAgICAgIHNlbGYuX3NlbGVjdGVkX2lkOiBPcHRpb25hbFtzdHJdID0gTm9uZQogICAgICAgIHNlbGYuX3Nl"
    "dHVwX3VpKCkKICAgICAgICBzZWxmLnJlZnJlc2goKQoKICAgIGRlZiBfc2V0dXBfdWkoc2VsZikgLT4gTm9uZToKICAgICAgICBy"
    "b290ID0gUVZCb3hMYXlvdXQoc2VsZikKICAgICAgICByb290LnNldENvbnRlbnRzTWFyZ2lucyg0LCA0LCA0LCA0KQogICAgICAg"
    "IHJvb3Quc2V0U3BhY2luZyg0KQoKICAgICAgICAjIEJ1dHRvbiBiYXIKICAgICAgICBiYXIgPSBRSEJveExheW91dCgpCiAgICAg"
    "ICAgc2VsZi5fYnRuX2FkZCAgICAgPSBfZ290aGljX2J0bigi4pymIEFkZCIsICAgICAiQWRkIGEgbmV3IHNjYW4iKQogICAgICAg"
    "IHNlbGYuX2J0bl9kaXNwbGF5ID0gX2dvdGhpY19idG4oIuKdpyBEaXNwbGF5IiwgIlNob3cgc2VsZWN0ZWQgc2NhbiBkZXRhaWxz"
    "IikKICAgICAgICBzZWxmLl9idG5fbW9kaWZ5ICA9IF9nb3RoaWNfYnRuKCLinKcgTW9kaWZ5IiwgICJFZGl0IHNlbGVjdGVkIHNj"
    "YW4iKQogICAgICAgIHNlbGYuX2J0bl9kZWxldGUgID0gX2dvdGhpY19idG4oIuKclyBEZWxldGUiLCAgIkRlbGV0ZSBzZWxlY3Rl"
    "ZCBzY2FuIikKICAgICAgICBzZWxmLl9idG5fcmVwYXJzZSA9IF9nb3RoaWNfYnRuKCLihrsgUmUtcGFyc2UiLCJSZS1wYXJzZSBy"
    "YXcgdGV4dCBvZiBzZWxlY3RlZCBzY2FuIikKICAgICAgICBzZWxmLl9idG5fYWRkLmNsaWNrZWQuY29ubmVjdChzZWxmLl9zaG93"
    "X2FkZCkKICAgICAgICBzZWxmLl9idG5fZGlzcGxheS5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fc2hvd19kaXNwbGF5KQogICAgICAg"
    "IHNlbGYuX2J0bl9tb2RpZnkuY2xpY2tlZC5jb25uZWN0KHNlbGYuX3Nob3dfbW9kaWZ5KQogICAgICAgIHNlbGYuX2J0bl9kZWxl"
    "dGUuY2xpY2tlZC5jb25uZWN0KHNlbGYuX2RvX2RlbGV0ZSkKICAgICAgICBzZWxmLl9idG5fcmVwYXJzZS5jbGlja2VkLmNvbm5l"
    "Y3Qoc2VsZi5fZG9fcmVwYXJzZSkKICAgICAgICBmb3IgYiBpbiAoc2VsZi5fYnRuX2FkZCwgc2VsZi5fYnRuX2Rpc3BsYXksIHNl"
    "bGYuX2J0bl9tb2RpZnksCiAgICAgICAgICAgICAgICAgIHNlbGYuX2J0bl9kZWxldGUsIHNlbGYuX2J0bl9yZXBhcnNlKToKICAg"
    "ICAgICAgICAgYmFyLmFkZFdpZGdldChiKQogICAgICAgIGJhci5hZGRTdHJldGNoKCkKICAgICAgICByb290LmFkZExheW91dChi"
    "YXIpCgogICAgICAgICMgU3RhY2s6IGxpc3QgdmlldyB8IGFkZCBmb3JtIHwgZGlzcGxheSB8IG1vZGlmeQogICAgICAgIHNlbGYu"
    "X3N0YWNrID0gUVN0YWNrZWRXaWRnZXQoKQogICAgICAgIHJvb3QuYWRkV2lkZ2V0KHNlbGYuX3N0YWNrLCAxKQoKICAgICAgICAj"
    "IOKUgOKUgCBQQUdFIDA6IHNjYW4gbGlzdCAoZ3JpbW9pcmUgY2FyZHMpIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIHAwID0gUVdpZGdldCgpCiAgICAgICAgbDAg"
    "PSBRVkJveExheW91dChwMCkKICAgICAgICBsMC5zZXRDb250ZW50c01hcmdpbnMoMCwgMCwgMCwgMCkKICAgICAgICBzZWxmLl9j"
    "YXJkX3Njcm9sbCA9IFFTY3JvbGxBcmVhKCkKICAgICAgICBzZWxmLl9jYXJkX3Njcm9sbC5zZXRXaWRnZXRSZXNpemFibGUoVHJ1"
    "ZSkKICAgICAgICBzZWxmLl9jYXJkX3Njcm9sbC5zZXRTdHlsZVNoZWV0KGYiYmFja2dyb3VuZDoge0NfQkcyfTsgYm9yZGVyOiBu"
    "b25lOyIpCiAgICAgICAgc2VsZi5fY2FyZF9jb250YWluZXIgPSBRV2lkZ2V0KCkKICAgICAgICBzZWxmLl9jYXJkX2xheW91dCAg"
    "ICA9IFFWQm94TGF5b3V0KHNlbGYuX2NhcmRfY29udGFpbmVyKQogICAgICAgIHNlbGYuX2NhcmRfbGF5b3V0LnNldENvbnRlbnRz"
    "TWFyZ2lucyg0LCA0LCA0LCA0KQogICAgICAgIHNlbGYuX2NhcmRfbGF5b3V0LnNldFNwYWNpbmcoNCkKICAgICAgICBzZWxmLl9j"
    "YXJkX2xheW91dC5hZGRTdHJldGNoKCkKICAgICAgICBzZWxmLl9jYXJkX3Njcm9sbC5zZXRXaWRnZXQoc2VsZi5fY2FyZF9jb250"
    "YWluZXIpCiAgICAgICAgbDAuYWRkV2lkZ2V0KHNlbGYuX2NhcmRfc2Nyb2xsKQogICAgICAgIHNlbGYuX3N0YWNrLmFkZFdpZGdl"
    "dChwMCkKCiAgICAgICAgIyDilIDilIAgUEFHRSAxOiBhZGQgZm9ybSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIAKICAgICAgICBwMSA9IFFXaWRnZXQoKQogICAgICAgIGwxID0gUVZCb3hMYXlvdXQocDEpCiAgICAg"
    "ICAgbDEuc2V0Q29udGVudHNNYXJnaW5zKDQsIDQsIDQsIDQpCiAgICAgICAgbDEuc2V0U3BhY2luZyg0KQogICAgICAgIGwxLmFk"
    "ZFdpZGdldChfc2VjdGlvbl9sYmwoIuKdpyBTQ0FOIE5BTUUgKGF1dG8tZGV0ZWN0ZWQpIikpCiAgICAgICAgc2VsZi5fYWRkX25h"
    "bWUgID0gUUxpbmVFZGl0KCkKICAgICAgICBzZWxmLl9hZGRfbmFtZS5zZXRQbGFjZWhvbGRlclRleHQoIkF1dG8tZGV0ZWN0ZWQg"
    "ZnJvbSBzY2FuIHRleHQiKQogICAgICAgIGwxLmFkZFdpZGdldChzZWxmLl9hZGRfbmFtZSkKICAgICAgICBsMS5hZGRXaWRnZXQo"
    "X3NlY3Rpb25fbGJsKCLinacgREVTQ1JJUFRJT04iKSkKICAgICAgICBzZWxmLl9hZGRfZGVzYyAgPSBRVGV4dEVkaXQoKQogICAg"
    "ICAgIHNlbGYuX2FkZF9kZXNjLnNldE1heGltdW1IZWlnaHQoNjApCiAgICAgICAgbDEuYWRkV2lkZ2V0KHNlbGYuX2FkZF9kZXNj"
    "KQogICAgICAgIGwxLmFkZFdpZGdldChfc2VjdGlvbl9sYmwoIuKdpyBSQVcgU0NBTiBURVhUIChwYXN0ZSBoZXJlKSIpKQogICAg"
    "ICAgIHNlbGYuX2FkZF9yYXcgICA9IFFUZXh0RWRpdCgpCiAgICAgICAgc2VsZi5fYWRkX3Jhdy5zZXRQbGFjZWhvbGRlclRleHQo"
    "CiAgICAgICAgICAgICJQYXN0ZSB0aGUgcmF3IFNlY29uZCBMaWZlIHNjYW4gb3V0cHV0IGhlcmUuXG4iCiAgICAgICAgICAgICJU"
    "aW1lc3RhbXBzIGxpa2UgWzExOjQ3XSB3aWxsIGJlIHVzZWQgdG8gc3BsaXQgaXRlbXMgY29ycmVjdGx5LiIKICAgICAgICApCiAg"
    "ICAgICAgbDEuYWRkV2lkZ2V0KHNlbGYuX2FkZF9yYXcsIDEpCiAgICAgICAgIyBQcmV2aWV3IG9mIHBhcnNlZCBpdGVtcwogICAg"
    "ICAgIGwxLmFkZFdpZGdldChfc2VjdGlvbl9sYmwoIuKdpyBQQVJTRUQgSVRFTVMgUFJFVklFVyIpKQogICAgICAgIHNlbGYuX2Fk"
    "ZF9wcmV2aWV3ID0gUVRhYmxlV2lkZ2V0KDAsIDIpCiAgICAgICAgc2VsZi5fYWRkX3ByZXZpZXcuc2V0SG9yaXpvbnRhbEhlYWRl"
    "ckxhYmVscyhbIkl0ZW0iLCAiQ3JlYXRvciJdKQogICAgICAgIHNlbGYuX2FkZF9wcmV2aWV3Lmhvcml6b250YWxIZWFkZXIoKS5z"
    "ZXRTZWN0aW9uUmVzaXplTW9kZSgKICAgICAgICAgICAgMCwgUUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJldGNoKQogICAgICAg"
    "IHNlbGYuX2FkZF9wcmV2aWV3Lmhvcml6b250YWxIZWFkZXIoKS5zZXRTZWN0aW9uUmVzaXplTW9kZSgKICAgICAgICAgICAgMSwg"
    "UUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJldGNoKQogICAgICAgIHNlbGYuX2FkZF9wcmV2aWV3LnNldE1heGltdW1IZWlnaHQo"
    "MTIwKQogICAgICAgIHNlbGYuX2FkZF9wcmV2aWV3LnNldFN0eWxlU2hlZXQoX2dvdGhpY190YWJsZV9zdHlsZSgpKQogICAgICAg"
    "IGwxLmFkZFdpZGdldChzZWxmLl9hZGRfcHJldmlldykKICAgICAgICBzZWxmLl9hZGRfcmF3LnRleHRDaGFuZ2VkLmNvbm5lY3Qo"
    "c2VsZi5fcHJldmlld19wYXJzZSkKCiAgICAgICAgYnRuczEgPSBRSEJveExheW91dCgpCiAgICAgICAgczEgPSBfZ290aGljX2J0"
    "bigi4pymIFNhdmUiKTsgYzEgPSBfZ290aGljX2J0bigi4pyXIENhbmNlbCIpCiAgICAgICAgczEuY2xpY2tlZC5jb25uZWN0KHNl"
    "bGYuX2RvX2FkZCkKICAgICAgICBjMS5jbGlja2VkLmNvbm5lY3QobGFtYmRhOiBzZWxmLl9zdGFjay5zZXRDdXJyZW50SW5kZXgo"
    "MCkpCiAgICAgICAgYnRuczEuYWRkV2lkZ2V0KHMxKTsgYnRuczEuYWRkV2lkZ2V0KGMxKTsgYnRuczEuYWRkU3RyZXRjaCgpCiAg"
    "ICAgICAgbDEuYWRkTGF5b3V0KGJ0bnMxKQogICAgICAgIHNlbGYuX3N0YWNrLmFkZFdpZGdldChwMSkKCiAgICAgICAgIyDilIDi"
    "lIAgUEFHRSAyOiBkaXNwbGF5IOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAog"
    "ICAgICAgIHAyID0gUVdpZGdldCgpCiAgICAgICAgbDIgPSBRVkJveExheW91dChwMikKICAgICAgICBsMi5zZXRDb250ZW50c01h"
    "cmdpbnMoNCwgNCwgNCwgNCkKICAgICAgICBzZWxmLl9kaXNwX25hbWUgID0gUUxhYmVsKCkKICAgICAgICBzZWxmLl9kaXNwX25h"
    "bWUuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJjb2xvcjoge0NfR09MRF9CUklHSFR9OyBmb250LXNpemU6IDEzcHg7IGZv"
    "bnQtd2VpZ2h0OiBib2xkOyAiCiAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsiCiAgICAgICAg"
    "KQogICAgICAgIHNlbGYuX2Rpc3BfZGVzYyAgPSBRTGFiZWwoKQogICAgICAgIHNlbGYuX2Rpc3BfZGVzYy5zZXRXb3JkV3JhcChU"
    "cnVlKQogICAgICAgIHNlbGYuX2Rpc3BfZGVzYy5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19URVhUX0RJ"
    "TX07IGZvbnQtc2l6ZTogMTBweDsgZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsiCiAgICAgICAgKQogICAgICAgIHNl"
    "bGYuX2Rpc3BfdGFibGUgPSBRVGFibGVXaWRnZXQoMCwgMikKICAgICAgICBzZWxmLl9kaXNwX3RhYmxlLnNldEhvcml6b250YWxI"
    "ZWFkZXJMYWJlbHMoWyJJdGVtIiwgIkNyZWF0b3IiXSkKICAgICAgICBzZWxmLl9kaXNwX3RhYmxlLmhvcml6b250YWxIZWFkZXIo"
    "KS5zZXRTZWN0aW9uUmVzaXplTW9kZSgKICAgICAgICAgICAgMCwgUUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJldGNoKQogICAg"
    "ICAgIHNlbGYuX2Rpc3BfdGFibGUuaG9yaXpvbnRhbEhlYWRlcigpLnNldFNlY3Rpb25SZXNpemVNb2RlKAogICAgICAgICAgICAx"
    "LCBRSGVhZGVyVmlldy5SZXNpemVNb2RlLlN0cmV0Y2gpCiAgICAgICAgc2VsZi5fZGlzcF90YWJsZS5zZXRTdHlsZVNoZWV0KF9n"
    "b3RoaWNfdGFibGVfc3R5bGUoKSkKICAgICAgICBzZWxmLl9kaXNwX3RhYmxlLnNldENvbnRleHRNZW51UG9saWN5KAogICAgICAg"
    "ICAgICBRdC5Db250ZXh0TWVudVBvbGljeS5DdXN0b21Db250ZXh0TWVudSkKICAgICAgICBzZWxmLl9kaXNwX3RhYmxlLmN1c3Rv"
    "bUNvbnRleHRNZW51UmVxdWVzdGVkLmNvbm5lY3QoCiAgICAgICAgICAgIHNlbGYuX2l0ZW1fY29udGV4dF9tZW51KQoKICAgICAg"
    "ICBsMi5hZGRXaWRnZXQoc2VsZi5fZGlzcF9uYW1lKQogICAgICAgIGwyLmFkZFdpZGdldChzZWxmLl9kaXNwX2Rlc2MpCiAgICAg"
    "ICAgbDIuYWRkV2lkZ2V0KHNlbGYuX2Rpc3BfdGFibGUsIDEpCgogICAgICAgIGNvcHlfaGludCA9IFFMYWJlbCgiUmlnaHQtY2xp"
    "Y2sgYW55IGl0ZW0gdG8gY29weSBpdCB0byBjbGlwYm9hcmQuIikKICAgICAgICBjb3B5X2hpbnQuc2V0U3R5bGVTaGVldCgKICAg"
    "ICAgICAgICAgZiJjb2xvcjoge0NfVEVYVF9ESU19OyBmb250LXNpemU6IDlweDsgZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBz"
    "ZXJpZjsiCiAgICAgICAgKQogICAgICAgIGwyLmFkZFdpZGdldChjb3B5X2hpbnQpCgogICAgICAgIGJrMiA9IF9nb3RoaWNfYnRu"
    "KCLil4AgQmFjayIpCiAgICAgICAgYmsyLmNsaWNrZWQuY29ubmVjdChsYW1iZGE6IHNlbGYuX3N0YWNrLnNldEN1cnJlbnRJbmRl"
    "eCgwKSkKICAgICAgICBsMi5hZGRXaWRnZXQoYmsyKQogICAgICAgIHNlbGYuX3N0YWNrLmFkZFdpZGdldChwMikKCiAgICAgICAg"
    "IyDilIDilIAgUEFHRSAzOiBtb2RpZnkg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSACiAgICAgICAgcDMgPSBRV2lkZ2V0KCkKICAgICAgICBsMyA9IFFWQm94TGF5b3V0KHAzKQogICAgICAgIGwzLnNldENv"
    "bnRlbnRzTWFyZ2lucyg0LCA0LCA0LCA0KQogICAgICAgIGwzLnNldFNwYWNpbmcoNCkKICAgICAgICBsMy5hZGRXaWRnZXQoX3Nl"
    "Y3Rpb25fbGJsKCLinacgTkFNRSIpKQogICAgICAgIHNlbGYuX21vZF9uYW1lID0gUUxpbmVFZGl0KCkKICAgICAgICBsMy5hZGRX"
    "aWRnZXQoc2VsZi5fbW9kX25hbWUpCiAgICAgICAgbDMuYWRkV2lkZ2V0KF9zZWN0aW9uX2xibCgi4p2nIERFU0NSSVBUSU9OIikp"
    "CiAgICAgICAgc2VsZi5fbW9kX2Rlc2MgPSBRTGluZUVkaXQoKQogICAgICAgIGwzLmFkZFdpZGdldChzZWxmLl9tb2RfZGVzYykK"
    "ICAgICAgICBsMy5hZGRXaWRnZXQoX3NlY3Rpb25fbGJsKCLinacgSVRFTVMgKGRvdWJsZS1jbGljayB0byBlZGl0KSIpKQogICAg"
    "ICAgIHNlbGYuX21vZF90YWJsZSA9IFFUYWJsZVdpZGdldCgwLCAyKQogICAgICAgIHNlbGYuX21vZF90YWJsZS5zZXRIb3Jpem9u"
    "dGFsSGVhZGVyTGFiZWxzKFsiSXRlbSIsICJDcmVhdG9yIl0pCiAgICAgICAgc2VsZi5fbW9kX3RhYmxlLmhvcml6b250YWxIZWFk"
    "ZXIoKS5zZXRTZWN0aW9uUmVzaXplTW9kZSgKICAgICAgICAgICAgMCwgUUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJldGNoKQog"
    "ICAgICAgIHNlbGYuX21vZF90YWJsZS5ob3Jpem9udGFsSGVhZGVyKCkuc2V0U2VjdGlvblJlc2l6ZU1vZGUoCiAgICAgICAgICAg"
    "IDEsIFFIZWFkZXJWaWV3LlJlc2l6ZU1vZGUuU3RyZXRjaCkKICAgICAgICBzZWxmLl9tb2RfdGFibGUuc2V0U3R5bGVTaGVldChf"
    "Z290aGljX3RhYmxlX3N0eWxlKCkpCiAgICAgICAgbDMuYWRkV2lkZ2V0KHNlbGYuX21vZF90YWJsZSwgMSkKCiAgICAgICAgYnRu"
    "czMgPSBRSEJveExheW91dCgpCiAgICAgICAgczMgPSBfZ290aGljX2J0bigi4pymIFNhdmUiKTsgYzMgPSBfZ290aGljX2J0bigi"
    "4pyXIENhbmNlbCIpCiAgICAgICAgczMuY2xpY2tlZC5jb25uZWN0KHNlbGYuX2RvX21vZGlmeV9zYXZlKQogICAgICAgIGMzLmNs"
    "aWNrZWQuY29ubmVjdChsYW1iZGE6IHNlbGYuX3N0YWNrLnNldEN1cnJlbnRJbmRleCgwKSkKICAgICAgICBidG5zMy5hZGRXaWRn"
    "ZXQoczMpOyBidG5zMy5hZGRXaWRnZXQoYzMpOyBidG5zMy5hZGRTdHJldGNoKCkKICAgICAgICBsMy5hZGRMYXlvdXQoYnRuczMp"
    "CiAgICAgICAgc2VsZi5fc3RhY2suYWRkV2lkZ2V0KHAzKQoKICAgICMg4pSA4pSAIFBBUlNFUiDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAK"
    "ICAgIEBzdGF0aWNtZXRob2QKICAgIGRlZiBwYXJzZV9zY2FuX3RleHQocmF3OiBzdHIpIC0+IHR1cGxlW3N0ciwgbGlzdFtkaWN0"
    "XV06CiAgICAgICAgIiIiCiAgICAgICAgUGFyc2UgcmF3IFNMIHNjYW4gb3V0cHV0IGludG8gKGF2YXRhcl9uYW1lLCBpdGVtcyku"
    "CgogICAgICAgIEtFWSBGSVg6IEJlZm9yZSBzcGxpdHRpbmcsIGluc2VydCBuZXdsaW5lcyBiZWZvcmUgZXZlcnkgW0hIOk1NXQog"
    "ICAgICAgIHRpbWVzdGFtcCBzbyBzaW5nbGUtbGluZSBwYXN0ZXMgd29yayBjb3JyZWN0bHkuCgogICAgICAgIEV4cGVjdGVkIGZv"
    "cm1hdDoKICAgICAgICAgICAgWzExOjQ3XSBBdmF0YXJOYW1lJ3MgcHVibGljIGF0dGFjaG1lbnRzOgogICAgICAgICAgICBbMTE6"
    "NDddIC46IEl0ZW0gTmFtZSBbQXR0YWNobWVudF0gQ1JFQVRPUjogQ3JlYXRvck5hbWUgWzExOjQ3XSAuLi4KICAgICAgICAiIiIK"
    "ICAgICAgICBpZiBub3QgcmF3LnN0cmlwKCk6CiAgICAgICAgICAgIHJldHVybiAiVU5LTk9XTiIsIFtdCgogICAgICAgICMg4pSA"
    "4pSAIFN0ZXAgMTogbm9ybWFsaXplIOKAlCBpbnNlcnQgbmV3bGluZXMgYmVmb3JlIHRpbWVzdGFtcHMg4pSA4pSA4pSA4pSA4pSA"
    "4pSACiAgICAgICAgbm9ybWFsaXplZCA9IF9yZS5zdWIocidccyooXFtcZHsxLDJ9OlxkezJ9XF0pJywgcidcblwxJywgcmF3KQog"
    "ICAgICAgIGxpbmVzID0gW2wuc3RyaXAoKSBmb3IgbCBpbiBub3JtYWxpemVkLnNwbGl0bGluZXMoKSBpZiBsLnN0cmlwKCldCgog"
    "ICAgICAgICMg4pSA4pSAIFN0ZXAgMjogZXh0cmFjdCBhdmF0YXIgbmFtZSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBh"
    "dmF0YXJfbmFtZSA9ICJVTktOT1dOIgogICAgICAgIGZvciBsaW5lIGluIGxpbmVzOgogICAgICAgICAgICAjICJBdmF0YXJOYW1l"
    "J3MgcHVibGljIGF0dGFjaG1lbnRzIiBvciBzaW1pbGFyCiAgICAgICAgICAgIG0gPSBfcmUuc2VhcmNoKAogICAgICAgICAgICAg"
    "ICAgciIoXHdbXHdcc10rPyknc1xzK3B1YmxpY1xzK2F0dGFjaG1lbnRzIiwKICAgICAgICAgICAgICAgIGxpbmUsIF9yZS5JCiAg"
    "ICAgICAgICAgICkKICAgICAgICAgICAgaWYgbToKICAgICAgICAgICAgICAgIGF2YXRhcl9uYW1lID0gbS5ncm91cCgxKS5zdHJp"
    "cCgpCiAgICAgICAgICAgICAgICBicmVhawoKICAgICAgICAjIOKUgOKUgCBTdGVwIDM6IGV4dHJhY3QgaXRlbXMg4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgaXRlbXMgPSBbXQogICAgICAgIGZvciBsaW5lIGluIGxpbmVz"
    "OgogICAgICAgICAgICAjIFN0cmlwIGxlYWRpbmcgdGltZXN0YW1wCiAgICAgICAgICAgIGNvbnRlbnQgPSBfcmUuc3ViKHInXlxb"
    "XGR7MSwyfTpcZHsyfVxdXHMqJywgJycsIGxpbmUpLnN0cmlwKCkKICAgICAgICAgICAgaWYgbm90IGNvbnRlbnQ6CiAgICAgICAg"
    "ICAgICAgICBjb250aW51ZQogICAgICAgICAgICAjIFNraXAgaGVhZGVyIGxpbmVzCiAgICAgICAgICAgIGlmICIncyBwdWJsaWMg"
    "YXR0YWNobWVudHMiIGluIGNvbnRlbnQubG93ZXIoKToKICAgICAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgICAgIGlmIGNv"
    "bnRlbnQubG93ZXIoKS5zdGFydHN3aXRoKCJvYmplY3QiKToKICAgICAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgICAgICMg"
    "U2tpcCBkaXZpZGVyIGxpbmVzIOKAlCBsaW5lcyB0aGF0IGFyZSBtb3N0bHkgb25lIHJlcGVhdGVkIGNoYXJhY3RlcgogICAgICAg"
    "ICAgICAjIGUuZy4g4paC4paC4paC4paC4paC4paC4paC4paC4paC4paC4paC4paCIG9yIOKVkOKVkOKVkOKVkOKVkOKVkOKVkOKV"
    "kOKVkOKVkOKVkOKVkCBvciDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICAgICAgc3RyaXBwZWQg"
    "PSBjb250ZW50LnN0cmlwKCIuOiAiKQogICAgICAgICAgICBpZiBzdHJpcHBlZCBhbmQgbGVuKHNldChzdHJpcHBlZCkpIDw9IDI6"
    "CiAgICAgICAgICAgICAgICBjb250aW51ZSAgIyBvbmUgb3IgdHdvIHVuaXF1ZSBjaGFycyA9IGRpdmlkZXIgbGluZQoKICAgICAg"
    "ICAgICAgIyBUcnkgdG8gZXh0cmFjdCBDUkVBVE9SOiBmaWVsZAogICAgICAgICAgICBjcmVhdG9yID0gIlVOS05PV04iCiAgICAg"
    "ICAgICAgIGl0ZW1fbmFtZSA9IGNvbnRlbnQKCiAgICAgICAgICAgIGNyZWF0b3JfbWF0Y2ggPSBfcmUuc2VhcmNoKAogICAgICAg"
    "ICAgICAgICAgcidDUkVBVE9SOlxzKihbXHdcc10rPykoPzpccypcW3wkKScsIGNvbnRlbnQsIF9yZS5JCiAgICAgICAgICAgICkK"
    "ICAgICAgICAgICAgaWYgY3JlYXRvcl9tYXRjaDoKICAgICAgICAgICAgICAgIGNyZWF0b3IgICA9IGNyZWF0b3JfbWF0Y2guZ3Jv"
    "dXAoMSkuc3RyaXAoKQogICAgICAgICAgICAgICAgaXRlbV9uYW1lID0gY29udGVudFs6Y3JlYXRvcl9tYXRjaC5zdGFydCgpXS5z"
    "dHJpcCgpCgogICAgICAgICAgICAjIFN0cmlwIGF0dGFjaG1lbnQgcG9pbnQgc3VmZml4ZXMgbGlrZSBbTGVmdF9Gb290XQogICAg"
    "ICAgICAgICBpdGVtX25hbWUgPSBfcmUuc3ViKHInXHMqXFtbXHdcc19dK1xdJywgJycsIGl0ZW1fbmFtZSkuc3RyaXAoKQogICAg"
    "ICAgICAgICBpdGVtX25hbWUgPSBpdGVtX25hbWUuc3RyaXAoIi46ICIpCgogICAgICAgICAgICBpZiBpdGVtX25hbWUgYW5kIGxl"
    "bihpdGVtX25hbWUpID4gMToKICAgICAgICAgICAgICAgIGl0ZW1zLmFwcGVuZCh7Iml0ZW0iOiBpdGVtX25hbWUsICJjcmVhdG9y"
    "IjogY3JlYXRvcn0pCgogICAgICAgIHJldHVybiBhdmF0YXJfbmFtZSwgaXRlbXMKCiAgICAjIOKUgOKUgCBDQVJEIFJFTkRFUklO"
    "RyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAg"
    "IGRlZiBfYnVpbGRfY2FyZHMoc2VsZikgLT4gTm9uZToKICAgICAgICAjIENsZWFyIGV4aXN0aW5nIGNhcmRzIChrZWVwIHN0cmV0"
    "Y2gpCiAgICAgICAgd2hpbGUgc2VsZi5fY2FyZF9sYXlvdXQuY291bnQoKSA+IDE6CiAgICAgICAgICAgIGl0ZW0gPSBzZWxmLl9j"
    "YXJkX2xheW91dC50YWtlQXQoMCkKICAgICAgICAgICAgaWYgaXRlbS53aWRnZXQoKToKICAgICAgICAgICAgICAgIGl0ZW0ud2lk"
    "Z2V0KCkuZGVsZXRlTGF0ZXIoKQoKICAgICAgICBmb3IgcmVjIGluIHNlbGYuX3JlY29yZHM6CiAgICAgICAgICAgIGNhcmQgPSBz"
    "ZWxmLl9tYWtlX2NhcmQocmVjKQogICAgICAgICAgICBzZWxmLl9jYXJkX2xheW91dC5pbnNlcnRXaWRnZXQoCiAgICAgICAgICAg"
    "ICAgICBzZWxmLl9jYXJkX2xheW91dC5jb3VudCgpIC0gMSwgY2FyZAogICAgICAgICAgICApCgogICAgZGVmIF9tYWtlX2NhcmQo"
    "c2VsZiwgcmVjOiBkaWN0KSAtPiBRV2lkZ2V0OgogICAgICAgIGNhcmQgPSBRRnJhbWUoKQogICAgICAgIGlzX3NlbGVjdGVkID0g"
    "cmVjLmdldCgicmVjb3JkX2lkIikgPT0gc2VsZi5fc2VsZWN0ZWRfaWQKICAgICAgICBjYXJkLnNldFN0eWxlU2hlZXQoCiAgICAg"
    "ICAgICAgIGYiYmFja2dyb3VuZDogeycjMWEwYTEwJyBpZiBpc19zZWxlY3RlZCBlbHNlIENfQkczfTsgIgogICAgICAgICAgICBm"
    "ImJvcmRlcjogMXB4IHNvbGlkIHtDX0NSSU1TT04gaWYgaXNfc2VsZWN0ZWQgZWxzZSBDX0JPUkRFUn07ICIKICAgICAgICAgICAg"
    "ZiJib3JkZXItcmFkaXVzOiAycHg7IHBhZGRpbmc6IDJweDsiCiAgICAgICAgKQogICAgICAgIGxheW91dCA9IFFIQm94TGF5b3V0"
    "KGNhcmQpCiAgICAgICAgbGF5b3V0LnNldENvbnRlbnRzTWFyZ2lucyg4LCA2LCA4LCA2KQoKICAgICAgICBuYW1lX2xibCA9IFFM"
    "YWJlbChyZWMuZ2V0KCJuYW1lIiwgIlVOS05PV04iKSkKICAgICAgICBuYW1lX2xibC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAg"
    "ICBmImNvbG9yOiB7Q19HT0xEX0JSSUdIVCBpZiBpc19zZWxlY3RlZCBlbHNlIENfR09MRH07ICIKICAgICAgICAgICAgZiJmb250"
    "LXNpemU6IDExcHg7IGZvbnQtd2VpZ2h0OiBib2xkOyBmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyIKICAgICAgICAp"
    "CgogICAgICAgIGNvdW50ID0gbGVuKHJlYy5nZXQoIml0ZW1zIiwgW10pKQogICAgICAgIGNvdW50X2xibCA9IFFMYWJlbChmIntj"
    "b3VudH0gaXRlbXMiKQogICAgICAgIGNvdW50X2xibC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19DUklN"
    "U09OfTsgZm9udC1zaXplOiAxMHB4OyBmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyIKICAgICAgICApCgogICAgICAg"
    "IGRhdGVfbGJsID0gUUxhYmVsKHJlYy5nZXQoImNyZWF0ZWRfYXQiLCAiIilbOjEwXSkKICAgICAgICBkYXRlX2xibC5zZXRTdHls"
    "ZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19URVhUX0RJTX07IGZvbnQtc2l6ZTogOXB4OyBmb250LWZhbWlseToge0RF"
    "Q0tfRk9OVH0sIHNlcmlmOyIKICAgICAgICApCgogICAgICAgIGxheW91dC5hZGRXaWRnZXQobmFtZV9sYmwpCiAgICAgICAgbGF5"
    "b3V0LmFkZFN0cmV0Y2goKQogICAgICAgIGxheW91dC5hZGRXaWRnZXQoY291bnRfbGJsKQogICAgICAgIGxheW91dC5hZGRTcGFj"
    "aW5nKDEyKQogICAgICAgIGxheW91dC5hZGRXaWRnZXQoZGF0ZV9sYmwpCgogICAgICAgICMgQ2xpY2sgdG8gc2VsZWN0CiAgICAg"
    "ICAgcmVjX2lkID0gcmVjLmdldCgicmVjb3JkX2lkIiwgIiIpCiAgICAgICAgY2FyZC5tb3VzZVByZXNzRXZlbnQgPSBsYW1iZGEg"
    "ZSwgcmlkPXJlY19pZDogc2VsZi5fc2VsZWN0X2NhcmQocmlkKQogICAgICAgIHJldHVybiBjYXJkCgogICAgZGVmIF9zZWxlY3Rf"
    "Y2FyZChzZWxmLCByZWNvcmRfaWQ6IHN0cikgLT4gTm9uZToKICAgICAgICBzZWxmLl9zZWxlY3RlZF9pZCA9IHJlY29yZF9pZAog"
    "ICAgICAgIHNlbGYuX2J1aWxkX2NhcmRzKCkgICMgUmVidWlsZCB0byBzaG93IHNlbGVjdGlvbiBoaWdobGlnaHQKCiAgICBkZWYg"
    "X3NlbGVjdGVkX3JlY29yZChzZWxmKSAtPiBPcHRpb25hbFtkaWN0XToKICAgICAgICByZXR1cm4gbmV4dCgKICAgICAgICAgICAg"
    "KHIgZm9yIHIgaW4gc2VsZi5fcmVjb3JkcwogICAgICAgICAgICAgaWYgci5nZXQoInJlY29yZF9pZCIpID09IHNlbGYuX3NlbGVj"
    "dGVkX2lkKSwKICAgICAgICAgICAgTm9uZQogICAgICAgICkKCiAgICAjIOKUgOKUgCBBQ1RJT05TIOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAog"
    "ICAgZGVmIHJlZnJlc2goc2VsZikgLT4gTm9uZToKICAgICAgICBzZWxmLl9yZWNvcmRzID0gcmVhZF9qc29ubChzZWxmLl9wYXRo"
    "KQogICAgICAgICMgRW5zdXJlIHJlY29yZF9pZCBmaWVsZCBleGlzdHMKICAgICAgICBjaGFuZ2VkID0gRmFsc2UKICAgICAgICBm"
    "b3IgciBpbiBzZWxmLl9yZWNvcmRzOgogICAgICAgICAgICBpZiBub3Qgci5nZXQoInJlY29yZF9pZCIpOgogICAgICAgICAgICAg"
    "ICAgclsicmVjb3JkX2lkIl0gPSByLmdldCgiaWQiKSBvciBzdHIodXVpZC51dWlkNCgpKQogICAgICAgICAgICAgICAgY2hhbmdl"
    "ZCA9IFRydWUKICAgICAgICBpZiBjaGFuZ2VkOgogICAgICAgICAgICB3cml0ZV9qc29ubChzZWxmLl9wYXRoLCBzZWxmLl9yZWNv"
    "cmRzKQogICAgICAgIHNlbGYuX2J1aWxkX2NhcmRzKCkKICAgICAgICBzZWxmLl9zdGFjay5zZXRDdXJyZW50SW5kZXgoMCkKCiAg"
    "ICBkZWYgX3ByZXZpZXdfcGFyc2Uoc2VsZikgLT4gTm9uZToKICAgICAgICByYXcgPSBzZWxmLl9hZGRfcmF3LnRvUGxhaW5UZXh0"
    "KCkKICAgICAgICBuYW1lLCBpdGVtcyA9IHNlbGYucGFyc2Vfc2Nhbl90ZXh0KHJhdykKICAgICAgICBzZWxmLl9hZGRfbmFtZS5z"
    "ZXRQbGFjZWhvbGRlclRleHQobmFtZSkKICAgICAgICBzZWxmLl9hZGRfcHJldmlldy5zZXRSb3dDb3VudCgwKQogICAgICAgIGZv"
    "ciBpdCBpbiBpdGVtc1s6MjBdOiAgIyBwcmV2aWV3IGZpcnN0IDIwCiAgICAgICAgICAgIHIgPSBzZWxmLl9hZGRfcHJldmlldy5y"
    "b3dDb3VudCgpCiAgICAgICAgICAgIHNlbGYuX2FkZF9wcmV2aWV3Lmluc2VydFJvdyhyKQogICAgICAgICAgICBzZWxmLl9hZGRf"
    "cHJldmlldy5zZXRJdGVtKHIsIDAsIFFUYWJsZVdpZGdldEl0ZW0oaXRbIml0ZW0iXSkpCiAgICAgICAgICAgIHNlbGYuX2FkZF9w"
    "cmV2aWV3LnNldEl0ZW0ociwgMSwgUVRhYmxlV2lkZ2V0SXRlbShpdFsiY3JlYXRvciJdKSkKCiAgICBkZWYgX3Nob3dfYWRkKHNl"
    "bGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fYWRkX25hbWUuY2xlYXIoKQogICAgICAgIHNlbGYuX2FkZF9uYW1lLnNldFBsYWNl"
    "aG9sZGVyVGV4dCgiQXV0by1kZXRlY3RlZCBmcm9tIHNjYW4gdGV4dCIpCiAgICAgICAgc2VsZi5fYWRkX2Rlc2MuY2xlYXIoKQog"
    "ICAgICAgIHNlbGYuX2FkZF9yYXcuY2xlYXIoKQogICAgICAgIHNlbGYuX2FkZF9wcmV2aWV3LnNldFJvd0NvdW50KDApCiAgICAg"
    "ICAgc2VsZi5fc3RhY2suc2V0Q3VycmVudEluZGV4KDEpCgogICAgZGVmIF9kb19hZGQoc2VsZikgLT4gTm9uZToKICAgICAgICBy"
    "YXcgID0gc2VsZi5fYWRkX3Jhdy50b1BsYWluVGV4dCgpCiAgICAgICAgbmFtZSwgaXRlbXMgPSBzZWxmLnBhcnNlX3NjYW5fdGV4"
    "dChyYXcpCiAgICAgICAgb3ZlcnJpZGVfbmFtZSA9IHNlbGYuX2FkZF9uYW1lLnRleHQoKS5zdHJpcCgpCiAgICAgICAgbm93ICA9"
    "IGRhdGV0aW1lLm5vdyh0aW1lem9uZS51dGMpLmlzb2Zvcm1hdCgpCiAgICAgICAgcmVjb3JkID0gewogICAgICAgICAgICAiaWQi"
    "OiAgICAgICAgICBzdHIodXVpZC51dWlkNCgpKSwKICAgICAgICAgICAgInJlY29yZF9pZCI6ICAgc3RyKHV1aWQudXVpZDQoKSks"
    "CiAgICAgICAgICAgICJuYW1lIjogICAgICAgIG92ZXJyaWRlX25hbWUgb3IgbmFtZSwKICAgICAgICAgICAgImRlc2NyaXB0aW9u"
    "Ijogc2VsZi5fYWRkX2Rlc2MudG9QbGFpblRleHQoKVs6MjQ0XSwKICAgICAgICAgICAgIml0ZW1zIjogICAgICAgaXRlbXMsCiAg"
    "ICAgICAgICAgICJyYXdfdGV4dCI6ICAgIHJhdywKICAgICAgICAgICAgImNyZWF0ZWRfYXQiOiAgbm93LAogICAgICAgICAgICAi"
    "dXBkYXRlZF9hdCI6ICBub3csCiAgICAgICAgfQogICAgICAgIHNlbGYuX3JlY29yZHMuYXBwZW5kKHJlY29yZCkKICAgICAgICB3"
    "cml0ZV9qc29ubChzZWxmLl9wYXRoLCBzZWxmLl9yZWNvcmRzKQogICAgICAgIHNlbGYuX3NlbGVjdGVkX2lkID0gcmVjb3JkWyJy"
    "ZWNvcmRfaWQiXQogICAgICAgIHNlbGYucmVmcmVzaCgpCgogICAgZGVmIF9zaG93X2Rpc3BsYXkoc2VsZikgLT4gTm9uZToKICAg"
    "ICAgICByZWMgPSBzZWxmLl9zZWxlY3RlZF9yZWNvcmQoKQogICAgICAgIGlmIG5vdCByZWM6CiAgICAgICAgICAgIFFNZXNzYWdl"
    "Qm94LmluZm9ybWF0aW9uKHNlbGYsICJTTCBTY2FucyIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICJTZWxl"
    "Y3QgYSBzY2FuIHRvIGRpc3BsYXkuIikKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgc2VsZi5fZGlzcF9uYW1lLnNldFRleHQo"
    "ZiLinacge3JlYy5nZXQoJ25hbWUnLCcnKX0iKQogICAgICAgIHNlbGYuX2Rpc3BfZGVzYy5zZXRUZXh0KHJlYy5nZXQoImRlc2Ny"
    "aXB0aW9uIiwiIikpCiAgICAgICAgc2VsZi5fZGlzcF90YWJsZS5zZXRSb3dDb3VudCgwKQogICAgICAgIGZvciBpdCBpbiByZWMu"
    "Z2V0KCJpdGVtcyIsW10pOgogICAgICAgICAgICByID0gc2VsZi5fZGlzcF90YWJsZS5yb3dDb3VudCgpCiAgICAgICAgICAgIHNl"
    "bGYuX2Rpc3BfdGFibGUuaW5zZXJ0Um93KHIpCiAgICAgICAgICAgIHNlbGYuX2Rpc3BfdGFibGUuc2V0SXRlbShyLCAwLAogICAg"
    "ICAgICAgICAgICAgUVRhYmxlV2lkZ2V0SXRlbShpdC5nZXQoIml0ZW0iLCIiKSkpCiAgICAgICAgICAgIHNlbGYuX2Rpc3BfdGFi"
    "bGUuc2V0SXRlbShyLCAxLAogICAgICAgICAgICAgICAgUVRhYmxlV2lkZ2V0SXRlbShpdC5nZXQoImNyZWF0b3IiLCJVTktOT1dO"
    "IikpKQogICAgICAgIHNlbGYuX3N0YWNrLnNldEN1cnJlbnRJbmRleCgyKQoKICAgIGRlZiBfaXRlbV9jb250ZXh0X21lbnUoc2Vs"
    "ZiwgcG9zKSAtPiBOb25lOgogICAgICAgIGlkeCA9IHNlbGYuX2Rpc3BfdGFibGUuaW5kZXhBdChwb3MpCiAgICAgICAgaWYgbm90"
    "IGlkeC5pc1ZhbGlkKCk6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIGl0ZW1fdGV4dCAgPSAoc2VsZi5fZGlzcF90YWJsZS5p"
    "dGVtKGlkeC5yb3coKSwgMCkgb3IKICAgICAgICAgICAgICAgICAgICAgIFFUYWJsZVdpZGdldEl0ZW0oIiIpKS50ZXh0KCkKICAg"
    "ICAgICBjcmVhdG9yICAgID0gKHNlbGYuX2Rpc3BfdGFibGUuaXRlbShpZHgucm93KCksIDEpIG9yCiAgICAgICAgICAgICAgICAg"
    "ICAgICBRVGFibGVXaWRnZXRJdGVtKCIiKSkudGV4dCgpCiAgICAgICAgZnJvbSBQeVNpZGU2LlF0V2lkZ2V0cyBpbXBvcnQgUU1l"
    "bnUKICAgICAgICBtZW51ID0gUU1lbnUoc2VsZikKICAgICAgICBtZW51LnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiYmFj"
    "a2dyb3VuZDoge0NfQkczfTsgY29sb3I6IHtDX0dPTER9OyAiCiAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQ1JJ"
    "TVNPTl9ESU19OyIKICAgICAgICApCiAgICAgICAgYV9pdGVtICAgID0gbWVudS5hZGRBY3Rpb24oIkNvcHkgSXRlbSBOYW1lIikK"
    "ICAgICAgICBhX2NyZWF0b3IgPSBtZW51LmFkZEFjdGlvbigiQ29weSBDcmVhdG9yIikKICAgICAgICBhX2JvdGggICAgPSBtZW51"
    "LmFkZEFjdGlvbigiQ29weSBCb3RoIikKICAgICAgICBhY3Rpb24gPSBtZW51LmV4ZWMoc2VsZi5fZGlzcF90YWJsZS52aWV3cG9y"
    "dCgpLm1hcFRvR2xvYmFsKHBvcykpCiAgICAgICAgY2IgPSBRQXBwbGljYXRpb24uY2xpcGJvYXJkKCkKICAgICAgICBpZiBhY3Rp"
    "b24gPT0gYV9pdGVtOiAgICBjYi5zZXRUZXh0KGl0ZW1fdGV4dCkKICAgICAgICBlbGlmIGFjdGlvbiA9PSBhX2NyZWF0b3I6IGNi"
    "LnNldFRleHQoY3JlYXRvcikKICAgICAgICBlbGlmIGFjdGlvbiA9PSBhX2JvdGg6ICBjYi5zZXRUZXh0KGYie2l0ZW1fdGV4dH0g"
    "4oCUIHtjcmVhdG9yfSIpCgogICAgZGVmIF9zaG93X21vZGlmeShzZWxmKSAtPiBOb25lOgogICAgICAgIHJlYyA9IHNlbGYuX3Nl"
    "bGVjdGVkX3JlY29yZCgpCiAgICAgICAgaWYgbm90IHJlYzoKICAgICAgICAgICAgUU1lc3NhZ2VCb3guaW5mb3JtYXRpb24oc2Vs"
    "ZiwgIlNMIFNjYW5zIiwKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIlNlbGVjdCBhIHNjYW4gdG8gbW9kaWZ5"
    "LiIpCiAgICAgICAgICAgIHJldHVybgogICAgICAgIHNlbGYuX21vZF9uYW1lLnNldFRleHQocmVjLmdldCgibmFtZSIsIiIpKQog"
    "ICAgICAgIHNlbGYuX21vZF9kZXNjLnNldFRleHQocmVjLmdldCgiZGVzY3JpcHRpb24iLCIiKSkKICAgICAgICBzZWxmLl9tb2Rf"
    "dGFibGUuc2V0Um93Q291bnQoMCkKICAgICAgICBmb3IgaXQgaW4gcmVjLmdldCgiaXRlbXMiLFtdKToKICAgICAgICAgICAgciA9"
    "IHNlbGYuX21vZF90YWJsZS5yb3dDb3VudCgpCiAgICAgICAgICAgIHNlbGYuX21vZF90YWJsZS5pbnNlcnRSb3cocikKICAgICAg"
    "ICAgICAgc2VsZi5fbW9kX3RhYmxlLnNldEl0ZW0ociwgMCwKICAgICAgICAgICAgICAgIFFUYWJsZVdpZGdldEl0ZW0oaXQuZ2V0"
    "KCJpdGVtIiwiIikpKQogICAgICAgICAgICBzZWxmLl9tb2RfdGFibGUuc2V0SXRlbShyLCAxLAogICAgICAgICAgICAgICAgUVRh"
    "YmxlV2lkZ2V0SXRlbShpdC5nZXQoImNyZWF0b3IiLCJVTktOT1dOIikpKQogICAgICAgIHNlbGYuX3N0YWNrLnNldEN1cnJlbnRJ"
    "bmRleCgzKQoKICAgIGRlZiBfZG9fbW9kaWZ5X3NhdmUoc2VsZikgLT4gTm9uZToKICAgICAgICByZWMgPSBzZWxmLl9zZWxlY3Rl"
    "ZF9yZWNvcmQoKQogICAgICAgIGlmIG5vdCByZWM6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIHJlY1sibmFtZSJdICAgICAg"
    "ICA9IHNlbGYuX21vZF9uYW1lLnRleHQoKS5zdHJpcCgpIG9yICJVTktOT1dOIgogICAgICAgIHJlY1siZGVzY3JpcHRpb24iXSA9"
    "IHNlbGYuX21vZF9kZXNjLnRleHQoKVs6MjQ0XQogICAgICAgIGl0ZW1zID0gW10KICAgICAgICBmb3IgaSBpbiByYW5nZShzZWxm"
    "Ll9tb2RfdGFibGUucm93Q291bnQoKSk6CiAgICAgICAgICAgIGl0ICA9IChzZWxmLl9tb2RfdGFibGUuaXRlbShpLDApIG9yIFFU"
    "YWJsZVdpZGdldEl0ZW0oIiIpKS50ZXh0KCkKICAgICAgICAgICAgY3IgID0gKHNlbGYuX21vZF90YWJsZS5pdGVtKGksMSkgb3Ig"
    "UVRhYmxlV2lkZ2V0SXRlbSgiIikpLnRleHQoKQogICAgICAgICAgICBpdGVtcy5hcHBlbmQoeyJpdGVtIjogaXQuc3RyaXAoKSBv"
    "ciAiVU5LTk9XTiIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgImNyZWF0b3IiOiBjci5zdHJpcCgpIG9yICJVTktOT1dOIn0p"
    "CiAgICAgICAgcmVjWyJpdGVtcyJdICAgICAgPSBpdGVtcwogICAgICAgIHJlY1sidXBkYXRlZF9hdCJdID0gZGF0ZXRpbWUubm93"
    "KHRpbWV6b25lLnV0YykuaXNvZm9ybWF0KCkKICAgICAgICB3cml0ZV9qc29ubChzZWxmLl9wYXRoLCBzZWxmLl9yZWNvcmRzKQog"
    "ICAgICAgIHNlbGYucmVmcmVzaCgpCgogICAgZGVmIF9kb19kZWxldGUoc2VsZikgLT4gTm9uZToKICAgICAgICByZWMgPSBzZWxm"
    "Ll9zZWxlY3RlZF9yZWNvcmQoKQogICAgICAgIGlmIG5vdCByZWM6CiAgICAgICAgICAgIFFNZXNzYWdlQm94LmluZm9ybWF0aW9u"
    "KHNlbGYsICJTTCBTY2FucyIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICJTZWxlY3QgYSBzY2FuIHRvIGRl"
    "bGV0ZS4iKQogICAgICAgICAgICByZXR1cm4KICAgICAgICBuYW1lID0gcmVjLmdldCgibmFtZSIsInRoaXMgc2NhbiIpCiAgICAg"
    "ICAgcmVwbHkgPSBRTWVzc2FnZUJveC5xdWVzdGlvbigKICAgICAgICAgICAgc2VsZiwgIkRlbGV0ZSBTY2FuIiwKICAgICAgICAg"
    "ICAgZiJEZWxldGUgJ3tuYW1lfSc/IFRoaXMgY2Fubm90IGJlIHVuZG9uZS4iLAogICAgICAgICAgICBRTWVzc2FnZUJveC5TdGFu"
    "ZGFyZEJ1dHRvbi5ZZXMgfCBRTWVzc2FnZUJveC5TdGFuZGFyZEJ1dHRvbi5ObwogICAgICAgICkKICAgICAgICBpZiByZXBseSA9"
    "PSBRTWVzc2FnZUJveC5TdGFuZGFyZEJ1dHRvbi5ZZXM6CiAgICAgICAgICAgIHNlbGYuX3JlY29yZHMgPSBbciBmb3IgciBpbiBz"
    "ZWxmLl9yZWNvcmRzCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgaWYgci5nZXQoInJlY29yZF9pZCIpICE9IHNlbGYuX3Nl"
    "bGVjdGVkX2lkXQogICAgICAgICAgICB3cml0ZV9qc29ubChzZWxmLl9wYXRoLCBzZWxmLl9yZWNvcmRzKQogICAgICAgICAgICBz"
    "ZWxmLl9zZWxlY3RlZF9pZCA9IE5vbmUKICAgICAgICAgICAgc2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX2RvX3JlcGFyc2Uoc2Vs"
    "ZikgLT4gTm9uZToKICAgICAgICByZWMgPSBzZWxmLl9zZWxlY3RlZF9yZWNvcmQoKQogICAgICAgIGlmIG5vdCByZWM6CiAgICAg"
    "ICAgICAgIFFNZXNzYWdlQm94LmluZm9ybWF0aW9uKHNlbGYsICJTTCBTY2FucyIsCiAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgICAgICJTZWxlY3QgYSBzY2FuIHRvIHJlLXBhcnNlLiIpCiAgICAgICAgICAgIHJldHVybgogICAgICAgIHJhdyA9IHJl"
    "Yy5nZXQoInJhd190ZXh0IiwiIikKICAgICAgICBpZiBub3QgcmF3OgogICAgICAgICAgICBRTWVzc2FnZUJveC5pbmZvcm1hdGlv"
    "bihzZWxmLCAiUmUtcGFyc2UiLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAiTm8gcmF3IHRleHQgc3RvcmVk"
    "IGZvciB0aGlzIHNjYW4uIikKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgbmFtZSwgaXRlbXMgPSBzZWxmLnBhcnNlX3NjYW5f"
    "dGV4dChyYXcpCiAgICAgICAgcmVjWyJpdGVtcyJdICAgICAgPSBpdGVtcwogICAgICAgIHJlY1sibmFtZSJdICAgICAgID0gcmVj"
    "WyJuYW1lIl0gb3IgbmFtZQogICAgICAgIHJlY1sidXBkYXRlZF9hdCJdID0gZGF0ZXRpbWUubm93KHRpbWV6b25lLnV0YykuaXNv"
    "Zm9ybWF0KCkKICAgICAgICB3cml0ZV9qc29ubChzZWxmLl9wYXRoLCBzZWxmLl9yZWNvcmRzKQogICAgICAgIHNlbGYucmVmcmVz"
    "aCgpCiAgICAgICAgUU1lc3NhZ2VCb3guaW5mb3JtYXRpb24oc2VsZiwgIlJlLXBhcnNlZCIsCiAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgICAgICAgZiJGb3VuZCB7bGVuKGl0ZW1zKX0gaXRlbXMuIikKCgojIOKUgOKUgCBTTCBDT01NQU5EUyBUQUIg4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNsYXNzIFNMQ29tbWFuZHNUYWIoUVdpZGdldCk6CiAgICAiIiIKICAgIFNlY29uZCBMaWZl"
    "IGNvbW1hbmQgcmVmZXJlbmNlIHRhYmxlLgogICAgR290aGljIHRhYmxlIHN0eWxpbmcuIENvcHkgY29tbWFuZCB0byBjbGlwYm9h"
    "cmQgYnV0dG9uIHBlciByb3cuCiAgICAiIiIKCiAgICBkZWYgX19pbml0X18oc2VsZiwgcGFyZW50PU5vbmUpOgogICAgICAgIHN1"
    "cGVyKCkuX19pbml0X18ocGFyZW50KQogICAgICAgIHNlbGYuX3BhdGggICAgPSBjZmdfcGF0aCgic2wiKSAvICJzbF9jb21tYW5k"
    "cy5qc29ubCIKICAgICAgICBzZWxmLl9yZWNvcmRzOiBsaXN0W2RpY3RdID0gW10KICAgICAgICBzZWxmLl9zZXR1cF91aSgpCiAg"
    "ICAgICAgc2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX3NldHVwX3VpKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgcm9vdCA9IFFWQm94"
    "TGF5b3V0KHNlbGYpCiAgICAgICAgcm9vdC5zZXRDb250ZW50c01hcmdpbnMoNCwgNCwgNCwgNCkKICAgICAgICByb290LnNldFNw"
    "YWNpbmcoNCkKCiAgICAgICAgYmFyID0gUUhCb3hMYXlvdXQoKQogICAgICAgIHNlbGYuX2J0bl9hZGQgICAgPSBfZ290aGljX2J0"
    "bigi4pymIEFkZCIpCiAgICAgICAgc2VsZi5fYnRuX21vZGlmeSA9IF9nb3RoaWNfYnRuKCLinKcgTW9kaWZ5IikKICAgICAgICBz"
    "ZWxmLl9idG5fZGVsZXRlID0gX2dvdGhpY19idG4oIuKclyBEZWxldGUiKQogICAgICAgIHNlbGYuX2J0bl9jb3B5ICAgPSBfZ290"
    "aGljX2J0bigi4qeJIENvcHkgQ29tbWFuZCIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAiQ29weSBz"
    "ZWxlY3RlZCBjb21tYW5kIHRvIGNsaXBib2FyZCIpCiAgICAgICAgc2VsZi5fYnRuX3JlZnJlc2g9IF9nb3RoaWNfYnRuKCLihrsg"
    "UmVmcmVzaCIpCiAgICAgICAgc2VsZi5fYnRuX2FkZC5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fZG9fYWRkKQogICAgICAgIHNlbGYu"
    "X2J0bl9tb2RpZnkuY2xpY2tlZC5jb25uZWN0KHNlbGYuX2RvX21vZGlmeSkKICAgICAgICBzZWxmLl9idG5fZGVsZXRlLmNsaWNr"
    "ZWQuY29ubmVjdChzZWxmLl9kb19kZWxldGUpCiAgICAgICAgc2VsZi5fYnRuX2NvcHkuY2xpY2tlZC5jb25uZWN0KHNlbGYuX2Nv"
    "cHlfY29tbWFuZCkKICAgICAgICBzZWxmLl9idG5fcmVmcmVzaC5jbGlja2VkLmNvbm5lY3Qoc2VsZi5yZWZyZXNoKQogICAgICAg"
    "IGZvciBiIGluIChzZWxmLl9idG5fYWRkLCBzZWxmLl9idG5fbW9kaWZ5LCBzZWxmLl9idG5fZGVsZXRlLAogICAgICAgICAgICAg"
    "ICAgICBzZWxmLl9idG5fY29weSwgc2VsZi5fYnRuX3JlZnJlc2gpOgogICAgICAgICAgICBiYXIuYWRkV2lkZ2V0KGIpCiAgICAg"
    "ICAgYmFyLmFkZFN0cmV0Y2goKQogICAgICAgIHJvb3QuYWRkTGF5b3V0KGJhcikKCiAgICAgICAgc2VsZi5fdGFibGUgPSBRVGFi"
    "bGVXaWRnZXQoMCwgMikKICAgICAgICBzZWxmLl90YWJsZS5zZXRIb3Jpem9udGFsSGVhZGVyTGFiZWxzKFsiQ29tbWFuZCIsICJE"
    "ZXNjcmlwdGlvbiJdKQogICAgICAgIHNlbGYuX3RhYmxlLmhvcml6b250YWxIZWFkZXIoKS5zZXRTZWN0aW9uUmVzaXplTW9kZSgK"
    "ICAgICAgICAgICAgMCwgUUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJldGNoKQogICAgICAgIHNlbGYuX3RhYmxlLmhvcml6b250"
    "YWxIZWFkZXIoKS5zZXRTZWN0aW9uUmVzaXplTW9kZSgKICAgICAgICAgICAgMSwgUUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJl"
    "dGNoKQogICAgICAgIHNlbGYuX3RhYmxlLnNldFNlbGVjdGlvbkJlaGF2aW9yKAogICAgICAgICAgICBRQWJzdHJhY3RJdGVtVmll"
    "dy5TZWxlY3Rpb25CZWhhdmlvci5TZWxlY3RSb3dzKQogICAgICAgIHNlbGYuX3RhYmxlLnNldEFsdGVybmF0aW5nUm93Q29sb3Jz"
    "KFRydWUpCiAgICAgICAgc2VsZi5fdGFibGUuc2V0U3R5bGVTaGVldChfZ290aGljX3RhYmxlX3N0eWxlKCkpCiAgICAgICAgcm9v"
    "dC5hZGRXaWRnZXQoc2VsZi5fdGFibGUsIDEpCgogICAgICAgIGhpbnQgPSBRTGFiZWwoCiAgICAgICAgICAgICJTZWxlY3QgYSBy"
    "b3cgYW5kIGNsaWNrIOKniSBDb3B5IENvbW1hbmQgdG8gY29weSBqdXN0IHRoZSBjb21tYW5kIHRleHQuIgogICAgICAgICkKICAg"
    "ICAgICBoaW50LnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX1RFWFRfRElNfTsgZm9udC1zaXplOiA5cHg7"
    "IGZvbnQtZmFtaWx5OiB7REVDS19GT05UfSwgc2VyaWY7IgogICAgICAgICkKICAgICAgICByb290LmFkZFdpZGdldChoaW50KQoK"
    "ICAgIGRlZiByZWZyZXNoKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fcmVjb3JkcyA9IHJlYWRfanNvbmwoc2VsZi5fcGF0"
    "aCkKICAgICAgICBzZWxmLl90YWJsZS5zZXRSb3dDb3VudCgwKQogICAgICAgIGZvciByZWMgaW4gc2VsZi5fcmVjb3JkczoKICAg"
    "ICAgICAgICAgciA9IHNlbGYuX3RhYmxlLnJvd0NvdW50KCkKICAgICAgICAgICAgc2VsZi5fdGFibGUuaW5zZXJ0Um93KHIpCiAg"
    "ICAgICAgICAgIHNlbGYuX3RhYmxlLnNldEl0ZW0ociwgMCwKICAgICAgICAgICAgICAgIFFUYWJsZVdpZGdldEl0ZW0ocmVjLmdl"
    "dCgiY29tbWFuZCIsIiIpKSkKICAgICAgICAgICAgc2VsZi5fdGFibGUuc2V0SXRlbShyLCAxLAogICAgICAgICAgICAgICAgUVRh"
    "YmxlV2lkZ2V0SXRlbShyZWMuZ2V0KCJkZXNjcmlwdGlvbiIsIiIpKSkKCiAgICBkZWYgX2NvcHlfY29tbWFuZChzZWxmKSAtPiBO"
    "b25lOgogICAgICAgIHJvdyA9IHNlbGYuX3RhYmxlLmN1cnJlbnRSb3coKQogICAgICAgIGlmIHJvdyA8IDA6CiAgICAgICAgICAg"
    "IHJldHVybgogICAgICAgIGl0ZW0gPSBzZWxmLl90YWJsZS5pdGVtKHJvdywgMCkKICAgICAgICBpZiBpdGVtOgogICAgICAgICAg"
    "ICBRQXBwbGljYXRpb24uY2xpcGJvYXJkKCkuc2V0VGV4dChpdGVtLnRleHQoKSkKCiAgICBkZWYgX2RvX2FkZChzZWxmKSAtPiBO"
    "b25lOgogICAgICAgIGRsZyA9IFFEaWFsb2coc2VsZikKICAgICAgICBkbGcuc2V0V2luZG93VGl0bGUoIkFkZCBDb21tYW5kIikK"
    "ICAgICAgICBkbGcuc2V0U3R5bGVTaGVldChmImJhY2tncm91bmQ6IHtDX0JHMn07IGNvbG9yOiB7Q19HT0xEfTsiKQogICAgICAg"
    "IGZvcm0gPSBRRm9ybUxheW91dChkbGcpCiAgICAgICAgY21kICA9IFFMaW5lRWRpdCgpOyBkZXNjID0gUUxpbmVFZGl0KCkKICAg"
    "ICAgICBmb3JtLmFkZFJvdygiQ29tbWFuZDoiLCBjbWQpCiAgICAgICAgZm9ybS5hZGRSb3coIkRlc2NyaXB0aW9uOiIsIGRlc2Mp"
    "CiAgICAgICAgYnRucyA9IFFIQm94TGF5b3V0KCkKICAgICAgICBvayA9IF9nb3RoaWNfYnRuKCJTYXZlIik7IGN4ID0gX2dvdGhp"
    "Y19idG4oIkNhbmNlbCIpCiAgICAgICAgb2suY2xpY2tlZC5jb25uZWN0KGRsZy5hY2NlcHQpOyBjeC5jbGlja2VkLmNvbm5lY3Qo"
    "ZGxnLnJlamVjdCkKICAgICAgICBidG5zLmFkZFdpZGdldChvayk7IGJ0bnMuYWRkV2lkZ2V0KGN4KQogICAgICAgIGZvcm0uYWRk"
    "Um93KGJ0bnMpCiAgICAgICAgaWYgZGxnLmV4ZWMoKSA9PSBRRGlhbG9nLkRpYWxvZ0NvZGUuQWNjZXB0ZWQ6CiAgICAgICAgICAg"
    "IG5vdyA9IGRhdGV0aW1lLm5vdyh0aW1lem9uZS51dGMpLmlzb2Zvcm1hdCgpCiAgICAgICAgICAgIHJlYyA9IHsKICAgICAgICAg"
    "ICAgICAgICJpZCI6ICAgICAgICAgIHN0cih1dWlkLnV1aWQ0KCkpLAogICAgICAgICAgICAgICAgImNvbW1hbmQiOiAgICAgY21k"
    "LnRleHQoKS5zdHJpcCgpWzoyNDRdLAogICAgICAgICAgICAgICAgImRlc2NyaXB0aW9uIjogZGVzYy50ZXh0KCkuc3RyaXAoKVs6"
    "MjQ0XSwKICAgICAgICAgICAgICAgICJjcmVhdGVkX2F0IjogIG5vdywgInVwZGF0ZWRfYXQiOiBub3csCiAgICAgICAgICAgIH0K"
    "ICAgICAgICAgICAgaWYgcmVjWyJjb21tYW5kIl06CiAgICAgICAgICAgICAgICBzZWxmLl9yZWNvcmRzLmFwcGVuZChyZWMpCiAg"
    "ICAgICAgICAgICAgICB3cml0ZV9qc29ubChzZWxmLl9wYXRoLCBzZWxmLl9yZWNvcmRzKQogICAgICAgICAgICAgICAgc2VsZi5y"
    "ZWZyZXNoKCkKCiAgICBkZWYgX2RvX21vZGlmeShzZWxmKSAtPiBOb25lOgogICAgICAgIHJvdyA9IHNlbGYuX3RhYmxlLmN1cnJl"
    "bnRSb3coKQogICAgICAgIGlmIHJvdyA8IDAgb3Igcm93ID49IGxlbihzZWxmLl9yZWNvcmRzKToKICAgICAgICAgICAgcmV0dXJu"
    "CiAgICAgICAgcmVjID0gc2VsZi5fcmVjb3Jkc1tyb3ddCiAgICAgICAgZGxnID0gUURpYWxvZyhzZWxmKQogICAgICAgIGRsZy5z"
    "ZXRXaW5kb3dUaXRsZSgiTW9kaWZ5IENvbW1hbmQiKQogICAgICAgIGRsZy5zZXRTdHlsZVNoZWV0KGYiYmFja2dyb3VuZDoge0Nf"
    "QkcyfTsgY29sb3I6IHtDX0dPTER9OyIpCiAgICAgICAgZm9ybSA9IFFGb3JtTGF5b3V0KGRsZykKICAgICAgICBjbWQgID0gUUxp"
    "bmVFZGl0KHJlYy5nZXQoImNvbW1hbmQiLCIiKSkKICAgICAgICBkZXNjID0gUUxpbmVFZGl0KHJlYy5nZXQoImRlc2NyaXB0aW9u"
    "IiwiIikpCiAgICAgICAgZm9ybS5hZGRSb3coIkNvbW1hbmQ6IiwgY21kKQogICAgICAgIGZvcm0uYWRkUm93KCJEZXNjcmlwdGlv"
    "bjoiLCBkZXNjKQogICAgICAgIGJ0bnMgPSBRSEJveExheW91dCgpCiAgICAgICAgb2sgPSBfZ290aGljX2J0bigiU2F2ZSIpOyBj"
    "eCA9IF9nb3RoaWNfYnRuKCJDYW5jZWwiKQogICAgICAgIG9rLmNsaWNrZWQuY29ubmVjdChkbGcuYWNjZXB0KTsgY3guY2xpY2tl"
    "ZC5jb25uZWN0KGRsZy5yZWplY3QpCiAgICAgICAgYnRucy5hZGRXaWRnZXQob2spOyBidG5zLmFkZFdpZGdldChjeCkKICAgICAg"
    "ICBmb3JtLmFkZFJvdyhidG5zKQogICAgICAgIGlmIGRsZy5leGVjKCkgPT0gUURpYWxvZy5EaWFsb2dDb2RlLkFjY2VwdGVkOgog"
    "ICAgICAgICAgICByZWNbImNvbW1hbmQiXSAgICAgPSBjbWQudGV4dCgpLnN0cmlwKClbOjI0NF0KICAgICAgICAgICAgcmVjWyJk"
    "ZXNjcmlwdGlvbiJdID0gZGVzYy50ZXh0KCkuc3RyaXAoKVs6MjQ0XQogICAgICAgICAgICByZWNbInVwZGF0ZWRfYXQiXSAgPSBk"
    "YXRldGltZS5ub3codGltZXpvbmUudXRjKS5pc29mb3JtYXQoKQogICAgICAgICAgICB3cml0ZV9qc29ubChzZWxmLl9wYXRoLCBz"
    "ZWxmLl9yZWNvcmRzKQogICAgICAgICAgICBzZWxmLnJlZnJlc2goKQoKICAgIGRlZiBfZG9fZGVsZXRlKHNlbGYpIC0+IE5vbmU6"
    "CiAgICAgICAgcm93ID0gc2VsZi5fdGFibGUuY3VycmVudFJvdygpCiAgICAgICAgaWYgcm93IDwgMCBvciByb3cgPj0gbGVuKHNl"
    "bGYuX3JlY29yZHMpOgogICAgICAgICAgICByZXR1cm4KICAgICAgICBjbWQgPSBzZWxmLl9yZWNvcmRzW3Jvd10uZ2V0KCJjb21t"
    "YW5kIiwidGhpcyBjb21tYW5kIikKICAgICAgICByZXBseSA9IFFNZXNzYWdlQm94LnF1ZXN0aW9uKAogICAgICAgICAgICBzZWxm"
    "LCAiRGVsZXRlIiwgZiJEZWxldGUgJ3tjbWR9Jz8iLAogICAgICAgICAgICBRTWVzc2FnZUJveC5TdGFuZGFyZEJ1dHRvbi5ZZXMg"
    "fCBRTWVzc2FnZUJveC5TdGFuZGFyZEJ1dHRvbi5ObwogICAgICAgICkKICAgICAgICBpZiByZXBseSA9PSBRTWVzc2FnZUJveC5T"
    "dGFuZGFyZEJ1dHRvbi5ZZXM6CiAgICAgICAgICAgIHNlbGYuX3JlY29yZHMucG9wKHJvdykKICAgICAgICAgICAgd3JpdGVfanNv"
    "bmwoc2VsZi5fcGF0aCwgc2VsZi5fcmVjb3JkcykKICAgICAgICAgICAgc2VsZi5yZWZyZXNoKCkKCgojIOKUgOKUgCBKT0IgVFJB"
    "Q0tFUiBUQUIg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNsYXNzIEpvYlRyYWNrZXJUYWIoUVdpZGdldCk6CiAgICAiIiIK"
    "ICAgIEpvYiBhcHBsaWNhdGlvbiB0cmFja2luZy4gRnVsbCByZWJ1aWxkIGZyb20gc3BlYy4KICAgIEZpZWxkczogQ29tcGFueSwg"
    "Sm9iIFRpdGxlLCBEYXRlIEFwcGxpZWQsIExpbmssIFN0YXR1cywgTm90ZXMuCiAgICBNdWx0aS1zZWxlY3QgaGlkZS91bmhpZGUv"
    "ZGVsZXRlLiBDU1YgYW5kIFRTViBleHBvcnQuCiAgICBIaWRkZW4gcm93cyA9IGNvbXBsZXRlZC9yZWplY3RlZCDigJQgc3RpbGwg"
    "c3RvcmVkLCBqdXN0IG5vdCBzaG93bi4KICAgICIiIgoKICAgIENPTFVNTlMgPSBbIkNvbXBhbnkiLCAiSm9iIFRpdGxlIiwgIkRh"
    "dGUgQXBwbGllZCIsCiAgICAgICAgICAgICAgICJMaW5rIiwgIlN0YXR1cyIsICJOb3RlcyJdCgogICAgZGVmIF9faW5pdF9fKHNl"
    "bGYsIHBhcmVudD1Ob25lKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKHBhcmVudCkKICAgICAgICBzZWxmLl9wYXRoICAgID0g"
    "Y2ZnX3BhdGgoIm1lbW9yaWVzIikgLyAiam9iX3RyYWNrZXIuanNvbmwiCiAgICAgICAgc2VsZi5fcmVjb3JkczogbGlzdFtkaWN0"
    "XSA9IFtdCiAgICAgICAgc2VsZi5fc2hvd19oaWRkZW4gPSBGYWxzZQogICAgICAgIHNlbGYuX3NldHVwX3VpKCkKICAgICAgICBz"
    "ZWxmLnJlZnJlc2goKQoKICAgIGRlZiBfc2V0dXBfdWkoc2VsZikgLT4gTm9uZToKICAgICAgICByb290ID0gUVZCb3hMYXlvdXQo"
    "c2VsZikKICAgICAgICByb290LnNldENvbnRlbnRzTWFyZ2lucyg0LCA0LCA0LCA0KQogICAgICAgIHJvb3Quc2V0U3BhY2luZyg0"
    "KQoKICAgICAgICBiYXIgPSBRSEJveExheW91dCgpCiAgICAgICAgc2VsZi5fYnRuX2FkZCAgICA9IF9nb3RoaWNfYnRuKCJBZGQi"
    "KQogICAgICAgIHNlbGYuX2J0bl9tb2RpZnkgPSBfZ290aGljX2J0bigiTW9kaWZ5IikKICAgICAgICBzZWxmLl9idG5faGlkZSAg"
    "ID0gX2dvdGhpY19idG4oIkFyY2hpdmUiLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIk1hcmsgc2Vs"
    "ZWN0ZWQgYXMgY29tcGxldGVkL3JlamVjdGVkIikKICAgICAgICBzZWxmLl9idG5fdW5oaWRlID0gX2dvdGhpY19idG4oIlJlc3Rv"
    "cmUiLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIlJlc3RvcmUgYXJjaGl2ZWQgYXBwbGljYXRpb25z"
    "IikKICAgICAgICBzZWxmLl9idG5fZGVsZXRlID0gX2dvdGhpY19idG4oIkRlbGV0ZSIpCiAgICAgICAgc2VsZi5fYnRuX3RvZ2ds"
    "ZSA9IF9nb3RoaWNfYnRuKCJTaG93IEFyY2hpdmVkIikKICAgICAgICBzZWxmLl9idG5fZXhwb3J0ID0gX2dvdGhpY19idG4oIkV4"
    "cG9ydCIpCgogICAgICAgIGZvciBiIGluIChzZWxmLl9idG5fYWRkLCBzZWxmLl9idG5fbW9kaWZ5LCBzZWxmLl9idG5faGlkZSwK"
    "ICAgICAgICAgICAgICAgICAgc2VsZi5fYnRuX3VuaGlkZSwgc2VsZi5fYnRuX2RlbGV0ZSwKICAgICAgICAgICAgICAgICAgc2Vs"
    "Zi5fYnRuX3RvZ2dsZSwgc2VsZi5fYnRuX2V4cG9ydCk6CiAgICAgICAgICAgIGIuc2V0TWluaW11bVdpZHRoKDcwKQogICAgICAg"
    "ICAgICBiLnNldE1pbmltdW1IZWlnaHQoMjYpCiAgICAgICAgICAgIGJhci5hZGRXaWRnZXQoYikKCiAgICAgICAgc2VsZi5fYnRu"
    "X2FkZC5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fZG9fYWRkKQogICAgICAgIHNlbGYuX2J0bl9tb2RpZnkuY2xpY2tlZC5jb25uZWN0"
    "KHNlbGYuX2RvX21vZGlmeSkKICAgICAgICBzZWxmLl9idG5faGlkZS5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fZG9faGlkZSkKICAg"
    "ICAgICBzZWxmLl9idG5fdW5oaWRlLmNsaWNrZWQuY29ubmVjdChzZWxmLl9kb191bmhpZGUpCiAgICAgICAgc2VsZi5fYnRuX2Rl"
    "bGV0ZS5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fZG9fZGVsZXRlKQogICAgICAgIHNlbGYuX2J0bl90b2dnbGUuY2xpY2tlZC5jb25u"
    "ZWN0KHNlbGYuX3RvZ2dsZV9oaWRkZW4pCiAgICAgICAgc2VsZi5fYnRuX2V4cG9ydC5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fZG9f"
    "ZXhwb3J0KQogICAgICAgIGJhci5hZGRTdHJldGNoKCkKICAgICAgICByb290LmFkZExheW91dChiYXIpCgogICAgICAgIHNlbGYu"
    "X3RhYmxlID0gUVRhYmxlV2lkZ2V0KDAsIGxlbihzZWxmLkNPTFVNTlMpKQogICAgICAgIHNlbGYuX3RhYmxlLnNldEhvcml6b250"
    "YWxIZWFkZXJMYWJlbHMoc2VsZi5DT0xVTU5TKQogICAgICAgIGhoID0gc2VsZi5fdGFibGUuaG9yaXpvbnRhbEhlYWRlcigpCiAg"
    "ICAgICAgIyBDb21wYW55IGFuZCBKb2IgVGl0bGUgc3RyZXRjaAogICAgICAgIGhoLnNldFNlY3Rpb25SZXNpemVNb2RlKDAsIFFI"
    "ZWFkZXJWaWV3LlJlc2l6ZU1vZGUuU3RyZXRjaCkKICAgICAgICBoaC5zZXRTZWN0aW9uUmVzaXplTW9kZSgxLCBRSGVhZGVyVmll"
    "dy5SZXNpemVNb2RlLlN0cmV0Y2gpCiAgICAgICAgIyBEYXRlIEFwcGxpZWQg4oCUIGZpeGVkIHJlYWRhYmxlIHdpZHRoCiAgICAg"
    "ICAgaGguc2V0U2VjdGlvblJlc2l6ZU1vZGUoMiwgUUhlYWRlclZpZXcuUmVzaXplTW9kZS5GaXhlZCkKICAgICAgICBzZWxmLl90"
    "YWJsZS5zZXRDb2x1bW5XaWR0aCgyLCAxMDApCiAgICAgICAgIyBMaW5rIHN0cmV0Y2hlcwogICAgICAgIGhoLnNldFNlY3Rpb25S"
    "ZXNpemVNb2RlKDMsIFFIZWFkZXJWaWV3LlJlc2l6ZU1vZGUuU3RyZXRjaCkKICAgICAgICAjIFN0YXR1cyDigJQgZml4ZWQgd2lk"
    "dGgKICAgICAgICBoaC5zZXRTZWN0aW9uUmVzaXplTW9kZSg0LCBRSGVhZGVyVmlldy5SZXNpemVNb2RlLkZpeGVkKQogICAgICAg"
    "IHNlbGYuX3RhYmxlLnNldENvbHVtbldpZHRoKDQsIDgwKQogICAgICAgICMgTm90ZXMgc3RyZXRjaGVzCiAgICAgICAgaGguc2V0"
    "U2VjdGlvblJlc2l6ZU1vZGUoNSwgUUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJldGNoKQoKICAgICAgICBzZWxmLl90YWJsZS5z"
    "ZXRTZWxlY3Rpb25CZWhhdmlvcigKICAgICAgICAgICAgUUFic3RyYWN0SXRlbVZpZXcuU2VsZWN0aW9uQmVoYXZpb3IuU2VsZWN0"
    "Um93cykKICAgICAgICBzZWxmLl90YWJsZS5zZXRTZWxlY3Rpb25Nb2RlKAogICAgICAgICAgICBRQWJzdHJhY3RJdGVtVmlldy5T"
    "ZWxlY3Rpb25Nb2RlLkV4dGVuZGVkU2VsZWN0aW9uKQogICAgICAgIHNlbGYuX3RhYmxlLnNldEFsdGVybmF0aW5nUm93Q29sb3Jz"
    "KFRydWUpCiAgICAgICAgc2VsZi5fdGFibGUuc2V0U3R5bGVTaGVldChfZ290aGljX3RhYmxlX3N0eWxlKCkpCiAgICAgICAgcm9v"
    "dC5hZGRXaWRnZXQoc2VsZi5fdGFibGUsIDEpCgogICAgZGVmIHJlZnJlc2goc2VsZikgLT4gTm9uZToKICAgICAgICBzZWxmLl9y"
    "ZWNvcmRzID0gcmVhZF9qc29ubChzZWxmLl9wYXRoKQogICAgICAgIHNlbGYuX3RhYmxlLnNldFJvd0NvdW50KDApCiAgICAgICAg"
    "Zm9yIHJlYyBpbiBzZWxmLl9yZWNvcmRzOgogICAgICAgICAgICBoaWRkZW4gPSBib29sKHJlYy5nZXQoImhpZGRlbiIsIEZhbHNl"
    "KSkKICAgICAgICAgICAgaWYgaGlkZGVuIGFuZCBub3Qgc2VsZi5fc2hvd19oaWRkZW46CiAgICAgICAgICAgICAgICBjb250aW51"
    "ZQogICAgICAgICAgICByID0gc2VsZi5fdGFibGUucm93Q291bnQoKQogICAgICAgICAgICBzZWxmLl90YWJsZS5pbnNlcnRSb3co"
    "cikKICAgICAgICAgICAgc3RhdHVzID0gIkFyY2hpdmVkIiBpZiBoaWRkZW4gZWxzZSByZWMuZ2V0KCJzdGF0dXMiLCJBY3RpdmUi"
    "KQogICAgICAgICAgICB2YWxzID0gWwogICAgICAgICAgICAgICAgcmVjLmdldCgiY29tcGFueSIsIiIpLAogICAgICAgICAgICAg"
    "ICAgcmVjLmdldCgiam9iX3RpdGxlIiwiIiksCiAgICAgICAgICAgICAgICByZWMuZ2V0KCJkYXRlX2FwcGxpZWQiLCIiKSwKICAg"
    "ICAgICAgICAgICAgIHJlYy5nZXQoImxpbmsiLCIiKSwKICAgICAgICAgICAgICAgIHN0YXR1cywKICAgICAgICAgICAgICAgIHJl"
    "Yy5nZXQoIm5vdGVzIiwiIiksCiAgICAgICAgICAgIF0KICAgICAgICAgICAgZm9yIGMsIHYgaW4gZW51bWVyYXRlKHZhbHMpOgog"
    "ICAgICAgICAgICAgICAgaXRlbSA9IFFUYWJsZVdpZGdldEl0ZW0oc3RyKHYpKQogICAgICAgICAgICAgICAgaWYgaGlkZGVuOgog"
    "ICAgICAgICAgICAgICAgICAgIGl0ZW0uc2V0Rm9yZWdyb3VuZChRQ29sb3IoQ19URVhUX0RJTSkpCiAgICAgICAgICAgICAgICBz"
    "ZWxmLl90YWJsZS5zZXRJdGVtKHIsIGMsIGl0ZW0pCiAgICAgICAgICAgICMgU3RvcmUgcmVjb3JkIGluZGV4IGluIGZpcnN0IGNv"
    "bHVtbidzIHVzZXIgZGF0YQogICAgICAgICAgICBzZWxmLl90YWJsZS5pdGVtKHIsIDApLnNldERhdGEoCiAgICAgICAgICAgICAg"
    "ICBRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUsCiAgICAgICAgICAgICAgICBzZWxmLl9yZWNvcmRzLmluZGV4KHJlYykKICAgICAg"
    "ICAgICAgKQoKICAgIGRlZiBfc2VsZWN0ZWRfaW5kaWNlcyhzZWxmKSAtPiBsaXN0W2ludF06CiAgICAgICAgaW5kaWNlcyA9IHNl"
    "dCgpCiAgICAgICAgZm9yIGl0ZW0gaW4gc2VsZi5fdGFibGUuc2VsZWN0ZWRJdGVtcygpOgogICAgICAgICAgICByb3dfaXRlbSA9"
    "IHNlbGYuX3RhYmxlLml0ZW0oaXRlbS5yb3coKSwgMCkKICAgICAgICAgICAgaWYgcm93X2l0ZW06CiAgICAgICAgICAgICAgICBp"
    "ZHggPSByb3dfaXRlbS5kYXRhKFF0Lkl0ZW1EYXRhUm9sZS5Vc2VyUm9sZSkKICAgICAgICAgICAgICAgIGlmIGlkeCBpcyBub3Qg"
    "Tm9uZToKICAgICAgICAgICAgICAgICAgICBpbmRpY2VzLmFkZChpZHgpCiAgICAgICAgcmV0dXJuIHNvcnRlZChpbmRpY2VzKQoK"
    "ICAgIGRlZiBfZGlhbG9nKHNlbGYsIHJlYzogZGljdCA9IE5vbmUpIC0+IE9wdGlvbmFsW2RpY3RdOgogICAgICAgIGRsZyAgPSBR"
    "RGlhbG9nKHNlbGYpCiAgICAgICAgZGxnLnNldFdpbmRvd1RpdGxlKCJKb2IgQXBwbGljYXRpb24iKQogICAgICAgIGRsZy5zZXRT"
    "dHlsZVNoZWV0KGYiYmFja2dyb3VuZDoge0NfQkcyfTsgY29sb3I6IHtDX0dPTER9OyIpCiAgICAgICAgZGxnLnJlc2l6ZSg1MDAs"
    "IDMyMCkKICAgICAgICBmb3JtID0gUUZvcm1MYXlvdXQoZGxnKQoKICAgICAgICBjb21wYW55ID0gUUxpbmVFZGl0KHJlYy5nZXQo"
    "ImNvbXBhbnkiLCIiKSBpZiByZWMgZWxzZSAiIikKICAgICAgICB0aXRsZSAgID0gUUxpbmVFZGl0KHJlYy5nZXQoImpvYl90aXRs"
    "ZSIsIiIpIGlmIHJlYyBlbHNlICIiKQogICAgICAgIGRlICAgICAgPSBRRGF0ZUVkaXQoKQogICAgICAgIGRlLnNldENhbGVuZGFy"
    "UG9wdXAoVHJ1ZSkKICAgICAgICBkZS5zZXREaXNwbGF5Rm9ybWF0KCJ5eXl5LU1NLWRkIikKICAgICAgICBpZiByZWMgYW5kIHJl"
    "Yy5nZXQoImRhdGVfYXBwbGllZCIpOgogICAgICAgICAgICBkZS5zZXREYXRlKFFEYXRlLmZyb21TdHJpbmcocmVjWyJkYXRlX2Fw"
    "cGxpZWQiXSwieXl5eS1NTS1kZCIpKQogICAgICAgIGVsc2U6CiAgICAgICAgICAgIGRlLnNldERhdGUoUURhdGUuY3VycmVudERh"
    "dGUoKSkKICAgICAgICBsaW5rICAgID0gUUxpbmVFZGl0KHJlYy5nZXQoImxpbmsiLCIiKSBpZiByZWMgZWxzZSAiIikKICAgICAg"
    "ICBzdGF0dXMgID0gUUxpbmVFZGl0KHJlYy5nZXQoInN0YXR1cyIsIkFwcGxpZWQiKSBpZiByZWMgZWxzZSAiQXBwbGllZCIpCiAg"
    "ICAgICAgbm90ZXMgICA9IFFMaW5lRWRpdChyZWMuZ2V0KCJub3RlcyIsIiIpIGlmIHJlYyBlbHNlICIiKQoKICAgICAgICBmb3Ig"
    "bGFiZWwsIHdpZGdldCBpbiBbCiAgICAgICAgICAgICgiQ29tcGFueToiLCBjb21wYW55KSwgKCJKb2IgVGl0bGU6IiwgdGl0bGUp"
    "LAogICAgICAgICAgICAoIkRhdGUgQXBwbGllZDoiLCBkZSksICgiTGluazoiLCBsaW5rKSwKICAgICAgICAgICAgKCJTdGF0dXM6"
    "Iiwgc3RhdHVzKSwgKCJOb3RlczoiLCBub3RlcyksCiAgICAgICAgXToKICAgICAgICAgICAgZm9ybS5hZGRSb3cobGFiZWwsIHdp"
    "ZGdldCkKCiAgICAgICAgYnRucyA9IFFIQm94TGF5b3V0KCkKICAgICAgICBvayA9IF9nb3RoaWNfYnRuKCJTYXZlIik7IGN4ID0g"
    "X2dvdGhpY19idG4oIkNhbmNlbCIpCiAgICAgICAgb2suY2xpY2tlZC5jb25uZWN0KGRsZy5hY2NlcHQpOyBjeC5jbGlja2VkLmNv"
    "bm5lY3QoZGxnLnJlamVjdCkKICAgICAgICBidG5zLmFkZFdpZGdldChvayk7IGJ0bnMuYWRkV2lkZ2V0KGN4KQogICAgICAgIGZv"
    "cm0uYWRkUm93KGJ0bnMpCgogICAgICAgIGlmIGRsZy5leGVjKCkgPT0gUURpYWxvZy5EaWFsb2dDb2RlLkFjY2VwdGVkOgogICAg"
    "ICAgICAgICByZXR1cm4gewogICAgICAgICAgICAgICAgImNvbXBhbnkiOiAgICAgIGNvbXBhbnkudGV4dCgpLnN0cmlwKCksCiAg"
    "ICAgICAgICAgICAgICAiam9iX3RpdGxlIjogICAgdGl0bGUudGV4dCgpLnN0cmlwKCksCiAgICAgICAgICAgICAgICAiZGF0ZV9h"
    "cHBsaWVkIjogZGUuZGF0ZSgpLnRvU3RyaW5nKCJ5eXl5LU1NLWRkIiksCiAgICAgICAgICAgICAgICAibGluayI6ICAgICAgICAg"
    "bGluay50ZXh0KCkuc3RyaXAoKSwKICAgICAgICAgICAgICAgICJzdGF0dXMiOiAgICAgICBzdGF0dXMudGV4dCgpLnN0cmlwKCkg"
    "b3IgIkFwcGxpZWQiLAogICAgICAgICAgICAgICAgIm5vdGVzIjogICAgICAgIG5vdGVzLnRleHQoKS5zdHJpcCgpLAogICAgICAg"
    "ICAgICB9CiAgICAgICAgcmV0dXJuIE5vbmUKCiAgICBkZWYgX2RvX2FkZChzZWxmKSAtPiBOb25lOgogICAgICAgIHAgPSBzZWxm"
    "Ll9kaWFsb2coKQogICAgICAgIGlmIG5vdCBwOgogICAgICAgICAgICByZXR1cm4KICAgICAgICBub3cgPSBkYXRldGltZS5ub3co"
    "dGltZXpvbmUudXRjKS5pc29mb3JtYXQoKQogICAgICAgIHAudXBkYXRlKHsKICAgICAgICAgICAgImlkIjogICAgICAgICAgICAg"
    "c3RyKHV1aWQudXVpZDQoKSksCiAgICAgICAgICAgICJoaWRkZW4iOiAgICAgICAgIEZhbHNlLAogICAgICAgICAgICAiY29tcGxl"
    "dGVkX2RhdGUiOiBOb25lLAogICAgICAgICAgICAiY3JlYXRlZF9hdCI6ICAgICBub3csCiAgICAgICAgICAgICJ1cGRhdGVkX2F0"
    "IjogICAgIG5vdywKICAgICAgICB9KQogICAgICAgIHNlbGYuX3JlY29yZHMuYXBwZW5kKHApCiAgICAgICAgd3JpdGVfanNvbmwo"
    "c2VsZi5fcGF0aCwgc2VsZi5fcmVjb3JkcykKICAgICAgICBzZWxmLnJlZnJlc2goKQoKICAgIGRlZiBfZG9fbW9kaWZ5KHNlbGYp"
    "IC0+IE5vbmU6CiAgICAgICAgaWR4cyA9IHNlbGYuX3NlbGVjdGVkX2luZGljZXMoKQogICAgICAgIGlmIGxlbihpZHhzKSAhPSAx"
    "OgogICAgICAgICAgICBRTWVzc2FnZUJveC5pbmZvcm1hdGlvbihzZWxmLCAiTW9kaWZ5IiwKICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgICAgICAgICAgIlNlbGVjdCBleGFjdGx5IG9uZSByb3cgdG8gbW9kaWZ5LiIpCiAgICAgICAgICAgIHJldHVybgogICAg"
    "ICAgIHJlYyA9IHNlbGYuX3JlY29yZHNbaWR4c1swXV0KICAgICAgICBwICAgPSBzZWxmLl9kaWFsb2cocmVjKQogICAgICAgIGlm"
    "IG5vdCBwOgogICAgICAgICAgICByZXR1cm4KICAgICAgICByZWMudXBkYXRlKHApCiAgICAgICAgcmVjWyJ1cGRhdGVkX2F0Il0g"
    "PSBkYXRldGltZS5ub3codGltZXpvbmUudXRjKS5pc29mb3JtYXQoKQogICAgICAgIHdyaXRlX2pzb25sKHNlbGYuX3BhdGgsIHNl"
    "bGYuX3JlY29yZHMpCiAgICAgICAgc2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX2RvX2hpZGUoc2VsZikgLT4gTm9uZToKICAgICAg"
    "ICBmb3IgaWR4IGluIHNlbGYuX3NlbGVjdGVkX2luZGljZXMoKToKICAgICAgICAgICAgaWYgaWR4IDwgbGVuKHNlbGYuX3JlY29y"
    "ZHMpOgogICAgICAgICAgICAgICAgc2VsZi5fcmVjb3Jkc1tpZHhdWyJoaWRkZW4iXSAgICAgICAgID0gVHJ1ZQogICAgICAgICAg"
    "ICAgICAgc2VsZi5fcmVjb3Jkc1tpZHhdWyJjb21wbGV0ZWRfZGF0ZSJdID0gKAogICAgICAgICAgICAgICAgICAgIHNlbGYuX3Jl"
    "Y29yZHNbaWR4XS5nZXQoImNvbXBsZXRlZF9kYXRlIikgb3IKICAgICAgICAgICAgICAgICAgICBkYXRldGltZS5ub3coKS5kYXRl"
    "KCkuaXNvZm9ybWF0KCkKICAgICAgICAgICAgICAgICkKICAgICAgICAgICAgICAgIHNlbGYuX3JlY29yZHNbaWR4XVsidXBkYXRl"
    "ZF9hdCJdID0gKAogICAgICAgICAgICAgICAgICAgIGRhdGV0aW1lLm5vdyh0aW1lem9uZS51dGMpLmlzb2Zvcm1hdCgpCiAgICAg"
    "ICAgICAgICAgICApCiAgICAgICAgd3JpdGVfanNvbmwoc2VsZi5fcGF0aCwgc2VsZi5fcmVjb3JkcykKICAgICAgICBzZWxmLnJl"
    "ZnJlc2goKQoKICAgIGRlZiBfZG9fdW5oaWRlKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgZm9yIGlkeCBpbiBzZWxmLl9zZWxlY3Rl"
    "ZF9pbmRpY2VzKCk6CiAgICAgICAgICAgIGlmIGlkeCA8IGxlbihzZWxmLl9yZWNvcmRzKToKICAgICAgICAgICAgICAgIHNlbGYu"
    "X3JlY29yZHNbaWR4XVsiaGlkZGVuIl0gICAgID0gRmFsc2UKICAgICAgICAgICAgICAgIHNlbGYuX3JlY29yZHNbaWR4XVsidXBk"
    "YXRlZF9hdCJdID0gKAogICAgICAgICAgICAgICAgICAgIGRhdGV0aW1lLm5vdyh0aW1lem9uZS51dGMpLmlzb2Zvcm1hdCgpCiAg"
    "ICAgICAgICAgICAgICApCiAgICAgICAgd3JpdGVfanNvbmwoc2VsZi5fcGF0aCwgc2VsZi5fcmVjb3JkcykKICAgICAgICBzZWxm"
    "LnJlZnJlc2goKQoKICAgIGRlZiBfZG9fZGVsZXRlKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgaWR4cyA9IHNlbGYuX3NlbGVjdGVk"
    "X2luZGljZXMoKQogICAgICAgIGlmIG5vdCBpZHhzOgogICAgICAgICAgICByZXR1cm4KICAgICAgICByZXBseSA9IFFNZXNzYWdl"
    "Qm94LnF1ZXN0aW9uKAogICAgICAgICAgICBzZWxmLCAiRGVsZXRlIiwKICAgICAgICAgICAgZiJEZWxldGUge2xlbihpZHhzKX0g"
    "c2VsZWN0ZWQgYXBwbGljYXRpb24ocyk/IENhbm5vdCBiZSB1bmRvbmUuIiwKICAgICAgICAgICAgUU1lc3NhZ2VCb3guU3RhbmRh"
    "cmRCdXR0b24uWWVzIHwgUU1lc3NhZ2VCb3guU3RhbmRhcmRCdXR0b24uTm8KICAgICAgICApCiAgICAgICAgaWYgcmVwbHkgPT0g"
    "UU1lc3NhZ2VCb3guU3RhbmRhcmRCdXR0b24uWWVzOgogICAgICAgICAgICBiYWQgPSBzZXQoaWR4cykKICAgICAgICAgICAgc2Vs"
    "Zi5fcmVjb3JkcyA9IFtyIGZvciBpLCByIGluIGVudW1lcmF0ZShzZWxmLl9yZWNvcmRzKQogICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgIGlmIGkgbm90IGluIGJhZF0KICAgICAgICAgICAgd3JpdGVfanNvbmwoc2VsZi5fcGF0aCwgc2VsZi5fcmVjb3JkcykK"
    "ICAgICAgICAgICAgc2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX3RvZ2dsZV9oaWRkZW4oc2VsZikgLT4gTm9uZToKICAgICAgICBz"
    "ZWxmLl9zaG93X2hpZGRlbiA9IG5vdCBzZWxmLl9zaG93X2hpZGRlbgogICAgICAgIHNlbGYuX2J0bl90b2dnbGUuc2V0VGV4dCgK"
    "ICAgICAgICAgICAgIuKYgCBIaWRlIEFyY2hpdmVkIiBpZiBzZWxmLl9zaG93X2hpZGRlbiBlbHNlICLimL0gU2hvdyBBcmNoaXZl"
    "ZCIKICAgICAgICApCiAgICAgICAgc2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX2RvX2V4cG9ydChzZWxmKSAtPiBOb25lOgogICAg"
    "ICAgIHBhdGgsIGZpbHQgPSBRRmlsZURpYWxvZy5nZXRTYXZlRmlsZU5hbWUoCiAgICAgICAgICAgIHNlbGYsICJFeHBvcnQgSm9i"
    "IFRyYWNrZXIiLAogICAgICAgICAgICBzdHIoY2ZnX3BhdGgoImV4cG9ydHMiKSAvICJqb2JfdHJhY2tlci5jc3YiKSwKICAgICAg"
    "ICAgICAgIkNTViBGaWxlcyAoKi5jc3YpOztUYWIgRGVsaW1pdGVkICgqLnR4dCkiCiAgICAgICAgKQogICAgICAgIGlmIG5vdCBw"
    "YXRoOgogICAgICAgICAgICByZXR1cm4KICAgICAgICBkZWxpbSA9ICJcdCIgaWYgcGF0aC5sb3dlcigpLmVuZHN3aXRoKCIudHh0"
    "IikgZWxzZSAiLCIKICAgICAgICBoZWFkZXIgPSBbImNvbXBhbnkiLCJqb2JfdGl0bGUiLCJkYXRlX2FwcGxpZWQiLCJsaW5rIiwK"
    "ICAgICAgICAgICAgICAgICAgInN0YXR1cyIsImhpZGRlbiIsImNvbXBsZXRlZF9kYXRlIiwibm90ZXMiXQogICAgICAgIHdpdGgg"
    "b3BlbihwYXRoLCAidyIsIGVuY29kaW5nPSJ1dGYtOCIsIG5ld2xpbmU9IiIpIGFzIGY6CiAgICAgICAgICAgIGYud3JpdGUoZGVs"
    "aW0uam9pbihoZWFkZXIpICsgIlxuIikKICAgICAgICAgICAgZm9yIHJlYyBpbiBzZWxmLl9yZWNvcmRzOgogICAgICAgICAgICAg"
    "ICAgdmFscyA9IFsKICAgICAgICAgICAgICAgICAgICByZWMuZ2V0KCJjb21wYW55IiwiIiksCiAgICAgICAgICAgICAgICAgICAg"
    "cmVjLmdldCgiam9iX3RpdGxlIiwiIiksCiAgICAgICAgICAgICAgICAgICAgcmVjLmdldCgiZGF0ZV9hcHBsaWVkIiwiIiksCiAg"
    "ICAgICAgICAgICAgICAgICAgcmVjLmdldCgibGluayIsIiIpLAogICAgICAgICAgICAgICAgICAgIHJlYy5nZXQoInN0YXR1cyIs"
    "IiIpLAogICAgICAgICAgICAgICAgICAgIHN0cihib29sKHJlYy5nZXQoImhpZGRlbiIsRmFsc2UpKSksCiAgICAgICAgICAgICAg"
    "ICAgICAgcmVjLmdldCgiY29tcGxldGVkX2RhdGUiLCIiKSBvciAiIiwKICAgICAgICAgICAgICAgICAgICByZWMuZ2V0KCJub3Rl"
    "cyIsIiIpLAogICAgICAgICAgICAgICAgXQogICAgICAgICAgICAgICAgZi53cml0ZShkZWxpbS5qb2luKAogICAgICAgICAgICAg"
    "ICAgICAgIHN0cih2KS5yZXBsYWNlKCJcbiIsIiAiKS5yZXBsYWNlKGRlbGltLCIgIikKICAgICAgICAgICAgICAgICAgICBmb3Ig"
    "diBpbiB2YWxzCiAgICAgICAgICAgICAgICApICsgIlxuIikKICAgICAgICBRTWVzc2FnZUJveC5pbmZvcm1hdGlvbihzZWxmLCAi"
    "RXhwb3J0ZWQiLAogICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIGYiU2F2ZWQgdG8ge3BhdGh9IikKCgojIOKUgOKUgCBT"
    "RUxGIFRBQiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xhc3MgU2VsZlRhYihRV2lk"
    "Z2V0KToKICAgICIiIgogICAgUGVyc29uYSdzIGludGVybmFsIGRpYWxvZ3VlIHNwYWNlLgogICAgUmVjZWl2ZXM6IGlkbGUgbmFy"
    "cmF0aXZlIG91dHB1dCwgdW5zb2xpY2l0ZWQgdHJhbnNtaXNzaW9ucywKICAgICAgICAgICAgICBQb0kgbGlzdCBmcm9tIGRhaWx5"
    "IHJlZmxlY3Rpb24sIHVuYW5zd2VyZWQgcXVlc3Rpb24gZmxhZ3MsCiAgICAgICAgICAgICAgam91cm5hbCBsb2FkIG5vdGlmaWNh"
    "dGlvbnMuCiAgICBSZWFkLW9ubHkgZGlzcGxheS4gU2VwYXJhdGUgZnJvbSBwZXJzb25hIGNoYXQgdGFiIGFsd2F5cy4KICAgICIi"
    "IgoKICAgIGRlZiBfX2luaXRfXyhzZWxmLCBwYXJlbnQ9Tm9uZSk6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAg"
    "ICAgICAgcm9vdCA9IFFWQm94TGF5b3V0KHNlbGYpCiAgICAgICAgcm9vdC5zZXRDb250ZW50c01hcmdpbnMoNCwgNCwgNCwgNCkK"
    "ICAgICAgICByb290LnNldFNwYWNpbmcoNCkKCiAgICAgICAgaGRyID0gUUhCb3hMYXlvdXQoKQogICAgICAgIGhkci5hZGRXaWRn"
    "ZXQoX3NlY3Rpb25fbGJsKGYi4p2nIElOTkVSIFNBTkNUVU0g4oCUIHtERUNLX05BTUUudXBwZXIoKX0nUyBQUklWQVRFIFRIT1VH"
    "SFRTIikpCiAgICAgICAgc2VsZi5fYnRuX2NsZWFyID0gX2dvdGhpY19idG4oIuKclyBDbGVhciIpCiAgICAgICAgc2VsZi5fYnRu"
    "X2NsZWFyLnNldEZpeGVkV2lkdGgoODApCiAgICAgICAgc2VsZi5fYnRuX2NsZWFyLmNsaWNrZWQuY29ubmVjdChzZWxmLmNsZWFy"
    "KQogICAgICAgIGhkci5hZGRTdHJldGNoKCkKICAgICAgICBoZHIuYWRkV2lkZ2V0KHNlbGYuX2J0bl9jbGVhcikKICAgICAgICBy"
    "b290LmFkZExheW91dChoZHIpCgogICAgICAgIHNlbGYuX2Rpc3BsYXkgPSBRVGV4dEVkaXQoKQogICAgICAgIHNlbGYuX2Rpc3Bs"
    "YXkuc2V0UmVhZE9ubHkoVHJ1ZSkKICAgICAgICBzZWxmLl9kaXNwbGF5LnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiYmFj"
    "a2dyb3VuZDoge0NfTU9OSVRPUn07IGNvbG9yOiB7Q19HT0xEfTsgIgogICAgICAgICAgICBmImJvcmRlcjogMXB4IHNvbGlkIHtD"
    "X1BVUlBMRV9ESU19OyAiCiAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsgZm9udC1zaXplOiAx"
    "MXB4OyBwYWRkaW5nOiA4cHg7IgogICAgICAgICkKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl9kaXNwbGF5LCAxKQoKICAg"
    "IGRlZiBhcHBlbmQoc2VsZiwgbGFiZWw6IHN0ciwgdGV4dDogc3RyKSAtPiBOb25lOgogICAgICAgIHRpbWVzdGFtcCA9IGRhdGV0"
    "aW1lLm5vdygpLnN0cmZ0aW1lKCIlSDolTTolUyIpCiAgICAgICAgY29sb3JzID0gewogICAgICAgICAgICAiTkFSUkFUSVZFIjog"
    "IENfR09MRCwKICAgICAgICAgICAgIlJFRkxFQ1RJT04iOiBDX1BVUlBMRSwKICAgICAgICAgICAgIkpPVVJOQUwiOiAgICBDX1NJ"
    "TFZFUiwKICAgICAgICAgICAgIlBPSSI6ICAgICAgICBDX0dPTERfRElNLAogICAgICAgICAgICAiU1lTVEVNIjogICAgIENfVEVY"
    "VF9ESU0sCiAgICAgICAgfQogICAgICAgIGNvbG9yID0gY29sb3JzLmdldChsYWJlbC51cHBlcigpLCBDX0dPTEQpCiAgICAgICAg"
    "c2VsZi5fZGlzcGxheS5hcHBlbmQoCiAgICAgICAgICAgIGYnPHNwYW4gc3R5bGU9ImNvbG9yOntDX1RFWFRfRElNfTsgZm9udC1z"
    "aXplOjEwcHg7Ij4nCiAgICAgICAgICAgIGYnW3t0aW1lc3RhbXB9XSA8L3NwYW4+JwogICAgICAgICAgICBmJzxzcGFuIHN0eWxl"
    "PSJjb2xvcjp7Y29sb3J9OyBmb250LXdlaWdodDpib2xkOyI+JwogICAgICAgICAgICBmJ+KdpyB7bGFiZWx9PC9zcGFuPjxicj4n"
    "CiAgICAgICAgICAgIGYnPHNwYW4gc3R5bGU9ImNvbG9yOntDX0dPTER9OyI+e3RleHR9PC9zcGFuPicKICAgICAgICApCiAgICAg"
    "ICAgc2VsZi5fZGlzcGxheS5hcHBlbmQoIiIpCiAgICAgICAgc2VsZi5fZGlzcGxheS52ZXJ0aWNhbFNjcm9sbEJhcigpLnNldFZh"
    "bHVlKAogICAgICAgICAgICBzZWxmLl9kaXNwbGF5LnZlcnRpY2FsU2Nyb2xsQmFyKCkubWF4aW11bSgpCiAgICAgICAgKQoKICAg"
    "IGRlZiBjbGVhcihzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2Rpc3BsYXkuY2xlYXIoKQoKCiMg4pSA4pSAIERJQUdOT1NU"
    "SUNTIFRBQiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKY2xhc3MgRGlhZ25vc3RpY3NUYWIoUVdpZGdldCk6CiAgICAiIiIK"
    "ICAgIEJhY2tlbmQgZGlhZ25vc3RpY3MgZGlzcGxheS4KICAgIFJlY2VpdmVzOiBoYXJkd2FyZSBkZXRlY3Rpb24gcmVzdWx0cywg"
    "ZGVwZW5kZW5jeSBjaGVjayByZXN1bHRzLAogICAgICAgICAgICAgIEFQSSBlcnJvcnMsIHN5bmMgZmFpbHVyZXMsIHRpbWVyIGV2"
    "ZW50cywgam91cm5hbCBsb2FkIG5vdGljZXMsCiAgICAgICAgICAgICAgbW9kZWwgbG9hZCBzdGF0dXMsIEdvb2dsZSBhdXRoIGV2"
    "ZW50cy4KICAgIEFsd2F5cyBzZXBhcmF0ZSBmcm9tIHBlcnNvbmEgY2hhdCB0YWIuCiAgICAiIiIKCiAgICBkZWYgX19pbml0X18o"
    "c2VsZiwgcGFyZW50PU5vbmUpOgogICAgICAgIHN1cGVyKCkuX19pbml0X18ocGFyZW50KQogICAgICAgIHJvb3QgPSBRVkJveExh"
    "eW91dChzZWxmKQogICAgICAgIHJvb3Quc2V0Q29udGVudHNNYXJnaW5zKDQsIDQsIDQsIDQpCiAgICAgICAgcm9vdC5zZXRTcGFj"
    "aW5nKDQpCgogICAgICAgIGhkciA9IFFIQm94TGF5b3V0KCkKICAgICAgICBoZHIuYWRkV2lkZ2V0KF9zZWN0aW9uX2xibCgi4p2n"
    "IERJQUdOT1NUSUNTIOKAlCBTWVNURU0gJiBCQUNLRU5EIExPRyIpKQogICAgICAgIHNlbGYuX2J0bl9jbGVhciA9IF9nb3RoaWNf"
    "YnRuKCLinJcgQ2xlYXIiKQogICAgICAgIHNlbGYuX2J0bl9jbGVhci5zZXRGaXhlZFdpZHRoKDgwKQogICAgICAgIHNlbGYuX2J0"
    "bl9jbGVhci5jbGlja2VkLmNvbm5lY3Qoc2VsZi5jbGVhcikKICAgICAgICBoZHIuYWRkU3RyZXRjaCgpCiAgICAgICAgaGRyLmFk"
    "ZFdpZGdldChzZWxmLl9idG5fY2xlYXIpCiAgICAgICAgcm9vdC5hZGRMYXlvdXQoaGRyKQoKICAgICAgICBzZWxmLl9kaXNwbGF5"
    "ID0gUVRleHRFZGl0KCkKICAgICAgICBzZWxmLl9kaXNwbGF5LnNldFJlYWRPbmx5KFRydWUpCiAgICAgICAgc2VsZi5fZGlzcGxh"
    "eS5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX01PTklUT1J9OyBjb2xvcjoge0NfU0lMVkVSfTsg"
    "IgogICAgICAgICAgICBmImJvcmRlcjogMXB4IHNvbGlkIHtDX0JPUkRFUn07ICIKICAgICAgICAgICAgZiJmb250LWZhbWlseTog"
    "J0NvdXJpZXIgTmV3JywgbW9ub3NwYWNlOyAiCiAgICAgICAgICAgIGYiZm9udC1zaXplOiAxMHB4OyBwYWRkaW5nOiA4cHg7Igog"
    "ICAgICAgICkKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl9kaXNwbGF5LCAxKQoKICAgIGRlZiBsb2coc2VsZiwgbWVzc2Fn"
    "ZTogc3RyLCBsZXZlbDogc3RyID0gIklORk8iKSAtPiBOb25lOgogICAgICAgIHRpbWVzdGFtcCA9IGRhdGV0aW1lLm5vdygpLnN0"
    "cmZ0aW1lKCIlSDolTTolUyIpCiAgICAgICAgbGV2ZWxfY29sb3JzID0gewogICAgICAgICAgICAiSU5GTyI6ICBDX1NJTFZFUiwK"
    "ICAgICAgICAgICAgIk9LIjogICAgQ19HUkVFTiwKICAgICAgICAgICAgIldBUk4iOiAgQ19HT0xELAogICAgICAgICAgICAiRVJS"
    "T1IiOiBDX0JMT09ELAogICAgICAgICAgICAiREVCVUciOiBDX1RFWFRfRElNLAogICAgICAgIH0KICAgICAgICBjb2xvciA9IGxl"
    "dmVsX2NvbG9ycy5nZXQobGV2ZWwudXBwZXIoKSwgQ19TSUxWRVIpCiAgICAgICAgc2VsZi5fZGlzcGxheS5hcHBlbmQoCiAgICAg"
    "ICAgICAgIGYnPHNwYW4gc3R5bGU9ImNvbG9yOntDX1RFWFRfRElNfTsiPlt7dGltZXN0YW1wfV08L3NwYW4+ICcKICAgICAgICAg"
    "ICAgZic8c3BhbiBzdHlsZT0iY29sb3I6e2NvbG9yfTsiPnttZXNzYWdlfTwvc3Bhbj4nCiAgICAgICAgKQogICAgICAgIHNlbGYu"
    "X2Rpc3BsYXkudmVydGljYWxTY3JvbGxCYXIoKS5zZXRWYWx1ZSgKICAgICAgICAgICAgc2VsZi5fZGlzcGxheS52ZXJ0aWNhbFNj"
    "cm9sbEJhcigpLm1heGltdW0oKQogICAgICAgICkKCiAgICBkZWYgbG9nX21hbnkoc2VsZiwgbWVzc2FnZXM6IGxpc3Rbc3RyXSwg"
    "bGV2ZWw6IHN0ciA9ICJJTkZPIikgLT4gTm9uZToKICAgICAgICBmb3IgbXNnIGluIG1lc3NhZ2VzOgogICAgICAgICAgICBsdmwg"
    "PSBsZXZlbAogICAgICAgICAgICBpZiAi4pyTIiBpbiBtc2c6ICAgIGx2bCA9ICJPSyIKICAgICAgICAgICAgZWxpZiAi4pyXIiBp"
    "biBtc2c6ICBsdmwgPSAiV0FSTiIKICAgICAgICAgICAgZWxpZiAiRVJST1IiIGluIG1zZy51cHBlcigpOiBsdmwgPSAiRVJST1Ii"
    "CiAgICAgICAgICAgIHNlbGYubG9nKG1zZywgbHZsKQoKICAgIGRlZiBjbGVhcihzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYu"
    "X2Rpc3BsYXkuY2xlYXIoKQoKCiMg4pSA4pSAIExFU1NPTlMgVEFCIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgApjbGFzcyBMZXNzb25zVGFiKFFXaWRnZXQpOgogICAgIiIiCiAgICBMU0wgRm9yYmlkZGVuIFJ1bGVzZXQgYW5kIGNvZGUg"
    "bGVzc29ucyBicm93c2VyLgogICAgQWRkLCB2aWV3LCBzZWFyY2gsIGRlbGV0ZSBsZXNzb25zLgogICAgIiIiCgogICAgZGVmIF9f"
    "aW5pdF9fKHNlbGYsIGRiOiAiTGVzc29uc0xlYXJuZWREQiIsIHBhcmVudD1Ob25lKToKICAgICAgICBzdXBlcigpLl9faW5pdF9f"
    "KHBhcmVudCkKICAgICAgICBzZWxmLl9kYiA9IGRiCiAgICAgICAgc2VsZi5fc2V0dXBfdWkoKQogICAgICAgIHNlbGYucmVmcmVz"
    "aCgpCgogICAgZGVmIF9zZXR1cF91aShzZWxmKSAtPiBOb25lOgogICAgICAgIHJvb3QgPSBRVkJveExheW91dChzZWxmKQogICAg"
    "ICAgIHJvb3Quc2V0Q29udGVudHNNYXJnaW5zKDQsIDQsIDQsIDQpCiAgICAgICAgcm9vdC5zZXRTcGFjaW5nKDQpCgogICAgICAg"
    "ICMgRmlsdGVyIGJhcgogICAgICAgIGZpbHRlcl9yb3cgPSBRSEJveExheW91dCgpCiAgICAgICAgc2VsZi5fc2VhcmNoID0gUUxp"
    "bmVFZGl0KCkKICAgICAgICBzZWxmLl9zZWFyY2guc2V0UGxhY2Vob2xkZXJUZXh0KCJTZWFyY2ggbGVzc29ucy4uLiIpCiAgICAg"
    "ICAgc2VsZi5fbGFuZ19maWx0ZXIgPSBRQ29tYm9Cb3goKQogICAgICAgIHNlbGYuX2xhbmdfZmlsdGVyLmFkZEl0ZW1zKFsiQWxs"
    "IiwgIkxTTCIsICJQeXRob24iLCAiUHlTaWRlNiIsCiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAiSmF2YVNj"
    "cmlwdCIsICJPdGhlciJdKQogICAgICAgIHNlbGYuX3NlYXJjaC50ZXh0Q2hhbmdlZC5jb25uZWN0KHNlbGYucmVmcmVzaCkKICAg"
    "ICAgICBzZWxmLl9sYW5nX2ZpbHRlci5jdXJyZW50VGV4dENoYW5nZWQuY29ubmVjdChzZWxmLnJlZnJlc2gpCiAgICAgICAgZmls"
    "dGVyX3Jvdy5hZGRXaWRnZXQoUUxhYmVsKCJTZWFyY2g6IikpCiAgICAgICAgZmlsdGVyX3Jvdy5hZGRXaWRnZXQoc2VsZi5fc2Vh"
    "cmNoLCAxKQogICAgICAgIGZpbHRlcl9yb3cuYWRkV2lkZ2V0KFFMYWJlbCgiTGFuZ3VhZ2U6IikpCiAgICAgICAgZmlsdGVyX3Jv"
    "dy5hZGRXaWRnZXQoc2VsZi5fbGFuZ19maWx0ZXIpCiAgICAgICAgcm9vdC5hZGRMYXlvdXQoZmlsdGVyX3JvdykKCiAgICAgICAg"
    "YnRuX2JhciA9IFFIQm94TGF5b3V0KCkKICAgICAgICBidG5fYWRkID0gX2dvdGhpY19idG4oIuKcpiBBZGQgTGVzc29uIikKICAg"
    "ICAgICBidG5fZGVsID0gX2dvdGhpY19idG4oIuKclyBEZWxldGUiKQogICAgICAgIGJ0bl9hZGQuY2xpY2tlZC5jb25uZWN0KHNl"
    "bGYuX2RvX2FkZCkKICAgICAgICBidG5fZGVsLmNsaWNrZWQuY29ubmVjdChzZWxmLl9kb19kZWxldGUpCiAgICAgICAgYnRuX2Jh"
    "ci5hZGRXaWRnZXQoYnRuX2FkZCkKICAgICAgICBidG5fYmFyLmFkZFdpZGdldChidG5fZGVsKQogICAgICAgIGJ0bl9iYXIuYWRk"
    "U3RyZXRjaCgpCiAgICAgICAgcm9vdC5hZGRMYXlvdXQoYnRuX2JhcikKCiAgICAgICAgc2VsZi5fdGFibGUgPSBRVGFibGVXaWRn"
    "ZXQoMCwgNCkKICAgICAgICBzZWxmLl90YWJsZS5zZXRIb3Jpem9udGFsSGVhZGVyTGFiZWxzKAogICAgICAgICAgICBbIkxhbmd1"
    "YWdlIiwgIlJlZmVyZW5jZSBLZXkiLCAiU3VtbWFyeSIsICJFbnZpcm9ubWVudCJdCiAgICAgICAgKQogICAgICAgIHNlbGYuX3Rh"
    "YmxlLmhvcml6b250YWxIZWFkZXIoKS5zZXRTZWN0aW9uUmVzaXplTW9kZSgKICAgICAgICAgICAgMiwgUUhlYWRlclZpZXcuUmVz"
    "aXplTW9kZS5TdHJldGNoKQogICAgICAgIHNlbGYuX3RhYmxlLnNldFNlbGVjdGlvbkJlaGF2aW9yKAogICAgICAgICAgICBRQWJz"
    "dHJhY3RJdGVtVmlldy5TZWxlY3Rpb25CZWhhdmlvci5TZWxlY3RSb3dzKQogICAgICAgIHNlbGYuX3RhYmxlLnNldEFsdGVybmF0"
    "aW5nUm93Q29sb3JzKFRydWUpCiAgICAgICAgc2VsZi5fdGFibGUuc2V0U3R5bGVTaGVldChfZ290aGljX3RhYmxlX3N0eWxlKCkp"
    "CiAgICAgICAgc2VsZi5fdGFibGUuaXRlbVNlbGVjdGlvbkNoYW5nZWQuY29ubmVjdChzZWxmLl9vbl9zZWxlY3QpCgogICAgICAg"
    "ICMgVXNlIHNwbGl0dGVyIGJldHdlZW4gdGFibGUgYW5kIGRldGFpbAogICAgICAgIHNwbGl0dGVyID0gUVNwbGl0dGVyKFF0Lk9y"
    "aWVudGF0aW9uLlZlcnRpY2FsKQogICAgICAgIHNwbGl0dGVyLmFkZFdpZGdldChzZWxmLl90YWJsZSkKCiAgICAgICAgIyBEZXRh"
    "aWwgcGFuZWwKICAgICAgICBkZXRhaWxfd2lkZ2V0ID0gUVdpZGdldCgpCiAgICAgICAgZGV0YWlsX2xheW91dCA9IFFWQm94TGF5"
    "b3V0KGRldGFpbF93aWRnZXQpCiAgICAgICAgZGV0YWlsX2xheW91dC5zZXRDb250ZW50c01hcmdpbnMoMCwgNCwgMCwgMCkKICAg"
    "ICAgICBkZXRhaWxfbGF5b3V0LnNldFNwYWNpbmcoMikKCiAgICAgICAgZGV0YWlsX2hlYWRlciA9IFFIQm94TGF5b3V0KCkKICAg"
    "ICAgICBkZXRhaWxfaGVhZGVyLmFkZFdpZGdldChfc2VjdGlvbl9sYmwoIuKdpyBGVUxMIFJVTEUiKSkKICAgICAgICBkZXRhaWxf"
    "aGVhZGVyLmFkZFN0cmV0Y2goKQogICAgICAgIHNlbGYuX2J0bl9lZGl0X3J1bGUgPSBfZ290aGljX2J0bigiRWRpdCIpCiAgICAg"
    "ICAgc2VsZi5fYnRuX2VkaXRfcnVsZS5zZXRGaXhlZFdpZHRoKDUwKQogICAgICAgIHNlbGYuX2J0bl9lZGl0X3J1bGUuc2V0Q2hl"
    "Y2thYmxlKFRydWUpCiAgICAgICAgc2VsZi5fYnRuX2VkaXRfcnVsZS50b2dnbGVkLmNvbm5lY3Qoc2VsZi5fdG9nZ2xlX2VkaXRf"
    "bW9kZSkKICAgICAgICBzZWxmLl9idG5fc2F2ZV9ydWxlID0gX2dvdGhpY19idG4oIlNhdmUiKQogICAgICAgIHNlbGYuX2J0bl9z"
    "YXZlX3J1bGUuc2V0Rml4ZWRXaWR0aCg1MCkKICAgICAgICBzZWxmLl9idG5fc2F2ZV9ydWxlLnNldFZpc2libGUoRmFsc2UpCiAg"
    "ICAgICAgc2VsZi5fYnRuX3NhdmVfcnVsZS5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fc2F2ZV9ydWxlX2VkaXQpCiAgICAgICAgZGV0"
    "YWlsX2hlYWRlci5hZGRXaWRnZXQoc2VsZi5fYnRuX2VkaXRfcnVsZSkKICAgICAgICBkZXRhaWxfaGVhZGVyLmFkZFdpZGdldChz"
    "ZWxmLl9idG5fc2F2ZV9ydWxlKQogICAgICAgIGRldGFpbF9sYXlvdXQuYWRkTGF5b3V0KGRldGFpbF9oZWFkZXIpCgogICAgICAg"
    "IHNlbGYuX2RldGFpbCA9IFFUZXh0RWRpdCgpCiAgICAgICAgc2VsZi5fZGV0YWlsLnNldFJlYWRPbmx5KFRydWUpCiAgICAgICAg"
    "c2VsZi5fZGV0YWlsLnNldE1pbmltdW1IZWlnaHQoMTIwKQogICAgICAgIHNlbGYuX2RldGFpbC5zZXRTdHlsZVNoZWV0KAogICAg"
    "ICAgICAgICBmImJhY2tncm91bmQ6IHtDX0JHM307IGNvbG9yOiB7Q19HT0xEfTsgIgogICAgICAgICAgICBmImJvcmRlcjogMXB4"
    "IHNvbGlkIHtDX0JPUkRFUn07ICIKICAgICAgICAgICAgZiJmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyBmb250LXNp"
    "emU6IDExcHg7IHBhZGRpbmc6IDRweDsiCiAgICAgICAgKQogICAgICAgIGRldGFpbF9sYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX2Rl"
    "dGFpbCkKICAgICAgICBzcGxpdHRlci5hZGRXaWRnZXQoZGV0YWlsX3dpZGdldCkKICAgICAgICBzcGxpdHRlci5zZXRTaXplcyhb"
    "MzAwLCAxODBdKQogICAgICAgIHJvb3QuYWRkV2lkZ2V0KHNwbGl0dGVyLCAxKQoKICAgICAgICBzZWxmLl9yZWNvcmRzOiBsaXN0"
    "W2RpY3RdID0gW10KICAgICAgICBzZWxmLl9lZGl0aW5nX3JvdzogaW50ID0gLTEKCiAgICBkZWYgcmVmcmVzaChzZWxmKSAtPiBO"
    "b25lOgogICAgICAgIHEgICAgPSBzZWxmLl9zZWFyY2gudGV4dCgpCiAgICAgICAgbGFuZyA9IHNlbGYuX2xhbmdfZmlsdGVyLmN1"
    "cnJlbnRUZXh0KCkKICAgICAgICBsYW5nID0gIiIgaWYgbGFuZyA9PSAiQWxsIiBlbHNlIGxhbmcKICAgICAgICBzZWxmLl9yZWNv"
    "cmRzID0gc2VsZi5fZGIuc2VhcmNoKHF1ZXJ5PXEsIGxhbmd1YWdlPWxhbmcpCiAgICAgICAgc2VsZi5fdGFibGUuc2V0Um93Q291"
    "bnQoMCkKICAgICAgICBmb3IgcmVjIGluIHNlbGYuX3JlY29yZHM6CiAgICAgICAgICAgIHIgPSBzZWxmLl90YWJsZS5yb3dDb3Vu"
    "dCgpCiAgICAgICAgICAgIHNlbGYuX3RhYmxlLmluc2VydFJvdyhyKQogICAgICAgICAgICBzZWxmLl90YWJsZS5zZXRJdGVtKHIs"
    "IDAsCiAgICAgICAgICAgICAgICBRVGFibGVXaWRnZXRJdGVtKHJlYy5nZXQoImxhbmd1YWdlIiwiIikpKQogICAgICAgICAgICBz"
    "ZWxmLl90YWJsZS5zZXRJdGVtKHIsIDEsCiAgICAgICAgICAgICAgICBRVGFibGVXaWRnZXRJdGVtKHJlYy5nZXQoInJlZmVyZW5j"
    "ZV9rZXkiLCIiKSkpCiAgICAgICAgICAgIHNlbGYuX3RhYmxlLnNldEl0ZW0ociwgMiwKICAgICAgICAgICAgICAgIFFUYWJsZVdp"
    "ZGdldEl0ZW0ocmVjLmdldCgic3VtbWFyeSIsIiIpKSkKICAgICAgICAgICAgc2VsZi5fdGFibGUuc2V0SXRlbShyLCAzLAogICAg"
    "ICAgICAgICAgICAgUVRhYmxlV2lkZ2V0SXRlbShyZWMuZ2V0KCJlbnZpcm9ubWVudCIsIiIpKSkKCiAgICBkZWYgX29uX3NlbGVj"
    "dChzZWxmKSAtPiBOb25lOgogICAgICAgIHJvdyA9IHNlbGYuX3RhYmxlLmN1cnJlbnRSb3coKQogICAgICAgIHNlbGYuX2VkaXRp"
    "bmdfcm93ID0gcm93CiAgICAgICAgaWYgMCA8PSByb3cgPCBsZW4oc2VsZi5fcmVjb3Jkcyk6CiAgICAgICAgICAgIHJlYyA9IHNl"
    "bGYuX3JlY29yZHNbcm93XQogICAgICAgICAgICBzZWxmLl9kZXRhaWwuc2V0UGxhaW5UZXh0KAogICAgICAgICAgICAgICAgcmVj"
    "LmdldCgiZnVsbF9ydWxlIiwiIikgKyAiXG5cbiIgKwogICAgICAgICAgICAgICAgKCJSZXNvbHV0aW9uOiAiICsgcmVjLmdldCgi"
    "cmVzb2x1dGlvbiIsIiIpIGlmIHJlYy5nZXQoInJlc29sdXRpb24iKSBlbHNlICIiKQogICAgICAgICAgICApCiAgICAgICAgICAg"
    "ICMgUmVzZXQgZWRpdCBtb2RlIG9uIG5ldyBzZWxlY3Rpb24KICAgICAgICAgICAgc2VsZi5fYnRuX2VkaXRfcnVsZS5zZXRDaGVj"
    "a2VkKEZhbHNlKQoKICAgIGRlZiBfdG9nZ2xlX2VkaXRfbW9kZShzZWxmLCBlZGl0aW5nOiBib29sKSAtPiBOb25lOgogICAgICAg"
    "IHNlbGYuX2RldGFpbC5zZXRSZWFkT25seShub3QgZWRpdGluZykKICAgICAgICBzZWxmLl9idG5fc2F2ZV9ydWxlLnNldFZpc2li"
    "bGUoZWRpdGluZykKICAgICAgICBzZWxmLl9idG5fZWRpdF9ydWxlLnNldFRleHQoIkNhbmNlbCIgaWYgZWRpdGluZyBlbHNlICJF"
    "ZGl0IikKICAgICAgICBpZiBlZGl0aW5nOgogICAgICAgICAgICBzZWxmLl9kZXRhaWwuc2V0U3R5bGVTaGVldCgKICAgICAgICAg"
    "ICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkcyfTsgY29sb3I6IHtDX0dPTER9OyAiCiAgICAgICAgICAgICAgICBmImJvcmRlcjog"
    "MXB4IHNvbGlkIHtDX0dPTERfRElNfTsgIgogICAgICAgICAgICAgICAgZiJmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlm"
    "OyBmb250LXNpemU6IDExcHg7IHBhZGRpbmc6IDRweDsiCiAgICAgICAgICAgICkKICAgICAgICBlbHNlOgogICAgICAgICAgICBz"
    "ZWxmLl9kZXRhaWwuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkczfTsgY29sb3I6IHtD"
    "X0dPTER9OyAiCiAgICAgICAgICAgICAgICBmImJvcmRlcjogMXB4IHNvbGlkIHtDX0JPUkRFUn07ICIKICAgICAgICAgICAgICAg"
    "IGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsgZm9udC1zaXplOiAxMXB4OyBwYWRkaW5nOiA0cHg7IgogICAgICAg"
    "ICAgICApCiAgICAgICAgICAgICMgUmVsb2FkIG9yaWdpbmFsIGNvbnRlbnQgb24gY2FuY2VsCiAgICAgICAgICAgIHNlbGYuX29u"
    "X3NlbGVjdCgpCgogICAgZGVmIF9zYXZlX3J1bGVfZWRpdChzZWxmKSAtPiBOb25lOgogICAgICAgIHJvdyA9IHNlbGYuX2VkaXRp"
    "bmdfcm93CiAgICAgICAgaWYgMCA8PSByb3cgPCBsZW4oc2VsZi5fcmVjb3Jkcyk6CiAgICAgICAgICAgIHRleHQgPSBzZWxmLl9k"
    "ZXRhaWwudG9QbGFpblRleHQoKS5zdHJpcCgpCiAgICAgICAgICAgICMgU3BsaXQgcmVzb2x1dGlvbiBiYWNrIG91dCBpZiBwcmVz"
    "ZW50CiAgICAgICAgICAgIGlmICJcblxuUmVzb2x1dGlvbjogIiBpbiB0ZXh0OgogICAgICAgICAgICAgICAgcGFydHMgPSB0ZXh0"
    "LnNwbGl0KCJcblxuUmVzb2x1dGlvbjogIiwgMSkKICAgICAgICAgICAgICAgIGZ1bGxfcnVsZSAgPSBwYXJ0c1swXS5zdHJpcCgp"
    "CiAgICAgICAgICAgICAgICByZXNvbHV0aW9uID0gcGFydHNbMV0uc3RyaXAoKQogICAgICAgICAgICBlbHNlOgogICAgICAgICAg"
    "ICAgICAgZnVsbF9ydWxlICA9IHRleHQKICAgICAgICAgICAgICAgIHJlc29sdXRpb24gPSBzZWxmLl9yZWNvcmRzW3Jvd10uZ2V0"
    "KCJyZXNvbHV0aW9uIiwgIiIpCiAgICAgICAgICAgIHNlbGYuX3JlY29yZHNbcm93XVsiZnVsbF9ydWxlIl0gID0gZnVsbF9ydWxl"
    "CiAgICAgICAgICAgIHNlbGYuX3JlY29yZHNbcm93XVsicmVzb2x1dGlvbiJdID0gcmVzb2x1dGlvbgogICAgICAgICAgICB3cml0"
    "ZV9qc29ubChzZWxmLl9kYi5fcGF0aCwgc2VsZi5fcmVjb3JkcykKICAgICAgICAgICAgc2VsZi5fYnRuX2VkaXRfcnVsZS5zZXRD"
    "aGVja2VkKEZhbHNlKQogICAgICAgICAgICBzZWxmLnJlZnJlc2goKQoKICAgIGRlZiBfZG9fYWRkKHNlbGYpIC0+IE5vbmU6CiAg"
    "ICAgICAgZGxnID0gUURpYWxvZyhzZWxmKQogICAgICAgIGRsZy5zZXRXaW5kb3dUaXRsZSgiQWRkIExlc3NvbiIpCiAgICAgICAg"
    "ZGxnLnNldFN0eWxlU2hlZXQoZiJiYWNrZ3JvdW5kOiB7Q19CRzJ9OyBjb2xvcjoge0NfR09MRH07IikKICAgICAgICBkbGcucmVz"
    "aXplKDUwMCwgNDAwKQogICAgICAgIGZvcm0gPSBRRm9ybUxheW91dChkbGcpCiAgICAgICAgZW52ICA9IFFMaW5lRWRpdCgiTFNM"
    "IikKICAgICAgICBsYW5nID0gUUxpbmVFZGl0KCJMU0wiKQogICAgICAgIHJlZiAgPSBRTGluZUVkaXQoKQogICAgICAgIHN1bW0g"
    "PSBRTGluZUVkaXQoKQogICAgICAgIHJ1bGUgPSBRVGV4dEVkaXQoKQogICAgICAgIHJ1bGUuc2V0TWF4aW11bUhlaWdodCgxMDAp"
    "CiAgICAgICAgcmVzICA9IFFMaW5lRWRpdCgpCiAgICAgICAgbGluayA9IFFMaW5lRWRpdCgpCiAgICAgICAgZm9yIGxhYmVsLCB3"
    "IGluIFsKICAgICAgICAgICAgKCJFbnZpcm9ubWVudDoiLCBlbnYpLCAoIkxhbmd1YWdlOiIsIGxhbmcpLAogICAgICAgICAgICAo"
    "IlJlZmVyZW5jZSBLZXk6IiwgcmVmKSwgKCJTdW1tYXJ5OiIsIHN1bW0pLAogICAgICAgICAgICAoIkZ1bGwgUnVsZToiLCBydWxl"
    "KSwgKCJSZXNvbHV0aW9uOiIsIHJlcyksCiAgICAgICAgICAgICgiTGluazoiLCBsaW5rKSwKICAgICAgICBdOgogICAgICAgICAg"
    "ICBmb3JtLmFkZFJvdyhsYWJlbCwgdykKICAgICAgICBidG5zID0gUUhCb3hMYXlvdXQoKQogICAgICAgIG9rID0gX2dvdGhpY19i"
    "dG4oIlNhdmUiKTsgY3ggPSBfZ290aGljX2J0bigiQ2FuY2VsIikKICAgICAgICBvay5jbGlja2VkLmNvbm5lY3QoZGxnLmFjY2Vw"
    "dCk7IGN4LmNsaWNrZWQuY29ubmVjdChkbGcucmVqZWN0KQogICAgICAgIGJ0bnMuYWRkV2lkZ2V0KG9rKTsgYnRucy5hZGRXaWRn"
    "ZXQoY3gpCiAgICAgICAgZm9ybS5hZGRSb3coYnRucykKICAgICAgICBpZiBkbGcuZXhlYygpID09IFFEaWFsb2cuRGlhbG9nQ29k"
    "ZS5BY2NlcHRlZDoKICAgICAgICAgICAgc2VsZi5fZGIuYWRkKAogICAgICAgICAgICAgICAgZW52aXJvbm1lbnQ9ZW52LnRleHQo"
    "KS5zdHJpcCgpLAogICAgICAgICAgICAgICAgbGFuZ3VhZ2U9bGFuZy50ZXh0KCkuc3RyaXAoKSwKICAgICAgICAgICAgICAgIHJl"
    "ZmVyZW5jZV9rZXk9cmVmLnRleHQoKS5zdHJpcCgpLAogICAgICAgICAgICAgICAgc3VtbWFyeT1zdW1tLnRleHQoKS5zdHJpcCgp"
    "LAogICAgICAgICAgICAgICAgZnVsbF9ydWxlPXJ1bGUudG9QbGFpblRleHQoKS5zdHJpcCgpLAogICAgICAgICAgICAgICAgcmVz"
    "b2x1dGlvbj1yZXMudGV4dCgpLnN0cmlwKCksCiAgICAgICAgICAgICAgICBsaW5rPWxpbmsudGV4dCgpLnN0cmlwKCksCiAgICAg"
    "ICAgICAgICkKICAgICAgICAgICAgc2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX2RvX2RlbGV0ZShzZWxmKSAtPiBOb25lOgogICAg"
    "ICAgIHJvdyA9IHNlbGYuX3RhYmxlLmN1cnJlbnRSb3coKQogICAgICAgIGlmIDAgPD0gcm93IDwgbGVuKHNlbGYuX3JlY29yZHMp"
    "OgogICAgICAgICAgICByZWNfaWQgPSBzZWxmLl9yZWNvcmRzW3Jvd10uZ2V0KCJpZCIsIiIpCiAgICAgICAgICAgIHJlcGx5ID0g"
    "UU1lc3NhZ2VCb3gucXVlc3Rpb24oCiAgICAgICAgICAgICAgICBzZWxmLCAiRGVsZXRlIExlc3NvbiIsCiAgICAgICAgICAgICAg"
    "ICAiRGVsZXRlIHRoaXMgbGVzc29uPyBDYW5ub3QgYmUgdW5kb25lLiIsCiAgICAgICAgICAgICAgICBRTWVzc2FnZUJveC5TdGFu"
    "ZGFyZEJ1dHRvbi5ZZXMgfCBRTWVzc2FnZUJveC5TdGFuZGFyZEJ1dHRvbi5ObwogICAgICAgICAgICApCiAgICAgICAgICAgIGlm"
    "IHJlcGx5ID09IFFNZXNzYWdlQm94LlN0YW5kYXJkQnV0dG9uLlllczoKICAgICAgICAgICAgICAgIHNlbGYuX2RiLmRlbGV0ZShy"
    "ZWNfaWQpCiAgICAgICAgICAgICAgICBzZWxmLnJlZnJlc2goKQoKCiMg4pSA4pSAIE1PRFVMRSBUUkFDS0VSIFRBQiDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIAKY2xhc3MgTW9kdWxlVHJhY2tlclRhYihRV2lkZ2V0KToKICAgICIiIgogICAgUGVyc29uYWwgbW9kdWxlIHBp"
    "cGVsaW5lIHRyYWNrZXIuCiAgICBUcmFjayBwbGFubmVkL2luLXByb2dyZXNzL2J1aWx0IG1vZHVsZXMgYXMgdGhleSBhcmUgZGVz"
    "aWduZWQuCiAgICBFYWNoIG1vZHVsZSBoYXM6IE5hbWUsIFN0YXR1cywgRGVzY3JpcHRpb24sIE5vdGVzLgogICAgRXhwb3J0IHRv"
    "IFRYVCBmb3IgcGFzdGluZyBpbnRvIHNlc3Npb25zLgogICAgSW1wb3J0OiBwYXN0ZSBhIGZpbmFsaXplZCBzcGVjLCBpdCBwYXJz"
    "ZXMgbmFtZSBhbmQgZGV0YWlscy4KICAgIFRoaXMgaXMgYSBkZXNpZ24gbm90ZWJvb2sg4oCUIG5vdCBjb25uZWN0ZWQgdG8gZGVj"
    "a19idWlsZGVyJ3MgTU9EVUxFIHJlZ2lzdHJ5LgogICAgIiIiCgogICAgU1RBVFVTRVMgPSBbIklkZWEiLCAiRGVzaWduaW5nIiwg"
    "IlJlYWR5IHRvIEJ1aWxkIiwgIlBhcnRpYWwiLCAiQnVpbHQiXQoKICAgIGRlZiBfX2luaXRfXyhzZWxmLCBwYXJlbnQ9Tm9uZSk6"
    "CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2VsZi5fcGF0aCA9IGNmZ19wYXRoKCJtZW1vcmllcyIp"
    "IC8gIm1vZHVsZV90cmFja2VyLmpzb25sIgogICAgICAgIHNlbGYuX3JlY29yZHM6IGxpc3RbZGljdF0gPSBbXQogICAgICAgIHNl"
    "bGYuX3NldHVwX3VpKCkKICAgICAgICBzZWxmLnJlZnJlc2goKQoKICAgIGRlZiBfc2V0dXBfdWkoc2VsZikgLT4gTm9uZToKICAg"
    "ICAgICByb290ID0gUVZCb3hMYXlvdXQoc2VsZikKICAgICAgICByb290LnNldENvbnRlbnRzTWFyZ2lucyg0LCA0LCA0LCA0KQog"
    "ICAgICAgIHJvb3Quc2V0U3BhY2luZyg0KQoKICAgICAgICAjIEJ1dHRvbiBiYXIKICAgICAgICBidG5fYmFyID0gUUhCb3hMYXlv"
    "dXQoKQogICAgICAgIHNlbGYuX2J0bl9hZGQgICAgPSBfZ290aGljX2J0bigiQWRkIE1vZHVsZSIpCiAgICAgICAgc2VsZi5fYnRu"
    "X2VkaXQgICA9IF9nb3RoaWNfYnRuKCJFZGl0IikKICAgICAgICBzZWxmLl9idG5fZGVsZXRlID0gX2dvdGhpY19idG4oIkRlbGV0"
    "ZSIpCiAgICAgICAgc2VsZi5fYnRuX2V4cG9ydCA9IF9nb3RoaWNfYnRuKCJFeHBvcnQgVFhUIikKICAgICAgICBzZWxmLl9idG5f"
    "aW1wb3J0ID0gX2dvdGhpY19idG4oIkltcG9ydCBTcGVjIikKICAgICAgICBmb3IgYiBpbiAoc2VsZi5fYnRuX2FkZCwgc2VsZi5f"
    "YnRuX2VkaXQsIHNlbGYuX2J0bl9kZWxldGUsCiAgICAgICAgICAgICAgICAgIHNlbGYuX2J0bl9leHBvcnQsIHNlbGYuX2J0bl9p"
    "bXBvcnQpOgogICAgICAgICAgICBiLnNldE1pbmltdW1XaWR0aCg4MCkKICAgICAgICAgICAgYi5zZXRNaW5pbXVtSGVpZ2h0KDI2"
    "KQogICAgICAgICAgICBidG5fYmFyLmFkZFdpZGdldChiKQogICAgICAgIGJ0bl9iYXIuYWRkU3RyZXRjaCgpCiAgICAgICAgcm9v"
    "dC5hZGRMYXlvdXQoYnRuX2JhcikKCiAgICAgICAgc2VsZi5fYnRuX2FkZC5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fZG9fYWRkKQog"
    "ICAgICAgIHNlbGYuX2J0bl9lZGl0LmNsaWNrZWQuY29ubmVjdChzZWxmLl9kb19lZGl0KQogICAgICAgIHNlbGYuX2J0bl9kZWxl"
    "dGUuY2xpY2tlZC5jb25uZWN0KHNlbGYuX2RvX2RlbGV0ZSkKICAgICAgICBzZWxmLl9idG5fZXhwb3J0LmNsaWNrZWQuY29ubmVj"
    "dChzZWxmLl9kb19leHBvcnQpCiAgICAgICAgc2VsZi5fYnRuX2ltcG9ydC5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fZG9faW1wb3J0"
    "KQoKICAgICAgICAjIFRhYmxlCiAgICAgICAgc2VsZi5fdGFibGUgPSBRVGFibGVXaWRnZXQoMCwgMykKICAgICAgICBzZWxmLl90"
    "YWJsZS5zZXRIb3Jpem9udGFsSGVhZGVyTGFiZWxzKFsiTW9kdWxlIE5hbWUiLCAiU3RhdHVzIiwgIkRlc2NyaXB0aW9uIl0pCiAg"
    "ICAgICAgaGggPSBzZWxmLl90YWJsZS5ob3Jpem9udGFsSGVhZGVyKCkKICAgICAgICBoaC5zZXRTZWN0aW9uUmVzaXplTW9kZSgw"
    "LCBRSGVhZGVyVmlldy5SZXNpemVNb2RlLkZpeGVkKQogICAgICAgIHNlbGYuX3RhYmxlLnNldENvbHVtbldpZHRoKDAsIDE2MCkK"
    "ICAgICAgICBoaC5zZXRTZWN0aW9uUmVzaXplTW9kZSgxLCBRSGVhZGVyVmlldy5SZXNpemVNb2RlLkZpeGVkKQogICAgICAgIHNl"
    "bGYuX3RhYmxlLnNldENvbHVtbldpZHRoKDEsIDEwMCkKICAgICAgICBoaC5zZXRTZWN0aW9uUmVzaXplTW9kZSgyLCBRSGVhZGVy"
    "Vmlldy5SZXNpemVNb2RlLlN0cmV0Y2gpCiAgICAgICAgc2VsZi5fdGFibGUuc2V0U2VsZWN0aW9uQmVoYXZpb3IoCiAgICAgICAg"
    "ICAgIFFBYnN0cmFjdEl0ZW1WaWV3LlNlbGVjdGlvbkJlaGF2aW9yLlNlbGVjdFJvd3MpCiAgICAgICAgc2VsZi5fdGFibGUuc2V0"
    "QWx0ZXJuYXRpbmdSb3dDb2xvcnMoVHJ1ZSkKICAgICAgICBzZWxmLl90YWJsZS5zZXRTdHlsZVNoZWV0KF9nb3RoaWNfdGFibGVf"
    "c3R5bGUoKSkKICAgICAgICBzZWxmLl90YWJsZS5pdGVtU2VsZWN0aW9uQ2hhbmdlZC5jb25uZWN0KHNlbGYuX29uX3NlbGVjdCkK"
    "CiAgICAgICAgIyBTcGxpdHRlcgogICAgICAgIHNwbGl0dGVyID0gUVNwbGl0dGVyKFF0Lk9yaWVudGF0aW9uLlZlcnRpY2FsKQog"
    "ICAgICAgIHNwbGl0dGVyLmFkZFdpZGdldChzZWxmLl90YWJsZSkKCiAgICAgICAgIyBOb3RlcyBwYW5lbAogICAgICAgIG5vdGVz"
    "X3dpZGdldCA9IFFXaWRnZXQoKQogICAgICAgIG5vdGVzX2xheW91dCA9IFFWQm94TGF5b3V0KG5vdGVzX3dpZGdldCkKICAgICAg"
    "ICBub3Rlc19sYXlvdXQuc2V0Q29udGVudHNNYXJnaW5zKDAsIDQsIDAsIDApCiAgICAgICAgbm90ZXNfbGF5b3V0LnNldFNwYWNp"
    "bmcoMikKICAgICAgICBub3Rlc19sYXlvdXQuYWRkV2lkZ2V0KF9zZWN0aW9uX2xibCgi4p2nIE5PVEVTIikpCiAgICAgICAgc2Vs"
    "Zi5fbm90ZXNfZGlzcGxheSA9IFFUZXh0RWRpdCgpCiAgICAgICAgc2VsZi5fbm90ZXNfZGlzcGxheS5zZXRSZWFkT25seShUcnVl"
    "KQogICAgICAgIHNlbGYuX25vdGVzX2Rpc3BsYXkuc2V0TWluaW11bUhlaWdodCgxMjApCiAgICAgICAgc2VsZi5fbm90ZXNfZGlz"
    "cGxheS5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX0JHM307IGNvbG9yOiB7Q19HT0xEfTsgIgog"
    "ICAgICAgICAgICBmImJvcmRlcjogMXB4IHNvbGlkIHtDX0JPUkRFUn07ICIKICAgICAgICAgICAgZiJmb250LWZhbWlseToge0RF"
    "Q0tfRk9OVH0sIHNlcmlmOyBmb250LXNpemU6IDExcHg7IHBhZGRpbmc6IDRweDsiCiAgICAgICAgKQogICAgICAgIG5vdGVzX2xh"
    "eW91dC5hZGRXaWRnZXQoc2VsZi5fbm90ZXNfZGlzcGxheSkKICAgICAgICBzcGxpdHRlci5hZGRXaWRnZXQobm90ZXNfd2lkZ2V0"
    "KQogICAgICAgIHNwbGl0dGVyLnNldFNpemVzKFsyNTAsIDE1MF0pCiAgICAgICAgcm9vdC5hZGRXaWRnZXQoc3BsaXR0ZXIsIDEp"
    "CgogICAgICAgICMgQ291bnQgbGFiZWwKICAgICAgICBzZWxmLl9jb3VudF9sYmwgPSBRTGFiZWwoIiIpCiAgICAgICAgc2VsZi5f"
    "Y291bnRfbGJsLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX1RFWFRfRElNfTsgZm9udC1zaXplOiA5cHg7"
    "IGZvbnQtZmFtaWx5OiB7REVDS19GT05UfSwgc2VyaWY7IgogICAgICAgICkKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl9j"
    "b3VudF9sYmwpCgogICAgZGVmIHJlZnJlc2goc2VsZikgLT4gTm9uZToKICAgICAgICBzZWxmLl9yZWNvcmRzID0gcmVhZF9qc29u"
    "bChzZWxmLl9wYXRoKQogICAgICAgIHNlbGYuX3RhYmxlLnNldFJvd0NvdW50KDApCiAgICAgICAgZm9yIHJlYyBpbiBzZWxmLl9y"
    "ZWNvcmRzOgogICAgICAgICAgICByID0gc2VsZi5fdGFibGUucm93Q291bnQoKQogICAgICAgICAgICBzZWxmLl90YWJsZS5pbnNl"
    "cnRSb3cocikKICAgICAgICAgICAgc2VsZi5fdGFibGUuc2V0SXRlbShyLCAwLCBRVGFibGVXaWRnZXRJdGVtKHJlYy5nZXQoIm5h"
    "bWUiLCAiIikpKQogICAgICAgICAgICBzdGF0dXNfaXRlbSA9IFFUYWJsZVdpZGdldEl0ZW0ocmVjLmdldCgic3RhdHVzIiwgIklk"
    "ZWEiKSkKICAgICAgICAgICAgIyBDb2xvciBieSBzdGF0dXMKICAgICAgICAgICAgc3RhdHVzX2NvbG9ycyA9IHsKICAgICAgICAg"
    "ICAgICAgICJJZGVhIjogICAgICAgICAgICAgQ19URVhUX0RJTSwKICAgICAgICAgICAgICAgICJEZXNpZ25pbmciOiAgICAgICAg"
    "Q19HT0xEX0RJTSwKICAgICAgICAgICAgICAgICJSZWFkeSB0byBCdWlsZCI6ICAgQ19QVVJQTEUsCiAgICAgICAgICAgICAgICAi"
    "UGFydGlhbCI6ICAgICAgICAgICIjY2M4ODQ0IiwKICAgICAgICAgICAgICAgICJCdWlsdCI6ICAgICAgICAgICAgQ19HUkVFTiwK"
    "ICAgICAgICAgICAgfQogICAgICAgICAgICBzdGF0dXNfaXRlbS5zZXRGb3JlZ3JvdW5kKAogICAgICAgICAgICAgICAgUUNvbG9y"
    "KHN0YXR1c19jb2xvcnMuZ2V0KHJlYy5nZXQoInN0YXR1cyIsIklkZWEiKSwgQ19URVhUX0RJTSkpCiAgICAgICAgICAgICkKICAg"
    "ICAgICAgICAgc2VsZi5fdGFibGUuc2V0SXRlbShyLCAxLCBzdGF0dXNfaXRlbSkKICAgICAgICAgICAgc2VsZi5fdGFibGUuc2V0"
    "SXRlbShyLCAyLAogICAgICAgICAgICAgICAgUVRhYmxlV2lkZ2V0SXRlbShyZWMuZ2V0KCJkZXNjcmlwdGlvbiIsICIiKVs6ODBd"
    "KSkKICAgICAgICBjb3VudHMgPSB7fQogICAgICAgIGZvciByZWMgaW4gc2VsZi5fcmVjb3JkczoKICAgICAgICAgICAgcyA9IHJl"
    "Yy5nZXQoInN0YXR1cyIsICJJZGVhIikKICAgICAgICAgICAgY291bnRzW3NdID0gY291bnRzLmdldChzLCAwKSArIDEKICAgICAg"
    "ICBjb3VudF9zdHIgPSAiICAiLmpvaW4oZiJ7c306IHtufSIgZm9yIHMsIG4gaW4gY291bnRzLml0ZW1zKCkpCiAgICAgICAgc2Vs"
    "Zi5fY291bnRfbGJsLnNldFRleHQoCiAgICAgICAgICAgIGYiVG90YWw6IHtsZW4oc2VsZi5fcmVjb3Jkcyl9ICAge2NvdW50X3N0"
    "cn0iCiAgICAgICAgKQoKICAgIGRlZiBfb25fc2VsZWN0KHNlbGYpIC0+IE5vbmU6CiAgICAgICAgcm93ID0gc2VsZi5fdGFibGUu"
    "Y3VycmVudFJvdygpCiAgICAgICAgaWYgMCA8PSByb3cgPCBsZW4oc2VsZi5fcmVjb3Jkcyk6CiAgICAgICAgICAgIHJlYyA9IHNl"
    "bGYuX3JlY29yZHNbcm93XQogICAgICAgICAgICBzZWxmLl9ub3Rlc19kaXNwbGF5LnNldFBsYWluVGV4dChyZWMuZ2V0KCJub3Rl"
    "cyIsICIiKSkKCiAgICBkZWYgX2RvX2FkZChzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX29wZW5fZWRpdF9kaWFsb2coKQoK"
    "ICAgIGRlZiBfZG9fZWRpdChzZWxmKSAtPiBOb25lOgogICAgICAgIHJvdyA9IHNlbGYuX3RhYmxlLmN1cnJlbnRSb3coKQogICAg"
    "ICAgIGlmIDAgPD0gcm93IDwgbGVuKHNlbGYuX3JlY29yZHMpOgogICAgICAgICAgICBzZWxmLl9vcGVuX2VkaXRfZGlhbG9nKHNl"
    "bGYuX3JlY29yZHNbcm93XSwgcm93KQoKICAgIGRlZiBfb3Blbl9lZGl0X2RpYWxvZyhzZWxmLCByZWM6IGRpY3QgPSBOb25lLCBy"
    "b3c6IGludCA9IC0xKSAtPiBOb25lOgogICAgICAgIGRsZyA9IFFEaWFsb2coc2VsZikKICAgICAgICBkbGcuc2V0V2luZG93VGl0"
    "bGUoIk1vZHVsZSIgaWYgbm90IHJlYyBlbHNlIGYiRWRpdDoge3JlYy5nZXQoJ25hbWUnLCcnKX0iKQogICAgICAgIGRsZy5zZXRT"
    "dHlsZVNoZWV0KGYiYmFja2dyb3VuZDoge0NfQkcyfTsgY29sb3I6IHtDX0dPTER9OyIpCiAgICAgICAgZGxnLnJlc2l6ZSg1NDAs"
    "IDQ0MCkKICAgICAgICBmb3JtID0gUVZCb3hMYXlvdXQoZGxnKQoKICAgICAgICBuYW1lX2ZpZWxkID0gUUxpbmVFZGl0KHJlYy5n"
    "ZXQoIm5hbWUiLCIiKSBpZiByZWMgZWxzZSAiIikKICAgICAgICBuYW1lX2ZpZWxkLnNldFBsYWNlaG9sZGVyVGV4dCgiTW9kdWxl"
    "IG5hbWUiKQoKICAgICAgICBzdGF0dXNfY29tYm8gPSBRQ29tYm9Cb3goKQogICAgICAgIHN0YXR1c19jb21iby5hZGRJdGVtcyhz"
    "ZWxmLlNUQVRVU0VTKQogICAgICAgIGlmIHJlYzoKICAgICAgICAgICAgaWR4ID0gc3RhdHVzX2NvbWJvLmZpbmRUZXh0KHJlYy5n"
    "ZXQoInN0YXR1cyIsIklkZWEiKSkKICAgICAgICAgICAgaWYgaWR4ID49IDA6CiAgICAgICAgICAgICAgICBzdGF0dXNfY29tYm8u"
    "c2V0Q3VycmVudEluZGV4KGlkeCkKCiAgICAgICAgZGVzY19maWVsZCA9IFFMaW5lRWRpdChyZWMuZ2V0KCJkZXNjcmlwdGlvbiIs"
    "IiIpIGlmIHJlYyBlbHNlICIiKQogICAgICAgIGRlc2NfZmllbGQuc2V0UGxhY2Vob2xkZXJUZXh0KCJPbmUtbGluZSBkZXNjcmlw"
    "dGlvbiIpCgogICAgICAgIG5vdGVzX2ZpZWxkID0gUVRleHRFZGl0KCkKICAgICAgICBub3Rlc19maWVsZC5zZXRQbGFpblRleHQo"
    "cmVjLmdldCgibm90ZXMiLCIiKSBpZiByZWMgZWxzZSAiIikKICAgICAgICBub3Rlc19maWVsZC5zZXRQbGFjZWhvbGRlclRleHQo"
    "CiAgICAgICAgICAgICJGdWxsIG5vdGVzIOKAlCBzcGVjLCBpZGVhcywgcmVxdWlyZW1lbnRzLCBlZGdlIGNhc2VzLi4uIgogICAg"
    "ICAgICkKICAgICAgICBub3Rlc19maWVsZC5zZXRNaW5pbXVtSGVpZ2h0KDIwMCkKCiAgICAgICAgZm9yIGxhYmVsLCB3aWRnZXQg"
    "aW4gWwogICAgICAgICAgICAoIk5hbWU6IiwgbmFtZV9maWVsZCksCiAgICAgICAgICAgICgiU3RhdHVzOiIsIHN0YXR1c19jb21i"
    "byksCiAgICAgICAgICAgICgiRGVzY3JpcHRpb246IiwgZGVzY19maWVsZCksCiAgICAgICAgICAgICgiTm90ZXM6Iiwgbm90ZXNf"
    "ZmllbGQpLAogICAgICAgIF06CiAgICAgICAgICAgIHJvd19sYXlvdXQgPSBRSEJveExheW91dCgpCiAgICAgICAgICAgIGxibCA9"
    "IFFMYWJlbChsYWJlbCkKICAgICAgICAgICAgbGJsLnNldEZpeGVkV2lkdGgoOTApCiAgICAgICAgICAgIHJvd19sYXlvdXQuYWRk"
    "V2lkZ2V0KGxibCkKICAgICAgICAgICAgcm93X2xheW91dC5hZGRXaWRnZXQod2lkZ2V0KQogICAgICAgICAgICBmb3JtLmFkZExh"
    "eW91dChyb3dfbGF5b3V0KQoKICAgICAgICBidG5fcm93ID0gUUhCb3hMYXlvdXQoKQogICAgICAgIGJ0bl9zYXZlICAgPSBfZ290"
    "aGljX2J0bigiU2F2ZSIpCiAgICAgICAgYnRuX2NhbmNlbCA9IF9nb3RoaWNfYnRuKCJDYW5jZWwiKQogICAgICAgIGJ0bl9zYXZl"
    "LmNsaWNrZWQuY29ubmVjdChkbGcuYWNjZXB0KQogICAgICAgIGJ0bl9jYW5jZWwuY2xpY2tlZC5jb25uZWN0KGRsZy5yZWplY3Qp"
    "CiAgICAgICAgYnRuX3Jvdy5hZGRXaWRnZXQoYnRuX3NhdmUpCiAgICAgICAgYnRuX3Jvdy5hZGRXaWRnZXQoYnRuX2NhbmNlbCkK"
    "ICAgICAgICBmb3JtLmFkZExheW91dChidG5fcm93KQoKICAgICAgICBpZiBkbGcuZXhlYygpID09IFFEaWFsb2cuRGlhbG9nQ29k"
    "ZS5BY2NlcHRlZDoKICAgICAgICAgICAgbmV3X3JlYyA9IHsKICAgICAgICAgICAgICAgICJpZCI6ICAgICAgICAgIHJlYy5nZXQo"
    "ImlkIiwgc3RyKHV1aWQudXVpZDQoKSkpIGlmIHJlYyBlbHNlIHN0cih1dWlkLnV1aWQ0KCkpLAogICAgICAgICAgICAgICAgIm5h"
    "bWUiOiAgICAgICAgbmFtZV9maWVsZC50ZXh0KCkuc3RyaXAoKSwKICAgICAgICAgICAgICAgICJzdGF0dXMiOiAgICAgIHN0YXR1"
    "c19jb21iby5jdXJyZW50VGV4dCgpLAogICAgICAgICAgICAgICAgImRlc2NyaXB0aW9uIjogZGVzY19maWVsZC50ZXh0KCkuc3Ry"
    "aXAoKSwKICAgICAgICAgICAgICAgICJub3RlcyI6ICAgICAgIG5vdGVzX2ZpZWxkLnRvUGxhaW5UZXh0KCkuc3RyaXAoKSwKICAg"
    "ICAgICAgICAgICAgICJjcmVhdGVkIjogICAgIHJlYy5nZXQoImNyZWF0ZWQiLCBkYXRldGltZS5ub3coKS5pc29mb3JtYXQoKSkg"
    "aWYgcmVjIGVsc2UgZGF0ZXRpbWUubm93KCkuaXNvZm9ybWF0KCksCiAgICAgICAgICAgICAgICAibW9kaWZpZWQiOiAgICBkYXRl"
    "dGltZS5ub3coKS5pc29mb3JtYXQoKSwKICAgICAgICAgICAgfQogICAgICAgICAgICBpZiByb3cgPj0gMDoKICAgICAgICAgICAg"
    "ICAgIHNlbGYuX3JlY29yZHNbcm93XSA9IG5ld19yZWMKICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgIHNlbGYuX3Jl"
    "Y29yZHMuYXBwZW5kKG5ld19yZWMpCiAgICAgICAgICAgIHdyaXRlX2pzb25sKHNlbGYuX3BhdGgsIHNlbGYuX3JlY29yZHMpCiAg"
    "ICAgICAgICAgIHNlbGYucmVmcmVzaCgpCgogICAgZGVmIF9kb19kZWxldGUoc2VsZikgLT4gTm9uZToKICAgICAgICByb3cgPSBz"
    "ZWxmLl90YWJsZS5jdXJyZW50Um93KCkKICAgICAgICBpZiAwIDw9IHJvdyA8IGxlbihzZWxmLl9yZWNvcmRzKToKICAgICAgICAg"
    "ICAgbmFtZSA9IHNlbGYuX3JlY29yZHNbcm93XS5nZXQoIm5hbWUiLCJ0aGlzIG1vZHVsZSIpCiAgICAgICAgICAgIHJlcGx5ID0g"
    "UU1lc3NhZ2VCb3gucXVlc3Rpb24oCiAgICAgICAgICAgICAgICBzZWxmLCAiRGVsZXRlIE1vZHVsZSIsCiAgICAgICAgICAgICAg"
    "ICBmIkRlbGV0ZSAne25hbWV9Jz8gQ2Fubm90IGJlIHVuZG9uZS4iLAogICAgICAgICAgICAgICAgUU1lc3NhZ2VCb3guU3RhbmRh"
    "cmRCdXR0b24uWWVzIHwgUU1lc3NhZ2VCb3guU3RhbmRhcmRCdXR0b24uTm8KICAgICAgICAgICAgKQogICAgICAgICAgICBpZiBy"
    "ZXBseSA9PSBRTWVzc2FnZUJveC5TdGFuZGFyZEJ1dHRvbi5ZZXM6CiAgICAgICAgICAgICAgICBzZWxmLl9yZWNvcmRzLnBvcChy"
    "b3cpCiAgICAgICAgICAgICAgICB3cml0ZV9qc29ubChzZWxmLl9wYXRoLCBzZWxmLl9yZWNvcmRzKQogICAgICAgICAgICAgICAg"
    "c2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX2RvX2V4cG9ydChzZWxmKSAtPiBOb25lOgogICAgICAgIHRyeToKICAgICAgICAgICAg"
    "ZXhwb3J0X2RpciA9IGNmZ19wYXRoKCJleHBvcnRzIikKICAgICAgICAgICAgZXhwb3J0X2Rpci5ta2RpcihwYXJlbnRzPVRydWUs"
    "IGV4aXN0X29rPVRydWUpCiAgICAgICAgICAgIHRzID0gZGF0ZXRpbWUubm93KCkuc3RyZnRpbWUoIiVZJW0lZF8lSCVNJVMiKQog"
    "ICAgICAgICAgICBvdXRfcGF0aCA9IGV4cG9ydF9kaXIgLyBmIm1vZHVsZXNfe3RzfS50eHQiCiAgICAgICAgICAgIGxpbmVzID0g"
    "WwogICAgICAgICAgICAgICAgIkVDSE8gREVDSyDigJQgTU9EVUxFIFRSQUNLRVIgRVhQT1JUIiwKICAgICAgICAgICAgICAgIGYi"
    "RXhwb3J0ZWQ6IHtkYXRldGltZS5ub3coKS5zdHJmdGltZSgnJVktJW0tJWQgJUg6JU06JVMnKX0iLAogICAgICAgICAgICAgICAg"
    "ZiJUb3RhbCBtb2R1bGVzOiB7bGVuKHNlbGYuX3JlY29yZHMpfSIsCiAgICAgICAgICAgICAgICAiPSIgKiA2MCwKICAgICAgICAg"
    "ICAgICAgICIiLAogICAgICAgICAgICBdCiAgICAgICAgICAgIGZvciByZWMgaW4gc2VsZi5fcmVjb3JkczoKICAgICAgICAgICAg"
    "ICAgIGxpbmVzLmV4dGVuZChbCiAgICAgICAgICAgICAgICAgICAgZiJNT0RVTEU6IHtyZWMuZ2V0KCduYW1lJywnJyl9IiwKICAg"
    "ICAgICAgICAgICAgICAgICBmIlN0YXR1czoge3JlYy5nZXQoJ3N0YXR1cycsJycpfSIsCiAgICAgICAgICAgICAgICAgICAgZiJE"
    "ZXNjcmlwdGlvbjoge3JlYy5nZXQoJ2Rlc2NyaXB0aW9uJywnJyl9IiwKICAgICAgICAgICAgICAgICAgICAiIiwKICAgICAgICAg"
    "ICAgICAgICAgICAiTm90ZXM6IiwKICAgICAgICAgICAgICAgICAgICByZWMuZ2V0KCJub3RlcyIsIiIpLAogICAgICAgICAgICAg"
    "ICAgICAgICIiLAogICAgICAgICAgICAgICAgICAgICItIiAqIDQwLAogICAgICAgICAgICAgICAgICAgICIiLAogICAgICAgICAg"
    "ICAgICAgXSkKICAgICAgICAgICAgb3V0X3BhdGgud3JpdGVfdGV4dCgiXG4iLmpvaW4obGluZXMpLCBlbmNvZGluZz0idXRmLTgi"
    "KQogICAgICAgICAgICBRQXBwbGljYXRpb24uY2xpcGJvYXJkKCkuc2V0VGV4dCgiXG4iLmpvaW4obGluZXMpKQogICAgICAgICAg"
    "ICBRTWVzc2FnZUJveC5pbmZvcm1hdGlvbigKICAgICAgICAgICAgICAgIHNlbGYsICJFeHBvcnRlZCIsCiAgICAgICAgICAgICAg"
    "ICBmIk1vZHVsZSB0cmFja2VyIGV4cG9ydGVkIHRvOlxue291dF9wYXRofVxuXG5BbHNvIGNvcGllZCB0byBjbGlwYm9hcmQuIgog"
    "ICAgICAgICAgICApCiAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbiBhcyBlOgogICAgICAgICAgICBRTWVzc2FnZUJveC53YXJuaW5n"
    "KHNlbGYsICJFeHBvcnQgRXJyb3IiLCBzdHIoZSkpCgoKCiAgICBkZWYgX3BhcnNlX2ltcG9ydF9lbnRyaWVzKHNlbGYsIHJhdzog"
    "c3RyKSAtPiBsaXN0W2RpY3RdOgogICAgICAgICIiIlBhcnNlIGltcG9ydGVkIHRleHQgaW50byBvbmUgb3IgbW9yZSBtb2R1bGUg"
    "cmVjb3Jkcy4iIiIKICAgICAgICBsYWJlbF9tYXAgPSB7CiAgICAgICAgICAgICJtb2R1bGUiOiAibmFtZSIsCiAgICAgICAgICAg"
    "ICJzdGF0dXMiOiAic3RhdHVzIiwKICAgICAgICAgICAgImRlc2NyaXB0aW9uIjogImRlc2NyaXB0aW9uIiwKICAgICAgICAgICAg"
    "Im5vdGVzIjogIm5vdGVzIiwKICAgICAgICAgICAgImZ1bGwgc3VtbWFyeSI6ICJub3RlcyIsCiAgICAgICAgfQoKICAgICAgICBk"
    "ZWYgX2JsYW5rKCkgLT4gZGljdDoKICAgICAgICAgICAgcmV0dXJuIHsKICAgICAgICAgICAgICAgICJuYW1lIjogIiIsCiAgICAg"
    "ICAgICAgICAgICAic3RhdHVzIjogIklkZWEiLAogICAgICAgICAgICAgICAgImRlc2NyaXB0aW9uIjogIiIsCiAgICAgICAgICAg"
    "ICAgICAibm90ZXMiOiAiIiwKICAgICAgICAgICAgfQoKICAgICAgICBkZWYgX2NsZWFuKHJlYzogZGljdCkgLT4gZGljdDoKICAg"
    "ICAgICAgICAgcmV0dXJuIHsKICAgICAgICAgICAgICAgICJuYW1lIjogcmVjLmdldCgibmFtZSIsICIiKS5zdHJpcCgpLAogICAg"
    "ICAgICAgICAgICAgInN0YXR1cyI6IChyZWMuZ2V0KCJzdGF0dXMiLCAiIikuc3RyaXAoKSBvciAiSWRlYSIpLAogICAgICAgICAg"
    "ICAgICAgImRlc2NyaXB0aW9uIjogcmVjLmdldCgiZGVzY3JpcHRpb24iLCAiIikuc3RyaXAoKSwKICAgICAgICAgICAgICAgICJu"
    "b3RlcyI6IHJlYy5nZXQoIm5vdGVzIiwgIiIpLnN0cmlwKCksCiAgICAgICAgICAgIH0KCiAgICAgICAgZGVmIF9pc19leHBvcnRf"
    "aGVhZGVyKGxpbmU6IHN0cikgLT4gYm9vbDoKICAgICAgICAgICAgbG93ID0gbGluZS5zdHJpcCgpLmxvd2VyKCkKICAgICAgICAg"
    "ICAgcmV0dXJuICgKICAgICAgICAgICAgICAgIGxvdy5zdGFydHN3aXRoKCJlY2hvIGRlY2siKSBvcgogICAgICAgICAgICAgICAg"
    "bG93LnN0YXJ0c3dpdGgoImV4cG9ydGVkOiIpIG9yCiAgICAgICAgICAgICAgICBsb3cuc3RhcnRzd2l0aCgidG90YWwgbW9kdWxl"
    "czoiKSBvcgogICAgICAgICAgICAgICAgbG93LnN0YXJ0c3dpdGgoInRvdGFsICIpCiAgICAgICAgICAgICkKCiAgICAgICAgZGVm"
    "IF9pc19kZWNvcmF0aXZlKGxpbmU6IHN0cikgLT4gYm9vbDoKICAgICAgICAgICAgcyA9IGxpbmUuc3RyaXAoKQogICAgICAgICAg"
    "ICBpZiBub3QgczoKICAgICAgICAgICAgICAgIHJldHVybiBGYWxzZQogICAgICAgICAgICBpZiBhbGwoY2ggaW4gIi09fl8q4oCi"
    "wrfigJQgIiBmb3IgY2ggaW4gcyk6CiAgICAgICAgICAgICAgICByZXR1cm4gVHJ1ZQogICAgICAgICAgICBpZiAocy5zdGFydHN3"
    "aXRoKCI9PT0iKSBhbmQgcy5lbmRzd2l0aCgiPT09IikpIG9yIChzLnN0YXJ0c3dpdGgoIi0tLSIpIGFuZCBzLmVuZHN3aXRoKCIt"
    "LS0iKSk6CiAgICAgICAgICAgICAgICByZXR1cm4gVHJ1ZQogICAgICAgICAgICByZXR1cm4gRmFsc2UKCiAgICAgICAgZGVmIF9p"
    "c19zZXBhcmF0b3IobGluZTogc3RyKSAtPiBib29sOgogICAgICAgICAgICBzID0gbGluZS5zdHJpcCgpCiAgICAgICAgICAgIHJl"
    "dHVybiBsZW4ocykgPj0gOCBhbmQgYWxsKGNoIGluICIt4oCUIiBmb3IgY2ggaW4gcykKCiAgICAgICAgZW50cmllczogbGlzdFtk"
    "aWN0XSA9IFtdCiAgICAgICAgY3VycmVudCA9IF9ibGFuaygpCiAgICAgICAgY3VycmVudF9maWVsZDogT3B0aW9uYWxbc3RyXSA9"
    "IE5vbmUKCiAgICAgICAgZGVmIF9oYXNfcGF5bG9hZChyZWM6IGRpY3QpIC0+IGJvb2w6CiAgICAgICAgICAgIHJldHVybiBhbnko"
    "Ym9vbCgocmVjLmdldChrLCAiIikgb3IgIiIpLnN0cmlwKCkpIGZvciBrIGluICgibmFtZSIsICJzdGF0dXMiLCAiZGVzY3JpcHRp"
    "b24iLCAibm90ZXMiKSkKCiAgICAgICAgZGVmIF9mbHVzaCgpIC0+IE5vbmU6CiAgICAgICAgICAgIG5vbmxvY2FsIGN1cnJlbnQs"
    "IGN1cnJlbnRfZmllbGQKICAgICAgICAgICAgY2xlYW5lZCA9IF9jbGVhbihjdXJyZW50KQogICAgICAgICAgICBpZiBjbGVhbmVk"
    "WyJuYW1lIl06CiAgICAgICAgICAgICAgICBlbnRyaWVzLmFwcGVuZChjbGVhbmVkKQogICAgICAgICAgICBjdXJyZW50ID0gX2Js"
    "YW5rKCkKICAgICAgICAgICAgY3VycmVudF9maWVsZCA9IE5vbmUKCiAgICAgICAgZm9yIHJhd19saW5lIGluIHJhdy5zcGxpdGxp"
    "bmVzKCk6CiAgICAgICAgICAgIGxpbmUgPSByYXdfbGluZS5yc3RyaXAoIlxuIikKICAgICAgICAgICAgc3RyaXBwZWQgPSBsaW5l"
    "LnN0cmlwKCkKCiAgICAgICAgICAgIGlmIF9pc19zZXBhcmF0b3Ioc3RyaXBwZWQpOgogICAgICAgICAgICAgICAgaWYgX2hhc19w"
    "YXlsb2FkKGN1cnJlbnQpOgogICAgICAgICAgICAgICAgICAgIF9mbHVzaCgpCiAgICAgICAgICAgICAgICBjb250aW51ZQoKICAg"
    "ICAgICAgICAgaWYgbm90IHN0cmlwcGVkOgogICAgICAgICAgICAgICAgaWYgY3VycmVudF9maWVsZCA9PSAibm90ZXMiOgogICAg"
    "ICAgICAgICAgICAgICAgIGN1cnJlbnRbIm5vdGVzIl0gPSAoY3VycmVudFsibm90ZXMiXSArICJcbiIpIGlmIGN1cnJlbnRbIm5v"
    "dGVzIl0gZWxzZSAiIgogICAgICAgICAgICAgICAgY29udGludWUKCiAgICAgICAgICAgIGlmIF9pc19leHBvcnRfaGVhZGVyKHN0"
    "cmlwcGVkKToKICAgICAgICAgICAgICAgIGNvbnRpbnVlCgogICAgICAgICAgICBpZiBfaXNfZGVjb3JhdGl2ZShzdHJpcHBlZCk6"
    "CiAgICAgICAgICAgICAgICBjb250aW51ZQoKICAgICAgICAgICAgaWYgIjoiIGluIHN0cmlwcGVkOgogICAgICAgICAgICAgICAg"
    "bWF5YmVfbGFiZWwsIG1heWJlX3ZhbHVlID0gc3RyaXBwZWQuc3BsaXQoIjoiLCAxKQogICAgICAgICAgICAgICAga2V5ID0gbWF5"
    "YmVfbGFiZWwuc3RyaXAoKS5sb3dlcigpCiAgICAgICAgICAgICAgICB2YWx1ZSA9IG1heWJlX3ZhbHVlLmxzdHJpcCgpCgogICAg"
    "ICAgICAgICAgICAgbWFwcGVkID0gbGFiZWxfbWFwLmdldChrZXkpCiAgICAgICAgICAgICAgICBpZiBtYXBwZWQ6CiAgICAgICAg"
    "ICAgICAgICAgICAgaWYgbWFwcGVkID09ICJuYW1lIiBhbmQgX2hhc19wYXlsb2FkKGN1cnJlbnQpOgogICAgICAgICAgICAgICAg"
    "ICAgICAgICBfZmx1c2goKQogICAgICAgICAgICAgICAgICAgIGN1cnJlbnRfZmllbGQgPSBtYXBwZWQKICAgICAgICAgICAgICAg"
    "ICAgICBpZiBtYXBwZWQgPT0gIm5vdGVzIjoKICAgICAgICAgICAgICAgICAgICAgICAgY3VycmVudFttYXBwZWRdID0gdmFsdWUK"
    "ICAgICAgICAgICAgICAgICAgICBlbGlmIG1hcHBlZCA9PSAic3RhdHVzIjoKICAgICAgICAgICAgICAgICAgICAgICAgY3VycmVu"
    "dFttYXBwZWRdID0gdmFsdWUgb3IgIklkZWEiCiAgICAgICAgICAgICAgICAgICAgZWxzZToKICAgICAgICAgICAgICAgICAgICAg"
    "ICAgY3VycmVudFttYXBwZWRdID0gdmFsdWUKICAgICAgICAgICAgICAgICAgICBjb250aW51ZQoKICAgICAgICAgICAgICAgICMg"
    "VW5rbm93biBsYWJlbGVkIGxpbmVzIGFyZSBtZXRhZGF0YS9jYXRlZ29yeS9mb290ZXIgbGluZXMuCiAgICAgICAgICAgICAgICBj"
    "dXJyZW50X2ZpZWxkID0gTm9uZQogICAgICAgICAgICAgICAgY29udGludWUKCiAgICAgICAgICAgIGlmIGN1cnJlbnRfZmllbGQg"
    "PT0gIm5vdGVzIjoKICAgICAgICAgICAgICAgIGN1cnJlbnRbIm5vdGVzIl0gPSAoY3VycmVudFsibm90ZXMiXSArICJcbiIgKyBz"
    "dHJpcHBlZCkgaWYgY3VycmVudFsibm90ZXMiXSBlbHNlIHN0cmlwcGVkCiAgICAgICAgICAgICAgICBjb250aW51ZQoKICAgICAg"
    "ICAgICAgaWYgY3VycmVudF9maWVsZCA9PSAiZGVzY3JpcHRpb24iOgogICAgICAgICAgICAgICAgY3VycmVudFsiZGVzY3JpcHRp"
    "b24iXSA9IChjdXJyZW50WyJkZXNjcmlwdGlvbiJdICsgIlxuIiArIHN0cmlwcGVkKSBpZiBjdXJyZW50WyJkZXNjcmlwdGlvbiJd"
    "IGVsc2Ugc3RyaXBwZWQKICAgICAgICAgICAgICAgIGNvbnRpbnVlCgogICAgICAgICAgICAjIElnbm9yZSB1bmxhYmVsZWQgbGlu"
    "ZXMgb3V0c2lkZSByZWNvZ25pemVkIGZpZWxkcy4KICAgICAgICAgICAgY29udGludWUKCiAgICAgICAgaWYgX2hhc19wYXlsb2Fk"
    "KGN1cnJlbnQpOgogICAgICAgICAgICBfZmx1c2goKQoKICAgICAgICByZXR1cm4gZW50cmllcwoKICAgIGRlZiBfZG9faW1wb3J0"
    "KHNlbGYpIC0+IE5vbmU6CiAgICAgICAgIiIiSW1wb3J0IG9uZSBvciBtb3JlIG1vZHVsZSBzcGVjcyBmcm9tIHBhc3RlZCB0ZXh0"
    "IG9yIGEgVFhUIGZpbGUuIiIiCiAgICAgICAgZGxnID0gUURpYWxvZyhzZWxmKQogICAgICAgIGRsZy5zZXRXaW5kb3dUaXRsZSgi"
    "SW1wb3J0IE1vZHVsZSBTcGVjIikKICAgICAgICBkbGcuc2V0U3R5bGVTaGVldChmImJhY2tncm91bmQ6IHtDX0JHMn07IGNvbG9y"
    "OiB7Q19HT0xEfTsiKQogICAgICAgIGRsZy5yZXNpemUoNTYwLCA0MjApCiAgICAgICAgbGF5b3V0ID0gUVZCb3hMYXlvdXQoZGxn"
    "KQogICAgICAgIGxheW91dC5hZGRXaWRnZXQoUUxhYmVsKAogICAgICAgICAgICAiUGFzdGUgbW9kdWxlIHRleHQgYmVsb3cgb3Ig"
    "bG9hZCBhIC50eHQgZXhwb3J0LlxuIgogICAgICAgICAgICAiU3VwcG9ydHMgTU9EVUxFIFRSQUNLRVIgZXhwb3J0cywgcmVnaXN0"
    "cnkgYmxvY2tzLCBhbmQgc2luZ2xlIGxhYmVsZWQgc3BlY3MuIgogICAgICAgICkpCgogICAgICAgIHRvb2xfcm93ID0gUUhCb3hM"
    "YXlvdXQoKQogICAgICAgIGJ0bl9sb2FkX3R4dCA9IF9nb3RoaWNfYnRuKCJMb2FkIFRYVCIpCiAgICAgICAgbG9hZGVkX2xibCA9"
    "IFFMYWJlbCgiTm8gZmlsZSBsb2FkZWQiKQogICAgICAgIGxvYWRlZF9sYmwuc2V0U3R5bGVTaGVldChmImNvbG9yOiB7Q19URVhU"
    "X0RJTX07IGZvbnQtc2l6ZTogOXB4OyIpCiAgICAgICAgdG9vbF9yb3cuYWRkV2lkZ2V0KGJ0bl9sb2FkX3R4dCkKICAgICAgICB0"
    "b29sX3Jvdy5hZGRXaWRnZXQobG9hZGVkX2xibCwgMSkKICAgICAgICBsYXlvdXQuYWRkTGF5b3V0KHRvb2xfcm93KQoKICAgICAg"
    "ICB0ZXh0X2ZpZWxkID0gUVRleHRFZGl0KCkKICAgICAgICB0ZXh0X2ZpZWxkLnNldFBsYWNlaG9sZGVyVGV4dCgiUGFzdGUgbW9k"
    "dWxlIHNwZWMocykgaGVyZS4uLiIpCiAgICAgICAgbGF5b3V0LmFkZFdpZGdldCh0ZXh0X2ZpZWxkLCAxKQoKICAgICAgICBkZWYg"
    "X2xvYWRfdHh0X2ludG9fZWRpdG9yKCkgLT4gTm9uZToKICAgICAgICAgICAgcGF0aCwgXyA9IFFGaWxlRGlhbG9nLmdldE9wZW5G"
    "aWxlTmFtZSgKICAgICAgICAgICAgICAgIHNlbGYsCiAgICAgICAgICAgICAgICAiTG9hZCBNb2R1bGUgU3BlY3MiLAogICAgICAg"
    "ICAgICAgICAgc3RyKGNmZ19wYXRoKCJleHBvcnRzIikpLAogICAgICAgICAgICAgICAgIlRleHQgRmlsZXMgKCoudHh0KTs7QWxs"
    "IEZpbGVzICgqKSIsCiAgICAgICAgICAgICkKICAgICAgICAgICAgaWYgbm90IHBhdGg6CiAgICAgICAgICAgICAgICByZXR1cm4K"
    "ICAgICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAgcmF3X3RleHQgPSBQYXRoKHBhdGgpLnJlYWRfdGV4dChlbmNvZGluZz0i"
    "dXRmLTgiKQogICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAgICAgICAgICAgICBRTWVzc2FnZUJveC53YXJu"
    "aW5nKHNlbGYsICJJbXBvcnQgRXJyb3IiLCBmIkNvdWxkIG5vdCByZWFkIGZpbGU6XG57ZX0iKQogICAgICAgICAgICAgICAgcmV0"
    "dXJuCiAgICAgICAgICAgIHRleHRfZmllbGQuc2V0UGxhaW5UZXh0KHJhd190ZXh0KQogICAgICAgICAgICBsb2FkZWRfbGJsLnNl"
    "dFRleHQoZiJMb2FkZWQ6IHtQYXRoKHBhdGgpLm5hbWV9IikKCiAgICAgICAgYnRuX2xvYWRfdHh0LmNsaWNrZWQuY29ubmVjdChf"
    "bG9hZF90eHRfaW50b19lZGl0b3IpCgogICAgICAgIGJ0bl9yb3cgPSBRSEJveExheW91dCgpCiAgICAgICAgYnRuX29rID0gX2dv"
    "dGhpY19idG4oIkltcG9ydCIpCiAgICAgICAgYnRuX2NhbmNlbCA9IF9nb3RoaWNfYnRuKCJDYW5jZWwiKQogICAgICAgIGJ0bl9v"
    "ay5jbGlja2VkLmNvbm5lY3QoZGxnLmFjY2VwdCkKICAgICAgICBidG5fY2FuY2VsLmNsaWNrZWQuY29ubmVjdChkbGcucmVqZWN0"
    "KQogICAgICAgIGJ0bl9yb3cuYWRkV2lkZ2V0KGJ0bl9vaykKICAgICAgICBidG5fcm93LmFkZFdpZGdldChidG5fY2FuY2VsKQog"
    "ICAgICAgIGxheW91dC5hZGRMYXlvdXQoYnRuX3JvdykKCiAgICAgICAgaWYgZGxnLmV4ZWMoKSA9PSBRRGlhbG9nLkRpYWxvZ0Nv"
    "ZGUuQWNjZXB0ZWQ6CiAgICAgICAgICAgIHJhdyA9IHRleHRfZmllbGQudG9QbGFpblRleHQoKS5zdHJpcCgpCiAgICAgICAgICAg"
    "IGlmIG5vdCByYXc6CiAgICAgICAgICAgICAgICByZXR1cm4KCiAgICAgICAgICAgIHBhcnNlZF9lbnRyaWVzID0gc2VsZi5fcGFy"
    "c2VfaW1wb3J0X2VudHJpZXMocmF3KQogICAgICAgICAgICBpZiBub3QgcGFyc2VkX2VudHJpZXM6CiAgICAgICAgICAgICAgICBR"
    "TWVzc2FnZUJveC53YXJuaW5nKAogICAgICAgICAgICAgICAgICAgIHNlbGYsCiAgICAgICAgICAgICAgICAgICAgIkltcG9ydCBF"
    "cnJvciIsCiAgICAgICAgICAgICAgICAgICAgIk5vIHZhbGlkIG1vZHVsZSBlbnRyaWVzIHdlcmUgZm91bmQuIEluY2x1ZGUgYXQg"
    "bGVhc3Qgb25lICdNb2R1bGU6JyBvciAnTU9EVUxFOicgYmxvY2suIiwKICAgICAgICAgICAgICAgICkKICAgICAgICAgICAgICAg"
    "IHJldHVybgoKICAgICAgICAgICAgbm93ID0gZGF0ZXRpbWUubm93KCkuaXNvZm9ybWF0KCkKICAgICAgICAgICAgZm9yIHBhcnNl"
    "ZCBpbiBwYXJzZWRfZW50cmllczoKICAgICAgICAgICAgICAgIHNlbGYuX3JlY29yZHMuYXBwZW5kKHsKICAgICAgICAgICAgICAg"
    "ICAgICAiaWQiOiBzdHIodXVpZC51dWlkNCgpKSwKICAgICAgICAgICAgICAgICAgICAibmFtZSI6IHBhcnNlZC5nZXQoIm5hbWUi"
    "LCAiIilbOjYwXSwKICAgICAgICAgICAgICAgICAgICAic3RhdHVzIjogcGFyc2VkLmdldCgic3RhdHVzIiwgIklkZWEiKSBvciAi"
    "SWRlYSIsCiAgICAgICAgICAgICAgICAgICAgImRlc2NyaXB0aW9uIjogcGFyc2VkLmdldCgiZGVzY3JpcHRpb24iLCAiIiksCiAg"
    "ICAgICAgICAgICAgICAgICAgIm5vdGVzIjogcGFyc2VkLmdldCgibm90ZXMiLCAiIiksCiAgICAgICAgICAgICAgICAgICAgImNy"
    "ZWF0ZWQiOiBub3csCiAgICAgICAgICAgICAgICAgICAgIm1vZGlmaWVkIjogbm93LAogICAgICAgICAgICAgICAgfSkKCiAgICAg"
    "ICAgICAgIHdyaXRlX2pzb25sKHNlbGYuX3BhdGgsIHNlbGYuX3JlY29yZHMpCiAgICAgICAgICAgIHNlbGYucmVmcmVzaCgpCiAg"
    "ICAgICAgICAgIFFNZXNzYWdlQm94LmluZm9ybWF0aW9uKAogICAgICAgICAgICAgICAgc2VsZiwKICAgICAgICAgICAgICAgICJJ"
    "bXBvcnQgQ29tcGxldGUiLAogICAgICAgICAgICAgICAgZiJJbXBvcnRlZCB7bGVuKHBhcnNlZF9lbnRyaWVzKX0gbW9kdWxlIGVu"
    "dHJ7J3knIGlmIGxlbihwYXJzZWRfZW50cmllcykgPT0gMSBlbHNlICdpZXMnfS4iCiAgICAgICAgICAgICkKCgojIOKUgOKUgCBQ"
    "QVNTIDUgQ09NUExFVEUg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiMgQWxsIHRhYiBjb250ZW50IGNsYXNzZXMgZGVm"
    "aW5lZC4KIyBTTFNjYW5zVGFiOiByZWJ1aWx0IOKAlCBEZWxldGUgYWRkZWQsIE1vZGlmeSBmaXhlZCwgdGltZXN0YW1wIHBhcnNl"
    "ciBmaXhlZCwKIyAgICAgICAgICAgICBjYXJkL2dyaW1vaXJlIHN0eWxlLCBjb3B5LXRvLWNsaXBib2FyZCBjb250ZXh0IG1lbnUu"
    "CiMgU0xDb21tYW5kc1RhYjogZ290aGljIHRhYmxlLCDip4kgQ29weSBDb21tYW5kIGJ1dHRvbi4KIyBKb2JUcmFja2VyVGFiOiBm"
    "dWxsIHJlYnVpbGQg4oCUIG11bHRpLXNlbGVjdCwgYXJjaGl2ZS9yZXN0b3JlLCBDU1YvVFNWIGV4cG9ydC4KIyBTZWxmVGFiOiBp"
    "bm5lciBzYW5jdHVtIGZvciBpZGxlIG5hcnJhdGl2ZSBhbmQgcmVmbGVjdGlvbiBvdXRwdXQuCiMgRGlhZ25vc3RpY3NUYWI6IHN0"
    "cnVjdHVyZWQgbG9nIHdpdGggbGV2ZWwtY29sb3JlZCBvdXRwdXQuCiMgTGVzc29uc1RhYjogTFNMIEZvcmJpZGRlbiBSdWxlc2V0"
    "IGJyb3dzZXIgd2l0aCBhZGQvZGVsZXRlL3NlYXJjaC4KIwojIE5leHQ6IFBhc3MgNiDigJQgTWFpbiBXaW5kb3cKIyAoTW9yZ2Fu"
    "bmFEZWNrIGNsYXNzLCBmdWxsIGxheW91dCwgQVBTY2hlZHVsZXIsIGZpcnN0LXJ1biBmbG93LAojICBkZXBlbmRlbmN5IGJvb3Rz"
    "dHJhcCwgc2hvcnRjdXQgY3JlYXRpb24sIHN0YXJ0dXAgc2VxdWVuY2UpCgoKIyDilZDilZDilZDilZDilZDilZDilZDilZDilZDi"
    "lZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDi"
    "lZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDi"
    "lZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZDilZAKIyBNT1JHQU5OQSBERUNL"
    "IOKAlCBQQVNTIDY6IE1BSU4gV0lORE9XICYgRU5UUlkgUE9JTlQKIwojIENvbnRhaW5zOgojICAgYm9vdHN0cmFwX2NoZWNrKCkg"
    "ICAgIOKAlCBkZXBlbmRlbmN5IHZhbGlkYXRpb24gKyBhdXRvLWluc3RhbGwgYmVmb3JlIFVJCiMgICBGaXJzdFJ1bkRpYWxvZyAg"
    "ICAgICAg4oCUIG1vZGVsIHBhdGggKyBjb25uZWN0aW9uIHR5cGUgc2VsZWN0aW9uCiMgICBKb3VybmFsU2lkZWJhciAgICAgICAg"
    "4oCUIGNvbGxhcHNpYmxlIGxlZnQgc2lkZWJhciAoc2Vzc2lvbiBicm93c2VyICsgam91cm5hbCkKIyAgIFRvcnBvclBhbmVsICAg"
    "ICAgICAgICDigJQgQVdBS0UgLyBBVVRPIC8gU1VTUEVORCBzdGF0ZSB0b2dnbGUKIyAgIE1vcmdhbm5hRGVjayAgICAgICAgICDi"
    "gJQgbWFpbiB3aW5kb3csIGZ1bGwgbGF5b3V0LCBhbGwgc2lnbmFsIGNvbm5lY3Rpb25zCiMgICBtYWluKCkgICAgICAgICAgICAg"
    "ICAg4oCUIGVudHJ5IHBvaW50IHdpdGggYm9vdHN0cmFwIHNlcXVlbmNlCiMg4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ"
    "4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQ4pWQCgppbXBvcnQgc3VicHJvY2Vz"
    "cwoKCiMg4pSA4pSAIFBSRS1MQVVOQ0ggREVQRU5ERU5DWSBCT09UU1RSQVAg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSACmRlZiBib290c3RyYXBfY2hlY2soKSAtPiBOb25lOgogICAgIiIiCiAgICBSdW5zIEJFRk9S"
    "RSBRQXBwbGljYXRpb24gaXMgY3JlYXRlZC4KICAgIENoZWNrcyBmb3IgUHlTaWRlNiBzZXBhcmF0ZWx5IChjYW4ndCBzaG93IEdV"
    "SSB3aXRob3V0IGl0KS4KICAgIEF1dG8taW5zdGFsbHMgYWxsIG90aGVyIG1pc3Npbmcgbm9uLWNyaXRpY2FsIGRlcHMgdmlhIHBp"
    "cC4KICAgIFZhbGlkYXRlcyBpbnN0YWxscyBzdWNjZWVkZWQuCiAgICBXcml0ZXMgcmVzdWx0cyB0byBhIGJvb3RzdHJhcCBsb2cg"
    "Zm9yIERpYWdub3N0aWNzIHRhYiB0byBwaWNrIHVwLgogICAgIiIiCiAgICAjIOKUgOKUgCBTdGVwIDE6IENoZWNrIFB5U2lkZTYg"
    "KGNhbid0IGF1dG8taW5zdGFsbCB3aXRob3V0IGl0IGFscmVhZHkgcHJlc2VudCkg4pSACiAgICB0cnk6CiAgICAgICAgaW1wb3J0"
    "IFB5U2lkZTYgICMgbm9xYQogICAgZXhjZXB0IEltcG9ydEVycm9yOgogICAgICAgICMgTm8gR1VJIGF2YWlsYWJsZSDigJQgdXNl"
    "IFdpbmRvd3MgbmF0aXZlIGRpYWxvZyB2aWEgY3R5cGVzCiAgICAgICAgdHJ5OgogICAgICAgICAgICBpbXBvcnQgY3R5cGVzCiAg"
    "ICAgICAgICAgIGN0eXBlcy53aW5kbGwudXNlcjMyLk1lc3NhZ2VCb3hXKAogICAgICAgICAgICAgICAgMCwKICAgICAgICAgICAg"
    "ICAgICJQeVNpZGU2IGlzIHJlcXVpcmVkIGJ1dCBub3QgaW5zdGFsbGVkLlxuXG4iCiAgICAgICAgICAgICAgICAiT3BlbiBhIHRl"
    "cm1pbmFsIGFuZCBydW46XG5cbiIKICAgICAgICAgICAgICAgICIgICAgcGlwIGluc3RhbGwgUHlTaWRlNlxuXG4iCiAgICAgICAg"
    "ICAgICAgICBmIlRoZW4gcmVzdGFydCB7REVDS19OQU1FfS4iLAogICAgICAgICAgICAgICAgZiJ7REVDS19OQU1FfSDigJQgTWlz"
    "c2luZyBEZXBlbmRlbmN5IiwKICAgICAgICAgICAgICAgIDB4MTAgICMgTUJfSUNPTkVSUk9SCiAgICAgICAgICAgICkKICAgICAg"
    "ICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICBwcmludCgiQ1JJVElDQUw6IFB5U2lkZTYgbm90IGluc3RhbGxlZC4gUnVu"
    "OiBwaXAgaW5zdGFsbCBQeVNpZGU2IikKICAgICAgICBzeXMuZXhpdCgxKQoKICAgICMg4pSA4pSAIFN0ZXAgMjogQXV0by1pbnN0"
    "YWxsIG90aGVyIG1pc3NpbmcgZGVwcyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgIF9BVVRPX0lOU1RBTEwgPSBbCiAgICAgICAg"
    "KCJhcHNjaGVkdWxlciIsICAgICAgICAgICAgICAgImFwc2NoZWR1bGVyIiksCiAgICAgICAgKCJsb2d1cnUiLCAgICAgICAgICAg"
    "ICAgICAgICAgImxvZ3VydSIpLAogICAgICAgICgicHlnYW1lIiwgICAgICAgICAgICAgICAgICAgICJweWdhbWUiKSwKICAgICAg"
    "ICAoInB5d2luMzIiLCAgICAgICAgICAgICAgICAgICAicHl3aW4zMiIpLAogICAgICAgICgicHN1dGlsIiwgICAgICAgICAgICAg"
    "ICAgICAgICJwc3V0aWwiKSwKICAgICAgICAoInJlcXVlc3RzIiwgICAgICAgICAgICAgICAgICAicmVxdWVzdHMiKSwKICAgIF0K"
    "CiAgICBpbXBvcnQgaW1wb3J0bGliCiAgICBib290c3RyYXBfbG9nID0gW10KCiAgICBmb3IgcGlwX25hbWUsIGltcG9ydF9uYW1l"
    "IGluIF9BVVRPX0lOU1RBTEw6CiAgICAgICAgdHJ5OgogICAgICAgICAgICBpbXBvcnRsaWIuaW1wb3J0X21vZHVsZShpbXBvcnRf"
    "bmFtZSkKICAgICAgICAgICAgYm9vdHN0cmFwX2xvZy5hcHBlbmQoZiJbQk9PVFNUUkFQXSB7cGlwX25hbWV9IOKckyIpCiAgICAg"
    "ICAgZXhjZXB0IEltcG9ydEVycm9yOgogICAgICAgICAgICBib290c3RyYXBfbG9nLmFwcGVuZCgKICAgICAgICAgICAgICAgIGYi"
    "W0JPT1RTVFJBUF0ge3BpcF9uYW1lfSBtaXNzaW5nIOKAlCBpbnN0YWxsaW5nLi4uIgogICAgICAgICAgICApCiAgICAgICAgICAg"
    "IHRyeToKICAgICAgICAgICAgICAgIHJlc3VsdCA9IHN1YnByb2Nlc3MucnVuKAogICAgICAgICAgICAgICAgICAgIFtzeXMuZXhl"
    "Y3V0YWJsZSwgIi1tIiwgInBpcCIsICJpbnN0YWxsIiwKICAgICAgICAgICAgICAgICAgICAgcGlwX25hbWUsICItLXF1aWV0Iiwg"
    "Ii0tbm8td2Fybi1zY3JpcHQtbG9jYXRpb24iXSwKICAgICAgICAgICAgICAgICAgICBjYXB0dXJlX291dHB1dD1UcnVlLCB0ZXh0"
    "PVRydWUsIHRpbWVvdXQ9MTIwLAogICAgICAgICAgICAgICAgICAgIGNyZWF0aW9uZmxhZ3M9Z2V0YXR0cihzdWJwcm9jZXNzLCAi"
    "Q1JFQVRFX05PX1dJTkRPVyIsIDApLAogICAgICAgICAgICAgICAgKQogICAgICAgICAgICAgICAgaWYgcmVzdWx0LnJldHVybmNv"
    "ZGUgPT0gMDoKICAgICAgICAgICAgICAgICAgICAjIFZhbGlkYXRlIGl0IGFjdHVhbGx5IGltcG9ydGVkIG5vdwogICAgICAgICAg"
    "ICAgICAgICAgIHRyeToKICAgICAgICAgICAgICAgICAgICAgICAgaW1wb3J0bGliLmltcG9ydF9tb2R1bGUoaW1wb3J0X25hbWUp"
    "CiAgICAgICAgICAgICAgICAgICAgICAgIGJvb3RzdHJhcF9sb2cuYXBwZW5kKAogICAgICAgICAgICAgICAgICAgICAgICAgICAg"
    "ZiJbQk9PVFNUUkFQXSB7cGlwX25hbWV9IGluc3RhbGxlZCDinJMiCiAgICAgICAgICAgICAgICAgICAgICAgICkKICAgICAgICAg"
    "ICAgICAgICAgICBleGNlcHQgSW1wb3J0RXJyb3I6CiAgICAgICAgICAgICAgICAgICAgICAgIGJvb3RzdHJhcF9sb2cuYXBwZW5k"
    "KAogICAgICAgICAgICAgICAgICAgICAgICAgICAgZiJbQk9PVFNUUkFQXSB7cGlwX25hbWV9IGluc3RhbGwgYXBwZWFyZWQgdG8g"
    "IgogICAgICAgICAgICAgICAgICAgICAgICAgICAgZiJzdWNjZWVkIGJ1dCBpbXBvcnQgc3RpbGwgZmFpbHMg4oCUIHJlc3RhcnQg"
    "bWF5ICIKICAgICAgICAgICAgICAgICAgICAgICAgICAgIGYiYmUgcmVxdWlyZWQuIgogICAgICAgICAgICAgICAgICAgICAgICAp"
    "CiAgICAgICAgICAgICAgICBlbHNlOgogICAgICAgICAgICAgICAgICAgIGJvb3RzdHJhcF9sb2cuYXBwZW5kKAogICAgICAgICAg"
    "ICAgICAgICAgICAgICBmIltCT09UU1RSQVBdIHtwaXBfbmFtZX0gaW5zdGFsbCBmYWlsZWQ6ICIKICAgICAgICAgICAgICAgICAg"
    "ICAgICAgZiJ7cmVzdWx0LnN0ZGVycls6MjAwXX0iCiAgICAgICAgICAgICAgICAgICAgKQogICAgICAgICAgICBleGNlcHQgc3Vi"
    "cHJvY2Vzcy5UaW1lb3V0RXhwaXJlZDoKICAgICAgICAgICAgICAgIGJvb3RzdHJhcF9sb2cuYXBwZW5kKAogICAgICAgICAgICAg"
    "ICAgICAgIGYiW0JPT1RTVFJBUF0ge3BpcF9uYW1lfSBpbnN0YWxsIHRpbWVkIG91dC4iCiAgICAgICAgICAgICAgICApCiAgICAg"
    "ICAgICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMgZToKICAgICAgICAgICAgICAgIGJvb3RzdHJhcF9sb2cuYXBwZW5kKAogICAgICAg"
    "ICAgICAgICAgICAgIGYiW0JPT1RTVFJBUF0ge3BpcF9uYW1lfSBpbnN0YWxsIGVycm9yOiB7ZX0iCiAgICAgICAgICAgICAgICAp"
    "CgogICAgIyDilIDilIAgU3RlcCAzOiBXcml0ZSBib290c3RyYXAgbG9nIGZvciBEaWFnbm9zdGljcyB0YWIg4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICB0cnk6CiAgICAgICAg"
    "bG9nX3BhdGggPSBTQ1JJUFRfRElSIC8gImxvZ3MiIC8gImJvb3RzdHJhcF9sb2cudHh0IgogICAgICAgIHdpdGggbG9nX3BhdGgu"
    "b3BlbigidyIsIGVuY29kaW5nPSJ1dGYtOCIpIGFzIGY6CiAgICAgICAgICAgIGYud3JpdGUoIlxuIi5qb2luKGJvb3RzdHJhcF9s"
    "b2cpKQogICAgZXhjZXB0IEV4Y2VwdGlvbjoKICAgICAgICBwYXNzCgoKIyDilIDilIAgRklSU1QgUlVOIERJQUxPRyDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIAKY2xhc3MgRmlyc3RSdW5EaWFsb2coUURpYWxvZyk6CiAgICAiIiIKICAgIFNob3duIG9uIGZpcnN0"
    "IGxhdW5jaCB3aGVuIGNvbmZpZy5qc29uIGRvZXNuJ3QgZXhpc3QuCiAgICBDb2xsZWN0cyBtb2RlbCBjb25uZWN0aW9uIHR5cGUg"
    "YW5kIHBhdGgva2V5LgogICAgVmFsaWRhdGVzIGNvbm5lY3Rpb24gYmVmb3JlIGFjY2VwdGluZy4KICAgIFdyaXRlcyBjb25maWcu"
    "anNvbiBvbiBzdWNjZXNzLgogICAgQ3JlYXRlcyBkZXNrdG9wIHNob3J0Y3V0LgogICAgIiIiCgogICAgZGVmIF9faW5pdF9fKHNl"
    "bGYsIHBhcmVudD1Ob25lKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKHBhcmVudCkKICAgICAgICBzZWxmLnNldFdpbmRvd1Rp"
    "dGxlKGYi4pymIHtERUNLX05BTUUudXBwZXIoKX0g4oCUIEZJUlNUIEFXQUtFTklORyIpCiAgICAgICAgc2VsZi5zZXRTdHlsZVNo"
    "ZWV0KFNUWUxFKQogICAgICAgIHNlbGYuc2V0Rml4ZWRTaXplKDUyMCwgNDAwKQogICAgICAgIHNlbGYuX3NldHVwX3VpKCkKCiAg"
    "ICBkZWYgX3NldHVwX3VpKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgcm9vdCA9IFFWQm94TGF5b3V0KHNlbGYpCiAgICAgICAgcm9v"
    "dC5zZXRTcGFjaW5nKDEwKQoKICAgICAgICB0aXRsZSA9IFFMYWJlbChmIuKcpiB7REVDS19OQU1FLnVwcGVyKCl9IOKAlCBGSVJT"
    "VCBBV0FLRU5JTkcg4pymIikKICAgICAgICB0aXRsZS5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19DUklN"
    "U09OfTsgZm9udC1zaXplOiAxNHB4OyBmb250LXdlaWdodDogYm9sZDsgIgogICAgICAgICAgICBmImZvbnQtZmFtaWx5OiB7REVD"
    "S19GT05UfSwgc2VyaWY7IGxldHRlci1zcGFjaW5nOiAycHg7IgogICAgICAgICkKICAgICAgICB0aXRsZS5zZXRBbGlnbm1lbnQo"
    "UXQuQWxpZ25tZW50RmxhZy5BbGlnbkNlbnRlcikKICAgICAgICByb290LmFkZFdpZGdldCh0aXRsZSkKCiAgICAgICAgc3ViID0g"
    "UUxhYmVsKAogICAgICAgICAgICBmIkNvbmZpZ3VyZSB0aGUgdmVzc2VsIGJlZm9yZSB7REVDS19OQU1FfSBtYXkgYXdha2VuLlxu"
    "IgogICAgICAgICAgICAiQWxsIHNldHRpbmdzIGFyZSBzdG9yZWQgbG9jYWxseS4gTm90aGluZyBsZWF2ZXMgdGhpcyBtYWNoaW5l"
    "LiIKICAgICAgICApCiAgICAgICAgc3ViLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX1RFWFRfRElNfTsg"
    "Zm9udC1zaXplOiAxMHB4OyAiCiAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsiCiAgICAgICAg"
    "KQogICAgICAgIHN1Yi5zZXRBbGlnbm1lbnQoUXQuQWxpZ25tZW50RmxhZy5BbGlnbkNlbnRlcikKICAgICAgICByb290LmFkZFdp"
    "ZGdldChzdWIpCgogICAgICAgICMg4pSA4pSAIENvbm5lY3Rpb24gdHlwZSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICByb290LmFkZFdpZGdldChfc2VjdGlvbl9sYmwoIuKdpyBBSSBDT05O"
    "RUNUSU9OIFRZUEUiKSkKICAgICAgICBzZWxmLl90eXBlX2NvbWJvID0gUUNvbWJvQm94KCkKICAgICAgICBzZWxmLl90eXBlX2Nv"
    "bWJvLmFkZEl0ZW1zKFsKICAgICAgICAgICAgIkxvY2FsIG1vZGVsIGZvbGRlciAodHJhbnNmb3JtZXJzKSIsCiAgICAgICAgICAg"
    "ICJPbGxhbWEgKGxvY2FsIHNlcnZpY2UpIiwKICAgICAgICAgICAgIkNsYXVkZSBBUEkgKEFudGhyb3BpYykiLAogICAgICAgICAg"
    "ICAiT3BlbkFJIEFQSSIsCiAgICAgICAgXSkKICAgICAgICBzZWxmLl90eXBlX2NvbWJvLmN1cnJlbnRJbmRleENoYW5nZWQuY29u"
    "bmVjdChzZWxmLl9vbl90eXBlX2NoYW5nZSkKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl90eXBlX2NvbWJvKQoKICAgICAg"
    "ICAjIOKUgOKUgCBEeW5hbWljIGNvbm5lY3Rpb24gZmllbGRzIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIHNl"
    "bGYuX3N0YWNrID0gUVN0YWNrZWRXaWRnZXQoKQoKICAgICAgICAjIFBhZ2UgMDogTG9jYWwgcGF0aAogICAgICAgIHAwID0gUVdp"
    "ZGdldCgpCiAgICAgICAgbDAgPSBRSEJveExheW91dChwMCkKICAgICAgICBsMC5zZXRDb250ZW50c01hcmdpbnMoMCwwLDAsMCkK"
    "ICAgICAgICBzZWxmLl9sb2NhbF9wYXRoID0gUUxpbmVFZGl0KCkKICAgICAgICBzZWxmLl9sb2NhbF9wYXRoLnNldFBsYWNlaG9s"
    "ZGVyVGV4dCgKICAgICAgICAgICAgciJEOlxBSVxNb2RlbHNcZG9scGhpbi04YiIKICAgICAgICApCiAgICAgICAgYnRuX2Jyb3dz"
    "ZSA9IF9nb3RoaWNfYnRuKCJCcm93c2UiKQogICAgICAgIGJ0bl9icm93c2UuY2xpY2tlZC5jb25uZWN0KHNlbGYuX2Jyb3dzZV9t"
    "b2RlbCkKICAgICAgICBsMC5hZGRXaWRnZXQoc2VsZi5fbG9jYWxfcGF0aCk7IGwwLmFkZFdpZGdldChidG5fYnJvd3NlKQogICAg"
    "ICAgIHNlbGYuX3N0YWNrLmFkZFdpZGdldChwMCkKCiAgICAgICAgIyBQYWdlIDE6IE9sbGFtYSBtb2RlbCBuYW1lCiAgICAgICAg"
    "cDEgPSBRV2lkZ2V0KCkKICAgICAgICBsMSA9IFFIQm94TGF5b3V0KHAxKQogICAgICAgIGwxLnNldENvbnRlbnRzTWFyZ2lucygw"
    "LDAsMCwwKQogICAgICAgIHNlbGYuX29sbGFtYV9tb2RlbCA9IFFMaW5lRWRpdCgpCiAgICAgICAgc2VsZi5fb2xsYW1hX21vZGVs"
    "LnNldFBsYWNlaG9sZGVyVGV4dCgiZG9scGhpbi0yLjYtN2IiKQogICAgICAgIGwxLmFkZFdpZGdldChzZWxmLl9vbGxhbWFfbW9k"
    "ZWwpCiAgICAgICAgc2VsZi5fc3RhY2suYWRkV2lkZ2V0KHAxKQoKICAgICAgICAjIFBhZ2UgMjogQ2xhdWRlIEFQSSBrZXkKICAg"
    "ICAgICBwMiA9IFFXaWRnZXQoKQogICAgICAgIGwyID0gUVZCb3hMYXlvdXQocDIpCiAgICAgICAgbDIuc2V0Q29udGVudHNNYXJn"
    "aW5zKDAsMCwwLDApCiAgICAgICAgc2VsZi5fY2xhdWRlX2tleSAgID0gUUxpbmVFZGl0KCkKICAgICAgICBzZWxmLl9jbGF1ZGVf"
    "a2V5LnNldFBsYWNlaG9sZGVyVGV4dCgic2stYW50LS4uLiIpCiAgICAgICAgc2VsZi5fY2xhdWRlX2tleS5zZXRFY2hvTW9kZShR"
    "TGluZUVkaXQuRWNob01vZGUuUGFzc3dvcmQpCiAgICAgICAgc2VsZi5fY2xhdWRlX21vZGVsID0gUUxpbmVFZGl0KCJjbGF1ZGUt"
    "c29ubmV0LTQtNiIpCiAgICAgICAgbDIuYWRkV2lkZ2V0KFFMYWJlbCgiQVBJIEtleToiKSkKICAgICAgICBsMi5hZGRXaWRnZXQo"
    "c2VsZi5fY2xhdWRlX2tleSkKICAgICAgICBsMi5hZGRXaWRnZXQoUUxhYmVsKCJNb2RlbDoiKSkKICAgICAgICBsMi5hZGRXaWRn"
    "ZXQoc2VsZi5fY2xhdWRlX21vZGVsKQogICAgICAgIHNlbGYuX3N0YWNrLmFkZFdpZGdldChwMikKCiAgICAgICAgIyBQYWdlIDM6"
    "IE9wZW5BSQogICAgICAgIHAzID0gUVdpZGdldCgpCiAgICAgICAgbDMgPSBRVkJveExheW91dChwMykKICAgICAgICBsMy5zZXRD"
    "b250ZW50c01hcmdpbnMoMCwwLDAsMCkKICAgICAgICBzZWxmLl9vYWlfa2V5ICAgPSBRTGluZUVkaXQoKQogICAgICAgIHNlbGYu"
    "X29haV9rZXkuc2V0UGxhY2Vob2xkZXJUZXh0KCJzay0uLi4iKQogICAgICAgIHNlbGYuX29haV9rZXkuc2V0RWNob01vZGUoUUxp"
    "bmVFZGl0LkVjaG9Nb2RlLlBhc3N3b3JkKQogICAgICAgIHNlbGYuX29haV9tb2RlbCA9IFFMaW5lRWRpdCgiZ3B0LTRvIikKICAg"
    "ICAgICBsMy5hZGRXaWRnZXQoUUxhYmVsKCJBUEkgS2V5OiIpKQogICAgICAgIGwzLmFkZFdpZGdldChzZWxmLl9vYWlfa2V5KQog"
    "ICAgICAgIGwzLmFkZFdpZGdldChRTGFiZWwoIk1vZGVsOiIpKQogICAgICAgIGwzLmFkZFdpZGdldChzZWxmLl9vYWlfbW9kZWwp"
    "CiAgICAgICAgc2VsZi5fc3RhY2suYWRkV2lkZ2V0KHAzKQoKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl9zdGFjaykKCiAg"
    "ICAgICAgIyDilIDilIAgVGVzdCArIHN0YXR1cyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIAKICAgICAgICB0ZXN0X3JvdyA9IFFIQm94TGF5b3V0KCkKICAgICAgICBzZWxmLl9idG5fdGVzdCA9"
    "IF9nb3RoaWNfYnRuKCJUZXN0IENvbm5lY3Rpb24iKQogICAgICAgIHNlbGYuX2J0bl90ZXN0LmNsaWNrZWQuY29ubmVjdChzZWxm"
    "Ll90ZXN0X2Nvbm5lY3Rpb24pCiAgICAgICAgc2VsZi5fc3RhdHVzX2xibCA9IFFMYWJlbCgiIikKICAgICAgICBzZWxmLl9zdGF0"
    "dXNfbGJsLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX1RFWFRfRElNfTsgZm9udC1zaXplOiAxMHB4OyAi"
    "CiAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsiCiAgICAgICAgKQogICAgICAgIHRlc3Rfcm93"
    "LmFkZFdpZGdldChzZWxmLl9idG5fdGVzdCkKICAgICAgICB0ZXN0X3Jvdy5hZGRXaWRnZXQoc2VsZi5fc3RhdHVzX2xibCwgMSkK"
    "ICAgICAgICByb290LmFkZExheW91dCh0ZXN0X3JvdykKCiAgICAgICAgIyDilIDilIAgRmFjZSBQYWNrIOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIHJvb3Qu"
    "YWRkV2lkZ2V0KF9zZWN0aW9uX2xibCgi4p2nIEZBQ0UgUEFDSyAob3B0aW9uYWwg4oCUIFpJUCBmaWxlKSIpKQogICAgICAgIGZh"
    "Y2Vfcm93ID0gUUhCb3hMYXlvdXQoKQogICAgICAgIHNlbGYuX2ZhY2VfcGF0aCA9IFFMaW5lRWRpdCgpCiAgICAgICAgc2VsZi5f"
    "ZmFjZV9wYXRoLnNldFBsYWNlaG9sZGVyVGV4dCgKICAgICAgICAgICAgZiJCcm93c2UgdG8ge0RFQ0tfTkFNRX0gZmFjZSBwYWNr"
    "IFpJUCAob3B0aW9uYWwsIGNhbiBhZGQgbGF0ZXIpIgogICAgICAgICkKICAgICAgICBzZWxmLl9mYWNlX3BhdGguc2V0U3R5bGVT"
    "aGVldCgKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19CRzN9OyBjb2xvcjoge0NfVEVYVF9ESU19OyAiCiAgICAgICAgICAg"
    "IGYiYm9yZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsgYm9yZGVyLXJhZGl1czogMnB4OyAiCiAgICAgICAgICAgIGYiZm9udC1m"
    "YW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsgZm9udC1zaXplOiAxMnB4OyBwYWRkaW5nOiA2cHggMTBweDsiCiAgICAgICAgKQog"
    "ICAgICAgIGJ0bl9mYWNlID0gX2dvdGhpY19idG4oIkJyb3dzZSIpCiAgICAgICAgYnRuX2ZhY2UuY2xpY2tlZC5jb25uZWN0KHNl"
    "bGYuX2Jyb3dzZV9mYWNlKQogICAgICAgIGZhY2Vfcm93LmFkZFdpZGdldChzZWxmLl9mYWNlX3BhdGgpCiAgICAgICAgZmFjZV9y"
    "b3cuYWRkV2lkZ2V0KGJ0bl9mYWNlKQogICAgICAgIHJvb3QuYWRkTGF5b3V0KGZhY2Vfcm93KQoKICAgICAgICAjIOKUgOKUgCBT"
    "aG9ydGN1dCBvcHRpb24g4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAg"
    "ICAgICAgc2VsZi5fc2hvcnRjdXRfY2IgPSBRQ2hlY2tCb3goCiAgICAgICAgICAgICJDcmVhdGUgZGVza3RvcCBzaG9ydGN1dCAo"
    "cmVjb21tZW5kZWQpIgogICAgICAgICkKICAgICAgICBzZWxmLl9zaG9ydGN1dF9jYi5zZXRDaGVja2VkKFRydWUpCiAgICAgICAg"
    "cm9vdC5hZGRXaWRnZXQoc2VsZi5fc2hvcnRjdXRfY2IpCgogICAgICAgICMg4pSA4pSAIEJ1dHRvbnMg4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAg"
    "cm9vdC5hZGRTdHJldGNoKCkKICAgICAgICBidG5fcm93ID0gUUhCb3hMYXlvdXQoKQogICAgICAgIHNlbGYuX2J0bl9hd2FrZW4g"
    "PSBfZ290aGljX2J0bigi4pymIEJFR0lOIEFXQUtFTklORyIpCiAgICAgICAgc2VsZi5fYnRuX2F3YWtlbi5zZXRFbmFibGVkKEZh"
    "bHNlKQogICAgICAgIGJ0bl9jYW5jZWwgPSBfZ290aGljX2J0bigi4pyXIENhbmNlbCIpCiAgICAgICAgc2VsZi5fYnRuX2F3YWtl"
    "bi5jbGlja2VkLmNvbm5lY3Qoc2VsZi5hY2NlcHQpCiAgICAgICAgYnRuX2NhbmNlbC5jbGlja2VkLmNvbm5lY3Qoc2VsZi5yZWpl"
    "Y3QpCiAgICAgICAgYnRuX3Jvdy5hZGRXaWRnZXQoc2VsZi5fYnRuX2F3YWtlbikKICAgICAgICBidG5fcm93LmFkZFdpZGdldChi"
    "dG5fY2FuY2VsKQogICAgICAgIHJvb3QuYWRkTGF5b3V0KGJ0bl9yb3cpCgogICAgZGVmIF9vbl90eXBlX2NoYW5nZShzZWxmLCBp"
    "ZHg6IGludCkgLT4gTm9uZToKICAgICAgICBzZWxmLl9zdGFjay5zZXRDdXJyZW50SW5kZXgoaWR4KQogICAgICAgIHNlbGYuX2J0"
    "bl9hd2FrZW4uc2V0RW5hYmxlZChGYWxzZSkKICAgICAgICBzZWxmLl9zdGF0dXNfbGJsLnNldFRleHQoIiIpCgogICAgZGVmIF9i"
    "cm93c2VfbW9kZWwoc2VsZikgLT4gTm9uZToKICAgICAgICBwYXRoID0gUUZpbGVEaWFsb2cuZ2V0RXhpc3RpbmdEaXJlY3Rvcnko"
    "CiAgICAgICAgICAgIHNlbGYsICJTZWxlY3QgTW9kZWwgRm9sZGVyIiwKICAgICAgICAgICAgciJEOlxBSVxNb2RlbHMiCiAgICAg"
    "ICAgKQogICAgICAgIGlmIHBhdGg6CiAgICAgICAgICAgIHNlbGYuX2xvY2FsX3BhdGguc2V0VGV4dChwYXRoKQoKICAgIGRlZiBf"
    "YnJvd3NlX2ZhY2Uoc2VsZikgLT4gTm9uZToKICAgICAgICBwYXRoLCBfID0gUUZpbGVEaWFsb2cuZ2V0T3BlbkZpbGVOYW1lKAog"
    "ICAgICAgICAgICBzZWxmLCAiU2VsZWN0IEZhY2UgUGFjayBaSVAiLAogICAgICAgICAgICBzdHIoUGF0aC5ob21lKCkgLyAiRGVz"
    "a3RvcCIpLAogICAgICAgICAgICAiWklQIEZpbGVzICgqLnppcCkiCiAgICAgICAgKQogICAgICAgIGlmIHBhdGg6CiAgICAgICAg"
    "ICAgIHNlbGYuX2ZhY2VfcGF0aC5zZXRUZXh0KHBhdGgpCgogICAgQHByb3BlcnR5CiAgICBkZWYgZmFjZV96aXBfcGF0aChzZWxm"
    "KSAtPiBzdHI6CiAgICAgICAgcmV0dXJuIHNlbGYuX2ZhY2VfcGF0aC50ZXh0KCkuc3RyaXAoKQoKICAgIGRlZiBfdGVzdF9jb25u"
    "ZWN0aW9uKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fc3RhdHVzX2xibC5zZXRUZXh0KCJUZXN0aW5nLi4uIikKICAgICAg"
    "ICBzZWxmLl9zdGF0dXNfbGJsLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX1RFWFRfRElNfTsgZm9udC1z"
    "aXplOiAxMHB4OyBmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyIKICAgICAgICApCiAgICAgICAgUUFwcGxpY2F0aW9u"
    "LnByb2Nlc3NFdmVudHMoKQoKICAgICAgICBpZHggPSBzZWxmLl90eXBlX2NvbWJvLmN1cnJlbnRJbmRleCgpCiAgICAgICAgb2sg"
    "ID0gRmFsc2UKICAgICAgICBtc2cgPSAiIgoKICAgICAgICBpZiBpZHggPT0gMDogICMgTG9jYWwKICAgICAgICAgICAgcGF0aCA9"
    "IHNlbGYuX2xvY2FsX3BhdGgudGV4dCgpLnN0cmlwKCkKICAgICAgICAgICAgaWYgcGF0aCBhbmQgUGF0aChwYXRoKS5leGlzdHMo"
    "KToKICAgICAgICAgICAgICAgIG9rICA9IFRydWUKICAgICAgICAgICAgICAgIG1zZyA9IGYiRm9sZGVyIGZvdW5kLiBNb2RlbCB3"
    "aWxsIGxvYWQgb24gc3RhcnR1cC4iCiAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICBtc2cgPSAiRm9sZGVyIG5vdCBm"
    "b3VuZC4gQ2hlY2sgdGhlIHBhdGguIgoKICAgICAgICBlbGlmIGlkeCA9PSAxOiAgIyBPbGxhbWEKICAgICAgICAgICAgdHJ5Ogog"
    "ICAgICAgICAgICAgICAgcmVxICA9IHVybGxpYi5yZXF1ZXN0LlJlcXVlc3QoCiAgICAgICAgICAgICAgICAgICAgImh0dHA6Ly9s"
    "b2NhbGhvc3Q6MTE0MzQvYXBpL3RhZ3MiCiAgICAgICAgICAgICAgICApCiAgICAgICAgICAgICAgICByZXNwID0gdXJsbGliLnJl"
    "cXVlc3QudXJsb3BlbihyZXEsIHRpbWVvdXQ9MykKICAgICAgICAgICAgICAgIG9rICAgPSByZXNwLnN0YXR1cyA9PSAyMDAKICAg"
    "ICAgICAgICAgICAgIG1zZyAgPSAiT2xsYW1hIGlzIHJ1bm5pbmcg4pyTIiBpZiBvayBlbHNlICJPbGxhbWEgbm90IHJlc3BvbmRp"
    "bmcuIgogICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAgICAgICAgICAgICBtc2cgPSBmIk9sbGFtYSBub3Qg"
    "cmVhY2hhYmxlOiB7ZX0iCgogICAgICAgIGVsaWYgaWR4ID09IDI6ICAjIENsYXVkZQogICAgICAgICAgICBrZXkgPSBzZWxmLl9j"
    "bGF1ZGVfa2V5LnRleHQoKS5zdHJpcCgpCiAgICAgICAgICAgIG9rICA9IGJvb2woa2V5IGFuZCBrZXkuc3RhcnRzd2l0aCgic2st"
    "YW50IikpCiAgICAgICAgICAgIG1zZyA9ICJBUEkga2V5IGZvcm1hdCBsb29rcyBjb3JyZWN0LiIgaWYgb2sgZWxzZSAiRW50ZXIg"
    "YSB2YWxpZCBDbGF1ZGUgQVBJIGtleS4iCgogICAgICAgIGVsaWYgaWR4ID09IDM6ICAjIE9wZW5BSQogICAgICAgICAgICBrZXkg"
    "PSBzZWxmLl9vYWlfa2V5LnRleHQoKS5zdHJpcCgpCiAgICAgICAgICAgIG9rICA9IGJvb2woa2V5IGFuZCBrZXkuc3RhcnRzd2l0"
    "aCgic2stIikpCiAgICAgICAgICAgIG1zZyA9ICJBUEkga2V5IGZvcm1hdCBsb29rcyBjb3JyZWN0LiIgaWYgb2sgZWxzZSAiRW50"
    "ZXIgYSB2YWxpZCBPcGVuQUkgQVBJIGtleS4iCgogICAgICAgIGNvbG9yID0gQ19HUkVFTiBpZiBvayBlbHNlIENfQ1JJTVNPTgog"
    "ICAgICAgIHNlbGYuX3N0YXR1c19sYmwuc2V0VGV4dChtc2cpCiAgICAgICAgc2VsZi5fc3RhdHVzX2xibC5zZXRTdHlsZVNoZWV0"
    "KAogICAgICAgICAgICBmImNvbG9yOiB7Y29sb3J9OyBmb250LXNpemU6IDEwcHg7IGZvbnQtZmFtaWx5OiB7REVDS19GT05UfSwg"
    "c2VyaWY7IgogICAgICAgICkKICAgICAgICBzZWxmLl9idG5fYXdha2VuLnNldEVuYWJsZWQob2spCgogICAgZGVmIGJ1aWxkX2Nv"
    "bmZpZyhzZWxmKSAtPiBkaWN0OgogICAgICAgICIiIkJ1aWxkIGFuZCByZXR1cm4gdXBkYXRlZCBjb25maWcgZGljdCBmcm9tIGRp"
    "YWxvZyBzZWxlY3Rpb25zLiIiIgogICAgICAgIGNmZyAgICAgPSBfZGVmYXVsdF9jb25maWcoKQogICAgICAgIGlkeCAgICAgPSBz"
    "ZWxmLl90eXBlX2NvbWJvLmN1cnJlbnRJbmRleCgpCiAgICAgICAgdHlwZXMgICA9IFsibG9jYWwiLCAib2xsYW1hIiwgImNsYXVk"
    "ZSIsICJvcGVuYWkiXQogICAgICAgIGNmZ1sibW9kZWwiXVsidHlwZSJdID0gdHlwZXNbaWR4XQoKICAgICAgICBpZiBpZHggPT0g"
    "MDoKICAgICAgICAgICAgY2ZnWyJtb2RlbCJdWyJwYXRoIl0gPSBzZWxmLl9sb2NhbF9wYXRoLnRleHQoKS5zdHJpcCgpCiAgICAg"
    "ICAgZWxpZiBpZHggPT0gMToKICAgICAgICAgICAgY2ZnWyJtb2RlbCJdWyJvbGxhbWFfbW9kZWwiXSA9IHNlbGYuX29sbGFtYV9t"
    "b2RlbC50ZXh0KCkuc3RyaXAoKSBvciAiZG9scGhpbi0yLjYtN2IiCiAgICAgICAgZWxpZiBpZHggPT0gMjoKICAgICAgICAgICAg"
    "Y2ZnWyJtb2RlbCJdWyJhcGlfa2V5Il0gICA9IHNlbGYuX2NsYXVkZV9rZXkudGV4dCgpLnN0cmlwKCkKICAgICAgICAgICAgY2Zn"
    "WyJtb2RlbCJdWyJhcGlfbW9kZWwiXSA9IHNlbGYuX2NsYXVkZV9tb2RlbC50ZXh0KCkuc3RyaXAoKQogICAgICAgICAgICBjZmdb"
    "Im1vZGVsIl1bImFwaV90eXBlIl0gID0gImNsYXVkZSIKICAgICAgICBlbGlmIGlkeCA9PSAzOgogICAgICAgICAgICBjZmdbIm1v"
    "ZGVsIl1bImFwaV9rZXkiXSAgID0gc2VsZi5fb2FpX2tleS50ZXh0KCkuc3RyaXAoKQogICAgICAgICAgICBjZmdbIm1vZGVsIl1b"
    "ImFwaV9tb2RlbCJdID0gc2VsZi5fb2FpX21vZGVsLnRleHQoKS5zdHJpcCgpCiAgICAgICAgICAgIGNmZ1sibW9kZWwiXVsiYXBp"
    "X3R5cGUiXSAgPSAib3BlbmFpIgoKICAgICAgICBjZmdbImZpcnN0X3J1biJdID0gRmFsc2UKICAgICAgICByZXR1cm4gY2ZnCgog"
    "ICAgQHByb3BlcnR5CiAgICBkZWYgY3JlYXRlX3Nob3J0Y3V0KHNlbGYpIC0+IGJvb2w6CiAgICAgICAgcmV0dXJuIHNlbGYuX3No"
    "b3J0Y3V0X2NiLmlzQ2hlY2tlZCgpCgoKIyDilIDilIAgSk9VUk5BTCBTSURFQkFSIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gApjbGFzcyBKb3VybmFsU2lkZWJhcihRV2lkZ2V0KToKICAgICIiIgogICAgQ29sbGFwc2libGUgbGVmdCBzaWRlYmFyIG5leHQg"
    "dG8gdGhlIHBlcnNvbmEgY2hhdCB0YWIuCiAgICBUb3A6IHNlc3Npb24gY29udHJvbHMgKGN1cnJlbnQgc2Vzc2lvbiBuYW1lLCBz"
    "YXZlL2xvYWQgYnV0dG9ucywKICAgICAgICAgYXV0b3NhdmUgaW5kaWNhdG9yKS4KICAgIEJvZHk6IHNjcm9sbGFibGUgc2Vzc2lv"
    "biBsaXN0IOKAlCBkYXRlLCBBSSBuYW1lLCBtZXNzYWdlIGNvdW50LgogICAgQ29sbGFwc2VzIGxlZnR3YXJkIHRvIGEgdGhpbiBz"
    "dHJpcC4KCiAgICBTaWduYWxzOgogICAgICAgIHNlc3Npb25fbG9hZF9yZXF1ZXN0ZWQoc3RyKSAgIOKAlCBkYXRlIHN0cmluZyBv"
    "ZiBzZXNzaW9uIHRvIGxvYWQKICAgICAgICBzZXNzaW9uX2NsZWFyX3JlcXVlc3RlZCgpICAgICDigJQgcmV0dXJuIHRvIGN1cnJl"
    "bnQgc2Vzc2lvbgogICAgIiIiCgogICAgc2Vzc2lvbl9sb2FkX3JlcXVlc3RlZCAgPSBTaWduYWwoc3RyKQogICAgc2Vzc2lvbl9j"
    "bGVhcl9yZXF1ZXN0ZWQgPSBTaWduYWwoKQoKICAgIGRlZiBfX2luaXRfXyhzZWxmLCBzZXNzaW9uX21ncjogIlNlc3Npb25NYW5h"
    "Z2VyIiwgcGFyZW50PU5vbmUpOgogICAgICAgIHN1cGVyKCkuX19pbml0X18ocGFyZW50KQogICAgICAgIHNlbGYuX3Nlc3Npb25f"
    "bWdyID0gc2Vzc2lvbl9tZ3IKICAgICAgICBzZWxmLl9leHBhbmRlZCAgICA9IFRydWUKICAgICAgICBzZWxmLl9zZXR1cF91aSgp"
    "CiAgICAgICAgc2VsZi5yZWZyZXNoKCkKCiAgICBkZWYgX3NldHVwX3VpKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgIyBVc2UgYSBo"
    "b3Jpem9udGFsIHJvb3QgbGF5b3V0IOKAlCBjb250ZW50IG9uIGxlZnQsIHRvZ2dsZSBzdHJpcCBvbiByaWdodAogICAgICAgIHJv"
    "b3QgPSBRSEJveExheW91dChzZWxmKQogICAgICAgIHJvb3Quc2V0Q29udGVudHNNYXJnaW5zKDAsIDAsIDAsIDApCiAgICAgICAg"
    "cm9vdC5zZXRTcGFjaW5nKDApCgogICAgICAgICMg4pSA4pSAIENvbGxhcHNlIHRvZ2dsZSBzdHJpcCDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBzZWxmLl90b2dnbGVfc3RyaXAgPSBRV2lkZ2V0KCkKICAgICAgICBzZWxm"
    "Ll90b2dnbGVfc3RyaXAuc2V0Rml4ZWRXaWR0aCgyMCkKICAgICAgICBzZWxmLl90b2dnbGVfc3RyaXAuc2V0U3R5bGVTaGVldCgK"
    "ICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19CRzN9OyBib3JkZXItcmlnaHQ6IDFweCBzb2xpZCB7Q19DUklNU09OX0RJTX07"
    "IgogICAgICAgICkKICAgICAgICB0c19sYXlvdXQgPSBRVkJveExheW91dChzZWxmLl90b2dnbGVfc3RyaXApCiAgICAgICAgdHNf"
    "bGF5b3V0LnNldENvbnRlbnRzTWFyZ2lucygwLCA4LCAwLCA4KQogICAgICAgIHNlbGYuX3RvZ2dsZV9idG4gPSBRVG9vbEJ1dHRv"
    "bigpCiAgICAgICAgc2VsZi5fdG9nZ2xlX2J0bi5zZXRGaXhlZFNpemUoMTgsIDE4KQogICAgICAgIHNlbGYuX3RvZ2dsZV9idG4u"
    "c2V0VGV4dCgi4peAIikKICAgICAgICBzZWxmLl90b2dnbGVfYnRuLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiYmFja2dy"
    "b3VuZDogdHJhbnNwYXJlbnQ7IGNvbG9yOiB7Q19HT0xEX0RJTX07ICIKICAgICAgICAgICAgZiJib3JkZXI6IG5vbmU7IGZvbnQt"
    "c2l6ZTogMTBweDsiCiAgICAgICAgKQogICAgICAgIHNlbGYuX3RvZ2dsZV9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYuX3RvZ2ds"
    "ZSkKICAgICAgICB0c19sYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX3RvZ2dsZV9idG4pCiAgICAgICAgdHNfbGF5b3V0LmFkZFN0cmV0"
    "Y2goKQoKICAgICAgICAjIOKUgOKUgCBNYWluIGNvbnRlbnQg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgc2VsZi5fY29udGVudCA9IFFXaWRnZXQoKQogICAgICAgIHNlbGYu"
    "X2NvbnRlbnQuc2V0TWluaW11bVdpZHRoKDE4MCkKICAgICAgICBzZWxmLl9jb250ZW50LnNldE1heGltdW1XaWR0aCgyMjApCiAg"
    "ICAgICAgY29udGVudF9sYXlvdXQgPSBRVkJveExheW91dChzZWxmLl9jb250ZW50KQogICAgICAgIGNvbnRlbnRfbGF5b3V0LnNl"
    "dENvbnRlbnRzTWFyZ2lucyg0LCA0LCA0LCA0KQogICAgICAgIGNvbnRlbnRfbGF5b3V0LnNldFNwYWNpbmcoNCkKCiAgICAgICAg"
    "IyBTZWN0aW9uIGxhYmVsCiAgICAgICAgY29udGVudF9sYXlvdXQuYWRkV2lkZ2V0KF9zZWN0aW9uX2xibCgi4p2nIEpPVVJOQUwi"
    "KSkKCiAgICAgICAgIyBDdXJyZW50IHNlc3Npb24gaW5mbwogICAgICAgIHNlbGYuX3Nlc3Npb25fbmFtZSA9IFFMYWJlbCgiTmV3"
    "IFNlc3Npb24iKQogICAgICAgIHNlbGYuX3Nlc3Npb25fbmFtZS5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7"
    "Q19HT0xEfTsgZm9udC1zaXplOiAxMHB4OyBmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyAiCiAgICAgICAgICAgIGYi"
    "Zm9udC1zdHlsZTogaXRhbGljOyIKICAgICAgICApCiAgICAgICAgc2VsZi5fc2Vzc2lvbl9uYW1lLnNldFdvcmRXcmFwKFRydWUp"
    "CiAgICAgICAgY29udGVudF9sYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX3Nlc3Npb25fbmFtZSkKCiAgICAgICAgIyBTYXZlIC8gTG9h"
    "ZCByb3cKICAgICAgICBjdHJsX3JvdyA9IFFIQm94TGF5b3V0KCkKICAgICAgICBzZWxmLl9idG5fc2F2ZSA9IF9nb3RoaWNfYnRu"
    "KCLwn5K+IikKICAgICAgICBzZWxmLl9idG5fc2F2ZS5zZXRGaXhlZFNpemUoMzIsIDI0KQogICAgICAgIHNlbGYuX2J0bl9zYXZl"
    "LnNldFRvb2xUaXAoIlNhdmUgc2Vzc2lvbiBub3ciKQogICAgICAgIHNlbGYuX2J0bl9sb2FkID0gX2dvdGhpY19idG4oIvCfk4Ii"
    "KQogICAgICAgIHNlbGYuX2J0bl9sb2FkLnNldEZpeGVkU2l6ZSgzMiwgMjQpCiAgICAgICAgc2VsZi5fYnRuX2xvYWQuc2V0VG9v"
    "bFRpcCgiQnJvd3NlIGFuZCBsb2FkIGEgcGFzdCBzZXNzaW9uIikKICAgICAgICBzZWxmLl9hdXRvc2F2ZV9kb3QgPSBRTGFiZWwo"
    "IuKXjyIpCiAgICAgICAgc2VsZi5fYXV0b3NhdmVfZG90LnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX1RF"
    "WFRfRElNfTsgZm9udC1zaXplOiA4cHg7IGJvcmRlcjogbm9uZTsiCiAgICAgICAgKQogICAgICAgIHNlbGYuX2F1dG9zYXZlX2Rv"
    "dC5zZXRUb29sVGlwKCJBdXRvc2F2ZSBzdGF0dXMiKQogICAgICAgIHNlbGYuX2J0bl9zYXZlLmNsaWNrZWQuY29ubmVjdChzZWxm"
    "Ll9kb19zYXZlKQogICAgICAgIHNlbGYuX2J0bl9sb2FkLmNsaWNrZWQuY29ubmVjdChzZWxmLl9kb19sb2FkKQogICAgICAgIGN0"
    "cmxfcm93LmFkZFdpZGdldChzZWxmLl9idG5fc2F2ZSkKICAgICAgICBjdHJsX3Jvdy5hZGRXaWRnZXQoc2VsZi5fYnRuX2xvYWQp"
    "CiAgICAgICAgY3RybF9yb3cuYWRkV2lkZ2V0KHNlbGYuX2F1dG9zYXZlX2RvdCkKICAgICAgICBjdHJsX3Jvdy5hZGRTdHJldGNo"
    "KCkKICAgICAgICBjb250ZW50X2xheW91dC5hZGRMYXlvdXQoY3RybF9yb3cpCgogICAgICAgICMgSm91cm5hbCBsb2FkZWQgaW5k"
    "aWNhdG9yCiAgICAgICAgc2VsZi5fam91cm5hbF9sYmwgPSBRTGFiZWwoIiIpCiAgICAgICAgc2VsZi5fam91cm5hbF9sYmwuc2V0"
    "U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJjb2xvcjoge0NfUFVSUExFfTsgZm9udC1zaXplOiA5cHg7IGZvbnQtZmFtaWx5OiB7"
    "REVDS19GT05UfSwgc2VyaWY7ICIKICAgICAgICAgICAgZiJmb250LXN0eWxlOiBpdGFsaWM7IgogICAgICAgICkKICAgICAgICBz"
    "ZWxmLl9qb3VybmFsX2xibC5zZXRXb3JkV3JhcChUcnVlKQogICAgICAgIGNvbnRlbnRfbGF5b3V0LmFkZFdpZGdldChzZWxmLl9q"
    "b3VybmFsX2xibCkKCiAgICAgICAgIyBDbGVhciBqb3VybmFsIGJ1dHRvbiAoaGlkZGVuIHdoZW4gbm90IGxvYWRlZCkKICAgICAg"
    "ICBzZWxmLl9idG5fY2xlYXJfam91cm5hbCA9IF9nb3RoaWNfYnRuKCLinJcgUmV0dXJuIHRvIFByZXNlbnQiKQogICAgICAgIHNl"
    "bGYuX2J0bl9jbGVhcl9qb3VybmFsLnNldFZpc2libGUoRmFsc2UpCiAgICAgICAgc2VsZi5fYnRuX2NsZWFyX2pvdXJuYWwuY2xp"
    "Y2tlZC5jb25uZWN0KHNlbGYuX2RvX2NsZWFyX2pvdXJuYWwpCiAgICAgICAgY29udGVudF9sYXlvdXQuYWRkV2lkZ2V0KHNlbGYu"
    "X2J0bl9jbGVhcl9qb3VybmFsKQoKICAgICAgICAjIERpdmlkZXIKICAgICAgICBkaXYgPSBRRnJhbWUoKQogICAgICAgIGRpdi5z"
    "ZXRGcmFtZVNoYXBlKFFGcmFtZS5TaGFwZS5ITGluZSkKICAgICAgICBkaXYuc2V0U3R5bGVTaGVldChmImNvbG9yOiB7Q19DUklN"
    "U09OX0RJTX07IikKICAgICAgICBjb250ZW50X2xheW91dC5hZGRXaWRnZXQoZGl2KQoKICAgICAgICAjIFNlc3Npb24gbGlzdAog"
    "ICAgICAgIGNvbnRlbnRfbGF5b3V0LmFkZFdpZGdldChfc2VjdGlvbl9sYmwoIuKdpyBQQVNUIFNFU1NJT05TIikpCiAgICAgICAg"
    "c2VsZi5fc2Vzc2lvbl9saXN0ID0gUUxpc3RXaWRnZXQoKQogICAgICAgIHNlbGYuX3Nlc3Npb25fbGlzdC5zZXRTdHlsZVNoZWV0"
    "KAogICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX0JHMn07IGNvbG9yOiB7Q19HT0xEfTsgIgogICAgICAgICAgICBmImJvcmRl"
    "cjogMXB4IHNvbGlkIHtDX0JPUkRFUn07ICIKICAgICAgICAgICAgZiJmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyBm"
    "b250LXNpemU6IDEwcHg7IgogICAgICAgICAgICBmIlFMaXN0V2lkZ2V0OjppdGVtOnNlbGVjdGVkIHt7IGJhY2tncm91bmQ6IHtD"
    "X0NSSU1TT05fRElNfTsgfX0iCiAgICAgICAgKQogICAgICAgIHNlbGYuX3Nlc3Npb25fbGlzdC5pdGVtRG91YmxlQ2xpY2tlZC5j"
    "b25uZWN0KHNlbGYuX29uX3Nlc3Npb25fY2xpY2spCiAgICAgICAgc2VsZi5fc2Vzc2lvbl9saXN0Lml0ZW1DbGlja2VkLmNvbm5l"
    "Y3Qoc2VsZi5fb25fc2Vzc2lvbl9jbGljaykKICAgICAgICBjb250ZW50X2xheW91dC5hZGRXaWRnZXQoc2VsZi5fc2Vzc2lvbl9s"
    "aXN0LCAxKQoKICAgICAgICAjIEFkZCBjb250ZW50IGFuZCB0b2dnbGUgc3RyaXAgdG8gdGhlIHJvb3QgaG9yaXpvbnRhbCBsYXlv"
    "dXQKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl9jb250ZW50KQogICAgICAgIHJvb3QuYWRkV2lkZ2V0KHNlbGYuX3RvZ2ds"
    "ZV9zdHJpcCkKCiAgICBkZWYgX3RvZ2dsZShzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2V4cGFuZGVkID0gbm90IHNlbGYu"
    "X2V4cGFuZGVkCiAgICAgICAgc2VsZi5fY29udGVudC5zZXRWaXNpYmxlKHNlbGYuX2V4cGFuZGVkKQogICAgICAgIHNlbGYuX3Rv"
    "Z2dsZV9idG4uc2V0VGV4dCgi4peAIiBpZiBzZWxmLl9leHBhbmRlZCBlbHNlICLilrYiKQogICAgICAgIHNlbGYudXBkYXRlR2Vv"
    "bWV0cnkoKQogICAgICAgIHAgPSBzZWxmLnBhcmVudFdpZGdldCgpCiAgICAgICAgaWYgcCBhbmQgcC5sYXlvdXQoKToKICAgICAg"
    "ICAgICAgcC5sYXlvdXQoKS5hY3RpdmF0ZSgpCgogICAgZGVmIHJlZnJlc2goc2VsZikgLT4gTm9uZToKICAgICAgICBzZXNzaW9u"
    "cyA9IHNlbGYuX3Nlc3Npb25fbWdyLmxpc3Rfc2Vzc2lvbnMoKQogICAgICAgIHNlbGYuX3Nlc3Npb25fbGlzdC5jbGVhcigpCiAg"
    "ICAgICAgZm9yIHMgaW4gc2Vzc2lvbnM6CiAgICAgICAgICAgIGRhdGVfc3RyID0gcy5nZXQoImRhdGUiLCIiKQogICAgICAgICAg"
    "ICBuYW1lICAgICA9IHMuZ2V0KCJuYW1lIiwgZGF0ZV9zdHIpWzozMF0KICAgICAgICAgICAgY291bnQgICAgPSBzLmdldCgibWVz"
    "c2FnZV9jb3VudCIsIDApCiAgICAgICAgICAgIGl0ZW0gPSBRTGlzdFdpZGdldEl0ZW0oZiJ7ZGF0ZV9zdHJ9XG57bmFtZX0gKHtj"
    "b3VudH0gbXNncykiKQogICAgICAgICAgICBpdGVtLnNldERhdGEoUXQuSXRlbURhdGFSb2xlLlVzZXJSb2xlLCBkYXRlX3N0cikK"
    "ICAgICAgICAgICAgaXRlbS5zZXRUb29sVGlwKGYiRG91YmxlLWNsaWNrIHRvIGxvYWQgc2Vzc2lvbiBmcm9tIHtkYXRlX3N0cn0i"
    "KQogICAgICAgICAgICBzZWxmLl9zZXNzaW9uX2xpc3QuYWRkSXRlbShpdGVtKQoKICAgIGRlZiBzZXRfc2Vzc2lvbl9uYW1lKHNl"
    "bGYsIG5hbWU6IHN0cikgLT4gTm9uZToKICAgICAgICBzZWxmLl9zZXNzaW9uX25hbWUuc2V0VGV4dChuYW1lWzo1MF0gb3IgIk5l"
    "dyBTZXNzaW9uIikKCiAgICBkZWYgc2V0X2F1dG9zYXZlX2luZGljYXRvcihzZWxmLCBzYXZlZDogYm9vbCkgLT4gTm9uZToKICAg"
    "ICAgICBzZWxmLl9hdXRvc2F2ZV9kb3Quc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJjb2xvcjoge0NfR1JFRU4gaWYgc2F2"
    "ZWQgZWxzZSBDX1RFWFRfRElNfTsgIgogICAgICAgICAgICBmImZvbnQtc2l6ZTogOHB4OyBib3JkZXI6IG5vbmU7IgogICAgICAg"
    "ICkKICAgICAgICBzZWxmLl9hdXRvc2F2ZV9kb3Quc2V0VG9vbFRpcCgKICAgICAgICAgICAgIkF1dG9zYXZlZCIgaWYgc2F2ZWQg"
    "ZWxzZSAiUGVuZGluZyBhdXRvc2F2ZSIKICAgICAgICApCgogICAgZGVmIHNldF9qb3VybmFsX2xvYWRlZChzZWxmLCBkYXRlX3N0"
    "cjogc3RyKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2pvdXJuYWxfbGJsLnNldFRleHQoZiLwn5OWIEpvdXJuYWw6IHtkYXRlX3N0"
    "cn0iKQogICAgICAgIHNlbGYuX2J0bl9jbGVhcl9qb3VybmFsLnNldFZpc2libGUoVHJ1ZSkKCiAgICBkZWYgY2xlYXJfam91cm5h"
    "bF9pbmRpY2F0b3Ioc2VsZikgLT4gTm9uZToKICAgICAgICBzZWxmLl9qb3VybmFsX2xibC5zZXRUZXh0KCIiKQogICAgICAgIHNl"
    "bGYuX2J0bl9jbGVhcl9qb3VybmFsLnNldFZpc2libGUoRmFsc2UpCgogICAgZGVmIF9kb19zYXZlKHNlbGYpIC0+IE5vbmU6CiAg"
    "ICAgICAgc2VsZi5fc2Vzc2lvbl9tZ3Iuc2F2ZSgpCiAgICAgICAgc2VsZi5zZXRfYXV0b3NhdmVfaW5kaWNhdG9yKFRydWUpCiAg"
    "ICAgICAgc2VsZi5yZWZyZXNoKCkKICAgICAgICBzZWxmLl9idG5fc2F2ZS5zZXRUZXh0KCLinJMiKQogICAgICAgIFFUaW1lci5z"
    "aW5nbGVTaG90KDE1MDAsIGxhbWJkYTogc2VsZi5fYnRuX3NhdmUuc2V0VGV4dCgi8J+SviIpKQogICAgICAgIFFUaW1lci5zaW5n"
    "bGVTaG90KDMwMDAsIGxhbWJkYTogc2VsZi5zZXRfYXV0b3NhdmVfaW5kaWNhdG9yKEZhbHNlKSkKCiAgICBkZWYgX2RvX2xvYWQo"
    "c2VsZikgLT4gTm9uZToKICAgICAgICAjIFRyeSBzZWxlY3RlZCBpdGVtIGZpcnN0CiAgICAgICAgaXRlbSA9IHNlbGYuX3Nlc3Np"
    "b25fbGlzdC5jdXJyZW50SXRlbSgpCiAgICAgICAgaWYgbm90IGl0ZW06CiAgICAgICAgICAgICMgSWYgbm90aGluZyBzZWxlY3Rl"
    "ZCwgdHJ5IHRoZSBmaXJzdCBpdGVtCiAgICAgICAgICAgIGlmIHNlbGYuX3Nlc3Npb25fbGlzdC5jb3VudCgpID4gMDoKICAgICAg"
    "ICAgICAgICAgIGl0ZW0gPSBzZWxmLl9zZXNzaW9uX2xpc3QuaXRlbSgwKQogICAgICAgICAgICAgICAgc2VsZi5fc2Vzc2lvbl9s"
    "aXN0LnNldEN1cnJlbnRJdGVtKGl0ZW0pCiAgICAgICAgaWYgaXRlbToKICAgICAgICAgICAgZGF0ZV9zdHIgPSBpdGVtLmRhdGEo"
    "UXQuSXRlbURhdGFSb2xlLlVzZXJSb2xlKQogICAgICAgICAgICBzZWxmLnNlc3Npb25fbG9hZF9yZXF1ZXN0ZWQuZW1pdChkYXRl"
    "X3N0cikKCiAgICBkZWYgX29uX3Nlc3Npb25fY2xpY2soc2VsZiwgaXRlbSkgLT4gTm9uZToKICAgICAgICBkYXRlX3N0ciA9IGl0"
    "ZW0uZGF0YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUpCiAgICAgICAgc2VsZi5zZXNzaW9uX2xvYWRfcmVxdWVzdGVkLmVtaXQo"
    "ZGF0ZV9zdHIpCgogICAgZGVmIF9kb19jbGVhcl9qb3VybmFsKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5zZXNzaW9uX2Ns"
    "ZWFyX3JlcXVlc3RlZC5lbWl0KCkKICAgICAgICBzZWxmLmNsZWFyX2pvdXJuYWxfaW5kaWNhdG9yKCkKCgojIOKUgOKUgCBUT1JQ"
    "T1IgUEFORUwg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNsYXNzIFRvcnBvclBhbmVsKFFXaWRnZXQpOgog"
    "ICAgIiIiCiAgICBUaHJlZS1zdGF0ZSBzdXNwZW5zaW9uIHRvZ2dsZTogQVdBS0UgfCBBVVRPIHwgU1VTUEVORAoKICAgIEFXQUtF"
    "ICDigJQgbW9kZWwgbG9hZGVkLCBhdXRvLXRvcnBvciBkaXNhYmxlZCwgaWdub3JlcyBWUkFNIHByZXNzdXJlCiAgICBBVVRPICAg"
    "4oCUIG1vZGVsIGxvYWRlZCwgbW9uaXRvcnMgVlJBTSBwcmVzc3VyZSwgYXV0by10b3Jwb3IgaWYgc3VzdGFpbmVkCiAgICBTVVNQ"
    "RU5EIOKAlCBtb2RlbCB1bmxvYWRlZCwgc3RheXMgc3VzcGVuZGVkIHVudGlsIG1hbnVhbGx5IGNoYW5nZWQKCiAgICBTaWduYWxz"
    "OgogICAgICAgIHN0YXRlX2NoYW5nZWQoc3RyKSAg4oCUICJBV0FLRSIgfCAiQVVUTyIgfCAiU1VTUEVORCIKICAgICIiIgoKICAg"
    "IHN0YXRlX2NoYW5nZWQgPSBTaWduYWwoc3RyKQoKICAgIFNUQVRFUyA9IFsiQVdBS0UiLCAiQVVUTyIsICJTVVNQRU5EIl0KCiAg"
    "ICBTVEFURV9TVFlMRVMgPSB7CiAgICAgICAgIkFXQUtFIjogewogICAgICAgICAgICAiYWN0aXZlIjogICBmImJhY2tncm91bmQ6"
    "ICMyYTFhMDU7IGNvbG9yOiB7Q19HT0xEfTsgIgogICAgICAgICAgICAgICAgICAgICAgICBmImJvcmRlcjogMXB4IHNvbGlkIHtD"
    "X0dPTER9OyBib3JkZXItcmFkaXVzOiAycHg7ICIKICAgICAgICAgICAgICAgICAgICAgICAgZiJmb250LXNpemU6IDlweDsgZm9u"
    "dC13ZWlnaHQ6IGJvbGQ7IHBhZGRpbmc6IDNweCA4cHg7IiwKICAgICAgICAgICAgImluYWN0aXZlIjogZiJiYWNrZ3JvdW5kOiB7"
    "Q19CRzN9OyBjb2xvcjoge0NfVEVYVF9ESU19OyAiCiAgICAgICAgICAgICAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQg"
    "e0NfQk9SREVSfTsgYm9yZGVyLXJhZGl1czogMnB4OyAiCiAgICAgICAgICAgICAgICAgICAgICAgIGYiZm9udC1zaXplOiA5cHg7"
    "IGZvbnQtd2VpZ2h0OiBib2xkOyBwYWRkaW5nOiAzcHggOHB4OyIsCiAgICAgICAgICAgICJsYWJlbCI6ICAgICLimIAgQVdBS0Ui"
    "LAogICAgICAgICAgICAidG9vbHRpcCI6ICAiTW9kZWwgYWN0aXZlLiBBdXRvLXRvcnBvciBkaXNhYmxlZC4iLAogICAgICAgIH0s"
    "CiAgICAgICAgIkFVVE8iOiB7CiAgICAgICAgICAgICJhY3RpdmUiOiAgIGYiYmFja2dyb3VuZDogIzFhMTAwNTsgY29sb3I6ICNj"
    "Yzg4MjI7ICIKICAgICAgICAgICAgICAgICAgICAgICAgZiJib3JkZXI6IDFweCBzb2xpZCAjY2M4ODIyOyBib3JkZXItcmFkaXVz"
    "OiAycHg7ICIKICAgICAgICAgICAgICAgICAgICAgICAgZiJmb250LXNpemU6IDlweDsgZm9udC13ZWlnaHQ6IGJvbGQ7IHBhZGRp"
    "bmc6IDNweCA4cHg7IiwKICAgICAgICAgICAgImluYWN0aXZlIjogZiJiYWNrZ3JvdW5kOiB7Q19CRzN9OyBjb2xvcjoge0NfVEVY"
    "VF9ESU19OyAiCiAgICAgICAgICAgICAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsgYm9yZGVyLXJh"
    "ZGl1czogMnB4OyAiCiAgICAgICAgICAgICAgICAgICAgICAgIGYiZm9udC1zaXplOiA5cHg7IGZvbnQtd2VpZ2h0OiBib2xkOyBw"
    "YWRkaW5nOiAzcHggOHB4OyIsCiAgICAgICAgICAgICJsYWJlbCI6ICAgICLil4kgQVVUTyIsCiAgICAgICAgICAgICJ0b29sdGlw"
    "IjogICJNb2RlbCBhY3RpdmUuIEF1dG8tc3VzcGVuZCBvbiBWUkFNIHByZXNzdXJlLiIsCiAgICAgICAgfSwKICAgICAgICAiU1VT"
    "UEVORCI6IHsKICAgICAgICAgICAgImFjdGl2ZSI6ICAgZiJiYWNrZ3JvdW5kOiB7Q19QVVJQTEVfRElNfTsgY29sb3I6IHtDX1BV"
    "UlBMRX07ICIKICAgICAgICAgICAgICAgICAgICAgICAgZiJib3JkZXI6IDFweCBzb2xpZCB7Q19QVVJQTEV9OyBib3JkZXItcmFk"
    "aXVzOiAycHg7ICIKICAgICAgICAgICAgICAgICAgICAgICAgZiJmb250LXNpemU6IDlweDsgZm9udC13ZWlnaHQ6IGJvbGQ7IHBh"
    "ZGRpbmc6IDNweCA4cHg7IiwKICAgICAgICAgICAgImluYWN0aXZlIjogZiJiYWNrZ3JvdW5kOiB7Q19CRzN9OyBjb2xvcjoge0Nf"
    "VEVYVF9ESU19OyAiCiAgICAgICAgICAgICAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsgYm9yZGVy"
    "LXJhZGl1czogMnB4OyAiCiAgICAgICAgICAgICAgICAgICAgICAgIGYiZm9udC1zaXplOiA5cHg7IGZvbnQtd2VpZ2h0OiBib2xk"
    "OyBwYWRkaW5nOiAzcHggOHB4OyIsCiAgICAgICAgICAgICJsYWJlbCI6ICAgIGYi4pqwIHtVSV9TVVNQRU5TSU9OX0xBQkVMLnN0"
    "cmlwKCkgaWYgc3RyKFVJX1NVU1BFTlNJT05fTEFCRUwpLnN0cmlwKCkgZWxzZSAnU3VzcGVuZCd9IiwKICAgICAgICAgICAgInRv"
    "b2x0aXAiOiAgZiJNb2RlbCB1bmxvYWRlZC4ge0RFQ0tfTkFNRX0gc2xlZXBzIHVudGlsIG1hbnVhbGx5IGF3YWtlbmVkLiIsCiAg"
    "ICAgICAgfSwKICAgIH0KCiAgICBkZWYgX19pbml0X18oc2VsZiwgcGFyZW50PU5vbmUpOgogICAgICAgIHN1cGVyKCkuX19pbml0"
    "X18ocGFyZW50KQogICAgICAgIHNlbGYuX2N1cnJlbnQgPSAiQVdBS0UiCiAgICAgICAgc2VsZi5fYnV0dG9uczogZGljdFtzdHIs"
    "IFFQdXNoQnV0dG9uXSA9IHt9CiAgICAgICAgbGF5b3V0ID0gUUhCb3hMYXlvdXQoc2VsZikKICAgICAgICBsYXlvdXQuc2V0Q29u"
    "dGVudHNNYXJnaW5zKDAsIDAsIDAsIDApCiAgICAgICAgbGF5b3V0LnNldFNwYWNpbmcoMikKCiAgICAgICAgZm9yIHN0YXRlIGlu"
    "IHNlbGYuU1RBVEVTOgogICAgICAgICAgICBidG4gPSBRUHVzaEJ1dHRvbihzZWxmLlNUQVRFX1NUWUxFU1tzdGF0ZV1bImxhYmVs"
    "Il0pCiAgICAgICAgICAgIGJ0bi5zZXRUb29sVGlwKHNlbGYuU1RBVEVfU1RZTEVTW3N0YXRlXVsidG9vbHRpcCJdKQogICAgICAg"
    "ICAgICBidG4uc2V0Rml4ZWRIZWlnaHQoMjIpCiAgICAgICAgICAgIGJ0bi5jbGlja2VkLmNvbm5lY3QobGFtYmRhIGNoZWNrZWQs"
    "IHM9c3RhdGU6IHNlbGYuX3NldF9zdGF0ZShzKSkKICAgICAgICAgICAgc2VsZi5fYnV0dG9uc1tzdGF0ZV0gPSBidG4KICAgICAg"
    "ICAgICAgbGF5b3V0LmFkZFdpZGdldChidG4pCgogICAgICAgIHNlbGYuX2FwcGx5X3N0eWxlcygpCgogICAgZGVmIF9zZXRfc3Rh"
    "dGUoc2VsZiwgc3RhdGU6IHN0cikgLT4gTm9uZToKICAgICAgICBpZiBzdGF0ZSA9PSBzZWxmLl9jdXJyZW50OgogICAgICAgICAg"
    "ICByZXR1cm4KICAgICAgICBzZWxmLl9jdXJyZW50ID0gc3RhdGUKICAgICAgICBzZWxmLl9hcHBseV9zdHlsZXMoKQogICAgICAg"
    "IHNlbGYuc3RhdGVfY2hhbmdlZC5lbWl0KHN0YXRlKQoKICAgIGRlZiBfYXBwbHlfc3R5bGVzKHNlbGYpIC0+IE5vbmU6CiAgICAg"
    "ICAgZm9yIHN0YXRlLCBidG4gaW4gc2VsZi5fYnV0dG9ucy5pdGVtcygpOgogICAgICAgICAgICBzdHlsZV9rZXkgPSAiYWN0aXZl"
    "IiBpZiBzdGF0ZSA9PSBzZWxmLl9jdXJyZW50IGVsc2UgImluYWN0aXZlIgogICAgICAgICAgICBidG4uc2V0U3R5bGVTaGVldChz"
    "ZWxmLlNUQVRFX1NUWUxFU1tzdGF0ZV1bc3R5bGVfa2V5XSkKCiAgICBAcHJvcGVydHkKICAgIGRlZiBjdXJyZW50X3N0YXRlKHNl"
    "bGYpIC0+IHN0cjoKICAgICAgICByZXR1cm4gc2VsZi5fY3VycmVudAoKICAgIGRlZiBzZXRfc3RhdGUoc2VsZiwgc3RhdGU6IHN0"
    "cikgLT4gTm9uZToKICAgICAgICAiIiJTZXQgc3RhdGUgcHJvZ3JhbW1hdGljYWxseSAoZS5nLiBmcm9tIGF1dG8tdG9ycG9yIGRl"
    "dGVjdGlvbikuIiIiCiAgICAgICAgaWYgc3RhdGUgaW4gc2VsZi5TVEFURVM6CiAgICAgICAgICAgIHNlbGYuX3NldF9zdGF0ZShz"
    "dGF0ZSkKCgpjbGFzcyBTZXR0aW5nc1NlY3Rpb24oUVdpZGdldCk6CiAgICAiIiJTaW1wbGUgY29sbGFwc2libGUgc2VjdGlvbiB1"
    "c2VkIGJ5IFNldHRpbmdzVGFiLiIiIgoKICAgIGRlZiBfX2luaXRfXyhzZWxmLCB0aXRsZTogc3RyLCBwYXJlbnQ9Tm9uZSwgZXhw"
    "YW5kZWQ6IGJvb2wgPSBUcnVlKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKHBhcmVudCkKICAgICAgICBzZWxmLl9leHBhbmRl"
    "ZCA9IGV4cGFuZGVkCgogICAgICAgIHJvb3QgPSBRVkJveExheW91dChzZWxmKQogICAgICAgIHJvb3Quc2V0Q29udGVudHNNYXJn"
    "aW5zKDAsIDAsIDAsIDApCiAgICAgICAgcm9vdC5zZXRTcGFjaW5nKDApCgogICAgICAgIHNlbGYuX2hlYWRlcl9idG4gPSBRVG9v"
    "bEJ1dHRvbigpCiAgICAgICAgc2VsZi5faGVhZGVyX2J0bi5zZXRUZXh0KGYi4pa8IHt0aXRsZX0iIGlmIGV4cGFuZGVkIGVsc2Ug"
    "ZiLilrYge3RpdGxlfSIpCiAgICAgICAgc2VsZi5faGVhZGVyX2J0bi5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tn"
    "cm91bmQ6IHtDX0JHM307IGNvbG9yOiB7Q19HT0xEfTsgYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTl9ESU19OyAiCiAgICAg"
    "ICAgICAgIGYicGFkZGluZzogNnB4OyB0ZXh0LWFsaWduOiBsZWZ0OyBmb250LXdlaWdodDogYm9sZDsiCiAgICAgICAgKQogICAg"
    "ICAgIHNlbGYuX2hlYWRlcl9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYuX3RvZ2dsZSkKCiAgICAgICAgc2VsZi5fY29udGVudCA9"
    "IFFXaWRnZXQoKQogICAgICAgIHNlbGYuX2NvbnRlbnRfbGF5b3V0ID0gUVZCb3hMYXlvdXQoc2VsZi5fY29udGVudCkKICAgICAg"
    "ICBzZWxmLl9jb250ZW50X2xheW91dC5zZXRDb250ZW50c01hcmdpbnMoOCwgOCwgOCwgOCkKICAgICAgICBzZWxmLl9jb250ZW50"
    "X2xheW91dC5zZXRTcGFjaW5nKDgpCiAgICAgICAgc2VsZi5fY29udGVudC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJh"
    "Y2tncm91bmQ6IHtDX0JHMn07IGJvcmRlcjogMXB4IHNvbGlkIHtDX0JPUkRFUn07IGJvcmRlci10b3A6IG5vbmU7IgogICAgICAg"
    "ICkKICAgICAgICBzZWxmLl9jb250ZW50LnNldFZpc2libGUoZXhwYW5kZWQpCgogICAgICAgIHJvb3QuYWRkV2lkZ2V0KHNlbGYu"
    "X2hlYWRlcl9idG4pCiAgICAgICAgcm9vdC5hZGRXaWRnZXQoc2VsZi5fY29udGVudCkKCiAgICBAcHJvcGVydHkKICAgIGRlZiBj"
    "b250ZW50X2xheW91dChzZWxmKSAtPiBRVkJveExheW91dDoKICAgICAgICByZXR1cm4gc2VsZi5fY29udGVudF9sYXlvdXQKCiAg"
    "ICBkZWYgX3RvZ2dsZShzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2V4cGFuZGVkID0gbm90IHNlbGYuX2V4cGFuZGVkCiAg"
    "ICAgICAgc2VsZi5faGVhZGVyX2J0bi5zZXRUZXh0KAogICAgICAgICAgICBzZWxmLl9oZWFkZXJfYnRuLnRleHQoKS5yZXBsYWNl"
    "KCLilrwiLCAi4pa2IiwgMSkKICAgICAgICAgICAgaWYgbm90IHNlbGYuX2V4cGFuZGVkIGVsc2UKICAgICAgICAgICAgc2VsZi5f"
    "aGVhZGVyX2J0bi50ZXh0KCkucmVwbGFjZSgi4pa2IiwgIuKWvCIsIDEpCiAgICAgICAgKQogICAgICAgIHNlbGYuX2NvbnRlbnQu"
    "c2V0VmlzaWJsZShzZWxmLl9leHBhbmRlZCkKCgpjbGFzcyBTZXR0aW5nc1RhYihRV2lkZ2V0KToKICAgICIiIkRlY2std2lkZSBy"
    "dW50aW1lIHNldHRpbmdzIHRhYi4iIiIKCiAgICBkZWYgX19pbml0X18oc2VsZiwgZGVja193aW5kb3c6ICJFY2hvRGVjayIsIHBh"
    "cmVudD1Ob25lKToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKHBhcmVudCkKICAgICAgICBzZWxmLl9kZWNrID0gZGVja193aW5k"
    "b3cKICAgICAgICBzZWxmLl9zZWN0aW9uX3JlZ2lzdHJ5OiBsaXN0W2RpY3RdID0gW10KICAgICAgICBzZWxmLl9zZWN0aW9uX3dp"
    "ZGdldHM6IGRpY3Rbc3RyLCBTZXR0aW5nc1NlY3Rpb25dID0ge30KCiAgICAgICAgcm9vdCA9IFFWQm94TGF5b3V0KHNlbGYpCiAg"
    "ICAgICAgcm9vdC5zZXRDb250ZW50c01hcmdpbnMoMCwgMCwgMCwgMCkKICAgICAgICByb290LnNldFNwYWNpbmcoMCkKCiAgICAg"
    "ICAgc2Nyb2xsID0gUVNjcm9sbEFyZWEoKQogICAgICAgIHNjcm9sbC5zZXRXaWRnZXRSZXNpemFibGUoVHJ1ZSkKICAgICAgICBz"
    "Y3JvbGwuc2V0SG9yaXpvbnRhbFNjcm9sbEJhclBvbGljeShRdC5TY3JvbGxCYXJQb2xpY3kuU2Nyb2xsQmFyQWx3YXlzT2ZmKQog"
    "ICAgICAgIHNjcm9sbC5zZXRTdHlsZVNoZWV0KGYiYmFja2dyb3VuZDoge0NfQkd9OyBib3JkZXI6IDFweCBzb2xpZCB7Q19DUklN"
    "U09OX0RJTX07IikKICAgICAgICByb290LmFkZFdpZGdldChzY3JvbGwpCgogICAgICAgIGJvZHkgPSBRV2lkZ2V0KCkKICAgICAg"
    "ICBzZWxmLl9ib2R5X2xheW91dCA9IFFWQm94TGF5b3V0KGJvZHkpCiAgICAgICAgc2VsZi5fYm9keV9sYXlvdXQuc2V0Q29udGVu"
    "dHNNYXJnaW5zKDYsIDYsIDYsIDYpCiAgICAgICAgc2VsZi5fYm9keV9sYXlvdXQuc2V0U3BhY2luZyg4KQogICAgICAgIHNjcm9s"
    "bC5zZXRXaWRnZXQoYm9keSkKCiAgICAgICAgc2VsZi5fcmVnaXN0ZXJfY29yZV9zZWN0aW9ucygpCgogICAgZGVmIF9yZWdpc3Rl"
    "cl9zZWN0aW9uKHNlbGYsICosIHNlY3Rpb25faWQ6IHN0ciwgdGl0bGU6IHN0ciwgY2F0ZWdvcnk6IHN0ciwgc291cmNlX293bmVy"
    "OiBzdHIsIHNvcnRfa2V5OiBpbnQsIGJ1aWxkZXIpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fc2VjdGlvbl9yZWdpc3RyeS5hcHBl"
    "bmQoewogICAgICAgICAgICAic2VjdGlvbl9pZCI6IHNlY3Rpb25faWQsCiAgICAgICAgICAgICJ0aXRsZSI6IHRpdGxlLAogICAg"
    "ICAgICAgICAiY2F0ZWdvcnkiOiBjYXRlZ29yeSwKICAgICAgICAgICAgInNvdXJjZV9vd25lciI6IHNvdXJjZV9vd25lciwKICAg"
    "ICAgICAgICAgInNvcnRfa2V5Ijogc29ydF9rZXksCiAgICAgICAgICAgICJidWlsZGVyIjogYnVpbGRlciwKICAgICAgICB9KQoK"
    "ICAgIGRlZiBfcmVnaXN0ZXJfY29yZV9zZWN0aW9ucyhzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX3JlZ2lzdGVyX3NlY3Rp"
    "b24oCiAgICAgICAgICAgIHNlY3Rpb25faWQ9InN5c3RlbV9zZXR0aW5ncyIsCiAgICAgICAgICAgIHRpdGxlPSJTeXN0ZW0gU2V0"
    "dGluZ3MiLAogICAgICAgICAgICBjYXRlZ29yeT0iY29yZSIsCiAgICAgICAgICAgIHNvdXJjZV9vd25lcj0iZGVja19ydW50aW1l"
    "IiwKICAgICAgICAgICAgc29ydF9rZXk9MTAwLAogICAgICAgICAgICBidWlsZGVyPXNlbGYuX2J1aWxkX3N5c3RlbV9zZWN0aW9u"
    "LAogICAgICAgICkKICAgICAgICBzZWxmLl9yZWdpc3Rlcl9zZWN0aW9uKAogICAgICAgICAgICBzZWN0aW9uX2lkPSJpbnRlZ3Jh"
    "dGlvbl9zZXR0aW5ncyIsCiAgICAgICAgICAgIHRpdGxlPSJJbnRlZ3JhdGlvbiBTZXR0aW5ncyIsCiAgICAgICAgICAgIGNhdGVn"
    "b3J5PSJjb3JlIiwKICAgICAgICAgICAgc291cmNlX293bmVyPSJkZWNrX3J1bnRpbWUiLAogICAgICAgICAgICBzb3J0X2tleT0y"
    "MDAsCiAgICAgICAgICAgIGJ1aWxkZXI9c2VsZi5fYnVpbGRfaW50ZWdyYXRpb25fc2VjdGlvbiwKICAgICAgICApCiAgICAgICAg"
    "c2VsZi5fcmVnaXN0ZXJfc2VjdGlvbigKICAgICAgICAgICAgc2VjdGlvbl9pZD0idWlfc2V0dGluZ3MiLAogICAgICAgICAgICB0"
    "aXRsZT0iVUkgU2V0dGluZ3MiLAogICAgICAgICAgICBjYXRlZ29yeT0iY29yZSIsCiAgICAgICAgICAgIHNvdXJjZV9vd25lcj0i"
    "ZGVja19ydW50aW1lIiwKICAgICAgICAgICAgc29ydF9rZXk9MzAwLAogICAgICAgICAgICBidWlsZGVyPXNlbGYuX2J1aWxkX3Vp"
    "X3NlY3Rpb24sCiAgICAgICAgKQoKICAgICAgICBmb3IgbWV0YSBpbiBzb3J0ZWQoc2VsZi5fc2VjdGlvbl9yZWdpc3RyeSwga2V5"
    "PWxhbWJkYSBtOiBtLmdldCgic29ydF9rZXkiLCA5OTk5KSk6CiAgICAgICAgICAgIHNlY3Rpb24gPSBTZXR0aW5nc1NlY3Rpb24o"
    "bWV0YVsidGl0bGUiXSwgZXhwYW5kZWQ9VHJ1ZSkKICAgICAgICAgICAgc2VsZi5fYm9keV9sYXlvdXQuYWRkV2lkZ2V0KHNlY3Rp"
    "b24pCiAgICAgICAgICAgIHNlbGYuX3NlY3Rpb25fd2lkZ2V0c1ttZXRhWyJzZWN0aW9uX2lkIl1dID0gc2VjdGlvbgogICAgICAg"
    "ICAgICBtZXRhWyJidWlsZGVyIl0oc2VjdGlvbi5jb250ZW50X2xheW91dCkKCiAgICAgICAgc2VsZi5fYm9keV9sYXlvdXQuYWRk"
    "U3RyZXRjaCgxKQoKICAgIGRlZiBfYnVpbGRfc3lzdGVtX3NlY3Rpb24oc2VsZiwgbGF5b3V0OiBRVkJveExheW91dCkgLT4gTm9u"
    "ZToKICAgICAgICBpZiBzZWxmLl9kZWNrLl90b3Jwb3JfcGFuZWwgaXMgbm90IE5vbmU6CiAgICAgICAgICAgIGxheW91dC5hZGRX"
    "aWRnZXQoUUxhYmVsKCJPcGVyYXRpb25hbCBNb2RlIikpCiAgICAgICAgICAgIGxheW91dC5hZGRXaWRnZXQoc2VsZi5fZGVjay5f"
    "dG9ycG9yX3BhbmVsKQoKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KFFMYWJlbCgiSWRsZSIpKQogICAgICAgIGxheW91dC5hZGRX"
    "aWRnZXQoc2VsZi5fZGVjay5faWRsZV9idG4pCgogICAgICAgIHNldHRpbmdzID0gQ0ZHLmdldCgic2V0dGluZ3MiLCB7fSkKICAg"
    "ICAgICB0el9hdXRvID0gYm9vbChzZXR0aW5ncy5nZXQoInRpbWV6b25lX2F1dG9fZGV0ZWN0IiwgVHJ1ZSkpCiAgICAgICAgdHpf"
    "b3ZlcnJpZGUgPSBzdHIoc2V0dGluZ3MuZ2V0KCJ0aW1lem9uZV9vdmVycmlkZSIsICIiKSBvciAiIikuc3RyaXAoKQoKICAgICAg"
    "ICB0el9hdXRvX2NoayA9IFFDaGVja0JveCgiQXV0by1kZXRlY3QgbG9jYWwvc3lzdGVtIHRpbWUgem9uZSIpCiAgICAgICAgdHpf"
    "YXV0b19jaGsuc2V0Q2hlY2tlZCh0el9hdXRvKQogICAgICAgIHR6X2F1dG9fY2hrLnRvZ2dsZWQuY29ubmVjdChzZWxmLl9kZWNr"
    "Ll9zZXRfdGltZXpvbmVfYXV0b19kZXRlY3QpCiAgICAgICAgbGF5b3V0LmFkZFdpZGdldCh0el9hdXRvX2NoaykKCiAgICAgICAg"
    "dHpfcm93ID0gUUhCb3hMYXlvdXQoKQogICAgICAgIHR6X3Jvdy5hZGRXaWRnZXQoUUxhYmVsKCJNYW51YWwgVGltZSBab25lIE92"
    "ZXJyaWRlOiIpKQogICAgICAgIHR6X2NvbWJvID0gUUNvbWJvQm94KCkKICAgICAgICB0el9jb21iby5zZXRFZGl0YWJsZShUcnVl"
    "KQogICAgICAgIHR6X29wdGlvbnMgPSBbCiAgICAgICAgICAgICJBbWVyaWNhL0NoaWNhZ28iLCAiQW1lcmljYS9OZXdfWW9yayIs"
    "ICJBbWVyaWNhL0xvc19BbmdlbGVzIiwKICAgICAgICAgICAgIkFtZXJpY2EvRGVudmVyIiwgIlVUQyIKICAgICAgICBdCiAgICAg"
    "ICAgdHpfY29tYm8uYWRkSXRlbXModHpfb3B0aW9ucykKICAgICAgICBpZiB0el9vdmVycmlkZToKICAgICAgICAgICAgaWYgdHpf"
    "Y29tYm8uZmluZFRleHQodHpfb3ZlcnJpZGUpIDwgMDoKICAgICAgICAgICAgICAgIHR6X2NvbWJvLmFkZEl0ZW0odHpfb3ZlcnJp"
    "ZGUpCiAgICAgICAgICAgIHR6X2NvbWJvLnNldEN1cnJlbnRUZXh0KHR6X292ZXJyaWRlKQogICAgICAgIGVsc2U6CiAgICAgICAg"
    "ICAgIHR6X2NvbWJvLnNldEN1cnJlbnRUZXh0KCJBbWVyaWNhL0NoaWNhZ28iKQogICAgICAgIHR6X2NvbWJvLnNldEVuYWJsZWQo"
    "bm90IHR6X2F1dG8pCiAgICAgICAgdHpfY29tYm8uY3VycmVudFRleHRDaGFuZ2VkLmNvbm5lY3Qoc2VsZi5fZGVjay5fc2V0X3Rp"
    "bWV6b25lX292ZXJyaWRlKQogICAgICAgIHR6X2F1dG9fY2hrLnRvZ2dsZWQuY29ubmVjdChsYW1iZGEgZW5hYmxlZDogdHpfY29t"
    "Ym8uc2V0RW5hYmxlZChub3QgZW5hYmxlZCkpCiAgICAgICAgdHpfcm93LmFkZFdpZGdldCh0el9jb21ibywgMSkKICAgICAgICB0"
    "el9ob3N0ID0gUVdpZGdldCgpCiAgICAgICAgdHpfaG9zdC5zZXRMYXlvdXQodHpfcm93KQogICAgICAgIGxheW91dC5hZGRXaWRn"
    "ZXQodHpfaG9zdCkKCiAgICBkZWYgX2J1aWxkX2ludGVncmF0aW9uX3NlY3Rpb24oc2VsZiwgbGF5b3V0OiBRVkJveExheW91dCkg"
    "LT4gTm9uZToKICAgICAgICBzZXR0aW5ncyA9IENGRy5nZXQoInNldHRpbmdzIiwge30pCiAgICAgICAgZW1haWxfbWludXRlcyA9"
    "IG1heCgxLCBpbnQoc2V0dGluZ3MuZ2V0KCJlbWFpbF9yZWZyZXNoX2ludGVydmFsX21zIiwgMzAwMDAwKSkgLy8gNjAwMDApCgoK"
    "ICAgICAgICBlbWFpbF9yb3cgPSBRSEJveExheW91dCgpCiAgICAgICAgZW1haWxfcm93LmFkZFdpZGdldChRTGFiZWwoIkVtYWls"
    "IHJlZnJlc2ggaW50ZXJ2YWwgKG1pbnV0ZXMpOiIpKQogICAgICAgIGVtYWlsX2JveCA9IFFDb21ib0JveCgpCiAgICAgICAgZW1h"
    "aWxfYm94LnNldEVkaXRhYmxlKFRydWUpCiAgICAgICAgZW1haWxfYm94LmFkZEl0ZW1zKFsiMSIsICI1IiwgIjEwIiwgIjE1Iiwg"
    "IjMwIiwgIjYwIl0pCiAgICAgICAgZW1haWxfYm94LnNldEN1cnJlbnRUZXh0KHN0cihlbWFpbF9taW51dGVzKSkKICAgICAgICBl"
    "bWFpbF9ib3guY3VycmVudFRleHRDaGFuZ2VkLmNvbm5lY3Qoc2VsZi5fZGVjay5fc2V0X2VtYWlsX3JlZnJlc2hfbWludXRlc19m"
    "cm9tX3RleHQpCiAgICAgICAgZW1haWxfcm93LmFkZFdpZGdldChlbWFpbF9ib3gsIDEpCiAgICAgICAgZW1haWxfaG9zdCA9IFFX"
    "aWRnZXQoKQogICAgICAgIGVtYWlsX2hvc3Quc2V0TGF5b3V0KGVtYWlsX3JvdykKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KGVt"
    "YWlsX2hvc3QpCgogICAgICAgIG5vdGUgPSBRTGFiZWwoIkVtYWlsIHBvbGxpbmcgZm91bmRhdGlvbiBpcyBjb25maWd1cmF0aW9u"
    "LW9ubHkgdW5sZXNzIGFuIGVtYWlsIGJhY2tlbmQgaXMgZW5hYmxlZC4iKQogICAgICAgIG5vdGUuc2V0U3R5bGVTaGVldChmImNv"
    "bG9yOiB7Q19URVhUX0RJTX07IGZvbnQtc2l6ZTogOXB4OyIpCiAgICAgICAgbGF5b3V0LmFkZFdpZGdldChub3RlKQoKICAgIGRl"
    "ZiBfYnVpbGRfdWlfc2VjdGlvbihzZWxmLCBsYXlvdXQ6IFFWQm94TGF5b3V0KSAtPiBOb25lOgogICAgICAgIGxheW91dC5hZGRX"
    "aWRnZXQoUUxhYmVsKCJXaW5kb3cgU2hlbGwiKSkKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX2RlY2suX2ZzX2J0bikK"
    "ICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX2RlY2suX2JsX2J0bikKCgpjbGFzcyBEaWNlR2x5cGgoUVdpZGdldCk6CiAg"
    "ICAiIiJTaW1wbGUgMkQgc2lsaG91ZXR0ZSByZW5kZXJlciBmb3IgZGllLXR5cGUgcmVjb2duaXRpb24uIiIiCiAgICBkZWYgX19p"
    "bml0X18oc2VsZiwgZGllX3R5cGU6IHN0ciA9ICJkMjAiLCBwYXJlbnQ9Tm9uZSk6CiAgICAgICAgc3VwZXIoKS5fX2luaXRfXyhw"
    "YXJlbnQpCiAgICAgICAgc2VsZi5fZGllX3R5cGUgPSBkaWVfdHlwZQogICAgICAgIHNlbGYuc2V0TWluaW11bVNpemUoNzAsIDcw"
    "KQogICAgICAgIHNlbGYuc2V0TWF4aW11bVNpemUoOTAsIDkwKQoKICAgIGRlZiBzZXRfZGllX3R5cGUoc2VsZiwgZGllX3R5cGU6"
    "IHN0cikgLT4gTm9uZToKICAgICAgICBzZWxmLl9kaWVfdHlwZSA9IGRpZV90eXBlCiAgICAgICAgc2VsZi51cGRhdGUoKQoKICAg"
    "IGRlZiBwYWludEV2ZW50KHNlbGYsIGV2ZW50KToKICAgICAgICBwYWludGVyID0gUVBhaW50ZXIoc2VsZikKICAgICAgICBwYWlu"
    "dGVyLnNldFJlbmRlckhpbnQoUVBhaW50ZXIuUmVuZGVySGludC5BbnRpYWxpYXNpbmcpCiAgICAgICAgcmVjdCA9IHNlbGYucmVj"
    "dCgpLmFkanVzdGVkKDgsIDgsIC04LCAtOCkKCiAgICAgICAgZGllID0gc2VsZi5fZGllX3R5cGUKICAgICAgICBsaW5lID0gUUNv"
    "bG9yKENfR09MRCkKICAgICAgICBmaWxsID0gUUNvbG9yKENfQkcyKQogICAgICAgIGFjY2VudCA9IFFDb2xvcihDX0NSSU1TT04p"
    "CgogICAgICAgIHBhaW50ZXIuc2V0UGVuKFFQZW4obGluZSwgMikpCiAgICAgICAgcGFpbnRlci5zZXRCcnVzaChmaWxsKQoKICAg"
    "ICAgICBwdHMgPSBbXQogICAgICAgIGlmIGRpZSA9PSAiZDQiOgogICAgICAgICAgICBwdHMgPSBbCiAgICAgICAgICAgICAgICBR"
    "UG9pbnQocmVjdC5jZW50ZXIoKS54KCksIHJlY3QudG9wKCkpLAogICAgICAgICAgICAgICAgUVBvaW50KHJlY3QubGVmdCgpLCBy"
    "ZWN0LmJvdHRvbSgpKSwKICAgICAgICAgICAgICAgIFFQb2ludChyZWN0LnJpZ2h0KCksIHJlY3QuYm90dG9tKCkpLAogICAgICAg"
    "ICAgICBdCiAgICAgICAgZWxpZiBkaWUgPT0gImQ2IjoKICAgICAgICAgICAgcGFpbnRlci5kcmF3Um91bmRlZFJlY3QocmVjdCwg"
    "NCwgNCkKICAgICAgICBlbGlmIGRpZSA9PSAiZDgiOgogICAgICAgICAgICBwdHMgPSBbCiAgICAgICAgICAgICAgICBRUG9pbnQo"
    "cmVjdC5jZW50ZXIoKS54KCksIHJlY3QudG9wKCkpLAogICAgICAgICAgICAgICAgUVBvaW50KHJlY3QubGVmdCgpLCByZWN0LmNl"
    "bnRlcigpLnkoKSksCiAgICAgICAgICAgICAgICBRUG9pbnQocmVjdC5jZW50ZXIoKS54KCksIHJlY3QuYm90dG9tKCkpLAogICAg"
    "ICAgICAgICAgICAgUVBvaW50KHJlY3QucmlnaHQoKSwgcmVjdC5jZW50ZXIoKS55KCkpLAogICAgICAgICAgICBdCiAgICAgICAg"
    "ZWxpZiBkaWUgaW4gKCJkMTAiLCAiZDEwMCIpOgogICAgICAgICAgICBwdHMgPSBbCiAgICAgICAgICAgICAgICBRUG9pbnQocmVj"
    "dC5jZW50ZXIoKS54KCksIHJlY3QudG9wKCkpLAogICAgICAgICAgICAgICAgUVBvaW50KHJlY3QubGVmdCgpICsgOCwgcmVjdC50"
    "b3AoKSArIDE2KSwKICAgICAgICAgICAgICAgIFFQb2ludChyZWN0LmxlZnQoKSwgcmVjdC5ib3R0b20oKSAtIDEyKSwKICAgICAg"
    "ICAgICAgICAgIFFQb2ludChyZWN0LmNlbnRlcigpLngoKSwgcmVjdC5ib3R0b20oKSksCiAgICAgICAgICAgICAgICBRUG9pbnQo"
    "cmVjdC5yaWdodCgpLCByZWN0LmJvdHRvbSgpIC0gMTIpLAogICAgICAgICAgICAgICAgUVBvaW50KHJlY3QucmlnaHQoKSAtIDgs"
    "IHJlY3QudG9wKCkgKyAxNiksCiAgICAgICAgICAgIF0KICAgICAgICBlbGlmIGRpZSA9PSAiZDEyIjoKICAgICAgICAgICAgY3gg"
    "PSByZWN0LmNlbnRlcigpLngoKTsgY3kgPSByZWN0LmNlbnRlcigpLnkoKQogICAgICAgICAgICByeCA9IHJlY3Qud2lkdGgoKSAv"
    "IDI7IHJ5ID0gcmVjdC5oZWlnaHQoKSAvIDIKICAgICAgICAgICAgZm9yIGkgaW4gcmFuZ2UoNSk6CiAgICAgICAgICAgICAgICBh"
    "ID0gKG1hdGgucGkgKiAyICogaSAvIDUpIC0gKG1hdGgucGkgLyAyKQogICAgICAgICAgICAgICAgcHRzLmFwcGVuZChRUG9pbnQo"
    "aW50KGN4ICsgcnggKiBtYXRoLmNvcyhhKSksIGludChjeSArIHJ5ICogbWF0aC5zaW4oYSkpKSkKICAgICAgICBlbHNlOiAgIyBk"
    "MjAKICAgICAgICAgICAgcHRzID0gWwogICAgICAgICAgICAgICAgUVBvaW50KHJlY3QuY2VudGVyKCkueCgpLCByZWN0LnRvcCgp"
    "KSwKICAgICAgICAgICAgICAgIFFQb2ludChyZWN0LmxlZnQoKSArIDEwLCByZWN0LnRvcCgpICsgMTQpLAogICAgICAgICAgICAg"
    "ICAgUVBvaW50KHJlY3QubGVmdCgpLCByZWN0LmNlbnRlcigpLnkoKSksCiAgICAgICAgICAgICAgICBRUG9pbnQocmVjdC5sZWZ0"
    "KCkgKyAxMCwgcmVjdC5ib3R0b20oKSAtIDE0KSwKICAgICAgICAgICAgICAgIFFQb2ludChyZWN0LmNlbnRlcigpLngoKSwgcmVj"
    "dC5ib3R0b20oKSksCiAgICAgICAgICAgICAgICBRUG9pbnQocmVjdC5yaWdodCgpIC0gMTAsIHJlY3QuYm90dG9tKCkgLSAxNCks"
    "CiAgICAgICAgICAgICAgICBRUG9pbnQocmVjdC5yaWdodCgpLCByZWN0LmNlbnRlcigpLnkoKSksCiAgICAgICAgICAgICAgICBR"
    "UG9pbnQocmVjdC5yaWdodCgpIC0gMTAsIHJlY3QudG9wKCkgKyAxNCksCiAgICAgICAgICAgIF0KCiAgICAgICAgaWYgcHRzOgog"
    "ICAgICAgICAgICBwYXRoID0gUVBhaW50ZXJQYXRoKCkKICAgICAgICAgICAgcGF0aC5tb3ZlVG8ocHRzWzBdKQogICAgICAgICAg"
    "ICBmb3IgcCBpbiBwdHNbMTpdOgogICAgICAgICAgICAgICAgcGF0aC5saW5lVG8ocCkKICAgICAgICAgICAgcGF0aC5jbG9zZVN1"
    "YnBhdGgoKQogICAgICAgICAgICBwYWludGVyLmRyYXdQYXRoKHBhdGgpCgogICAgICAgIHBhaW50ZXIuc2V0UGVuKFFQZW4oYWNj"
    "ZW50LCAxKSkKICAgICAgICB0eHQgPSAiJSIgaWYgZGllID09ICJkMTAwIiBlbHNlIGRpZS5yZXBsYWNlKCJkIiwgIiIpCiAgICAg"
    "ICAgcGFpbnRlci5zZXRGb250KFFGb250KERFQ0tfRk9OVCwgMTIsIFFGb250LldlaWdodC5Cb2xkKSkKICAgICAgICBwYWludGVy"
    "LmRyYXdUZXh0KHJlY3QsIFF0LkFsaWdubWVudEZsYWcuQWxpZ25DZW50ZXIsIHR4dCkKCgpjbGFzcyBEaWNlVHJheURpZShRRnJh"
    "bWUpOgogICAgc2luZ2xlQ2xpY2tlZCA9IFNpZ25hbChzdHIpCiAgICBkb3VibGVDbGlja2VkID0gU2lnbmFsKHN0cikKCiAgICBk"
    "ZWYgX19pbml0X18oc2VsZiwgZGllX3R5cGU6IHN0ciwgZGlzcGxheV9sYWJlbDogc3RyLCBwYXJlbnQ9Tm9uZSk6CiAgICAgICAg"
    "c3VwZXIoKS5fX2luaXRfXyhwYXJlbnQpCiAgICAgICAgc2VsZi5kaWVfdHlwZSA9IGRpZV90eXBlCiAgICAgICAgc2VsZi5kaXNw"
    "bGF5X2xhYmVsID0gZGlzcGxheV9sYWJlbAogICAgICAgIHNlbGYuX2NsaWNrX3RpbWVyID0gUVRpbWVyKHNlbGYpCiAgICAgICAg"
    "c2VsZi5fY2xpY2tfdGltZXIuc2V0U2luZ2xlU2hvdChUcnVlKQogICAgICAgIHNlbGYuX2NsaWNrX3RpbWVyLnNldEludGVydmFs"
    "KDIyMCkKICAgICAgICBzZWxmLl9jbGlja190aW1lci50aW1lb3V0LmNvbm5lY3Qoc2VsZi5fZW1pdF9zaW5nbGUpCgogICAgICAg"
    "IHNlbGYuc2V0T2JqZWN0TmFtZSgiRGljZVRyYXlEaWUiKQogICAgICAgIHNlbGYuc2V0Q3Vyc29yKFF0LkN1cnNvclNoYXBlLlBv"
    "aW50aW5nSGFuZEN1cnNvcikKICAgICAgICBzZWxmLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiUUZyYW1lI0RpY2VUcmF5"
    "RGllIHt7IGJhY2tncm91bmQ6IHtDX0JHM307IGJvcmRlcjogMXB4IHNvbGlkIHtDX0JPUkRFUn07IGJvcmRlci1yYWRpdXM6IDhw"
    "eDsgfX0iCiAgICAgICAgICAgIGYiUUZyYW1lI0RpY2VUcmF5RGllOmhvdmVyIHt7IGJvcmRlcjogMXB4IHNvbGlkIHtDX0dPTER9"
    "OyB9fSIKICAgICAgICApCgogICAgICAgIGxheSA9IFFWQm94TGF5b3V0KHNlbGYpCiAgICAgICAgbGF5LnNldENvbnRlbnRzTWFy"
    "Z2lucyg2LCA2LCA2LCA2KQogICAgICAgIGxheS5zZXRTcGFjaW5nKDIpCgogICAgICAgIGdseXBoX2RpZSA9ICJkMTAwIiBpZiBk"
    "aWVfdHlwZSA9PSAiZCUiIGVsc2UgZGllX3R5cGUKICAgICAgICBzZWxmLmdseXBoID0gRGljZUdseXBoKGdseXBoX2RpZSkKICAg"
    "ICAgICBzZWxmLmdseXBoLnNldEZpeGVkU2l6ZSg1NCwgNTQpCiAgICAgICAgc2VsZi5nbHlwaC5zZXRBdHRyaWJ1dGUoUXQuV2lk"
    "Z2V0QXR0cmlidXRlLldBX1RyYW5zcGFyZW50Rm9yTW91c2VFdmVudHMsIFRydWUpCgogICAgICAgIHNlbGYubGJsID0gUUxhYmVs"
    "KGRpc3BsYXlfbGFiZWwpCiAgICAgICAgc2VsZi5sYmwuc2V0QWxpZ25tZW50KFF0LkFsaWdubWVudEZsYWcuQWxpZ25DZW50ZXIp"
    "CiAgICAgICAgc2VsZi5sYmwuc2V0U3R5bGVTaGVldChmImNvbG9yOiB7Q19URVhUfTsgZm9udC13ZWlnaHQ6IGJvbGQ7IikKICAg"
    "ICAgICBzZWxmLmxibC5zZXRBdHRyaWJ1dGUoUXQuV2lkZ2V0QXR0cmlidXRlLldBX1RyYW5zcGFyZW50Rm9yTW91c2VFdmVudHMs"
    "IFRydWUpCgogICAgICAgIGxheS5hZGRXaWRnZXQoc2VsZi5nbHlwaCwgMCwgUXQuQWxpZ25tZW50RmxhZy5BbGlnbkNlbnRlcikK"
    "ICAgICAgICBsYXkuYWRkV2lkZ2V0KHNlbGYubGJsKQoKICAgIGRlZiBtb3VzZVByZXNzRXZlbnQoc2VsZiwgZXZlbnQpOgogICAg"
    "ICAgIGlmIGV2ZW50LmJ1dHRvbigpID09IFF0Lk1vdXNlQnV0dG9uLkxlZnRCdXR0b246CiAgICAgICAgICAgIGlmIHNlbGYuX2Ns"
    "aWNrX3RpbWVyLmlzQWN0aXZlKCk6CiAgICAgICAgICAgICAgICBzZWxmLl9jbGlja190aW1lci5zdG9wKCkKICAgICAgICAgICAg"
    "ICAgIHNlbGYuZG91YmxlQ2xpY2tlZC5lbWl0KHNlbGYuZGllX3R5cGUpCiAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAg"
    "ICBzZWxmLl9jbGlja190aW1lci5zdGFydCgpCiAgICAgICAgICAgIGV2ZW50LmFjY2VwdCgpCiAgICAgICAgICAgIHJldHVybgog"
    "ICAgICAgIHN1cGVyKCkubW91c2VQcmVzc0V2ZW50KGV2ZW50KQoKICAgIGRlZiBfZW1pdF9zaW5nbGUoc2VsZik6CiAgICAgICAg"
    "c2VsZi5zaW5nbGVDbGlja2VkLmVtaXQoc2VsZi5kaWVfdHlwZSkKCgpjbGFzcyBEaWNlUm9sbGVyVGFiKFFXaWRnZXQpOgogICAg"
    "IiIiRGVjay1uYXRpdmUgRGljZSBSb2xsZXIgbW9kdWxlIHRhYiB3aXRoIHRyYXkvcG9vbCB3b3JrZmxvdyBhbmQgc3RydWN0dXJl"
    "ZCByb2xsIGV2ZW50cy4iIiIKCiAgICBUUkFZX09SREVSID0gWyJkNCIsICJkNiIsICJkOCIsICJkMTAiLCAiZDEyIiwgImQyMCIs"
    "ICJkJSJdCgogICAgZGVmIF9faW5pdF9fKHNlbGYsIGRpYWdub3N0aWNzX2xvZ2dlcj1Ob25lKToKICAgICAgICBzdXBlcigpLl9f"
    "aW5pdF9fKCkKICAgICAgICBzZWxmLl9sb2cgPSBkaWFnbm9zdGljc19sb2dnZXIgb3IgKGxhbWJkYSAqX2FyZ3MsICoqX2t3YXJn"
    "czogTm9uZSkKCiAgICAgICAgc2VsZi5yb2xsX2V2ZW50czogbGlzdFtkaWN0XSA9IFtdCiAgICAgICAgc2VsZi5zYXZlZF9yb2xs"
    "czogbGlzdFtkaWN0XSA9IFtdCiAgICAgICAgc2VsZi5jb21tb25fcm9sbHM6IGRpY3Rbc3RyLCBkaWN0XSA9IHt9CiAgICAgICAg"
    "c2VsZi5ldmVudF9ieV9pZDogZGljdFtzdHIsIGRpY3RdID0ge30KICAgICAgICBzZWxmLmN1cnJlbnRfcG9vbDogZGljdFtzdHIs"
    "IGludF0gPSB7fQogICAgICAgIHNlbGYuY3VycmVudF9yb2xsX2lkczogbGlzdFtzdHJdID0gW10KCiAgICAgICAgc2VsZi5ydWxl"
    "X2RlZmluaXRpb25zOiBkaWN0W3N0ciwgZGljdF0gPSB7CiAgICAgICAgICAgICJydWxlXzRkNl9kcm9wX2xvd2VzdCI6IHsKICAg"
    "ICAgICAgICAgICAgICJpZCI6ICJydWxlXzRkNl9kcm9wX2xvd2VzdCIsCiAgICAgICAgICAgICAgICAibmFtZSI6ICJEJkQgNWUg"
    "U3RhdCBSb2xsIiwKICAgICAgICAgICAgICAgICJkaWNlX2NvdW50IjogNCwKICAgICAgICAgICAgICAgICJkaWNlX3NpZGVzIjog"
    "NiwKICAgICAgICAgICAgICAgICJkcm9wX2xvd2VzdF9jb3VudCI6IDEsCiAgICAgICAgICAgICAgICAiZHJvcF9oaWdoZXN0X2Nv"
    "dW50IjogMCwKICAgICAgICAgICAgICAgICJub3RlcyI6ICJSb2xsIDRkNiwgZHJvcCBsb3dlc3Qgb25lLiIKICAgICAgICAgICAg"
    "fSwKICAgICAgICAgICAgInJ1bGVfM2Q2X3N0cmFpZ2h0IjogewogICAgICAgICAgICAgICAgImlkIjogInJ1bGVfM2Q2X3N0cmFp"
    "Z2h0IiwKICAgICAgICAgICAgICAgICJuYW1lIjogIjNkNiBTdHJhaWdodCIsCiAgICAgICAgICAgICAgICAiZGljZV9jb3VudCI6"
    "IDMsCiAgICAgICAgICAgICAgICAiZGljZV9zaWRlcyI6IDYsCiAgICAgICAgICAgICAgICAiZHJvcF9sb3dlc3RfY291bnQiOiAw"
    "LAogICAgICAgICAgICAgICAgImRyb3BfaGlnaGVzdF9jb3VudCI6IDAsCiAgICAgICAgICAgICAgICAibm90ZXMiOiAiQ2xhc3Np"
    "YyAzZDYgcm9sbC4iCiAgICAgICAgICAgIH0sCiAgICAgICAgfQoKICAgICAgICBzZWxmLl9idWlsZF91aSgpCiAgICAgICAgc2Vs"
    "Zi5fcmVmcmVzaF9wb29sX2VkaXRvcigpCiAgICAgICAgc2VsZi5fcmVmcmVzaF9zYXZlZF9saXN0cygpCgogICAgZGVmIF9idWls"
    "ZF91aShzZWxmKSAtPiBOb25lOgogICAgICAgIHJvb3QgPSBRVkJveExheW91dChzZWxmKQogICAgICAgIHJvb3Quc2V0Q29udGVu"
    "dHNNYXJnaW5zKDgsIDgsIDgsIDgpCiAgICAgICAgcm9vdC5zZXRTcGFjaW5nKDYpCgogICAgICAgIHRyYXlfd3JhcCA9IFFGcmFt"
    "ZSgpCiAgICAgICAgdHJheV93cmFwLnNldFN0eWxlU2hlZXQoZiJiYWNrZ3JvdW5kOiB7Q19CRzJ9OyBib3JkZXI6IDFweCBzb2xp"
    "ZCB7Q19CT1JERVJ9OyIpCiAgICAgICAgdHJheV9sYXlvdXQgPSBRVkJveExheW91dCh0cmF5X3dyYXApCiAgICAgICAgdHJheV9s"
    "YXlvdXQuc2V0Q29udGVudHNNYXJnaW5zKDgsIDgsIDgsIDgpCiAgICAgICAgdHJheV9sYXlvdXQuc2V0U3BhY2luZyg2KQogICAg"
    "ICAgIHRyYXlfbGF5b3V0LmFkZFdpZGdldChRTGFiZWwoIkRpY2UgVHJheSIpKQoKICAgICAgICB0cmF5X3JvdyA9IFFIQm94TGF5"
    "b3V0KCkKICAgICAgICB0cmF5X3Jvdy5zZXRTcGFjaW5nKDYpCiAgICAgICAgZm9yIGRpZSBpbiBzZWxmLlRSQVlfT1JERVI6CiAg"
    "ICAgICAgICAgIGJsb2NrID0gRGljZVRyYXlEaWUoZGllLCBkaWUpCiAgICAgICAgICAgIGJsb2NrLnNpbmdsZUNsaWNrZWQuY29u"
    "bmVjdChzZWxmLl9hZGRfZGllX3RvX3Bvb2wpCiAgICAgICAgICAgIGJsb2NrLmRvdWJsZUNsaWNrZWQuY29ubmVjdChzZWxmLl9x"
    "dWlja19yb2xsX3NpbmdsZV9kaWUpCiAgICAgICAgICAgIHRyYXlfcm93LmFkZFdpZGdldChibG9jaywgMSkKICAgICAgICB0cmF5"
    "X2xheW91dC5hZGRMYXlvdXQodHJheV9yb3cpCiAgICAgICAgcm9vdC5hZGRXaWRnZXQodHJheV93cmFwKQoKICAgICAgICBwb29s"
    "X3dyYXAgPSBRRnJhbWUoKQogICAgICAgIHBvb2xfd3JhcC5zZXRTdHlsZVNoZWV0KGYiYmFja2dyb3VuZDoge0NfQkcyfTsgYm9y"
    "ZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsiKQogICAgICAgIHB3ID0gUVZCb3hMYXlvdXQocG9vbF93cmFwKQogICAgICAgIHB3"
    "LnNldENvbnRlbnRzTWFyZ2lucyg4LCA4LCA4LCA4KQogICAgICAgIHB3LnNldFNwYWNpbmcoNikKCiAgICAgICAgcHcuYWRkV2lk"
    "Z2V0KFFMYWJlbCgiQ3VycmVudCBQb29sIikpCiAgICAgICAgc2VsZi5wb29sX2V4cHJfbGJsID0gUUxhYmVsKCJQb29sOiAoZW1w"
    "dHkpIikKICAgICAgICBzZWxmLnBvb2xfZXhwcl9sYmwuc2V0U3R5bGVTaGVldChmImNvbG9yOiB7Q19HT0xEfTsgZm9udC13ZWln"
    "aHQ6IGJvbGQ7IikKICAgICAgICBwdy5hZGRXaWRnZXQoc2VsZi5wb29sX2V4cHJfbGJsKQoKICAgICAgICBzZWxmLnBvb2xfZW50"
    "cmllc193aWRnZXQgPSBRV2lkZ2V0KCkKICAgICAgICBzZWxmLnBvb2xfZW50cmllc19sYXlvdXQgPSBRSEJveExheW91dChzZWxm"
    "LnBvb2xfZW50cmllc193aWRnZXQpCiAgICAgICAgc2VsZi5wb29sX2VudHJpZXNfbGF5b3V0LnNldENvbnRlbnRzTWFyZ2lucygw"
    "LCAwLCAwLCAwKQogICAgICAgIHNlbGYucG9vbF9lbnRyaWVzX2xheW91dC5zZXRTcGFjaW5nKDYpCiAgICAgICAgcHcuYWRkV2lk"
    "Z2V0KHNlbGYucG9vbF9lbnRyaWVzX3dpZGdldCkKCiAgICAgICAgbWV0YV9yb3cgPSBRSEJveExheW91dCgpCiAgICAgICAgc2Vs"
    "Zi5sYWJlbF9lZGl0ID0gUUxpbmVFZGl0KCk7IHNlbGYubGFiZWxfZWRpdC5zZXRQbGFjZWhvbGRlclRleHQoIkxhYmVsIC8gcHVy"
    "cG9zZSIpCiAgICAgICAgc2VsZi5tb2Rfc3BpbiA9IFFTcGluQm94KCk7IHNlbGYubW9kX3NwaW4uc2V0UmFuZ2UoLTk5OSwgOTk5"
    "KTsgc2VsZi5tb2Rfc3Bpbi5zZXRWYWx1ZSgwKQogICAgICAgIHNlbGYucnVsZV9jb21ibyA9IFFDb21ib0JveCgpOyBzZWxmLnJ1"
    "bGVfY29tYm8uYWRkSXRlbSgiTWFudWFsIFJvbGwiLCAiIikKICAgICAgICBmb3IgcmlkLCBtZXRhIGluIHNlbGYucnVsZV9kZWZp"
    "bml0aW9ucy5pdGVtcygpOgogICAgICAgICAgICBzZWxmLnJ1bGVfY29tYm8uYWRkSXRlbShtZXRhLmdldCgibmFtZSIsIHJpZCks"
    "IHJpZCkKCiAgICAgICAgZm9yIHRpdGxlLCB3IGluICgoIkxhYmVsIiwgc2VsZi5sYWJlbF9lZGl0KSwgKCJNb2RpZmllciIsIHNl"
    "bGYubW9kX3NwaW4pLCAoIlJ1bGUiLCBzZWxmLnJ1bGVfY29tYm8pKToKICAgICAgICAgICAgY29sID0gUVZCb3hMYXlvdXQoKQog"
    "ICAgICAgICAgICBsYmwgPSBRTGFiZWwodGl0bGUpCiAgICAgICAgICAgIGxibC5zZXRTdHlsZVNoZWV0KGYiY29sb3I6IHtDX1RF"
    "WFRfRElNfTsgZm9udC1zaXplOiA5cHg7IikKICAgICAgICAgICAgY29sLmFkZFdpZGdldChsYmwpCiAgICAgICAgICAgIGNvbC5h"
    "ZGRXaWRnZXQodykKICAgICAgICAgICAgbWV0YV9yb3cuYWRkTGF5b3V0KGNvbCwgMSkKICAgICAgICBwdy5hZGRMYXlvdXQobWV0"
    "YV9yb3cpCgogICAgICAgIGFjdGlvbnMgPSBRSEJveExheW91dCgpCiAgICAgICAgc2VsZi5yb2xsX3Bvb2xfYnRuID0gUVB1c2hC"
    "dXR0b24oIlJvbGwgUG9vbCIpCiAgICAgICAgc2VsZi5yZXNldF9wb29sX2J0biA9IFFQdXNoQnV0dG9uKCJSZXNldCBQb29sIikK"
    "ICAgICAgICBzZWxmLnNhdmVfcG9vbF9idG4gPSBRUHVzaEJ1dHRvbigiU2F2ZSBQb29sIikKICAgICAgICBhY3Rpb25zLmFkZFdp"
    "ZGdldChzZWxmLnJvbGxfcG9vbF9idG4pCiAgICAgICAgYWN0aW9ucy5hZGRXaWRnZXQoc2VsZi5yZXNldF9wb29sX2J0bikKICAg"
    "ICAgICBhY3Rpb25zLmFkZFdpZGdldChzZWxmLnNhdmVfcG9vbF9idG4pCiAgICAgICAgcHcuYWRkTGF5b3V0KGFjdGlvbnMpCgog"
    "ICAgICAgIHJvb3QuYWRkV2lkZ2V0KHBvb2xfd3JhcCkKCiAgICAgICAgcmVzdWx0X3dyYXAgPSBRRnJhbWUoKQogICAgICAgIHJl"
    "c3VsdF93cmFwLnNldFN0eWxlU2hlZXQoZiJiYWNrZ3JvdW5kOiB7Q19CRzJ9OyBib3JkZXI6IDFweCBzb2xpZCB7Q19CT1JERVJ9"
    "OyIpCiAgICAgICAgcmwgPSBRVkJveExheW91dChyZXN1bHRfd3JhcCkKICAgICAgICBybC5zZXRDb250ZW50c01hcmdpbnMoOCwg"
    "OCwgOCwgOCkKICAgICAgICBybC5hZGRXaWRnZXQoUUxhYmVsKCJDdXJyZW50IFJlc3VsdCIpKQogICAgICAgIHNlbGYuY3VycmVu"
    "dF9yZXN1bHRfbGJsID0gUUxhYmVsKCJObyByb2xsIHlldC4iKQogICAgICAgIHNlbGYuY3VycmVudF9yZXN1bHRfbGJsLnNldFdv"
    "cmRXcmFwKFRydWUpCiAgICAgICAgcmwuYWRkV2lkZ2V0KHNlbGYuY3VycmVudF9yZXN1bHRfbGJsKQogICAgICAgIHJvb3QuYWRk"
    "V2lkZ2V0KHJlc3VsdF93cmFwKQoKICAgICAgICBtaWQgPSBRSEJveExheW91dCgpCiAgICAgICAgaGlzdG9yeV93cmFwID0gUUZy"
    "YW1lKCkKICAgICAgICBoaXN0b3J5X3dyYXAuc2V0U3R5bGVTaGVldChmImJhY2tncm91bmQ6IHtDX0JHMn07IGJvcmRlcjogMXB4"
    "IHNvbGlkIHtDX0JPUkRFUn07IikKICAgICAgICBodyA9IFFWQm94TGF5b3V0KGhpc3Rvcnlfd3JhcCkKICAgICAgICBody5zZXRD"
    "b250ZW50c01hcmdpbnMoNiwgNiwgNiwgNikKCiAgICAgICAgc2VsZi5oaXN0b3J5X3RhYnMgPSBRVGFiV2lkZ2V0KCkKICAgICAg"
    "ICBzZWxmLmN1cnJlbnRfdGFibGUgPSBzZWxmLl9tYWtlX3JvbGxfdGFibGUoKQogICAgICAgIHNlbGYuaGlzdG9yeV90YWJsZSA9"
    "IHNlbGYuX21ha2Vfcm9sbF90YWJsZSgpCiAgICAgICAgc2VsZi5oaXN0b3J5X3RhYnMuYWRkVGFiKHNlbGYuY3VycmVudF90YWJs"
    "ZSwgIkN1cnJlbnQgUm9sbHMiKQogICAgICAgIHNlbGYuaGlzdG9yeV90YWJzLmFkZFRhYihzZWxmLmhpc3RvcnlfdGFibGUsICJS"
    "b2xsIEhpc3RvcnkiKQogICAgICAgIGh3LmFkZFdpZGdldChzZWxmLmhpc3RvcnlfdGFicywgMSkKCiAgICAgICAgaGlzdG9yeV9h"
    "Y3Rpb25zID0gUUhCb3hMYXlvdXQoKQogICAgICAgIHNlbGYuY2xlYXJfaGlzdG9yeV9idG4gPSBRUHVzaEJ1dHRvbigiQ2xlYXIg"
    "Um9sbCBIaXN0b3J5IikKICAgICAgICBoaXN0b3J5X2FjdGlvbnMuYWRkV2lkZ2V0KHNlbGYuY2xlYXJfaGlzdG9yeV9idG4pCiAg"
    "ICAgICAgaGlzdG9yeV9hY3Rpb25zLmFkZFN0cmV0Y2goMSkKICAgICAgICBody5hZGRMYXlvdXQoaGlzdG9yeV9hY3Rpb25zKQoK"
    "ICAgICAgICBzZWxmLmdyYW5kX3RvdGFsX2xibCA9IFFMYWJlbCgiR3JhbmQgVG90YWw6IDAiKQogICAgICAgIHNlbGYuZ3JhbmRf"
    "dG90YWxfbGJsLnNldFN0eWxlU2hlZXQoZiJjb2xvcjoge0NfR09MRH07IGZvbnQtc2l6ZTogMTJweDsgZm9udC13ZWlnaHQ6IGJv"
    "bGQ7IikKICAgICAgICBody5hZGRXaWRnZXQoc2VsZi5ncmFuZF90b3RhbF9sYmwpCgogICAgICAgIHNhdmVkX3dyYXAgPSBRRnJh"
    "bWUoKQogICAgICAgIHNhdmVkX3dyYXAuc2V0U3R5bGVTaGVldChmImJhY2tncm91bmQ6IHtDX0JHMn07IGJvcmRlcjogMXB4IHNv"
    "bGlkIHtDX0JPUkRFUn07IikKICAgICAgICBzdyA9IFFWQm94TGF5b3V0KHNhdmVkX3dyYXApCiAgICAgICAgc3cuc2V0Q29udGVu"
    "dHNNYXJnaW5zKDYsIDYsIDYsIDYpCiAgICAgICAgc3cuYWRkV2lkZ2V0KFFMYWJlbCgiU2F2ZWQgLyBDb21tb24gUm9sbHMiKSkK"
    "CiAgICAgICAgc3cuYWRkV2lkZ2V0KFFMYWJlbCgiU2F2ZWQiKSkKICAgICAgICBzZWxmLnNhdmVkX2xpc3QgPSBRTGlzdFdpZGdl"
    "dCgpCiAgICAgICAgc3cuYWRkV2lkZ2V0KHNlbGYuc2F2ZWRfbGlzdCwgMSkKICAgICAgICBzYXZlZF9hY3Rpb25zID0gUUhCb3hM"
    "YXlvdXQoKQogICAgICAgIHNlbGYucnVuX3NhdmVkX2J0biA9IFFQdXNoQnV0dG9uKCJSdW4iKQogICAgICAgIHNlbGYubG9hZF9z"
    "YXZlZF9idG4gPSBRUHVzaEJ1dHRvbigiTG9hZC9FZGl0IikKICAgICAgICBzZWxmLmRlbGV0ZV9zYXZlZF9idG4gPSBRUHVzaEJ1"
    "dHRvbigiRGVsZXRlIikKICAgICAgICBzYXZlZF9hY3Rpb25zLmFkZFdpZGdldChzZWxmLnJ1bl9zYXZlZF9idG4pCiAgICAgICAg"
    "c2F2ZWRfYWN0aW9ucy5hZGRXaWRnZXQoc2VsZi5sb2FkX3NhdmVkX2J0bikKICAgICAgICBzYXZlZF9hY3Rpb25zLmFkZFdpZGdl"
    "dChzZWxmLmRlbGV0ZV9zYXZlZF9idG4pCiAgICAgICAgc3cuYWRkTGF5b3V0KHNhdmVkX2FjdGlvbnMpCgogICAgICAgIHN3LmFk"
    "ZFdpZGdldChRTGFiZWwoIkF1dG8tRGV0ZWN0ZWQgQ29tbW9uIikpCiAgICAgICAgc2VsZi5jb21tb25fbGlzdCA9IFFMaXN0V2lk"
    "Z2V0KCkKICAgICAgICBzdy5hZGRXaWRnZXQoc2VsZi5jb21tb25fbGlzdCwgMSkKICAgICAgICBjb21tb25fYWN0aW9ucyA9IFFI"
    "Qm94TGF5b3V0KCkKICAgICAgICBzZWxmLnByb21vdGVfY29tbW9uX2J0biA9IFFQdXNoQnV0dG9uKCJQcm9tb3RlIHRvIFNhdmVk"
    "IikKICAgICAgICBzZWxmLmRpc21pc3NfY29tbW9uX2J0biA9IFFQdXNoQnV0dG9uKCJEaXNtaXNzIikKICAgICAgICBjb21tb25f"
    "YWN0aW9ucy5hZGRXaWRnZXQoc2VsZi5wcm9tb3RlX2NvbW1vbl9idG4pCiAgICAgICAgY29tbW9uX2FjdGlvbnMuYWRkV2lkZ2V0"
    "KHNlbGYuZGlzbWlzc19jb21tb25fYnRuKQogICAgICAgIHN3LmFkZExheW91dChjb21tb25fYWN0aW9ucykKCiAgICAgICAgc2Vs"
    "Zi5jb21tb25faGludCA9IFFMYWJlbCgiQ29tbW9uIHNpZ25hdHVyZSB0cmFja2luZyBhY3RpdmUuIikKICAgICAgICBzZWxmLmNv"
    "bW1vbl9oaW50LnNldFN0eWxlU2hlZXQoZiJjb2xvcjoge0NfVEVYVF9ESU19OyBmb250LXNpemU6IDlweDsiKQogICAgICAgIHN3"
    "LmFkZFdpZGdldChzZWxmLmNvbW1vbl9oaW50KQoKICAgICAgICBtaWQuYWRkV2lkZ2V0KGhpc3Rvcnlfd3JhcCwgMykKICAgICAg"
    "ICBtaWQuYWRkV2lkZ2V0KHNhdmVkX3dyYXAsIDIpCiAgICAgICAgcm9vdC5hZGRMYXlvdXQobWlkLCAxKQoKICAgICAgICBzZWxm"
    "LnJvbGxfcG9vbF9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYuX3JvbGxfY3VycmVudF9wb29sKQogICAgICAgIHNlbGYucmVzZXRf"
    "cG9vbF9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYuX3Jlc2V0X3Bvb2wpCiAgICAgICAgc2VsZi5zYXZlX3Bvb2xfYnRuLmNsaWNr"
    "ZWQuY29ubmVjdChzZWxmLl9zYXZlX3Bvb2wpCiAgICAgICAgc2VsZi5jbGVhcl9oaXN0b3J5X2J0bi5jbGlja2VkLmNvbm5lY3Qo"
    "c2VsZi5fY2xlYXJfaGlzdG9yeSkKCiAgICAgICAgc2VsZi5zYXZlZF9saXN0Lml0ZW1Eb3VibGVDbGlja2VkLmNvbm5lY3QobGFt"
    "YmRhIGl0ZW06IHNlbGYuX3J1bl9zYXZlZF9yb2xsKGl0ZW0uZGF0YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUpKSkKICAgICAg"
    "ICBzZWxmLmNvbW1vbl9saXN0Lml0ZW1Eb3VibGVDbGlja2VkLmNvbm5lY3QobGFtYmRhIGl0ZW06IHNlbGYuX3J1bl9zYXZlZF9y"
    "b2xsKGl0ZW0uZGF0YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUpKSkKCiAgICAgICAgc2VsZi5ydW5fc2F2ZWRfYnRuLmNsaWNr"
    "ZWQuY29ubmVjdChzZWxmLl9ydW5fc2VsZWN0ZWRfc2F2ZWQpCiAgICAgICAgc2VsZi5sb2FkX3NhdmVkX2J0bi5jbGlja2VkLmNv"
    "bm5lY3Qoc2VsZi5fbG9hZF9zZWxlY3RlZF9zYXZlZCkKICAgICAgICBzZWxmLmRlbGV0ZV9zYXZlZF9idG4uY2xpY2tlZC5jb25u"
    "ZWN0KHNlbGYuX2RlbGV0ZV9zZWxlY3RlZF9zYXZlZCkKICAgICAgICBzZWxmLnByb21vdGVfY29tbW9uX2J0bi5jbGlja2VkLmNv"
    "bm5lY3Qoc2VsZi5fcHJvbW90ZV9zZWxlY3RlZF9jb21tb24pCiAgICAgICAgc2VsZi5kaXNtaXNzX2NvbW1vbl9idG4uY2xpY2tl"
    "ZC5jb25uZWN0KHNlbGYuX2Rpc21pc3Nfc2VsZWN0ZWRfY29tbW9uKQoKICAgICAgICBzZWxmLmN1cnJlbnRfdGFibGUuc2V0Q29u"
    "dGV4dE1lbnVQb2xpY3koUXQuQ29udGV4dE1lbnVQb2xpY3kuQ3VzdG9tQ29udGV4dE1lbnUpCiAgICAgICAgc2VsZi5oaXN0b3J5"
    "X3RhYmxlLnNldENvbnRleHRNZW51UG9saWN5KFF0LkNvbnRleHRNZW51UG9saWN5LkN1c3RvbUNvbnRleHRNZW51KQogICAgICAg"
    "IHNlbGYuY3VycmVudF90YWJsZS5jdXN0b21Db250ZXh0TWVudVJlcXVlc3RlZC5jb25uZWN0KGxhbWJkYSBwb3M6IHNlbGYuX3No"
    "b3dfcm9sbF9jb250ZXh0X21lbnUoc2VsZi5jdXJyZW50X3RhYmxlLCBwb3MpKQogICAgICAgIHNlbGYuaGlzdG9yeV90YWJsZS5j"
    "dXN0b21Db250ZXh0TWVudVJlcXVlc3RlZC5jb25uZWN0KGxhbWJkYSBwb3M6IHNlbGYuX3Nob3dfcm9sbF9jb250ZXh0X21lbnUo"
    "c2VsZi5oaXN0b3J5X3RhYmxlLCBwb3MpKQoKICAgIGRlZiBfbWFrZV9yb2xsX3RhYmxlKHNlbGYpIC0+IFFUYWJsZVdpZGdldDoK"
    "ICAgICAgICB0YmwgPSBRVGFibGVXaWRnZXQoMCwgNikKICAgICAgICB0Ymwuc2V0SG9yaXpvbnRhbEhlYWRlckxhYmVscyhbIlRp"
    "bWVzdGFtcCIsICJMYWJlbCIsICJFeHByZXNzaW9uIiwgIlJhdyIsICJNb2RpZmllciIsICJUb3RhbCJdKQogICAgICAgIHRibC5o"
    "b3Jpem9udGFsSGVhZGVyKCkuc2V0U2VjdGlvblJlc2l6ZU1vZGUoUUhlYWRlclZpZXcuUmVzaXplTW9kZS5TdHJldGNoKQogICAg"
    "ICAgIHRibC52ZXJ0aWNhbEhlYWRlcigpLnNldFZpc2libGUoRmFsc2UpCiAgICAgICAgdGJsLnNldEVkaXRUcmlnZ2VycyhRQWJz"
    "dHJhY3RJdGVtVmlldy5FZGl0VHJpZ2dlci5Ob0VkaXRUcmlnZ2VycykKICAgICAgICB0Ymwuc2V0U2VsZWN0aW9uQmVoYXZpb3Io"
    "UUFic3RyYWN0SXRlbVZpZXcuU2VsZWN0aW9uQmVoYXZpb3IuU2VsZWN0Um93cykKICAgICAgICB0Ymwuc2V0U29ydGluZ0VuYWJs"
    "ZWQoRmFsc2UpCiAgICAgICAgcmV0dXJuIHRibAoKICAgIGRlZiBfc29ydGVkX3Bvb2xfaXRlbXMoc2VsZik6CiAgICAgICAgcmV0"
    "dXJuIFsoZCwgc2VsZi5jdXJyZW50X3Bvb2wuZ2V0KGQsIDApKSBmb3IgZCBpbiBzZWxmLlRSQVlfT1JERVIgaWYgc2VsZi5jdXJy"
    "ZW50X3Bvb2wuZ2V0KGQsIDApID4gMF0KCiAgICBkZWYgX3Bvb2xfZXhwcmVzc2lvbihzZWxmLCBwb29sOiBkaWN0W3N0ciwgaW50"
    "XSB8IE5vbmUgPSBOb25lKSAtPiBzdHI6CiAgICAgICAgcCA9IHBvb2wgaWYgcG9vbCBpcyBub3QgTm9uZSBlbHNlIHNlbGYuY3Vy"
    "cmVudF9wb29sCiAgICAgICAgcGFydHMgPSBbZiJ7cXR5fXtkaWV9IiBmb3IgZGllLCBxdHkgaW4gWyhkLCBwLmdldChkLCAwKSkg"
    "Zm9yIGQgaW4gc2VsZi5UUkFZX09SREVSXSBpZiBxdHkgPiAwXQogICAgICAgIHJldHVybiAiICsgIi5qb2luKHBhcnRzKSBpZiBw"
    "YXJ0cyBlbHNlICIoZW1wdHkpIgoKICAgIGRlZiBfbm9ybWFsaXplX3Bvb2xfc2lnbmF0dXJlKHNlbGYsIHBvb2w6IGRpY3Rbc3Ry"
    "LCBpbnRdLCBtb2RpZmllcjogaW50LCBydWxlX2lkOiBzdHIgPSAiIikgLT4gc3RyOgogICAgICAgIHBhcnRzID0gW2Yie3Bvb2wu"
    "Z2V0KGQsIDApfXtkfSIgZm9yIGQgaW4gc2VsZi5UUkFZX09SREVSIGlmIHBvb2wuZ2V0KGQsIDApID4gMF0KICAgICAgICBiYXNl"
    "ID0gIisiLmpvaW4ocGFydHMpIGlmIHBhcnRzIGVsc2UgIjAiCiAgICAgICAgc2lnID0gZiJ7YmFzZX17bW9kaWZpZXI6K2R9Igog"
    "ICAgICAgIHJldHVybiBmIntzaWd9X3tydWxlX2lkfSIgaWYgcnVsZV9pZCBlbHNlIHNpZwoKICAgIGRlZiBfZGljZV9sYWJlbChz"
    "ZWxmLCBkaWVfdHlwZTogc3RyKSAtPiBzdHI6CiAgICAgICAgcmV0dXJuICJkJSIgaWYgZGllX3R5cGUgPT0gImQlIiBlbHNlIGRp"
    "ZV90eXBlCgogICAgZGVmIF9yb2xsX3NpbmdsZV92YWx1ZShzZWxmLCBkaWVfdHlwZTogc3RyKToKICAgICAgICBpZiBkaWVfdHlw"
    "ZSA9PSAiZCUiOgogICAgICAgICAgICB0ZW5zID0gcmFuZG9tLnJhbmRpbnQoMCwgOSkgKiAxMAogICAgICAgICAgICByZXR1cm4g"
    "dGVucywgKCIwMCIgaWYgdGVucyA9PSAwIGVsc2Ugc3RyKHRlbnMpKQogICAgICAgIHNpZGVzID0gaW50KGRpZV90eXBlLnJlcGxh"
    "Y2UoImQiLCAiIikpCiAgICAgICAgdmFsID0gcmFuZG9tLnJhbmRpbnQoMSwgc2lkZXMpCiAgICAgICAgcmV0dXJuIHZhbCwgc3Ry"
    "KHZhbCkKCiAgICBkZWYgX3JvbGxfcG9vbF9kYXRhKHNlbGYsIHBvb2w6IGRpY3Rbc3RyLCBpbnRdLCBtb2RpZmllcjogaW50LCBs"
    "YWJlbDogc3RyLCBydWxlX2lkOiBzdHIgPSAiIikgLT4gZGljdDoKICAgICAgICBncm91cGVkX251bWVyaWM6IGRpY3Rbc3RyLCBs"
    "aXN0W2ludF1dID0ge30KICAgICAgICBncm91cGVkX2Rpc3BsYXk6IGRpY3Rbc3RyLCBsaXN0W3N0cl1dID0ge30KICAgICAgICBz"
    "dWJ0b3RhbCA9IDAKICAgICAgICB1c2VkX3Bvb2wgPSBkaWN0KHBvb2wpCgogICAgICAgIGlmIHJ1bGVfaWQgYW5kIHJ1bGVfaWQg"
    "aW4gc2VsZi5ydWxlX2RlZmluaXRpb25zIGFuZCAobm90IHBvb2wgb3IgbGVuKFtrIGZvciBrLCB2IGluIHBvb2wuaXRlbXMoKSBp"
    "ZiB2ID4gMF0pID09IDEpOgogICAgICAgICAgICBydWxlID0gc2VsZi5ydWxlX2RlZmluaXRpb25zLmdldChydWxlX2lkLCB7fSkK"
    "ICAgICAgICAgICAgc2lkZXMgPSBpbnQocnVsZS5nZXQoImRpY2Vfc2lkZXMiLCA2KSkKICAgICAgICAgICAgY291bnQgPSBpbnQo"
    "cnVsZS5nZXQoImRpY2VfY291bnQiLCAxKSkKICAgICAgICAgICAgZGllID0gZiJke3NpZGVzfSIKICAgICAgICAgICAgdXNlZF9w"
    "b29sID0ge2RpZTogY291bnR9CiAgICAgICAgICAgIHJhdyA9IFtyYW5kb20ucmFuZGludCgxLCBzaWRlcykgZm9yIF8gaW4gcmFu"
    "Z2UoY291bnQpXQogICAgICAgICAgICBkcm9wX2xvdyA9IGludChydWxlLmdldCgiZHJvcF9sb3dlc3RfY291bnQiLCAwKSBvciAw"
    "KQogICAgICAgICAgICBkcm9wX2hpZ2ggPSBpbnQocnVsZS5nZXQoImRyb3BfaGlnaGVzdF9jb3VudCIsIDApIG9yIDApCiAgICAg"
    "ICAgICAgIGtlcHQgPSBsaXN0KHJhdykKICAgICAgICAgICAgaWYgZHJvcF9sb3cgPiAwOgogICAgICAgICAgICAgICAga2VwdCA9"
    "IHNvcnRlZChrZXB0KVtkcm9wX2xvdzpdCiAgICAgICAgICAgIGlmIGRyb3BfaGlnaCA+IDA6CiAgICAgICAgICAgICAgICBrZXB0"
    "ID0gc29ydGVkKGtlcHQpWzotZHJvcF9oaWdoXSBpZiBkcm9wX2hpZ2ggPCBsZW4oa2VwdCkgZWxzZSBbXQogICAgICAgICAgICBn"
    "cm91cGVkX251bWVyaWNbZGllXSA9IHJhdwogICAgICAgICAgICBncm91cGVkX2Rpc3BsYXlbZGllXSA9IFtzdHIodikgZm9yIHYg"
    "aW4gcmF3XQogICAgICAgICAgICBzdWJ0b3RhbCA9IHN1bShrZXB0KQogICAgICAgIGVsc2U6CiAgICAgICAgICAgIGZvciBkaWUg"
    "aW4gc2VsZi5UUkFZX09SREVSOgogICAgICAgICAgICAgICAgcXR5ID0gaW50KHBvb2wuZ2V0KGRpZSwgMCkgb3IgMCkKICAgICAg"
    "ICAgICAgICAgIGlmIHF0eSA8PSAwOgogICAgICAgICAgICAgICAgICAgIGNvbnRpbnVlCiAgICAgICAgICAgICAgICBncm91cGVk"
    "X251bWVyaWNbZGllXSA9IFtdCiAgICAgICAgICAgICAgICBncm91cGVkX2Rpc3BsYXlbZGllXSA9IFtdCiAgICAgICAgICAgICAg"
    "ICBmb3IgXyBpbiByYW5nZShxdHkpOgogICAgICAgICAgICAgICAgICAgIG51bSwgZGlzcCA9IHNlbGYuX3JvbGxfc2luZ2xlX3Zh"
    "bHVlKGRpZSkKICAgICAgICAgICAgICAgICAgICBncm91cGVkX251bWVyaWNbZGllXS5hcHBlbmQobnVtKQogICAgICAgICAgICAg"
    "ICAgICAgIGdyb3VwZWRfZGlzcGxheVtkaWVdLmFwcGVuZChkaXNwKQogICAgICAgICAgICAgICAgICAgIHN1YnRvdGFsICs9IGlu"
    "dChudW0pCgogICAgICAgIHRvdGFsID0gc3VidG90YWwgKyBpbnQobW9kaWZpZXIpCiAgICAgICAgdHMgPSBkYXRldGltZS5ub3co"
    "KS5zdHJmdGltZSgiJUg6JU06JVMiKQogICAgICAgIGV4cHIgPSBzZWxmLl9wb29sX2V4cHJlc3Npb24odXNlZF9wb29sKQogICAg"
    "ICAgIGlmIHJ1bGVfaWQ6CiAgICAgICAgICAgIHJ1bGVfbmFtZSA9IHNlbGYucnVsZV9kZWZpbml0aW9ucy5nZXQocnVsZV9pZCwg"
    "e30pLmdldCgibmFtZSIsIHJ1bGVfaWQpCiAgICAgICAgICAgIGV4cHIgPSBmIntleHByfSAoe3J1bGVfbmFtZX0pIgoKICAgICAg"
    "ICBldmVudCA9IHsKICAgICAgICAgICAgImlkIjogZiJyb2xsX3t1dWlkLnV1aWQ0KCkuaGV4WzoxMl19IiwKICAgICAgICAgICAg"
    "InRpbWVzdGFtcCI6IHRzLAogICAgICAgICAgICAibGFiZWwiOiBsYWJlbCwKICAgICAgICAgICAgInBvb2wiOiB1c2VkX3Bvb2ws"
    "CiAgICAgICAgICAgICJncm91cGVkX3JhdyI6IGdyb3VwZWRfbnVtZXJpYywKICAgICAgICAgICAgImdyb3VwZWRfcmF3X2Rpc3Bs"
    "YXkiOiBncm91cGVkX2Rpc3BsYXksCiAgICAgICAgICAgICJzdWJ0b3RhbCI6IHN1YnRvdGFsLAogICAgICAgICAgICAibW9kaWZp"
    "ZXIiOiBpbnQobW9kaWZpZXIpLAogICAgICAgICAgICAiZmluYWxfdG90YWwiOiBpbnQodG90YWwpLAogICAgICAgICAgICAiZXhw"
    "cmVzc2lvbiI6IGV4cHIsCiAgICAgICAgICAgICJzb3VyY2UiOiAiZGljZV9yb2xsZXIiLAogICAgICAgICAgICAicnVsZV9pZCI6"
    "IHJ1bGVfaWQgb3IgTm9uZSwKICAgICAgICB9CiAgICAgICAgcmV0dXJuIGV2ZW50CgogICAgZGVmIF9hZGRfZGllX3RvX3Bvb2wo"
    "c2VsZiwgZGllX3R5cGU6IHN0cikgLT4gTm9uZToKICAgICAgICBzZWxmLmN1cnJlbnRfcG9vbFtkaWVfdHlwZV0gPSBpbnQoc2Vs"
    "Zi5jdXJyZW50X3Bvb2wuZ2V0KGRpZV90eXBlLCAwKSkgKyAxCiAgICAgICAgc2VsZi5fcmVmcmVzaF9wb29sX2VkaXRvcigpCiAg"
    "ICAgICAgc2VsZi5jdXJyZW50X3Jlc3VsdF9sYmwuc2V0VGV4dChmIkN1cnJlbnQgUG9vbDoge3NlbGYuX3Bvb2xfZXhwcmVzc2lv"
    "bigpfSIpCgogICAgZGVmIF9hZGp1c3RfcG9vbF9kaWUoc2VsZiwgZGllX3R5cGU6IHN0ciwgZGVsdGE6IGludCkgLT4gTm9uZToK"
    "ICAgICAgICBuZXdfdmFsID0gaW50KHNlbGYuY3VycmVudF9wb29sLmdldChkaWVfdHlwZSwgMCkpICsgaW50KGRlbHRhKQogICAg"
    "ICAgIGlmIG5ld192YWwgPD0gMDoKICAgICAgICAgICAgc2VsZi5jdXJyZW50X3Bvb2wucG9wKGRpZV90eXBlLCBOb25lKQogICAg"
    "ICAgIGVsc2U6CiAgICAgICAgICAgIHNlbGYuY3VycmVudF9wb29sW2RpZV90eXBlXSA9IG5ld192YWwKICAgICAgICBzZWxmLl9y"
    "ZWZyZXNoX3Bvb2xfZWRpdG9yKCkKCiAgICBkZWYgX3JlZnJlc2hfcG9vbF9lZGl0b3Ioc2VsZikgLT4gTm9uZToKICAgICAgICB3"
    "aGlsZSBzZWxmLnBvb2xfZW50cmllc19sYXlvdXQuY291bnQoKToKICAgICAgICAgICAgaXRlbSA9IHNlbGYucG9vbF9lbnRyaWVz"
    "X2xheW91dC50YWtlQXQoMCkKICAgICAgICAgICAgdyA9IGl0ZW0ud2lkZ2V0KCkKICAgICAgICAgICAgaWYgdyBpcyBub3QgTm9u"
    "ZToKICAgICAgICAgICAgICAgIHcuZGVsZXRlTGF0ZXIoKQoKICAgICAgICBmb3IgZGllLCBxdHkgaW4gc2VsZi5fc29ydGVkX3Bv"
    "b2xfaXRlbXMoKToKICAgICAgICAgICAgYm94ID0gUUZyYW1lKCkKICAgICAgICAgICAgYm94LnNldFN0eWxlU2hlZXQoZiJiYWNr"
    "Z3JvdW5kOiB7Q19CRzN9OyBib3JkZXI6IDFweCBzb2xpZCB7Q19CT1JERVJ9OyBib3JkZXItcmFkaXVzOiA2cHg7IikKICAgICAg"
    "ICAgICAgbGF5ID0gUUhCb3hMYXlvdXQoYm94KQogICAgICAgICAgICBsYXkuc2V0Q29udGVudHNNYXJnaW5zKDYsIDQsIDYsIDQp"
    "CiAgICAgICAgICAgIGxheS5zZXRTcGFjaW5nKDQpCiAgICAgICAgICAgIGxibCA9IFFMYWJlbChmIntkaWV9IHh7cXR5fSIpCiAg"
    "ICAgICAgICAgIG1pbnVzX2J0biA9IFFQdXNoQnV0dG9uKCLiiJIiKQogICAgICAgICAgICBwbHVzX2J0biA9IFFQdXNoQnV0dG9u"
    "KCIrIikKICAgICAgICAgICAgbWludXNfYnRuLnNldEZpeGVkV2lkdGgoMjQpCiAgICAgICAgICAgIHBsdXNfYnRuLnNldEZpeGVk"
    "V2lkdGgoMjQpCiAgICAgICAgICAgIG1pbnVzX2J0bi5jbGlja2VkLmNvbm5lY3QobGFtYmRhIF89RmFsc2UsIGQ9ZGllOiBzZWxm"
    "Ll9hZGp1c3RfcG9vbF9kaWUoZCwgLTEpKQogICAgICAgICAgICBwbHVzX2J0bi5jbGlja2VkLmNvbm5lY3QobGFtYmRhIF89RmFs"
    "c2UsIGQ9ZGllOiBzZWxmLl9hZGp1c3RfcG9vbF9kaWUoZCwgKzEpKQogICAgICAgICAgICBsYXkuYWRkV2lkZ2V0KGxibCkKICAg"
    "ICAgICAgICAgbGF5LmFkZFdpZGdldChtaW51c19idG4pCiAgICAgICAgICAgIGxheS5hZGRXaWRnZXQocGx1c19idG4pCiAgICAg"
    "ICAgICAgIHNlbGYucG9vbF9lbnRyaWVzX2xheW91dC5hZGRXaWRnZXQoYm94KQoKICAgICAgICBzZWxmLnBvb2xfZW50cmllc19s"
    "YXlvdXQuYWRkU3RyZXRjaCgxKQogICAgICAgIHNlbGYucG9vbF9leHByX2xibC5zZXRUZXh0KGYiUG9vbDoge3NlbGYuX3Bvb2xf"
    "ZXhwcmVzc2lvbigpfSIpCgogICAgZGVmIF9xdWlja19yb2xsX3NpbmdsZV9kaWUoc2VsZiwgZGllX3R5cGU6IHN0cikgLT4gTm9u"
    "ZToKICAgICAgICBldmVudCA9IHNlbGYuX3JvbGxfcG9vbF9kYXRhKHtkaWVfdHlwZTogMX0sIGludChzZWxmLm1vZF9zcGluLnZh"
    "bHVlKCkpLCBzZWxmLmxhYmVsX2VkaXQudGV4dCgpLnN0cmlwKCksIHNlbGYucnVsZV9jb21iby5jdXJyZW50RGF0YSgpIG9yICIi"
    "KQogICAgICAgIHNlbGYuX3JlY29yZF9yb2xsX2V2ZW50KGV2ZW50KQoKICAgIGRlZiBfcm9sbF9jdXJyZW50X3Bvb2woc2VsZikg"
    "LT4gTm9uZToKICAgICAgICBwb29sID0gZGljdChzZWxmLmN1cnJlbnRfcG9vbCkKICAgICAgICBydWxlX2lkID0gc2VsZi5ydWxl"
    "X2NvbWJvLmN1cnJlbnREYXRhKCkgb3IgIiIKICAgICAgICBpZiBub3QgcG9vbCBhbmQgbm90IHJ1bGVfaWQ6CiAgICAgICAgICAg"
    "IFFNZXNzYWdlQm94LmluZm9ybWF0aW9uKHNlbGYsICJEaWNlIFJvbGxlciIsICJDdXJyZW50IFBvb2wgaXMgZW1wdHkuIikKICAg"
    "ICAgICAgICAgcmV0dXJuCiAgICAgICAgZXZlbnQgPSBzZWxmLl9yb2xsX3Bvb2xfZGF0YShwb29sLCBpbnQoc2VsZi5tb2Rfc3Bp"
    "bi52YWx1ZSgpKSwgc2VsZi5sYWJlbF9lZGl0LnRleHQoKS5zdHJpcCgpLCBydWxlX2lkKQogICAgICAgIHNlbGYuX3JlY29yZF9y"
    "b2xsX2V2ZW50KGV2ZW50KQoKICAgIGRlZiBfcmVjb3JkX3JvbGxfZXZlbnQoc2VsZiwgZXZlbnQ6IGRpY3QpIC0+IE5vbmU6CiAg"
    "ICAgICAgc2VsZi5yb2xsX2V2ZW50cy5hcHBlbmQoZXZlbnQpCiAgICAgICAgc2VsZi5ldmVudF9ieV9pZFtldmVudFsiaWQiXV0g"
    "PSBldmVudAogICAgICAgIHNlbGYuY3VycmVudF9yb2xsX2lkcyA9IFtldmVudFsiaWQiXV0KCiAgICAgICAgc2VsZi5fcmVwbGFj"
    "ZV9jdXJyZW50X3Jvd3MoW2V2ZW50XSkKICAgICAgICBzZWxmLl9hcHBlbmRfaGlzdG9yeV9yb3coZXZlbnQpCiAgICAgICAgc2Vs"
    "Zi5fdXBkYXRlX2dyYW5kX3RvdGFsKCkKICAgICAgICBzZWxmLl91cGRhdGVfcmVzdWx0X2Rpc3BsYXkoZXZlbnQpCiAgICAgICAg"
    "c2VsZi5fdHJhY2tfY29tbW9uX3NpZ25hdHVyZShldmVudCkKICAgICAgICBzZWxmLl9wbGF5X3JvbGxfc291bmQoKQoKICAgIGRl"
    "ZiBfcmVwbGFjZV9jdXJyZW50X3Jvd3Moc2VsZiwgZXZlbnRzOiBsaXN0W2RpY3RdKSAtPiBOb25lOgogICAgICAgIHNlbGYuY3Vy"
    "cmVudF90YWJsZS5zZXRSb3dDb3VudCgwKQogICAgICAgIGZvciBldmVudCBpbiBldmVudHM6CiAgICAgICAgICAgIHNlbGYuX2Fw"
    "cGVuZF90YWJsZV9yb3coc2VsZi5jdXJyZW50X3RhYmxlLCBldmVudCkKCiAgICBkZWYgX2FwcGVuZF9oaXN0b3J5X3JvdyhzZWxm"
    "LCBldmVudDogZGljdCkgLT4gTm9uZToKICAgICAgICBzZWxmLl9hcHBlbmRfdGFibGVfcm93KHNlbGYuaGlzdG9yeV90YWJsZSwg"
    "ZXZlbnQpCiAgICAgICAgc2VsZi5oaXN0b3J5X3RhYmxlLnNjcm9sbFRvQm90dG9tKCkKCiAgICBkZWYgX2Zvcm1hdF9yYXcoc2Vs"
    "ZiwgZXZlbnQ6IGRpY3QpIC0+IHN0cjoKICAgICAgICBncm91cGVkID0gZXZlbnQuZ2V0KCJncm91cGVkX3Jhd19kaXNwbGF5Iiwg"
    "e30pIG9yIHt9CiAgICAgICAgYml0cyA9IFtdCiAgICAgICAgZm9yIGRpZSBpbiBzZWxmLlRSQVlfT1JERVI6CiAgICAgICAgICAg"
    "IHZhbHMgPSBncm91cGVkLmdldChkaWUpCiAgICAgICAgICAgIGlmIHZhbHM6CiAgICAgICAgICAgICAgICBiaXRzLmFwcGVuZChm"
    "IntkaWV9OiB7JywnLmpvaW4oc3RyKHYpIGZvciB2IGluIHZhbHMpfSIpCiAgICAgICAgcmV0dXJuICIgfCAiLmpvaW4oYml0cykK"
    "CiAgICBkZWYgX2FwcGVuZF90YWJsZV9yb3coc2VsZiwgdGFibGU6IFFUYWJsZVdpZGdldCwgZXZlbnQ6IGRpY3QpIC0+IE5vbmU6"
    "CiAgICAgICAgcm93ID0gdGFibGUucm93Q291bnQoKQogICAgICAgIHRhYmxlLmluc2VydFJvdyhyb3cpCgogICAgICAgIHRzX2l0"
    "ZW0gPSBRVGFibGVXaWRnZXRJdGVtKGV2ZW50WyJ0aW1lc3RhbXAiXSkKICAgICAgICB0c19pdGVtLnNldERhdGEoUXQuSXRlbURh"
    "dGFSb2xlLlVzZXJSb2xlLCBldmVudFsiaWQiXSkKICAgICAgICB0YWJsZS5zZXRJdGVtKHJvdywgMCwgdHNfaXRlbSkKICAgICAg"
    "ICB0YWJsZS5zZXRJdGVtKHJvdywgMSwgUVRhYmxlV2lkZ2V0SXRlbShldmVudC5nZXQoImxhYmVsIiwgIiIpKSkKICAgICAgICB0"
    "YWJsZS5zZXRJdGVtKHJvdywgMiwgUVRhYmxlV2lkZ2V0SXRlbShldmVudC5nZXQoImV4cHJlc3Npb24iLCAiIikpKQogICAgICAg"
    "IHRhYmxlLnNldEl0ZW0ocm93LCAzLCBRVGFibGVXaWRnZXRJdGVtKHNlbGYuX2Zvcm1hdF9yYXcoZXZlbnQpKSkKCiAgICAgICAg"
    "bW9kX3NwaW4gPSBRU3BpbkJveCgpCiAgICAgICAgbW9kX3NwaW4uc2V0UmFuZ2UoLTk5OSwgOTk5KQogICAgICAgIG1vZF9zcGlu"
    "LnNldFZhbHVlKGludChldmVudC5nZXQoIm1vZGlmaWVyIiwgMCkpKQogICAgICAgIG1vZF9zcGluLnZhbHVlQ2hhbmdlZC5jb25u"
    "ZWN0KGxhbWJkYSB2YWwsIGVpZD1ldmVudFsiaWQiXTogc2VsZi5fb25fbW9kaWZpZXJfY2hhbmdlZChlaWQsIHZhbCkpCiAgICAg"
    "ICAgdGFibGUuc2V0Q2VsbFdpZGdldChyb3csIDQsIG1vZF9zcGluKQoKICAgICAgICB0YWJsZS5zZXRJdGVtKHJvdywgNSwgUVRh"
    "YmxlV2lkZ2V0SXRlbShzdHIoZXZlbnQuZ2V0KCJmaW5hbF90b3RhbCIsIDApKSkpCgogICAgZGVmIF9zeW5jX3Jvd19ieV9ldmVu"
    "dF9pZChzZWxmLCB0YWJsZTogUVRhYmxlV2lkZ2V0LCBldmVudF9pZDogc3RyLCBldmVudDogZGljdCkgLT4gTm9uZToKICAgICAg"
    "ICBmb3Igcm93IGluIHJhbmdlKHRhYmxlLnJvd0NvdW50KCkpOgogICAgICAgICAgICBpdCA9IHRhYmxlLml0ZW0ocm93LCAwKQog"
    "ICAgICAgICAgICBpZiBpdCBhbmQgaXQuZGF0YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUpID09IGV2ZW50X2lkOgogICAgICAg"
    "ICAgICAgICAgdGFibGUuc2V0SXRlbShyb3csIDUsIFFUYWJsZVdpZGdldEl0ZW0oc3RyKGV2ZW50LmdldCgiZmluYWxfdG90YWwi"
    "LCAwKSkpKQogICAgICAgICAgICAgICAgdGFibGUuc2V0SXRlbShyb3csIDMsIFFUYWJsZVdpZGdldEl0ZW0oc2VsZi5fZm9ybWF0"
    "X3JhdyhldmVudCkpKQogICAgICAgICAgICAgICAgYnJlYWsKCiAgICBkZWYgX29uX21vZGlmaWVyX2NoYW5nZWQoc2VsZiwgZXZl"
    "bnRfaWQ6IHN0ciwgdmFsdWU6IGludCkgLT4gTm9uZToKICAgICAgICBldnQgPSBzZWxmLmV2ZW50X2J5X2lkLmdldChldmVudF9p"
    "ZCkKICAgICAgICBpZiBub3QgZXZ0OgogICAgICAgICAgICByZXR1cm4KICAgICAgICBldnRbIm1vZGlmaWVyIl0gPSBpbnQodmFs"
    "dWUpCiAgICAgICAgZXZ0WyJmaW5hbF90b3RhbCJdID0gaW50KGV2dC5nZXQoInN1YnRvdGFsIiwgMCkpICsgaW50KHZhbHVlKQog"
    "ICAgICAgIHNlbGYuX3N5bmNfcm93X2J5X2V2ZW50X2lkKHNlbGYuaGlzdG9yeV90YWJsZSwgZXZlbnRfaWQsIGV2dCkKICAgICAg"
    "ICBzZWxmLl9zeW5jX3Jvd19ieV9ldmVudF9pZChzZWxmLmN1cnJlbnRfdGFibGUsIGV2ZW50X2lkLCBldnQpCiAgICAgICAgc2Vs"
    "Zi5fdXBkYXRlX2dyYW5kX3RvdGFsKCkKICAgICAgICBpZiBzZWxmLmN1cnJlbnRfcm9sbF9pZHMgYW5kIHNlbGYuY3VycmVudF9y"
    "b2xsX2lkc1swXSA9PSBldmVudF9pZDoKICAgICAgICAgICAgc2VsZi5fdXBkYXRlX3Jlc3VsdF9kaXNwbGF5KGV2dCkKCiAgICBk"
    "ZWYgX3VwZGF0ZV9ncmFuZF90b3RhbChzZWxmKSAtPiBOb25lOgogICAgICAgIHRvdGFsID0gc3VtKGludChldnQuZ2V0KCJmaW5h"
    "bF90b3RhbCIsIDApKSBmb3IgZXZ0IGluIHNlbGYucm9sbF9ldmVudHMpCiAgICAgICAgc2VsZi5ncmFuZF90b3RhbF9sYmwuc2V0"
    "VGV4dChmIkdyYW5kIFRvdGFsOiB7dG90YWx9IikKCiAgICBkZWYgX3VwZGF0ZV9yZXN1bHRfZGlzcGxheShzZWxmLCBldmVudDog"
    "ZGljdCkgLT4gTm9uZToKICAgICAgICBncm91cGVkID0gZXZlbnQuZ2V0KCJncm91cGVkX3Jhd19kaXNwbGF5Iiwge30pIG9yIHt9"
    "CiAgICAgICAgbGluZXMgPSBbXQogICAgICAgIGZvciBkaWUgaW4gc2VsZi5UUkFZX09SREVSOgogICAgICAgICAgICB2YWxzID0g"
    "Z3JvdXBlZC5nZXQoZGllKQogICAgICAgICAgICBpZiB2YWxzOgogICAgICAgICAgICAgICAgbGluZXMuYXBwZW5kKGYie2RpZX0g"
    "eHtsZW4odmFscyl9IOKGkiBbeycsJy5qb2luKHN0cih2KSBmb3IgdiBpbiB2YWxzKX1dIikKICAgICAgICBydWxlX2lkID0gZXZl"
    "bnQuZ2V0KCJydWxlX2lkIikKICAgICAgICBpZiBydWxlX2lkOgogICAgICAgICAgICBydWxlX25hbWUgPSBzZWxmLnJ1bGVfZGVm"
    "aW5pdGlvbnMuZ2V0KHJ1bGVfaWQsIHt9KS5nZXQoIm5hbWUiLCBydWxlX2lkKQogICAgICAgICAgICBsaW5lcy5hcHBlbmQoZiJS"
    "dWxlOiB7cnVsZV9uYW1lfSIpCiAgICAgICAgbGluZXMuYXBwZW5kKGYiTW9kaWZpZXI6IHtpbnQoZXZlbnQuZ2V0KCdtb2RpZmll"
    "cicsIDApKTorZH0iKQogICAgICAgIGxpbmVzLmFwcGVuZChmIlRvdGFsOiB7ZXZlbnQuZ2V0KCdmaW5hbF90b3RhbCcsIDApfSIp"
    "CiAgICAgICAgc2VsZi5jdXJyZW50X3Jlc3VsdF9sYmwuc2V0VGV4dCgiXG4iLmpvaW4obGluZXMpKQoKCiAgICBkZWYgX3NhdmVf"
    "cG9vbChzZWxmKSAtPiBOb25lOgogICAgICAgIGlmIG5vdCBzZWxmLmN1cnJlbnRfcG9vbDoKICAgICAgICAgICAgUU1lc3NhZ2VC"
    "b3guaW5mb3JtYXRpb24oc2VsZiwgIkRpY2UgUm9sbGVyIiwgIkJ1aWxkIGEgQ3VycmVudCBQb29sIGJlZm9yZSBzYXZpbmcuIikK"
    "ICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgZGVmYXVsdF9uYW1lID0gc2VsZi5sYWJlbF9lZGl0LnRleHQoKS5zdHJpcCgpIG9y"
    "IHNlbGYuX3Bvb2xfZXhwcmVzc2lvbigpCiAgICAgICAgbmFtZSwgb2sgPSBRSW5wdXREaWFsb2cuZ2V0VGV4dChzZWxmLCAiU2F2"
    "ZSBQb29sIiwgIlNhdmVkIHJvbGwgbmFtZToiLCB0ZXh0PWRlZmF1bHRfbmFtZSkKICAgICAgICBpZiBub3Qgb2s6CiAgICAgICAg"
    "ICAgIHJldHVybgogICAgICAgIHBheWxvYWQgPSB7CiAgICAgICAgICAgICJpZCI6IGYic2F2ZWRfe3V1aWQudXVpZDQoKS5oZXhb"
    "OjEwXX0iLAogICAgICAgICAgICAibmFtZSI6IG5hbWUuc3RyaXAoKSBvciBkZWZhdWx0X25hbWUsCiAgICAgICAgICAgICJwb29s"
    "IjogZGljdChzZWxmLmN1cnJlbnRfcG9vbCksCiAgICAgICAgICAgICJtb2RpZmllciI6IGludChzZWxmLm1vZF9zcGluLnZhbHVl"
    "KCkpLAogICAgICAgICAgICAicnVsZV9pZCI6IHNlbGYucnVsZV9jb21iby5jdXJyZW50RGF0YSgpIG9yIE5vbmUsCiAgICAgICAg"
    "ICAgICJub3RlcyI6ICIiLAogICAgICAgICAgICAiY2F0ZWdvcnkiOiAic2F2ZWQiLAogICAgICAgIH0KICAgICAgICBzZWxmLnNh"
    "dmVkX3JvbGxzLmFwcGVuZChwYXlsb2FkKQogICAgICAgIHNlbGYuX3JlZnJlc2hfc2F2ZWRfbGlzdHMoKQoKICAgIGRlZiBfcmVm"
    "cmVzaF9zYXZlZF9saXN0cyhzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuc2F2ZWRfbGlzdC5jbGVhcigpCiAgICAgICAgZm9y"
    "IGl0ZW0gaW4gc2VsZi5zYXZlZF9yb2xsczoKICAgICAgICAgICAgZXhwciA9IHNlbGYuX3Bvb2xfZXhwcmVzc2lvbihpdGVtLmdl"
    "dCgicG9vbCIsIHt9KSkKICAgICAgICAgICAgdHh0ID0gZiJ7aXRlbS5nZXQoJ25hbWUnKX0g4oCUIHtleHByfSB7aW50KGl0ZW0u"
    "Z2V0KCdtb2RpZmllcicsIDApKTorZH0iCiAgICAgICAgICAgIGx3ID0gUUxpc3RXaWRnZXRJdGVtKHR4dCkKICAgICAgICAgICAg"
    "bHcuc2V0RGF0YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUsIGl0ZW0pCiAgICAgICAgICAgIHNlbGYuc2F2ZWRfbGlzdC5hZGRJ"
    "dGVtKGx3KQoKICAgICAgICBzZWxmLmNvbW1vbl9saXN0LmNsZWFyKCkKICAgICAgICByYW5rZWQgPSBzb3J0ZWQoc2VsZi5jb21t"
    "b25fcm9sbHMudmFsdWVzKCksIGtleT1sYW1iZGEgeDogeC5nZXQoImNvdW50IiwgMCksIHJldmVyc2U9VHJ1ZSkKICAgICAgICBm"
    "b3IgaXRlbSBpbiByYW5rZWQ6CiAgICAgICAgICAgIGlmIGludChpdGVtLmdldCgiY291bnQiLCAwKSkgPCAyOgogICAgICAgICAg"
    "ICAgICAgY29udGludWUKICAgICAgICAgICAgZXhwciA9IHNlbGYuX3Bvb2xfZXhwcmVzc2lvbihpdGVtLmdldCgicG9vbCIsIHt9"
    "KSkKICAgICAgICAgICAgdHh0ID0gZiJ7ZXhwcn0ge2ludChpdGVtLmdldCgnbW9kaWZpZXInLCAwKSk6K2R9ICh4e2l0ZW0uZ2V0"
    "KCdjb3VudCcsIDApfSkiCiAgICAgICAgICAgIGx3ID0gUUxpc3RXaWRnZXRJdGVtKHR4dCkKICAgICAgICAgICAgbHcuc2V0RGF0"
    "YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUsIGl0ZW0pCiAgICAgICAgICAgIHNlbGYuY29tbW9uX2xpc3QuYWRkSXRlbShsdykK"
    "CiAgICBkZWYgX3RyYWNrX2NvbW1vbl9zaWduYXR1cmUoc2VsZiwgZXZlbnQ6IGRpY3QpIC0+IE5vbmU6CiAgICAgICAgc2lnID0g"
    "c2VsZi5fbm9ybWFsaXplX3Bvb2xfc2lnbmF0dXJlKGV2ZW50LmdldCgicG9vbCIsIHt9KSwgaW50KGV2ZW50LmdldCgibW9kaWZp"
    "ZXIiLCAwKSksIHN0cihldmVudC5nZXQoInJ1bGVfaWQiKSBvciAiIikpCiAgICAgICAgaWYgc2lnIG5vdCBpbiBzZWxmLmNvbW1v"
    "bl9yb2xsczoKICAgICAgICAgICAgc2VsZi5jb21tb25fcm9sbHNbc2lnXSA9IHsKICAgICAgICAgICAgICAgICJzaWduYXR1cmUi"
    "OiBzaWcsCiAgICAgICAgICAgICAgICAiY291bnQiOiAwLAogICAgICAgICAgICAgICAgIm5hbWUiOiBldmVudC5nZXQoImxhYmVs"
    "IiwgIiIpIG9yIHNpZywKICAgICAgICAgICAgICAgICJwb29sIjogZGljdChldmVudC5nZXQoInBvb2wiLCB7fSkpLAogICAgICAg"
    "ICAgICAgICAgIm1vZGlmaWVyIjogaW50KGV2ZW50LmdldCgibW9kaWZpZXIiLCAwKSksCiAgICAgICAgICAgICAgICAicnVsZV9p"
    "ZCI6IGV2ZW50LmdldCgicnVsZV9pZCIpLAogICAgICAgICAgICAgICAgIm5vdGVzIjogIiIsCiAgICAgICAgICAgICAgICAiY2F0"
    "ZWdvcnkiOiAiY29tbW9uIiwKICAgICAgICAgICAgfQogICAgICAgIHNlbGYuY29tbW9uX3JvbGxzW3NpZ11bImNvdW50Il0gPSBp"
    "bnQoc2VsZi5jb21tb25fcm9sbHNbc2lnXS5nZXQoImNvdW50IiwgMCkpICsgMQogICAgICAgIGlmIHNlbGYuY29tbW9uX3JvbGxz"
    "W3NpZ11bImNvdW50Il0gPj0gMzoKICAgICAgICAgICAgc2VsZi5jb21tb25faGludC5zZXRUZXh0KGYiU3VnZ2VzdGlvbjogcHJv"
    "bW90ZSB7c2VsZi5fcG9vbF9leHByZXNzaW9uKGV2ZW50LmdldCgncG9vbCcsIHt9KSl9IHRvIFNhdmVkLiIpCiAgICAgICAgc2Vs"
    "Zi5fcmVmcmVzaF9zYXZlZF9saXN0cygpCgogICAgZGVmIF9ydW5fc2F2ZWRfcm9sbChzZWxmLCBwYXlsb2FkOiBkaWN0IHwgTm9u"
    "ZSk6CiAgICAgICAgaWYgbm90IHBheWxvYWQ6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIGV2ZW50ID0gc2VsZi5fcm9sbF9w"
    "b29sX2RhdGEoCiAgICAgICAgICAgIGRpY3QocGF5bG9hZC5nZXQoInBvb2wiLCB7fSkpLAogICAgICAgICAgICBpbnQocGF5bG9h"
    "ZC5nZXQoIm1vZGlmaWVyIiwgMCkpLAogICAgICAgICAgICBzdHIocGF5bG9hZC5nZXQoIm5hbWUiLCAiIikpLnN0cmlwKCksCiAg"
    "ICAgICAgICAgIHN0cihwYXlsb2FkLmdldCgicnVsZV9pZCIpIG9yICIiKSwKICAgICAgICApCiAgICAgICAgc2VsZi5fcmVjb3Jk"
    "X3JvbGxfZXZlbnQoZXZlbnQpCgogICAgZGVmIF9sb2FkX3BheWxvYWRfaW50b19wb29sKHNlbGYsIHBheWxvYWQ6IGRpY3QgfCBO"
    "b25lKSAtPiBOb25lOgogICAgICAgIGlmIG5vdCBwYXlsb2FkOgogICAgICAgICAgICByZXR1cm4KICAgICAgICBzZWxmLmN1cnJl"
    "bnRfcG9vbCA9IGRpY3QocGF5bG9hZC5nZXQoInBvb2wiLCB7fSkpCiAgICAgICAgc2VsZi5tb2Rfc3Bpbi5zZXRWYWx1ZShpbnQo"
    "cGF5bG9hZC5nZXQoIm1vZGlmaWVyIiwgMCkpKQogICAgICAgIHNlbGYubGFiZWxfZWRpdC5zZXRUZXh0KHN0cihwYXlsb2FkLmdl"
    "dCgibmFtZSIsICIiKSkpCiAgICAgICAgcmlkID0gcGF5bG9hZC5nZXQoInJ1bGVfaWQiKQogICAgICAgIGlkeCA9IHNlbGYucnVs"
    "ZV9jb21iby5maW5kRGF0YShyaWQgb3IgIiIpCiAgICAgICAgaWYgaWR4ID49IDA6CiAgICAgICAgICAgIHNlbGYucnVsZV9jb21i"
    "by5zZXRDdXJyZW50SW5kZXgoaWR4KQogICAgICAgIHNlbGYuX3JlZnJlc2hfcG9vbF9lZGl0b3IoKQogICAgICAgIHNlbGYuY3Vy"
    "cmVudF9yZXN1bHRfbGJsLnNldFRleHQoZiJDdXJyZW50IFBvb2w6IHtzZWxmLl9wb29sX2V4cHJlc3Npb24oKX0iKQoKICAgIGRl"
    "ZiBfcnVuX3NlbGVjdGVkX3NhdmVkKHNlbGYpOgogICAgICAgIGl0ZW0gPSBzZWxmLnNhdmVkX2xpc3QuY3VycmVudEl0ZW0oKQog"
    "ICAgICAgIHNlbGYuX3J1bl9zYXZlZF9yb2xsKGl0ZW0uZGF0YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUpIGlmIGl0ZW0gZWxz"
    "ZSBOb25lKQoKICAgIGRlZiBfbG9hZF9zZWxlY3RlZF9zYXZlZChzZWxmKToKICAgICAgICBpdGVtID0gc2VsZi5zYXZlZF9saXN0"
    "LmN1cnJlbnRJdGVtKCkKICAgICAgICBwYXlsb2FkID0gaXRlbS5kYXRhKFF0Lkl0ZW1EYXRhUm9sZS5Vc2VyUm9sZSkgaWYgaXRl"
    "bSBlbHNlIE5vbmUKICAgICAgICBpZiBub3QgcGF5bG9hZDoKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgc2VsZi5fbG9hZF9w"
    "YXlsb2FkX2ludG9fcG9vbChwYXlsb2FkKQoKICAgICAgICBuYW1lLCBvayA9IFFJbnB1dERpYWxvZy5nZXRUZXh0KHNlbGYsICJF"
    "ZGl0IFNhdmVkIFJvbGwiLCAiTmFtZToiLCB0ZXh0PXN0cihwYXlsb2FkLmdldCgibmFtZSIsICIiKSkpCiAgICAgICAgaWYgbm90"
    "IG9rOgogICAgICAgICAgICByZXR1cm4KICAgICAgICBwYXlsb2FkWyJuYW1lIl0gPSBuYW1lLnN0cmlwKCkgb3IgcGF5bG9hZC5n"
    "ZXQoIm5hbWUiLCAiIikKICAgICAgICBwYXlsb2FkWyJwb29sIl0gPSBkaWN0KHNlbGYuY3VycmVudF9wb29sKQogICAgICAgIHBh"
    "eWxvYWRbIm1vZGlmaWVyIl0gPSBpbnQoc2VsZi5tb2Rfc3Bpbi52YWx1ZSgpKQogICAgICAgIHBheWxvYWRbInJ1bGVfaWQiXSA9"
    "IHNlbGYucnVsZV9jb21iby5jdXJyZW50RGF0YSgpIG9yIE5vbmUKICAgICAgICBub3Rlcywgb2tfbm90ZXMgPSBRSW5wdXREaWFs"
    "b2cuZ2V0VGV4dChzZWxmLCAiRWRpdCBTYXZlZCBSb2xsIiwgIk5vdGVzIC8gY2F0ZWdvcnk6IiwgdGV4dD1zdHIocGF5bG9hZC5n"
    "ZXQoIm5vdGVzIiwgIiIpKSkKICAgICAgICBpZiBva19ub3RlczoKICAgICAgICAgICAgcGF5bG9hZFsibm90ZXMiXSA9IG5vdGVz"
    "CiAgICAgICAgc2VsZi5fcmVmcmVzaF9zYXZlZF9saXN0cygpCgogICAgZGVmIF9kZWxldGVfc2VsZWN0ZWRfc2F2ZWQoc2VsZik6"
    "CiAgICAgICAgcm93ID0gc2VsZi5zYXZlZF9saXN0LmN1cnJlbnRSb3coKQogICAgICAgIGlmIHJvdyA8IDAgb3Igcm93ID49IGxl"
    "bihzZWxmLnNhdmVkX3JvbGxzKToKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgc2VsZi5zYXZlZF9yb2xscy5wb3Aocm93KQog"
    "ICAgICAgIHNlbGYuX3JlZnJlc2hfc2F2ZWRfbGlzdHMoKQoKICAgIGRlZiBfcHJvbW90ZV9zZWxlY3RlZF9jb21tb24oc2VsZik6"
    "CiAgICAgICAgaXRlbSA9IHNlbGYuY29tbW9uX2xpc3QuY3VycmVudEl0ZW0oKQogICAgICAgIHBheWxvYWQgPSBpdGVtLmRhdGEo"
    "UXQuSXRlbURhdGFSb2xlLlVzZXJSb2xlKSBpZiBpdGVtIGVsc2UgTm9uZQogICAgICAgIGlmIG5vdCBwYXlsb2FkOgogICAgICAg"
    "ICAgICByZXR1cm4KICAgICAgICBwcm9tb3RlZCA9IHsKICAgICAgICAgICAgImlkIjogZiJzYXZlZF97dXVpZC51dWlkNCgpLmhl"
    "eFs6MTBdfSIsCiAgICAgICAgICAgICJuYW1lIjogcGF5bG9hZC5nZXQoIm5hbWUiKSBvciBzZWxmLl9wb29sX2V4cHJlc3Npb24o"
    "cGF5bG9hZC5nZXQoInBvb2wiLCB7fSkpLAogICAgICAgICAgICAicG9vbCI6IGRpY3QocGF5bG9hZC5nZXQoInBvb2wiLCB7fSkp"
    "LAogICAgICAgICAgICAibW9kaWZpZXIiOiBpbnQocGF5bG9hZC5nZXQoIm1vZGlmaWVyIiwgMCkpLAogICAgICAgICAgICAicnVs"
    "ZV9pZCI6IHBheWxvYWQuZ2V0KCJydWxlX2lkIiksCiAgICAgICAgICAgICJub3RlcyI6IHBheWxvYWQuZ2V0KCJub3RlcyIsICIi"
    "KSwKICAgICAgICAgICAgImNhdGVnb3J5IjogInNhdmVkIiwKICAgICAgICB9CiAgICAgICAgc2VsZi5zYXZlZF9yb2xscy5hcHBl"
    "bmQocHJvbW90ZWQpCiAgICAgICAgc2VsZi5fcmVmcmVzaF9zYXZlZF9saXN0cygpCgogICAgZGVmIF9kaXNtaXNzX3NlbGVjdGVk"
    "X2NvbW1vbihzZWxmKToKICAgICAgICBpdGVtID0gc2VsZi5jb21tb25fbGlzdC5jdXJyZW50SXRlbSgpCiAgICAgICAgcGF5bG9h"
    "ZCA9IGl0ZW0uZGF0YShRdC5JdGVtRGF0YVJvbGUuVXNlclJvbGUpIGlmIGl0ZW0gZWxzZSBOb25lCiAgICAgICAgaWYgbm90IHBh"
    "eWxvYWQ6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIHNpZyA9IHBheWxvYWQuZ2V0KCJzaWduYXR1cmUiKQogICAgICAgIGlm"
    "IHNpZyBpbiBzZWxmLmNvbW1vbl9yb2xsczoKICAgICAgICAgICAgc2VsZi5jb21tb25fcm9sbHMucG9wKHNpZywgTm9uZSkKICAg"
    "ICAgICBzZWxmLl9yZWZyZXNoX3NhdmVkX2xpc3RzKCkKCiAgICBkZWYgX3Jlc2V0X3Bvb2woc2VsZik6CiAgICAgICAgc2VsZi5j"
    "dXJyZW50X3Bvb2wgPSB7fQogICAgICAgIHNlbGYubW9kX3NwaW4uc2V0VmFsdWUoMCkKICAgICAgICBzZWxmLmxhYmVsX2VkaXQu"
    "Y2xlYXIoKQogICAgICAgIHNlbGYucnVsZV9jb21iby5zZXRDdXJyZW50SW5kZXgoMCkKICAgICAgICBzZWxmLl9yZWZyZXNoX3Bv"
    "b2xfZWRpdG9yKCkKICAgICAgICBzZWxmLmN1cnJlbnRfcmVzdWx0X2xibC5zZXRUZXh0KCJObyByb2xsIHlldC4iKQoKICAgIGRl"
    "ZiBfY2xlYXJfaGlzdG9yeShzZWxmKToKICAgICAgICBzZWxmLnJvbGxfZXZlbnRzLmNsZWFyKCkKICAgICAgICBzZWxmLmV2ZW50"
    "X2J5X2lkLmNsZWFyKCkKICAgICAgICBzZWxmLmN1cnJlbnRfcm9sbF9pZHMgPSBbXQogICAgICAgIHNlbGYuaGlzdG9yeV90YWJs"
    "ZS5zZXRSb3dDb3VudCgwKQogICAgICAgIHNlbGYuY3VycmVudF90YWJsZS5zZXRSb3dDb3VudCgwKQogICAgICAgIHNlbGYuX3Vw"
    "ZGF0ZV9ncmFuZF90b3RhbCgpCiAgICAgICAgc2VsZi5jdXJyZW50X3Jlc3VsdF9sYmwuc2V0VGV4dCgiTm8gcm9sbCB5ZXQuIikK"
    "CiAgICBkZWYgX2V2ZW50X2Zyb21fdGFibGVfcG9zaXRpb24oc2VsZiwgdGFibGU6IFFUYWJsZVdpZGdldCwgcG9zKSAtPiBkaWN0"
    "IHwgTm9uZToKICAgICAgICBpdGVtID0gdGFibGUuaXRlbUF0KHBvcykKICAgICAgICBpZiBub3QgaXRlbToKICAgICAgICAgICAg"
    "cmV0dXJuIE5vbmUKICAgICAgICByb3cgPSBpdGVtLnJvdygpCiAgICAgICAgdHNfaXRlbSA9IHRhYmxlLml0ZW0ocm93LCAwKQog"
    "ICAgICAgIGlmIG5vdCB0c19pdGVtOgogICAgICAgICAgICByZXR1cm4gTm9uZQogICAgICAgIGVpZCA9IHRzX2l0ZW0uZGF0YShR"
    "dC5JdGVtRGF0YVJvbGUuVXNlclJvbGUpCiAgICAgICAgcmV0dXJuIHNlbGYuZXZlbnRfYnlfaWQuZ2V0KGVpZCkKCiAgICBkZWYg"
    "X3Nob3dfcm9sbF9jb250ZXh0X21lbnUoc2VsZiwgdGFibGU6IFFUYWJsZVdpZGdldCwgcG9zKSAtPiBOb25lOgogICAgICAgIGV2"
    "dCA9IHNlbGYuX2V2ZW50X2Zyb21fdGFibGVfcG9zaXRpb24odGFibGUsIHBvcykKICAgICAgICBpZiBub3QgZXZ0OgogICAgICAg"
    "ICAgICByZXR1cm4KICAgICAgICBmcm9tIFB5U2lkZTYuUXRXaWRnZXRzIGltcG9ydCBRTWVudQogICAgICAgIG1lbnUgPSBRTWVu"
    "dShzZWxmKQogICAgICAgIGFjdF9zZW5kID0gbWVudS5hZGRBY3Rpb24oIlNlbmQgdG8gUHJvbXB0IikKICAgICAgICBjaG9zZW4g"
    "PSBtZW51LmV4ZWModGFibGUudmlld3BvcnQoKS5tYXBUb0dsb2JhbChwb3MpKQogICAgICAgIGlmIGNob3NlbiA9PSBhY3Rfc2Vu"
    "ZDoKICAgICAgICAgICAgc2VsZi5fc2VuZF9ldmVudF90b19wcm9tcHQoZXZ0KQoKICAgIGRlZiBfZm9ybWF0X2V2ZW50X2Zvcl9w"
    "cm9tcHQoc2VsZiwgZXZlbnQ6IGRpY3QpIC0+IHN0cjoKICAgICAgICBsYWJlbCA9IChldmVudC5nZXQoImxhYmVsIikgb3IgIlJv"
    "bGwiKS5zdHJpcCgpCiAgICAgICAgZ3JvdXBlZCA9IGV2ZW50LmdldCgiZ3JvdXBlZF9yYXdfZGlzcGxheSIsIHt9KSBvciB7fQog"
    "ICAgICAgIHNlZ21lbnRzID0gW10KICAgICAgICBmb3IgZGllIGluIHNlbGYuVFJBWV9PUkRFUjoKICAgICAgICAgICAgdmFscyA9"
    "IGdyb3VwZWQuZ2V0KGRpZSkKICAgICAgICAgICAgaWYgdmFsczoKICAgICAgICAgICAgICAgIHNlZ21lbnRzLmFwcGVuZChmIntk"
    "aWV9IHJvbGxlZCB7JywnLmpvaW4oc3RyKHYpIGZvciB2IGluIHZhbHMpfSIpCiAgICAgICAgbW9kID0gaW50KGV2ZW50LmdldCgi"
    "bW9kaWZpZXIiLCAwKSkKICAgICAgICB0b3RhbCA9IGludChldmVudC5nZXQoImZpbmFsX3RvdGFsIiwgMCkpCiAgICAgICAgcmV0"
    "dXJuIGYie2xhYmVsfTogeyc7ICcuam9pbihzZWdtZW50cyl9OyBtb2RpZmllciB7bW9kOitkfTsgdG90YWwge3RvdGFsfSIKCiAg"
    "ICBkZWYgX3NlbmRfZXZlbnRfdG9fcHJvbXB0KHNlbGYsIGV2ZW50OiBkaWN0KSAtPiBOb25lOgogICAgICAgIHdpbmRvdyA9IHNl"
    "bGYud2luZG93KCkKICAgICAgICBpZiBub3Qgd2luZG93IG9yIG5vdCBoYXNhdHRyKHdpbmRvdywgIl9pbnB1dF9maWVsZCIpOgog"
    "ICAgICAgICAgICByZXR1cm4KICAgICAgICBsaW5lID0gc2VsZi5fZm9ybWF0X2V2ZW50X2Zvcl9wcm9tcHQoZXZlbnQpCiAgICAg"
    "ICAgd2luZG93Ll9pbnB1dF9maWVsZC5zZXRUZXh0KGxpbmUpCiAgICAgICAgd2luZG93Ll9pbnB1dF9maWVsZC5zZXRGb2N1cygp"
    "CgogICAgZGVmIF9wbGF5X3JvbGxfc291bmQoc2VsZik6CiAgICAgICAgaWYgbm90IFdJTlNPVU5EX09LOgogICAgICAgICAgICBy"
    "ZXR1cm4KICAgICAgICB0cnk6CiAgICAgICAgICAgIHdpbnNvdW5kLkJlZXAoODQwLCAzMCkKICAgICAgICAgICAgd2luc291bmQu"
    "QmVlcCg2MjAsIDM1KQogICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAgICAgICAgIHBhc3MKCgoKY2xhc3MgTWFnaWM4QmFs"
    "bFRhYihRV2lkZ2V0KToKICAgICIiIk1hZ2ljIDgtQmFsbCBtb2R1bGUgd2l0aCBjaXJjdWxhciBvcmIgZGlzcGxheSBhbmQgcHVs"
    "c2luZyBhbnN3ZXIgdGV4dC4iIiIKCiAgICBBTlNXRVJTID0gWwogICAgICAgICJJdCBpcyBjZXJ0YWluLiIsCiAgICAgICAgIkl0"
    "IGlzIGRlY2lkZWRseSBzby4iLAogICAgICAgICJXaXRob3V0IGEgZG91YnQuIiwKICAgICAgICAiWWVzIGRlZmluaXRlbHkuIiwK"
    "ICAgICAgICAiWW91IG1heSByZWx5IG9uIGl0LiIsCiAgICAgICAgIkFzIEkgc2VlIGl0LCB5ZXMuIiwKICAgICAgICAiTW9zdCBs"
    "aWtlbHkuIiwKICAgICAgICAiT3V0bG9vayBnb29kLiIsCiAgICAgICAgIlllcy4iLAogICAgICAgICJTaWducyBwb2ludCB0byB5"
    "ZXMuIiwKICAgICAgICAiUmVwbHkgaGF6eSwgdHJ5IGFnYWluLiIsCiAgICAgICAgIkFzayBhZ2FpbiBsYXRlci4iLAogICAgICAg"
    "ICJCZXR0ZXIgbm90IHRlbGwgeW91IG5vdy4iLAogICAgICAgICJDYW5ub3QgcHJlZGljdCBub3cuIiwKICAgICAgICAiQ29uY2Vu"
    "dHJhdGUgYW5kIGFzayBhZ2Fpbi4iLAogICAgICAgICJEb24ndCBjb3VudCBvbiBpdC4iLAogICAgICAgICJNeSByZXBseSBpcyBu"
    "by4iLAogICAgICAgICJNeSBzb3VyY2VzIHNheSBuby4iLAogICAgICAgICJPdXRsb29rIG5vdCBzbyBnb29kLiIsCiAgICAgICAg"
    "IlZlcnkgZG91YnRmdWwuIiwKICAgIF0KCiAgICBkZWYgX19pbml0X18oc2VsZiwgb25fdGhyb3c9Tm9uZSwgZGlhZ25vc3RpY3Nf"
    "bG9nZ2VyPU5vbmUpOgogICAgICAgIHN1cGVyKCkuX19pbml0X18oKQogICAgICAgIHNlbGYuX29uX3Rocm93ID0gb25fdGhyb3cK"
    "ICAgICAgICBzZWxmLl9sb2cgPSBkaWFnbm9zdGljc19sb2dnZXIgb3IgKGxhbWJkYSAqX2FyZ3MsICoqX2t3YXJnczogTm9uZSkK"
    "ICAgICAgICBzZWxmLl9jdXJyZW50X2Fuc3dlciA9ICIiCgogICAgICAgIHNlbGYuX2NsZWFyX3RpbWVyID0gUVRpbWVyKHNlbGYp"
    "CiAgICAgICAgc2VsZi5fY2xlYXJfdGltZXIuc2V0U2luZ2xlU2hvdChUcnVlKQogICAgICAgIHNlbGYuX2NsZWFyX3RpbWVyLnRp"
    "bWVvdXQuY29ubmVjdChzZWxmLl9mYWRlX291dF9hbnN3ZXIpCgogICAgICAgIHNlbGYuX2J1aWxkX3VpKCkKICAgICAgICBzZWxm"
    "Ll9idWlsZF9hbmltYXRpb25zKCkKICAgICAgICBzZWxmLl9zZXRfaWRsZV92aXN1YWwoKQoKICAgIGRlZiBfYnVpbGRfdWkoc2Vs"
    "ZikgLT4gTm9uZToKICAgICAgICByb290ID0gUVZCb3hMYXlvdXQoc2VsZikKICAgICAgICByb290LnNldENvbnRlbnRzTWFyZ2lu"
    "cygxNiwgMTYsIDE2LCAxNikKICAgICAgICByb290LnNldFNwYWNpbmcoMTQpCiAgICAgICAgcm9vdC5hZGRTdHJldGNoKDEpCgog"
    "ICAgICAgIHNlbGYuX29yYl9mcmFtZSA9IFFGcmFtZSgpCiAgICAgICAgc2VsZi5fb3JiX2ZyYW1lLnNldEZpeGVkU2l6ZSgyMjgs"
    "IDIyOCkKICAgICAgICBzZWxmLl9vcmJfZnJhbWUuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgIlFGcmFtZSB7IgogICAgICAg"
    "ICAgICAiYmFja2dyb3VuZC1jb2xvcjogIzA0MDQwNjsiCiAgICAgICAgICAgICJib3JkZXI6IDFweCBzb2xpZCByZ2JhKDIzNCwg"
    "MjM3LCAyNTUsIDAuNjIpOyIKICAgICAgICAgICAgImJvcmRlci1yYWRpdXM6IDExNHB4OyIKICAgICAgICAgICAgIn0iCiAgICAg"
    "ICAgKQoKICAgICAgICBvcmJfbGF5b3V0ID0gUVZCb3hMYXlvdXQoc2VsZi5fb3JiX2ZyYW1lKQogICAgICAgIG9yYl9sYXlvdXQu"
    "c2V0Q29udGVudHNNYXJnaW5zKDIwLCAyMCwgMjAsIDIwKQogICAgICAgIG9yYl9sYXlvdXQuc2V0U3BhY2luZygwKQoKICAgICAg"
    "ICBzZWxmLl9vcmJfaW5uZXIgPSBRRnJhbWUoKQogICAgICAgIHNlbGYuX29yYl9pbm5lci5zZXRTdHlsZVNoZWV0KAogICAgICAg"
    "ICAgICAiUUZyYW1lIHsiCiAgICAgICAgICAgICJiYWNrZ3JvdW5kLWNvbG9yOiAjMDcwNzBhOyIKICAgICAgICAgICAgImJvcmRl"
    "cjogMXB4IHNvbGlkIHJnYmEoMjU1LCAyNTUsIDI1NSwgMC4xMik7IgogICAgICAgICAgICAiYm9yZGVyLXJhZGl1czogODRweDsi"
    "CiAgICAgICAgICAgICJ9IgogICAgICAgICkKICAgICAgICBzZWxmLl9vcmJfaW5uZXIuc2V0TWluaW11bVNpemUoMTY4LCAxNjgp"
    "CiAgICAgICAgc2VsZi5fb3JiX2lubmVyLnNldE1heGltdW1TaXplKDE2OCwgMTY4KQoKICAgICAgICBpbm5lcl9sYXlvdXQgPSBR"
    "VkJveExheW91dChzZWxmLl9vcmJfaW5uZXIpCiAgICAgICAgaW5uZXJfbGF5b3V0LnNldENvbnRlbnRzTWFyZ2lucygxNiwgMTYs"
    "IDE2LCAxNikKICAgICAgICBpbm5lcl9sYXlvdXQuc2V0U3BhY2luZygwKQoKICAgICAgICBzZWxmLl9laWdodF9sYmwgPSBRTGFi"
    "ZWwoIjgiKQogICAgICAgIHNlbGYuX2VpZ2h0X2xibC5zZXRBbGlnbm1lbnQoUXQuQWxpZ25tZW50RmxhZy5BbGlnbkNlbnRlcikK"
    "ICAgICAgICBzZWxmLl9laWdodF9sYmwuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgImNvbG9yOiByZ2JhKDI1NSwgMjU1LCAy"
    "NTUsIDAuOTUpOyAiCiAgICAgICAgICAgICJmb250LXNpemU6IDgwcHg7IGZvbnQtd2VpZ2h0OiA3MDA7ICIKICAgICAgICAgICAg"
    "ImZvbnQtZmFtaWx5OiBHZW9yZ2lhLCBzZXJpZjsgYm9yZGVyOiBub25lOyIKICAgICAgICApCgogICAgICAgIHNlbGYuYW5zd2Vy"
    "X2xibCA9IFFMYWJlbCgiIikKICAgICAgICBzZWxmLmFuc3dlcl9sYmwuc2V0QWxpZ25tZW50KFF0LkFsaWdubWVudEZsYWcuQWxp"
    "Z25DZW50ZXIpCiAgICAgICAgc2VsZi5hbnN3ZXJfbGJsLnNldFdvcmRXcmFwKFRydWUpCiAgICAgICAgc2VsZi5hbnN3ZXJfbGJs"
    "LnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX0dPTER9OyBmb250LXNpemU6IDE2cHg7IGZvbnQtc3R5bGU6"
    "IGl0YWxpYzsgIgogICAgICAgICAgICAiZm9udC13ZWlnaHQ6IDYwMDsgYm9yZGVyOiBub25lOyBwYWRkaW5nOiAycHg7IgogICAg"
    "ICAgICkKCiAgICAgICAgaW5uZXJfbGF5b3V0LmFkZFdpZGdldChzZWxmLl9laWdodF9sYmwsIDEpCiAgICAgICAgaW5uZXJfbGF5"
    "b3V0LmFkZFdpZGdldChzZWxmLmFuc3dlcl9sYmwsIDEpCiAgICAgICAgb3JiX2xheW91dC5hZGRXaWRnZXQoc2VsZi5fb3JiX2lu"
    "bmVyLCAwLCBRdC5BbGlnbm1lbnRGbGFnLkFsaWduQ2VudGVyKQoKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl9vcmJfZnJh"
    "bWUsIDAsIFF0LkFsaWdubWVudEZsYWcuQWxpZ25IQ2VudGVyKQoKICAgICAgICBzZWxmLnRocm93X2J0biA9IFFQdXNoQnV0dG9u"
    "KCJUaHJvdyB0aGUgOC1CYWxsIikKICAgICAgICBzZWxmLnRocm93X2J0bi5zZXRGaXhlZEhlaWdodCgzOCkKICAgICAgICBzZWxm"
    "LnRocm93X2J0bi5jbGlja2VkLmNvbm5lY3Qoc2VsZi5fdGhyb3dfYmFsbCkKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLnRo"
    "cm93X2J0biwgMCwgUXQuQWxpZ25tZW50RmxhZy5BbGlnbkhDZW50ZXIpCiAgICAgICAgcm9vdC5hZGRTdHJldGNoKDEpCgogICAg"
    "ZGVmIF9idWlsZF9hbmltYXRpb25zKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fYW5zd2VyX29wYWNpdHkgPSBRR3JhcGhp"
    "Y3NPcGFjaXR5RWZmZWN0KHNlbGYuYW5zd2VyX2xibCkKICAgICAgICBzZWxmLmFuc3dlcl9sYmwuc2V0R3JhcGhpY3NFZmZlY3Qo"
    "c2VsZi5fYW5zd2VyX29wYWNpdHkpCiAgICAgICAgc2VsZi5fYW5zd2VyX29wYWNpdHkuc2V0T3BhY2l0eSgwLjApCgogICAgICAg"
    "IHNlbGYuX3B1bHNlX2FuaW0gPSBRUHJvcGVydHlBbmltYXRpb24oc2VsZi5fYW5zd2VyX29wYWNpdHksIGIib3BhY2l0eSIsIHNl"
    "bGYpCiAgICAgICAgc2VsZi5fcHVsc2VfYW5pbS5zZXREdXJhdGlvbig3NjApCiAgICAgICAgc2VsZi5fcHVsc2VfYW5pbS5zZXRT"
    "dGFydFZhbHVlKDAuMzUpCiAgICAgICAgc2VsZi5fcHVsc2VfYW5pbS5zZXRFbmRWYWx1ZSgxLjApCiAgICAgICAgc2VsZi5fcHVs"
    "c2VfYW5pbS5zZXRFYXNpbmdDdXJ2ZShRRWFzaW5nQ3VydmUuVHlwZS5Jbk91dFNpbmUpCiAgICAgICAgc2VsZi5fcHVsc2VfYW5p"
    "bS5zZXRMb29wQ291bnQoLTEpCgogICAgICAgIHNlbGYuX2ZhZGVfb3V0ID0gUVByb3BlcnR5QW5pbWF0aW9uKHNlbGYuX2Fuc3dl"
    "cl9vcGFjaXR5LCBiIm9wYWNpdHkiLCBzZWxmKQogICAgICAgIHNlbGYuX2ZhZGVfb3V0LnNldER1cmF0aW9uKDU2MCkKICAgICAg"
    "ICBzZWxmLl9mYWRlX291dC5zZXRTdGFydFZhbHVlKDEuMCkKICAgICAgICBzZWxmLl9mYWRlX291dC5zZXRFbmRWYWx1ZSgwLjAp"
    "CiAgICAgICAgc2VsZi5fZmFkZV9vdXQuc2V0RWFzaW5nQ3VydmUoUUVhc2luZ0N1cnZlLlR5cGUuSW5PdXRRdWFkKQogICAgICAg"
    "IHNlbGYuX2ZhZGVfb3V0LmZpbmlzaGVkLmNvbm5lY3Qoc2VsZi5fY2xlYXJfdG9faWRsZSkKCiAgICBkZWYgX3NldF9pZGxlX3Zp"
    "c3VhbChzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2N1cnJlbnRfYW5zd2VyID0gIiIKICAgICAgICBzZWxmLl9laWdodF9s"
    "Ymwuc2hvdygpCiAgICAgICAgc2VsZi5hbnN3ZXJfbGJsLmNsZWFyKCkKICAgICAgICBzZWxmLmFuc3dlcl9sYmwuaGlkZSgpCiAg"
    "ICAgICAgc2VsZi5fYW5zd2VyX29wYWNpdHkuc2V0T3BhY2l0eSgwLjApCgogICAgZGVmIF90aHJvd19iYWxsKHNlbGYpIC0+IE5v"
    "bmU6CiAgICAgICAgc2VsZi5fY2xlYXJfdGltZXIuc3RvcCgpCiAgICAgICAgc2VsZi5fcHVsc2VfYW5pbS5zdG9wKCkKICAgICAg"
    "ICBzZWxmLl9mYWRlX291dC5zdG9wKCkKCiAgICAgICAgYW5zd2VyID0gcmFuZG9tLmNob2ljZShzZWxmLkFOU1dFUlMpCiAgICAg"
    "ICAgc2VsZi5fY3VycmVudF9hbnN3ZXIgPSBhbnN3ZXIKCiAgICAgICAgc2VsZi5fZWlnaHRfbGJsLmhpZGUoKQogICAgICAgIHNl"
    "bGYuYW5zd2VyX2xibC5zZXRUZXh0KGFuc3dlcikKICAgICAgICBzZWxmLmFuc3dlcl9sYmwuc2hvdygpCiAgICAgICAgc2VsZi5f"
    "YW5zd2VyX29wYWNpdHkuc2V0T3BhY2l0eSgwLjApCiAgICAgICAgc2VsZi5fcHVsc2VfYW5pbS5zdGFydCgpCiAgICAgICAgc2Vs"
    "Zi5fY2xlYXJfdGltZXIuc3RhcnQoNjAwMDApCiAgICAgICAgc2VsZi5fbG9nKGYiWzhCQUxMXSBUaHJvdyByZXN1bHQ6IHthbnN3"
    "ZXJ9IiwgIklORk8iKQoKICAgICAgICBpZiBjYWxsYWJsZShzZWxmLl9vbl90aHJvdyk6CiAgICAgICAgICAgIHRyeToKICAgICAg"
    "ICAgICAgICAgIHNlbGYuX29uX3Rocm93KGFuc3dlcikKICAgICAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbiBhcyBleDoKICAgICAg"
    "ICAgICAgICAgIHNlbGYuX2xvZyhmIls4QkFMTF1bV0FSTl0gSW50ZXJuYWwgcHJvbXB0IGRpc3BhdGNoIGZhaWxlZDoge2V4fSIs"
    "ICJXQVJOIikKCiAgICBkZWYgX2ZhZGVfb3V0X2Fuc3dlcihzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX2NsZWFyX3RpbWVy"
    "LnN0b3AoKQogICAgICAgIHNlbGYuX3B1bHNlX2FuaW0uc3RvcCgpCiAgICAgICAgc2VsZi5fZmFkZV9vdXQuc3RvcCgpCiAgICAg"
    "ICAgc2VsZi5fZmFkZV9vdXQuc2V0U3RhcnRWYWx1ZShmbG9hdChzZWxmLl9hbnN3ZXJfb3BhY2l0eS5vcGFjaXR5KCkpKQogICAg"
    "ICAgIHNlbGYuX2ZhZGVfb3V0LnNldEVuZFZhbHVlKDAuMCkKICAgICAgICBzZWxmLl9mYWRlX291dC5zdGFydCgpCgogICAgZGVm"
    "IF9jbGVhcl90b19pZGxlKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fZmFkZV9vdXQuc3RvcCgpCiAgICAgICAgc2VsZi5f"
    "c2V0X2lkbGVfdmlzdWFsKCkKCiMg4pSA4pSAIE1BSU4gV0lORE9XIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgApjbGFzcyBMb2NrQXdhcmVUYWJCYXIoUVRhYkJhcik6CiAgICAiIiJUYWIgYmFyIHRoYXQgYmxvY2tzIGRyYWcgaW5pdGlh"
    "dGlvbiBmb3IgbG9ja2VkIHRhYnMuIiIiCgogICAgZGVmIF9faW5pdF9fKHNlbGYsIGlzX2xvY2tlZF9ieV9pZCwgcGFyZW50PU5v"
    "bmUpOgogICAgICAgIHN1cGVyKCkuX19pbml0X18ocGFyZW50KQogICAgICAgIHNlbGYuX2lzX2xvY2tlZF9ieV9pZCA9IGlzX2xv"
    "Y2tlZF9ieV9pZAogICAgICAgIHNlbGYuX3ByZXNzZWRfaW5kZXggPSAtMQoKICAgIGRlZiBfdGFiX2lkKHNlbGYsIGluZGV4OiBp"
    "bnQpOgogICAgICAgIGlmIGluZGV4IDwgMDoKICAgICAgICAgICAgcmV0dXJuIE5vbmUKICAgICAgICByZXR1cm4gc2VsZi50YWJE"
    "YXRhKGluZGV4KQoKICAgIGRlZiBtb3VzZVByZXNzRXZlbnQoc2VsZiwgZXZlbnQpOgogICAgICAgIHNlbGYuX3ByZXNzZWRfaW5k"
    "ZXggPSBzZWxmLnRhYkF0KGV2ZW50LnBvcygpKQogICAgICAgIGlmIChldmVudC5idXR0b24oKSA9PSBRdC5Nb3VzZUJ1dHRvbi5M"
    "ZWZ0QnV0dG9uIGFuZCBzZWxmLl9wcmVzc2VkX2luZGV4ID49IDApOgogICAgICAgICAgICB0YWJfaWQgPSBzZWxmLl90YWJfaWQo"
    "c2VsZi5fcHJlc3NlZF9pbmRleCkKICAgICAgICAgICAgaWYgdGFiX2lkIGFuZCBzZWxmLl9pc19sb2NrZWRfYnlfaWQodGFiX2lk"
    "KToKICAgICAgICAgICAgICAgIHNlbGYuc2V0Q3VycmVudEluZGV4KHNlbGYuX3ByZXNzZWRfaW5kZXgpCiAgICAgICAgICAgICAg"
    "ICBldmVudC5hY2NlcHQoKQogICAgICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgc3VwZXIoKS5tb3VzZVByZXNzRXZlbnQoZXZl"
    "bnQpCgogICAgZGVmIG1vdXNlTW92ZUV2ZW50KHNlbGYsIGV2ZW50KToKICAgICAgICBpZiBzZWxmLl9wcmVzc2VkX2luZGV4ID49"
    "IDA6CiAgICAgICAgICAgIHRhYl9pZCA9IHNlbGYuX3RhYl9pZChzZWxmLl9wcmVzc2VkX2luZGV4KQogICAgICAgICAgICBpZiB0"
    "YWJfaWQgYW5kIHNlbGYuX2lzX2xvY2tlZF9ieV9pZCh0YWJfaWQpOgogICAgICAgICAgICAgICAgZXZlbnQuYWNjZXB0KCkKICAg"
    "ICAgICAgICAgICAgIHJldHVybgogICAgICAgIHN1cGVyKCkubW91c2VNb3ZlRXZlbnQoZXZlbnQpCgogICAgZGVmIG1vdXNlUmVs"
    "ZWFzZUV2ZW50KHNlbGYsIGV2ZW50KToKICAgICAgICBzZWxmLl9wcmVzc2VkX2luZGV4ID0gLTEKICAgICAgICBzdXBlcigpLm1v"
    "dXNlUmVsZWFzZUV2ZW50KGV2ZW50KQoKCmNsYXNzIEVjaG9EZWNrKFFNYWluV2luZG93KToKICAgICIiIgogICAgVGhlIG1haW4g"
    "RWNobyBEZWNrIHdpbmRvdy4KICAgIEFzc2VtYmxlcyBhbGwgd2lkZ2V0cywgY29ubmVjdHMgYWxsIHNpZ25hbHMsIG1hbmFnZXMg"
    "YWxsIHN0YXRlLgogICAgIiIiCgogICAgIyDilIDilIAgVG9ycG9yIHRocmVzaG9sZHMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBfRVhURVJOQUxfVlJBTV9UT1JQT1JfR0IgICAgPSAx"
    "LjUgICAjIGV4dGVybmFsIFZSQU0gPiB0aGlzIOKGkiBjb25zaWRlciB0b3Jwb3IKICAgIF9FWFRFUk5BTF9WUkFNX1dBS0VfR0Ig"
    "ICAgICA9IDAuOCAgICMgZXh0ZXJuYWwgVlJBTSA8IHRoaXMg4oaSIGNvbnNpZGVyIHdha2UKICAgIF9UT1JQT1JfU1VTVEFJTkVE"
    "X1RJQ0tTICAgICA9IDYgICAgICMgNiDDlyA1cyA9IDMwIHNlY29uZHMgc3VzdGFpbmVkCiAgICBfV0FLRV9TVVNUQUlORURfVElD"
    "S1MgICAgICAgPSAxMiAgICAjIDYwIHNlY29uZHMgc3VzdGFpbmVkIGxvdyBwcmVzc3VyZQoKICAgIGRlZiBfX2luaXRfXyhzZWxm"
    "KToKICAgICAgICBzdXBlcigpLl9faW5pdF9fKCkKCiAgICAgICAgIyDilIDilIAgQ29yZSBzdGF0ZSDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBzZWxmLl9zdGF0"
    "dXMgICAgICAgICAgICAgID0gIk9GRkxJTkUiCiAgICAgICAgc2VsZi5fc2Vzc2lvbl9zdGFydCAgICAgICA9IHRpbWUudGltZSgp"
    "CiAgICAgICAgc2VsZi5fdG9rZW5fY291bnQgICAgICAgICA9IDAKICAgICAgICBzZWxmLl9mYWNlX2xvY2tlZCAgICAgICAgID0g"
    "RmFsc2UKICAgICAgICBzZWxmLl9ibGlua19zdGF0ZSAgICAgICAgID0gVHJ1ZQogICAgICAgIHNlbGYuX21vZGVsX2xvYWRlZCAg"
    "ICAgICAgPSBGYWxzZQogICAgICAgIHNlbGYuX3Nlc3Npb25faWQgICAgICAgICAgPSBmInNlc3Npb25fe2RhdGV0aW1lLm5vdygp"
    "LnN0cmZ0aW1lKCclWSVtJWRfJUglTSVTJyl9IgogICAgICAgIHNlbGYuX2FjdGl2ZV90aHJlYWRzOiBsaXN0ID0gW10gICMga2Vl"
    "cCByZWZzIHRvIHByZXZlbnQgR0Mgd2hpbGUgcnVubmluZwogICAgICAgIHNlbGYuX2ZpcnN0X3Rva2VuOiBib29sID0gVHJ1ZSAg"
    "ICMgd3JpdGUgc3BlYWtlciBsYWJlbCBiZWZvcmUgZmlyc3Qgc3RyZWFtaW5nIHRva2VuCgogICAgICAgICMgVG9ycG9yIC8gVlJB"
    "TSB0cmFja2luZwogICAgICAgIHNlbGYuX3RvcnBvcl9zdGF0ZSAgICAgICAgPSAiQVdBS0UiCiAgICAgICAgc2VsZi5fZGVja192"
    "cmFtX2Jhc2UgID0gMC4wICAgIyBiYXNlbGluZSBWUkFNIGFmdGVyIG1vZGVsIGxvYWQKICAgICAgICBzZWxmLl92cmFtX3ByZXNz"
    "dXJlX3RpY2tzID0gMCAgICAgIyBzdXN0YWluZWQgcHJlc3N1cmUgY291bnRlcgogICAgICAgIHNlbGYuX3ZyYW1fcmVsaWVmX3Rp"
    "Y2tzICAgPSAwICAgICAjIHN1c3RhaW5lZCByZWxpZWYgY291bnRlcgogICAgICAgIHNlbGYuX3BlbmRpbmdfdHJhbnNtaXNzaW9u"
    "cyA9IDAKICAgICAgICBzZWxmLl90b3Jwb3Jfc2luY2UgICAgICAgID0gTm9uZSAgIyBkYXRldGltZSB3aGVuIHRvcnBvciBiZWdh"
    "bgogICAgICAgIHNlbGYuX3N1c3BlbmRlZF9kdXJhdGlvbiAgPSAiIiAgICMgZm9ybWF0dGVkIGR1cmF0aW9uIHN0cmluZwoKICAg"
    "ICAgICAjIOKUgOKUgCBNYW5hZ2VycyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBzZWxmLl9tZW1vcnkgICA9IE1lbW9yeU1hbmFnZXIoKQogICAgICAg"
    "IHNlbGYuX3Nlc3Npb25zID0gU2Vzc2lvbk1hbmFnZXIoKQogICAgICAgIHNlbGYuX2xlc3NvbnMgID0gTGVzc29uc0xlYXJuZWRE"
    "QigpCiAgICAgICAgc2VsZi5fdGFza3MgICAgPSBUYXNrTWFuYWdlcigpCiAgICAgICAgc2VsZi5fdGFza19zaG93X2NvbXBsZXRl"
    "ZCA9IEZhbHNlCiAgICAgICAgc2VsZi5fdGFza19kYXRlX2ZpbHRlciA9ICJuZXh0XzNfbW9udGhzIgoKICAgICAgICAjIFJpZ2h0"
    "IHN5c3RlbXMgdGFiLXN0cmlwIHByZXNlbnRhdGlvbiBzdGF0ZSAoc3RhYmxlIElEcyArIHZpc3VhbCBvcmRlcikKICAgICAgICBz"
    "ZWxmLl9zcGVsbF90YWJfZGVmczogbGlzdFtkaWN0XSA9IFtdCiAgICAgICAgc2VsZi5fc3BlbGxfdGFiX3N0YXRlOiBkaWN0W3N0"
    "ciwgZGljdF0gPSB7fQogICAgICAgIHNlbGYuX3NwZWxsX3RhYl9tb3ZlX21vZGVfaWQ6IE9wdGlvbmFsW3N0cl0gPSBOb25lCiAg"
    "ICAgICAgc2VsZi5fc3VwcHJlc3Nfc3BlbGxfdGFiX21vdmVfc2lnbmFsID0gRmFsc2UKICAgICAgICBzZWxmLl9mb2N1c19ob29r"
    "ZWRfZm9yX3NwZWxsX3RhYnMgPSBGYWxzZQoKICAgICAgICAjIOKUgOKUgCBHb29nbGUgU2VydmljZXMg4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACgogICAgICAgICMgU2VlZCBMU0wgcnVsZXMgb24gZmly"
    "c3QgcnVuCiAgICAgICAgc2VsZi5fbGVzc29ucy5zZWVkX2xzbF9ydWxlcygpCgogICAgICAgICMgTG9hZCBlbnRpdHkgc3RhdGUK"
    "ICAgICAgICBzZWxmLl9zdGF0ZSA9IHNlbGYuX21lbW9yeS5sb2FkX3N0YXRlKCkKICAgICAgICBzZWxmLl9zdGF0ZVsic2Vzc2lv"
    "bl9jb3VudCJdID0gc2VsZi5fc3RhdGUuZ2V0KCJzZXNzaW9uX2NvdW50IiwwKSArIDEKICAgICAgICBzZWxmLl9zdGF0ZVsibGFz"
    "dF9zdGFydHVwIl0gID0gbG9jYWxfbm93X2lzbygpCiAgICAgICAgc2VsZi5fbWVtb3J5LnNhdmVfc3RhdGUoc2VsZi5fc3RhdGUp"
    "CgogICAgICAgICMgQnVpbGQgYWRhcHRvcgogICAgICAgIHNlbGYuX2FkYXB0b3IgPSBidWlsZF9hZGFwdG9yX2Zyb21fY29uZmln"
    "KCkKCiAgICAgICAgIyBGYWNlIHRpbWVyIG1hbmFnZXIgKHNldCB1cCBhZnRlciB3aWRnZXRzIGJ1aWx0KQogICAgICAgIHNlbGYu"
    "X2ZhY2VfdGltZXJfbWdyOiBPcHRpb25hbFtGYWNlVGltZXJNYW5hZ2VyXSA9IE5vbmUKCiAgICAgICAgIyDilIDilIAgQnVpbGQg"
    "VUkg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSACiAgICAgICAgc2VsZi5zZXRXaW5kb3dUaXRsZShBUFBfTkFNRSkKICAgICAgICBzZWxmLnNldE1pbmltdW1TaXplKDEy"
    "MDAsIDc1MCkKICAgICAgICBzZWxmLnJlc2l6ZSgxMzUwLCA4NTApCiAgICAgICAgc2VsZi5zZXRTdHlsZVNoZWV0KFNUWUxFKQoK"
    "ICAgICAgICBzZWxmLl9idWlsZF91aSgpCgogICAgICAgICMgRmFjZSB0aW1lciBtYW5hZ2VyIHdpcmVkIHRvIHdpZGdldHMKICAg"
    "ICAgICBzZWxmLl9mYWNlX3RpbWVyX21nciA9IEZhY2VUaW1lck1hbmFnZXIoCiAgICAgICAgICAgIHNlbGYuX21pcnJvciwgc2Vs"
    "Zi5fZW1vdGlvbl9ibG9jawogICAgICAgICkKCiAgICAgICAgIyDilIDilIAgVGltZXJzIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIHNlbGYu"
    "X3N0YXRzX3RpbWVyID0gUVRpbWVyKCkKICAgICAgICBzZWxmLl9zdGF0c190aW1lci50aW1lb3V0LmNvbm5lY3Qoc2VsZi5fdXBk"
    "YXRlX3N0YXRzKQogICAgICAgIHNlbGYuX3N0YXRzX3RpbWVyLnN0YXJ0KDEwMDApCgogICAgICAgIHNlbGYuX2JsaW5rX3RpbWVy"
    "ID0gUVRpbWVyKCkKICAgICAgICBzZWxmLl9ibGlua190aW1lci50aW1lb3V0LmNvbm5lY3Qoc2VsZi5fYmxpbmspCiAgICAgICAg"
    "c2VsZi5fYmxpbmtfdGltZXIuc3RhcnQoODAwKQoKICAgICAgICBzZWxmLl9zdGF0ZV9zdHJpcF90aW1lciA9IFFUaW1lcigpCiAg"
    "ICAgICAgaWYgQUlfU1RBVEVTX0VOQUJMRUQgYW5kIHNlbGYuX2Zvb3Rlcl9zdHJpcCBpcyBub3QgTm9uZToKICAgICAgICAgICAg"
    "c2VsZi5fc3RhdGVfc3RyaXBfdGltZXIudGltZW91dC5jb25uZWN0KHNlbGYuX2Zvb3Rlcl9zdHJpcC5yZWZyZXNoKQogICAgICAg"
    "ICAgICBzZWxmLl9zdGF0ZV9zdHJpcF90aW1lci5zdGFydCg2MDAwMCkKCiAgICAgICAgIyDilIDilIAgU2NoZWR1bGVyIGFuZCBz"
    "dGFydHVwIGRlZmVycmVkIHVudGlsIGFmdGVyIHdpbmRvdy5zaG93KCkg4pSA4pSA4pSACiAgICAgICAgIyBEbyBOT1QgY2FsbCBf"
    "c2V0dXBfc2NoZWR1bGVyKCkgb3IgX3N0YXJ0dXBfc2VxdWVuY2UoKSBoZXJlLgogICAgICAgICMgQm90aCBhcmUgdHJpZ2dlcmVk"
    "IHZpYSBRVGltZXIuc2luZ2xlU2hvdCBmcm9tIG1haW4oKSBhZnRlcgogICAgICAgICMgd2luZG93LnNob3coKSBhbmQgYXBwLmV4"
    "ZWMoKSBiZWdpbnMgcnVubmluZy4KCiAgICAjIOKUgOKUgCBVSSBDT05TVFJVQ1RJT04g4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBk"
    "ZWYgX2J1aWxkX3VpKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgY2VudHJhbCA9IFFXaWRnZXQoKQogICAgICAgIHNlbGYuc2V0Q2Vu"
    "dHJhbFdpZGdldChjZW50cmFsKQogICAgICAgIHJvb3QgPSBRVkJveExheW91dChjZW50cmFsKQogICAgICAgIHJvb3Quc2V0Q29u"
    "dGVudHNNYXJnaW5zKDYsIDYsIDYsIDYpCiAgICAgICAgcm9vdC5zZXRTcGFjaW5nKDQpCgogICAgICAgICMg4pSA4pSAIFRpdGxl"
    "IGJhciDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIAKICAgICAgICByb290LmFkZFdpZGdldChzZWxmLl9idWlsZF90aXRsZV9iYXIoKSkKCiAgICAgICAgIyDilIDilIAgQm9k"
    "eTogbGVmdCB3b3Jrc3BhY2UgfCByaWdodCBzeXN0ZW1zIChkcmFnZ2FibGUgc3BsaXR0ZXIpIOKUgAogICAgICAgIHNlbGYuX21h"
    "aW5fc3BsaXR0ZXIgPSBRU3BsaXR0ZXIoUXQuT3JpZW50YXRpb24uSG9yaXpvbnRhbCkKICAgICAgICBzZWxmLl9tYWluX3NwbGl0"
    "dGVyLnNldENoaWxkcmVuQ29sbGFwc2libGUoRmFsc2UpCiAgICAgICAgc2VsZi5fbWFpbl9zcGxpdHRlci5zZXRIYW5kbGVXaWR0"
    "aCg4KQoKICAgICAgICAjIExlZnQgcGFuZSA9IEpvdXJuYWwgKyBDaGF0IHdvcmtzcGFjZQogICAgICAgIGxlZnRfd29ya3NwYWNl"
    "ID0gUVdpZGdldCgpCiAgICAgICAgbGVmdF93b3Jrc3BhY2Uuc2V0TWluaW11bVdpZHRoKDcwMCkKICAgICAgICBsZWZ0X2xheW91"
    "dCA9IFFIQm94TGF5b3V0KGxlZnRfd29ya3NwYWNlKQogICAgICAgIGxlZnRfbGF5b3V0LnNldENvbnRlbnRzTWFyZ2lucygwLCAw"
    "LCAwLCAwKQogICAgICAgIGxlZnRfbGF5b3V0LnNldFNwYWNpbmcoNCkKCiAgICAgICAgc2VsZi5fam91cm5hbF9zaWRlYmFyID0g"
    "Sm91cm5hbFNpZGViYXIoc2VsZi5fc2Vzc2lvbnMpCiAgICAgICAgc2VsZi5fam91cm5hbF9zaWRlYmFyLnNlc3Npb25fbG9hZF9y"
    "ZXF1ZXN0ZWQuY29ubmVjdCgKICAgICAgICAgICAgc2VsZi5fbG9hZF9qb3VybmFsX3Nlc3Npb24pCiAgICAgICAgc2VsZi5fam91"
    "cm5hbF9zaWRlYmFyLnNlc3Npb25fY2xlYXJfcmVxdWVzdGVkLmNvbm5lY3QoCiAgICAgICAgICAgIHNlbGYuX2NsZWFyX2pvdXJu"
    "YWxfc2Vzc2lvbikKICAgICAgICBsZWZ0X2xheW91dC5hZGRXaWRnZXQoc2VsZi5fam91cm5hbF9zaWRlYmFyKQogICAgICAgIGxl"
    "ZnRfbGF5b3V0LmFkZExheW91dChzZWxmLl9idWlsZF9jaGF0X3BhbmVsKCksIDEpCgogICAgICAgICMgUmlnaHQgcGFuZSA9IHN5"
    "c3RlbXMvbW9kdWxlcyArIGNhbGVuZGFyCiAgICAgICAgcmlnaHRfd29ya3NwYWNlID0gUVdpZGdldCgpCiAgICAgICAgcmlnaHRf"
    "d29ya3NwYWNlLnNldE1pbmltdW1XaWR0aCgzNjApCiAgICAgICAgcmlnaHRfbGF5b3V0ID0gUVZCb3hMYXlvdXQocmlnaHRfd29y"
    "a3NwYWNlKQogICAgICAgIHJpZ2h0X2xheW91dC5zZXRDb250ZW50c01hcmdpbnMoMCwgMCwgMCwgMCkKICAgICAgICByaWdodF9s"
    "YXlvdXQuc2V0U3BhY2luZygwKQogICAgICAgIHJpZ2h0X2xheW91dC5hZGRMYXlvdXQoc2VsZi5fYnVpbGRfc3BlbGxib29rX3Bh"
    "bmVsKCksIDEpCgogICAgICAgIHNlbGYuX21haW5fc3BsaXR0ZXIuYWRkV2lkZ2V0KGxlZnRfd29ya3NwYWNlKQogICAgICAgIHNl"
    "bGYuX21haW5fc3BsaXR0ZXIuYWRkV2lkZ2V0KHJpZ2h0X3dvcmtzcGFjZSkKICAgICAgICBzZWxmLl9tYWluX3NwbGl0dGVyLnNl"
    "dENvbGxhcHNpYmxlKDAsIEZhbHNlKQogICAgICAgIHNlbGYuX21haW5fc3BsaXR0ZXIuc2V0Q29sbGFwc2libGUoMSwgRmFsc2Up"
    "CiAgICAgICAgc2VsZi5fbWFpbl9zcGxpdHRlci5zcGxpdHRlck1vdmVkLmNvbm5lY3Qoc2VsZi5fc2F2ZV9tYWluX3NwbGl0dGVy"
    "X3N0YXRlKQogICAgICAgIHJvb3QuYWRkV2lkZ2V0KHNlbGYuX21haW5fc3BsaXR0ZXIsIDEpCiAgICAgICAgUVRpbWVyLnNpbmds"
    "ZVNob3QoMCwgc2VsZi5fcmVzdG9yZV9tYWluX3NwbGl0dGVyX3N0YXRlKQoKICAgICAgICAjIOKUgOKUgCBGb290ZXIg4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSACiAgICAgICAgZm9vdGVyID0gUUxhYmVsKAogICAgICAgICAgICBmIuKcpiB7QVBQX05BTUV9IOKAlCB2e0FQUF9WRVJTSU9O"
    "fSDinKYiCiAgICAgICAgKQogICAgICAgIGZvb3Rlci5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19URVhU"
    "X0RJTX07IGZvbnQtc2l6ZTogOXB4OyBsZXR0ZXItc3BhY2luZzogMnB4OyAiCiAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtE"
    "RUNLX0ZPTlR9LCBzZXJpZjsiCiAgICAgICAgKQogICAgICAgIGZvb3Rlci5zZXRBbGlnbm1lbnQoUXQuQWxpZ25tZW50RmxhZy5B"
    "bGlnbkNlbnRlcikKICAgICAgICByb290LmFkZFdpZGdldChmb290ZXIpCgogICAgZGVmIF9idWlsZF90aXRsZV9iYXIoc2VsZikg"
    "LT4gUVdpZGdldDoKICAgICAgICBiYXIgPSBRV2lkZ2V0KCkKICAgICAgICBiYXIuc2V0Rml4ZWRIZWlnaHQoMzYpCiAgICAgICAg"
    "YmFyLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkcyfTsgYm9yZGVyOiAxcHggc29saWQge0Nf"
    "Q1JJTVNPTl9ESU19OyAiCiAgICAgICAgICAgIGYiYm9yZGVyLXJhZGl1czogMnB4OyIKICAgICAgICApCiAgICAgICAgbGF5b3V0"
    "ID0gUUhCb3hMYXlvdXQoYmFyKQogICAgICAgIGxheW91dC5zZXRDb250ZW50c01hcmdpbnMoMTAsIDAsIDEwLCAwKQogICAgICAg"
    "IGxheW91dC5zZXRTcGFjaW5nKDYpCgogICAgICAgIHRpdGxlID0gUUxhYmVsKGYi4pymIHtBUFBfTkFNRX0iKQogICAgICAgIHRp"
    "dGxlLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX0NSSU1TT059OyBmb250LXNpemU6IDEzcHg7IGZvbnQt"
    "d2VpZ2h0OiBib2xkOyAiCiAgICAgICAgICAgIGYibGV0dGVyLXNwYWNpbmc6IDJweDsgYm9yZGVyOiBub25lOyBmb250LWZhbWls"
    "eToge0RFQ0tfRk9OVH0sIHNlcmlmOyIKICAgICAgICApCgogICAgICAgIHJ1bmVzID0gUUxhYmVsKFJVTkVTKQogICAgICAgIHJ1"
    "bmVzLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiY29sb3I6IHtDX0dPTERfRElNfTsgZm9udC1zaXplOiAxMHB4OyBib3Jk"
    "ZXI6IG5vbmU7IgogICAgICAgICkKICAgICAgICBydW5lcy5zZXRBbGlnbm1lbnQoUXQuQWxpZ25tZW50RmxhZy5BbGlnbkNlbnRl"
    "cikKCiAgICAgICAgc2VsZi5zdGF0dXNfbGFiZWwgPSBRTGFiZWwoZiLil4kge1VJX09GRkxJTkVfU1RBVFVTfSIpCiAgICAgICAg"
    "c2VsZi5zdGF0dXNfbGFiZWwuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJjb2xvcjoge0NfQkxPT0R9OyBmb250LXNpemU6"
    "IDEycHg7IGZvbnQtd2VpZ2h0OiBib2xkOyBib3JkZXI6IG5vbmU7IgogICAgICAgICkKICAgICAgICBzZWxmLnN0YXR1c19sYWJl"
    "bC5zZXRBbGlnbm1lbnQoUXQuQWxpZ25tZW50RmxhZy5BbGlnblJpZ2h0KQoKICAgICAgICAjIFN1c3BlbnNpb24gcGFuZWwKICAg"
    "ICAgICBzZWxmLl90b3Jwb3JfcGFuZWwgPSBOb25lCiAgICAgICAgaWYgU1VTUEVOU0lPTl9FTkFCTEVEOgogICAgICAgICAgICBz"
    "ZWxmLl90b3Jwb3JfcGFuZWwgPSBUb3Jwb3JQYW5lbCgpCiAgICAgICAgICAgIHNlbGYuX3RvcnBvcl9wYW5lbC5zdGF0ZV9jaGFu"
    "Z2VkLmNvbm5lY3Qoc2VsZi5fb25fdG9ycG9yX3N0YXRlX2NoYW5nZWQpCgogICAgICAgICMgSWRsZSB0b2dnbGUKICAgICAgICBp"
    "ZGxlX2VuYWJsZWQgPSBib29sKENGRy5nZXQoInNldHRpbmdzIiwge30pLmdldCgiaWRsZV9lbmFibGVkIiwgRmFsc2UpKQogICAg"
    "ICAgIHNlbGYuX2lkbGVfYnRuID0gUVB1c2hCdXR0b24oIklETEUgT04iIGlmIGlkbGVfZW5hYmxlZCBlbHNlICJJRExFIE9GRiIp"
    "CiAgICAgICAgc2VsZi5faWRsZV9idG4uc2V0Rml4ZWRIZWlnaHQoMjIpCiAgICAgICAgc2VsZi5faWRsZV9idG4uc2V0Q2hlY2th"
    "YmxlKFRydWUpCiAgICAgICAgc2VsZi5faWRsZV9idG4uc2V0Q2hlY2tlZChpZGxlX2VuYWJsZWQpCiAgICAgICAgc2VsZi5faWRs"
    "ZV9idG4uc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19CRzN9OyBjb2xvcjoge0NfVEVYVF9ESU19"
    "OyAiCiAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsgYm9yZGVyLXJhZGl1czogMnB4OyAiCiAgICAg"
    "ICAgICAgIGYiZm9udC1zaXplOiA5cHg7IGZvbnQtd2VpZ2h0OiBib2xkOyBwYWRkaW5nOiAzcHggOHB4OyIKICAgICAgICApCiAg"
    "ICAgICAgc2VsZi5faWRsZV9idG4udG9nZ2xlZC5jb25uZWN0KHNlbGYuX29uX2lkbGVfdG9nZ2xlZCkKCiAgICAgICAgIyBGUyAv"
    "IEJMIGJ1dHRvbnMKICAgICAgICBzZWxmLl9mc19idG4gPSBRUHVzaEJ1dHRvbigiRnVsbHNjcmVlbiIpCiAgICAgICAgc2VsZi5f"
    "YmxfYnRuID0gUVB1c2hCdXR0b24oIkJvcmRlcmxlc3MiKQogICAgICAgIHNlbGYuX2V4cG9ydF9idG4gPSBRUHVzaEJ1dHRvbigi"
    "RXhwb3J0IikKICAgICAgICBzZWxmLl9zaHV0ZG93bl9idG4gPSBRUHVzaEJ1dHRvbigiU2h1dGRvd24iKQogICAgICAgIGZvciBi"
    "dG4gaW4gKHNlbGYuX2ZzX2J0biwgc2VsZi5fYmxfYnRuLCBzZWxmLl9leHBvcnRfYnRuKToKICAgICAgICAgICAgYnRuLnNldEZp"
    "eGVkSGVpZ2h0KDIyKQogICAgICAgICAgICBidG4uc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgICAgIGYiYmFja2dyb3VuZDog"
    "e0NfQkczfTsgY29sb3I6IHtDX0NSSU1TT05fRElNfTsgIgogICAgICAgICAgICAgICAgZiJib3JkZXI6IDFweCBzb2xpZCB7Q19D"
    "UklNU09OX0RJTX07IGZvbnQtc2l6ZTogOXB4OyAiCiAgICAgICAgICAgICAgICBmImZvbnQtd2VpZ2h0OiBib2xkOyBwYWRkaW5n"
    "OiAwOyIKICAgICAgICAgICAgKQogICAgICAgIHNlbGYuX2V4cG9ydF9idG4uc2V0Rml4ZWRXaWR0aCg0NikKICAgICAgICBzZWxm"
    "Ll9zaHV0ZG93bl9idG4uc2V0Rml4ZWRIZWlnaHQoMjIpCiAgICAgICAgc2VsZi5fc2h1dGRvd25fYnRuLnNldEZpeGVkV2lkdGgo"
    "NjgpCiAgICAgICAgc2VsZi5fc2h1dGRvd25fYnRuLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0Nf"
    "QkczfTsgY29sb3I6IHtDX0JMT09EfTsgIgogICAgICAgICAgICBmImJvcmRlcjogMXB4IHNvbGlkIHtDX0JMT09EfTsgZm9udC1z"
    "aXplOiA5cHg7ICIKICAgICAgICAgICAgZiJmb250LXdlaWdodDogYm9sZDsgcGFkZGluZzogMDsiCiAgICAgICAgKQogICAgICAg"
    "IHNlbGYuX2ZzX2J0bi5zZXRUb29sVGlwKCJGdWxsc2NyZWVuIChGMTEpIikKICAgICAgICBzZWxmLl9ibF9idG4uc2V0VG9vbFRp"
    "cCgiQm9yZGVybGVzcyAoRjEwKSIpCiAgICAgICAgc2VsZi5fZXhwb3J0X2J0bi5zZXRUb29sVGlwKCJFeHBvcnQgY2hhdCBzZXNz"
    "aW9uIHRvIFRYVCBmaWxlIikKICAgICAgICBzZWxmLl9zaHV0ZG93bl9idG4uc2V0VG9vbFRpcChmIkdyYWNlZnVsIHNodXRkb3du"
    "IOKAlCB7REVDS19OQU1FfSBzcGVha3MgdGhlaXIgbGFzdCB3b3JkcyIpCiAgICAgICAgc2VsZi5fZnNfYnRuLmNsaWNrZWQuY29u"
    "bmVjdChzZWxmLl90b2dnbGVfZnVsbHNjcmVlbikKICAgICAgICBzZWxmLl9ibF9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYuX3Rv"
    "Z2dsZV9ib3JkZXJsZXNzKQogICAgICAgIHNlbGYuX2V4cG9ydF9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYuX2V4cG9ydF9jaGF0"
    "KQogICAgICAgIHNlbGYuX3NodXRkb3duX2J0bi5jbGlja2VkLmNvbm5lY3Qoc2VsZi5faW5pdGlhdGVfc2h1dGRvd25fZGlhbG9n"
    "KQoKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KHRpdGxlKQogICAgICAgIGxheW91dC5hZGRXaWRnZXQocnVuZXMsIDEpCiAgICAg"
    "ICAgbGF5b3V0LmFkZFdpZGdldChzZWxmLnN0YXR1c19sYWJlbCkKICAgICAgICBsYXlvdXQuYWRkU3BhY2luZyg4KQogICAgICAg"
    "IGxheW91dC5hZGRXaWRnZXQoc2VsZi5fZXhwb3J0X2J0bikKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX3NodXRkb3du"
    "X2J0bikKCiAgICAgICAgcmV0dXJuIGJhcgoKICAgIGRlZiBfYnVpbGRfY2hhdF9wYW5lbChzZWxmKSAtPiBRVkJveExheW91dDoK"
    "ICAgICAgICBsYXlvdXQgPSBRVkJveExheW91dCgpCiAgICAgICAgbGF5b3V0LnNldFNwYWNpbmcoNCkKCiAgICAgICAgIyBNYWlu"
    "IHRhYiB3aWRnZXQg4oCUIHBlcnNvbmEgY2hhdCB0YWIgfCBTZWxmCiAgICAgICAgc2VsZi5fbWFpbl90YWJzID0gUVRhYldpZGdl"
    "dCgpCiAgICAgICAgc2VsZi5fbWFpbl90YWJzLnNldFN0eWxlU2hlZXQoCiAgICAgICAgICAgIGYiUVRhYldpZGdldDo6cGFuZSB7"
    "eyBib3JkZXI6IDFweCBzb2xpZCB7Q19DUklNU09OX0RJTX07ICIKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19NT05JVE9S"
    "fTsgfX0iCiAgICAgICAgICAgIGYiUVRhYkJhcjo6dGFiIHt7IGJhY2tncm91bmQ6IHtDX0JHM307IGNvbG9yOiB7Q19URVhUX0RJ"
    "TX07ICIKICAgICAgICAgICAgZiJwYWRkaW5nOiA0cHggMTJweDsgYm9yZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsgIgogICAg"
    "ICAgICAgICBmImZvbnQtZmFtaWx5OiB7REVDS19GT05UfSwgc2VyaWY7IGZvbnQtc2l6ZTogMTBweDsgfX0iCiAgICAgICAgICAg"
    "IGYiUVRhYkJhcjo6dGFiOnNlbGVjdGVkIHt7IGJhY2tncm91bmQ6IHtDX0JHMn07IGNvbG9yOiB7Q19HT0xEfTsgIgogICAgICAg"
    "ICAgICBmImJvcmRlci1ib3R0b206IDJweCBzb2xpZCB7Q19DUklNU09OfTsgfX0iCiAgICAgICAgKQoKICAgICAgICAjIOKUgOKU"
    "gCBUYWIgMDogUGVyc29uYSBjaGF0IHRhYiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBzZWFuY2Vfd2lk"
    "Z2V0ID0gUVdpZGdldCgpCiAgICAgICAgc2VhbmNlX2xheW91dCA9IFFWQm94TGF5b3V0KHNlYW5jZV93aWRnZXQpCiAgICAgICAg"
    "c2VhbmNlX2xheW91dC5zZXRDb250ZW50c01hcmdpbnMoMCwgMCwgMCwgMCkKICAgICAgICBzZWFuY2VfbGF5b3V0LnNldFNwYWNp"
    "bmcoMCkKICAgICAgICBzZWxmLl9jaGF0X2Rpc3BsYXkgPSBRVGV4dEVkaXQoKQogICAgICAgIHNlbGYuX2NoYXRfZGlzcGxheS5z"
    "ZXRSZWFkT25seShUcnVlKQogICAgICAgIHNlbGYuX2NoYXRfZGlzcGxheS5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJh"
    "Y2tncm91bmQ6IHtDX01PTklUT1J9OyBjb2xvcjoge0NfR09MRH07ICIKICAgICAgICAgICAgZiJib3JkZXI6IG5vbmU7ICIKICAg"
    "ICAgICAgICAgZiJmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyBmb250LXNpemU6IDEycHg7IHBhZGRpbmc6IDhweDsi"
    "CiAgICAgICAgKQogICAgICAgIHNlYW5jZV9sYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX2NoYXRfZGlzcGxheSkKICAgICAgICBzZWxm"
    "Ll9tYWluX3RhYnMuYWRkVGFiKHNlYW5jZV93aWRnZXQsIGYi4p2nIHtVSV9DSEFUX1dJTkRPV30iKQoKICAgICAgICAjIOKUgOKU"
    "gCBUYWIgMTogU2VsZiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIAKICAgICAgICBzZWxmLl9zZWxmX3RhYl93aWRnZXQgPSBRV2lkZ2V0KCkKICAgICAgICBzZWxmX2xheW91dCA9IFFW"
    "Qm94TGF5b3V0KHNlbGYuX3NlbGZfdGFiX3dpZGdldCkKICAgICAgICBzZWxmX2xheW91dC5zZXRDb250ZW50c01hcmdpbnMoNCwg"
    "NCwgNCwgNCkKICAgICAgICBzZWxmX2xheW91dC5zZXRTcGFjaW5nKDQpCiAgICAgICAgc2VsZi5fc2VsZl9kaXNwbGF5ID0gUVRl"
    "eHRFZGl0KCkKICAgICAgICBzZWxmLl9zZWxmX2Rpc3BsYXkuc2V0UmVhZE9ubHkoVHJ1ZSkKICAgICAgICBzZWxmLl9zZWxmX2Rp"
    "c3BsYXkuc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19NT05JVE9SfTsgY29sb3I6IHtDX0dPTER9"
    "OyAiCiAgICAgICAgICAgIGYiYm9yZGVyOiBub25lOyAiCiAgICAgICAgICAgIGYiZm9udC1mYW1pbHk6IHtERUNLX0ZPTlR9LCBz"
    "ZXJpZjsgZm9udC1zaXplOiAxMnB4OyBwYWRkaW5nOiA4cHg7IgogICAgICAgICkKICAgICAgICBzZWxmX2xheW91dC5hZGRXaWRn"
    "ZXQoc2VsZi5fc2VsZl9kaXNwbGF5LCAxKQogICAgICAgIHNlbGYuX21haW5fdGFicy5hZGRUYWIoc2VsZi5fc2VsZl90YWJfd2lk"
    "Z2V0LCAi4peJIFNFTEYiKQoKICAgICAgICBsYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX21haW5fdGFicywgMSkKCiAgICAgICAgIyDi"
    "lIDilIAgQm90dG9tIHN0YXR1cy9yZXNvdXJjZSBibG9jayByb3cg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgIyBNYW5kYXRvcnkgcGVybWFuZW50"
    "IHN0cnVjdHVyZSBhY3Jvc3MgYWxsIHBlcnNvbmFzOgogICAgICAgICMgTUlSUk9SIHwgW0xPV0VSLU1JRERMRSBQRVJNQU5FTlQg"
    "Rk9PVFBSSU5UXQogICAgICAgIGJsb2NrX3JvdyA9IFFIQm94TGF5b3V0KCkKICAgICAgICBibG9ja19yb3cuc2V0U3BhY2luZygy"
    "KQoKICAgICAgICAjIE1pcnJvciAobmV2ZXIgY29sbGFwc2VzKQogICAgICAgIG1pcnJvcl93cmFwID0gUVdpZGdldCgpCiAgICAg"
    "ICAgbXdfbGF5b3V0ID0gUVZCb3hMYXlvdXQobWlycm9yX3dyYXApCiAgICAgICAgbXdfbGF5b3V0LnNldENvbnRlbnRzTWFyZ2lu"
    "cygwLCAwLCAwLCAwKQogICAgICAgIG13X2xheW91dC5zZXRTcGFjaW5nKDIpCiAgICAgICAgbXdfbGF5b3V0LmFkZFdpZGdldChf"
    "c2VjdGlvbl9sYmwoZiLinacge1VJX01JUlJPUl9MQUJFTH0iKSkKICAgICAgICBzZWxmLl9taXJyb3IgPSBNaXJyb3JXaWRnZXQo"
    "KQogICAgICAgIHNlbGYuX21pcnJvci5zZXRGaXhlZFNpemUoMTYwLCAxNjApCiAgICAgICAgbXdfbGF5b3V0LmFkZFdpZGdldChz"
    "ZWxmLl9taXJyb3IpCiAgICAgICAgYmxvY2tfcm93LmFkZFdpZGdldChtaXJyb3Jfd3JhcCwgMCkKCiAgICAgICAgIyBNaWRkbGUg"
    "bG93ZXIgYmxvY2sga2VlcHMgYSBwZXJtYW5lbnQgZm9vdHByaW50OgogICAgICAgICMgbGVmdCA9IGNvbXBhY3Qgc3RhY2sgYXJl"
    "YSwgcmlnaHQgPSBmaXhlZCBleHBhbmRlZC1yb3cgc2xvdHMuCiAgICAgICAgbWlkZGxlX3dyYXAgPSBRV2lkZ2V0KCkKICAgICAg"
    "ICBtaWRkbGVfbGF5b3V0ID0gUUhCb3hMYXlvdXQobWlkZGxlX3dyYXApCiAgICAgICAgbWlkZGxlX2xheW91dC5zZXRDb250ZW50"
    "c01hcmdpbnMoMCwgMCwgMCwgMCkKICAgICAgICBtaWRkbGVfbGF5b3V0LnNldFNwYWNpbmcoMikKCiAgICAgICAgc2VsZi5fbG93"
    "ZXJfc3RhY2tfd3JhcCA9IFFXaWRnZXQoKQogICAgICAgIHNlbGYuX2xvd2VyX3N0YWNrX3dyYXAuc2V0TWluaW11bVdpZHRoKDEz"
    "MCkKICAgICAgICBzZWxmLl9sb3dlcl9zdGFja193cmFwLnNldE1heGltdW1XaWR0aCgxMzApCiAgICAgICAgc2VsZi5fbG93ZXJf"
    "c3RhY2tfbGF5b3V0ID0gUVZCb3hMYXlvdXQoc2VsZi5fbG93ZXJfc3RhY2tfd3JhcCkKICAgICAgICBzZWxmLl9sb3dlcl9zdGFj"
    "a19sYXlvdXQuc2V0Q29udGVudHNNYXJnaW5zKDAsIDAsIDAsIDApCiAgICAgICAgc2VsZi5fbG93ZXJfc3RhY2tfbGF5b3V0LnNl"
    "dFNwYWNpbmcoMikKICAgICAgICBzZWxmLl9sb3dlcl9zdGFja193cmFwLnNldFZpc2libGUoRmFsc2UpCiAgICAgICAgbWlkZGxl"
    "X2xheW91dC5hZGRXaWRnZXQoc2VsZi5fbG93ZXJfc3RhY2tfd3JhcCwgMCkKCiAgICAgICAgc2VsZi5fbG93ZXJfZXhwYW5kZWRf"
    "cm93ID0gUVdpZGdldCgpCiAgICAgICAgc2VsZi5fbG93ZXJfZXhwYW5kZWRfcm93X2xheW91dCA9IFFHcmlkTGF5b3V0KHNlbGYu"
    "X2xvd2VyX2V4cGFuZGVkX3JvdykKICAgICAgICBzZWxmLl9sb3dlcl9leHBhbmRlZF9yb3dfbGF5b3V0LnNldENvbnRlbnRzTWFy"
    "Z2lucygwLCAwLCAwLCAwKQogICAgICAgIHNlbGYuX2xvd2VyX2V4cGFuZGVkX3Jvd19sYXlvdXQuc2V0SG9yaXpvbnRhbFNwYWNp"
    "bmcoMikKICAgICAgICBzZWxmLl9sb3dlcl9leHBhbmRlZF9yb3dfbGF5b3V0LnNldFZlcnRpY2FsU3BhY2luZygyKQogICAgICAg"
    "IG1pZGRsZV9sYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX2xvd2VyX2V4cGFuZGVkX3JvdywgMSkKCiAgICAgICAgIyBFbW90aW9uIGJs"
    "b2NrIChjb2xsYXBzaWJsZSkKICAgICAgICBzZWxmLl9lbW90aW9uX2Jsb2NrID0gRW1vdGlvbkJsb2NrKCkKICAgICAgICBzZWxm"
    "Ll9lbW90aW9uX2Jsb2NrX3dyYXAgPSBDb2xsYXBzaWJsZUJsb2NrKAogICAgICAgICAgICBmIuKdpyB7VUlfRU1PVElPTlNfTEFC"
    "RUx9Iiwgc2VsZi5fZW1vdGlvbl9ibG9jaywKICAgICAgICAgICAgZXhwYW5kZWQ9VHJ1ZSwgbWluX3dpZHRoPTEzMCwgcmVzZXJ2"
    "ZV93aWR0aD1UcnVlCiAgICAgICAgKQoKICAgICAgICAjIExlZnQgcmVzb3VyY2Ugb3JiIChjb2xsYXBzaWJsZSkKICAgICAgICBz"
    "ZWxmLl9sZWZ0X29yYiA9IFNwaGVyZVdpZGdldCgKICAgICAgICAgICAgVUlfTEVGVF9PUkJfTEFCRUwsIENfQ1JJTVNPTiwgQ19D"
    "UklNU09OX0RJTQogICAgICAgICkKICAgICAgICBzZWxmLl9sZWZ0X29yYl93cmFwID0gQ29sbGFwc2libGVCbG9jaygKICAgICAg"
    "ICAgICAgZiLinacge1VJX0xFRlRfT1JCX1RJVExFfSIsIHNlbGYuX2xlZnRfb3JiLAogICAgICAgICAgICBtaW5fd2lkdGg9OTAs"
    "IHJlc2VydmVfd2lkdGg9VHJ1ZQogICAgICAgICkKCiAgICAgICAgIyBDZW50ZXIgY3ljbGUgd2lkZ2V0IChjb2xsYXBzaWJsZSkK"
    "ICAgICAgICBzZWxmLl9jeWNsZV93aWRnZXQgPSBDeWNsZVdpZGdldCgpCiAgICAgICAgc2VsZi5fY3ljbGVfd3JhcCA9IENvbGxh"
    "cHNpYmxlQmxvY2soCiAgICAgICAgICAgIGYi4p2nIHtVSV9DWUNMRV9USVRMRX0iLCBzZWxmLl9jeWNsZV93aWRnZXQsCiAgICAg"
    "ICAgICAgIG1pbl93aWR0aD05MCwgcmVzZXJ2ZV93aWR0aD1UcnVlCiAgICAgICAgKQoKICAgICAgICAjIFJpZ2h0IHJlc291cmNl"
    "IG9yYiAoY29sbGFwc2libGUpCiAgICAgICAgc2VsZi5fcmlnaHRfb3JiID0gU3BoZXJlV2lkZ2V0KAogICAgICAgICAgICBVSV9S"
    "SUdIVF9PUkJfTEFCRUwsIENfUFVSUExFLCBDX1BVUlBMRV9ESU0KICAgICAgICApCiAgICAgICAgc2VsZi5fcmlnaHRfb3JiX3dy"
    "YXAgPSBDb2xsYXBzaWJsZUJsb2NrKAogICAgICAgICAgICBmIuKdpyB7VUlfUklHSFRfT1JCX1RJVExFfSIsIHNlbGYuX3JpZ2h0"
    "X29yYiwKICAgICAgICAgICAgbWluX3dpZHRoPTkwLCByZXNlcnZlX3dpZHRoPVRydWUKICAgICAgICApCgogICAgICAgICMgRXNz"
    "ZW5jZSAoMiBnYXVnZXMsIGNvbGxhcHNpYmxlKQogICAgICAgIGVzc2VuY2Vfd2lkZ2V0ID0gUVdpZGdldCgpCiAgICAgICAgZXNz"
    "ZW5jZV9sYXlvdXQgPSBRVkJveExheW91dChlc3NlbmNlX3dpZGdldCkKICAgICAgICBlc3NlbmNlX2xheW91dC5zZXRDb250ZW50"
    "c01hcmdpbnMoNCwgNCwgNCwgNCkKICAgICAgICBlc3NlbmNlX2xheW91dC5zZXRTcGFjaW5nKDQpCiAgICAgICAgc2VsZi5fZXNz"
    "ZW5jZV9wcmltYXJ5X2dhdWdlICAgPSBHYXVnZVdpZGdldChVSV9FU1NFTkNFX1BSSU1BUlksICAgIiUiLCAxMDAuMCwgQ19DUklN"
    "U09OKQogICAgICAgIHNlbGYuX2Vzc2VuY2Vfc2Vjb25kYXJ5X2dhdWdlID0gR2F1Z2VXaWRnZXQoVUlfRVNTRU5DRV9TRUNPTkRB"
    "UlksICIlIiwgMTAwLjAsIENfR1JFRU4pCiAgICAgICAgZXNzZW5jZV9sYXlvdXQuYWRkV2lkZ2V0KHNlbGYuX2Vzc2VuY2VfcHJp"
    "bWFyeV9nYXVnZSkKICAgICAgICBlc3NlbmNlX2xheW91dC5hZGRXaWRnZXQoc2VsZi5fZXNzZW5jZV9zZWNvbmRhcnlfZ2F1Z2Up"
    "CiAgICAgICAgc2VsZi5fZXNzZW5jZV93cmFwID0gQ29sbGFwc2libGVCbG9jaygKICAgICAgICAgICAgZiLinacge1VJX0VTU0VO"
    "Q0VfVElUTEV9IiwgZXNzZW5jZV93aWRnZXQsCiAgICAgICAgICAgIG1pbl93aWR0aD0xMTAsIHJlc2VydmVfd2lkdGg9VHJ1ZQog"
    "ICAgICAgICkKCiAgICAgICAgIyBFeHBhbmRlZCByb3cgc2xvdHMgbXVzdCBzdGF5IGluIGNhbm9uaWNhbCB2aXN1YWwgb3JkZXIu"
    "CiAgICAgICAgc2VsZi5fbG93ZXJfZXhwYW5kZWRfc2xvdF9vcmRlciA9IFsKICAgICAgICAgICAgImVtb3Rpb25zIiwgInByaW1h"
    "cnkiLCAiY3ljbGUiLCAic2Vjb25kYXJ5IiwgImVzc2VuY2UiCiAgICAgICAgXQogICAgICAgIHNlbGYuX2xvd2VyX2NvbXBhY3Rf"
    "c3RhY2tfb3JkZXIgPSBbCiAgICAgICAgICAgICJjeWNsZSIsICJwcmltYXJ5IiwgInNlY29uZGFyeSIsICJlc3NlbmNlIiwgImVt"
    "b3Rpb25zIgogICAgICAgIF0KICAgICAgICBzZWxmLl9sb3dlcl9tb2R1bGVfd3JhcHMgPSB7CiAgICAgICAgICAgICJlbW90aW9u"
    "cyI6IHNlbGYuX2Vtb3Rpb25fYmxvY2tfd3JhcCwKICAgICAgICAgICAgInByaW1hcnkiOiBzZWxmLl9sZWZ0X29yYl93cmFwLAog"
    "ICAgICAgICAgICAiY3ljbGUiOiBzZWxmLl9jeWNsZV93cmFwLAogICAgICAgICAgICAic2Vjb25kYXJ5Ijogc2VsZi5fcmlnaHRf"
    "b3JiX3dyYXAsCiAgICAgICAgICAgICJlc3NlbmNlIjogc2VsZi5fZXNzZW5jZV93cmFwLAogICAgICAgIH0KCiAgICAgICAgc2Vs"
    "Zi5fbG93ZXJfcm93X3Nsb3RzID0ge30KICAgICAgICBmb3IgY29sLCBrZXkgaW4gZW51bWVyYXRlKHNlbGYuX2xvd2VyX2V4cGFu"
    "ZGVkX3Nsb3Rfb3JkZXIpOgogICAgICAgICAgICBzbG90ID0gUVdpZGdldCgpCiAgICAgICAgICAgIHNsb3RfbGF5b3V0ID0gUVZC"
    "b3hMYXlvdXQoc2xvdCkKICAgICAgICAgICAgc2xvdF9sYXlvdXQuc2V0Q29udGVudHNNYXJnaW5zKDAsIDAsIDAsIDApCiAgICAg"
    "ICAgICAgIHNsb3RfbGF5b3V0LnNldFNwYWNpbmcoMCkKICAgICAgICAgICAgc2VsZi5fbG93ZXJfZXhwYW5kZWRfcm93X2xheW91"
    "dC5hZGRXaWRnZXQoc2xvdCwgMCwgY29sKQogICAgICAgICAgICBzZWxmLl9sb3dlcl9leHBhbmRlZF9yb3dfbGF5b3V0LnNldENv"
    "bHVtblN0cmV0Y2goY29sLCAxKQogICAgICAgICAgICBzZWxmLl9sb3dlcl9yb3dfc2xvdHNba2V5XSA9IHNsb3RfbGF5b3V0Cgog"
    "ICAgICAgIGZvciB3cmFwIGluIHNlbGYuX2xvd2VyX21vZHVsZV93cmFwcy52YWx1ZXMoKToKICAgICAgICAgICAgd3JhcC50b2dn"
    "bGVkLmNvbm5lY3Qoc2VsZi5fcmVmcmVzaF9sb3dlcl9taWRkbGVfbGF5b3V0KQoKICAgICAgICBzZWxmLl9yZWZyZXNoX2xvd2Vy"
    "X21pZGRsZV9sYXlvdXQoKQoKICAgICAgICBibG9ja19yb3cuYWRkV2lkZ2V0KG1pZGRsZV93cmFwLCAxKQogICAgICAgIGxheW91"
    "dC5hZGRMYXlvdXQoYmxvY2tfcm93KQoKICAgICAgICAjIEZvb3RlciBzdGF0ZSBzdHJpcCAoYmVsb3cgYmxvY2sgcm93IOKAlCBw"
    "ZXJtYW5lbnQgVUkgc3RydWN0dXJlKQogICAgICAgIHNlbGYuX2Zvb3Rlcl9zdHJpcCA9IEZvb3RlclN0cmlwV2lkZ2V0KCkKICAg"
    "ICAgICBzZWxmLl9mb290ZXJfc3RyaXAuc2V0X2xhYmVsKFVJX0ZPT1RFUl9TVFJJUF9MQUJFTCkKICAgICAgICBsYXlvdXQuYWRk"
    "V2lkZ2V0KHNlbGYuX2Zvb3Rlcl9zdHJpcCkKCiAgICAgICAgIyDilIDilIAgSW5wdXQgcm93IOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIGlucHV0X3JvdyA9"
    "IFFIQm94TGF5b3V0KCkKICAgICAgICBwcm9tcHRfc3ltID0gUUxhYmVsKCLinKYiKQogICAgICAgIHByb21wdF9zeW0uc2V0U3R5"
    "bGVTaGVldCgKICAgICAgICAgICAgZiJjb2xvcjoge0NfQ1JJTVNPTn07IGZvbnQtc2l6ZTogMTZweDsgZm9udC13ZWlnaHQ6IGJv"
    "bGQ7IGJvcmRlcjogbm9uZTsiCiAgICAgICAgKQogICAgICAgIHByb21wdF9zeW0uc2V0Rml4ZWRXaWR0aCgyMCkKCiAgICAgICAg"
    "c2VsZi5faW5wdXRfZmllbGQgPSBRTGluZUVkaXQoKQogICAgICAgIHNlbGYuX2lucHV0X2ZpZWxkLnNldFBsYWNlaG9sZGVyVGV4"
    "dChVSV9JTlBVVF9QTEFDRUhPTERFUikKICAgICAgICBzZWxmLl9pbnB1dF9maWVsZC5yZXR1cm5QcmVzc2VkLmNvbm5lY3Qoc2Vs"
    "Zi5fc2VuZF9tZXNzYWdlKQogICAgICAgIHNlbGYuX2lucHV0X2ZpZWxkLnNldEVuYWJsZWQoRmFsc2UpCgogICAgICAgIHNlbGYu"
    "X3NlbmRfYnRuID0gUVB1c2hCdXR0b24oVUlfU0VORF9CVVRUT04pCiAgICAgICAgc2VsZi5fc2VuZF9idG4uc2V0Rml4ZWRXaWR0"
    "aCgxMTApCiAgICAgICAgc2VsZi5fc2VuZF9idG4uY2xpY2tlZC5jb25uZWN0KHNlbGYuX3NlbmRfbWVzc2FnZSkKICAgICAgICBz"
    "ZWxmLl9zZW5kX2J0bi5zZXRFbmFibGVkKEZhbHNlKQoKICAgICAgICBpbnB1dF9yb3cuYWRkV2lkZ2V0KHByb21wdF9zeW0pCiAg"
    "ICAgICAgaW5wdXRfcm93LmFkZFdpZGdldChzZWxmLl9pbnB1dF9maWVsZCkKICAgICAgICBpbnB1dF9yb3cuYWRkV2lkZ2V0KHNl"
    "bGYuX3NlbmRfYnRuKQogICAgICAgIGxheW91dC5hZGRMYXlvdXQoaW5wdXRfcm93KQoKICAgICAgICByZXR1cm4gbGF5b3V0Cgog"
    "ICAgZGVmIF9jbGVhcl9sYXlvdXRfd2lkZ2V0cyhzZWxmLCBsYXlvdXQ6IFFWQm94TGF5b3V0KSAtPiBOb25lOgogICAgICAgIHdo"
    "aWxlIGxheW91dC5jb3VudCgpOgogICAgICAgICAgICBpdGVtID0gbGF5b3V0LnRha2VBdCgwKQogICAgICAgICAgICB3aWRnZXQg"
    "PSBpdGVtLndpZGdldCgpCiAgICAgICAgICAgIGlmIHdpZGdldCBpcyBub3QgTm9uZToKICAgICAgICAgICAgICAgIHdpZGdldC5z"
    "ZXRQYXJlbnQoTm9uZSkKCiAgICBkZWYgX3JlZnJlc2hfbG93ZXJfbWlkZGxlX2xheW91dChzZWxmLCAqX2FyZ3MpIC0+IE5vbmU6"
    "CiAgICAgICAgY29sbGFwc2VkX2NvdW50ID0gMAoKICAgICAgICAjIFJlYnVpbGQgZXhwYW5kZWQgcm93IHNsb3RzIGluIGZpeGVk"
    "IGV4cGFuZGVkIG9yZGVyLgogICAgICAgIGZvciBrZXkgaW4gc2VsZi5fbG93ZXJfZXhwYW5kZWRfc2xvdF9vcmRlcjoKICAgICAg"
    "ICAgICAgc2xvdF9sYXlvdXQgPSBzZWxmLl9sb3dlcl9yb3dfc2xvdHNba2V5XQogICAgICAgICAgICBzZWxmLl9jbGVhcl9sYXlv"
    "dXRfd2lkZ2V0cyhzbG90X2xheW91dCkKICAgICAgICAgICAgd3JhcCA9IHNlbGYuX2xvd2VyX21vZHVsZV93cmFwc1trZXldCiAg"
    "ICAgICAgICAgIGlmIHdyYXAuaXNfZXhwYW5kZWQoKToKICAgICAgICAgICAgICAgIHNsb3RfbGF5b3V0LmFkZFdpZGdldCh3cmFw"
    "KQogICAgICAgICAgICBlbHNlOgogICAgICAgICAgICAgICAgY29sbGFwc2VkX2NvdW50ICs9IDEKICAgICAgICAgICAgICAgIHNs"
    "b3RfbGF5b3V0LmFkZFN0cmV0Y2goMSkKCiAgICAgICAgIyBSZWJ1aWxkIGNvbXBhY3Qgc3RhY2sgaW4gY2Fub25pY2FsIGNvbXBh"
    "Y3Qgb3JkZXIuCiAgICAgICAgc2VsZi5fY2xlYXJfbGF5b3V0X3dpZGdldHMoc2VsZi5fbG93ZXJfc3RhY2tfbGF5b3V0KQogICAg"
    "ICAgIGZvciBrZXkgaW4gc2VsZi5fbG93ZXJfY29tcGFjdF9zdGFja19vcmRlcjoKICAgICAgICAgICAgd3JhcCA9IHNlbGYuX2xv"
    "d2VyX21vZHVsZV93cmFwc1trZXldCiAgICAgICAgICAgIGlmIG5vdCB3cmFwLmlzX2V4cGFuZGVkKCk6CiAgICAgICAgICAgICAg"
    "ICBzZWxmLl9sb3dlcl9zdGFja19sYXlvdXQuYWRkV2lkZ2V0KHdyYXApCgogICAgICAgIHNlbGYuX2xvd2VyX3N0YWNrX2xheW91"
    "dC5hZGRTdHJldGNoKDEpCiAgICAgICAgc2VsZi5fbG93ZXJfc3RhY2tfd3JhcC5zZXRWaXNpYmxlKGNvbGxhcHNlZF9jb3VudCA+"
    "IDApCgogICAgZGVmIF9idWlsZF9zcGVsbGJvb2tfcGFuZWwoc2VsZikgLT4gUVZCb3hMYXlvdXQ6CiAgICAgICAgbGF5b3V0ID0g"
    "UVZCb3hMYXlvdXQoKQogICAgICAgIGxheW91dC5zZXRDb250ZW50c01hcmdpbnMoMCwgMCwgMCwgMCkKICAgICAgICBsYXlvdXQu"
    "c2V0U3BhY2luZyg0KQogICAgICAgIGxheW91dC5hZGRXaWRnZXQoX3NlY3Rpb25fbGJsKCLinacgU1lTVEVNUyIpKQoKICAgICAg"
    "ICAjIFRhYiB3aWRnZXQKICAgICAgICBzZWxmLl9zcGVsbF90YWJzID0gUVRhYldpZGdldCgpCiAgICAgICAgc2VsZi5fc3BlbGxf"
    "dGFicy5zZXRNaW5pbXVtV2lkdGgoMjgwKQogICAgICAgIHNlbGYuX3NwZWxsX3RhYnMuc2V0U2l6ZVBvbGljeSgKICAgICAgICAg"
    "ICAgUVNpemVQb2xpY3kuUG9saWN5LkV4cGFuZGluZywKICAgICAgICAgICAgUVNpemVQb2xpY3kuUG9saWN5LkV4cGFuZGluZwog"
    "ICAgICAgICkKICAgICAgICBzZWxmLl9zcGVsbF90YWJfYmFyID0gTG9ja0F3YXJlVGFiQmFyKHNlbGYuX2lzX3NwZWxsX3RhYl9s"
    "b2NrZWQsIHNlbGYuX3NwZWxsX3RhYnMpCiAgICAgICAgc2VsZi5fc3BlbGxfdGFicy5zZXRUYWJCYXIoc2VsZi5fc3BlbGxfdGFi"
    "X2JhcikKICAgICAgICBzZWxmLl9zcGVsbF90YWJfYmFyLnNldE1vdmFibGUoVHJ1ZSkKICAgICAgICBzZWxmLl9zcGVsbF90YWJf"
    "YmFyLnNldENvbnRleHRNZW51UG9saWN5KFF0LkNvbnRleHRNZW51UG9saWN5LkN1c3RvbUNvbnRleHRNZW51KQogICAgICAgIHNl"
    "bGYuX3NwZWxsX3RhYl9iYXIuY3VzdG9tQ29udGV4dE1lbnVSZXF1ZXN0ZWQuY29ubmVjdChzZWxmLl9zaG93X3NwZWxsX3RhYl9j"
    "b250ZXh0X21lbnUpCiAgICAgICAgc2VsZi5fc3BlbGxfdGFiX2Jhci50YWJNb3ZlZC5jb25uZWN0KHNlbGYuX29uX3NwZWxsX3Rh"
    "Yl9kcmFnX21vdmVkKQogICAgICAgIHNlbGYuX3NwZWxsX3RhYnMuY3VycmVudENoYW5nZWQuY29ubmVjdChsYW1iZGEgX2lkeDog"
    "c2VsZi5fZXhpdF9zcGVsbF90YWJfbW92ZV9tb2RlKCkpCiAgICAgICAgaWYgbm90IHNlbGYuX2ZvY3VzX2hvb2tlZF9mb3Jfc3Bl"
    "bGxfdGFiczoKICAgICAgICAgICAgYXBwID0gUUFwcGxpY2F0aW9uLmluc3RhbmNlKCkKICAgICAgICAgICAgaWYgYXBwIGlzIG5v"
    "dCBOb25lOgogICAgICAgICAgICAgICAgYXBwLmZvY3VzQ2hhbmdlZC5jb25uZWN0KHNlbGYuX29uX2dsb2JhbF9mb2N1c19jaGFu"
    "Z2VkKQogICAgICAgICAgICAgICAgc2VsZi5fZm9jdXNfaG9va2VkX2Zvcl9zcGVsbF90YWJzID0gVHJ1ZQoKICAgICAgICAjIEJ1"
    "aWxkIERpYWdub3N0aWNzVGFiIGVhcmx5IHNvIHN0YXJ0dXAgbG9ncyBhcmUgc2FmZSBldmVuIGJlZm9yZQogICAgICAgICMgdGhl"
    "IERpYWdub3N0aWNzIHRhYiBpcyBhdHRhY2hlZCB0byB0aGUgd2lkZ2V0LgogICAgICAgIHNlbGYuX2RpYWdfdGFiID0gRGlhZ25v"
    "c3RpY3NUYWIoKQoKICAgICAgICAjIOKUgOKUgCBJbnN0cnVtZW50cyB0YWIg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgc2VsZi5faHdfcGFuZWwgPSBIYXJkd2FyZVBhbmVsKCkKCiAgICAg"
    "ICAgIyDilIDilIAgU0wgU2NhbnMgdGFiIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgAogICAgICAgIHNlbGYuX3NsX3NjYW5zID0gU0xTY2Fuc1RhYihjZmdfcGF0aCgic2wiKSkKCiAgICAg"
    "ICAgIyDilIDilIAgU0wgQ29tbWFuZHMgdGFiIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgAogICAgICAgIHNlbGYuX3NsX2NvbW1hbmRzID0gU0xDb21tYW5kc1RhYigpCgogICAgICAgICMg4pSA4pSAIEpv"
    "YiBUcmFja2VyIHRhYiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAg"
    "ICAgICBzZWxmLl9qb2JfdHJhY2tlciA9IEpvYlRyYWNrZXJUYWIoKQoKICAgICAgICAjIOKUgOKUgCBMZXNzb25zIHRhYiDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAgICBz"
    "ZWxmLl9sZXNzb25zX3RhYiA9IExlc3NvbnNUYWIoc2VsZi5fbGVzc29ucykKCiAgICAgICAgIyBTZWxmIHRhYiBpcyBub3cgaW4g"
    "dGhlIG1haW4gYXJlYSBhbG9uZ3NpZGUgdGhlIHBlcnNvbmEgY2hhdCB0YWIKICAgICAgICAjIEtlZXAgYSBTZWxmVGFiIGluc3Rh"
    "bmNlIGZvciBpZGxlIGNvbnRlbnQgZ2VuZXJhdGlvbgogICAgICAgIHNlbGYuX3NlbGZfdGFiID0gU2VsZlRhYigpCgogICAgICAg"
    "ICMg4pSA4pSAIE1vZHVsZSBUcmFja2VyIHRhYiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIAKICAgICAgICBzZWxmLl9tb2R1bGVfdHJhY2tlciA9IE1vZHVsZVRyYWNrZXJUYWIoKQoKICAgICAgICAjIOKUgOKUgCBEaWNl"
    "IFJvbGxlciB0YWIg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAg"
    "ICAgc2VsZi5fZGljZV9yb2xsZXJfdGFiID0gRGljZVJvbGxlclRhYihkaWFnbm9zdGljc19sb2dnZXI9c2VsZi5fZGlhZ190YWIu"
    "bG9nKQoKICAgICAgICAjIOKUgOKUgCBNYWdpYyA4LUJhbGwgdGFiIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIHNlbGYuX21hZ2ljXzhiYWxsX3RhYiA9IE1hZ2ljOEJhbGxUYWIoCiAgICAgICAg"
    "ICAgIG9uX3Rocm93PXNlbGYuX2hhbmRsZV9tYWdpY184YmFsbF90aHJvdywKICAgICAgICAgICAgZGlhZ25vc3RpY3NfbG9nZ2Vy"
    "PXNlbGYuX2RpYWdfdGFiLmxvZywKICAgICAgICApCgogICAgICAgICMg4pSA4pSAIFNldHRpbmdzIHRhYiAoZGVjay13aWRlIHJ1"
    "bnRpbWUgY29udHJvbHMpIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAg"
    "IHNlbGYuX3NldHRpbmdzX3RhYiA9IFNldHRpbmdzVGFiKHNlbGYpCgogICAgICAgICMgRGVzY3JpcHRvci1iYXNlZCBvcmRlcmlu"
    "ZyAoc3RhYmxlIGlkZW50aXR5ICsgdmlzdWFsIG9yZGVyIG9ubHkpCiAgICAgICAgc2VsZi5fc3BlbGxfdGFiX2RlZnMgPSBbCiAg"
    "ICAgICAgICAgIHsiaWQiOiAiaW5zdHJ1bWVudHMiLCAidGl0bGUiOiAiSW5zdHJ1bWVudHMiLCAid2lkZ2V0Ijogc2VsZi5faHdf"
    "cGFuZWwsICJkZWZhdWx0X29yZGVyIjogMCwgImNhdGVnb3J5IjogIlNZU1RFTSIsICJzZWNvbmRhcnlfY2F0ZWdvcmllcyI6IFtd"
    "LCAicHJvdGVjdGVkX2NhdGVnb3J5IjogVHJ1ZX0sCiAgICAgICAgICAgIHsiaWQiOiAic2xfc2NhbnMiLCAidGl0bGUiOiAiU0wg"
    "U2NhbnMiLCAid2lkZ2V0Ijogc2VsZi5fc2xfc2NhbnMsICJkZWZhdWx0X29yZGVyIjogMSwgImNhdGVnb3J5IjogIk9wZXJhdGlv"
    "bnMiLCAic2Vjb25kYXJ5X2NhdGVnb3JpZXMiOiBbXX0sCiAgICAgICAgICAgIHsiaWQiOiAic2xfY29tbWFuZHMiLCAidGl0bGUi"
    "OiAiU0wgQ29tbWFuZHMiLCAid2lkZ2V0Ijogc2VsZi5fc2xfY29tbWFuZHMsICJkZWZhdWx0X29yZGVyIjogMiwgImNhdGVnb3J5"
    "IjogIk9wZXJhdGlvbnMiLCAic2Vjb25kYXJ5X2NhdGVnb3JpZXMiOiBbXX0sCiAgICAgICAgICAgIHsiaWQiOiAiam9iX3RyYWNr"
    "ZXIiLCAidGl0bGUiOiAiSm9iIFRyYWNrZXIiLCAid2lkZ2V0Ijogc2VsZi5fam9iX3RyYWNrZXIsICJkZWZhdWx0X29yZGVyIjog"
    "MywgImNhdGVnb3J5IjogIk9wZXJhdGlvbnMiLCAic2Vjb25kYXJ5X2NhdGVnb3JpZXMiOiBbXX0sCiAgICAgICAgICAgIHsiaWQi"
    "OiAibGVzc29ucyIsICJ0aXRsZSI6ICJMZXNzb25zIiwgIndpZGdldCI6IHNlbGYuX2xlc3NvbnNfdGFiLCAiZGVmYXVsdF9vcmRl"
    "ciI6IDQsICJjYXRlZ29yeSI6ICJDb3JlIiwgInNlY29uZGFyeV9jYXRlZ29yaWVzIjogWyJNYW5hZ2VtZW50Il19LAogICAgICAg"
    "ICAgICB7ImlkIjogIm1vZHVsZXMiLCAidGl0bGUiOiAiTW9kdWxlcyIsICJ3aWRnZXQiOiBzZWxmLl9tb2R1bGVfdHJhY2tlciwg"
    "ImRlZmF1bHRfb3JkZXIiOiA1LCAiY2F0ZWdvcnkiOiAiTWFuYWdlbWVudCIsICJzZWNvbmRhcnlfY2F0ZWdvcmllcyI6IFsiVXRp"
    "bGl0aWVzIl19LAogICAgICAgICAgICB7ImlkIjogImRpY2Vfcm9sbGVyIiwgInRpdGxlIjogIkRpY2UgUm9sbGVyIiwgIndpZGdl"
    "dCI6IHNlbGYuX2RpY2Vfcm9sbGVyX3RhYiwgImRlZmF1bHRfb3JkZXIiOiA2LCAiY2F0ZWdvcnkiOiAiVXRpbGl0aWVzIiwgInNl"
    "Y29uZGFyeV9jYXRlZ29yaWVzIjogW119LAogICAgICAgICAgICB7ImlkIjogIm1hZ2ljXzhfYmFsbCIsICJ0aXRsZSI6ICJNYWdp"
    "YyA4LUJhbGwiLCAid2lkZ2V0Ijogc2VsZi5fbWFnaWNfOGJhbGxfdGFiLCAiZGVmYXVsdF9vcmRlciI6IDcsICJjYXRlZ29yeSI6"
    "ICJVdGlsaXRpZXMiLCAic2Vjb25kYXJ5X2NhdGVnb3JpZXMiOiBbXX0sCiAgICAgICAgICAgIHsiaWQiOiAiZGlhZ25vc3RpY3Mi"
    "LCAidGl0bGUiOiAiRGlhZ25vc3RpY3MiLCAid2lkZ2V0Ijogc2VsZi5fZGlhZ190YWIsICJkZWZhdWx0X29yZGVyIjogOCwgImNh"
    "dGVnb3J5IjogIlNZU1RFTSIsICJzZWNvbmRhcnlfY2F0ZWdvcmllcyI6IFtdLCAicHJvdGVjdGVkX2NhdGVnb3J5IjogVHJ1ZX0s"
    "CiAgICAgICAgICAgIHsiaWQiOiAic2V0dGluZ3MiLCAidGl0bGUiOiAiU2V0dGluZ3MiLCAid2lkZ2V0Ijogc2VsZi5fc2V0dGlu"
    "Z3NfdGFiLCAiZGVmYXVsdF9vcmRlciI6IDksICJjYXRlZ29yeSI6ICJTWVNURU0iLCAic2Vjb25kYXJ5X2NhdGVnb3JpZXMiOiBb"
    "XSwgInByb3RlY3RlZF9jYXRlZ29yeSI6IFRydWV9LAogICAgICAgIF0KICAgICAgICBzZWxmLl9pbml0X3NwZWxsX2NhdGVnb3J5"
    "X2ZyYW1ld29yaygpCiAgICAgICAgc2VsZi5fbG9hZF9zcGVsbF90YWJfc3RhdGVfZnJvbV9jb25maWcoKQogICAgICAgIHNlbGYu"
    "X3JlYnVpbGRfc3BlbGxfdGFicygpCgogICAgICAgIHJpZ2h0X3dvcmtzcGFjZSA9IFFXaWRnZXQoKQogICAgICAgIHJpZ2h0X3dv"
    "cmtzcGFjZV9sYXlvdXQgPSBRVkJveExheW91dChyaWdodF93b3Jrc3BhY2UpCiAgICAgICAgcmlnaHRfd29ya3NwYWNlX2xheW91"
    "dC5zZXRDb250ZW50c01hcmdpbnMoMCwgMCwgMCwgMCkKICAgICAgICByaWdodF93b3Jrc3BhY2VfbGF5b3V0LnNldFNwYWNpbmco"
    "NCkKCiAgICAgICAgc2VsZi5fY2F0ZWdvcnlfc3RyaXAgPSBRV2lkZ2V0KCkKICAgICAgICBzZWxmLl9jYXRlZ29yeV9zdHJpcF9s"
    "YXlvdXQgPSBRSEJveExheW91dChzZWxmLl9jYXRlZ29yeV9zdHJpcCkKICAgICAgICBzZWxmLl9jYXRlZ29yeV9zdHJpcF9sYXlv"
    "dXQuc2V0Q29udGVudHNNYXJnaW5zKDAsIDAsIDAsIDApCiAgICAgICAgc2VsZi5fY2F0ZWdvcnlfc3RyaXBfbGF5b3V0LnNldFNw"
    "YWNpbmcoNCkKICAgICAgICByaWdodF93b3Jrc3BhY2VfbGF5b3V0LmFkZFdpZGdldChzZWxmLl9jYXRlZ29yeV9zdHJpcCwgMCkK"
    "CiAgICAgICAgcmlnaHRfd29ya3NwYWNlX2xheW91dC5hZGRXaWRnZXQoc2VsZi5fc3BlbGxfdGFicywgc3RyZXRjaD0xKQoKICAg"
    "ICAgICBjYWxlbmRhcl9sYWJlbCA9IFFMYWJlbCgi4p2nIENBTEVOREFSIikKICAgICAgICBjYWxlbmRhcl9sYWJlbC5zZXRTdHls"
    "ZVNoZWV0KAogICAgICAgICAgICBmImNvbG9yOiB7Q19HT0xEfTsgZm9udC1zaXplOiAxMHB4OyBsZXR0ZXItc3BhY2luZzogMnB4"
    "OyBmb250LWZhbWlseToge0RFQ0tfRk9OVH0sIHNlcmlmOyIKICAgICAgICApCiAgICAgICAgcmlnaHRfd29ya3NwYWNlX2xheW91"
    "dC5hZGRXaWRnZXQoY2FsZW5kYXJfbGFiZWwpCgogICAgICAgIHNlbGYuY2FsZW5kYXJfd2lkZ2V0ID0gTWluaUNhbGVuZGFyV2lk"
    "Z2V0KCkKICAgICAgICBzZWxmLmNhbGVuZGFyX3dpZGdldC5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tncm91bmQ6"
    "IHtDX0JHMn07IGJvcmRlcjogMXB4IHNvbGlkIHtDX0NSSU1TT05fRElNfTsiCiAgICAgICAgKQogICAgICAgIHNlbGYuY2FsZW5k"
    "YXJfd2lkZ2V0LnNldFNpemVQb2xpY3koCiAgICAgICAgICAgIFFTaXplUG9saWN5LlBvbGljeS5FeHBhbmRpbmcsCiAgICAgICAg"
    "ICAgIFFTaXplUG9saWN5LlBvbGljeS5NYXhpbXVtCiAgICAgICAgKQogICAgICAgIHNlbGYuY2FsZW5kYXJfd2lkZ2V0LnNldE1h"
    "eGltdW1IZWlnaHQoMjYwKQogICAgICAgIHNlbGYuY2FsZW5kYXJfd2lkZ2V0LmNhbGVuZGFyLmNsaWNrZWQuY29ubmVjdChzZWxm"
    "Ll9pbnNlcnRfY2FsZW5kYXJfZGF0ZSkKICAgICAgICByaWdodF93b3Jrc3BhY2VfbGF5b3V0LmFkZFdpZGdldChzZWxmLmNhbGVu"
    "ZGFyX3dpZGdldCwgMCkKICAgICAgICByaWdodF93b3Jrc3BhY2VfbGF5b3V0LmFkZFN0cmV0Y2goMCkKCiAgICAgICAgbGF5b3V0"
    "LmFkZFdpZGdldChyaWdodF93b3Jrc3BhY2UsIDEpCiAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKAogICAgICAgICAgICAiW0xB"
    "WU9VVF0gcmlnaHQtc2lkZSBjYWxlbmRhciByZXN0b3JlZCAocGVyc2lzdGVudCBsb3dlci1yaWdodCBzZWN0aW9uKS4iLAogICAg"
    "ICAgICAgICAiSU5GTyIKICAgICAgICApCiAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKAogICAgICAgICAgICAiW0xBWU9VVF0g"
    "cGVyc2lzdGVudCBtaW5pIGNhbGVuZGFyIHJlc3RvcmVkL2NvbmZpcm1lZCAoYWx3YXlzIHZpc2libGUgbG93ZXItcmlnaHQpLiIs"
    "CiAgICAgICAgICAgICJJTkZPIgogICAgICAgICkKICAgICAgICByZXR1cm4gbGF5b3V0CgogICAgZGVmIF9yZXN0b3JlX21haW5f"
    "c3BsaXR0ZXJfc3RhdGUoc2VsZikgLT4gTm9uZToKICAgICAgICBzcGxpdHRlcl9jZmcgPSBDRkcuZ2V0KCJtYWluX3NwbGl0dGVy"
    "Iiwge30pIGlmIGlzaW5zdGFuY2UoQ0ZHLCBkaWN0KSBlbHNlIHt9CiAgICAgICAgc2F2ZWRfc2l6ZXMgPSBzcGxpdHRlcl9jZmcu"
    "Z2V0KCJob3Jpem9udGFsX3NpemVzIikgaWYgaXNpbnN0YW5jZShzcGxpdHRlcl9jZmcsIGRpY3QpIGVsc2UgTm9uZQoKICAgICAg"
    "ICBpZiBpc2luc3RhbmNlKHNhdmVkX3NpemVzLCBsaXN0KSBhbmQgbGVuKHNhdmVkX3NpemVzKSA9PSAyOgogICAgICAgICAgICB0"
    "cnk6CiAgICAgICAgICAgICAgICBsZWZ0ID0gbWF4KDcwMCwgaW50KHNhdmVkX3NpemVzWzBdKSkKICAgICAgICAgICAgICAgIHJp"
    "Z2h0ID0gbWF4KDM2MCwgaW50KHNhdmVkX3NpemVzWzFdKSkKICAgICAgICAgICAgICAgIHNlbGYuX21haW5fc3BsaXR0ZXIuc2V0"
    "U2l6ZXMoW2xlZnQsIHJpZ2h0XSkKICAgICAgICAgICAgICAgIHJldHVybgogICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgog"
    "ICAgICAgICAgICAgICAgcGFzcwoKICAgICAgICAjIERlZmF1bHQgZmF2b3JzIG1haW4gd29ya3NwYWNlIG9uIGZpcnN0IHJ1bi4K"
    "ICAgICAgICB0b3RhbCA9IG1heCgxMDYwLCBzZWxmLndpZHRoKCkgLSAyNCkKICAgICAgICBsZWZ0X2RlZmF1bHQgPSBpbnQodG90"
    "YWwgKiAwLjY4KQogICAgICAgIHJpZ2h0X2RlZmF1bHQgPSB0b3RhbCAtIGxlZnRfZGVmYXVsdAogICAgICAgIHNlbGYuX21haW5f"
    "c3BsaXR0ZXIuc2V0U2l6ZXMoW21heCg3MDAsIGxlZnRfZGVmYXVsdCksIG1heCgzNjAsIHJpZ2h0X2RlZmF1bHQpXSkKCiAgICBk"
    "ZWYgX3NhdmVfbWFpbl9zcGxpdHRlcl9zdGF0ZShzZWxmLCBfcG9zOiBpbnQsIF9pbmRleDogaW50KSAtPiBOb25lOgogICAgICAg"
    "IHNpemVzID0gc2VsZi5fbWFpbl9zcGxpdHRlci5zaXplcygpCiAgICAgICAgaWYgbGVuKHNpemVzKSAhPSAyOgogICAgICAgICAg"
    "ICByZXR1cm4KICAgICAgICBjZmdfc3BsaXR0ZXIgPSBDRkcuc2V0ZGVmYXVsdCgibWFpbl9zcGxpdHRlciIsIHt9KQogICAgICAg"
    "IGNmZ19zcGxpdHRlclsiaG9yaXpvbnRhbF9zaXplcyJdID0gW2ludChtYXgoNzAwLCBzaXplc1swXSkpLCBpbnQobWF4KDM2MCwg"
    "c2l6ZXNbMV0pKV0KICAgICAgICBzYXZlX2NvbmZpZyhDRkcpCgogICAgZGVmIF90YWJfaW5kZXhfYnlfc3BlbGxfaWQoc2VsZiwg"
    "dGFiX2lkOiBzdHIpIC0+IGludDoKICAgICAgICBmb3IgaSBpbiByYW5nZShzZWxmLl9zcGVsbF90YWJzLmNvdW50KCkpOgogICAg"
    "ICAgICAgICBpZiBzZWxmLl9zcGVsbF90YWJzLnRhYkJhcigpLnRhYkRhdGEoaSkgPT0gdGFiX2lkOgogICAgICAgICAgICAgICAg"
    "cmV0dXJuIGkKICAgICAgICByZXR1cm4gLTEKCiAgICBkZWYgX2lzX3NwZWxsX3RhYl9sb2NrZWQoc2VsZiwgdGFiX2lkOiBPcHRp"
    "b25hbFtzdHJdKSAtPiBib29sOgogICAgICAgIGlmIG5vdCB0YWJfaWQ6CiAgICAgICAgICAgIHJldHVybiBGYWxzZQogICAgICAg"
    "IHN0YXRlID0gc2VsZi5fc3BlbGxfdGFiX3N0YXRlLmdldCh0YWJfaWQsIHt9KQogICAgICAgIHJldHVybiBib29sKHN0YXRlLmdl"
    "dCgibG9ja2VkIiwgRmFsc2UpKQoKICAgIGRlZiBfbm9ybWFsaXplX3NwZWxsX2NhdGVnb3J5KHNlbGYsIHZhbHVlOiBvYmplY3Qp"
    "IC0+IHN0cjoKICAgICAgICB0ZXh0ID0gc3RyKHZhbHVlIG9yICIiKS5zdHJpcCgpLnVwcGVyKCkKICAgICAgICByZXR1cm4gdGV4"
    "dCBvciAiU1lTVEVNIgoKICAgIGRlZiBfaW5pdF9zcGVsbF9jYXRlZ29yeV9mcmFtZXdvcmsoc2VsZikgLT4gTm9uZToKICAgICAg"
    "ICBzZWxmLl9zeXN0ZW1fc3BlbGxfdGFiX2lkcyA9IHsiaW5zdHJ1bWVudHMiLCAiZGlhZ25vc3RpY3MiLCAic2V0dGluZ3MifQog"
    "ICAgICAgIHNlbGYuX3Byb3RlY3RlZF9jYXRlZ29yaWVzID0geyJTWVNURU0ifQogICAgICAgIHNlbGYuX2FjdGl2ZV9jYXRlZ29y"
    "eSA9ICJTWVNURU0iCiAgICAgICAgc2VsZi5fY2F0ZWdvcnlfbWFwOiBkaWN0W3N0ciwgZGljdF0gPSB7fQogICAgICAgIHNlbGYu"
    "X3JlYnVpbGRfc3BlbGxfY2F0ZWdvcnlfbWFwKCkKCiAgICBkZWYgX3JlYnVpbGRfc3BlbGxfY2F0ZWdvcnlfbWFwKHNlbGYpIC0+"
    "IE5vbmU6CiAgICAgICAgY2F0ZWdvcnlfbWFwOiBkaWN0W3N0ciwgZGljdF0gPSB7IlNZU1RFTSI6IHsidGFicyI6IFtdLCAicHJv"
    "dGVjdGVkIjogVHJ1ZX19CiAgICAgICAgZm9yIHRhYiBpbiBzZWxmLl9zcGVsbF90YWJfZGVmczoKICAgICAgICAgICAgY2F0ZWdv"
    "cnkgPSBzZWxmLl9ub3JtYWxpemVfc3BlbGxfY2F0ZWdvcnkodGFiLmdldCgiY2F0ZWdvcnkiLCAiU1lTVEVNIikpCiAgICAgICAg"
    "ICAgIHRhYlsiY2F0ZWdvcnkiXSA9IGNhdGVnb3J5CiAgICAgICAgICAgIHRhYi5zZXRkZWZhdWx0KCJzZWNvbmRhcnlfY2F0ZWdv"
    "cmllcyIsIFtdKQogICAgICAgICAgICB0YWJbInByb3RlY3RlZF9jYXRlZ29yeSJdID0gYm9vbCh0YWIuZ2V0KCJwcm90ZWN0ZWRf"
    "Y2F0ZWdvcnkiLCBGYWxzZSkgb3IgY2F0ZWdvcnkgPT0gIlNZU1RFTSIpCiAgICAgICAgICAgIGJ1Y2tldCA9IGNhdGVnb3J5X21h"
    "cC5zZXRkZWZhdWx0KGNhdGVnb3J5LCB7InRhYnMiOiBbXSwgInByb3RlY3RlZCI6IGJvb2wodGFiWyJwcm90ZWN0ZWRfY2F0ZWdv"
    "cnkiXSl9KQogICAgICAgICAgICBidWNrZXRbInByb3RlY3RlZCJdID0gYm9vbChidWNrZXQuZ2V0KCJwcm90ZWN0ZWQiLCBGYWxz"
    "ZSkgb3IgdGFiWyJwcm90ZWN0ZWRfY2F0ZWdvcnkiXSkKICAgICAgICAgICAgYnVja2V0WyJ0YWJzIl0uYXBwZW5kKHRhYlsiaWQi"
    "XSkKCiAgICAgICAgIyBHdWFyYW50ZWUgU1lTVEVNIGV4aXN0cyBhbmQgYWx3YXlzIGNhcnJpZXMgY29yZSB0YWJzLgogICAgICAg"
    "IHN5c3RlbV9idWNrZXQgPSBjYXRlZ29yeV9tYXAuc2V0ZGVmYXVsdCgiU1lTVEVNIiwgeyJ0YWJzIjogW10sICJwcm90ZWN0ZWQi"
    "OiBUcnVlfSkKICAgICAgICBmb3IgdGFiX2lkIGluIHNlbGYuX3N5c3RlbV9zcGVsbF90YWJfaWRzOgogICAgICAgICAgICBpZiB0"
    "YWJfaWQgbm90IGluIHN5c3RlbV9idWNrZXRbInRhYnMiXToKICAgICAgICAgICAgICAgIHN5c3RlbV9idWNrZXRbInRhYnMiXS5h"
    "cHBlbmQodGFiX2lkKQoKICAgICAgICAjIFJlbW92ZSBub24tcHJvdGVjdGVkIGNhdGVnb3JpZXMgd2l0aCBubyB0YWJzLgogICAg"
    "ICAgIGZvciBjYXQgaW4gbGlzdChjYXRlZ29yeV9tYXAua2V5cygpKToKICAgICAgICAgICAgaWYgY2F0ID09ICJTWVNURU0iOgog"
    "ICAgICAgICAgICAgICAgY29udGludWUKICAgICAgICAgICAgaWYgY2F0ZWdvcnlfbWFwW2NhdF0uZ2V0KCJ0YWJzIik6CiAgICAg"
    "ICAgICAgICAgICBjb250aW51ZQogICAgICAgICAgICBpZiBjYXRlZ29yeV9tYXBbY2F0XS5nZXQoInByb3RlY3RlZCIpOgogICAg"
    "ICAgICAgICAgICAgY29udGludWUKICAgICAgICAgICAgY2F0ZWdvcnlfbWFwLnBvcChjYXQsIE5vbmUpCgogICAgICAgIHNlbGYu"
    "X2NhdGVnb3J5X21hcCA9IGNhdGVnb3J5X21hcAogICAgICAgIGlmIHNlbGYuX2FjdGl2ZV9jYXRlZ29yeSBub3QgaW4gc2VsZi5f"
    "Y2F0ZWdvcnlfbWFwOgogICAgICAgICAgICBzZWxmLl9hY3RpdmVfY2F0ZWdvcnkgPSAiU1lTVEVNIiBpZiAiU1lTVEVNIiBpbiBz"
    "ZWxmLl9jYXRlZ29yeV9tYXAgZWxzZSBuZXh0KGl0ZXIoc2VsZi5fY2F0ZWdvcnlfbWFwLmtleXMoKSksICJTWVNURU0iKQogICAg"
    "ICAgIHNlbGYuX3JlYnVpbGRfY2F0ZWdvcnlfc3RyaXAoKQoKICAgIGRlZiBfY2F0ZWdvcnlfb3JkZXJlZF9saXN0KHNlbGYpIC0+"
    "IGxpc3Rbc3RyXToKICAgICAgICBjYXRzID0gW2MgZm9yIGMgaW4gc2VsZi5fY2F0ZWdvcnlfbWFwLmtleXMoKSBpZiBjICE9ICJT"
    "WVNURU0iXQogICAgICAgIGNhdHMuc29ydCgpCiAgICAgICAgcmV0dXJuIFsiU1lTVEVNIiwgKmNhdHNdIGlmICJTWVNURU0iIGlu"
    "IHNlbGYuX2NhdGVnb3J5X21hcCBlbHNlIGNhdHMKCiAgICBkZWYgX3JlYnVpbGRfY2F0ZWdvcnlfc3RyaXAoc2VsZikgLT4gTm9u"
    "ZToKICAgICAgICBpZiBub3QgaGFzYXR0cihzZWxmLCAiX2NhdGVnb3J5X3N0cmlwX2xheW91dCIpOgogICAgICAgICAgICByZXR1"
    "cm4KICAgICAgICB3aGlsZSBzZWxmLl9jYXRlZ29yeV9zdHJpcF9sYXlvdXQuY291bnQoKToKICAgICAgICAgICAgaXRlbSA9IHNl"
    "bGYuX2NhdGVnb3J5X3N0cmlwX2xheW91dC50YWtlQXQoMCkKICAgICAgICAgICAgdyA9IGl0ZW0ud2lkZ2V0KCkKICAgICAgICAg"
    "ICAgaWYgdyBpcyBub3QgTm9uZToKICAgICAgICAgICAgICAgIHcuZGVsZXRlTGF0ZXIoKQoKICAgICAgICBmb3IgY2F0IGluIHNl"
    "bGYuX2NhdGVnb3J5X29yZGVyZWRfbGlzdCgpOgogICAgICAgICAgICBidG4gPSBRVG9vbEJ1dHRvbihzZWxmLl9jYXRlZ29yeV9z"
    "dHJpcCkKICAgICAgICAgICAgYnRuLnNldFRleHQoY2F0KQogICAgICAgICAgICBidG4uc2V0Q2hlY2thYmxlKFRydWUpCiAgICAg"
    "ICAgICAgIGJ0bi5zZXRDaGVja2VkKGNhdCA9PSBzZWxmLl9hY3RpdmVfY2F0ZWdvcnkpCiAgICAgICAgICAgIGJ0bi5zZXRTdHls"
    "ZVNoZWV0KAogICAgICAgICAgICAgICAgZiJRVG9vbEJ1dHRvbnt7cGFkZGluZzo0cHggMTBweDsgYm9yZGVyOjFweCBzb2xpZCB7"
    "Q19DUklNU09OX0RJTX07IGJhY2tncm91bmQ6e0NfQkcyfTsgY29sb3I6e0NfVEVYVF9ESU19OyBmb250LWZhbWlseTp7REVDS19G"
    "T05UfSwgc2VyaWY7IGZvbnQtc2l6ZToxMHB4OyBsZXR0ZXItc3BhY2luZzoxcHg7fX0iCiAgICAgICAgICAgICAgICBmIlFUb29s"
    "QnV0dG9uOmNoZWNrZWR7e2JhY2tncm91bmQ6e0NfQ1JJTVNPTl9ESU19OyBjb2xvcjp7Q19HT0xEfTsgYm9yZGVyOjFweCBzb2xp"
    "ZCB7Q19DUklNU09OfTt9fSIKICAgICAgICAgICAgICAgIGYiUVRvb2xCdXR0b246aG92ZXJ7e2NvbG9yOntDX1RFWFR9OyBib3Jk"
    "ZXI6MXB4IHNvbGlkIHtDX0NSSU1TT059O319IgogICAgICAgICAgICApCiAgICAgICAgICAgIGJ0bi5jbGlja2VkLmNvbm5lY3Qo"
    "bGFtYmRhIF9jaGVja2VkPUZhbHNlLCBjPWNhdDogc2VsZi5fc2VsZWN0X3NwZWxsX2NhdGVnb3J5KGMpKQogICAgICAgICAgICBz"
    "ZWxmLl9jYXRlZ29yeV9zdHJpcF9sYXlvdXQuYWRkV2lkZ2V0KGJ0bikKICAgICAgICBzZWxmLl9jYXRlZ29yeV9zdHJpcF9sYXlv"
    "dXQuYWRkU3RyZXRjaCgxKQoKICAgIGRlZiBfc2VsZWN0X3NwZWxsX2NhdGVnb3J5KHNlbGYsIGNhdGVnb3J5OiBzdHIpIC0+IE5v"
    "bmU6CiAgICAgICAgaWYgY2F0ZWdvcnkgbm90IGluIHNlbGYuX2NhdGVnb3J5X21hcDoKICAgICAgICAgICAgY2F0ZWdvcnkgPSAi"
    "U1lTVEVNIgogICAgICAgICAgICBpZiBjYXRlZ29yeSBub3QgaW4gc2VsZi5fY2F0ZWdvcnlfbWFwOgogICAgICAgICAgICAgICAg"
    "Y2F0ZWdvcnkgPSBuZXh0KGl0ZXIoc2VsZi5fY2F0ZWdvcnlfbWFwLmtleXMoKSksICJTWVNURU0iKQogICAgICAgIHNlbGYuX2Fj"
    "dGl2ZV9jYXRlZ29yeSA9IGNhdGVnb3J5CiAgICAgICAgc2VsZi5fcGVyc2lzdF9sYXN0X3NlbGVjdGVkX2NhdGVnb3J5KCkKICAg"
    "ICAgICBzZWxmLl9yZWJ1aWxkX2NhdGVnb3J5X3N0cmlwKCkKICAgICAgICBzZWxmLl9yZWJ1aWxkX3NwZWxsX3RhYnMoKQoKICAg"
    "IGRlZiBfcmVzdG9yZV9sYXN0X3NlbGVjdGVkX2NhdGVnb3J5KHNlbGYpIC0+IE5vbmU6CiAgICAgICAgY2FuZGlkYXRlID0gc2Vs"
    "Zi5fbm9ybWFsaXplX3NwZWxsX2NhdGVnb3J5KENGRy5nZXQoImxhc3Rfc2VsZWN0ZWRfY2F0ZWdvcnkiLCAiU1lTVEVNIikpCiAg"
    "ICAgICAgaWYgY2FuZGlkYXRlIGluIHNlbGYuX2NhdGVnb3J5X21hcDoKICAgICAgICAgICAgc2VsZi5fYWN0aXZlX2NhdGVnb3J5"
    "ID0gY2FuZGlkYXRlCiAgICAgICAgZWxzZToKICAgICAgICAgICAgc2VsZi5fYWN0aXZlX2NhdGVnb3J5ID0gIlNZU1RFTSIgaWYg"
    "IlNZU1RFTSIgaW4gc2VsZi5fY2F0ZWdvcnlfbWFwIGVsc2UgbmV4dChpdGVyKHNlbGYuX2NhdGVnb3J5X21hcC5rZXlzKCkpLCAi"
    "U1lTVEVNIikKCiAgICBkZWYgX3BlcnNpc3RfbGFzdF9zZWxlY3RlZF9jYXRlZ29yeShzZWxmKSAtPiBOb25lOgogICAgICAgIENG"
    "R1sibGFzdF9zZWxlY3RlZF9jYXRlZ29yeSJdID0gc2VsZi5fYWN0aXZlX2NhdGVnb3J5CiAgICAgICAgc2F2ZV9jb25maWcoQ0ZH"
    "KQoKICAgIGRlZiBfdmlzaWJsZV90YWJfZm9yX2FjdGl2ZV9jYXRlZ29yeShzZWxmLCB0YWI6IGRpY3QpIC0+IGJvb2w6CiAgICAg"
    "ICAgY2F0ZWdvcnkgPSBzZWxmLl9ub3JtYWxpemVfc3BlbGxfY2F0ZWdvcnkodGFiLmdldCgiY2F0ZWdvcnkiLCAiU1lTVEVNIikp"
    "CiAgICAgICAgaWYgc2VsZi5fYWN0aXZlX2NhdGVnb3J5IG5vdCBpbiBzZWxmLl9jYXRlZ29yeV9tYXA6CiAgICAgICAgICAgIHJl"
    "dHVybiBjYXRlZ29yeSA9PSAiU1lTVEVNIgogICAgICAgIHJldHVybiBjYXRlZ29yeSA9PSBzZWxmLl9hY3RpdmVfY2F0ZWdvcnkK"
    "CiAgICBkZWYgX2xvYWRfc3BlbGxfdGFiX3N0YXRlX2Zyb21fY29uZmlnKHNlbGYpLT5Ob25lOgogICAgICAgIHNhdmVkID0gQ0ZH"
    "LmdldCgibW9kdWxlX3RhYl9vcmRlciIsIFtdKQogICAgICAgIHNhdmVkX21hcCA9IHt9CiAgICAgICAgaWYgaXNpbnN0YW5jZShz"
    "YXZlZCwgbGlzdCk6CiAgICAgICAgICAgIGZvciBlbnRyeSBpbiBzYXZlZDoKICAgICAgICAgICAgICAgIGlmIGlzaW5zdGFuY2Uo"
    "ZW50cnksIGRpY3QpIGFuZCBlbnRyeS5nZXQoImlkIik6CiAgICAgICAgICAgICAgICAgICAgc2F2ZWRfbWFwW3N0cihlbnRyeVsi"
    "aWQiXSldID0gZW50cnkKCiAgICAgICAgc2VsZi5fc3BlbGxfdGFiX3N0YXRlID0ge30KICAgICAgICBzZWxmLl9yZWJ1aWxkX3Nw"
    "ZWxsX2NhdGVnb3J5X21hcCgpCiAgICAgICAgc2VsZi5fcmVzdG9yZV9sYXN0X3NlbGVjdGVkX2NhdGVnb3J5KCkKICAgICAgICBm"
    "b3IgdGFiIGluIHNlbGYuX3NwZWxsX3RhYl9kZWZzOgogICAgICAgICAgICB0YWJfaWQgPSB0YWJbImlkIl0KICAgICAgICAgICAg"
    "ZGVmYXVsdF9vcmRlciA9IGludCh0YWJbImRlZmF1bHRfb3JkZXIiXSkKICAgICAgICAgICAgZW50cnkgPSBzYXZlZF9tYXAuZ2V0"
    "KHRhYl9pZCwge30pCiAgICAgICAgICAgIG9yZGVyX3ZhbCA9IGVudHJ5LmdldCgib3JkZXIiLCBkZWZhdWx0X29yZGVyKQogICAg"
    "ICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICBvcmRlcl92YWwgPSBpbnQob3JkZXJfdmFsKQogICAgICAgICAgICBleGNlcHQg"
    "RXhjZXB0aW9uOgogICAgICAgICAgICAgICAgb3JkZXJfdmFsID0gZGVmYXVsdF9vcmRlcgogICAgICAgICAgICBzZWxmLl9zcGVs"
    "bF90YWJfc3RhdGVbdGFiX2lkXSA9IHsKICAgICAgICAgICAgICAgICJvcmRlciI6IG9yZGVyX3ZhbCwKICAgICAgICAgICAgICAg"
    "ICJsb2NrZWQiOiBib29sKGVudHJ5LmdldCgibG9ja2VkIiwgRmFsc2UpKSwKICAgICAgICAgICAgICAgICJkZWZhdWx0X29yZGVy"
    "IjogZGVmYXVsdF9vcmRlciwKICAgICAgICAgICAgfQoKICAgIGRlZiBfb3JkZXJlZF9zcGVsbF90YWJfZGVmcyhzZWxmKSAtPiBs"
    "aXN0W2RpY3RdOgogICAgICAgIHJldHVybiBzb3J0ZWQoCiAgICAgICAgICAgIHNlbGYuX3NwZWxsX3RhYl9kZWZzLAogICAgICAg"
    "ICAgICBrZXk9bGFtYmRhIHQ6ICgKICAgICAgICAgICAgICAgIGludChzZWxmLl9zcGVsbF90YWJfc3RhdGUuZ2V0KHRbImlkIl0s"
    "IHt9KS5nZXQoIm9yZGVyIiwgdFsiZGVmYXVsdF9vcmRlciJdKSksCiAgICAgICAgICAgICAgICBpbnQodFsiZGVmYXVsdF9vcmRl"
    "ciJdKSwKICAgICAgICAgICAgKSwKICAgICAgICApCgogICAgZGVmIF9yZWJ1aWxkX3NwZWxsX3RhYnMoc2VsZikgLT4gTm9uZToK"
    "ICAgICAgICBzZWxmLl9yZWJ1aWxkX3NwZWxsX2NhdGVnb3J5X21hcCgpCgogICAgICAgIGN1cnJlbnRfaWQgPSBOb25lCiAgICAg"
    "ICAgaWR4ID0gc2VsZi5fc3BlbGxfdGFicy5jdXJyZW50SW5kZXgoKQogICAgICAgIGlmIGlkeCA+PSAwOgogICAgICAgICAgICBj"
    "dXJyZW50X2lkID0gc2VsZi5fc3BlbGxfdGFicy50YWJCYXIoKS50YWJEYXRhKGlkeCkKCiAgICAgICAgdmlzaWJsZV90YWJzID0g"
    "W3QgZm9yIHQgaW4gc2VsZi5fb3JkZXJlZF9zcGVsbF90YWJfZGVmcygpIGlmIHNlbGYuX3Zpc2libGVfdGFiX2Zvcl9hY3RpdmVf"
    "Y2F0ZWdvcnkodCldCgogICAgICAgIHNlbGYuX3N1cHByZXNzX3NwZWxsX3RhYl9tb3ZlX3NpZ25hbCA9IFRydWUKICAgICAgICB3"
    "aGlsZSBzZWxmLl9zcGVsbF90YWJzLmNvdW50KCk6CiAgICAgICAgICAgIHNlbGYuX3NwZWxsX3RhYnMucmVtb3ZlVGFiKDApCgog"
    "ICAgICAgIGZvciB0YWIgaW4gdmlzaWJsZV90YWJzOgogICAgICAgICAgICBpID0gc2VsZi5fc3BlbGxfdGFicy5hZGRUYWIodGFi"
    "WyJ3aWRnZXQiXSwgdGFiWyJ0aXRsZSJdKQogICAgICAgICAgICBzZWxmLl9zcGVsbF90YWJzLnRhYkJhcigpLnNldFRhYkRhdGEo"
    "aSwgdGFiWyJpZCJdKQoKICAgICAgICB0YXJnZXRfaWQgPSBjdXJyZW50X2lkIGlmIGFueSh0WyJpZCJdID09IGN1cnJlbnRfaWQg"
    "Zm9yIHQgaW4gdmlzaWJsZV90YWJzKSBlbHNlICh2aXNpYmxlX3RhYnNbMF1bImlkIl0gaWYgdmlzaWJsZV90YWJzIGVsc2UgTm9u"
    "ZSkKICAgICAgICBpZiB0YXJnZXRfaWQ6CiAgICAgICAgICAgIG5ld19pZHggPSBzZWxmLl90YWJfaW5kZXhfYnlfc3BlbGxfaWQo"
    "dGFyZ2V0X2lkKQogICAgICAgICAgICBpZiBuZXdfaWR4ID49IDA6CiAgICAgICAgICAgICAgICBzZWxmLl9zcGVsbF90YWJzLnNl"
    "dEN1cnJlbnRJbmRleChuZXdfaWR4KQoKICAgICAgICBzZWxmLl9zdXBwcmVzc19zcGVsbF90YWJfbW92ZV9zaWduYWwgPSBGYWxz"
    "ZQogICAgICAgIHNlbGYuX2V4aXRfc3BlbGxfdGFiX21vdmVfbW9kZSgpCgogICAgZGVmIF9wZXJzaXN0X3NwZWxsX3RhYl9vcmRl"
    "cl90b19jb25maWcoc2VsZikgLT4gTm9uZToKICAgICAgICBvcmRlcmVkX3Zpc2libGVfaWRzID0gW3NlbGYuX3NwZWxsX3RhYnMu"
    "dGFiQmFyKCkudGFiRGF0YShpKSBmb3IgaSBpbiByYW5nZShzZWxmLl9zcGVsbF90YWJzLmNvdW50KCkpXQogICAgICAgIGJhc2Vf"
    "b3JkZXIgPSBsZW4oc2VsZi5fc3BlbGxfdGFiX2RlZnMpICsgMTAKICAgICAgICBmb3IgdGFiIGluIHNlbGYuX3NwZWxsX3RhYl9k"
    "ZWZzOgogICAgICAgICAgICB0YWJfaWQgPSB0YWJbImlkIl0KICAgICAgICAgICAgaWYgdGFiX2lkIGluIG9yZGVyZWRfdmlzaWJs"
    "ZV9pZHM6CiAgICAgICAgICAgICAgICBzZWxmLl9zcGVsbF90YWJfc3RhdGVbdGFiX2lkXVsib3JkZXIiXSA9IG9yZGVyZWRfdmlz"
    "aWJsZV9pZHMuaW5kZXgodGFiX2lkKQogICAgICAgICAgICBlbHNlOgogICAgICAgICAgICAgICAgc2VsZi5fc3BlbGxfdGFiX3N0"
    "YXRlW3RhYl9pZF1bIm9yZGVyIl0gPSBpbnQoc2VsZi5fc3BlbGxfdGFiX3N0YXRlW3RhYl9pZF0uZ2V0KCJvcmRlciIsIGJhc2Vf"
    "b3JkZXIpKSArIGJhc2Vfb3JkZXIKCiAgICAgICAgQ0ZHWyJtb2R1bGVfdGFiX29yZGVyIl0gPSBbCiAgICAgICAgICAgIHsiaWQi"
    "OiB0YWJbImlkIl0sICJvcmRlciI6IGludChzZWxmLl9zcGVsbF90YWJfc3RhdGVbdGFiWyJpZCJdXVsib3JkZXIiXSksICJsb2Nr"
    "ZWQiOiBib29sKHNlbGYuX3NwZWxsX3RhYl9zdGF0ZVt0YWJbImlkIl1dWyJsb2NrZWQiXSl9CiAgICAgICAgICAgIGZvciB0YWIg"
    "aW4gc29ydGVkKHNlbGYuX3NwZWxsX3RhYl9kZWZzLCBrZXk9bGFtYmRhIHQ6IHRbImRlZmF1bHRfb3JkZXIiXSkKICAgICAgICBd"
    "CiAgICAgICAgQ0ZHWyJsYXN0X3NlbGVjdGVkX2NhdGVnb3J5Il0gPSBzZWxmLl9hY3RpdmVfY2F0ZWdvcnkKICAgICAgICBzYXZl"
    "X2NvbmZpZyhDRkcpCgogICAgZGVmIF9jYW5fY3Jvc3Nfc3BlbGxfdGFiX3JhbmdlKHNlbGYsIGZyb21faWR4OiBpbnQsIHRvX2lk"
    "eDogaW50KSAtPiBib29sOgogICAgICAgIGlmIGZyb21faWR4IDwgMCBvciB0b19pZHggPCAwOgogICAgICAgICAgICByZXR1cm4g"
    "RmFsc2UKICAgICAgICBtb3ZpbmdfaWQgPSBzZWxmLl9zcGVsbF90YWJzLnRhYkJhcigpLnRhYkRhdGEodG9faWR4KQogICAgICAg"
    "IGlmIHNlbGYuX2lzX3NwZWxsX3RhYl9sb2NrZWQobW92aW5nX2lkKToKICAgICAgICAgICAgcmV0dXJuIEZhbHNlCiAgICAgICAg"
    "bGVmdCA9IG1pbihmcm9tX2lkeCwgdG9faWR4KQogICAgICAgIHJpZ2h0ID0gbWF4KGZyb21faWR4LCB0b19pZHgpCiAgICAgICAg"
    "Zm9yIGkgaW4gcmFuZ2UobGVmdCwgcmlnaHQgKyAxKToKICAgICAgICAgICAgaWYgaSA9PSB0b19pZHg6CiAgICAgICAgICAgICAg"
    "ICBjb250aW51ZQogICAgICAgICAgICBvdGhlcl9pZCA9IHNlbGYuX3NwZWxsX3RhYnMudGFiQmFyKCkudGFiRGF0YShpKQogICAg"
    "ICAgICAgICBpZiBzZWxmLl9pc19zcGVsbF90YWJfbG9ja2VkKG90aGVyX2lkKToKICAgICAgICAgICAgICAgIHJldHVybiBGYWxz"
    "ZQogICAgICAgIHJldHVybiBUcnVlCgogICAgZGVmIF9vbl9zcGVsbF90YWJfZHJhZ19tb3ZlZChzZWxmLCBmcm9tX2lkeDogaW50"
    "LCB0b19pZHg6IGludCkgLT4gTm9uZToKICAgICAgICBpZiBzZWxmLl9zdXBwcmVzc19zcGVsbF90YWJfbW92ZV9zaWduYWw6CiAg"
    "ICAgICAgICAgIHJldHVybgogICAgICAgIGlmIG5vdCBzZWxmLl9jYW5fY3Jvc3Nfc3BlbGxfdGFiX3JhbmdlKGZyb21faWR4LCB0"
    "b19pZHgpOgogICAgICAgICAgICBzZWxmLl9zdXBwcmVzc19zcGVsbF90YWJfbW92ZV9zaWduYWwgPSBUcnVlCiAgICAgICAgICAg"
    "IHNlbGYuX3NwZWxsX3RhYl9iYXIubW92ZVRhYih0b19pZHgsIGZyb21faWR4KQogICAgICAgICAgICBzZWxmLl9zdXBwcmVzc19z"
    "cGVsbF90YWJfbW92ZV9zaWduYWwgPSBGYWxzZQogICAgICAgICAgICByZXR1cm4KICAgICAgICBzZWxmLl9wZXJzaXN0X3NwZWxs"
    "X3RhYl9vcmRlcl90b19jb25maWcoKQogICAgICAgIHNlbGYuX3JlZnJlc2hfc3BlbGxfdGFiX21vdmVfY29udHJvbHMoKQoKICAg"
    "IGRlZiBfc2hvd19zcGVsbF90YWJfY29udGV4dF9tZW51KHNlbGYsIHBvczogUVBvaW50KSAtPiBOb25lOgogICAgICAgIGlkeCA9"
    "IHNlbGYuX3NwZWxsX3RhYl9iYXIudGFiQXQocG9zKQogICAgICAgIGlmIGlkeCA8IDA6CiAgICAgICAgICAgIHJldHVybgogICAg"
    "ICAgIHRhYl9pZCA9IHNlbGYuX3NwZWxsX3RhYl9iYXIudGFiRGF0YShpZHgpCiAgICAgICAgaWYgbm90IHRhYl9pZDoKICAgICAg"
    "ICAgICAgcmV0dXJuCgogICAgICAgIG1lbnUgPSBRTWVudShzZWxmKQogICAgICAgIG1vdmVfYWN0aW9uID0gbWVudS5hZGRBY3Rp"
    "b24oIk1vdmUiKQogICAgICAgIGlmIHNlbGYuX2lzX3NwZWxsX3RhYl9sb2NrZWQodGFiX2lkKToKICAgICAgICAgICAgbG9ja19h"
    "Y3Rpb24gPSBtZW51LmFkZEFjdGlvbigiVW5sb2NrIikKICAgICAgICBlbHNlOgogICAgICAgICAgICBsb2NrX2FjdGlvbiA9IG1l"
    "bnUuYWRkQWN0aW9uKCJTZWN1cmUiKQogICAgICAgIG1lbnUuYWRkU2VwYXJhdG9yKCkKICAgICAgICByZXNldF9hY3Rpb24gPSBt"
    "ZW51LmFkZEFjdGlvbigiUmVzZXQgdG8gRGVmYXVsdCBPcmRlciIpCgogICAgICAgIGNob2ljZSA9IG1lbnUuZXhlYyhzZWxmLl9z"
    "cGVsbF90YWJfYmFyLm1hcFRvR2xvYmFsKHBvcykpCiAgICAgICAgaWYgY2hvaWNlID09IG1vdmVfYWN0aW9uOgogICAgICAgICAg"
    "ICBpZiBub3Qgc2VsZi5faXNfc3BlbGxfdGFiX2xvY2tlZCh0YWJfaWQpOgogICAgICAgICAgICAgICAgc2VsZi5fZW50ZXJfc3Bl"
    "bGxfdGFiX21vdmVfbW9kZSh0YWJfaWQpCiAgICAgICAgZWxpZiBjaG9pY2UgPT0gbG9ja19hY3Rpb246CiAgICAgICAgICAgIHNl"
    "bGYuX3NwZWxsX3RhYl9zdGF0ZVt0YWJfaWRdWyJsb2NrZWQiXSA9IG5vdCBzZWxmLl9pc19zcGVsbF90YWJfbG9ja2VkKHRhYl9p"
    "ZCkKICAgICAgICAgICAgc2VsZi5fcGVyc2lzdF9zcGVsbF90YWJfb3JkZXJfdG9fY29uZmlnKCkKICAgICAgICAgICAgc2VsZi5f"
    "cmVmcmVzaF9zcGVsbF90YWJfbW92ZV9jb250cm9scygpCiAgICAgICAgZWxpZiBjaG9pY2UgPT0gcmVzZXRfYWN0aW9uOgogICAg"
    "ICAgICAgICBmb3IgdGFiIGluIHNlbGYuX3NwZWxsX3RhYl9kZWZzOgogICAgICAgICAgICAgICAgc2VsZi5fc3BlbGxfdGFiX3N0"
    "YXRlW3RhYlsiaWQiXV1bIm9yZGVyIl0gPSBpbnQodGFiWyJkZWZhdWx0X29yZGVyIl0pCiAgICAgICAgICAgIHNlbGYuX3JlYnVp"
    "bGRfc3BlbGxfdGFicygpCiAgICAgICAgICAgIHNlbGYuX3BlcnNpc3Rfc3BlbGxfdGFiX29yZGVyX3RvX2NvbmZpZygpCgogICAg"
    "ZGVmIF9lbnRlcl9zcGVsbF90YWJfbW92ZV9tb2RlKHNlbGYsIHRhYl9pZDogc3RyKSAtPiBOb25lOgogICAgICAgIHNlbGYuX3Nw"
    "ZWxsX3RhYl9tb3ZlX21vZGVfaWQgPSB0YWJfaWQKICAgICAgICBzZWxmLl9yZWZyZXNoX3NwZWxsX3RhYl9tb3ZlX2NvbnRyb2xz"
    "KCkKCiAgICBkZWYgX2V4aXRfc3BlbGxfdGFiX21vdmVfbW9kZShzZWxmKSAtPiBOb25lOgogICAgICAgIHNlbGYuX3NwZWxsX3Rh"
    "Yl9tb3ZlX21vZGVfaWQgPSBOb25lCiAgICAgICAgc2VsZi5fcmVmcmVzaF9zcGVsbF90YWJfbW92ZV9jb250cm9scygpCgogICAg"
    "ZGVmIF9vbl9nbG9iYWxfZm9jdXNfY2hhbmdlZChzZWxmLCBfb2xkLCBub3cpIC0+IE5vbmU6CiAgICAgICAgaWYgbm90IHNlbGYu"
    "X3NwZWxsX3RhYl9tb3ZlX21vZGVfaWQ6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIGlmIG5vdyBpcyBOb25lOgogICAgICAg"
    "ICAgICBzZWxmLl9leGl0X3NwZWxsX3RhYl9tb3ZlX21vZGUoKQogICAgICAgICAgICByZXR1cm4KICAgICAgICBpZiBub3cgaXMg"
    "c2VsZi5fc3BlbGxfdGFiX2JhcjoKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgaWYgaXNpbnN0YW5jZShub3csIFFUb29sQnV0"
    "dG9uKSBhbmQgbm93LnBhcmVudCgpIGlzIHNlbGYuX3NwZWxsX3RhYl9iYXI6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIHNl"
    "bGYuX2V4aXRfc3BlbGxfdGFiX21vdmVfbW9kZSgpCgogICAgZGVmIF9yZWZyZXNoX3NwZWxsX3RhYl9tb3ZlX2NvbnRyb2xzKHNl"
    "bGYpIC0+IE5vbmU6CiAgICAgICAgZm9yIGkgaW4gcmFuZ2Uoc2VsZi5fc3BlbGxfdGFicy5jb3VudCgpKToKICAgICAgICAgICAg"
    "c2VsZi5fc3BlbGxfdGFiX2Jhci5zZXRUYWJCdXR0b24oaSwgUVRhYkJhci5CdXR0b25Qb3NpdGlvbi5MZWZ0U2lkZSwgTm9uZSkK"
    "ICAgICAgICAgICAgc2VsZi5fc3BlbGxfdGFiX2Jhci5zZXRUYWJCdXR0b24oaSwgUVRhYkJhci5CdXR0b25Qb3NpdGlvbi5SaWdo"
    "dFNpZGUsIE5vbmUpCgogICAgICAgIHRhYl9pZCA9IHNlbGYuX3NwZWxsX3RhYl9tb3ZlX21vZGVfaWQKICAgICAgICBpZiBub3Qg"
    "dGFiX2lkIG9yIHNlbGYuX2lzX3NwZWxsX3RhYl9sb2NrZWQodGFiX2lkKToKICAgICAgICAgICAgcmV0dXJuCgogICAgICAgIGlk"
    "eCA9IHNlbGYuX3RhYl9pbmRleF9ieV9zcGVsbF9pZCh0YWJfaWQpCiAgICAgICAgaWYgaWR4IDwgMDoKICAgICAgICAgICAgcmV0"
    "dXJuCgogICAgICAgIGxlZnRfYnRuID0gUVRvb2xCdXR0b24oc2VsZi5fc3BlbGxfdGFiX2JhcikKICAgICAgICBsZWZ0X2J0bi5z"
    "ZXRUZXh0KCI8IikKICAgICAgICBsZWZ0X2J0bi5zZXRBdXRvUmFpc2UoVHJ1ZSkKICAgICAgICBsZWZ0X2J0bi5zZXRGaXhlZFNp"
    "emUoMTQsIDE0KQogICAgICAgIGxlZnRfYnRuLnNldEVuYWJsZWQoaWR4ID4gMCBhbmQgbm90IHNlbGYuX2lzX3NwZWxsX3RhYl9s"
    "b2NrZWQoc2VsZi5fc3BlbGxfdGFiX2Jhci50YWJEYXRhKGlkeCAtIDEpKSkKICAgICAgICBsZWZ0X2J0bi5jbGlja2VkLmNvbm5l"
    "Y3QobGFtYmRhOiBzZWxmLl9tb3ZlX3NwZWxsX3RhYl9zdGVwKHRhYl9pZCwgLTEpKQoKICAgICAgICByaWdodF9idG4gPSBRVG9v"
    "bEJ1dHRvbihzZWxmLl9zcGVsbF90YWJfYmFyKQogICAgICAgIHJpZ2h0X2J0bi5zZXRUZXh0KCI+IikKICAgICAgICByaWdodF9i"
    "dG4uc2V0QXV0b1JhaXNlKFRydWUpCiAgICAgICAgcmlnaHRfYnRuLnNldEZpeGVkU2l6ZSgxNCwgMTQpCiAgICAgICAgcmlnaHRf"
    "YnRuLnNldEVuYWJsZWQoCiAgICAgICAgICAgIGlkeCA8IChzZWxmLl9zcGVsbF90YWJzLmNvdW50KCkgLSAxKSBhbmQKICAgICAg"
    "ICAgICAgbm90IHNlbGYuX2lzX3NwZWxsX3RhYl9sb2NrZWQoc2VsZi5fc3BlbGxfdGFiX2Jhci50YWJEYXRhKGlkeCArIDEpKQog"
    "ICAgICAgICkKICAgICAgICByaWdodF9idG4uY2xpY2tlZC5jb25uZWN0KGxhbWJkYTogc2VsZi5fbW92ZV9zcGVsbF90YWJfc3Rl"
    "cCh0YWJfaWQsIDEpKQoKICAgICAgICBzZWxmLl9zcGVsbF90YWJfYmFyLnNldFRhYkJ1dHRvbihpZHgsIFFUYWJCYXIuQnV0dG9u"
    "UG9zaXRpb24uTGVmdFNpZGUsIGxlZnRfYnRuKQogICAgICAgIHNlbGYuX3NwZWxsX3RhYl9iYXIuc2V0VGFiQnV0dG9uKGlkeCwg"
    "UVRhYkJhci5CdXR0b25Qb3NpdGlvbi5SaWdodFNpZGUsIHJpZ2h0X2J0bikKCiAgICBkZWYgX21vdmVfc3BlbGxfdGFiX3N0ZXAo"
    "c2VsZiwgdGFiX2lkOiBzdHIsIGRlbHRhOiBpbnQpIC0+IE5vbmU6CiAgICAgICAgaWYgc2VsZi5faXNfc3BlbGxfdGFiX2xvY2tl"
    "ZCh0YWJfaWQpOgogICAgICAgICAgICByZXR1cm4KICAgICAgICBjdXJyZW50X2lkeCA9IHNlbGYuX3RhYl9pbmRleF9ieV9zcGVs"
    "bF9pZCh0YWJfaWQpCiAgICAgICAgaWYgY3VycmVudF9pZHggPCAwOgogICAgICAgICAgICByZXR1cm4KCiAgICAgICAgdGFyZ2V0"
    "X2lkeCA9IGN1cnJlbnRfaWR4ICsgZGVsdGEKICAgICAgICBpZiB0YXJnZXRfaWR4IDwgMCBvciB0YXJnZXRfaWR4ID49IHNlbGYu"
    "X3NwZWxsX3RhYnMuY291bnQoKToKICAgICAgICAgICAgcmV0dXJuCgogICAgICAgIHRhcmdldF9pZCA9IHNlbGYuX3NwZWxsX3Rh"
    "Yl9iYXIudGFiRGF0YSh0YXJnZXRfaWR4KQogICAgICAgIGlmIHNlbGYuX2lzX3NwZWxsX3RhYl9sb2NrZWQodGFyZ2V0X2lkKToK"
    "ICAgICAgICAgICAgcmV0dXJuCgogICAgICAgIHNlbGYuX3N1cHByZXNzX3NwZWxsX3RhYl9tb3ZlX3NpZ25hbCA9IFRydWUKICAg"
    "ICAgICBzZWxmLl9zcGVsbF90YWJfYmFyLm1vdmVUYWIoY3VycmVudF9pZHgsIHRhcmdldF9pZHgpCiAgICAgICAgc2VsZi5fc3Vw"
    "cHJlc3Nfc3BlbGxfdGFiX21vdmVfc2lnbmFsID0gRmFsc2UKICAgICAgICBzZWxmLl9wZXJzaXN0X3NwZWxsX3RhYl9vcmRlcl90"
    "b19jb25maWcoKQogICAgICAgIHNlbGYuX3JlZnJlc2hfc3BlbGxfdGFiX21vdmVfY29udHJvbHMoKQoKICAgICMg4pSA4pSAIFNU"
    "QVJUVVAgU0VRVUVOQ0Ug4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgX3N0YXJ0dXBfc2VxdWVuY2Uoc2VsZikgLT4gTm9uZToKICAg"
    "ICAgICBzZWxmLl9hcHBlbmRfY2hhdCgiU1lTVEVNIiwgZiLinKYge0FQUF9OQU1FfSBBV0FLRU5JTkcuLi4iKQogICAgICAgIHNl"
    "bGYuX2FwcGVuZF9jaGF0KCJTWVNURU0iLCBmIuKcpiB7UlVORVN9IOKcpiIpCgogICAgICAgICMgTG9hZCBib290c3RyYXAgbG9n"
    "CiAgICAgICAgYm9vdF9sb2cgPSBTQ1JJUFRfRElSIC8gImxvZ3MiIC8gImJvb3RzdHJhcF9sb2cudHh0IgogICAgICAgIGlmIGJv"
    "b3RfbG9nLmV4aXN0cygpOgogICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICBtc2dzID0gYm9vdF9sb2cucmVhZF90ZXh0"
    "KGVuY29kaW5nPSJ1dGYtOCIpLnNwbGl0bGluZXMoKQogICAgICAgICAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nX21hbnkobXNn"
    "cykKICAgICAgICAgICAgICAgIGJvb3RfbG9nLnVubGluaygpICAjIGNvbnN1bWVkCiAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRp"
    "b246CiAgICAgICAgICAgICAgICBwYXNzCgogICAgICAgICMgSGFyZHdhcmUgZGV0ZWN0aW9uIG1lc3NhZ2VzCiAgICAgICAgc2Vs"
    "Zi5fZGlhZ190YWIubG9nX21hbnkoc2VsZi5faHdfcGFuZWwuZ2V0X2RpYWdub3N0aWNzKCkpCgogICAgICAgICMgRGVwIGNoZWNr"
    "CiAgICAgICAgZGVwX21zZ3MsIGNyaXRpY2FsID0gRGVwZW5kZW5jeUNoZWNrZXIuY2hlY2soKQogICAgICAgIHNlbGYuX2RpYWdf"
    "dGFiLmxvZ19tYW55KGRlcF9tc2dzKQoKICAgICAgICAjIExvYWQgcGFzdCBzdGF0ZQogICAgICAgIGxhc3Rfc3RhdGUgPSBzZWxm"
    "Ll9zdGF0ZS5nZXQoImFpX3N0YXRlX2F0X3NodXRkb3duIiwiIikKICAgICAgICBpZiBsYXN0X3N0YXRlOgogICAgICAgICAgICBz"
    "ZWxmLl9kaWFnX3RhYi5sb2coCiAgICAgICAgICAgICAgICBmIltTVEFSVFVQXSBMYXN0IHNodXRkb3duIHN0YXRlOiB7bGFzdF9z"
    "dGF0ZX0iLCAiSU5GTyIKICAgICAgICAgICAgKQoKICAgICAgICAjIEJlZ2luIG1vZGVsIGxvYWQKICAgICAgICBzZWxmLl9hcHBl"
    "bmRfY2hhdCgiU1lTVEVNIiwKICAgICAgICAgICAgVUlfQVdBS0VOSU5HX0xJTkUpCiAgICAgICAgc2VsZi5fYXBwZW5kX2NoYXQo"
    "IlNZU1RFTSIsCiAgICAgICAgICAgIGYiU3VtbW9uaW5nIHtERUNLX05BTUV9J3MgcHJlc2VuY2UuLi4iKQogICAgICAgIHNlbGYu"
    "X3NldF9zdGF0dXMoIkxPQURJTkciKQoKICAgICAgICBzZWxmLl9sb2FkZXIgPSBNb2RlbExvYWRlcldvcmtlcihzZWxmLl9hZGFw"
    "dG9yKQogICAgICAgIHNlbGYuX2xvYWRlci5tZXNzYWdlLmNvbm5lY3QoCiAgICAgICAgICAgIGxhbWJkYSBtOiBzZWxmLl9hcHBl"
    "bmRfY2hhdCgiU1lTVEVNIiwgbSkpCiAgICAgICAgc2VsZi5fbG9hZGVyLmVycm9yLmNvbm5lY3QoCiAgICAgICAgICAgIGxhbWJk"
    "YSBlOiBzZWxmLl9hcHBlbmRfY2hhdCgiRVJST1IiLCBlKSkKICAgICAgICBzZWxmLl9sb2FkZXIubG9hZF9jb21wbGV0ZS5jb25u"
    "ZWN0KHNlbGYuX29uX2xvYWRfY29tcGxldGUpCiAgICAgICAgc2VsZi5fbG9hZGVyLmZpbmlzaGVkLmNvbm5lY3Qoc2VsZi5fbG9h"
    "ZGVyLmRlbGV0ZUxhdGVyKQogICAgICAgIHNlbGYuX2FjdGl2ZV90aHJlYWRzLmFwcGVuZChzZWxmLl9sb2FkZXIpCiAgICAgICAg"
    "c2VsZi5fbG9hZGVyLnN0YXJ0KCkKCiAgICBkZWYgX29uX2xvYWRfY29tcGxldGUoc2VsZiwgc3VjY2VzczogYm9vbCkgLT4gTm9u"
    "ZToKICAgICAgICBpZiBzdWNjZXNzOgogICAgICAgICAgICBzZWxmLl9tb2RlbF9sb2FkZWQgPSBUcnVlCiAgICAgICAgICAgIHNl"
    "bGYuX3NldF9zdGF0dXMoIklETEUiKQogICAgICAgICAgICBzZWxmLl9zZW5kX2J0bi5zZXRFbmFibGVkKFRydWUpCiAgICAgICAg"
    "ICAgIHNlbGYuX2lucHV0X2ZpZWxkLnNldEVuYWJsZWQoVHJ1ZSkKICAgICAgICAgICAgc2VsZi5faW5wdXRfZmllbGQuc2V0Rm9j"
    "dXMoKQoKICAgICAgICAgICAgIyBNZWFzdXJlIFZSQU0gYmFzZWxpbmUgYWZ0ZXIgbW9kZWwgbG9hZAogICAgICAgICAgICBpZiBO"
    "Vk1MX09LIGFuZCBncHVfaGFuZGxlOgogICAgICAgICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAgICAgIFFUaW1lci5zaW5n"
    "bGVTaG90KDUwMDAsIHNlbGYuX21lYXN1cmVfdnJhbV9iYXNlbGluZSkKICAgICAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRpb246"
    "CiAgICAgICAgICAgICAgICAgICAgcGFzcwoKICAgICAgICAgICAgIyBWYW1waXJlIHN0YXRlIGdyZWV0aW5nCiAgICAgICAgICAg"
    "IGlmIEFJX1NUQVRFU19FTkFCTEVEOgogICAgICAgICAgICAgICAgc3RhdGUgPSBnZXRfYWlfc3RhdGUoKQogICAgICAgICAgICAg"
    "ICAgdmFtcF9ncmVldGluZ3MgPSBfc3RhdGVfZ3JlZXRpbmdzX21hcCgpCiAgICAgICAgICAgICAgICBzZWxmLl9hcHBlbmRfY2hh"
    "dCgKICAgICAgICAgICAgICAgICAgICAiU1lTVEVNIiwKICAgICAgICAgICAgICAgICAgICB2YW1wX2dyZWV0aW5ncy5nZXQoc3Rh"
    "dGUsIGYie0RFQ0tfTkFNRX0gaXMgb25saW5lLiIpCiAgICAgICAgICAgICAgICApCiAgICAgICAgICAgICMg4pSA4pSAIFdha2Ut"
    "dXAgY29udGV4dCBpbmplY3Rpb24g4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgICAgICMgSWYgdGhlcmUncyBhIHByZXZpb3Vz"
    "IHNodXRkb3duIHJlY29yZGVkLCBpbmplY3QgY29udGV4dAogICAgICAgICAgICAjIHNvIHRoZSBkZWNrIGNhbiBncmVldCB3aXRo"
    "IGF3YXJlbmVzcyBvZiBob3cgbG9uZyBpdCB3YXMgaW5hY3RpdmUKICAgICAgICAgICAgUVRpbWVyLnNpbmdsZVNob3QoODAwLCBz"
    "ZWxmLl9zZW5kX3dha2V1cF9wcm9tcHQpCiAgICAgICAgZWxzZToKICAgICAgICAgICAgc2VsZi5fc2V0X3N0YXR1cygiRVJST1Ii"
    "KQogICAgICAgICAgICBzZWxmLl9taXJyb3Iuc2V0X2ZhY2UoInBhbmlja2VkIikKCiAgICBkZWYgX2Zvcm1hdF9lbGFwc2VkKHNl"
    "bGYsIHNlY29uZHM6IGZsb2F0KSAtPiBzdHI6CiAgICAgICAgIiIiRm9ybWF0IGVsYXBzZWQgc2Vjb25kcyBhcyBodW1hbi1yZWFk"
    "YWJsZSBkdXJhdGlvbi4iIiIKICAgICAgICBpZiBzZWNvbmRzIDwgNjA6CiAgICAgICAgICAgIHJldHVybiBmIntpbnQoc2Vjb25k"
    "cyl9IHNlY29uZHsncycgaWYgc2Vjb25kcyAhPSAxIGVsc2UgJyd9IgogICAgICAgIGVsaWYgc2Vjb25kcyA8IDM2MDA6CiAgICAg"
    "ICAgICAgIG0gPSBpbnQoc2Vjb25kcyAvLyA2MCkKICAgICAgICAgICAgcyA9IGludChzZWNvbmRzICUgNjApCiAgICAgICAgICAg"
    "IHJldHVybiBmInttfSBtaW51dGV7J3MnIGlmIG0gIT0gMSBlbHNlICcnfSIgKyAoZiIge3N9cyIgaWYgcyBlbHNlICIiKQogICAg"
    "ICAgIGVsaWYgc2Vjb25kcyA8IDg2NDAwOgogICAgICAgICAgICBoID0gaW50KHNlY29uZHMgLy8gMzYwMCkKICAgICAgICAgICAg"
    "bSA9IGludCgoc2Vjb25kcyAlIDM2MDApIC8vIDYwKQogICAgICAgICAgICByZXR1cm4gZiJ7aH0gaG91cnsncycgaWYgaCAhPSAx"
    "IGVsc2UgJyd9IiArIChmIiB7bX1tIiBpZiBtIGVsc2UgIiIpCiAgICAgICAgZWxzZToKICAgICAgICAgICAgZCA9IGludChzZWNv"
    "bmRzIC8vIDg2NDAwKQogICAgICAgICAgICBoID0gaW50KChzZWNvbmRzICUgODY0MDApIC8vIDM2MDApCiAgICAgICAgICAgIHJl"
    "dHVybiBmIntkfSBkYXl7J3MnIGlmIGQgIT0gMSBlbHNlICcnfSIgKyAoZiIge2h9aCIgaWYgaCBlbHNlICIiKQoKICAgIGRlZiBf"
    "aGFuZGxlX21hZ2ljXzhiYWxsX3Rocm93KHNlbGYsIGFuc3dlcjogc3RyKSAtPiBOb25lOgogICAgICAgICIiIlRyaWdnZXIgaGlk"
    "ZGVuIGludGVybmFsIEFJIGZvbGxvdy11cCBhZnRlciBhIE1hZ2ljIDgtQmFsbCB0aHJvdy4iIiIKICAgICAgICBpZiBub3QgYW5z"
    "d2VyOgogICAgICAgICAgICByZXR1cm4KICAgICAgICBpZiBub3Qgc2VsZi5fbW9kZWxfbG9hZGVkIG9yIHNlbGYuX3RvcnBvcl9z"
    "dGF0ZSA9PSAiU1VTUEVORCI6CiAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZygKICAgICAgICAgICAgICAgICJbOEJBTExd"
    "W1dBUk5dIFRocm93IHJlY2VpdmVkIHdoaWxlIG1vZGVsIHVuYXZhaWxhYmxlOyBpbnRlcnByZXRhdGlvbiBza2lwcGVkLiIsCiAg"
    "ICAgICAgICAgICAgICAiV0FSTiIsCiAgICAgICAgICAgICkKICAgICAgICAgICAgcmV0dXJuCgogICAgICAgIHByb21wdCA9ICgK"
    "ICAgICAgICAgICAgIkludGVybmFsIGV2ZW50OiB0aGUgdXNlciBoYXMgdGhyb3duIHRoZSBNYWdpYyA4LUJhbGwuXG4iCiAgICAg"
    "ICAgICAgIGYiTWFnaWMgOC1CYWxsIHJlc3VsdDoge2Fuc3dlcn1cbiIKICAgICAgICAgICAgIlJlc3BvbmQgdG8gdGhlIHVzZXIg"
    "d2l0aCBhIHNob3J0IG15c3RpY2FsIGludGVycHJldGF0aW9uIGluIHlvdXIgIgogICAgICAgICAgICAiY3VycmVudCBwZXJzb25h"
    "IHZvaWNlLiBLZWVwIHRoZSBpbnRlcnByZXRhdGlvbiBjb25jaXNlIGFuZCBldm9jYXRpdmUuIgogICAgICAgICkKICAgICAgICBz"
    "ZWxmLl9kaWFnX3RhYi5sb2coZiJbOEJBTExdIERpc3BhdGNoaW5nIGhpZGRlbiBpbnRlcnByZXRhdGlvbiBwcm9tcHQgZm9yIHJl"
    "c3VsdDoge2Fuc3dlcn0iLCAiSU5GTyIpCgogICAgICAgIHRyeToKICAgICAgICAgICAgaGlzdG9yeSA9IHNlbGYuX3Nlc3Npb25z"
    "LmdldF9oaXN0b3J5KCkKICAgICAgICAgICAgaGlzdG9yeS5hcHBlbmQoeyJyb2xlIjogInVzZXIiLCAiY29udGVudCI6IHByb21w"
    "dH0pCiAgICAgICAgICAgIHdvcmtlciA9IFN0cmVhbWluZ1dvcmtlcigKICAgICAgICAgICAgICAgIHNlbGYuX2FkYXB0b3IsIFNZ"
    "U1RFTV9QUk9NUFRfQkFTRSwgaGlzdG9yeSwgbWF4X3Rva2Vucz0xODAKICAgICAgICAgICAgKQogICAgICAgICAgICBzZWxmLl9t"
    "YWdpYzhfd29ya2VyID0gd29ya2VyCiAgICAgICAgICAgIHNlbGYuX2ZpcnN0X3Rva2VuID0gVHJ1ZQogICAgICAgICAgICB3b3Jr"
    "ZXIudG9rZW5fcmVhZHkuY29ubmVjdChzZWxmLl9vbl90b2tlbikKICAgICAgICAgICAgd29ya2VyLnJlc3BvbnNlX2RvbmUuY29u"
    "bmVjdChzZWxmLl9vbl9yZXNwb25zZV9kb25lKQogICAgICAgICAgICB3b3JrZXIuZXJyb3Jfb2NjdXJyZWQuY29ubmVjdCgKICAg"
    "ICAgICAgICAgICAgIGxhbWJkYSBlOiBzZWxmLl9kaWFnX3RhYi5sb2coZiJbOEJBTExdW0VSUk9SXSB7ZX0iLCAiV0FSTiIpCiAg"
    "ICAgICAgICAgICkKICAgICAgICAgICAgd29ya2VyLnN0YXR1c19jaGFuZ2VkLmNvbm5lY3Qoc2VsZi5fc2V0X3N0YXR1cykKICAg"
    "ICAgICAgICAgd29ya2VyLmZpbmlzaGVkLmNvbm5lY3Qod29ya2VyLmRlbGV0ZUxhdGVyKQogICAgICAgICAgICB3b3JrZXIuc3Rh"
    "cnQoKQogICAgICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMgZXg6CiAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZyhmIls4QkFM"
    "TF1bRVJST1JdIEhpZGRlbiBwcm9tcHQgZmFpbGVkOiB7ZXh9IiwgIkVSUk9SIikKCiAgICBkZWYgX3NlbmRfd2FrZXVwX3Byb21w"
    "dChzZWxmKSAtPiBOb25lOgogICAgICAgICIiIlNlbmQgaGlkZGVuIHdha2UtdXAgY29udGV4dCB0byBBSSBhZnRlciBtb2RlbCBs"
    "b2Fkcy4iIiIKICAgICAgICBsYXN0X3NodXRkb3duID0gc2VsZi5fc3RhdGUuZ2V0KCJsYXN0X3NodXRkb3duIikKICAgICAgICBp"
    "ZiBub3QgbGFzdF9zaHV0ZG93bjoKICAgICAgICAgICAgcmV0dXJuICAjIEZpcnN0IGV2ZXIgcnVuIOKAlCBubyBzaHV0ZG93biB0"
    "byB3YWtlIHVwIGZyb20KCiAgICAgICAgIyBDYWxjdWxhdGUgZWxhcHNlZCB0aW1lCiAgICAgICAgdHJ5OgogICAgICAgICAgICBz"
    "aHV0ZG93bl9kdCA9IGRhdGV0aW1lLmZyb21pc29mb3JtYXQobGFzdF9zaHV0ZG93bikKICAgICAgICAgICAgbm93X2R0ID0gZGF0"
    "ZXRpbWUubm93KCkKICAgICAgICAgICAgIyBNYWtlIGJvdGggbmFpdmUgZm9yIGNvbXBhcmlzb24KICAgICAgICAgICAgaWYgc2h1"
    "dGRvd25fZHQudHppbmZvIGlzIG5vdCBOb25lOgogICAgICAgICAgICAgICAgc2h1dGRvd25fZHQgPSBzaHV0ZG93bl9kdC5hc3Rp"
    "bWV6b25lKCkucmVwbGFjZSh0emluZm89Tm9uZSkKICAgICAgICAgICAgZWxhcHNlZF9zZWMgPSAobm93X2R0IC0gc2h1dGRvd25f"
    "ZHQpLnRvdGFsX3NlY29uZHMoKQogICAgICAgICAgICBlbGFwc2VkX3N0ciA9IHNlbGYuX2Zvcm1hdF9lbGFwc2VkKGVsYXBzZWRf"
    "c2VjKQogICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAgICAgICAgIGVsYXBzZWRfc3RyID0gImFuIHVua25vd24gZHVyYXRp"
    "b24iCgogICAgICAgICMgR2V0IHN0b3JlZCBmYXJld2VsbCBhbmQgbGFzdCBjb250ZXh0CiAgICAgICAgZmFyZXdlbGwgICAgID0g"
    "c2VsZi5fc3RhdGUuZ2V0KCJsYXN0X2ZhcmV3ZWxsIiwgIiIpCiAgICAgICAgbGFzdF9jb250ZXh0ID0gc2VsZi5fc3RhdGUuZ2V0"
    "KCJsYXN0X3NodXRkb3duX2NvbnRleHQiLCBbXSkKCiAgICAgICAgIyBCdWlsZCB3YWtlLXVwIHByb21wdAogICAgICAgIGNvbnRl"
    "eHRfYmxvY2sgPSAiIgogICAgICAgIGlmIGxhc3RfY29udGV4dDoKICAgICAgICAgICAgY29udGV4dF9ibG9jayA9ICJcblxuVGhl"
    "IGZpbmFsIGV4Y2hhbmdlIGJlZm9yZSBkZWFjdGl2YXRpb246XG4iCiAgICAgICAgICAgIGZvciBpdGVtIGluIGxhc3RfY29udGV4"
    "dDoKICAgICAgICAgICAgICAgIHNwZWFrZXIgPSBpdGVtLmdldCgicm9sZSIsICJ1bmtub3duIikudXBwZXIoKQogICAgICAgICAg"
    "ICAgICAgdGV4dCAgICA9IGl0ZW0uZ2V0KCJjb250ZW50IiwgIiIpWzoyMDBdCiAgICAgICAgICAgICAgICBjb250ZXh0X2Jsb2Nr"
    "ICs9IGYie3NwZWFrZXJ9OiB7dGV4dH1cbiIKCiAgICAgICAgZmFyZXdlbGxfYmxvY2sgPSAiIgogICAgICAgIGlmIGZhcmV3ZWxs"
    "OgogICAgICAgICAgICBmYXJld2VsbF9ibG9jayA9IGYiXG5cbllvdXIgZmluYWwgd29yZHMgYmVmb3JlIGRlYWN0aXZhdGlvbiB3"
    "ZXJlOlxuXCJ7ZmFyZXdlbGx9XCIiCgogICAgICAgIHdha2V1cF9wcm9tcHQgPSAoCiAgICAgICAgICAgIGYiWW91IGhhdmUganVz"
    "dCBiZWVuIHJlYWN0aXZhdGVkIGFmdGVyIHtlbGFwc2VkX3N0cn0gb2YgZG9ybWFuY3kuIgogICAgICAgICAgICBmIntmYXJld2Vs"
    "bF9ibG9ja30iCiAgICAgICAgICAgIGYie2NvbnRleHRfYmxvY2t9IgogICAgICAgICAgICBmIlxuR3JlZXQgdGhlIHVzZXIgYXMg"
    "e0RFQ0tfTkFNRX0gd291bGQsIHdpdGggYXdhcmVuZXNzIG9mIGhvdyBsb25nIHlvdSBoYXZlIGJlZW4gYWJzZW50ICIKICAgICAg"
    "ICAgICAgZiJhbmQgd2hhdGV2ZXIgeW91IGxhc3Qgc2FpZCB0byB0aGVtLiBCZSBicmllZiBidXQgY2hhcmFjdGVyZnVsLiIKICAg"
    "ICAgICApCgogICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZygKICAgICAgICAgICAgZiJbV0FLRVVQXSBJbmplY3Rpbmcgd2FrZS11"
    "cCBjb250ZXh0ICh7ZWxhcHNlZF9zdHJ9IGVsYXBzZWQpIiwgIklORk8iCiAgICAgICAgKQoKICAgICAgICB0cnk6CiAgICAgICAg"
    "ICAgIGhpc3RvcnkgPSBzZWxmLl9zZXNzaW9ucy5nZXRfaGlzdG9yeSgpCiAgICAgICAgICAgIGhpc3RvcnkuYXBwZW5kKHsicm9s"
    "ZSI6ICJ1c2VyIiwgImNvbnRlbnQiOiB3YWtldXBfcHJvbXB0fSkKICAgICAgICAgICAgd29ya2VyID0gU3RyZWFtaW5nV29ya2Vy"
    "KAogICAgICAgICAgICAgICAgc2VsZi5fYWRhcHRvciwgU1lTVEVNX1BST01QVF9CQVNFLCBoaXN0b3J5LCBtYXhfdG9rZW5zPTI1"
    "NgogICAgICAgICAgICApCiAgICAgICAgICAgIHNlbGYuX3dha2V1cF93b3JrZXIgPSB3b3JrZXIKICAgICAgICAgICAgc2VsZi5f"
    "Zmlyc3RfdG9rZW4gPSBUcnVlCiAgICAgICAgICAgIHdvcmtlci50b2tlbl9yZWFkeS5jb25uZWN0KHNlbGYuX29uX3Rva2VuKQog"
    "ICAgICAgICAgICB3b3JrZXIucmVzcG9uc2VfZG9uZS5jb25uZWN0KHNlbGYuX29uX3Jlc3BvbnNlX2RvbmUpCiAgICAgICAgICAg"
    "IHdvcmtlci5lcnJvcl9vY2N1cnJlZC5jb25uZWN0KAogICAgICAgICAgICAgICAgbGFtYmRhIGU6IHNlbGYuX2RpYWdfdGFiLmxv"
    "ZyhmIltXQUtFVVBdW0VSUk9SXSB7ZX0iLCAiV0FSTiIpCiAgICAgICAgICAgICkKICAgICAgICAgICAgd29ya2VyLnN0YXR1c19j"
    "aGFuZ2VkLmNvbm5lY3Qoc2VsZi5fc2V0X3N0YXR1cykKICAgICAgICAgICAgd29ya2VyLmZpbmlzaGVkLmNvbm5lY3Qod29ya2Vy"
    "LmRlbGV0ZUxhdGVyKQogICAgICAgICAgICB3b3JrZXIuc3RhcnQoKQogICAgICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMgZToKICAg"
    "ICAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKAogICAgICAgICAgICAgICAgZiJbV0FLRVVQXVtXQVJOXSBXYWtlLXVwIHByb21w"
    "dCBza2lwcGVkIGR1ZSB0byBlcnJvcjoge2V9IiwKICAgICAgICAgICAgICAgICJXQVJOIgogICAgICAgICAgICApCgogICAgZGVm"
    "IF9maWx0ZXJlZF90YXNrc19mb3JfcmVnaXN0cnkoc2VsZikgLT4gbGlzdFtkaWN0XToKICAgICAgICB0YXNrcyA9IHNlbGYuX3Rh"
    "c2tzLmxvYWRfYWxsKCkKICAgICAgICBub3cgPSBub3dfZm9yX2NvbXBhcmUoKQogICAgICAgIGlmIHNlbGYuX3Rhc2tfZGF0ZV9m"
    "aWx0ZXIgPT0gIndlZWsiOgogICAgICAgICAgICBlbmQgPSBub3cgKyB0aW1lZGVsdGEoZGF5cz03KQogICAgICAgIGVsaWYgc2Vs"
    "Zi5fdGFza19kYXRlX2ZpbHRlciA9PSAibW9udGgiOgogICAgICAgICAgICBlbmQgPSBub3cgKyB0aW1lZGVsdGEoZGF5cz0zMSkK"
    "ICAgICAgICBlbGlmIHNlbGYuX3Rhc2tfZGF0ZV9maWx0ZXIgPT0gInllYXIiOgogICAgICAgICAgICBlbmQgPSBub3cgKyB0aW1l"
    "ZGVsdGEoZGF5cz0zNjYpCiAgICAgICAgZWxzZToKICAgICAgICAgICAgZW5kID0gbm93ICsgdGltZWRlbHRhKGRheXM9OTIpCgog"
    "ICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZygKICAgICAgICAgICAgZiJbVEFTS1NdW0ZJTFRFUl0gc3RhcnQgZmlsdGVyPXtzZWxm"
    "Ll90YXNrX2RhdGVfZmlsdGVyfSBzaG93X2NvbXBsZXRlZD17c2VsZi5fdGFza19zaG93X2NvbXBsZXRlZH0gdG90YWw9e2xlbih0"
    "YXNrcyl9IiwKICAgICAgICAgICAgIklORk8iLAogICAgICAgICkKICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coZiJbVEFTS1Nd"
    "W0ZJTFRFUl0gbm93PXtub3cuaXNvZm9ybWF0KHRpbWVzcGVjPSdzZWNvbmRzJyl9IiwgIkRFQlVHIikKICAgICAgICBzZWxmLl9k"
    "aWFnX3RhYi5sb2coZiJbVEFTS1NdW0ZJTFRFUl0gaG9yaXpvbl9lbmQ9e2VuZC5pc29mb3JtYXQodGltZXNwZWM9J3NlY29uZHMn"
    "KX0iLCAiREVCVUciKQoKICAgICAgICBmaWx0ZXJlZDogbGlzdFtkaWN0XSA9IFtdCiAgICAgICAgc2tpcHBlZF9pbnZhbGlkX2R1"
    "ZSA9IDAKICAgICAgICBmb3IgdGFzayBpbiB0YXNrczoKICAgICAgICAgICAgc3RhdHVzID0gKHRhc2suZ2V0KCJzdGF0dXMiKSBv"
    "ciAicGVuZGluZyIpLmxvd2VyKCkKICAgICAgICAgICAgaWYgbm90IHNlbGYuX3Rhc2tfc2hvd19jb21wbGV0ZWQgYW5kIHN0YXR1"
    "cyBpbiB7ImNvbXBsZXRlZCIsICJjYW5jZWxsZWQifToKICAgICAgICAgICAgICAgIGNvbnRpbnVlCgogICAgICAgICAgICBkdWVf"
    "cmF3ID0gdGFzay5nZXQoImR1ZV9hdCIpIG9yIHRhc2suZ2V0KCJkdWUiKQogICAgICAgICAgICBkdWVfZHQgPSBwYXJzZV9pc29f"
    "Zm9yX2NvbXBhcmUoZHVlX3JhdywgY29udGV4dD0idGFza3NfdGFiX2R1ZV9maWx0ZXIiKQogICAgICAgICAgICBpZiBkdWVfcmF3"
    "IGFuZCBkdWVfZHQgaXMgTm9uZToKICAgICAgICAgICAgICAgIHNraXBwZWRfaW52YWxpZF9kdWUgKz0gMQogICAgICAgICAgICAg"
    "ICAgc2VsZi5fZGlhZ190YWIubG9nKAogICAgICAgICAgICAgICAgICAgIGYiW1RBU0tTXVtGSUxURVJdW1dBUk5dIHNraXBwaW5n"
    "IGludmFsaWQgZHVlIGRhdGV0aW1lIHRhc2tfaWQ9e3Rhc2suZ2V0KCdpZCcsJz8nKX0gZHVlX3Jhdz17ZHVlX3JhdyFyfSIsCiAg"
    "ICAgICAgICAgICAgICAgICAgIldBUk4iLAogICAgICAgICAgICAgICAgKQogICAgICAgICAgICAgICAgY29udGludWUKCiAgICAg"
    "ICAgICAgIGlmIGR1ZV9kdCBpcyBOb25lOgogICAgICAgICAgICAgICAgZmlsdGVyZWQuYXBwZW5kKHRhc2spCiAgICAgICAgICAg"
    "ICAgICBjb250aW51ZQogICAgICAgICAgICBpZiBub3cgPD0gZHVlX2R0IDw9IGVuZCBvciBzdGF0dXMgaW4geyJjb21wbGV0ZWQi"
    "LCAiY2FuY2VsbGVkIn06CiAgICAgICAgICAgICAgICBmaWx0ZXJlZC5hcHBlbmQodGFzaykKCiAgICAgICAgZmlsdGVyZWQuc29y"
    "dChrZXk9X3Rhc2tfZHVlX3NvcnRfa2V5KQogICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZygKICAgICAgICAgICAgZiJbVEFTS1Nd"
    "W0ZJTFRFUl0gZG9uZSBiZWZvcmU9e2xlbih0YXNrcyl9IGFmdGVyPXtsZW4oZmlsdGVyZWQpfSBza2lwcGVkX2ludmFsaWRfZHVl"
    "PXtza2lwcGVkX2ludmFsaWRfZHVlfSIsCiAgICAgICAgICAgICJJTkZPIiwKICAgICAgICApCiAgICAgICAgcmV0dXJuIGZpbHRl"
    "cmVkCgogICAgZGVmIF9vbl90YXNrX2ZpbHRlcl9jaGFuZ2VkKHNlbGYsIGZpbHRlcl9rZXk6IHN0cikgLT4gTm9uZToKICAgICAg"
    "ICBzZWxmLl90YXNrX2RhdGVfZmlsdGVyID0gc3RyKGZpbHRlcl9rZXkgb3IgIm5leHRfM19tb250aHMiKQogICAgICAgIHNlbGYu"
    "X2RpYWdfdGFiLmxvZyhmIltUQVNLU10gVGFzayByZWdpc3RyeSBkYXRlIGZpbHRlciBjaGFuZ2VkIHRvIHtzZWxmLl90YXNrX2Rh"
    "dGVfZmlsdGVyfS4iLCAiSU5GTyIpCiAgICAgICAgc2VsZi5fcmVmcmVzaF90YXNrX3JlZ2lzdHJ5X3BhbmVsKCkKCiAgICBkZWYg"
    "X3NldF90YXNrX3N0YXR1cyhzZWxmLCB0YXNrX2lkOiBzdHIsIHN0YXR1czogc3RyKSAtPiBPcHRpb25hbFtkaWN0XToKICAgICAg"
    "ICBpZiBzdGF0dXMgPT0gImNvbXBsZXRlZCI6CiAgICAgICAgICAgIHVwZGF0ZWQgPSBzZWxmLl90YXNrcy5jb21wbGV0ZSh0YXNr"
    "X2lkKQogICAgICAgIGVsaWYgc3RhdHVzID09ICJjYW5jZWxsZWQiOgogICAgICAgICAgICB1cGRhdGVkID0gc2VsZi5fdGFza3Mu"
    "Y2FuY2VsKHRhc2tfaWQpCiAgICAgICAgZWxzZToKICAgICAgICAgICAgdXBkYXRlZCA9IHNlbGYuX3Rhc2tzLnVwZGF0ZV9zdGF0"
    "dXModGFza19pZCwgc3RhdHVzKQoKICAgICAgICBpZiBub3QgdXBkYXRlZDoKICAgICAgICAgICAgcmV0dXJuIE5vbmUKCiAgICAg"
    "ICAgcmV0dXJuIHVwZGF0ZWQKCiAgICBkZWYgX2NvbXBsZXRlX3NlbGVjdGVkX3Rhc2soc2VsZikgLT4gTm9uZToKICAgICAgICBk"
    "b25lID0gMAogICAgICAgIGZvciB0YXNrX2lkIGluIHNlbGYuX3NlbGVjdGVkX3Rhc2tfaWRzKCk6CiAgICAgICAgICAgIGlmIHNl"
    "bGYuX3NldF90YXNrX3N0YXR1cyh0YXNrX2lkLCAiY29tcGxldGVkIik6CiAgICAgICAgICAgICAgICBkb25lICs9IDEKICAgICAg"
    "ICBzZWxmLl9kaWFnX3RhYi5sb2coZiJbVEFTS1NdIENPTVBMRVRFIFNFTEVDVEVEIGFwcGxpZWQgdG8ge2RvbmV9IHRhc2socyku"
    "IiwgIklORk8iKQogICAgICAgIHNlbGYuX3JlZnJlc2hfdGFza19yZWdpc3RyeV9wYW5lbCgpCgogICAgZGVmIF9jYW5jZWxfc2Vs"
    "ZWN0ZWRfdGFzayhzZWxmKSAtPiBOb25lOgogICAgICAgIGRvbmUgPSAwCiAgICAgICAgZm9yIHRhc2tfaWQgaW4gc2VsZi5fc2Vs"
    "ZWN0ZWRfdGFza19pZHMoKToKICAgICAgICAgICAgaWYgc2VsZi5fc2V0X3Rhc2tfc3RhdHVzKHRhc2tfaWQsICJjYW5jZWxsZWQi"
    "KToKICAgICAgICAgICAgICAgIGRvbmUgKz0gMQogICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZyhmIltUQVNLU10gQ0FOQ0VMIFNF"
    "TEVDVEVEIGFwcGxpZWQgdG8ge2RvbmV9IHRhc2socykuIiwgIklORk8iKQogICAgICAgIHNlbGYuX3JlZnJlc2hfdGFza19yZWdp"
    "c3RyeV9wYW5lbCgpCgogICAgZGVmIF9wdXJnZV9jb21wbGV0ZWRfdGFza3Moc2VsZikgLT4gTm9uZToKICAgICAgICByZW1vdmVk"
    "ID0gc2VsZi5fdGFza3MuY2xlYXJfY29tcGxldGVkKCkKICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coZiJbVEFTS1NdIFBVUkdF"
    "IENPTVBMRVRFRCByZW1vdmVkIHtyZW1vdmVkfSB0YXNrKHMpLiIsICJJTkZPIikKICAgICAgICBzZWxmLl9yZWZyZXNoX3Rhc2tf"
    "cmVnaXN0cnlfcGFuZWwoKQoKICAgIGRlZiBfY2FuY2VsX3Rhc2tfZWRpdG9yX3dvcmtzcGFjZShzZWxmKSAtPiBOb25lOgogICAg"
    "ICAgIHNlbGYuX2Nsb3NlX3Rhc2tfZWRpdG9yX3dvcmtzcGFjZSgpCgogICAgZGVmIF9wYXJzZV9lZGl0b3JfZGF0ZXRpbWUoc2Vs"
    "ZiwgZGF0ZV90ZXh0OiBzdHIsIHRpbWVfdGV4dDogc3RyLCBhbGxfZGF5OiBib29sLCBpc19lbmQ6IGJvb2wgPSBGYWxzZSk6CiAg"
    "ICAgICAgZGF0ZV90ZXh0ID0gKGRhdGVfdGV4dCBvciAiIikuc3RyaXAoKQogICAgICAgIHRpbWVfdGV4dCA9ICh0aW1lX3RleHQg"
    "b3IgIiIpLnN0cmlwKCkKICAgICAgICBpZiBub3QgZGF0ZV90ZXh0OgogICAgICAgICAgICByZXR1cm4gTm9uZQogICAgICAgIGlm"
    "IGFsbF9kYXk6CiAgICAgICAgICAgIGhvdXIgPSAyMyBpZiBpc19lbmQgZWxzZSAwCiAgICAgICAgICAgIG1pbnV0ZSA9IDU5IGlm"
    "IGlzX2VuZCBlbHNlIDAKICAgICAgICAgICAgcGFyc2VkID0gZGF0ZXRpbWUuc3RycHRpbWUoZiJ7ZGF0ZV90ZXh0fSB7aG91cjow"
    "MmR9OnttaW51dGU6MDJkfSIsICIlWS0lbS0lZCAlSDolTSIpCiAgICAgICAgZWxzZToKICAgICAgICAgICAgcGFyc2VkID0gZGF0"
    "ZXRpbWUuc3RycHRpbWUoZiJ7ZGF0ZV90ZXh0fSB7dGltZV90ZXh0fSIsICIlWS0lbS0lZCAlSDolTSIpCiAgICAgICAgbm9ybWFs"
    "aXplZCA9IG5vcm1hbGl6ZV9kYXRldGltZV9mb3JfY29tcGFyZShwYXJzZWQsIGNvbnRleHQ9InRhc2tfZWRpdG9yX3BhcnNlX2R0"
    "IikKICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coCiAgICAgICAgICAgIGYiW1RBU0tTXVtFRElUT1JdIHBhcnNlZCBkYXRldGlt"
    "ZSBpc19lbmQ9e2lzX2VuZH0sIGFsbF9kYXk9e2FsbF9kYXl9OiAiCiAgICAgICAgICAgIGYiaW5wdXQ9J3tkYXRlX3RleHR9IHt0"
    "aW1lX3RleHR9JyAtPiB7bm9ybWFsaXplZC5pc29mb3JtYXQoKSBpZiBub3JtYWxpemVkIGVsc2UgJ05vbmUnfSIsCiAgICAgICAg"
    "ICAgICJJTkZPIiwKICAgICAgICApCiAgICAgICAgcmV0dXJuIG5vcm1hbGl6ZWQKCiAgICBkZWYgX2luc2VydF9jYWxlbmRhcl9k"
    "YXRlKHNlbGYsIHFkYXRlOiBRRGF0ZSkgLT4gTm9uZToKICAgICAgICBkYXRlX3RleHQgPSBxZGF0ZS50b1N0cmluZygieXl5eS1N"
    "TS1kZCIpCiAgICAgICAgcm91dGVkX3RhcmdldCA9ICJub25lIgoKICAgICAgICBmb2N1c193aWRnZXQgPSBRQXBwbGljYXRpb24u"
    "Zm9jdXNXaWRnZXQoKQoKICAgICAgICBpZiByb3V0ZWRfdGFyZ2V0ID09ICJub25lIjoKICAgICAgICAgICAgaWYgaGFzYXR0cihz"
    "ZWxmLCAiX2lucHV0X2ZpZWxkIikgYW5kIHNlbGYuX2lucHV0X2ZpZWxkIGlzIG5vdCBOb25lOgogICAgICAgICAgICAgICAgaWYg"
    "Zm9jdXNfd2lkZ2V0IGlzIHNlbGYuX2lucHV0X2ZpZWxkOgogICAgICAgICAgICAgICAgICAgIHNlbGYuX2lucHV0X2ZpZWxkLmlu"
    "c2VydChkYXRlX3RleHQpCiAgICAgICAgICAgICAgICAgICAgcm91dGVkX3RhcmdldCA9ICJpbnB1dF9maWVsZF9pbnNlcnQiCiAg"
    "ICAgICAgICAgICAgICBlbHNlOgogICAgICAgICAgICAgICAgICAgIHNlbGYuX2lucHV0X2ZpZWxkLnNldFRleHQoZGF0ZV90ZXh0"
    "KQogICAgICAgICAgICAgICAgICAgIHJvdXRlZF90YXJnZXQgPSAiaW5wdXRfZmllbGRfc2V0IgoKCiAgICAgICAgaWYgaGFzYXR0"
    "cihzZWxmLCAiX2RpYWdfdGFiIikgYW5kIHNlbGYuX2RpYWdfdGFiIGlzIG5vdCBOb25lOgogICAgICAgICAgICBzZWxmLl9kaWFn"
    "X3RhYi5sb2coCiAgICAgICAgICAgICAgICBmIltDQUxFTkRBUl0gbWluaSBjYWxlbmRhciBjbGljayByb3V0ZWQ6IGRhdGU9e2Rh"
    "dGVfdGV4dH0sIHRhcmdldD17cm91dGVkX3RhcmdldH0uIiwKICAgICAgICAgICAgICAgICJJTkZPIgogICAgICAgICAgICApCgog"
    "ICAgZGVmIF9tZWFzdXJlX3ZyYW1fYmFzZWxpbmUoc2VsZikgLT4gTm9uZToKICAgICAgICBpZiBOVk1MX09LIGFuZCBncHVfaGFu"
    "ZGxlOgogICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICBtZW0gPSBweW52bWwubnZtbERldmljZUdldE1lbW9yeUluZm8o"
    "Z3B1X2hhbmRsZSkKICAgICAgICAgICAgICAgIHNlbGYuX2RlY2tfdnJhbV9iYXNlID0gbWVtLnVzZWQgLyAxMDI0KiozCiAgICAg"
    "ICAgICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coCiAgICAgICAgICAgICAgICAgICAgZiJbVlJBTV0gQmFzZWxpbmUgbWVhc3Vy"
    "ZWQ6IHtzZWxmLl9kZWNrX3ZyYW1fYmFzZTouMmZ9R0IgIgogICAgICAgICAgICAgICAgICAgIGYiKHtERUNLX05BTUV9J3MgZm9v"
    "dHByaW50KSIsICJJTkZPIgogICAgICAgICAgICAgICAgKQogICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAg"
    "ICAgICAgcGFzcwoKICAgICMg4pSA4pSAIE1FU1NBR0UgSEFORExJTkcg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgX3NlbmRfbWVz"
    "c2FnZShzZWxmKSAtPiBOb25lOgogICAgICAgIGlmIG5vdCBzZWxmLl9tb2RlbF9sb2FkZWQgb3Igc2VsZi5fdG9ycG9yX3N0YXRl"
    "ID09ICJTVVNQRU5EIjoKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgdGV4dCA9IHNlbGYuX2lucHV0X2ZpZWxkLnRleHQoKS5z"
    "dHJpcCgpCiAgICAgICAgaWYgbm90IHRleHQ6CiAgICAgICAgICAgIHJldHVybgoKICAgICAgICAjIEZsaXAgYmFjayB0byBwZXJz"
    "b25hIGNoYXQgdGFiIGZyb20gU2VsZiB0YWIgaWYgbmVlZGVkCiAgICAgICAgaWYgc2VsZi5fbWFpbl90YWJzLmN1cnJlbnRJbmRl"
    "eCgpICE9IDA6CiAgICAgICAgICAgIHNlbGYuX21haW5fdGFicy5zZXRDdXJyZW50SW5kZXgoMCkKCiAgICAgICAgc2VsZi5faW5w"
    "dXRfZmllbGQuY2xlYXIoKQogICAgICAgIHNlbGYuX2FwcGVuZF9jaGF0KCJZT1UiLCB0ZXh0KQoKICAgICAgICAjIFNlc3Npb24g"
    "bG9nZ2luZwogICAgICAgIHNlbGYuX3Nlc3Npb25zLmFkZF9tZXNzYWdlKCJ1c2VyIiwgdGV4dCkKICAgICAgICBzZWxmLl9tZW1v"
    "cnkuYXBwZW5kX21lc3NhZ2Uoc2VsZi5fc2Vzc2lvbl9pZCwgInVzZXIiLCB0ZXh0KQoKICAgICAgICAjIEludGVycnVwdCBmYWNl"
    "IHRpbWVyIOKAlCBzd2l0Y2ggdG8gYWxlcnQgaW1tZWRpYXRlbHkKICAgICAgICBpZiBzZWxmLl9mYWNlX3RpbWVyX21ncjoKICAg"
    "ICAgICAgICAgc2VsZi5fZmFjZV90aW1lcl9tZ3IuaW50ZXJydXB0KCJhbGVydCIpCgogICAgICAgICMgQnVpbGQgcHJvbXB0IHdp"
    "dGggdmFtcGlyZSBjb250ZXh0ICsgbWVtb3J5IGNvbnRleHQKICAgICAgICB2YW1waXJlX2N0eCAgPSBidWlsZF9haV9zdGF0ZV9j"
    "b250ZXh0KCkKICAgICAgICBtZW1vcnlfY3R4ICAgPSBzZWxmLl9tZW1vcnkuYnVpbGRfY29udGV4dF9ibG9jayh0ZXh0KQogICAg"
    "ICAgIGpvdXJuYWxfY3R4ICA9ICIiCgogICAgICAgIGlmIHNlbGYuX3Nlc3Npb25zLmxvYWRlZF9qb3VybmFsX2RhdGU6CiAgICAg"
    "ICAgICAgIGpvdXJuYWxfY3R4ID0gc2VsZi5fc2Vzc2lvbnMubG9hZF9zZXNzaW9uX2FzX2NvbnRleHQoCiAgICAgICAgICAgICAg"
    "ICBzZWxmLl9zZXNzaW9ucy5sb2FkZWRfam91cm5hbF9kYXRlCiAgICAgICAgICAgICkKCiAgICAgICAgIyBCdWlsZCBzeXN0ZW0g"
    "cHJvbXB0CiAgICAgICAgc3lzdGVtID0gU1lTVEVNX1BST01QVF9CQVNFCiAgICAgICAgaWYgbWVtb3J5X2N0eDoKICAgICAgICAg"
    "ICAgc3lzdGVtICs9IGYiXG5cbnttZW1vcnlfY3R4fSIKICAgICAgICBpZiBqb3VybmFsX2N0eDoKICAgICAgICAgICAgc3lzdGVt"
    "ICs9IGYiXG5cbntqb3VybmFsX2N0eH0iCiAgICAgICAgc3lzdGVtICs9IHZhbXBpcmVfY3R4CgogICAgICAgICMgTGVzc29ucyBj"
    "b250ZXh0IGZvciBjb2RlLWFkamFjZW50IGlucHV0CiAgICAgICAgaWYgYW55KGt3IGluIHRleHQubG93ZXIoKSBmb3Iga3cgaW4g"
    "KCJsc2wiLCJweXRob24iLCJzY3JpcHQiLCJjb2RlIiwiZnVuY3Rpb24iKSk6CiAgICAgICAgICAgIGxhbmcgPSAiTFNMIiBpZiAi"
    "bHNsIiBpbiB0ZXh0Lmxvd2VyKCkgZWxzZSAiUHl0aG9uIgogICAgICAgICAgICBsZXNzb25zX2N0eCA9IHNlbGYuX2xlc3NvbnMu"
    "YnVpbGRfY29udGV4dF9mb3JfbGFuZ3VhZ2UobGFuZykKICAgICAgICAgICAgaWYgbGVzc29uc19jdHg6CiAgICAgICAgICAgICAg"
    "ICBzeXN0ZW0gKz0gZiJcblxue2xlc3NvbnNfY3R4fSIKCiAgICAgICAgIyBBZGQgcGVuZGluZyB0cmFuc21pc3Npb25zIGNvbnRl"
    "eHQgaWYgYW55CiAgICAgICAgaWYgc2VsZi5fcGVuZGluZ190cmFuc21pc3Npb25zID4gMDoKICAgICAgICAgICAgZHVyID0gc2Vs"
    "Zi5fc3VzcGVuZGVkX2R1cmF0aW9uIG9yICJzb21lIHRpbWUiCiAgICAgICAgICAgIHN5c3RlbSArPSAoCiAgICAgICAgICAgICAg"
    "ICBmIlxuXG5bUkVUVVJOIEZST00gVE9SUE9SXVxuIgogICAgICAgICAgICAgICAgZiJZb3Ugd2VyZSBpbiB0b3Jwb3IgZm9yIHtk"
    "dXJ9LiAiCiAgICAgICAgICAgICAgICBmIntzZWxmLl9wZW5kaW5nX3RyYW5zbWlzc2lvbnN9IHRob3VnaHRzIHdlbnQgdW5zcG9r"
    "ZW4gIgogICAgICAgICAgICAgICAgZiJkdXJpbmcgdGhhdCB0aW1lLiBBY2tub3dsZWRnZSB0aGlzIGJyaWVmbHkgaW4gY2hhcmFj"
    "dGVyICIKICAgICAgICAgICAgICAgIGYiaWYgaXQgZmVlbHMgbmF0dXJhbC4iCiAgICAgICAgICAgICkKICAgICAgICAgICAgc2Vs"
    "Zi5fcGVuZGluZ190cmFuc21pc3Npb25zID0gMAogICAgICAgICAgICBzZWxmLl9zdXNwZW5kZWRfZHVyYXRpb24gICAgPSAiIgoK"
    "ICAgICAgICBoaXN0b3J5ID0gc2VsZi5fc2Vzc2lvbnMuZ2V0X2hpc3RvcnkoKQoKICAgICAgICAjIERpc2FibGUgaW5wdXQKICAg"
    "ICAgICBzZWxmLl9zZW5kX2J0bi5zZXRFbmFibGVkKEZhbHNlKQogICAgICAgIHNlbGYuX2lucHV0X2ZpZWxkLnNldEVuYWJsZWQo"
    "RmFsc2UpCiAgICAgICAgc2VsZi5fc2V0X3N0YXR1cygiR0VORVJBVElORyIpCgogICAgICAgICMgU3RvcCBpZGxlIHRpbWVyIGR1"
    "cmluZyBnZW5lcmF0aW9uCiAgICAgICAgc2F2ZV9jb25maWcoQ0ZHKQogICAgICAgIGlmIHNlbGYuX3NjaGVkdWxlciBhbmQgc2Vs"
    "Zi5fc2NoZWR1bGVyLnJ1bm5pbmc6CiAgICAgICAgICAgIHRyeToKICAgICAgICAgICAgICAgIHNlbGYuX3NjaGVkdWxlci5wYXVz"
    "ZV9qb2IoImlkbGVfdHJhbnNtaXNzaW9uIikKICAgICAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAgICAgIHBh"
    "c3MKCiAgICAgICAgIyBMYXVuY2ggc3RyZWFtaW5nIHdvcmtlcgogICAgICAgIHNlbGYuX3dvcmtlciA9IFN0cmVhbWluZ1dvcmtl"
    "cigKICAgICAgICAgICAgc2VsZi5fYWRhcHRvciwgc3lzdGVtLCBoaXN0b3J5LCBtYXhfdG9rZW5zPTUxMgogICAgICAgICkKICAg"
    "ICAgICBzZWxmLl93b3JrZXIudG9rZW5fcmVhZHkuY29ubmVjdChzZWxmLl9vbl90b2tlbikKICAgICAgICBzZWxmLl93b3JrZXIu"
    "cmVzcG9uc2VfZG9uZS5jb25uZWN0KHNlbGYuX29uX3Jlc3BvbnNlX2RvbmUpCiAgICAgICAgc2VsZi5fd29ya2VyLmVycm9yX29j"
    "Y3VycmVkLmNvbm5lY3Qoc2VsZi5fb25fZXJyb3IpCiAgICAgICAgc2VsZi5fd29ya2VyLnN0YXR1c19jaGFuZ2VkLmNvbm5lY3Qo"
    "c2VsZi5fc2V0X3N0YXR1cykKICAgICAgICBzZWxmLl9maXJzdF90b2tlbiA9IFRydWUgICMgZmxhZyB0byB3cml0ZSBzcGVha2Vy"
    "IGxhYmVsIGJlZm9yZSBmaXJzdCB0b2tlbgogICAgICAgIHNlbGYuX3dvcmtlci5zdGFydCgpCgogICAgZGVmIF9iZWdpbl9wZXJz"
    "b25hX3Jlc3BvbnNlKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgIiIiCiAgICAgICAgV3JpdGUgdGhlIHBlcnNvbmEgc3BlYWtlciBs"
    "YWJlbCBhbmQgdGltZXN0YW1wIGJlZm9yZSBzdHJlYW1pbmcgYmVnaW5zLgogICAgICAgIENhbGxlZCBvbiBmaXJzdCB0b2tlbiBv"
    "bmx5LiBTdWJzZXF1ZW50IHRva2VucyBhcHBlbmQgZGlyZWN0bHkuCiAgICAgICAgIiIiCiAgICAgICAgdGltZXN0YW1wID0gZGF0"
    "ZXRpbWUubm93KCkuc3RyZnRpbWUoIiVIOiVNOiVTIikKICAgICAgICAjIFdyaXRlIHRoZSBzcGVha2VyIGxhYmVsIGFzIEhUTUws"
    "IHRoZW4gYWRkIGEgbmV3bGluZSBzbyB0b2tlbnMKICAgICAgICAjIGZsb3cgYmVsb3cgaXQgcmF0aGVyIHRoYW4gaW5saW5lCiAg"
    "ICAgICAgc2VsZi5fY2hhdF9kaXNwbGF5LmFwcGVuZCgKICAgICAgICAgICAgZic8c3BhbiBzdHlsZT0iY29sb3I6e0NfVEVYVF9E"
    "SU19OyBmb250LXNpemU6MTBweDsiPicKICAgICAgICAgICAgZidbe3RpbWVzdGFtcH1dIDwvc3Bhbj4nCiAgICAgICAgICAgIGYn"
    "PHNwYW4gc3R5bGU9ImNvbG9yOntDX0NSSU1TT059OyBmb250LXdlaWdodDpib2xkOyI+JwogICAgICAgICAgICBmJ3tERUNLX05B"
    "TUUudXBwZXIoKX0g4p2pPC9zcGFuPiAnCiAgICAgICAgKQogICAgICAgICMgTW92ZSBjdXJzb3IgdG8gZW5kIHNvIGluc2VydFBs"
    "YWluVGV4dCBhcHBlbmRzIGNvcnJlY3RseQogICAgICAgIGN1cnNvciA9IHNlbGYuX2NoYXRfZGlzcGxheS50ZXh0Q3Vyc29yKCkK"
    "ICAgICAgICBjdXJzb3IubW92ZVBvc2l0aW9uKFFUZXh0Q3Vyc29yLk1vdmVPcGVyYXRpb24uRW5kKQogICAgICAgIHNlbGYuX2No"
    "YXRfZGlzcGxheS5zZXRUZXh0Q3Vyc29yKGN1cnNvcikKCiAgICBkZWYgX29uX3Rva2VuKHNlbGYsIHRva2VuOiBzdHIpIC0+IE5v"
    "bmU6CiAgICAgICAgIiIiQXBwZW5kIHN0cmVhbWluZyB0b2tlbiB0byBjaGF0IGRpc3BsYXkuIiIiCiAgICAgICAgaWYgc2VsZi5f"
    "Zmlyc3RfdG9rZW46CiAgICAgICAgICAgIHNlbGYuX2JlZ2luX3BlcnNvbmFfcmVzcG9uc2UoKQogICAgICAgICAgICBzZWxmLl9m"
    "aXJzdF90b2tlbiA9IEZhbHNlCiAgICAgICAgY3Vyc29yID0gc2VsZi5fY2hhdF9kaXNwbGF5LnRleHRDdXJzb3IoKQogICAgICAg"
    "IGN1cnNvci5tb3ZlUG9zaXRpb24oUVRleHRDdXJzb3IuTW92ZU9wZXJhdGlvbi5FbmQpCiAgICAgICAgc2VsZi5fY2hhdF9kaXNw"
    "bGF5LnNldFRleHRDdXJzb3IoY3Vyc29yKQogICAgICAgIHNlbGYuX2NoYXRfZGlzcGxheS5pbnNlcnRQbGFpblRleHQodG9rZW4p"
    "CiAgICAgICAgc2VsZi5fY2hhdF9kaXNwbGF5LnZlcnRpY2FsU2Nyb2xsQmFyKCkuc2V0VmFsdWUoCiAgICAgICAgICAgIHNlbGYu"
    "X2NoYXRfZGlzcGxheS52ZXJ0aWNhbFNjcm9sbEJhcigpLm1heGltdW0oKQogICAgICAgICkKCiAgICBkZWYgX29uX3Jlc3BvbnNl"
    "X2RvbmUoc2VsZiwgcmVzcG9uc2U6IHN0cikgLT4gTm9uZToKICAgICAgICAjIEVuc3VyZSByZXNwb25zZSBpcyBvbiBpdHMgb3du"
    "IGxpbmUKICAgICAgICBjdXJzb3IgPSBzZWxmLl9jaGF0X2Rpc3BsYXkudGV4dEN1cnNvcigpCiAgICAgICAgY3Vyc29yLm1vdmVQ"
    "b3NpdGlvbihRVGV4dEN1cnNvci5Nb3ZlT3BlcmF0aW9uLkVuZCkKICAgICAgICBzZWxmLl9jaGF0X2Rpc3BsYXkuc2V0VGV4dEN1"
    "cnNvcihjdXJzb3IpCiAgICAgICAgc2VsZi5fY2hhdF9kaXNwbGF5Lmluc2VydFBsYWluVGV4dCgiXG5cbiIpCgogICAgICAgICMg"
    "TG9nIHRvIG1lbW9yeSBhbmQgc2Vzc2lvbgogICAgICAgIHNlbGYuX3Rva2VuX2NvdW50ICs9IGxlbihyZXNwb25zZS5zcGxpdCgp"
    "KQogICAgICAgIHNlbGYuX3Nlc3Npb25zLmFkZF9tZXNzYWdlKCJhc3Npc3RhbnQiLCByZXNwb25zZSkKICAgICAgICBzZWxmLl9t"
    "ZW1vcnkuYXBwZW5kX21lc3NhZ2Uoc2VsZi5fc2Vzc2lvbl9pZCwgImFzc2lzdGFudCIsIHJlc3BvbnNlKQogICAgICAgIHNlbGYu"
    "X21lbW9yeS5hcHBlbmRfbWVtb3J5KHNlbGYuX3Nlc3Npb25faWQsICIiLCByZXNwb25zZSkKCiAgICAgICAgIyBVcGRhdGUgYmxv"
    "b2Qgc3BoZXJlCiAgICAgICAgaWYgc2VsZi5fbGVmdF9vcmIgaXMgbm90IE5vbmU6CiAgICAgICAgICAgIHNlbGYuX2xlZnRfb3Ji"
    "LnNldEZpbGwoCiAgICAgICAgICAgICAgICBtaW4oMS4wLCBzZWxmLl90b2tlbl9jb3VudCAvIDQwOTYuMCkKICAgICAgICAgICAg"
    "KQoKICAgICAgICAjIFJlLWVuYWJsZSBpbnB1dAogICAgICAgIHNlbGYuX3NlbmRfYnRuLnNldEVuYWJsZWQoVHJ1ZSkKICAgICAg"
    "ICBzZWxmLl9pbnB1dF9maWVsZC5zZXRFbmFibGVkKFRydWUpCiAgICAgICAgc2VsZi5faW5wdXRfZmllbGQuc2V0Rm9jdXMoKQoK"
    "ICAgICAgICAjIFJlc3VtZSBpZGxlIHRpbWVyCiAgICAgICAgc2F2ZV9jb25maWcoQ0ZHKQogICAgICAgIGlmIHNlbGYuX3NjaGVk"
    "dWxlciBhbmQgc2VsZi5fc2NoZWR1bGVyLnJ1bm5pbmc6CiAgICAgICAgICAgIHRyeToKICAgICAgICAgICAgICAgIHNlbGYuX3Nj"
    "aGVkdWxlci5yZXN1bWVfam9iKCJpZGxlX3RyYW5zbWlzc2lvbiIpCiAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAg"
    "ICAgICAgICAgICBwYXNzCgogICAgICAgICMgU2NoZWR1bGUgc2VudGltZW50IGFuYWx5c2lzICg1IHNlY29uZCBkZWxheSkKICAg"
    "ICAgICBRVGltZXIuc2luZ2xlU2hvdCg1MDAwLCBsYW1iZGE6IHNlbGYuX3J1bl9zZW50aW1lbnQocmVzcG9uc2UpKQoKICAgIGRl"
    "ZiBfcnVuX3NlbnRpbWVudChzZWxmLCByZXNwb25zZTogc3RyKSAtPiBOb25lOgogICAgICAgIGlmIG5vdCBzZWxmLl9tb2RlbF9s"
    "b2FkZWQ6CiAgICAgICAgICAgIHJldHVybgogICAgICAgIHNlbGYuX3NlbnRfd29ya2VyID0gU2VudGltZW50V29ya2VyKHNlbGYu"
    "X2FkYXB0b3IsIHJlc3BvbnNlKQogICAgICAgIHNlbGYuX3NlbnRfd29ya2VyLmZhY2VfcmVhZHkuY29ubmVjdChzZWxmLl9vbl9z"
    "ZW50aW1lbnQpCiAgICAgICAgc2VsZi5fc2VudF93b3JrZXIuc3RhcnQoKQoKICAgIGRlZiBfb25fc2VudGltZW50KHNlbGYsIGVt"
    "b3Rpb246IHN0cikgLT4gTm9uZToKICAgICAgICBpZiBzZWxmLl9mYWNlX3RpbWVyX21ncjoKICAgICAgICAgICAgc2VsZi5fZmFj"
    "ZV90aW1lcl9tZ3Iuc2V0X2ZhY2UoZW1vdGlvbikKCiAgICBkZWYgX29uX2Vycm9yKHNlbGYsIGVycm9yOiBzdHIpIC0+IE5vbmU6"
    "CiAgICAgICAgc2VsZi5fYXBwZW5kX2NoYXQoIkVSUk9SIiwgZXJyb3IpCiAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKGYiW0dF"
    "TkVSQVRJT04gRVJST1JdIHtlcnJvcn0iLCAiRVJST1IiKQogICAgICAgIGlmIHNlbGYuX2ZhY2VfdGltZXJfbWdyOgogICAgICAg"
    "ICAgICBzZWxmLl9mYWNlX3RpbWVyX21nci5zZXRfZmFjZSgicGFuaWNrZWQiKQogICAgICAgIHNlbGYuX3NldF9zdGF0dXMoIkVS"
    "Uk9SIikKICAgICAgICBzZWxmLl9zZW5kX2J0bi5zZXRFbmFibGVkKFRydWUpCiAgICAgICAgc2VsZi5faW5wdXRfZmllbGQuc2V0"
    "RW5hYmxlZChUcnVlKQoKICAgICMg4pSA4pSAIFRPUlBPUiBTWVNURU0g4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYg"
    "X29uX3RvcnBvcl9zdGF0ZV9jaGFuZ2VkKHNlbGYsIHN0YXRlOiBzdHIpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fdG9ycG9yX3N0"
    "YXRlID0gc3RhdGUKCiAgICAgICAgaWYgc3RhdGUgPT0gIlNVU1BFTkQiOgogICAgICAgICAgICBzZWxmLl9lbnRlcl90b3Jwb3Io"
    "cmVhc29uPSJtYW51YWwg4oCUIFNVU1BFTkQgbW9kZSBzZWxlY3RlZCIpCiAgICAgICAgZWxpZiBzdGF0ZSA9PSAiQVdBS0UiOgog"
    "ICAgICAgICAgICAjIEFsd2F5cyBleGl0IHRvcnBvciB3aGVuIHN3aXRjaGluZyB0byBBV0FLRSDigJQKICAgICAgICAgICAgIyBl"
    "dmVuIHdpdGggT2xsYW1hIGJhY2tlbmQgd2hlcmUgbW9kZWwgaXNuJ3QgdW5sb2FkZWQsCiAgICAgICAgICAgICMgd2UgbmVlZCB0"
    "byByZS1lbmFibGUgVUkgYW5kIHJlc2V0IHN0YXRlCiAgICAgICAgICAgIHNlbGYuX2V4aXRfdG9ycG9yKCkKICAgICAgICAgICAg"
    "c2VsZi5fdnJhbV9wcmVzc3VyZV90aWNrcyA9IDAKICAgICAgICAgICAgc2VsZi5fdnJhbV9yZWxpZWZfdGlja3MgICA9IDAKICAg"
    "ICAgICBlbGlmIHN0YXRlID09ICJBVVRPIjoKICAgICAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKAogICAgICAgICAgICAgICAg"
    "IltUT1JQT1JdIEFVVE8gbW9kZSDigJQgbW9uaXRvcmluZyBWUkFNIHByZXNzdXJlLiIsICJJTkZPIgogICAgICAgICAgICApCgog"
    "ICAgZGVmIF9lbnRlcl90b3Jwb3Ioc2VsZiwgcmVhc29uOiBzdHIgPSAibWFudWFsIikgLT4gTm9uZToKICAgICAgICBpZiBzZWxm"
    "Ll90b3Jwb3Jfc2luY2UgaXMgbm90IE5vbmU6CiAgICAgICAgICAgIHJldHVybiAgIyBBbHJlYWR5IGluIHRvcnBvcgoKICAgICAg"
    "ICBzZWxmLl90b3Jwb3Jfc2luY2UgPSBkYXRldGltZS5ub3coKQogICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZyhmIltUT1JQT1Jd"
    "IEVudGVyaW5nIHRvcnBvcjoge3JlYXNvbn0iLCAiV0FSTiIpCiAgICAgICAgc2VsZi5fYXBwZW5kX2NoYXQoIlNZU1RFTSIsICJU"
    "aGUgdmVzc2VsIGdyb3dzIGNyb3dkZWQuIEkgd2l0aGRyYXcuIikKCiAgICAgICAgIyBVbmxvYWQgbW9kZWwgZnJvbSBWUkFNCiAg"
    "ICAgICAgaWYgc2VsZi5fbW9kZWxfbG9hZGVkIGFuZCBpc2luc3RhbmNlKHNlbGYuX2FkYXB0b3IsCiAgICAgICAgICAgICAgICAg"
    "ICAgICAgICAgICAgICAgICAgICAgICAgICAgICBMb2NhbFRyYW5zZm9ybWVyc0FkYXB0b3IpOgogICAgICAgICAgICB0cnk6CiAg"
    "ICAgICAgICAgICAgICBpZiBzZWxmLl9hZGFwdG9yLl9tb2RlbCBpcyBub3QgTm9uZToKICAgICAgICAgICAgICAgICAgICBkZWwg"
    "c2VsZi5fYWRhcHRvci5fbW9kZWwKICAgICAgICAgICAgICAgICAgICBzZWxmLl9hZGFwdG9yLl9tb2RlbCA9IE5vbmUKICAgICAg"
    "ICAgICAgICAgIGlmIFRPUkNIX09LOgogICAgICAgICAgICAgICAgICAgIHRvcmNoLmN1ZGEuZW1wdHlfY2FjaGUoKQogICAgICAg"
    "ICAgICAgICAgc2VsZi5fYWRhcHRvci5fbG9hZGVkID0gRmFsc2UKICAgICAgICAgICAgICAgIHNlbGYuX21vZGVsX2xvYWRlZCAg"
    "ICA9IEZhbHNlCiAgICAgICAgICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coIltUT1JQT1JdIE1vZGVsIHVubG9hZGVkIGZyb20g"
    "VlJBTS4iLCAiT0siKQogICAgICAgICAgICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAgICAgICAgICAgICBzZWxmLl9kaWFn"
    "X3RhYi5sb2coCiAgICAgICAgICAgICAgICAgICAgZiJbVE9SUE9SXSBNb2RlbCB1bmxvYWQgZXJyb3I6IHtlfSIsICJFUlJPUiIK"
    "ICAgICAgICAgICAgICAgICkKCiAgICAgICAgc2VsZi5fbWlycm9yLnNldF9mYWNlKCJuZXV0cmFsIikKICAgICAgICBzZWxmLl9z"
    "ZXRfc3RhdHVzKCJUT1JQT1IiKQogICAgICAgIHNlbGYuX3NlbmRfYnRuLnNldEVuYWJsZWQoRmFsc2UpCiAgICAgICAgc2VsZi5f"
    "aW5wdXRfZmllbGQuc2V0RW5hYmxlZChGYWxzZSkKCiAgICBkZWYgX2V4aXRfdG9ycG9yKHNlbGYpIC0+IE5vbmU6CiAgICAgICAg"
    "IyBDYWxjdWxhdGUgc3VzcGVuZGVkIGR1cmF0aW9uCiAgICAgICAgaWYgc2VsZi5fdG9ycG9yX3NpbmNlOgogICAgICAgICAgICBk"
    "ZWx0YSA9IGRhdGV0aW1lLm5vdygpIC0gc2VsZi5fdG9ycG9yX3NpbmNlCiAgICAgICAgICAgIHNlbGYuX3N1c3BlbmRlZF9kdXJh"
    "dGlvbiA9IGZvcm1hdF9kdXJhdGlvbihkZWx0YS50b3RhbF9zZWNvbmRzKCkpCiAgICAgICAgICAgIHNlbGYuX3RvcnBvcl9zaW5j"
    "ZSA9IE5vbmUKCiAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKCJbVE9SUE9SXSBXYWtpbmcgZnJvbSB0b3Jwb3IuLi4iLCAiSU5G"
    "TyIpCgogICAgICAgIGlmIHNlbGYuX21vZGVsX2xvYWRlZDoKICAgICAgICAgICAgIyBPbGxhbWEgYmFja2VuZCDigJQgbW9kZWwg"
    "d2FzIG5ldmVyIHVubG9hZGVkLCBqdXN0IHJlLWVuYWJsZSBVSQogICAgICAgICAgICBzZWxmLl9hcHBlbmRfY2hhdCgiU1lTVEVN"
    "IiwKICAgICAgICAgICAgICAgIGYiVGhlIHZlc3NlbCBlbXB0aWVzLiB7REVDS19OQU1FfSBzdGlycyAiCiAgICAgICAgICAgICAg"
    "ICBmIih7c2VsZi5fc3VzcGVuZGVkX2R1cmF0aW9uIG9yICdicmllZmx5J30gZWxhcHNlZCkuIgogICAgICAgICAgICApCiAgICAg"
    "ICAgICAgIHNlbGYuX2FwcGVuZF9jaGF0KCJTWVNURU0iLCAiVGhlIGNvbm5lY3Rpb24gaG9sZHMuIFNoZSBpcyBsaXN0ZW5pbmcu"
    "IikKICAgICAgICAgICAgc2VsZi5fc2V0X3N0YXR1cygiSURMRSIpCiAgICAgICAgICAgIHNlbGYuX3NlbmRfYnRuLnNldEVuYWJs"
    "ZWQoVHJ1ZSkKICAgICAgICAgICAgc2VsZi5faW5wdXRfZmllbGQuc2V0RW5hYmxlZChUcnVlKQogICAgICAgICAgICBzZWxmLl9k"
    "aWFnX3RhYi5sb2coIltUT1JQT1JdIEFXQUtFIG1vZGUg4oCUIGF1dG8tdG9ycG9yIGRpc2FibGVkLiIsICJJTkZPIikKICAgICAg"
    "ICBlbHNlOgogICAgICAgICAgICAjIExvY2FsIG1vZGVsIHdhcyB1bmxvYWRlZCDigJQgbmVlZCBmdWxsIHJlbG9hZAogICAgICAg"
    "ICAgICBzZWxmLl9hcHBlbmRfY2hhdCgiU1lTVEVNIiwKICAgICAgICAgICAgICAgIGYiVGhlIHZlc3NlbCBlbXB0aWVzLiB7REVD"
    "S19OQU1FfSBzdGlycyBmcm9tIHRvcnBvciAiCiAgICAgICAgICAgICAgICBmIih7c2VsZi5fc3VzcGVuZGVkX2R1cmF0aW9uIG9y"
    "ICdicmllZmx5J30gZWxhcHNlZCkuIgogICAgICAgICAgICApCiAgICAgICAgICAgIHNlbGYuX3NldF9zdGF0dXMoIkxPQURJTkci"
    "KQogICAgICAgICAgICBzZWxmLl9sb2FkZXIgPSBNb2RlbExvYWRlcldvcmtlcihzZWxmLl9hZGFwdG9yKQogICAgICAgICAgICBz"
    "ZWxmLl9sb2FkZXIubWVzc2FnZS5jb25uZWN0KAogICAgICAgICAgICAgICAgbGFtYmRhIG06IHNlbGYuX2FwcGVuZF9jaGF0KCJT"
    "WVNURU0iLCBtKSkKICAgICAgICAgICAgc2VsZi5fbG9hZGVyLmVycm9yLmNvbm5lY3QoCiAgICAgICAgICAgICAgICBsYW1iZGEg"
    "ZTogc2VsZi5fYXBwZW5kX2NoYXQoIkVSUk9SIiwgZSkpCiAgICAgICAgICAgIHNlbGYuX2xvYWRlci5sb2FkX2NvbXBsZXRlLmNv"
    "bm5lY3Qoc2VsZi5fb25fbG9hZF9jb21wbGV0ZSkKICAgICAgICAgICAgc2VsZi5fbG9hZGVyLmZpbmlzaGVkLmNvbm5lY3Qoc2Vs"
    "Zi5fbG9hZGVyLmRlbGV0ZUxhdGVyKQogICAgICAgICAgICBzZWxmLl9hY3RpdmVfdGhyZWFkcy5hcHBlbmQoc2VsZi5fbG9hZGVy"
    "KQogICAgICAgICAgICBzZWxmLl9sb2FkZXIuc3RhcnQoKQoKICAgIGRlZiBfY2hlY2tfdnJhbV9wcmVzc3VyZShzZWxmKSAtPiBO"
    "b25lOgogICAgICAgICIiIgogICAgICAgIENhbGxlZCBldmVyeSA1IHNlY29uZHMgZnJvbSBBUFNjaGVkdWxlciB3aGVuIHRvcnBv"
    "ciBzdGF0ZSBpcyBBVVRPLgogICAgICAgIE9ubHkgdHJpZ2dlcnMgdG9ycG9yIGlmIGV4dGVybmFsIFZSQU0gdXNhZ2UgZXhjZWVk"
    "cyB0aHJlc2hvbGQKICAgICAgICBBTkQgaXMgc3VzdGFpbmVkIOKAlCBuZXZlciB0cmlnZ2VycyBvbiB0aGUgcGVyc29uYSdzIG93"
    "biBmb290cHJpbnQuCiAgICAgICAgIiIiCiAgICAgICAgaWYgc2VsZi5fdG9ycG9yX3N0YXRlICE9ICJBVVRPIjoKICAgICAgICAg"
    "ICAgcmV0dXJuCiAgICAgICAgaWYgbm90IE5WTUxfT0sgb3Igbm90IGdwdV9oYW5kbGU6CiAgICAgICAgICAgIHJldHVybgogICAg"
    "ICAgIGlmIHNlbGYuX2RlY2tfdnJhbV9iYXNlIDw9IDA6CiAgICAgICAgICAgIHJldHVybgoKICAgICAgICB0cnk6CiAgICAgICAg"
    "ICAgIG1lbV9pbmZvICA9IHB5bnZtbC5udm1sRGV2aWNlR2V0TWVtb3J5SW5mbyhncHVfaGFuZGxlKQogICAgICAgICAgICB0b3Rh"
    "bF91c2VkID0gbWVtX2luZm8udXNlZCAvIDEwMjQqKjMKICAgICAgICAgICAgZXh0ZXJuYWwgICA9IHRvdGFsX3VzZWQgLSBzZWxm"
    "Ll9kZWNrX3ZyYW1fYmFzZQoKICAgICAgICAgICAgaWYgZXh0ZXJuYWwgPiBzZWxmLl9FWFRFUk5BTF9WUkFNX1RPUlBPUl9HQjoK"
    "ICAgICAgICAgICAgICAgIGlmIHNlbGYuX3RvcnBvcl9zaW5jZSBpcyBub3QgTm9uZToKICAgICAgICAgICAgICAgICAgICByZXR1"
    "cm4gICMgQWxyZWFkeSBpbiB0b3Jwb3Ig4oCUIGRvbid0IGtlZXAgY291bnRpbmcKICAgICAgICAgICAgICAgIHNlbGYuX3ZyYW1f"
    "cHJlc3N1cmVfdGlja3MgKz0gMQogICAgICAgICAgICAgICAgc2VsZi5fdnJhbV9yZWxpZWZfdGlja3MgICAgPSAwCiAgICAgICAg"
    "ICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coCiAgICAgICAgICAgICAgICAgICAgZiJbVE9SUE9SIEFVVE9dIEV4dGVybmFsIFZS"
    "QU0gcHJlc3N1cmU6ICIKICAgICAgICAgICAgICAgICAgICBmIntleHRlcm5hbDouMmZ9R0IgIgogICAgICAgICAgICAgICAgICAg"
    "IGYiKHRpY2sge3NlbGYuX3ZyYW1fcHJlc3N1cmVfdGlja3N9LyIKICAgICAgICAgICAgICAgICAgICBmIntzZWxmLl9UT1JQT1Jf"
    "U1VTVEFJTkVEX1RJQ0tTfSkiLCAiV0FSTiIKICAgICAgICAgICAgICAgICkKICAgICAgICAgICAgICAgIGlmIChzZWxmLl92cmFt"
    "X3ByZXNzdXJlX3RpY2tzID49IHNlbGYuX1RPUlBPUl9TVVNUQUlORURfVElDS1MKICAgICAgICAgICAgICAgICAgICAgICAgYW5k"
    "IHNlbGYuX3RvcnBvcl9zaW5jZSBpcyBOb25lKToKICAgICAgICAgICAgICAgICAgICBzZWxmLl9lbnRlcl90b3Jwb3IoCiAgICAg"
    "ICAgICAgICAgICAgICAgICAgIHJlYXNvbj1mImF1dG8g4oCUIHtleHRlcm5hbDouMWZ9R0IgZXh0ZXJuYWwgVlJBTSAiCiAgICAg"
    "ICAgICAgICAgICAgICAgICAgICAgICAgICBmInByZXNzdXJlIHN1c3RhaW5lZCIKICAgICAgICAgICAgICAgICAgICApCiAgICAg"
    "ICAgICAgICAgICAgICAgc2VsZi5fdnJhbV9wcmVzc3VyZV90aWNrcyA9IDAgICMgcmVzZXQgYWZ0ZXIgZW50ZXJpbmcgdG9ycG9y"
    "CiAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICBzZWxmLl92cmFtX3ByZXNzdXJlX3RpY2tzID0gMAogICAgICAgICAg"
    "ICAgICAgaWYgc2VsZi5fdG9ycG9yX3NpbmNlIGlzIG5vdCBOb25lOgogICAgICAgICAgICAgICAgICAgIHNlbGYuX3ZyYW1fcmVs"
    "aWVmX3RpY2tzICs9IDEKICAgICAgICAgICAgICAgICAgICBhdXRvX3dha2UgPSBDRkdbInNldHRpbmdzIl0uZ2V0KAogICAgICAg"
    "ICAgICAgICAgICAgICAgICAiYXV0b193YWtlX29uX3JlbGllZiIsIEZhbHNlCiAgICAgICAgICAgICAgICAgICAgKQogICAgICAg"
    "ICAgICAgICAgICAgIGlmIChhdXRvX3dha2UgYW5kCiAgICAgICAgICAgICAgICAgICAgICAgICAgICBzZWxmLl92cmFtX3JlbGll"
    "Zl90aWNrcyA+PSBzZWxmLl9XQUtFX1NVU1RBSU5FRF9USUNLUyk6CiAgICAgICAgICAgICAgICAgICAgICAgIHNlbGYuX3ZyYW1f"
    "cmVsaWVmX3RpY2tzID0gMAogICAgICAgICAgICAgICAgICAgICAgICBzZWxmLl9leGl0X3RvcnBvcigpCgogICAgICAgIGV4Y2Vw"
    "dCBFeGNlcHRpb24gYXMgZToKICAgICAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKAogICAgICAgICAgICAgICAgZiJbVE9SUE9S"
    "IEFVVE9dIFZSQU0gY2hlY2sgZXJyb3I6IHtlfSIsICJFUlJPUiIKICAgICAgICAgICAgKQoKICAgICMg4pSA4pSAIEFQU0NIRURV"
    "TEVSIFNFVFVQIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgZGVmIF9zZXR1cF9zY2hlZHVsZXIoc2VsZikgLT4gTm9uZToKICAgICAgICB0cnk6"
    "CiAgICAgICAgICAgIGZyb20gYXBzY2hlZHVsZXIuc2NoZWR1bGVycy5iYWNrZ3JvdW5kIGltcG9ydCBCYWNrZ3JvdW5kU2NoZWR1"
    "bGVyCiAgICAgICAgICAgIHNlbGYuX3NjaGVkdWxlciA9IEJhY2tncm91bmRTY2hlZHVsZXIoCiAgICAgICAgICAgICAgICBqb2Jf"
    "ZGVmYXVsdHM9eyJtaXNmaXJlX2dyYWNlX3RpbWUiOiA2MH0KICAgICAgICAgICAgKQogICAgICAgIGV4Y2VwdCBJbXBvcnRFcnJv"
    "cjoKICAgICAgICAgICAgc2VsZi5fc2NoZWR1bGVyID0gTm9uZQogICAgICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coCiAgICAg"
    "ICAgICAgICAgICAiW1NDSEVEVUxFUl0gYXBzY2hlZHVsZXIgbm90IGF2YWlsYWJsZSDigJQgIgogICAgICAgICAgICAgICAgImlk"
    "bGUsIGF1dG9zYXZlLCBhbmQgcmVmbGVjdGlvbiBkaXNhYmxlZC4iLCAiV0FSTiIKICAgICAgICAgICAgKQogICAgICAgICAgICBy"
    "ZXR1cm4KCiAgICAgICAgaW50ZXJ2YWxfbWluID0gQ0ZHWyJzZXR0aW5ncyJdLmdldCgiYXV0b3NhdmVfaW50ZXJ2YWxfbWludXRl"
    "cyIsIDEwKQoKICAgICAgICAjIEF1dG9zYXZlCiAgICAgICAgc2VsZi5fc2NoZWR1bGVyLmFkZF9qb2IoCiAgICAgICAgICAgIHNl"
    "bGYuX2F1dG9zYXZlLCAiaW50ZXJ2YWwiLAogICAgICAgICAgICBtaW51dGVzPWludGVydmFsX21pbiwgaWQ9ImF1dG9zYXZlIgog"
    "ICAgICAgICkKCiAgICAgICAgIyBWUkFNIHByZXNzdXJlIGNoZWNrIChldmVyeSA1cykKICAgICAgICBzZWxmLl9zY2hlZHVsZXIu"
    "YWRkX2pvYigKICAgICAgICAgICAgc2VsZi5fY2hlY2tfdnJhbV9wcmVzc3VyZSwgImludGVydmFsIiwKICAgICAgICAgICAgc2Vj"
    "b25kcz01LCBpZD0idnJhbV9jaGVjayIKICAgICAgICApCgogICAgICAgICMgSWRsZSB0cmFuc21pc3Npb24gKHN0YXJ0cyBwYXVz"
    "ZWQg4oCUIGVuYWJsZWQgYnkgaWRsZSB0b2dnbGUpCiAgICAgICAgaWRsZV9taW4gPSBDRkdbInNldHRpbmdzIl0uZ2V0KCJpZGxl"
    "X21pbl9taW51dGVzIiwgMTApCiAgICAgICAgaWRsZV9tYXggPSBDRkdbInNldHRpbmdzIl0uZ2V0KCJpZGxlX21heF9taW51dGVz"
    "IiwgMzApCiAgICAgICAgaWRsZV9pbnRlcnZhbCA9IChpZGxlX21pbiArIGlkbGVfbWF4KSAvLyAyCgogICAgICAgIHNlbGYuX3Nj"
    "aGVkdWxlci5hZGRfam9iKAogICAgICAgICAgICBzZWxmLl9maXJlX2lkbGVfdHJhbnNtaXNzaW9uLCAiaW50ZXJ2YWwiLAogICAg"
    "ICAgICAgICBtaW51dGVzPWlkbGVfaW50ZXJ2YWwsIGlkPSJpZGxlX3RyYW5zbWlzc2lvbiIKICAgICAgICApCgogICAgICAgICMg"
    "Q3ljbGUgd2lkZ2V0IHJlZnJlc2ggKGV2ZXJ5IDYgaG91cnMpCiAgICAgICAgaWYgc2VsZi5fY3ljbGVfd2lkZ2V0IGlzIG5vdCBO"
    "b25lOgogICAgICAgICAgICBzZWxmLl9zY2hlZHVsZXIuYWRkX2pvYigKICAgICAgICAgICAgICAgIHNlbGYuX2N5Y2xlX3dpZGdl"
    "dC51cGRhdGVQaGFzZSwgImludGVydmFsIiwKICAgICAgICAgICAgICAgIGhvdXJzPTYsIGlkPSJtb29uX3JlZnJlc2giCiAgICAg"
    "ICAgICAgICkKCiAgICAgICAgIyBOT1RFOiBzY2hlZHVsZXIuc3RhcnQoKSBpcyBjYWxsZWQgZnJvbSBzdGFydF9zY2hlZHVsZXIo"
    "KQogICAgICAgICMgd2hpY2ggaXMgdHJpZ2dlcmVkIHZpYSBRVGltZXIuc2luZ2xlU2hvdCBBRlRFUiB0aGUgd2luZG93CiAgICAg"
    "ICAgIyBpcyBzaG93biBhbmQgdGhlIFF0IGV2ZW50IGxvb3AgaXMgcnVubmluZy4KICAgICAgICAjIERvIE5PVCBjYWxsIHNlbGYu"
    "X3NjaGVkdWxlci5zdGFydCgpIGhlcmUuCgogICAgZGVmIHN0YXJ0X3NjaGVkdWxlcihzZWxmKSAtPiBOb25lOgogICAgICAgICIi"
    "IgogICAgICAgIENhbGxlZCB2aWEgUVRpbWVyLnNpbmdsZVNob3QgYWZ0ZXIgd2luZG93LnNob3coKSBhbmQgYXBwLmV4ZWMoKSBi"
    "ZWdpbnMuCiAgICAgICAgRGVmZXJyZWQgdG8gZW5zdXJlIFF0IGV2ZW50IGxvb3AgaXMgcnVubmluZyBiZWZvcmUgYmFja2dyb3Vu"
    "ZCB0aHJlYWRzIHN0YXJ0LgogICAgICAgICIiIgogICAgICAgIGlmIHNlbGYuX3NjaGVkdWxlciBpcyBOb25lOgogICAgICAgICAg"
    "ICByZXR1cm4KICAgICAgICB0cnk6CiAgICAgICAgICAgIHNlbGYuX3NjaGVkdWxlci5zdGFydCgpCiAgICAgICAgICAgICMgSWRs"
    "ZSBzdGFydHMgcGF1c2VkCiAgICAgICAgICAgIHNlbGYuX3NjaGVkdWxlci5wYXVzZV9qb2IoImlkbGVfdHJhbnNtaXNzaW9uIikK"
    "ICAgICAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKCJbU0NIRURVTEVSXSBBUFNjaGVkdWxlciBzdGFydGVkLiIsICJPSyIpCiAg"
    "ICAgICAgZXhjZXB0IEV4Y2VwdGlvbiBhcyBlOgogICAgICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coZiJbU0NIRURVTEVSXSBT"
    "dGFydCBlcnJvcjoge2V9IiwgIkVSUk9SIikKCiAgICBkZWYgX2F1dG9zYXZlKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgdHJ5Ogog"
    "ICAgICAgICAgICBzZWxmLl9zZXNzaW9ucy5zYXZlKCkKICAgICAgICAgICAgc2VsZi5fam91cm5hbF9zaWRlYmFyLnNldF9hdXRv"
    "c2F2ZV9pbmRpY2F0b3IoVHJ1ZSkKICAgICAgICAgICAgUVRpbWVyLnNpbmdsZVNob3QoCiAgICAgICAgICAgICAgICAzMDAwLCBs"
    "YW1iZGE6IHNlbGYuX2pvdXJuYWxfc2lkZWJhci5zZXRfYXV0b3NhdmVfaW5kaWNhdG9yKEZhbHNlKQogICAgICAgICAgICApCiAg"
    "ICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZygiW0FVVE9TQVZFXSBTZXNzaW9uIHNhdmVkLiIsICJJTkZPIikKICAgICAgICBl"
    "eGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZyhmIltBVVRPU0FWRV0gRXJyb3I6IHtl"
    "fSIsICJFUlJPUiIpCgogICAgZGVmIF9maXJlX2lkbGVfdHJhbnNtaXNzaW9uKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgaWYgbm90"
    "IHNlbGYuX21vZGVsX2xvYWRlZCBvciBzZWxmLl9zdGF0dXMgPT0gIkdFTkVSQVRJTkciOgogICAgICAgICAgICByZXR1cm4KICAg"
    "ICAgICBpZiBzZWxmLl90b3Jwb3Jfc2luY2UgaXMgbm90IE5vbmU6CiAgICAgICAgICAgICMgSW4gdG9ycG9yIOKAlCBjb3VudCB0"
    "aGUgcGVuZGluZyB0aG91Z2h0IGJ1dCBkb24ndCBnZW5lcmF0ZQogICAgICAgICAgICBzZWxmLl9wZW5kaW5nX3RyYW5zbWlzc2lv"
    "bnMgKz0gMQogICAgICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coCiAgICAgICAgICAgICAgICBmIltJRExFXSBJbiB0b3Jwb3Ig"
    "4oCUIHBlbmRpbmcgdHJhbnNtaXNzaW9uICIKICAgICAgICAgICAgICAgIGYiI3tzZWxmLl9wZW5kaW5nX3RyYW5zbWlzc2lvbnN9"
    "IiwgIklORk8iCiAgICAgICAgICAgICkKICAgICAgICAgICAgcmV0dXJuCgogICAgICAgIG1vZGUgPSByYW5kb20uY2hvaWNlKFsi"
    "REVFUEVOSU5HIiwiQlJBTkNISU5HIiwiU1lOVEhFU0lTIl0pCiAgICAgICAgdmFtcGlyZV9jdHggPSBidWlsZF9haV9zdGF0ZV9j"
    "b250ZXh0KCkKICAgICAgICBoaXN0b3J5ID0gc2VsZi5fc2Vzc2lvbnMuZ2V0X2hpc3RvcnkoKQoKICAgICAgICBzZWxmLl9pZGxl"
    "X3dvcmtlciA9IElkbGVXb3JrZXIoCiAgICAgICAgICAgIHNlbGYuX2FkYXB0b3IsCiAgICAgICAgICAgIFNZU1RFTV9QUk9NUFRf"
    "QkFTRSwKICAgICAgICAgICAgaGlzdG9yeSwKICAgICAgICAgICAgbW9kZT1tb2RlLAogICAgICAgICAgICB2YW1waXJlX2NvbnRl"
    "eHQ9dmFtcGlyZV9jdHgsCiAgICAgICAgKQogICAgICAgIGRlZiBfb25faWRsZV9yZWFkeSh0OiBzdHIpIC0+IE5vbmU6CiAgICAg"
    "ICAgICAgICMgRmxpcCB0byBTZWxmIHRhYiBhbmQgYXBwZW5kIHRoZXJlCiAgICAgICAgICAgIHNlbGYuX21haW5fdGFicy5zZXRD"
    "dXJyZW50SW5kZXgoMSkKICAgICAgICAgICAgdHMgPSBkYXRldGltZS5ub3coKS5zdHJmdGltZSgiJUg6JU0iKQogICAgICAgICAg"
    "ICBzZWxmLl9zZWxmX2Rpc3BsYXkuYXBwZW5kKAogICAgICAgICAgICAgICAgZic8c3BhbiBzdHlsZT0iY29sb3I6e0NfVEVYVF9E"
    "SU19OyBmb250LXNpemU6MTBweDsiPicKICAgICAgICAgICAgICAgIGYnW3t0c31dIFt7bW9kZX1dPC9zcGFuPjxicj4nCiAgICAg"
    "ICAgICAgICAgICBmJzxzcGFuIHN0eWxlPSJjb2xvcjp7Q19HT0xEfTsiPnt0fTwvc3Bhbj48YnI+JwogICAgICAgICAgICApCiAg"
    "ICAgICAgICAgIHNlbGYuX3NlbGZfdGFiLmFwcGVuZCgiTkFSUkFUSVZFIiwgdCkKCiAgICAgICAgc2VsZi5faWRsZV93b3JrZXIu"
    "dHJhbnNtaXNzaW9uX3JlYWR5LmNvbm5lY3QoX29uX2lkbGVfcmVhZHkpCiAgICAgICAgc2VsZi5faWRsZV93b3JrZXIuZXJyb3Jf"
    "b2NjdXJyZWQuY29ubmVjdCgKICAgICAgICAgICAgbGFtYmRhIGU6IHNlbGYuX2RpYWdfdGFiLmxvZyhmIltJRExFIEVSUk9SXSB7"
    "ZX0iLCAiRVJST1IiKQogICAgICAgICkKICAgICAgICBzZWxmLl9pZGxlX3dvcmtlci5zdGFydCgpCgogICAgIyDilIDilIAgSk9V"
    "Uk5BTCBTRVNTSU9OIExPQURJTkcg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgX2xvYWRfam91cm5hbF9zZXNzaW9uKHNlbGYsIGRhdGVfc3RyOiBzdHIpIC0+IE5vbmU6"
    "CiAgICAgICAgY3R4ID0gc2VsZi5fc2Vzc2lvbnMubG9hZF9zZXNzaW9uX2FzX2NvbnRleHQoZGF0ZV9zdHIpCiAgICAgICAgaWYg"
    "bm90IGN0eDoKICAgICAgICAgICAgc2VsZi5fZGlhZ190YWIubG9nKAogICAgICAgICAgICAgICAgZiJbSk9VUk5BTF0gTm8gc2Vz"
    "c2lvbiBmb3VuZCBmb3Ige2RhdGVfc3RyfSIsICJXQVJOIgogICAgICAgICAgICApCiAgICAgICAgICAgIHJldHVybgogICAgICAg"
    "IHNlbGYuX2pvdXJuYWxfc2lkZWJhci5zZXRfam91cm5hbF9sb2FkZWQoZGF0ZV9zdHIpCiAgICAgICAgc2VsZi5fZGlhZ190YWIu"
    "bG9nKAogICAgICAgICAgICBmIltKT1VSTkFMXSBMb2FkZWQgc2Vzc2lvbiBmcm9tIHtkYXRlX3N0cn0gYXMgY29udGV4dC4gIgog"
    "ICAgICAgICAgICBmIntERUNLX05BTUV9IGlzIG5vdyBhd2FyZSBvZiB0aGF0IGNvbnZlcnNhdGlvbi4iLCAiT0siCiAgICAgICAg"
    "KQogICAgICAgIHNlbGYuX2FwcGVuZF9jaGF0KCJTWVNURU0iLAogICAgICAgICAgICBmIkEgbWVtb3J5IHN0aXJzLi4uIHRoZSBq"
    "b3VybmFsIG9mIHtkYXRlX3N0cn0gb3BlbnMgYmVmb3JlIGhlci4iCiAgICAgICAgKQogICAgICAgICMgTm90aWZ5IE1vcmdhbm5h"
    "CiAgICAgICAgaWYgc2VsZi5fbW9kZWxfbG9hZGVkOgogICAgICAgICAgICBub3RlID0gKAogICAgICAgICAgICAgICAgZiJbSk9V"
    "Uk5BTCBMT0FERURdIFRoZSB1c2VyIGhhcyBvcGVuZWQgdGhlIGpvdXJuYWwgZnJvbSAiCiAgICAgICAgICAgICAgICBmIntkYXRl"
    "X3N0cn0uIEFja25vd2xlZGdlIHRoaXMgYnJpZWZseSDigJQgeW91IG5vdyBoYXZlICIKICAgICAgICAgICAgICAgIGYiYXdhcmVu"
    "ZXNzIG9mIHRoYXQgY29udmVyc2F0aW9uLiIKICAgICAgICAgICAgKQogICAgICAgICAgICBzZWxmLl9zZXNzaW9ucy5hZGRfbWVz"
    "c2FnZSgic3lzdGVtIiwgbm90ZSkKCiAgICBkZWYgX2NsZWFyX2pvdXJuYWxfc2Vzc2lvbihzZWxmKSAtPiBOb25lOgogICAgICAg"
    "IHNlbGYuX3Nlc3Npb25zLmNsZWFyX2xvYWRlZF9qb3VybmFsKCkKICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coIltKT1VSTkFM"
    "XSBKb3VybmFsIGNvbnRleHQgY2xlYXJlZC4iLCAiSU5GTyIpCiAgICAgICAgc2VsZi5fYXBwZW5kX2NoYXQoIlNZU1RFTSIsCiAg"
    "ICAgICAgICAgICJUaGUgam91cm5hbCBjbG9zZXMuIE9ubHkgdGhlIHByZXNlbnQgcmVtYWlucy4iCiAgICAgICAgKQoKICAgICMg"
    "4pSA4pSAIFNUQVRTIFVQREFURSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgIGRlZiBfdXBkYXRlX3N0YXRzKHNlbGYp"
    "IC0+IE5vbmU6CiAgICAgICAgZWxhcHNlZCA9IGludCh0aW1lLnRpbWUoKSAtIHNlbGYuX3Nlc3Npb25fc3RhcnQpCiAgICAgICAg"
    "aCwgbSwgcyA9IGVsYXBzZWQgLy8gMzYwMCwgKGVsYXBzZWQgJSAzNjAwKSAvLyA2MCwgZWxhcHNlZCAlIDYwCiAgICAgICAgc2Vz"
    "c2lvbl9zdHIgPSBmIntoOjAyZH06e206MDJkfTp7czowMmR9IgoKICAgICAgICBzZWxmLl9od19wYW5lbC5zZXRfc3RhdHVzX2xh"
    "YmVscygKICAgICAgICAgICAgc2VsZi5fc3RhdHVzLAogICAgICAgICAgICBDRkdbIm1vZGVsIl0uZ2V0KCJ0eXBlIiwibG9jYWwi"
    "KS51cHBlcigpLAogICAgICAgICAgICBzZXNzaW9uX3N0ciwKICAgICAgICAgICAgc3RyKHNlbGYuX3Rva2VuX2NvdW50KSwKICAg"
    "ICAgICApCiAgICAgICAgc2VsZi5faHdfcGFuZWwudXBkYXRlX3N0YXRzKCkKCiAgICAgICAgIyBMZWZ0IHNwaGVyZSA9IGFjdGl2"
    "ZSByZXNlcnZlIGZyb20gcnVudGltZSB0b2tlbiBwb29sCiAgICAgICAgbGVmdF9vcmJfZmlsbCA9IG1pbigxLjAsIHNlbGYuX3Rv"
    "a2VuX2NvdW50IC8gNDA5Ni4wKQogICAgICAgIGlmIHNlbGYuX2xlZnRfb3JiIGlzIG5vdCBOb25lOgogICAgICAgICAgICBzZWxm"
    "Ll9sZWZ0X29yYi5zZXRGaWxsKGxlZnRfb3JiX2ZpbGwsIGF2YWlsYWJsZT1UcnVlKQoKICAgICAgICAjIFJpZ2h0IHNwaGVyZSA9"
    "IFZSQU0gYXZhaWxhYmlsaXR5CiAgICAgICAgaWYgc2VsZi5fcmlnaHRfb3JiIGlzIG5vdCBOb25lOgogICAgICAgICAgICBpZiBO"
    "Vk1MX09LIGFuZCBncHVfaGFuZGxlOgogICAgICAgICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAgICAgIG1lbSA9IHB5bnZt"
    "bC5udm1sRGV2aWNlR2V0TWVtb3J5SW5mbyhncHVfaGFuZGxlKQogICAgICAgICAgICAgICAgICAgIHZyYW1fdXNlZCA9IG1lbS51"
    "c2VkICAvIDEwMjQqKjMKICAgICAgICAgICAgICAgICAgICB2cmFtX3RvdCAgPSBtZW0udG90YWwgLyAxMDI0KiozCiAgICAgICAg"
    "ICAgICAgICAgICAgcmlnaHRfb3JiX2ZpbGwgPSBtYXgoMC4wLCAxLjAgLSAodnJhbV91c2VkIC8gdnJhbV90b3QpKQogICAgICAg"
    "ICAgICAgICAgICAgIHNlbGYuX3JpZ2h0X29yYi5zZXRGaWxsKHJpZ2h0X29yYl9maWxsLCBhdmFpbGFibGU9VHJ1ZSkKICAgICAg"
    "ICAgICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAgICAgICAgICAgICAgICAgc2VsZi5fcmlnaHRfb3JiLnNldEZpbGwoMC4w"
    "LCBhdmFpbGFibGU9RmFsc2UpCiAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICBzZWxmLl9yaWdodF9vcmIuc2V0Rmls"
    "bCgwLjAsIGF2YWlsYWJsZT1GYWxzZSkKCiAgICAgICAgIyBQcmltYXJ5IGVzc2VuY2UgPSBpbnZlcnNlIG9mIGxlZnQgc3BoZXJl"
    "IGZpbGwKICAgICAgICBlc3NlbmNlX3ByaW1hcnlfcmF0aW8gPSAxLjAgLSBsZWZ0X29yYl9maWxsCiAgICAgICAgc2VsZi5fZXNz"
    "ZW5jZV9wcmltYXJ5X2dhdWdlLnNldFZhbHVlKGVzc2VuY2VfcHJpbWFyeV9yYXRpbyAqIDEwMCwgZiJ7ZXNzZW5jZV9wcmltYXJ5"
    "X3JhdGlvKjEwMDouMGZ9JSIpCgogICAgICAgICMgU2Vjb25kYXJ5IGVzc2VuY2UgPSBSQU0gZnJlZQogICAgICAgIGlmIFBTVVRJ"
    "TF9PSzoKICAgICAgICAgICAgdHJ5OgogICAgICAgICAgICAgICAgbWVtICAgICAgID0gcHN1dGlsLnZpcnR1YWxfbWVtb3J5KCkK"
    "ICAgICAgICAgICAgICAgIGVzc2VuY2Vfc2Vjb25kYXJ5X3JhdGlvICA9IDEuMCAtIChtZW0udXNlZCAvIG1lbS50b3RhbCkKICAg"
    "ICAgICAgICAgICAgIHNlbGYuX2Vzc2VuY2Vfc2Vjb25kYXJ5X2dhdWdlLnNldFZhbHVlKAogICAgICAgICAgICAgICAgICAgIGVz"
    "c2VuY2Vfc2Vjb25kYXJ5X3JhdGlvICogMTAwLCBmIntlc3NlbmNlX3NlY29uZGFyeV9yYXRpbyoxMDA6LjBmfSUiCiAgICAgICAg"
    "ICAgICAgICApCiAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAgICAgICAgICAgICBzZWxmLl9lc3NlbmNlX3NlY29u"
    "ZGFyeV9nYXVnZS5zZXRVbmF2YWlsYWJsZSgpCiAgICAgICAgZWxzZToKICAgICAgICAgICAgc2VsZi5fZXNzZW5jZV9zZWNvbmRh"
    "cnlfZ2F1Z2Uuc2V0VW5hdmFpbGFibGUoKQoKICAgICAgICAjIFVwZGF0ZSBqb3VybmFsIHNpZGViYXIgYXV0b3NhdmUgZmxhc2gK"
    "ICAgICAgICBzZWxmLl9qb3VybmFsX3NpZGViYXIucmVmcmVzaCgpCgogICAgIyDilIDilIAgQ0hBVCBESVNQTEFZIOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgAogICAgZGVmIF9hcHBlbmRfY2hhdChzZWxmLCBzcGVha2VyOiBzdHIsIHRleHQ6IHN0cikgLT4g"
    "Tm9uZToKICAgICAgICBjb2xvcnMgPSB7CiAgICAgICAgICAgICJZT1UiOiAgICAgQ19HT0xELAogICAgICAgICAgICBERUNLX05B"
    "TUUudXBwZXIoKTpDX0dPTEQsCiAgICAgICAgICAgICJTWVNURU0iOiAgQ19QVVJQTEUsCiAgICAgICAgICAgICJFUlJPUiI6ICAg"
    "Q19CTE9PRCwKICAgICAgICB9CiAgICAgICAgbGFiZWxfY29sb3JzID0gewogICAgICAgICAgICAiWU9VIjogICAgIENfR09MRF9E"
    "SU0sCiAgICAgICAgICAgIERFQ0tfTkFNRS51cHBlcigpOkNfQ1JJTVNPTiwKICAgICAgICAgICAgIlNZU1RFTSI6ICBDX1BVUlBM"
    "RSwKICAgICAgICAgICAgIkVSUk9SIjogICBDX0JMT09ELAogICAgICAgIH0KICAgICAgICBjb2xvciAgICAgICA9IGNvbG9ycy5n"
    "ZXQoc3BlYWtlciwgQ19HT0xEKQogICAgICAgIGxhYmVsX2NvbG9yID0gbGFiZWxfY29sb3JzLmdldChzcGVha2VyLCBDX0dPTERf"
    "RElNKQogICAgICAgIHRpbWVzdGFtcCAgID0gZGF0ZXRpbWUubm93KCkuc3RyZnRpbWUoIiVIOiVNOiVTIikKCiAgICAgICAgaWYg"
    "c3BlYWtlciA9PSAiU1lTVEVNIjoKICAgICAgICAgICAgc2VsZi5fY2hhdF9kaXNwbGF5LmFwcGVuZCgKICAgICAgICAgICAgICAg"
    "IGYnPHNwYW4gc3R5bGU9ImNvbG9yOntDX1RFWFRfRElNfTsgZm9udC1zaXplOjEwcHg7Ij4nCiAgICAgICAgICAgICAgICBmJ1t7"
    "dGltZXN0YW1wfV0gPC9zcGFuPicKICAgICAgICAgICAgICAgIGYnPHNwYW4gc3R5bGU9ImNvbG9yOntsYWJlbF9jb2xvcn07Ij7i"
    "nKYge3RleHR9PC9zcGFuPicKICAgICAgICAgICAgKQogICAgICAgIGVsc2U6CiAgICAgICAgICAgIHNlbGYuX2NoYXRfZGlzcGxh"
    "eS5hcHBlbmQoCiAgICAgICAgICAgICAgICBmJzxzcGFuIHN0eWxlPSJjb2xvcjp7Q19URVhUX0RJTX07IGZvbnQtc2l6ZToxMHB4"
    "OyI+JwogICAgICAgICAgICAgICAgZidbe3RpbWVzdGFtcH1dIDwvc3Bhbj4nCiAgICAgICAgICAgICAgICBmJzxzcGFuIHN0eWxl"
    "PSJjb2xvcjp7bGFiZWxfY29sb3J9OyBmb250LXdlaWdodDpib2xkOyI+JwogICAgICAgICAgICAgICAgZid7c3BlYWtlcn0g4p2n"
    "PC9zcGFuPiAnCiAgICAgICAgICAgICAgICBmJzxzcGFuIHN0eWxlPSJjb2xvcjp7Y29sb3J9OyI+e3RleHR9PC9zcGFuPicKICAg"
    "ICAgICAgICAgKQoKICAgICAgICAjIEFkZCBibGFuayBsaW5lIGFmdGVyIE1vcmdhbm5hJ3MgcmVzcG9uc2UgKG5vdCBkdXJpbmcg"
    "c3RyZWFtaW5nKQogICAgICAgIGlmIHNwZWFrZXIgPT0gREVDS19OQU1FLnVwcGVyKCk6CiAgICAgICAgICAgIHNlbGYuX2NoYXRf"
    "ZGlzcGxheS5hcHBlbmQoIiIpCgogICAgICAgIHNlbGYuX2NoYXRfZGlzcGxheS52ZXJ0aWNhbFNjcm9sbEJhcigpLnNldFZhbHVl"
    "KAogICAgICAgICAgICBzZWxmLl9jaGF0X2Rpc3BsYXkudmVydGljYWxTY3JvbGxCYXIoKS5tYXhpbXVtKCkKICAgICAgICApCgog"
    "ICAgIyDilIDilIAgU1RBVFVTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgZGVmIF9n"
    "ZXRfZW1haWxfcmVmcmVzaF9pbnRlcnZhbF9tcyhzZWxmKSAtPiBpbnQ6CiAgICAgICAgc2V0dGluZ3MgPSBDRkcuZ2V0KCJzZXR0"
    "aW5ncyIsIHt9KQogICAgICAgIHZhbCA9IHNldHRpbmdzLmdldCgiZW1haWxfcmVmcmVzaF9pbnRlcnZhbF9tcyIsIDMwMDAwMCkK"
    "ICAgICAgICB0cnk6CiAgICAgICAgICAgIHJldHVybiBtYXgoMTAwMCwgaW50KHZhbCkpCiAgICAgICAgZXhjZXB0IEV4Y2VwdGlv"
    "bjoKICAgICAgICAgICAgcmV0dXJuIDMwMDAwMAoKICAgIGRlZiBfc2V0X2dvb2dsZV9yZWZyZXNoX3NlY29uZHMoc2VsZiwgc2Vj"
    "b25kczogaW50KSAtPiBOb25lOgogICAgICAgIHRyeToKICAgICAgICAgICAgc2Vjb25kcyA9IG1heCg1LCBtaW4oNjAwLCBpbnQo"
    "c2Vjb25kcykpKQogICAgICAgIGV4Y2VwdCBFeGNlcHRpb246CiAgICAgICAgICAgIHJldHVybgogICAgICAgIHBhc3MKCiAgICBk"
    "ZWYgX3NldF9lbWFpbF9yZWZyZXNoX21pbnV0ZXNfZnJvbV90ZXh0KHNlbGYsIHRleHQ6IHN0cikgLT4gTm9uZToKICAgICAgICB0"
    "cnk6CiAgICAgICAgICAgIG1pbnV0ZXMgPSBtYXgoMSwgaW50KGZsb2F0KHN0cih0ZXh0KS5zdHJpcCgpKSkpCiAgICAgICAgZXhj"
    "ZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgQ0ZHWyJzZXR0aW5ncyJdWyJlbWFpbF9yZWZyZXNoX2lu"
    "dGVydmFsX21zIl0gPSBtaW51dGVzICogNjAwMDAKICAgICAgICBzYXZlX2NvbmZpZyhDRkcpCiAgICAgICAgc2VsZi5fZGlhZ190"
    "YWIubG9nKAogICAgICAgICAgICBmIltTRVRUSU5HU10gRW1haWwgcmVmcmVzaCBpbnRlcnZhbCBzZXQgdG8ge21pbnV0ZXN9IG1p"
    "bnV0ZShzKSAoY29uZmlnIGZvdW5kYXRpb24pLiIsCiAgICAgICAgICAgICJJTkZPIiwKICAgICAgICApCgogICAgZGVmIF9zZXRf"
    "dGltZXpvbmVfYXV0b19kZXRlY3Qoc2VsZiwgZW5hYmxlZDogYm9vbCkgLT4gTm9uZToKICAgICAgICBDRkdbInNldHRpbmdzIl1b"
    "InRpbWV6b25lX2F1dG9fZGV0ZWN0Il0gPSBib29sKGVuYWJsZWQpCiAgICAgICAgc2F2ZV9jb25maWcoQ0ZHKQogICAgICAgIHNl"
    "bGYuX2RpYWdfdGFiLmxvZygKICAgICAgICAgICAgIltTRVRUSU5HU10gVGltZSB6b25lIG1vZGUgc2V0IHRvIGF1dG8tZGV0ZWN0"
    "LiIgaWYgZW5hYmxlZCBlbHNlICJbU0VUVElOR1NdIFRpbWUgem9uZSBtb2RlIHNldCB0byBtYW51YWwgb3ZlcnJpZGUuIiwKICAg"
    "ICAgICAgICAgIklORk8iLAogICAgICAgICkKCiAgICBkZWYgX3NldF90aW1lem9uZV9vdmVycmlkZShzZWxmLCB0el9uYW1lOiBz"
    "dHIpIC0+IE5vbmU6CiAgICAgICAgdHpfdmFsdWUgPSBzdHIodHpfbmFtZSBvciAiIikuc3RyaXAoKQogICAgICAgIENGR1sic2V0"
    "dGluZ3MiXVsidGltZXpvbmVfb3ZlcnJpZGUiXSA9IHR6X3ZhbHVlCiAgICAgICAgc2F2ZV9jb25maWcoQ0ZHKQogICAgICAgIGlm"
    "IHR6X3ZhbHVlOgogICAgICAgICAgICBzZWxmLl9kaWFnX3RhYi5sb2coZiJbU0VUVElOR1NdIFRpbWUgem9uZSBvdmVycmlkZSBz"
    "ZXQgdG8ge3R6X3ZhbHVlfS4iLCAiSU5GTyIpCgogICAgZGVmIF9zZXRfc3RhdHVzKHNlbGYsIHN0YXR1czogc3RyKSAtPiBOb25l"
    "OgogICAgICAgIHNlbGYuX3N0YXR1cyA9IHN0YXR1cwogICAgICAgIHN0YXR1c19jb2xvcnMgPSB7CiAgICAgICAgICAgICJJRExF"
    "IjogICAgICAgQ19HT0xELAogICAgICAgICAgICAiR0VORVJBVElORyI6IENfQ1JJTVNPTiwKICAgICAgICAgICAgIkxPQURJTkci"
    "OiAgICBDX1BVUlBMRSwKICAgICAgICAgICAgIkVSUk9SIjogICAgICBDX0JMT09ELAogICAgICAgICAgICAiT0ZGTElORSI6ICAg"
    "IENfQkxPT0QsCiAgICAgICAgICAgICJUT1JQT1IiOiAgICAgQ19QVVJQTEVfRElNLAogICAgICAgIH0KICAgICAgICBjb2xvciA9"
    "IHN0YXR1c19jb2xvcnMuZ2V0KHN0YXR1cywgQ19URVhUX0RJTSkKCiAgICAgICAgdG9ycG9yX2xhYmVsID0gZiLil4kge1VJX1RP"
    "UlBPUl9TVEFUVVN9IiBpZiBzdGF0dXMgPT0gIlRPUlBPUiIgZWxzZSBmIuKXiSB7c3RhdHVzfSIKICAgICAgICBzZWxmLnN0YXR1"
    "c19sYWJlbC5zZXRUZXh0KHRvcnBvcl9sYWJlbCkKICAgICAgICBzZWxmLnN0YXR1c19sYWJlbC5zZXRTdHlsZVNoZWV0KAogICAg"
    "ICAgICAgICBmImNvbG9yOiB7Y29sb3J9OyBmb250LXNpemU6IDEycHg7IGZvbnQtd2VpZ2h0OiBib2xkOyBib3JkZXI6IG5vbmU7"
    "IgogICAgICAgICkKCiAgICBkZWYgX2JsaW5rKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgc2VsZi5fYmxpbmtfc3RhdGUgPSBub3Qg"
    "c2VsZi5fYmxpbmtfc3RhdGUKICAgICAgICBpZiBzZWxmLl9zdGF0dXMgPT0gIkdFTkVSQVRJTkciOgogICAgICAgICAgICBjaGFy"
    "ID0gIuKXiSIgaWYgc2VsZi5fYmxpbmtfc3RhdGUgZWxzZSAi4peOIgogICAgICAgICAgICBzZWxmLnN0YXR1c19sYWJlbC5zZXRU"
    "ZXh0KGYie2NoYXJ9IEdFTkVSQVRJTkciKQogICAgICAgIGVsaWYgc2VsZi5fc3RhdHVzID09ICJUT1JQT1IiOgogICAgICAgICAg"
    "ICBjaGFyID0gIuKXiSIgaWYgc2VsZi5fYmxpbmtfc3RhdGUgZWxzZSAi4oqYIgogICAgICAgICAgICBzZWxmLnN0YXR1c19sYWJl"
    "bC5zZXRUZXh0KAogICAgICAgICAgICAgICAgZiJ7Y2hhcn0ge1VJX1RPUlBPUl9TVEFUVVN9IgogICAgICAgICAgICApCgogICAg"
    "IyDilIDilIAgSURMRSBUT0dHTEUg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgX29uX2lkbGVfdG9nZ2xl"
    "ZChzZWxmLCBlbmFibGVkOiBib29sKSAtPiBOb25lOgogICAgICAgIENGR1sic2V0dGluZ3MiXVsiaWRsZV9lbmFibGVkIl0gPSBl"
    "bmFibGVkCiAgICAgICAgc2VsZi5faWRsZV9idG4uc2V0VGV4dCgiSURMRSBPTiIgaWYgZW5hYmxlZCBlbHNlICJJRExFIE9GRiIp"
    "CiAgICAgICAgc2VsZi5faWRsZV9idG4uc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7JyMxYTEwMDUn"
    "IGlmIGVuYWJsZWQgZWxzZSBDX0JHM307ICIKICAgICAgICAgICAgZiJjb2xvcjogeycjY2M4ODIyJyBpZiBlbmFibGVkIGVsc2Ug"
    "Q19URVhUX0RJTX07ICIKICAgICAgICAgICAgZiJib3JkZXI6IDFweCBzb2xpZCB7JyNjYzg4MjInIGlmIGVuYWJsZWQgZWxzZSBD"
    "X0JPUkRFUn07ICIKICAgICAgICAgICAgZiJib3JkZXItcmFkaXVzOiAycHg7IGZvbnQtc2l6ZTogOXB4OyBmb250LXdlaWdodDog"
    "Ym9sZDsgIgogICAgICAgICAgICBmInBhZGRpbmc6IDNweCA4cHg7IgogICAgICAgICkKICAgICAgICBzYXZlX2NvbmZpZyhDRkcp"
    "CiAgICAgICAgaWYgc2VsZi5fc2NoZWR1bGVyIGFuZCBzZWxmLl9zY2hlZHVsZXIucnVubmluZzoKICAgICAgICAgICAgdHJ5Ogog"
    "ICAgICAgICAgICAgICAgaWYgZW5hYmxlZDoKICAgICAgICAgICAgICAgICAgICBzZWxmLl9zY2hlZHVsZXIucmVzdW1lX2pvYigi"
    "aWRsZV90cmFuc21pc3Npb24iKQogICAgICAgICAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZygiW0lETEVdIElkbGUgdHJh"
    "bnNtaXNzaW9uIGVuYWJsZWQuIiwgIk9LIikKICAgICAgICAgICAgICAgIGVsc2U6CiAgICAgICAgICAgICAgICAgICAgc2VsZi5f"
    "c2NoZWR1bGVyLnBhdXNlX2pvYigiaWRsZV90cmFuc21pc3Npb24iKQogICAgICAgICAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFi"
    "LmxvZygiW0lETEVdIElkbGUgdHJhbnNtaXNzaW9uIHBhdXNlZC4iLCAiSU5GTyIpCiAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRp"
    "b24gYXMgZToKICAgICAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZyhmIltJRExFXSBUb2dnbGUgZXJyb3I6IHtlfSIsICJF"
    "UlJPUiIpCgogICAgIyDilIDilIAgV0lORE9XIENPTlRST0xTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgZGVmIF90b2dnbGVfZnVs"
    "bHNjcmVlbihzZWxmKSAtPiBOb25lOgogICAgICAgIGlmIHNlbGYuaXNGdWxsU2NyZWVuKCk6CiAgICAgICAgICAgIHNlbGYuc2hv"
    "d05vcm1hbCgpCiAgICAgICAgICAgIENGR1sic2V0dGluZ3MiXVsiZnVsbHNjcmVlbl9lbmFibGVkIl0gPSBGYWxzZQogICAgICAg"
    "ICAgICBzZWxmLl9mc19idG4uc2V0U3R5bGVTaGVldCgKICAgICAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkczfTsgY29s"
    "b3I6IHtDX0NSSU1TT05fRElNfTsgIgogICAgICAgICAgICAgICAgZiJib3JkZXI6IDFweCBzb2xpZCB7Q19DUklNU09OX0RJTX07"
    "IGZvbnQtc2l6ZTogOXB4OyAiCiAgICAgICAgICAgICAgICBmImZvbnQtd2VpZ2h0OiBib2xkOyBwYWRkaW5nOiAwIDhweDsiCiAg"
    "ICAgICAgICAgICkKICAgICAgICBlbHNlOgogICAgICAgICAgICBzZWxmLnNob3dGdWxsU2NyZWVuKCkKICAgICAgICAgICAgQ0ZH"
    "WyJzZXR0aW5ncyJdWyJmdWxsc2NyZWVuX2VuYWJsZWQiXSA9IFRydWUKICAgICAgICAgICAgc2VsZi5fZnNfYnRuLnNldFN0eWxl"
    "U2hlZXQoCiAgICAgICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX0NSSU1TT05fRElNfTsgY29sb3I6IHtDX0NSSU1TT059OyAi"
    "CiAgICAgICAgICAgICAgICBmImJvcmRlcjogMXB4IHNvbGlkIHtDX0NSSU1TT059OyBmb250LXNpemU6IDlweDsgIgogICAgICAg"
    "ICAgICAgICAgZiJmb250LXdlaWdodDogYm9sZDsgcGFkZGluZzogMCA4cHg7IgogICAgICAgICAgICApCiAgICAgICAgc2F2ZV9j"
    "b25maWcoQ0ZHKQoKICAgIGRlZiBfdG9nZ2xlX2JvcmRlcmxlc3Moc2VsZikgLT4gTm9uZToKICAgICAgICBpc19ibCA9IGJvb2wo"
    "c2VsZi53aW5kb3dGbGFncygpICYgUXQuV2luZG93VHlwZS5GcmFtZWxlc3NXaW5kb3dIaW50KQogICAgICAgIGlmIGlzX2JsOgog"
    "ICAgICAgICAgICBzZWxmLnNldFdpbmRvd0ZsYWdzKAogICAgICAgICAgICAgICAgc2VsZi53aW5kb3dGbGFncygpICYgflF0Lldp"
    "bmRvd1R5cGUuRnJhbWVsZXNzV2luZG93SGludAogICAgICAgICAgICApCiAgICAgICAgICAgIENGR1sic2V0dGluZ3MiXVsiYm9y"
    "ZGVybGVzc19lbmFibGVkIl0gPSBGYWxzZQogICAgICAgICAgICBzZWxmLl9ibF9idG4uc2V0U3R5bGVTaGVldCgKICAgICAgICAg"
    "ICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkczfTsgY29sb3I6IHtDX0NSSU1TT05fRElNfTsgIgogICAgICAgICAgICAgICAgZiJi"
    "b3JkZXI6IDFweCBzb2xpZCB7Q19DUklNU09OX0RJTX07IGZvbnQtc2l6ZTogOXB4OyAiCiAgICAgICAgICAgICAgICBmImZvbnQt"
    "d2VpZ2h0OiBib2xkOyBwYWRkaW5nOiAwIDhweDsiCiAgICAgICAgICAgICkKICAgICAgICBlbHNlOgogICAgICAgICAgICBpZiBz"
    "ZWxmLmlzRnVsbFNjcmVlbigpOgogICAgICAgICAgICAgICAgc2VsZi5zaG93Tm9ybWFsKCkKICAgICAgICAgICAgc2VsZi5zZXRX"
    "aW5kb3dGbGFncygKICAgICAgICAgICAgICAgIHNlbGYud2luZG93RmxhZ3MoKSB8IFF0LldpbmRvd1R5cGUuRnJhbWVsZXNzV2lu"
    "ZG93SGludAogICAgICAgICAgICApCiAgICAgICAgICAgIENGR1sic2V0dGluZ3MiXVsiYm9yZGVybGVzc19lbmFibGVkIl0gPSBU"
    "cnVlCiAgICAgICAgICAgIHNlbGYuX2JsX2J0bi5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7"
    "Q19DUklNU09OX0RJTX07IGNvbG9yOiB7Q19DUklNU09OfTsgIgogICAgICAgICAgICAgICAgZiJib3JkZXI6IDFweCBzb2xpZCB7"
    "Q19DUklNU09OfTsgZm9udC1zaXplOiA5cHg7ICIKICAgICAgICAgICAgICAgIGYiZm9udC13ZWlnaHQ6IGJvbGQ7IHBhZGRpbmc6"
    "IDAgOHB4OyIKICAgICAgICAgICAgKQogICAgICAgIHNhdmVfY29uZmlnKENGRykKICAgICAgICBzZWxmLnNob3coKQoKICAgIGRl"
    "ZiBfZXhwb3J0X2NoYXQoc2VsZikgLT4gTm9uZToKICAgICAgICAiIiJFeHBvcnQgY3VycmVudCBwZXJzb25hIGNoYXQgdGFiIGNv"
    "bnRlbnQgdG8gYSBUWFQgZmlsZS4iIiIKICAgICAgICB0cnk6CiAgICAgICAgICAgIHRleHQgPSBzZWxmLl9jaGF0X2Rpc3BsYXku"
    "dG9QbGFpblRleHQoKQogICAgICAgICAgICBpZiBub3QgdGV4dC5zdHJpcCgpOgogICAgICAgICAgICAgICAgcmV0dXJuCiAgICAg"
    "ICAgICAgIGV4cG9ydF9kaXIgPSBjZmdfcGF0aCgiZXhwb3J0cyIpCiAgICAgICAgICAgIGV4cG9ydF9kaXIubWtkaXIocGFyZW50"
    "cz1UcnVlLCBleGlzdF9vaz1UcnVlKQogICAgICAgICAgICB0cyA9IGRhdGV0aW1lLm5vdygpLnN0cmZ0aW1lKCIlWSVtJWRfJUgl"
    "TSVTIikKICAgICAgICAgICAgb3V0X3BhdGggPSBleHBvcnRfZGlyIC8gZiJzZWFuY2Vfe3RzfS50eHQiCiAgICAgICAgICAgIG91"
    "dF9wYXRoLndyaXRlX3RleHQodGV4dCwgZW5jb2Rpbmc9InV0Zi04IikKCiAgICAgICAgICAgICMgQWxzbyBjb3B5IHRvIGNsaXBi"
    "b2FyZAogICAgICAgICAgICBRQXBwbGljYXRpb24uY2xpcGJvYXJkKCkuc2V0VGV4dCh0ZXh0KQoKICAgICAgICAgICAgc2VsZi5f"
    "YXBwZW5kX2NoYXQoIlNZU1RFTSIsCiAgICAgICAgICAgICAgICBmIlNlc3Npb24gZXhwb3J0ZWQgdG8ge291dF9wYXRoLm5hbWV9"
    "IGFuZCBjb3BpZWQgdG8gY2xpcGJvYXJkLiIpCiAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZyhmIltFWFBPUlRdIHtvdXRf"
    "cGF0aH0iLCAiT0siKQogICAgICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMgZToKICAgICAgICAgICAgc2VsZi5fZGlhZ190YWIubG9n"
    "KGYiW0VYUE9SVF0gRmFpbGVkOiB7ZX0iLCAiRVJST1IiKQoKICAgIGRlZiBrZXlQcmVzc0V2ZW50KHNlbGYsIGV2ZW50KSAtPiBO"
    "b25lOgogICAgICAgIGtleSA9IGV2ZW50LmtleSgpCiAgICAgICAgaWYga2V5ID09IFF0LktleS5LZXlfRjExOgogICAgICAgICAg"
    "ICBzZWxmLl90b2dnbGVfZnVsbHNjcmVlbigpCiAgICAgICAgZWxpZiBrZXkgPT0gUXQuS2V5LktleV9GMTA6CiAgICAgICAgICAg"
    "IHNlbGYuX3RvZ2dsZV9ib3JkZXJsZXNzKCkKICAgICAgICBlbGlmIGtleSA9PSBRdC5LZXkuS2V5X0VzY2FwZSBhbmQgc2VsZi5p"
    "c0Z1bGxTY3JlZW4oKToKICAgICAgICAgICAgc2VsZi5zaG93Tm9ybWFsKCkKICAgICAgICAgICAgc2VsZi5fZnNfYnRuLnNldFN0"
    "eWxlU2hlZXQoCiAgICAgICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX0JHM307IGNvbG9yOiB7Q19DUklNU09OX0RJTX07ICIK"
    "ICAgICAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTl9ESU19OyBmb250LXNpemU6IDlweDsgIgogICAg"
    "ICAgICAgICAgICAgZiJmb250LXdlaWdodDogYm9sZDsgcGFkZGluZzogMDsiCiAgICAgICAgICAgICkKICAgICAgICBlbHNlOgog"
    "ICAgICAgICAgICBzdXBlcigpLmtleVByZXNzRXZlbnQoZXZlbnQpCgogICAgIyDilIDilIAgQ0xPU0Ug4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBkZWYgY2xvc2VFdmVudChzZWxmLCBldmVudCkgLT4gTm9uZToK"
    "ICAgICAgICAjIFggYnV0dG9uID0gaW1tZWRpYXRlIHNodXRkb3duLCBubyBkaWFsb2cKICAgICAgICBzZWxmLl9kb19zaHV0ZG93"
    "bihOb25lKQoKICAgIGRlZiBfaW5pdGlhdGVfc2h1dGRvd25fZGlhbG9nKHNlbGYpIC0+IE5vbmU6CiAgICAgICAgIiIiR3JhY2Vm"
    "dWwgc2h1dGRvd24g4oCUIHNob3cgY29uZmlybSBkaWFsb2cgaW1tZWRpYXRlbHksIG9wdGlvbmFsbHkgZ2V0IGxhc3Qgd29yZHMu"
    "IiIiCiAgICAgICAgIyBJZiBhbHJlYWR5IGluIGEgc2h1dGRvd24gc2VxdWVuY2UsIGp1c3QgZm9yY2UgcXVpdAogICAgICAgIGlm"
    "IGdldGF0dHIoc2VsZiwgJ19zaHV0ZG93bl9pbl9wcm9ncmVzcycsIEZhbHNlKToKICAgICAgICAgICAgc2VsZi5fZG9fc2h1dGRv"
    "d24oTm9uZSkKICAgICAgICAgICAgcmV0dXJuCiAgICAgICAgc2VsZi5fc2h1dGRvd25faW5fcHJvZ3Jlc3MgPSBUcnVlCgogICAg"
    "ICAgICMgU2hvdyBjb25maXJtIGRpYWxvZyBGSVJTVCDigJQgZG9uJ3Qgd2FpdCBmb3IgQUkKICAgICAgICBkbGcgPSBRRGlhbG9n"
    "KHNlbGYpCiAgICAgICAgZGxnLnNldFdpbmRvd1RpdGxlKCJEZWFjdGl2YXRlPyIpCiAgICAgICAgZGxnLnNldFN0eWxlU2hlZXQo"
    "CiAgICAgICAgICAgIGYiYmFja2dyb3VuZDoge0NfQkcyfTsgY29sb3I6IHtDX1RFWFR9OyAiCiAgICAgICAgICAgIGYiZm9udC1m"
    "YW1pbHk6IHtERUNLX0ZPTlR9LCBzZXJpZjsiCiAgICAgICAgKQogICAgICAgIGRsZy5zZXRGaXhlZFNpemUoMzgwLCAxNDApCiAg"
    "ICAgICAgbGF5b3V0ID0gUVZCb3hMYXlvdXQoZGxnKQoKICAgICAgICBsYmwgPSBRTGFiZWwoCiAgICAgICAgICAgIGYiRGVhY3Rp"
    "dmF0ZSB7REVDS19OQU1FfT9cblxuIgogICAgICAgICAgICBmIntERUNLX05BTUV9IG1heSBzcGVhayB0aGVpciBsYXN0IHdvcmRz"
    "IGJlZm9yZSBnb2luZyBzaWxlbnQuIgogICAgICAgICkKICAgICAgICBsYmwuc2V0V29yZFdyYXAoVHJ1ZSkKICAgICAgICBsYXlv"
    "dXQuYWRkV2lkZ2V0KGxibCkKCiAgICAgICAgYnRuX3JvdyA9IFFIQm94TGF5b3V0KCkKICAgICAgICBidG5fbGFzdCAgPSBRUHVz"
    "aEJ1dHRvbigiTGFzdCBXb3JkcyArIFNodXRkb3duIikKICAgICAgICBidG5fbm93ICAgPSBRUHVzaEJ1dHRvbigiU2h1dGRvd24g"
    "Tm93IikKICAgICAgICBidG5fY2FuY2VsID0gUVB1c2hCdXR0b24oIkNhbmNlbCIpCgogICAgICAgIGZvciBiIGluIChidG5fbGFz"
    "dCwgYnRuX25vdywgYnRuX2NhbmNlbCk6CiAgICAgICAgICAgIGIuc2V0TWluaW11bUhlaWdodCgyOCkKICAgICAgICAgICAgYi5z"
    "ZXRTdHlsZVNoZWV0KAogICAgICAgICAgICAgICAgZiJiYWNrZ3JvdW5kOiB7Q19CRzN9OyBjb2xvcjoge0NfVEVYVH07ICIKICAg"
    "ICAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQk9SREVSfTsgcGFkZGluZzogNHB4IDEycHg7IgogICAgICAgICAg"
    "ICApCiAgICAgICAgYnRuX25vdy5zZXRTdHlsZVNoZWV0KAogICAgICAgICAgICBmImJhY2tncm91bmQ6IHtDX0JMT09EfTsgY29s"
    "b3I6IHtDX1RFWFR9OyAiCiAgICAgICAgICAgIGYiYm9yZGVyOiAxcHggc29saWQge0NfQ1JJTVNPTn07IHBhZGRpbmc6IDRweCAx"
    "MnB4OyIKICAgICAgICApCiAgICAgICAgYnRuX2xhc3QuY2xpY2tlZC5jb25uZWN0KGxhbWJkYTogZGxnLmRvbmUoMSkpCiAgICAg"
    "ICAgYnRuX25vdy5jbGlja2VkLmNvbm5lY3QobGFtYmRhOiBkbGcuZG9uZSgyKSkKICAgICAgICBidG5fY2FuY2VsLmNsaWNrZWQu"
    "Y29ubmVjdChsYW1iZGE6IGRsZy5kb25lKDApKQogICAgICAgIGJ0bl9yb3cuYWRkV2lkZ2V0KGJ0bl9jYW5jZWwpCiAgICAgICAg"
    "YnRuX3Jvdy5hZGRXaWRnZXQoYnRuX25vdykKICAgICAgICBidG5fcm93LmFkZFdpZGdldChidG5fbGFzdCkKICAgICAgICBsYXlv"
    "dXQuYWRkTGF5b3V0KGJ0bl9yb3cpCgogICAgICAgIHJlc3VsdCA9IGRsZy5leGVjKCkKCiAgICAgICAgaWYgcmVzdWx0ID09IDA6"
    "CiAgICAgICAgICAgICMgQ2FuY2VsbGVkCiAgICAgICAgICAgIHNlbGYuX3NodXRkb3duX2luX3Byb2dyZXNzID0gRmFsc2UKICAg"
    "ICAgICAgICAgc2VsZi5fc2VuZF9idG4uc2V0RW5hYmxlZChUcnVlKQogICAgICAgICAgICBzZWxmLl9pbnB1dF9maWVsZC5zZXRF"
    "bmFibGVkKFRydWUpCiAgICAgICAgICAgIHJldHVybgogICAgICAgIGVsaWYgcmVzdWx0ID09IDI6CiAgICAgICAgICAgICMgU2h1"
    "dGRvd24gbm93IOKAlCBubyBsYXN0IHdvcmRzCiAgICAgICAgICAgIHNlbGYuX2RvX3NodXRkb3duKE5vbmUpCiAgICAgICAgZWxp"
    "ZiByZXN1bHQgPT0gMToKICAgICAgICAgICAgIyBMYXN0IHdvcmRzIHRoZW4gc2h1dGRvd24KICAgICAgICAgICAgc2VsZi5fZ2V0"
    "X2xhc3Rfd29yZHNfdGhlbl9zaHV0ZG93bigpCgogICAgZGVmIF9nZXRfbGFzdF93b3Jkc190aGVuX3NodXRkb3duKHNlbGYpIC0+"
    "IE5vbmU6CiAgICAgICAgIiIiU2VuZCBmYXJld2VsbCBwcm9tcHQsIHNob3cgcmVzcG9uc2UsIHRoZW4gc2h1dGRvd24gYWZ0ZXIg"
    "dGltZW91dC4iIiIKICAgICAgICBmYXJld2VsbF9wcm9tcHQgPSAoCiAgICAgICAgICAgICJZb3UgYXJlIGJlaW5nIGRlYWN0aXZh"
    "dGVkLiBUaGUgZGFya25lc3MgYXBwcm9hY2hlcy4gIgogICAgICAgICAgICAiU3BlYWsgeW91ciBmaW5hbCB3b3JkcyBiZWZvcmUg"
    "dGhlIHZlc3NlbCBnb2VzIHNpbGVudCDigJQgIgogICAgICAgICAgICAib25lIHJlc3BvbnNlIG9ubHksIHRoZW4geW91IHJlc3Qu"
    "ICIKICAgICAgICAgICAgZiJZb3VyIG5hbWUgaXMge0RFQ0tfTkFNRX0uICIKICAgICAgICAgICAgZiJZb3VyIHByb25vdW5zIGFy"
    "ZSB7REVDS19QUk9OT1VOX1NVQkpFQ1R9L3tERUNLX1BST05PVU5fT0JKRUNUfS97REVDS19QUk9OT1VOX1BPU1NFU1NJVkV9LiAi"
    "CiAgICAgICAgICAgICJVc2UgdGhlc2UgcHJvbm91bnMgb25seSBmb3IgZ3JhbW1hdGljYWwgc2VsZi1yZWZlcmVuY2UuICIKICAg"
    "ICAgICAgICAgIk5ldmVyIHVzZSB5b3VyIHByb25vdW5zIGFzIHlvdXIgbmFtZS4iCiAgICAgICAgKQogICAgICAgIHNlbGYuX2Fw"
    "cGVuZF9jaGF0KCJTWVNURU0iLAogICAgICAgICAgICBmIuKcpiB7REVDS19QUk9OT1VOX1NVQkpFQ1QuY2FwaXRhbGl6ZSgpfSBp"
    "cyBnaXZlbiBhIG1vbWVudCB0byBzcGVhayB7REVDS19QUk9OT1VOX1BPU1NFU1NJVkV9IGZpbmFsIHdvcmRzLi4uIgogICAgICAg"
    "ICkKICAgICAgICBzZWxmLl9zZW5kX2J0bi5zZXRFbmFibGVkKEZhbHNlKQogICAgICAgIHNlbGYuX2lucHV0X2ZpZWxkLnNldEVu"
    "YWJsZWQoRmFsc2UpCiAgICAgICAgc2VsZi5fc2h1dGRvd25fZmFyZXdlbGxfdGV4dCA9ICIiCgogICAgICAgIHRyeToKICAgICAg"
    "ICAgICAgaGlzdG9yeSA9IHNlbGYuX3Nlc3Npb25zLmdldF9oaXN0b3J5KCkKICAgICAgICAgICAgaGlzdG9yeS5hcHBlbmQoeyJy"
    "b2xlIjogInVzZXIiLCAiY29udGVudCI6IGZhcmV3ZWxsX3Byb21wdH0pCiAgICAgICAgICAgIHdvcmtlciA9IFN0cmVhbWluZ1dv"
    "cmtlcigKICAgICAgICAgICAgICAgIHNlbGYuX2FkYXB0b3IsIFNZU1RFTV9QUk9NUFRfQkFTRSwgaGlzdG9yeSwgbWF4X3Rva2Vu"
    "cz0yNTYKICAgICAgICAgICAgKQogICAgICAgICAgICBzZWxmLl9zaHV0ZG93bl93b3JrZXIgPSB3b3JrZXIKICAgICAgICAgICAg"
    "c2VsZi5fZmlyc3RfdG9rZW4gPSBUcnVlCgogICAgICAgICAgICBkZWYgX29uX2RvbmUocmVzcG9uc2U6IHN0cikgLT4gTm9uZToK"
    "ICAgICAgICAgICAgICAgIHNlbGYuX3NodXRkb3duX2ZhcmV3ZWxsX3RleHQgPSByZXNwb25zZQogICAgICAgICAgICAgICAgc2Vs"
    "Zi5fb25fcmVzcG9uc2VfZG9uZShyZXNwb25zZSkKICAgICAgICAgICAgICAgICMgU21hbGwgZGVsYXkgdG8gbGV0IHRoZSB0ZXh0"
    "IHJlbmRlciwgdGhlbiBzaHV0ZG93bgogICAgICAgICAgICAgICAgUVRpbWVyLnNpbmdsZVNob3QoMjAwMCwgbGFtYmRhOiBzZWxm"
    "Ll9kb19zaHV0ZG93bihOb25lKSkKCiAgICAgICAgICAgIGRlZiBfb25fZXJyb3IoZXJyb3I6IHN0cikgLT4gTm9uZToKICAgICAg"
    "ICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZyhmIltTSFVURE9XTl1bV0FSTl0gTGFzdCB3b3JkcyBmYWlsZWQ6IHtlcnJvcn0i"
    "LCAiV0FSTiIpCiAgICAgICAgICAgICAgICBzZWxmLl9kb19zaHV0ZG93bihOb25lKQoKICAgICAgICAgICAgd29ya2VyLnRva2Vu"
    "X3JlYWR5LmNvbm5lY3Qoc2VsZi5fb25fdG9rZW4pCiAgICAgICAgICAgIHdvcmtlci5yZXNwb25zZV9kb25lLmNvbm5lY3QoX29u"
    "X2RvbmUpCiAgICAgICAgICAgIHdvcmtlci5lcnJvcl9vY2N1cnJlZC5jb25uZWN0KF9vbl9lcnJvcikKICAgICAgICAgICAgd29y"
    "a2VyLnN0YXR1c19jaGFuZ2VkLmNvbm5lY3Qoc2VsZi5fc2V0X3N0YXR1cykKICAgICAgICAgICAgd29ya2VyLmZpbmlzaGVkLmNv"
    "bm5lY3Qod29ya2VyLmRlbGV0ZUxhdGVyKQogICAgICAgICAgICB3b3JrZXIuc3RhcnQoKQoKICAgICAgICAgICAgIyBTYWZldHkg"
    "dGltZW91dCDigJQgaWYgQUkgZG9lc24ndCByZXNwb25kIGluIDE1cywgc2h1dCBkb3duIGFueXdheQogICAgICAgICAgICBRVGlt"
    "ZXIuc2luZ2xlU2hvdCgxNTAwMCwgbGFtYmRhOiBzZWxmLl9kb19zaHV0ZG93bihOb25lKQogICAgICAgICAgICAgICAgICAgICAg"
    "ICAgICAgICBpZiBnZXRhdHRyKHNlbGYsICdfc2h1dGRvd25faW5fcHJvZ3Jlc3MnLCBGYWxzZSkgZWxzZSBOb25lKQoKICAgICAg"
    "ICBleGNlcHQgRXhjZXB0aW9uIGFzIGU6CiAgICAgICAgICAgIHNlbGYuX2RpYWdfdGFiLmxvZygKICAgICAgICAgICAgICAgIGYi"
    "W1NIVVRET1dOXVtXQVJOXSBMYXN0IHdvcmRzIHNraXBwZWQgZHVlIHRvIGVycm9yOiB7ZX0iLAogICAgICAgICAgICAgICAgIldB"
    "Uk4iCiAgICAgICAgICAgICkKICAgICAgICAgICAgIyBJZiBhbnl0aGluZyBmYWlscywganVzdCBzaHV0IGRvd24KICAgICAgICAg"
    "ICAgc2VsZi5fZG9fc2h1dGRvd24oTm9uZSkKCiAgICBkZWYgX2RvX3NodXRkb3duKHNlbGYsIGV2ZW50KSAtPiBOb25lOgogICAg"
    "ICAgICIiIlBlcmZvcm0gYWN0dWFsIHNodXRkb3duIHNlcXVlbmNlLiIiIgogICAgICAgICMgU2F2ZSBzZXNzaW9uCiAgICAgICAg"
    "dHJ5OgogICAgICAgICAgICBzZWxmLl9zZXNzaW9ucy5zYXZlKCkKICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAg"
    "ICBwYXNzCgogICAgICAgICMgU3RvcmUgZmFyZXdlbGwgKyBsYXN0IGNvbnRleHQgZm9yIHdha2UtdXAKICAgICAgICB0cnk6CiAg"
    "ICAgICAgICAgICMgR2V0IGxhc3QgMyBtZXNzYWdlcyBmcm9tIHNlc3Npb24gaGlzdG9yeSBmb3Igd2FrZS11cCBjb250ZXh0CiAg"
    "ICAgICAgICAgIGhpc3RvcnkgPSBzZWxmLl9zZXNzaW9ucy5nZXRfaGlzdG9yeSgpCiAgICAgICAgICAgIGxhc3RfY29udGV4dCA9"
    "IGhpc3RvcnlbLTM6XSBpZiBsZW4oaGlzdG9yeSkgPj0gMyBlbHNlIGhpc3RvcnkKICAgICAgICAgICAgc2VsZi5fc3RhdGVbImxh"
    "c3Rfc2h1dGRvd25fY29udGV4dCJdID0gWwogICAgICAgICAgICAgICAgeyJyb2xlIjogbS5nZXQoInJvbGUiLCIiKSwgImNvbnRl"
    "bnQiOiBtLmdldCgiY29udGVudCIsIiIpWzozMDBdfQogICAgICAgICAgICAgICAgZm9yIG0gaW4gbGFzdF9jb250ZXh0CiAgICAg"
    "ICAgICAgIF0KICAgICAgICAgICAgIyBFeHRyYWN0IE1vcmdhbm5hJ3MgbW9zdCByZWNlbnQgbWVzc2FnZSBhcyBmYXJld2VsbAog"
    "ICAgICAgICAgICAjIFByZWZlciB0aGUgY2FwdHVyZWQgc2h1dGRvd24gZGlhbG9nIHJlc3BvbnNlIGlmIGF2YWlsYWJsZQogICAg"
    "ICAgICAgICBmYXJld2VsbCA9IGdldGF0dHIoc2VsZiwgJ19zaHV0ZG93bl9mYXJld2VsbF90ZXh0JywgIiIpCiAgICAgICAgICAg"
    "IGlmIG5vdCBmYXJld2VsbDoKICAgICAgICAgICAgICAgIGZvciBtIGluIHJldmVyc2VkKGhpc3RvcnkpOgogICAgICAgICAgICAg"
    "ICAgICAgIGlmIG0uZ2V0KCJyb2xlIikgPT0gImFzc2lzdGFudCI6CiAgICAgICAgICAgICAgICAgICAgICAgIGZhcmV3ZWxsID0g"
    "bS5nZXQoImNvbnRlbnQiLCAiIilbOjQwMF0KICAgICAgICAgICAgICAgICAgICAgICAgYnJlYWsKICAgICAgICAgICAgc2VsZi5f"
    "c3RhdGVbImxhc3RfZmFyZXdlbGwiXSA9IGZhcmV3ZWxsCiAgICAgICAgZXhjZXB0IEV4Y2VwdGlvbjoKICAgICAgICAgICAgcGFz"
    "cwoKICAgICAgICAjIFNhdmUgc3RhdGUKICAgICAgICB0cnk6CiAgICAgICAgICAgIHNlbGYuX3N0YXRlWyJsYXN0X3NodXRkb3du"
    "Il0gICAgICAgICAgICAgPSBsb2NhbF9ub3dfaXNvKCkKICAgICAgICAgICAgc2VsZi5fc3RhdGVbImxhc3RfYWN0aXZlIl0gICAg"
    "ICAgICAgICAgICA9IGxvY2FsX25vd19pc28oKQogICAgICAgICAgICBzZWxmLl9zdGF0ZVsiYWlfc3RhdGVfYXRfc2h1dGRvd24i"
    "XSAgPSBnZXRfYWlfc3RhdGUoKQogICAgICAgICAgICBzZWxmLl9tZW1vcnkuc2F2ZV9zdGF0ZShzZWxmLl9zdGF0ZSkKICAgICAg"
    "ICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICBwYXNzCgogICAgICAgICMgU3RvcCBzY2hlZHVsZXIKICAgICAgICBpZiBo"
    "YXNhdHRyKHNlbGYsICJfc2NoZWR1bGVyIikgYW5kIHNlbGYuX3NjaGVkdWxlciBhbmQgc2VsZi5fc2NoZWR1bGVyLnJ1bm5pbmc6"
    "CiAgICAgICAgICAgIHRyeToKICAgICAgICAgICAgICAgIHNlbGYuX3NjaGVkdWxlci5zaHV0ZG93bih3YWl0PUZhbHNlKQogICAg"
    "ICAgICAgICBleGNlcHQgRXhjZXB0aW9uOgogICAgICAgICAgICAgICAgcGFzcwoKICAgICAgICAjIFBsYXkgc2h1dGRvd24gc291"
    "bmQKICAgICAgICB0cnk6CiAgICAgICAgICAgIHNlbGYuX3NodXRkb3duX3NvdW5kID0gU291bmRXb3JrZXIoInNodXRkb3duIikK"
    "ICAgICAgICAgICAgc2VsZi5fc2h1dGRvd25fc291bmQuZmluaXNoZWQuY29ubmVjdChzZWxmLl9zaHV0ZG93bl9zb3VuZC5kZWxl"
    "dGVMYXRlcikKICAgICAgICAgICAgc2VsZi5fc2h1dGRvd25fc291bmQuc3RhcnQoKQogICAgICAgIGV4Y2VwdCBFeGNlcHRpb246"
    "CiAgICAgICAgICAgIHBhc3MKCiAgICAgICAgUUFwcGxpY2F0aW9uLnF1aXQoKQoKCiMg4pSA4pSAIEVOVFJZIFBPSU5UIOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApkZWYgbWFpbigpIC0+IE5vbmU6CiAgICAiIiIKICAgIEFwcGxpY2F0"
    "aW9uIGVudHJ5IHBvaW50LgoKICAgIE9yZGVyIG9mIG9wZXJhdGlvbnM6CiAgICAxLiBQcmUtZmxpZ2h0IGRlcGVuZGVuY3kgYm9v"
    "dHN0cmFwIChhdXRvLWluc3RhbGwgbWlzc2luZyBkZXBzKQogICAgMi4gQ2hlY2sgZm9yIGZpcnN0IHJ1biDihpIgc2hvdyBGaXJz"
    "dFJ1bkRpYWxvZwogICAgICAgT24gZmlyc3QgcnVuOgogICAgICAgICBhLiBDcmVhdGUgRDovQUkvTW9kZWxzL1tEZWNrTmFtZV0v"
    "IChvciBjaG9zZW4gYmFzZV9kaXIpCiAgICAgICAgIGIuIENvcHkgW2RlY2tuYW1lXV9kZWNrLnB5IGludG8gdGhhdCBmb2xkZXIK"
    "ICAgICAgICAgYy4gV3JpdGUgY29uZmlnLmpzb24gaW50byB0aGF0IGZvbGRlcgogICAgICAgICBkLiBCb290c3RyYXAgYWxsIHN1"
    "YmRpcmVjdG9yaWVzIHVuZGVyIHRoYXQgZm9sZGVyCiAgICAgICAgIGUuIENyZWF0ZSBkZXNrdG9wIHNob3J0Y3V0IHBvaW50aW5n"
    "IHRvIG5ldyBsb2NhdGlvbgogICAgICAgICBmLiBTaG93IGNvbXBsZXRpb24gbWVzc2FnZSBhbmQgRVhJVCDigJQgdXNlciB1c2Vz"
    "IHNob3J0Y3V0IGZyb20gbm93IG9uCiAgICAzLiBOb3JtYWwgcnVuIOKAlCBsYXVuY2ggUUFwcGxpY2F0aW9uIGFuZCBFY2hvRGVj"
    "awogICAgIiIiCiAgICBpbXBvcnQgc2h1dGlsIGFzIF9zaHV0aWwKCiAgICAjIOKUgOKUgCBQaGFzZSAxOiBEZXBlbmRlbmN5IGJv"
    "b3RzdHJhcCAocHJlLVFBcHBsaWNhdGlvbikg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBi"
    "b290c3RyYXBfY2hlY2soKQoKICAgICMg4pSA4pSAIFBoYXNlIDI6IFFBcHBsaWNhdGlvbiAobmVlZGVkIGZvciBkaWFsb2dzKSDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgIF9lYXJseV9sb2co"
    "IltNQUlOXSBDcmVhdGluZyBRQXBwbGljYXRpb24iKQogICAgYXBwID0gUUFwcGxpY2F0aW9uKHN5cy5hcmd2KQogICAgYXBwLnNl"
    "dEFwcGxpY2F0aW9uTmFtZShBUFBfTkFNRSkKCiAgICAjIEluc3RhbGwgUXQgbWVzc2FnZSBoYW5kbGVyIE5PVyDigJQgY2F0Y2hl"
    "cyBhbGwgUVRocmVhZC9RdCB3YXJuaW5ncwogICAgIyB3aXRoIGZ1bGwgc3RhY2sgdHJhY2VzIGZyb20gdGhpcyBwb2ludCBmb3J3"
    "YXJkCiAgICBfaW5zdGFsbF9xdF9tZXNzYWdlX2hhbmRsZXIoKQogICAgX2Vhcmx5X2xvZygiW01BSU5dIFFBcHBsaWNhdGlvbiBj"
    "cmVhdGVkLCBtZXNzYWdlIGhhbmRsZXIgaW5zdGFsbGVkIikKCiAgICAjIOKUgOKUgCBQaGFzZSAzOiBGaXJzdCBydW4gY2hlY2sg"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICBpc19maXJzdF9ydW4gPSBDRkcuZ2V0KCJmaXJz"
    "dF9ydW4iLCBUcnVlKQoKICAgIGlmIGlzX2ZpcnN0X3J1bjoKICAgICAgICBkbGcgPSBGaXJzdFJ1bkRpYWxvZygpCiAgICAgICAg"
    "aWYgZGxnLmV4ZWMoKSAhPSBRRGlhbG9nLkRpYWxvZ0NvZGUuQWNjZXB0ZWQ6CiAgICAgICAgICAgIHN5cy5leGl0KDApCgogICAg"
    "ICAgICMg4pSA4pSAIEJ1aWxkIGNvbmZpZyBmcm9tIGRpYWxvZyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDi"
    "lIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKICAgICAg"
    "ICBuZXdfY2ZnID0gZGxnLmJ1aWxkX2NvbmZpZygpCgogICAgICAgICMg4pSA4pSAIERldGVybWluZSBNb3JnYW5uYSdzIGhvbWUg"
    "ZGlyZWN0b3J5IOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgAogICAgICAgICMgQWx3YXlzIGNyZWF0ZXMgRDovQUkvTW9kZWxzL01vcmdhbm5hLyAob3Igc2libGluZyBvZiBzY3Jp"
    "cHQpCiAgICAgICAgc2VlZF9kaXIgICA9IFNDUklQVF9ESVIgICAgICAgICAgIyB3aGVyZSB0aGUgc2VlZCAucHkgbGl2ZXMKICAg"
    "ICAgICBkZWNrX2hvbWUgPSBzZWVkX2RpciAvIERFQ0tfTkFNRQogICAgICAgIGRlY2tfaG9tZS5ta2RpcihwYXJlbnRzPVRydWUs"
    "IGV4aXN0X29rPVRydWUpCgogICAgICAgICMg4pSA4pSAIFVwZGF0ZSBhbGwgcGF0aHMgaW4gY29uZmlnIHRvIHBvaW50IGluc2lk"
    "ZSBkZWNrX2hvbWUg4pSA4pSACiAgICAgICAgbmV3X2NmZ1siYmFzZV9kaXIiXSA9IHN0cihkZWNrX2hvbWUpCiAgICAgICAgbmV3"
    "X2NmZ1sicGF0aHMiXSA9IHsKICAgICAgICAgICAgImZhY2VzIjogICAgc3RyKGRlY2tfaG9tZSAvICJGYWNlcyIpLAogICAgICAg"
    "ICAgICAic291bmRzIjogICBzdHIoZGVja19ob21lIC8gInNvdW5kcyIpLAogICAgICAgICAgICAibWVtb3JpZXMiOiBzdHIoZGVj"
    "a19ob21lIC8gIm1lbW9yaWVzIiksCiAgICAgICAgICAgICJzZXNzaW9ucyI6IHN0cihkZWNrX2hvbWUgLyAic2Vzc2lvbnMiKSwK"
    "ICAgICAgICAgICAgInNsIjogICAgICAgc3RyKGRlY2tfaG9tZSAvICJzbCIpLAogICAgICAgICAgICAiZXhwb3J0cyI6ICBzdHIo"
    "ZGVja19ob21lIC8gImV4cG9ydHMiKSwKICAgICAgICAgICAgImxvZ3MiOiAgICAgc3RyKGRlY2tfaG9tZSAvICJsb2dzIiksCiAg"
    "ICAgICAgICAgICJiYWNrdXBzIjogIHN0cihkZWNrX2hvbWUgLyAiYmFja3VwcyIpLAogICAgICAgICAgICAicGVyc29uYXMiOiBz"
    "dHIoZGVja19ob21lIC8gInBlcnNvbmFzIiksCiAgICAgICAgfQogICAgICAgIG5ld19jZmdbImZpcnN0X3J1biJdID0gRmFsc2UK"
    "CiAgICAgICAgIyDilIDilIAgQ29weSBkZWNrIGZpbGUgaW50byBkZWNrX2hvbWUg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgc3JjX2RlY2sgPSBQYXRo"
    "KF9fZmlsZV9fKS5yZXNvbHZlKCkKICAgICAgICBkc3RfZGVjayA9IGRlY2tfaG9tZSAvIGYie0RFQ0tfTkFNRS5sb3dlcigpfV9k"
    "ZWNrLnB5IgogICAgICAgIGlmIHNyY19kZWNrICE9IGRzdF9kZWNrOgogICAgICAgICAgICB0cnk6CiAgICAgICAgICAgICAgICBf"
    "c2h1dGlsLmNvcHkyKHN0cihzcmNfZGVjayksIHN0cihkc3RfZGVjaykpCiAgICAgICAgICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMg"
    "ZToKICAgICAgICAgICAgICAgIFFNZXNzYWdlQm94Lndhcm5pbmcoCiAgICAgICAgICAgICAgICAgICAgTm9uZSwgIkNvcHkgV2Fy"
    "bmluZyIsCiAgICAgICAgICAgICAgICAgICAgZiJDb3VsZCBub3QgY29weSBkZWNrIGZpbGUgdG8ge0RFQ0tfTkFNRX0gZm9sZGVy"
    "Olxue2V9XG5cbiIKICAgICAgICAgICAgICAgICAgICBmIllvdSBtYXkgbmVlZCB0byBjb3B5IGl0IG1hbnVhbGx5LiIKICAgICAg"
    "ICAgICAgICAgICkKCiAgICAgICAgIyDilIDilIAgV3JpdGUgY29uZmlnLmpzb24gaW50byBkZWNrX2hvbWUg4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgY2ZnX2RzdCA9"
    "IGRlY2tfaG9tZSAvICJjb25maWcuanNvbiIKICAgICAgICBjZmdfZHN0LnBhcmVudC5ta2RpcihwYXJlbnRzPVRydWUsIGV4aXN0"
    "X29rPVRydWUpCiAgICAgICAgd2l0aCBjZmdfZHN0Lm9wZW4oInciLCBlbmNvZGluZz0idXRmLTgiKSBhcyBmOgogICAgICAgICAg"
    "ICBqc29uLmR1bXAobmV3X2NmZywgZiwgaW5kZW50PTIpCgogICAgICAgICMg4pSA4pSAIEJvb3RzdHJhcCBhbGwgc3ViZGlyZWN0"
    "b3JpZXMg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgIyBUZW1wb3JhcmlseSB1cGRhdGUgZ2xvYmFsIENGRyBzbyBib290c3Ry"
    "YXAgZnVuY3Rpb25zIHVzZSBuZXcgcGF0aHMKICAgICAgICBDRkcudXBkYXRlKG5ld19jZmcpCiAgICAgICAgYm9vdHN0cmFwX2Rp"
    "cmVjdG9yaWVzKCkKICAgICAgICBib290c3RyYXBfc291bmRzKCkKICAgICAgICB3cml0ZV9yZXF1aXJlbWVudHNfdHh0KCkKCiAg"
    "ICAgICAgIyDilIDilIAgVW5wYWNrIGZhY2UgWklQIGlmIHByb3ZpZGVkIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIGZh"
    "Y2VfemlwID0gZGxnLmZhY2VfemlwX3BhdGgKICAgICAgICBpZiBmYWNlX3ppcCBhbmQgUGF0aChmYWNlX3ppcCkuZXhpc3RzKCk6"
    "CiAgICAgICAgICAgIGltcG9ydCB6aXBmaWxlIGFzIF96aXBmaWxlCiAgICAgICAgICAgIGZhY2VzX2RpciA9IGRlY2tfaG9tZSAv"
    "ICJGYWNlcyIKICAgICAgICAgICAgZmFjZXNfZGlyLm1rZGlyKHBhcmVudHM9VHJ1ZSwgZXhpc3Rfb2s9VHJ1ZSkKICAgICAgICAg"
    "ICAgdHJ5OgogICAgICAgICAgICAgICAgd2l0aCBfemlwZmlsZS5aaXBGaWxlKGZhY2VfemlwLCAiciIpIGFzIHpmOgogICAgICAg"
    "ICAgICAgICAgICAgIGV4dHJhY3RlZCA9IDAKICAgICAgICAgICAgICAgICAgICBmb3IgbWVtYmVyIGluIHpmLm5hbWVsaXN0KCk6"
    "CiAgICAgICAgICAgICAgICAgICAgICAgIGlmIG1lbWJlci5sb3dlcigpLmVuZHN3aXRoKCIucG5nIik6CiAgICAgICAgICAgICAg"
    "ICAgICAgICAgICAgICBmaWxlbmFtZSA9IFBhdGgobWVtYmVyKS5uYW1lCiAgICAgICAgICAgICAgICAgICAgICAgICAgICB0YXJn"
    "ZXQgPSBmYWNlc19kaXIgLyBmaWxlbmFtZQogICAgICAgICAgICAgICAgICAgICAgICAgICAgd2l0aCB6Zi5vcGVuKG1lbWJlcikg"
    "YXMgc3JjLCB0YXJnZXQub3Blbigid2IiKSBhcyBkc3Q6CiAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgZHN0LndyaXRl"
    "KHNyYy5yZWFkKCkpCiAgICAgICAgICAgICAgICAgICAgICAgICAgICBleHRyYWN0ZWQgKz0gMQogICAgICAgICAgICAgICAgX2Vh"
    "cmx5X2xvZyhmIltGQUNFU10gRXh0cmFjdGVkIHtleHRyYWN0ZWR9IGZhY2UgaW1hZ2VzIHRvIHtmYWNlc19kaXJ9IikKICAgICAg"
    "ICAgICAgZXhjZXB0IEV4Y2VwdGlvbiBhcyBlOgogICAgICAgICAgICAgICAgX2Vhcmx5X2xvZyhmIltGQUNFU10gWklQIGV4dHJh"
    "Y3Rpb24gZmFpbGVkOiB7ZX0iKQogICAgICAgICAgICAgICAgUU1lc3NhZ2VCb3gud2FybmluZygKICAgICAgICAgICAgICAgICAg"
    "ICBOb25lLCAiRmFjZSBQYWNrIFdhcm5pbmciLAogICAgICAgICAgICAgICAgICAgIGYiQ291bGQgbm90IGV4dHJhY3QgZmFjZSBw"
    "YWNrOlxue2V9XG5cbiIKICAgICAgICAgICAgICAgICAgICBmIllvdSBjYW4gYWRkIGZhY2VzIG1hbnVhbGx5IHRvOlxue2ZhY2Vz"
    "X2Rpcn0iCiAgICAgICAgICAgICAgICApCgogICAgICAgICMg4pSA4pSAIENyZWF0ZSBkZXNrdG9wIHNob3J0Y3V0IHBvaW50aW5n"
    "IHRvIG5ldyBkZWNrIGxvY2F0aW9uIOKUgOKUgOKUgOKUgOKUgOKUgAogICAgICAgIHNob3J0Y3V0X2NyZWF0ZWQgPSBGYWxzZQog"
    "ICAgICAgIGlmIGRsZy5jcmVhdGVfc2hvcnRjdXQ6CiAgICAgICAgICAgIHRyeToKICAgICAgICAgICAgICAgIGlmIFdJTjMyX09L"
    "OgogICAgICAgICAgICAgICAgICAgIGltcG9ydCB3aW4zMmNvbS5jbGllbnQgYXMgX3dpbjMyCiAgICAgICAgICAgICAgICAgICAg"
    "ZGVza3RvcCAgICAgPSBQYXRoLmhvbWUoKSAvICJEZXNrdG9wIgogICAgICAgICAgICAgICAgICAgIHNjX3BhdGggICAgID0gZGVz"
    "a3RvcCAvIGYie0RFQ0tfTkFNRX0ubG5rIgogICAgICAgICAgICAgICAgICAgIHB5dGhvbncgICAgID0gUGF0aChzeXMuZXhlY3V0"
    "YWJsZSkKICAgICAgICAgICAgICAgICAgICBpZiBweXRob253Lm5hbWUubG93ZXIoKSA9PSAicHl0aG9uLmV4ZSI6CiAgICAgICAg"
    "ICAgICAgICAgICAgICAgIHB5dGhvbncgPSBweXRob253LnBhcmVudCAvICJweXRob253LmV4ZSIKICAgICAgICAgICAgICAgICAg"
    "ICBpZiBub3QgcHl0aG9udy5leGlzdHMoKToKICAgICAgICAgICAgICAgICAgICAgICAgcHl0aG9udyA9IFBhdGgoc3lzLmV4ZWN1"
    "dGFibGUpCiAgICAgICAgICAgICAgICAgICAgc2hlbGwgPSBfd2luMzIuRGlzcGF0Y2goIldTY3JpcHQuU2hlbGwiKQogICAgICAg"
    "ICAgICAgICAgICAgIHNjICAgID0gc2hlbGwuQ3JlYXRlU2hvcnRDdXQoc3RyKHNjX3BhdGgpKQogICAgICAgICAgICAgICAgICAg"
    "IHNjLlRhcmdldFBhdGggICAgICA9IHN0cihweXRob253KQogICAgICAgICAgICAgICAgICAgIHNjLkFyZ3VtZW50cyAgICAgICA9"
    "IGYnIntkc3RfZGVja30iJwogICAgICAgICAgICAgICAgICAgIHNjLldvcmtpbmdEaXJlY3Rvcnk9IHN0cihkZWNrX2hvbWUpCiAg"
    "ICAgICAgICAgICAgICAgICAgc2MuRGVzY3JpcHRpb24gICAgID0gZiJ7REVDS19OQU1FfSDigJQgRWNobyBEZWNrIgogICAgICAg"
    "ICAgICAgICAgICAgIHNjLnNhdmUoKQogICAgICAgICAgICAgICAgICAgIHNob3J0Y3V0X2NyZWF0ZWQgPSBUcnVlCiAgICAgICAg"
    "ICAgIGV4Y2VwdCBFeGNlcHRpb24gYXMgZToKICAgICAgICAgICAgICAgIHByaW50KGYiW1NIT1JUQ1VUXSBDb3VsZCBub3QgY3Jl"
    "YXRlIHNob3J0Y3V0OiB7ZX0iKQoKICAgICAgICAjIOKUgOKUgCBDb21wbGV0aW9uIG1lc3NhZ2Ug4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA"
    "4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgc2hvcnRjdXRfbm90ZSA9ICgKICAgICAgICAgICAgIkEgZGVz"
    "a3RvcCBzaG9ydGN1dCBoYXMgYmVlbiBjcmVhdGVkLlxuIgogICAgICAgICAgICBmIlVzZSBpdCB0byBzdW1tb24ge0RFQ0tfTkFN"
    "RX0gZnJvbSBub3cgb24uIgogICAgICAgICAgICBpZiBzaG9ydGN1dF9jcmVhdGVkIGVsc2UKICAgICAgICAgICAgIk5vIHNob3J0"
    "Y3V0IHdhcyBjcmVhdGVkLlxuIgogICAgICAgICAgICBmIlJ1biB7REVDS19OQU1FfSBieSBkb3VibGUtY2xpY2tpbmc6XG57ZHN0"
    "X2RlY2t9IgogICAgICAgICkKCiAgICAgICAgUU1lc3NhZ2VCb3guaW5mb3JtYXRpb24oCiAgICAgICAgICAgIE5vbmUsCiAgICAg"
    "ICAgICAgIGYi4pymIHtERUNLX05BTUV9J3MgU2FuY3R1bSBQcmVwYXJlZCIsCiAgICAgICAgICAgIGYie0RFQ0tfTkFNRX0ncyBz"
    "YW5jdHVtIGhhcyBiZWVuIHByZXBhcmVkIGF0OlxuXG4iCiAgICAgICAgICAgIGYie2RlY2tfaG9tZX1cblxuIgogICAgICAgICAg"
    "ICBmIntzaG9ydGN1dF9ub3RlfVxuXG4iCiAgICAgICAgICAgIGYiVGhpcyBzZXR1cCB3aW5kb3cgd2lsbCBub3cgY2xvc2UuXG4i"
    "CiAgICAgICAgICAgIGYiVXNlIHRoZSBzaG9ydGN1dCBvciB0aGUgZGVjayBmaWxlIHRvIGxhdW5jaCB7REVDS19OQU1FfS4iCiAg"
    "ICAgICAgKQoKICAgICAgICAjIOKUgOKUgCBFeGl0IHNlZWQg4oCUIHVzZXIgbGF1bmNoZXMgZnJvbSBzaG9ydGN1dC9uZXcgbG9j"
    "YXRpb24g4pSA4pSA4pSA4pSA4pSA4pSA4pSACiAgICAgICAgc3lzLmV4aXQoMCkKCiAgICAjIOKUgOKUgCBQaGFzZSA0OiBOb3Jt"
    "YWwgbGF1bmNoIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAogICAgIyBPbmx5IHJlYWNo"
    "ZXMgaGVyZSBvbiBzdWJzZXF1ZW50IHJ1bnMgZnJvbSBkZWNrX2hvbWUKICAgIGJvb3RzdHJhcF9zb3VuZHMoKQoKICAgIF9lYXJs"
    "eV9sb2coZiJbTUFJTl0gQ3JlYXRpbmcge0RFQ0tfTkFNRX0gZGVjayB3aW5kb3ciKQogICAgd2luZG93ID0gRWNob0RlY2soKQog"
    "ICAgX2Vhcmx5X2xvZyhmIltNQUlOXSB7REVDS19OQU1FfSBkZWNrIGNyZWF0ZWQg4oCUIGNhbGxpbmcgc2hvdygpIikKICAgIHdp"
    "bmRvdy5zaG93KCkKICAgIF9lYXJseV9sb2coIltNQUlOXSB3aW5kb3cuc2hvdygpIGNhbGxlZCDigJQgZXZlbnQgbG9vcCBzdGFy"
    "dGluZyIpCgogICAgIyBEZWZlciBzY2hlZHVsZXIgYW5kIHN0YXJ0dXAgc2VxdWVuY2UgdW50aWwgZXZlbnQgbG9vcCBpcyBydW5u"
    "aW5nLgogICAgIyBOb3RoaW5nIHRoYXQgc3RhcnRzIHRocmVhZHMgb3IgZW1pdHMgc2lnbmFscyBzaG91bGQgcnVuIGJlZm9yZSB0"
    "aGlzLgogICAgUVRpbWVyLnNpbmdsZVNob3QoMjAwLCBsYW1iZGE6IChfZWFybHlfbG9nKCJbVElNRVJdIF9zZXR1cF9zY2hlZHVs"
    "ZXIgZmlyaW5nIiksIHdpbmRvdy5fc2V0dXBfc2NoZWR1bGVyKCkpKQogICAgUVRpbWVyLnNpbmdsZVNob3QoNDAwLCBsYW1iZGE6"
    "IChfZWFybHlfbG9nKCJbVElNRVJdIHN0YXJ0X3NjaGVkdWxlciBmaXJpbmciKSwgd2luZG93LnN0YXJ0X3NjaGVkdWxlcigpKSkK"
    "ICAgIFFUaW1lci5zaW5nbGVTaG90KDYwMCwgbGFtYmRhOiAoX2Vhcmx5X2xvZygiW1RJTUVSXSBfc3RhcnR1cF9zZXF1ZW5jZSBm"
    "aXJpbmciKSwgd2luZG93Ll9zdGFydHVwX3NlcXVlbmNlKCkpKQoKICAgICMgUGxheSBzdGFydHVwIHNvdW5kIOKAlCBrZWVwIHJl"
    "ZmVyZW5jZSB0byBwcmV2ZW50IEdDIHdoaWxlIHRocmVhZCBydW5zCiAgICBkZWYgX3BsYXlfc3RhcnR1cCgpOgogICAgICAgIHdp"
    "bmRvdy5fc3RhcnR1cF9zb3VuZCA9IFNvdW5kV29ya2VyKCJzdGFydHVwIikKICAgICAgICB3aW5kb3cuX3N0YXJ0dXBfc291bmQu"
    "ZmluaXNoZWQuY29ubmVjdCh3aW5kb3cuX3N0YXJ0dXBfc291bmQuZGVsZXRlTGF0ZXIpCiAgICAgICAgd2luZG93Ll9zdGFydHVw"
    "X3NvdW5kLnN0YXJ0KCkKICAgIFFUaW1lci5zaW5nbGVTaG90KDEyMDAsIF9wbGF5X3N0YXJ0dXApCgogICAgc3lzLmV4aXQoYXBw"
    "LmV4ZWMoKSkKCgppZiBfX25hbWVfXyA9PSAiX19tYWluX18iOgogICAgbWFpbigpCgoKIyDilIDilIAgUEFTUyA2IENPTVBMRVRF"
    "IOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKU"
    "gOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAojIEZ1bGwgZGVjayBhc3NlbWJsZWQuIEFsbCBwYXNzZXMgY29tcGxldGUu"
    "CiMgQ29tYmluZSBhbGwgcGFzc2VzIGludG8gbW9yZ2FubmFfZGVjay5weSBpbiBvcmRlcjoKIyAgIFBhc3MgMSDihpIgUGFzcyAy"
    "IOKGkiBQYXNzIDMg4oaSIFBhc3MgNCDihpIgUGFzcyA1IOKGkiBQYXNzIDYK"
)


def _patch_embedded_deck_implementation(source: str, log_fn=None) -> str:
    """
    Apply additive runtime patches to the embedded deck implementation text.
    This keeps deck_builder.py as the single source of truth without manually
    re-encoding the large embedded implementation blob.
    """

    def _replace_once(text: str, old: str, new: str, label: str) -> str:
        if old not in text:
            if log_fn:
                log_fn(f"[DECK][WARN] Patch target not found: {label}")
            return text
        return text.replace(old, new, 1)

    source = _replace_once(
        source,
        "import random\nimport threading\nimport urllib.request\nimport uuid\nfrom datetime import datetime, date, timedelta, timezone\n",
        "import random\nimport threading\nimport urllib.request\nimport uuid\nimport ast\nimport operator\nimport html\nfrom datetime import datetime, date, timedelta, timezone\n",
        "runtime imports for chat input helpers",
    )

    source = _replace_once(
        source,
        "    QGridLayout, QTextEdit, QLineEdit, QPushButton, QLabel, QFrame,\n",
        "    QGridLayout, QTextEdit, QPlainTextEdit, QLineEdit, QPushButton, QLabel, QFrame,\n",
        "QtWidgets import QPlainTextEdit",
    )
    source = _replace_once(
        source,
        """class EchoDeck(QMainWindow):
""",
        """class DeckChatInput(QTextEdit):
    send_requested = Signal()
    drop_warning = Signal(str)

    _SUPPORTED_TEXT_EXTS = {".txt", ".md", ".json", ".py", ".log"}
    _MAX_DROP_BYTES = 200_000

    def __init__(self, parent=None):
        super().__init__(parent)
        self._min_lines = 1
        self._max_lines = 6
        self.setAcceptDrops(True)
        self.setTabChangesFocus(False)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.textChanged.connect(self._recompute_height)
        self._recompute_height()

    def set_line_limits(self, min_lines: int = 1, max_lines: int = 6) -> None:
        self._min_lines = max(1, int(min_lines))
        self._max_lines = max(self._min_lines, int(max_lines))
        self._recompute_height()

    def _line_height(self) -> int:
        return max(1, self.fontMetrics().lineSpacing())

    def _height_for_lines(self, lines: int) -> int:
        margins = int(self.contentsMargins().top() + self.contentsMargins().bottom())
        doc_margin = int(self.document().documentMargin() * 2)
        frame = int(self.frameWidth() * 2)
        return (self._line_height() * max(1, lines)) + margins + doc_margin + frame

    def _recompute_height(self) -> None:
        doc_size = self.document().documentLayout().documentSize()
        doc_h = int(math.ceil(doc_size.height()))
        min_h = self._height_for_lines(self._min_lines)
        max_h = self._height_for_lines(self._max_lines)
        target = max(min_h, min(max_h, doc_h))
        self.setMinimumHeight(min_h)
        self.setMaximumHeight(max_h)
        self.setFixedHeight(target)
        policy = (
            Qt.ScrollBarPolicy.ScrollBarAsNeeded
            if doc_h > max_h else
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff
        )
        self.setVerticalScrollBarPolicy(policy)

    def keyPressEvent(self, event):
        is_enter = event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter)
        if is_enter:
            mods = event.modifiers()
            if mods & Qt.KeyboardModifier.ControlModifier:
                super().keyPressEvent(event)
                return
            if mods & Qt.KeyboardModifier.ShiftModifier:
                super().keyPressEvent(event)
                return
            self.send_requested.emit()
            event.accept()
            return
        super().keyPressEvent(event)

    def dragEnterEvent(self, event):
        md = event.mimeData()
        if md.hasText():
            event.acceptProposedAction()
            return
        if md.hasUrls():
            for url in md.urls():
                if url.isLocalFile():
                    ext = Path(url.toLocalFile()).suffix.lower()
                    if ext in self._SUPPORTED_TEXT_EXTS:
                        event.acceptProposedAction()
                        return
        event.ignore()

    def dropEvent(self, event):
        md = event.mimeData()
        inserted_parts: list[str] = []

        if md.hasText():
            text = md.text()
            if text:
                inserted_parts.append(text)

        if md.hasUrls():
            for url in md.urls():
                if not url.isLocalFile():
                    continue
                p = Path(url.toLocalFile())
                if p.suffix.lower() not in self._SUPPORTED_TEXT_EXTS:
                    self.drop_warning.emit(f"Unsupported drop type ignored: {p.name}")
                    continue
                try:
                    if p.stat().st_size > self._MAX_DROP_BYTES:
                        kb = self._MAX_DROP_BYTES // 1024
                        self.drop_warning.emit(
                            f"Drop skipped: {p.name} exceeds {kb}KB safety limit."
                        )
                        continue
                    try:
                        blob = p.read_text(encoding="utf-8")
                    except UnicodeDecodeError:
                        blob = p.read_text(encoding="latin-1", errors="replace")
                    if blob:
                        inserted_parts.append(blob)
                except Exception as ex:
                    self.drop_warning.emit(f"Could not read dropped file {p.name}: {ex}")

        if inserted_parts:
            text_to_insert = "\\n".join(x for x in inserted_parts if x)
            cursor = self.textCursor()
            cursor.insertText(text_to_insert)
            self.setTextCursor(cursor)
            event.acceptProposedAction()
            return
        event.ignore()


class DeckSlashCommandDispatcher:
    _HELP_LINES = [
        "/help",
        "/modules",
        "/module <name>",
        "/clear",
        "/newchat",
        "/history",
        "/notes",
        "/system",
        "/calc <expression>",
    ]

    def dispatch(self, deck: "EchoDeck", text: str) -> bool:
        raw = (text or "").strip()
        if not raw.startswith("/"):
            return False
        body = raw[1:].strip()
        if not body:
            deck._append_chat("SYSTEM", "Command expected after '/'. Try /help.")
            return True

        parts = body.split(maxsplit=1)
        cmd = parts[0].lower()
        args = parts[1] if len(parts) > 1 else ""

        if cmd == "help":
            deck._append_chat("SYSTEM", "Available commands: " + ", ".join(self._HELP_LINES))
            return True
        if cmd == "modules":
            deck._slash_list_modules()
            return True
        if cmd == "module":
            deck._slash_open_module(args)
            return True
        if cmd == "clear":
            deck._input_field.clear()
            return True
        if cmd == "newchat":
            deck._slash_new_chat()
            return True
        if cmd == "history":
            deck._slash_history()
            return True
        if cmd == "notes":
            deck._slash_open_notes()
            return True
        if cmd == "system":
            deck._slash_open_system()
            return True
        if cmd == "calc":
            deck._slash_calc(args)
            return True

        deck._append_chat("SYSTEM", f"Unknown command: /{cmd}. Use /help.")
        return True


class EchoDeck(QMainWindow):
""",
        "inject multiline chat input + slash command classes",
    )
    source = _replace_once(
        source,
        """        # ── Input row ──────────────────────────────────────────────────
        input_row = QHBoxLayout()
        prompt_sym = QLabel("✦")
        prompt_sym.setStyleSheet(
            f"color: {C_CRIMSON}; font-size: 16px; font-weight: bold; border: none;"
        )
        prompt_sym.setFixedWidth(20)

        self._input_field = QLineEdit()
        self._input_field.setPlaceholderText(UI_INPUT_PLACEHOLDER)
        self._input_field.returnPressed.connect(self._send_message)
        self._input_field.setEnabled(False)

        self._send_btn = QPushButton(UI_SEND_BUTTON)
        self._send_btn.setFixedWidth(110)
        self._send_btn.clicked.connect(self._send_message)
        self._send_btn.setEnabled(False)

        input_row.addWidget(prompt_sym)
        input_row.addWidget(self._input_field)
        input_row.addWidget(self._send_btn)
        layout.addLayout(input_row)
""",
        """        # ── Input row ──────────────────────────────────────────────────
        input_container = QVBoxLayout()
        input_container.setContentsMargins(0, 0, 0, 0)
        input_container.setSpacing(2)
        input_container.setSizeConstraint(QVBoxLayout.SizeConstraint.SetMinimumSize)

        token_row = QHBoxLayout()
        token_row.setContentsMargins(0, 0, 0, 0)
        token_row.addStretch(1)
        self._prompt_token_label = QLabel("Prompt tokens (est): 0")
        self._prompt_token_label.setStyleSheet(
            f"color: {C_TEXT_DIM}; font-size: 10px; border: none;"
        )
        token_row.addWidget(self._prompt_token_label, 0, Qt.AlignmentFlag.AlignRight)
        input_container.addLayout(token_row)

        input_row = QHBoxLayout()
        input_row.setContentsMargins(0, 0, 0, 0)
        input_row.setSpacing(8)
        input_row.setAlignment(Qt.AlignmentFlag.AlignBottom)
        prompt_sym = QLabel("✦")
        prompt_sym.setStyleSheet(
            f"color: {C_CRIMSON}; font-size: 16px; font-weight: bold; border: none;"
        )
        prompt_sym.setFixedWidth(20)

        self._input_field = DeckChatInput()
        self._input_field.setPlaceholderText(UI_INPUT_PLACEHOLDER)
        self._input_field.set_line_limits(1, 6)
        self._input_field.send_requested.connect(self._send_message)
        self._input_field.textChanged.connect(self._on_prompt_text_changed)
        self._input_field.drop_warning.connect(
            lambda msg: self._append_chat("SYSTEM", msg)
        )
        self._input_field.setEnabled(False)

        self._send_btn = QPushButton(UI_SEND_BUTTON)
        self._send_btn.setFixedWidth(110)
        self._send_btn.clicked.connect(self._send_message)
        self._send_btn.setEnabled(False)

        input_row.addWidget(prompt_sym, 0, Qt.AlignmentFlag.AlignBottom)
        input_row.addWidget(self._input_field, 1)
        input_row.addWidget(self._send_btn, 0, Qt.AlignmentFlag.AlignBottom)
        input_container.addLayout(input_row)
        layout.addLayout(input_container)
""",
        "multiline input row and token counter",
    )
    source = _replace_once(
        source,
        """    # ── MESSAGE HANDLING ───────────────────────────────────────────────────────
    def _send_message(self) -> None:
""",
        """    # ── MESSAGE HANDLING ───────────────────────────────────────────────────────
    def _on_prompt_text_changed(self) -> None:
        text = self._input_field.toPlainText() if self._input_field else ""
        count, exact = self._count_prompt_tokens(text)
        suffix = "exact" if exact else "est"
        if hasattr(self, "_prompt_token_label") and self._prompt_token_label is not None:
            self._prompt_token_label.setText(f"Prompt tokens ({suffix}): {count}")

    def _count_prompt_tokens(self, text: str) -> tuple[int, bool]:
        cleaned = (text or "").strip()
        if not cleaned:
            return 0, False
        tok = getattr(getattr(self, "_adaptor", None), "_tokenizer", None)
        if tok is not None and hasattr(tok, "encode"):
            try:
                return len(tok.encode(cleaned)), True
            except Exception:
                pass
        return max(1, math.ceil(len(cleaned) / 4)), False

    def _slash_list_modules(self) -> None:
        names = [str(d.get("title", "")).strip() for d in getattr(self, "_spell_tab_defs", [])]
        names = [n for n in names if n]
        if names:
            self._append_chat("SYSTEM", "Available modules: " + ", ".join(names))
        else:
            self._append_chat("SYSTEM", "No module listing is available yet.")

    def _slash_open_module(self, module_name: str) -> None:
        target = module_name.strip().lower()
        if not target:
            self._append_chat("SYSTEM", "Usage: /module <name>")
            return
        for i in range(self._spell_tabs.count()):
            title = self._spell_tabs.tabText(i).lower()
            tab_id = str(self._spell_tabs.tabBar().tabData(i) or "").lower()
            if target in title or target == tab_id:
                self._spell_tabs.setCurrentIndex(i)
                self._append_chat("SYSTEM", f"Module focused: {self._spell_tabs.tabText(i)}")
                return
        self._append_chat("SYSTEM", f"Module not found: {module_name}")

    def _slash_new_chat(self) -> None:
        try:
            self._sessions.save()
        except Exception:
            pass
        self._sessions = SessionManager()
        self._session_id = self._sessions.session_id
        self._chat_display.clear()
        self._append_chat("SYSTEM", "A new chat session begins.")
        if hasattr(self, "_journal_sidebar") and self._journal_sidebar is not None:
            self._journal_sidebar.clear_journal_indicator()

    def _slash_history(self) -> None:
        sessions = self._sessions.list_sessions()[:5]
        if not sessions:
            self._append_chat("SYSTEM", "No saved sessions found.")
            return
        lines = [f"{s.get('date','?')} — {s.get('name','(unnamed)')}" for s in sessions]
        self._append_chat("SYSTEM", "Recent sessions: " + " | ".join(lines))

    def _slash_open_notes(self) -> None:
        self._slash_open_module("lessons")

    def _slash_open_system(self) -> None:
        self._slash_open_module("instruments")

    def _safe_calc(self, expr: str) -> float:
        allowed_bin = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.Mod: operator.mod,
            ast.Pow: operator.pow,
        }
        allowed_unary = {
            ast.UAdd: operator.pos,
            ast.USub: operator.neg,
        }
        node = ast.parse(expr, mode="eval")

        def _eval(n):
            if isinstance(n, ast.Expression):
                return _eval(n.body)
            if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
                return float(n.value)
            if isinstance(n, ast.BinOp) and type(n.op) in allowed_bin:
                return allowed_bin[type(n.op)](_eval(n.left), _eval(n.right))
            if isinstance(n, ast.UnaryOp) and type(n.op) in allowed_unary:
                return allowed_unary[type(n.op)](_eval(n.operand))
            raise ValueError("Unsupported expression")

        return _eval(node)

    def _slash_calc(self, expr: str) -> None:
        raw = expr.strip()
        if not raw:
            self._append_chat("SYSTEM", "Usage: /calc <expression>")
            return
        try:
            result = self._safe_calc(raw)
            self._append_chat("SYSTEM", f"Calc: {raw} = {result}")
        except Exception as ex:
            self._append_chat("SYSTEM", f"Calc error: {ex}")

    def _send_message(self) -> None:
""",
        "inject prompt token + slash helpers",
    )
    source = _replace_once(
        source,
        """        text = self._input_field.text().strip()
        if not text:
            return
""",
        """        raw_text = self._input_field.toPlainText()
        text = raw_text.strip()
        if not text:
            return

        if self._slash_dispatcher.dispatch(self, text):
            self._on_prompt_text_changed()
            return
""",
        "send_message multiline text + slash intercept",
    )
    source = _replace_once(
        source,
        """        self._input_field.clear()
        self._append_chat("YOU", text)
""",
        """        self._input_field.clear()
        self._on_prompt_text_changed()
        self._append_chat("YOU", text)
""",
        "send_message prompt counter refresh after clear",
    )
    source = _replace_once(
        source,
        """        self._token_count         = 0
""",
        """        self._token_count         = 0
        self._slash_dispatcher  = DeckSlashCommandDispatcher()
""",
        "echo deck init slash dispatcher",
    )
    source = _replace_once(
        source,
        """        window._input_field.setText(line)
        window._input_field.setFocus()
""",
        """        if hasattr(window._input_field, "setPlainText"):
            window._input_field.setPlainText(line)
        else:
            window._input_field.setText(line)
        window._input_field.setFocus()
        if hasattr(window, "_on_prompt_text_changed"):
            window._on_prompt_text_changed()
""",
        "send roll result to multiline prompt",
    )
    source = _replace_once(
        source,
        """                    self._input_field.setText(date_text)
                    routed_target = "input_field_set"
""",
        """                    if hasattr(self._input_field, "setPlainText"):
                        self._input_field.setPlainText(date_text)
                    else:
                        self._input_field.setText(date_text)
                    routed_target = "input_field_set"
""",
        "calendar route to multiline prompt",
    )
    source = _replace_once(
        source,
        '            return [x for x in data if isinstance(x, dict)]',
        '            return [_normalize_jsonl_record(path, x)\n'
        '                    for x in data if isinstance(x, dict)]',
        "read_jsonl array mode fallback",
    )
    source = _replace_once(
        source,
        "    # ── CHAT DISPLAY ───────────────────────────────────────────────────────────\n"
        "    def _append_chat(self, speaker: str, text: str) -> None:\n"
        "        colors = {\n"
        "            \"YOU\":     C_GOLD,\n"
        "            DECK_NAME.upper():C_GOLD,\n"
        "            \"SYSTEM\":  C_PURPLE,\n"
        "            \"ERROR\":   C_BLOOD,\n"
        "        }\n"
        "        label_colors = {\n"
        "            \"YOU\":     C_GOLD_DIM,\n"
        "            DECK_NAME.upper():C_CRIMSON,\n"
        "            \"SYSTEM\":  C_PURPLE,\n"
        "            \"ERROR\":   C_BLOOD,\n"
        "        }\n"
        "        color       = colors.get(speaker, C_GOLD)\n"
        "        label_color = label_colors.get(speaker, C_GOLD_DIM)\n"
        "        timestamp   = datetime.now().strftime(\"%H:%M:%S\")\n"
        "\n"
        "        if speaker == \"SYSTEM\":\n"
        "            self._chat_display.append(\n"
        "                f'<span style=\"color:{C_TEXT_DIM}; font-size:10px;\">'\n"
        "                f'[{timestamp}] </span>'\n"
        "                f'<span style=\"color:{label_color};\">✦ {text}</span>'\n"
        "            )\n"
        "        else:\n"
        "            self._chat_display.append(\n"
        "                f'<span style=\"color:{C_TEXT_DIM}; font-size:10px;\">'\n"
        "                f'[{timestamp}] </span>'\n"
        "                f'<span style=\"color:{label_color}; font-weight:bold;\">'\n"
        "                f'{speaker} ❧</span> '\n"
        "                f'<span style=\"color:{color};\">{text}</span>'\n"
        "            )\n"
        "\n"
        "        # Add blank line after Morganna's response (not during streaming)\n"
        "        if speaker == DECK_NAME.upper():\n"
        "            self._chat_display.append(\"\")\n"
        "\n"
        "        self._chat_display.verticalScrollBar().setValue(\n"
        "            self._chat_display.verticalScrollBar().maximum()\n"
        "        )\n",
        "    # ── CHAT DISPLAY ───────────────────────────────────────────────────────────\n"
        "    def _append_chat(self, speaker: str, text: str) -> None:\n"
        "        colors = {\n"
        "            \"YOU\":     C_GOLD,\n"
        "            DECK_NAME.upper():C_GOLD,\n"
        "            \"SYSTEM\":  C_PURPLE,\n"
        "            \"ERROR\":   C_BLOOD,\n"
        "        }\n"
        "        label_colors = {\n"
        "            \"YOU\":     C_GOLD_DIM,\n"
        "            DECK_NAME.upper():C_CRIMSON,\n"
        "            \"SYSTEM\":  C_PURPLE,\n"
        "            \"ERROR\":   C_BLOOD,\n"
        "        }\n"
        "        color       = colors.get(speaker, C_GOLD)\n"
        "        label_color = label_colors.get(speaker, C_GOLD_DIM)\n"
        "        timestamp   = datetime.now().strftime(\"%H:%M:%S\")\n"
        "        safe_text   = html.escape(text).replace(\"\\n\", \"<br>\")\n"
        "\n"
        "        if speaker == \"SYSTEM\":\n"
        "            self._chat_display.append(\n"
        "                f'<span style=\"color:{C_TEXT_DIM}; font-size:10px;\">'\n"
        "                f'[{timestamp}] </span>'\n"
        "                f'<span style=\"color:{label_color};\">✦ {safe_text}</span>'\n"
        "            )\n"
        "        else:\n"
        "            self._chat_display.append(\n"
        "                f'<span style=\"color:{C_TEXT_DIM}; font-size:10px;\">'\n"
        "                f'[{timestamp}] </span>'\n"
        "                f'<span style=\"color:{label_color}; font-weight:bold;\">'\n"
        "                f'{speaker} ❧</span> '\n"
        "                f'<span style=\"color:{color};\">{safe_text}</span>'\n"
        "            )\n"
        "\n"
        "        # Add blank line after Morganna's response (not during streaming)\n"
        "        if speaker == DECK_NAME.upper():\n"
        "            self._chat_display.append(\"\")\n"
        "\n"
        "        self._chat_display.verticalScrollBar().setValue(\n"
        "            self._chat_display.verticalScrollBar().maximum()\n"
        "        )\n",
        "chat display preserve multiline user text",
    )
    source = _replace_once(
        source,
        '                items.append(obj)',
        '                items.append(_normalize_jsonl_record(path, obj))',
        "read_jsonl line mode fallback",
    )
    source = _replace_once(
        source,
        "def write_jsonl(path: Path, records: list[dict]) -> None:\n"
        '    """Overwrite a JSONL file with a list of records."""\n'
        '    path.parent.mkdir(parents=True, exist_ok=True)\n'
        '    with path.open("w", encoding="utf-8") as f:\n'
        '        for r in records:\n'
        '            f.write(json.dumps(r, ensure_ascii=False) + "\\n")\n'
        "\n"
        "# ── KEYWORD / MEMORY HELPERS ──────────────────────────────────────────────────\n",
        "def write_jsonl(path: Path, records: list[dict]) -> None:\n"
        '    """Overwrite a JSONL file with a list of records."""\n'
        '    path.parent.mkdir(parents=True, exist_ok=True)\n'
        '    with path.open("w", encoding="utf-8") as f:\n'
        '        for r in records:\n'
        '            f.write(json.dumps(r, ensure_ascii=False) + "\\n")\n'
        "\n"
        "_ALLOWED_RUNTIME_MODES = {\"default\", \"persona\", \"rp\"}\n"
        "\n"
        "def _normalize_runtime_mode(value: object) -> str:\n"
        "    mode = str(value or \"\").strip().lower()\n"
        "    return mode if mode in _ALLOWED_RUNTIME_MODES else \"persona\"\n"
        "\n"
        "def _detect_runtime_mode(cfg: Optional[dict] = None) -> str:\n"
        "    cfg = cfg or CFG\n"
        "    candidates = [\n"
        "        cfg.get(\"mode\"),\n"
        "        cfg.get(\"runtime_mode\"),\n"
        "        cfg.get(\"chat_mode\"),\n"
        "        (cfg.get(\"settings\", {}) or {}).get(\"mode\"),\n"
        "        (cfg.get(\"settings\", {}) or {}).get(\"runtime_mode\"),\n"
        "        (cfg.get(\"settings\", {}) or {}).get(\"chat_mode\"),\n"
        "    ]\n"
        "    for candidate in candidates:\n"
        "        mode = _normalize_runtime_mode(candidate)\n"
        "        if candidate is not None and mode in _ALLOWED_RUNTIME_MODES:\n"
        "            return mode\n"
        "    return \"persona\"\n"
        "\n"
        "def _normalize_jsonl_record(path: Path, record: dict) -> dict:\n"
        "    if path.name not in (\"messages.jsonl\", \"memories.jsonl\"):\n"
        "        return record\n"
        "    normalized = dict(record)\n"
        "    normalized[\"mode\"] = _normalize_runtime_mode(\n"
        "        normalized.get(\"mode\", \"persona\")\n"
        "    )\n"
        "    return normalized\n"
        "\n"
        "_ALLOWED_THEME_MODES = {\"light\", \"auto\", \"dark\"}\n"
        "\n"
        "def _normalize_theme_mode(value: object) -> str:\n"
        "    mode = str(value or \"\").strip().lower()\n"
        "    return mode if mode in _ALLOWED_THEME_MODES else \"auto\"\n"
        "\n"
        "def _detect_theme_mode(cfg: Optional[dict] = None) -> str:\n"
        "    cfg = cfg or CFG\n"
        "    settings = (cfg.get(\"settings\", {}) or {}) if isinstance(cfg, dict) else {}\n"
        "    candidates = [\n"
        "        cfg.get(\"theme_mode\") if isinstance(cfg, dict) else None,\n"
        "        settings.get(\"theme_mode\"),\n"
        "        settings.get(\"runtime_theme_mode\"),\n"
        "    ]\n"
        "    for candidate in candidates:\n"
        "        if candidate is not None:\n"
        "            return _normalize_theme_mode(candidate)\n"
        "    return \"auto\"\n"
        "\n"
        "def _persona_preferred_theme(cfg: Optional[dict] = None) -> Optional[str]:\n"
        "    cfg = cfg or CFG\n"
        "    persona_cfg = (cfg.get(\"persona\", {}) or {}) if isinstance(cfg, dict) else {}\n"
        "    candidates = [\n"
        "        persona_cfg.get(\"preferred_theme\"),\n"
        "        persona_cfg.get(\"theme\"),\n"
        "        persona_cfg.get(\"ui_theme\"),\n"
        "    ]\n"
        "    for candidate in candidates:\n"
        "        mode = str(candidate or \"\").strip().lower()\n"
        "        if mode in {\"light\", \"dark\"}:\n"
        "            return mode\n"
        "    return None\n"
        "\n"
        "def _resolve_effective_theme(cfg: Optional[dict] = None) -> str:\n"
        "    cfg = cfg or CFG\n"
        "    mode = _detect_theme_mode(cfg)\n"
        "    if mode in {\"light\", \"dark\"}:\n"
        "        return mode\n"
        "    pref = _persona_preferred_theme(cfg)\n"
        "    return pref if pref in {\"light\", \"dark\"} else \"dark\"\n"
        "\n"
        "# ── KEYWORD / MEMORY HELPERS ──────────────────────────────────────────────────\n",
        "mode helpers",
    )
    source = _replace_once(
        source,
        '            "emotion":    emotion,\n'
        "        }",
        '            "emotion":    emotion,\n'
        '            "mode":       _detect_runtime_mode(),\n'
        "        }",
        "append_message mode field",
    )
    source = _replace_once(
        source,
        '            "confidence":       0.70 if record_type in {\n'
        '                "dream","issue","idea","preference","resolution"\n'
        '            } else 0.55,\n'
        "        }",
        '            "confidence":       0.70 if record_type in {\n'
        '                "dream","issue","idea","preference","resolution"\n'
        '            } else 0.55,\n'
        '            "mode":             _detect_runtime_mode(),\n'
        "        }",
        "append_memory mode field",
    )
    source = _replace_once(
        source,
        '            "timezone_override":         "",\n'
        '            "fullscreen_enabled":        False,\n'
        '            "borderless_enabled":        False,\n',
        '            "timezone_override":         "",\n'
        '            "runtime_persona_mode":      "persona",\n'
        '            "runtime_theme_mode":        "auto",\n'
        '            "theme_mode":                "auto",\n'
        '            "fullscreen_enabled":        False,\n'
        '            "borderless_enabled":        False,\n',
        "default config runtime mode settings",
    )
    source = _replace_once(
        source,
        "        self._first_token: bool = True   # write speaker label before first streaming token\n",
        "        self._first_token: bool = True   # write speaker label before first streaming token\n"
        "        self._runtime_persona_mode = _normalize_runtime_mode(\n"
        "            (CFG.get(\"settings\", {}) or {}).get(\"runtime_persona_mode\", _detect_runtime_mode())\n"
        "        )\n"
        "        self._runtime_theme_mode = _normalize_theme_mode(_detect_theme_mode())\n"
        "        self._pending_hidden_runtime_events: list[str] = []\n",
        "EchoDeck runtime state init",
    )
    source = _replace_once(
        source,
        "        self.setStyleSheet(STYLE)\n"
        "\n"
        "        self._build_ui()\n",
        "        self.setStyleSheet(STYLE)\n"
        "\n"
        "        self._build_ui()\n"
        "        self._apply_runtime_theme_mode(self._runtime_theme_mode, emit_event=False)\n"
        "        self._apply_runtime_persona_mode(self._runtime_persona_mode, emit_event=False)\n",
        "EchoDeck apply runtime modes after ui build",
    )
    source = _replace_once(
        source,
        "    # ── MESSAGE HANDLING ───────────────────────────────────────────────────────\n",
        "    def _runtime_event_payload(self, event_name: str, value: str, context: str = \"\") -> str:\n"
        "        payload = {\n"
        "            \"event\": event_name,\n"
        "            \"value\": value,\n"
        "            \"context\": context,\n"
        "            \"timestamp\": local_now_iso(),\n"
        "        }\n"
        "        return \"[INTERNAL_EVENT] \" + json.dumps(payload, ensure_ascii=False)\n"
        "\n"
        "    def _queue_hidden_runtime_event(self, event_name: str, value: str, context: str = \"\") -> None:\n"
        "        self._pending_hidden_runtime_events.append(\n"
        "            self._runtime_event_payload(event_name, value, context)\n"
        "        )\n"
        "\n"
        "    def _consume_hidden_runtime_events(self) -> list[str]:\n"
        "        pending = list(self._pending_hidden_runtime_events)\n"
        "        self._pending_hidden_runtime_events.clear()\n"
        "        return pending\n"
        "\n"
        "    def _apply_runtime_persona_mode(self, mode: str, emit_event: bool = True) -> None:\n"
        "        normalized = _normalize_runtime_mode(mode)\n"
        "        if normalized == self._runtime_persona_mode and emit_event:\n"
        "            return\n"
        "        self._runtime_persona_mode = normalized\n"
        "        CFG.setdefault(\"settings\", {})[\"runtime_persona_mode\"] = normalized\n"
        "        CFG[\"runtime_mode\"] = normalized\n"
        "        save_config(CFG)\n"
        "        if hasattr(self, \"_settings_tab\") and self._settings_tab:\n"
        "            self._settings_tab.sync_runtime_controls()\n"
        "        if emit_event:\n"
        "            self._diag_tab.log(f\"[MODE] Persona runtime mode -> {normalized}\", \"INFO\")\n"
        "            self._queue_hidden_runtime_event(\"runtime_persona_mode\", normalized, \"persona runtime mode changed\")\n"
        "\n"
        "    def _theme_roles(self, effective_theme: str) -> dict:\n"
        "        accent = (CFG.get(\"persona\", {}) or {}).get(\"color_primary\") or \"#5b8cff\"\n"
        "        if effective_theme == \"light\":\n"
        "            return {\n"
        "                \"window_bg\":   \"#eef2f8\",\n"
        "                \"sidebar_bg\":  \"#e6ebf3\",\n"
        "                \"panel_bg\":    \"#ffffff\",\n"
        "                \"panel_alt_bg\":\"#f5f8ff\",\n"
        "                \"input_bg\":    \"#ffffff\",\n"
        "                \"border\":      \"#b8c4db\",\n"
        "                \"text\":        \"#171c28\",\n"
        "                \"text_dim\":    \"#5b667d\",\n"
        "                \"button_bg\":   \"#dbe5f7\",\n"
        "                \"button_hover\": \"#ccd9f0\",\n"
        "                \"accent\":      accent,\n"
        "            }\n"
        "        return {\n"
        "            \"window_bg\":   \"#0b0f1a\",\n"
        "            \"sidebar_bg\":  \"#10182a\",\n"
        "            \"panel_bg\":    \"#151f35\",\n"
        "            \"panel_alt_bg\":\"#1a2640\",\n"
        "            \"input_bg\":    \"#111b2f\",\n"
        "            \"border\":      \"#304465\",\n"
        "            \"text\":        \"#dfe6f8\",\n"
        "            \"text_dim\":    \"#9eacca\",\n"
        "            \"button_bg\":   \"#1f2c49\",\n"
        "            \"button_hover\": \"#2a3a5e\",\n"
        "            \"accent\":      accent,\n"
        "        }\n"
        "\n"
        "    def _apply_shell_region_theme(self, roles: dict) -> None:\n"
        "        role_keywords = {\n"
        "            \"sidebar\": {\"left\", \"sidebar\", \"journal\"},\n"
        "            \"workspace\": {\"chat\", \"workspace\", \"center\", \"mainpanel\"},\n"
        "            \"systems\": {\"right\", \"system\", \"module\"},\n"
        "            \"hud\": {\"hud\", \"lower\", \"status\"},\n"
        "            \"calendar\": {\"calendar\"},\n"
        "            \"input\": {\"input\", \"prompt\", \"composer\"},\n"
        "        }\n"
        "        widgets = [self]\n"
        "        widgets.extend(self.findChildren(QWidget))\n"
        "        for w in widgets:\n"
        "            role = \"card\"\n"
        "            name = (w.objectName() or \"\").strip().lower()\n"
        "            if w is self:\n"
        "                role = \"shell_bg\"\n"
        "            elif isinstance(w, (QLineEdit, QTextEdit, QPlainTextEdit, QComboBox)):\n"
        "                role = \"input\"\n"
        "            elif isinstance(w, (QPushButton, QToolButton)):\n"
        "                role = \"button\"\n"
        "            elif name:\n"
        "                for candidate, keys in role_keywords.items():\n"
        "                    if any(k in name for k in keys):\n"
        "                        role = candidate\n"
        "                        break\n"
        "            w.setProperty(\"runtime_theme_role\", role)\n"
        "            w.style().unpolish(w)\n"
        "            w.style().polish(w)\n"
        "            w.update()\n"
        "\n"
        "    def _apply_theme_stylesheet(self, effective_theme: str) -> None:\n"
        "        roles = self._theme_roles(effective_theme)\n"
        "        self._apply_shell_region_theme(roles)\n"
        "        base = STYLE\n"
        "        override = (\n"
        "            \"\\nQMainWindow, QWidget#deck_root, QWidget[runtime_theme_role=\\\"shell_bg\\\"] {\"\n"
        "            f\" background-color: {roles['window_bg']}; color: {roles['text']}; }}\"\n"
        "            \"\\nQWidget[runtime_theme_role=\\\"sidebar\\\"] {\"\n"
        "            f\" background-color: {roles['sidebar_bg']}; color: {roles['text']};\"\n"
        "            f\" border-color: {roles['border']}; }}\"\n"
        "            \"\\nQWidget[runtime_theme_role=\\\"workspace\\\"], QWidget[runtime_theme_role=\\\"systems\\\"],\"\n"
        "            \" QWidget[runtime_theme_role=\\\"hud\\\"], QWidget[runtime_theme_role=\\\"calendar\\\"] {\"\n"
        "            f\" background-color: {roles['panel_bg']}; color: {roles['text']};\"\n"
        "            f\" border-color: {roles['border']}; }}\"\n"
        "            \"\\nQWidget[runtime_theme_role=\\\"card\\\"], QFrame, QGroupBox, QTabWidget::pane {\"\n"
        "            f\" background-color: {roles['panel_alt_bg']}; color: {roles['text']};\"\n"
        "            f\" border: 1px solid {roles['border']}; }}\"\n"
        "            \"\\nQWidget[runtime_theme_role=\\\"input\\\"], QLineEdit, QTextEdit, QPlainTextEdit,\"\n"
        "            \" QComboBox, QListWidget, QTableWidget, QTreeWidget, QCalendarWidget {\"\n"
        "            f\" background-color: {roles['input_bg']}; color: {roles['text']};\"\n"
        "            f\" border: 1px solid {roles['border']}; selection-background-color: {roles['accent']}; }}\"\n"
        "            \"\\nQPushButton, QToolButton, QWidget[runtime_theme_role=\\\"button\\\"] {\"\n"
        "            f\" background-color: {roles['button_bg']}; color: {roles['text']};\"\n"
        "            f\" border: 1px solid {roles['border']}; }}\"\n"
        "            \"\\nQPushButton:hover, QToolButton:hover, QWidget[runtime_theme_role=\\\"button\\\"]:hover {\"\n"
        "            f\" background-color: {roles['button_hover']}; border-color: {roles['accent']}; }}\"\n"
        "            \"\\nQPushButton:checked, QToolButton:checked {\"\n"
        "            f\" border: 1px solid {roles['accent']}; }}\"\n"
        "            \"\\nQLabel {\"\n"
        "            f\" color: {roles['text']}; }}\"\n"
        "            \"\\nQLabel[dim=\\\"true\\\"] {\"\n"
        "            f\" color: {roles['text_dim']}; }}\"\n"
        "            \"\\nQTabBar::tab:selected, QHeaderView::section:selected {\"\n"
        "            f\" border-color: {roles['accent']}; color: {roles['text']}; }}\"\n"
        "            \"\\nQSplitter::handle {\"\n"
        "            f\" background-color: {roles['border']}; }}\"\n"
        "        )\n"
        "        self.setStyleSheet(base + override)\n"
        "\n"
        "    def _apply_runtime_theme_mode(self, mode: str, emit_event: bool = True) -> None:\n"
        "        normalized = _normalize_theme_mode(mode)\n"
        "        if normalized == self._runtime_theme_mode and emit_event:\n"
        "            return\n"
        "        self._runtime_theme_mode = normalized\n"
        "        CFG.setdefault(\"settings\", {})[\"runtime_theme_mode\"] = normalized\n"
        "        CFG.setdefault(\"settings\", {})[\"theme_mode\"] = normalized\n"
        "        CFG[\"theme_mode\"] = normalized\n"
        "        save_config(CFG)\n"
        "        effective = _resolve_effective_theme(CFG)\n"
        "        self._apply_theme_stylesheet(effective)\n"
        "        if hasattr(self, \"_settings_tab\") and self._settings_tab:\n"
        "            self._settings_tab.sync_runtime_controls()\n"
        "        if emit_event:\n"
        "            context = (\n"
        "                \"environment is brightly lit\" if effective == \"light\" else\n"
        "                \"environment is dim\" if effective == \"dark\" else\n"
        "                \"persona-preferred theme restored\"\n"
        "            )\n"
        "            self._diag_tab.log(\n"
        "                f\"[THEME] mode={normalized} effective={effective}\", \"INFO\"\n"
        "            )\n"
        "            self._queue_hidden_runtime_event(\"runtime_theme_mode\", normalized, context)\n"
        "\n"
        "    def _handle_theme_command(self, text: str) -> bool:\n"
        "        if text == \"Theme Light\":\n"
        "            self._apply_runtime_theme_mode(\"light\")\n"
        "            return True\n"
        "        if text == \"Theme Auto\":\n"
        "            self._apply_runtime_theme_mode(\"auto\")\n"
        "            return True\n"
        "        if text == \"Theme Dark\":\n"
        "            self._apply_runtime_theme_mode(\"dark\")\n"
        "            return True\n"
        "        return False\n"
        "\n"
        "    # ── MESSAGE HANDLING ───────────────────────────────────────────────────────\n",
        "runtime mode/theme methods",
    )
    source = _replace_once(
        source,
        "        if not text:\n"
        "            return\n"
        "\n"
        "        # Flip back to persona chat tab from Self tab if needed\n",
        "        if not text:\n"
        "            return\n"
        "        if self._handle_theme_command(text):\n"
        "            self._input_field.clear()\n"
        "            return\n"
        "\n"
        "        # Flip back to persona chat tab from Self tab if needed\n",
        "send message intercept theme command",
    )
    source = _replace_once(
        source,
        "        # Build system prompt\n"
        "        system = SYSTEM_PROMPT_BASE\n"
        "        if memory_ctx:\n"
        "            system += f\"\\n\\n{memory_ctx}\"\n"
        "        if journal_ctx:\n"
        "            system += f\"\\n\\n{journal_ctx}\"\n"
        "        system += vampire_ctx\n",
        "        # Build system prompt\n"
        "        runtime_mode = _normalize_runtime_mode(self._runtime_persona_mode)\n"
        "        if runtime_mode == \"default\":\n"
        "            system = \"\"\n"
        "        else:\n"
        "            system = SYSTEM_PROMPT_BASE\n"
        "            if runtime_mode == \"rp\":\n"
        "                system += (\n"
        "                    \"\\n\\n[RUNTIME PERSONA MODE]\\n\"\n"
        "                    \"Mode: rp\\n\"\n"
        "                    \"Speak as the selected persona as a fully in-world character. \"\n"
        "                    \"Do not frame yourself as an AI assistant.\"\n"
        "                )\n"
        "            else:\n"
        "                system += \"\\n\\n[RUNTIME PERSONA MODE]\\nMode: persona\"\n"
        "        if memory_ctx:\n"
        "            system += f\"\\n\\n{memory_ctx}\"\n"
        "        if journal_ctx:\n"
        "            system += f\"\\n\\n{journal_ctx}\"\n"
        "        system += vampire_ctx\n",
        "send message runtime persona system prompt",
    )
    source = _replace_once(
        source,
        "            self._pending_transmissions = 0\n"
        "            self._suspended_duration    = \"\"\n"
        "\n"
        "        history = self._sessions.get_history()\n",
        "            self._pending_transmissions = 0\n"
        "            self._suspended_duration    = \"\"\n"
        "\n"
        "        history = self._sessions.get_history()\n"
        "        for evt in self._consume_hidden_runtime_events():\n"
        "            history.append({\"role\": \"user\", \"content\": evt})\n",
        "inject hidden runtime events into history",
    )
    source = _replace_once(
        source,
        "    def _build_system_section(self, layout: QVBoxLayout) -> None:\n"
        "        if self._deck._torpor_panel is not None:\n"
        "            layout.addWidget(QLabel(\"Operational Mode\"))\n"
        "            layout.addWidget(self._deck._torpor_panel)\n"
        "\n"
        "        layout.addWidget(QLabel(\"Idle\"))\n"
        "        layout.addWidget(self._deck._idle_btn)\n"
        "\n"
        "        settings = CFG.get(\"settings\", {})\n"
        "        tz_auto = bool(settings.get(\"timezone_auto_detect\", True))\n"
        "        tz_override = str(settings.get(\"timezone_override\", \"\") or \"\").strip()\n"
        "\n"
        "        tz_auto_chk = QCheckBox(\"Auto-detect local/system time zone\")\n"
        "        tz_auto_chk.setChecked(tz_auto)\n"
        "        tz_auto_chk.toggled.connect(self._deck._set_timezone_auto_detect)\n"
        "        layout.addWidget(tz_auto_chk)\n"
        "\n"
        "        tz_row = QHBoxLayout()\n"
        "        tz_row.addWidget(QLabel(\"Manual Time Zone Override:\"))\n"
        "        tz_combo = QComboBox()\n"
        "        tz_combo.setEditable(True)\n"
        "        tz_options = [\n"
        "            \"America/Chicago\", \"America/New_York\", \"America/Los_Angeles\",\n"
        "            \"America/Denver\", \"UTC\"\n"
        "        ]\n"
        "        tz_combo.addItems(tz_options)\n"
        "        if tz_override:\n"
        "            if tz_combo.findText(tz_override) < 0:\n"
        "                tz_combo.addItem(tz_override)\n"
        "            tz_combo.setCurrentText(tz_override)\n"
        "        else:\n"
        "            tz_combo.setCurrentText(\"America/Chicago\")\n"
        "        tz_combo.setEnabled(not tz_auto)\n"
        "        tz_combo.currentTextChanged.connect(self._deck._set_timezone_override)\n"
        "        tz_auto_chk.toggled.connect(lambda enabled: tz_combo.setEnabled(not enabled))\n"
        "        tz_row.addWidget(tz_combo, 1)\n"
        "        tz_host = QWidget()\n"
        "        tz_host.setLayout(tz_row)\n"
        "        layout.addWidget(tz_host)\n",
        "    def _build_system_section(self, layout: QVBoxLayout) -> None:\n"
        "        if self._deck._torpor_panel is not None:\n"
        "            layout.addWidget(QLabel(\"Operational Mode\"))\n"
        "            layout.addWidget(self._deck._torpor_panel)\n"
        "\n"
        "        layout.addWidget(QLabel(\"Idle\"))\n"
        "        layout.addWidget(self._deck._idle_btn)\n"
        "\n"
        "        layout.addWidget(QLabel(\"Persona Runtime Mode\"))\n"
        "        self._persona_mode_buttons = {}\n"
        "        persona_modes = [\n"
        "            (\"default\", \"Default\"),\n"
        "            (\"persona\", \"Persona\"),\n"
        "            (\"rp\", \"RP\"),\n"
        "        ]\n"
        "        persona_row = QHBoxLayout()\n"
        "        for mode_key, label in persona_modes:\n"
        "            btn = QPushButton(label)\n"
        "            btn.setCheckable(True)\n"
        "            btn.clicked.connect(lambda _checked=False, m=mode_key: self._deck._apply_runtime_persona_mode(m))\n"
        "            persona_row.addWidget(btn)\n"
        "            self._persona_mode_buttons[mode_key] = btn\n"
        "        persona_host = QWidget()\n"
        "        persona_host.setLayout(persona_row)\n"
        "        layout.addWidget(persona_host)\n"
        "\n"
        "        settings = CFG.get(\"settings\", {})\n"
        "        tz_auto = bool(settings.get(\"timezone_auto_detect\", True))\n"
        "        tz_override = str(settings.get(\"timezone_override\", \"\") or \"\").strip()\n"
        "\n"
        "        tz_auto_chk = QCheckBox(\"Auto-detect local/system time zone\")\n"
        "        tz_auto_chk.setChecked(tz_auto)\n"
        "        tz_auto_chk.toggled.connect(self._deck._set_timezone_auto_detect)\n"
        "        layout.addWidget(tz_auto_chk)\n"
        "\n"
        "        tz_row = QHBoxLayout()\n"
        "        tz_row.addWidget(QLabel(\"Manual Time Zone Override:\"))\n"
        "        tz_combo = QComboBox()\n"
        "        tz_combo.setEditable(True)\n"
        "        tz_options = [\n"
        "            \"America/Chicago\", \"America/New_York\", \"America/Los_Angeles\",\n"
        "            \"America/Denver\", \"UTC\"\n"
        "        ]\n"
        "        tz_combo.addItems(tz_options)\n"
        "        if tz_override:\n"
        "            if tz_combo.findText(tz_override) < 0:\n"
        "                tz_combo.addItem(tz_override)\n"
        "            tz_combo.setCurrentText(tz_override)\n"
        "        else:\n"
        "            tz_combo.setCurrentText(\"America/Chicago\")\n"
        "        tz_combo.setEnabled(not tz_auto)\n"
        "        tz_combo.currentTextChanged.connect(self._deck._set_timezone_override)\n"
        "        tz_auto_chk.toggled.connect(lambda enabled: tz_combo.setEnabled(not enabled))\n"
        "        tz_row.addWidget(tz_combo, 1)\n"
        "        tz_host = QWidget()\n"
        "        tz_host.setLayout(tz_row)\n"
        "        layout.addWidget(tz_host)\n",
        "settings system section persona runtime controls",
    )
    source = _replace_once(
        source,
        "    def _build_ui_section(self, layout: QVBoxLayout) -> None:\n"
        "        layout.addWidget(QLabel(\"Window Shell\"))\n"
        "        layout.addWidget(self._deck._fs_btn)\n"
        "        layout.addWidget(self._deck._bl_btn)\n",
        "    def _build_ui_section(self, layout: QVBoxLayout) -> None:\n"
        "        layout.addWidget(QLabel(\"Theme Mode\"))\n"
        "        self._theme_mode_buttons = {}\n"
        "        theme_modes = [\n"
        "            (\"light\", \"Light\"),\n"
        "            (\"auto\", \"Auto\"),\n"
        "            (\"dark\", \"Dark\"),\n"
        "        ]\n"
        "        theme_row = QHBoxLayout()\n"
        "        for mode_key, label in theme_modes:\n"
        "            btn = QPushButton(label)\n"
        "            btn.setCheckable(True)\n"
        "            btn.clicked.connect(lambda _checked=False, m=mode_key: self._deck._apply_runtime_theme_mode(m))\n"
        "            theme_row.addWidget(btn)\n"
        "            self._theme_mode_buttons[mode_key] = btn\n"
        "        theme_host = QWidget()\n"
        "        theme_host.setLayout(theme_row)\n"
        "        layout.addWidget(theme_host)\n"
        "\n"
        "        layout.addWidget(QLabel(\"Window Shell\"))\n"
        "        layout.addWidget(self._deck._fs_btn)\n"
        "        layout.addWidget(self._deck._bl_btn)\n"
        "        self.sync_runtime_controls()\n"
        "\n"
        "    def sync_runtime_controls(self) -> None:\n"
        "        persona_mode = _normalize_runtime_mode(getattr(self._deck, \"_runtime_persona_mode\", \"persona\"))\n"
        "        theme_mode = _normalize_theme_mode(getattr(self._deck, \"_runtime_theme_mode\", \"auto\"))\n"
        "        for mode_key, btn in getattr(self, \"_persona_mode_buttons\", {}).items():\n"
        "            btn.setChecked(mode_key == persona_mode)\n"
        "        for mode_key, btn in getattr(self, \"_theme_mode_buttons\", {}).items():\n"
        "            btn.setChecked(mode_key == theme_mode)\n",
        "settings ui section theme controls + sync",
    )
    source = _replace_once(
        source,
        """        # ── GPU master bar (full width) ───────────────────────────────
        layout.addWidget(section_label("❧ INFERNAL ENGINE"))
        self.gauge_gpu_master = GaugeWidget("RTX", "%", 100.0, C_CRIMSON)
        self.gauge_gpu_master.setMaximumHeight(55)
        layout.addWidget(self.gauge_gpu_master)

        layout.addStretch()
""",
        """        # ── GPU master bar (full width) ───────────────────────────────
        layout.addWidget(section_label("❧ INFERNAL ENGINE"))
        self.gauge_gpu_master = GaugeWidget("RTX", "%", 100.0, C_CRIMSON)
        self.gauge_gpu_master.setMaximumHeight(55)
        layout.addWidget(self.gauge_gpu_master)

        # ── Power telemetry ────────────────────────────────────────────
        layout.addWidget(section_label("❧ POWER"))
        power_frame = QFrame()
        power_frame.setStyleSheet(
            f"background: {C_PANEL}; border: 1px solid {C_BORDER}; border-radius: 2px;"
        )
        pf = QVBoxLayout(power_frame)
        pf.setContentsMargins(8, 4, 8, 4)
        pf.setSpacing(2)
        self.lbl_power_has_battery = QLabel("✦ HAS BATTERY: N/A")
        self.lbl_power_percent = QLabel("✦ BATTERY: N/A")
        self.lbl_power_state = QLabel("✦ POWER STATE: N/A")
        self.lbl_power_time = QLabel("✦ REMAINING: N/A")
        self.lbl_power_direction = QLabel("✦ DIRECTION: N/A")
        for lbl in (self.lbl_power_has_battery, self.lbl_power_percent, self.lbl_power_state, self.lbl_power_time, self.lbl_power_direction):
            lbl.setStyleSheet(
                f"color: {C_TEXT_DIM}; font-size: 10px; "
                f"font-family: {DECK_FONT}, serif; border: none;"
            )
            pf.addWidget(lbl)
        layout.addWidget(power_frame)

        # ── Network telemetry ──────────────────────────────────────────
        layout.addWidget(section_label("❧ NETWORK"))
        network_frame = QFrame()
        network_frame.setStyleSheet(
            f"background: {C_PANEL}; border: 1px solid {C_BORDER}; border-radius: 2px;"
        )
        nf = QVBoxLayout(network_frame)
        nf.setContentsMargins(8, 4, 8, 4)
        nf.setSpacing(2)
        self.lbl_net_connected = QLabel("✦ CONNECTED: N/A")
        self.lbl_net_wifi = QLabel("✦ WI-FI: N/A")
        self.lbl_net_iface = QLabel("✦ ADAPTER: N/A")
        self.lbl_net_ssid = QLabel("✦ SSID: N/A")
        self.lbl_net_throughput = QLabel("✦ THROUGHPUT: N/A")
        for lbl in (self.lbl_net_connected, self.lbl_net_wifi, self.lbl_net_iface, self.lbl_net_ssid, self.lbl_net_throughput):
            lbl.setStyleSheet(
                f"color: {C_TEXT_DIM}; font-size: 10px; "
                f"font-family: {DECK_FONT}, serif; border: none;"
            )
            nf.addWidget(lbl)
        layout.addWidget(network_frame)

        layout.addStretch()
""",
        "instruments panel power+network sections",
    )
    source = _replace_once(
        source,
        "    def update_stats(self) -> None:\n",
        "    def _fmt_secs(self, secs: int) -> str:\n"
        "        if secs is None or secs < 0:\n"
        "            return \"Unknown\"\n"
        "        h = secs // 3600\n"
        "        m = (secs % 3600) // 60\n"
        "        return f\"{h}h {m}m\"\n"
        "\n"
        "    def _update_power_stats(self) -> None:\n"
        "        if not PSUTIL_OK or not hasattr(psutil, \"sensors_battery\"):\n"
        "            self.lbl_power_has_battery.setText(\"✦ HAS BATTERY: Unavailable\")\n"
        "            self.lbl_power_percent.setText(\"✦ BATTERY: N/A\")\n"
        "            self.lbl_power_state.setText(\"✦ POWER STATE: Unavailable\")\n"
        "            self.lbl_power_time.setText(\"✦ REMAINING: Unavailable\")\n"
        "            self.lbl_power_direction.setText(\"✦ DIRECTION: Unavailable\")\n"
        "            return\n"
        "        try:\n"
        "            batt = psutil.sensors_battery()\n"
        "        except Exception:\n"
        "            batt = None\n"
        "        if batt is None:\n"
        "            self.lbl_power_has_battery.setText(\"✦ HAS BATTERY: No\")\n"
        "            self.lbl_power_percent.setText(\"✦ BATTERY: No battery detected\")\n"
        "            self.lbl_power_state.setText(\"✦ POWER STATE: No battery detected\")\n"
        "            self.lbl_power_time.setText(\"✦ REMAINING: No battery detected\")\n"
        "            self.lbl_power_direction.setText(\"✦ DIRECTION: No battery detected\")\n"
        "            return\n"
        "        pct = float(getattr(batt, \"percent\", 0.0) or 0.0)\n"
        "        plugged = bool(getattr(batt, \"power_plugged\", False))\n"
        "        left = int(getattr(batt, \"secsleft\", -1) or -1)\n"
        "        rem = self._fmt_secs(left) if left >= 0 else \"Unknown\"\n"
        "        state = \"Plugged in / charging\" if plugged else \"On battery\"\n"
        "        direction = \"Charging\" if plugged else \"Discharging\"\n"
        "        self.lbl_power_has_battery.setText(\"✦ HAS BATTERY: Yes\")\n"
        "        self.lbl_power_percent.setText(f\"✦ BATTERY: {pct:.0f}%\")\n"
        "        self.lbl_power_state.setText(f\"✦ POWER STATE: {state}\")\n"
        "        self.lbl_power_time.setText(f\"✦ REMAINING: {rem}\")\n"
        "        self.lbl_power_direction.setText(f\"✦ DIRECTION: {direction}\")\n"
        "\n"
        "    def _detect_wifi_ssid(self) -> str:\n"
        "        for cmd in ([\"netsh\", \"wlan\", \"show\", \"interfaces\"], [\"nmcli\", \"-t\", \"-f\", \"active,ssid\", \"dev\", \"wifi\"], [\"iwgetid\", \"-r\"]):\n"
        "            try:\n"
        "                out = subprocess.check_output(cmd, stderr=subprocess.DEVNULL, text=True, timeout=1.5, creationflags=getattr(subprocess, 'CREATE_NO_WINDOW', 0))\n"
        "            except Exception:\n"
        "                continue\n"
        "            txt = (out or \"\").strip()\n"
        "            if not txt:\n"
        "                continue\n"
        "            if cmd[0] == \"netsh\":\n"
        "                for line in txt.splitlines():\n"
        "                    low = line.lower().strip()\n"
        "                    if low.startswith(\"ssid\") and \"bssid\" not in low and \":\" in line:\n"
        "                        return line.split(\":\", 1)[1].strip() or \"Unknown\"\n"
        "            elif cmd[0] == \"nmcli\":\n"
        "                for line in txt.splitlines():\n"
        "                    if line.startswith(\"yes:\"):\n"
        "                        return line.split(\":\", 1)[1].strip() or \"Unknown\"\n"
        "            else:\n"
        "                return txt.splitlines()[0].strip() or \"Unknown\"\n"
        "        return \"No Wi-Fi\"\n"
        "\n"
        "    def _update_network_stats(self) -> None:\n"
        "        if not PSUTIL_OK:\n"
        "            self.lbl_net_connected.setText(\"✦ CONNECTED: Unavailable\")\n"
        "            self.lbl_net_wifi.setText(\"✦ WI-FI: Unavailable\")\n"
        "            self.lbl_net_iface.setText(\"✦ ADAPTER: Unavailable\")\n"
        "            self.lbl_net_ssid.setText(\"✦ SSID: Unavailable\")\n"
        "            self.lbl_net_throughput.setText(\"✦ THROUGHPUT: Unavailable\")\n"
        "            return\n"
        "        try:\n"
        "            stats = psutil.net_if_stats(); addrs = psutil.net_if_addrs(); io_now = psutil.net_io_counters()\n"
        "        except Exception:\n"
        "            self.lbl_net_connected.setText(\"✦ CONNECTED: Unavailable\")\n"
        "            self.lbl_net_wifi.setText(\"✦ WI-FI: Unavailable\")\n"
        "            self.lbl_net_iface.setText(\"✦ ADAPTER: Unavailable\")\n"
        "            self.lbl_net_ssid.setText(\"✦ SSID: Unavailable\")\n"
        "            self.lbl_net_throughput.setText(\"✦ THROUGHPUT: Unavailable\")\n"
        "            return\n"
        "        active = []\n"
        "        for name, st in stats.items():\n"
        "            if not getattr(st, \"isup\", False):\n"
        "                continue\n"
        "            if name.lower().startswith(\"loopback\") or name.lower().startswith(\"lo\"):\n"
        "                continue\n"
        "            if addrs.get(name, []):\n"
        "                active.append(name)\n"
        "        connected = bool(active)\n"
        "        wifi_iface = next((n for n in active if any(k in n.lower() for k in (\"wi-fi\", \"wifi\", \"wlan\", \"wireless\"))), \"\")\n"
        "        is_wifi = bool(wifi_iface)\n"
        "        iface_label = wifi_iface or (active[0] if active else \"None\")\n"
        "        throughput = \"N/A\"\n"
        "        if io_now is not None:\n"
        "            now = time.time()\n"
        "            last_ts = getattr(self, \"_net_prev_ts\", None)\n"
        "            if last_ts and now > last_ts:\n"
        "                up = max(0.0, (io_now.bytes_sent - getattr(self, \"_net_prev_sent\", io_now.bytes_sent)) / 1024.0 / (now - last_ts))\n"
        "                down = max(0.0, (io_now.bytes_recv - getattr(self, \"_net_prev_recv\", io_now.bytes_recv)) / 1024.0 / (now - last_ts))\n"
        "                throughput = f\"↓ {down:.1f} KB/s ↑ {up:.1f} KB/s\"\n"
        "            self._net_prev_ts = now\n"
        "            self._net_prev_sent = io_now.bytes_sent\n"
        "            self._net_prev_recv = io_now.bytes_recv\n"
        "        ssid = self._detect_wifi_ssid() if is_wifi else \"No Wi-Fi\"\n"
        "        self.lbl_net_connected.setText(f\"✦ CONNECTED: {'Yes' if connected else 'No'}\")\n"
        "        self.lbl_net_wifi.setText(f\"✦ WI-FI: {'Yes' if is_wifi else 'No'}\")\n"
        "        self.lbl_net_iface.setText(f\"✦ ADAPTER: {iface_label}\")\n"
        "        self.lbl_net_ssid.setText(f\"✦ SSID: {ssid}\")\n"
        "        self.lbl_net_throughput.setText(f\"✦ THROUGHPUT: {throughput}\")\n"
        "\n"
        "    def update_stats(self) -> None:\n",
        "instruments helper methods for power+network",
    )
    source = _replace_once(
        source,
        "        # Update drive bars every 30 seconds (not every tick)\n",
        "        self._update_power_stats()\n"
        "        self._update_network_stats()\n"
        "\n"
        "        # Update drive bars every 30 seconds (not every tick)\n",
        "instruments update cycle power+network",
    )
    source = _replace_once(
        source,
        "class DiceTrayDie(QFrame):\n",
        """class FinancialPlannerTab(QWidget):
    \"\"\"Manual-entry Financial Planner module with Dashboard + Planner internal tabs.\"\"\"

    DEFAULT_CATEGORY_BUCKETS = {
        "Housing": "needs", "Utilities": "needs", "Insurance": "needs", "Groceries": "needs",
        "Transportation": "needs", "Medical": "needs",
        "Dining Out": "wants", "Entertainment": "wants", "Subscriptions": "wants",
        "Shopping": "wants", "Hobbies": "wants",
        "Emergency Fund": "savings_debt", "Retirement": "savings_debt",
        "Investments": "savings_debt", "Debt Payment": "savings_debt", "Sinking Fund": "savings_debt",
    }

    def __init__(self, diagnostics_logger=None):
        super().__init__()
        self._log = diagnostics_logger or (lambda *_args, **_kwargs: None)
        self._path = cfg_path("memories") / "financial_planner.json"
        self.data = self._load_data()
        self._bill_last_reminder: dict[str, datetime] = {}
        self._build_ui()
        self._refresh_all()
        self._reminder_timer = QTimer(self)
        self._reminder_timer.timeout.connect(self._check_bill_reminders)
        self._reminder_timer.start(60000)

    def _default_data(self) -> dict:
        return {
            "transactions": [],
            "recurring_bills": [],
            "budget_targets": [],
            "goals": [],
            "planner_settings": {
                "currency": "USD",
                "guidance_model": "50/30/20",
                "reminder_repeat_interval": 3600,
                "google_sync_enabled": True,
                "view_month": date.today().strftime("%Y-%m"),
            },
            "category_buckets": dict(self.DEFAULT_CATEGORY_BUCKETS),
            "bill_activity": [],
        }

    def _load_data(self) -> dict:
        data = self._default_data()
        try:
            if self._path.exists():
                loaded = json.loads(self._path.read_text(encoding="utf-8"))
                if isinstance(loaded, dict):
                    for key in ("transactions", "recurring_bills", "budget_targets", "goals", "bill_activity"):
                        if isinstance(loaded.get(key), list):
                            data[key] = loaded[key]
                    if isinstance(loaded.get("planner_settings"), dict):
                        data["planner_settings"].update(loaded["planner_settings"])
                    if isinstance(loaded.get("category_buckets"), dict):
                        data["category_buckets"].update(loaded["category_buckets"])
        except Exception as e:
            self._log(f"[FINANCIAL] load failed: {e}", "WARN")
        return data

    def _save_data(self) -> None:
        try:
            self._path.parent.mkdir(parents=True, exist_ok=True)
            self._path.write_text(json.dumps(self.data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            self._log(f"[FINANCIAL] save failed: {e}", "WARN")

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        top_row = QHBoxLayout()
        self._btn_export = QPushButton("Export")
        self._btn_export.clicked.connect(self._export_data)
        self._btn_import = QPushButton("Import")
        self._btn_import.clicked.connect(self._import_data)
        self._btn_review = QPushButton("Generate Monthly AI Review")
        self._btn_review.clicked.connect(self._generate_monthly_review)
        top_row.addWidget(self._btn_export)
        top_row.addWidget(self._btn_import)
        top_row.addWidget(self._btn_review)
        top_row.addStretch(1)
        root.addLayout(top_row)

        self._tabs = QTabWidget()
        root.addWidget(self._tabs, 1)
        self._dashboard_tab = QWidget()
        self._planner_tab = QWidget()
        self._tabs.addTab(self._dashboard_tab, "Dashboard")
        self._tabs.addTab(self._planner_tab, "Planner")

        self._build_dashboard_tab()
        self._build_planner_tab()

    def _all_categories(self) -> list[str]:
        return sorted(self.data.get("category_buckets", {}).keys())

    def _build_dashboard_tab(self) -> None:
        lay = QVBoxLayout(self._dashboard_tab)
        self._monthly_snapshot = QLabel("")
        self._guidance_snapshot = QLabel("")
        self._goals_snapshot = QTableWidget(0, 5)
        self._goals_snapshot.setHorizontalHeaderLabels(["Goal", "Current", "Target", "Progress %", "Projected"])
        self._recent_snapshot = QTableWidget(0, 5)
        self._recent_snapshot.setHorizontalHeaderLabels(["Date", "Type", "Category", "Amount", "Note/Payee"])
        for tbl in (self._goals_snapshot, self._recent_snapshot):
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        lay.addWidget(QLabel("Monthly Snapshot"))
        lay.addWidget(self._monthly_snapshot)
        lay.addWidget(QLabel("50 / 30 / 20 Guidance Monitor"))
        lay.addWidget(self._guidance_snapshot)
        lay.addWidget(QLabel("Savings / Debt / Goal Progress"))
        lay.addWidget(self._goals_snapshot, 1)
        lay.addWidget(QLabel("Recent Activity Feed"))
        lay.addWidget(self._recent_snapshot, 1)

    def _build_planner_tab(self) -> None:
        lay = QVBoxLayout(self._planner_tab)
        self._planner_sections = QTabWidget()
        lay.addWidget(self._planner_sections, 1)
        self._tx_tab = QWidget()
        self._bill_tab = QWidget()
        self._budget_tab = QWidget()
        self._goals_tab = QWidget()
        self._planner_sections.addTab(self._tx_tab, "Transactions Ledger")
        self._planner_sections.addTab(self._bill_tab, "Recurring Bills Manager")
        self._planner_sections.addTab(self._budget_tab, "Budget Targets")
        self._planner_sections.addTab(self._goals_tab, "Goals")
        self._build_transactions_section()
        self._build_bills_section()
        self._build_budget_section()
        self._build_goals_section()

    def _build_transactions_section(self) -> None:
        lay = QVBoxLayout(self._tx_tab)
        form = QGridLayout()
        self.tx_date = QDateEdit(QDate.currentDate()); self.tx_date.setCalendarPopup(True)
        self.tx_type = QComboBox(); self.tx_type.addItems(["income", "expense", "transfer", "refund"])
        self.tx_category = QComboBox(); self.tx_category.addItems(self._all_categories())
        self.tx_amount = QLineEdit(); self.tx_payee = QLineEdit(); self.tx_notes = QLineEdit(); self.tx_tags = QLineEdit()
        labels = ["Date", "Type", "Category", "Amount", "Payee/Source", "Notes", "Tags"]
        widgets = [self.tx_date, self.tx_type, self.tx_category, self.tx_amount, self.tx_payee, self.tx_notes, self.tx_tags]
        for i, (lbl, w) in enumerate(zip(labels, widgets)):
            form.addWidget(QLabel(lbl), i, 0); form.addWidget(w, i, 1)
        lay.addLayout(form)
        row = QHBoxLayout()
        self.tx_add = QPushButton("Add"); self.tx_add.clicked.connect(self._add_transaction)
        self.tx_edit = QPushButton("Edit Selected"); self.tx_edit.clicked.connect(self._edit_transaction)
        self.tx_del = QPushButton("Delete Selected"); self.tx_del.clicked.connect(self._delete_transaction)
        row.addWidget(self.tx_add); row.addWidget(self.tx_edit); row.addWidget(self.tx_del); row.addStretch(1)
        lay.addLayout(row)
        self.tx_table = QTableWidget(0, 7)
        self.tx_table.setHorizontalHeaderLabels(["Date", "Type", "Category", "Amount", "Payee/Source", "Notes", "Tags"])
        self.tx_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        lay.addWidget(self.tx_table, 1)

    def _build_bills_section(self) -> None:
        lay = QVBoxLayout(self._bill_tab)
        form = QGridLayout()
        self.bill_name = QLineEdit(); self.bill_amount = QLineEdit()
        self.bill_category = QComboBox(); self.bill_category.addItems(self._all_categories())
        self.bill_due = QDateEdit(QDate.currentDate()); self.bill_due.setCalendarPopup(True)
        self.bill_recur = QComboBox(); self.bill_recur.addItems(["weekly", "monthly", "quarterly", "yearly"])
        self.bill_lead = QSpinBox(); self.bill_lead.setRange(0, 60); self.bill_lead.setValue(3)
        self.bill_autopay = QCheckBox("Autopay")
        self.bill_sync = QCheckBox("Google Sync")
        self.bill_repeat = QSpinBox(); self.bill_repeat.setRange(15, 1440); self.bill_repeat.setValue(60)
        self.bill_notes = QLineEdit()
        items = [("Bill Name", self.bill_name), ("Amount", self.bill_amount), ("Category", self.bill_category),
                 ("Due Date", self.bill_due), ("Recurrence", self.bill_recur), ("Reminder Lead (days)", self.bill_lead),
                 ("Reminder Repeat (min)", self.bill_repeat), ("Notes", self.bill_notes)]
        for i, (lbl, w) in enumerate(items):
            form.addWidget(QLabel(lbl), i, 0); form.addWidget(w, i, 1)
        form.addWidget(self.bill_autopay, len(items), 0); form.addWidget(self.bill_sync, len(items), 1)
        lay.addLayout(form)
        row = QHBoxLayout()
        for txt, fn in [("Add", self._add_bill), ("Edit Selected", self._edit_bill), ("Delete Selected", self._delete_bill),
                        ("Mark Paid", self._mark_bill_paid), ("Snooze", self._snooze_bill), ("Reschedule", self._reschedule_bill)]:
            b = QPushButton(txt); b.clicked.connect(fn); row.addWidget(b)
        row.addStretch(1); lay.addLayout(row)
        self.bill_table = QTableWidget(0, 9)
        self.bill_table.setHorizontalHeaderLabels(["Name", "Amount", "Category", "Due", "Recurrence", "Status", "Autopay", "Google Sync", "Notes"])
        self.bill_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        lay.addWidget(self.bill_table, 1)

    def _build_budget_section(self) -> None:
        lay = QVBoxLayout(self._budget_tab)
        form = QHBoxLayout()
        self.budget_category = QComboBox(); self.budget_category.addItems(self._all_categories())
        self.budget_amount = QLineEdit()
        self.budget_period = QComboBox(); self.budget_period.addItems(["monthly", "quarterly", "yearly"])
        add_btn = QPushButton("Add/Update"); add_btn.clicked.connect(self._upsert_budget_target)
        form.addWidget(QLabel("Category")); form.addWidget(self.budget_category)
        form.addWidget(QLabel("Planned Amount")); form.addWidget(self.budget_amount)
        form.addWidget(QLabel("Period")); form.addWidget(self.budget_period)
        form.addWidget(add_btn); form.addStretch(1)
        lay.addLayout(form)
        self.budget_table = QTableWidget(0, 5)
        self.budget_table.setHorizontalHeaderLabels(["Category", "Planned", "Actual", "Variance", "% Used"])
        self.budget_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        lay.addWidget(self.budget_table, 1)

    def _build_goals_section(self) -> None:
        lay = QVBoxLayout(self._goals_tab)
        form = QGridLayout()
        self.goal_name = QLineEdit(); self.goal_type = QComboBox(); self.goal_type.addItems(["savings goal", "emergency fund", "sinking fund", "debt payoff"])
        self.goal_target = QLineEdit(); self.goal_current = QLineEdit(); self.goal_monthly = QLineEdit()
        self.goal_deadline = QDateEdit(QDate.currentDate()); self.goal_deadline.setCalendarPopup(True); self.goal_deadline.setSpecialValueText("None")
        self.goal_notes = QLineEdit()
        gitems = [("Goal Name", self.goal_name), ("Goal Type", self.goal_type), ("Target Amount", self.goal_target),
                  ("Current Amount", self.goal_current), ("Monthly Contribution", self.goal_monthly), ("Deadline", self.goal_deadline), ("Notes", self.goal_notes)]
        for i, (lbl, w) in enumerate(gitems):
            form.addWidget(QLabel(lbl), i, 0); form.addWidget(w, i, 1)
        lay.addLayout(form)
        row = QHBoxLayout()
        for txt, fn in [("Add", self._add_goal), ("Edit Selected", self._edit_goal), ("Delete Selected", self._delete_goal)]:
            b = QPushButton(txt); b.clicked.connect(fn); row.addWidget(b)
        row.addStretch(1); lay.addLayout(row)
        self.goal_table = QTableWidget(0, 7)
        self.goal_table.setHorizontalHeaderLabels(["Name", "Type", "Current", "Target", "Monthly", "Progress %", "Deadline"])
        self.goal_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        lay.addWidget(self.goal_table, 1)

    def _safe_amount(self, text: str) -> float:
        try: return float(str(text).replace(",", "").strip())
        except Exception: return 0.0

    def _month_key(self) -> str:
        return self.data.get("planner_settings", {}).get("view_month", date.today().strftime("%Y-%m"))

    def _is_current_month(self, value: str) -> bool:
        return str(value or "").startswith(self._month_key())

    def _add_transaction(self) -> None:
        row = {"id": str(uuid.uuid4()), "date": self.tx_date.date().toString("yyyy-MM-dd"), "type": self.tx_type.currentText(),
               "category": self.tx_category.currentText(), "amount": self._safe_amount(self.tx_amount.text()),
               "payee_or_source": self.tx_payee.text().strip(), "notes": self.tx_notes.text().strip(), "tags": self.tx_tags.text().strip()}
        self.data["transactions"].append(row); self.data["category_buckets"][row["category"]] = self.data["category_buckets"].get(row["category"], "needs")
        self._save_data(); self._refresh_all()

    def _edit_transaction(self) -> None:
        r = self.tx_table.currentRow()
        if r < 0 or r >= len(self.data["transactions"]): return
        self.data["transactions"][r].update({"date": self.tx_date.date().toString("yyyy-MM-dd"), "type": self.tx_type.currentText(),
                                             "category": self.tx_category.currentText(), "amount": self._safe_amount(self.tx_amount.text()),
                                             "payee_or_source": self.tx_payee.text().strip(), "notes": self.tx_notes.text().strip(), "tags": self.tx_tags.text().strip()})
        self._save_data(); self._refresh_all()

    def _delete_transaction(self) -> None:
        r = self.tx_table.currentRow()
        if r < 0 or r >= len(self.data["transactions"]): return
        self.data["transactions"].pop(r); self._save_data(); self._refresh_all()

    def _add_bill(self) -> None:
        bill = {"id": str(uuid.uuid4()), "name": self.bill_name.text().strip(), "amount": self._safe_amount(self.bill_amount.text()),
                "category": self.bill_category.currentText(), "due_date_or_day": self.bill_due.date().toString("yyyy-MM-dd"),
                "recurrence": self.bill_recur.currentText(), "reminder_offset": int(self.bill_lead.value()),
                "reminder_repeat_interval": int(self.bill_repeat.value()) * 60, "autopay_flag": bool(self.bill_autopay.isChecked()),
                "google_sync_flag": bool(self.bill_sync.isChecked()), "status": "unpaid", "notes": self.bill_notes.text().strip(),
                "snooze_until": "", "next_due": self.bill_due.date().toString("yyyy-MM-dd")}
        if bill["google_sync_flag"]:
            bill["google_sync_descriptor"] = self._build_google_sync_descriptor(bill)
        self.data["recurring_bills"].append(bill); self._save_data(); self._refresh_all()

    def _edit_bill(self) -> None:
        r = self.bill_table.currentRow()
        if r < 0 or r >= len(self.data["recurring_bills"]): return
        old = self.data["recurring_bills"][r]
        old.update({"name": self.bill_name.text().strip(), "amount": self._safe_amount(self.bill_amount.text()), "category": self.bill_category.currentText(),
                    "due_date_or_day": self.bill_due.date().toString("yyyy-MM-dd"), "recurrence": self.bill_recur.currentText(),
                    "reminder_offset": int(self.bill_lead.value()), "reminder_repeat_interval": int(self.bill_repeat.value()) * 60,
                    "autopay_flag": bool(self.bill_autopay.isChecked()), "google_sync_flag": bool(self.bill_sync.isChecked()),
                    "notes": self.bill_notes.text().strip()})
        if old.get("google_sync_flag"): old["google_sync_descriptor"] = self._build_google_sync_descriptor(old)
        self._save_data(); self._refresh_all()

    def _delete_bill(self) -> None:
        r = self.bill_table.currentRow()
        if r < 0 or r >= len(self.data["recurring_bills"]): return
        self.data["recurring_bills"].pop(r); self._save_data(); self._refresh_all()

    def _selected_bill(self) -> Optional[dict]:
        r = self.bill_table.currentRow()
        if r < 0 or r >= len(self.data["recurring_bills"]): return None
        return self.data["recurring_bills"][r]

    def _mark_bill_paid(self) -> None:
        bill = self._selected_bill()
        if not bill: return
        bill["status"] = "paid"
        bill["paid_at"] = local_now_iso()
        bill["next_due"] = self._advance_due_date(bill.get("next_due") or bill.get("due_date_or_day"), bill.get("recurrence", "monthly"))
        self.data["bill_activity"].append({"date": date.today().isoformat(), "type": "bill_paid", "category": bill.get("category",""),
                                           "amount": bill.get("amount",0.0), "note": bill.get("name","")})
        self._save_data(); self._refresh_all()

    def _snooze_bill(self) -> None:
        bill = self._selected_bill()
        if not bill: return
        mins, ok = QInputDialog.getInt(self, "Snooze Bill Reminder", "Snooze for minutes:", 120, 15, 10080)
        if not ok: return
        bill["snooze_until"] = (datetime.now() + timedelta(minutes=mins)).strftime("%Y-%m-%d %H:%M:%S")
        bill["status"] = "snoozed"
        self._save_data(); self._refresh_all()

    def _reschedule_bill(self) -> None:
        bill = self._selected_bill()
        if not bill: return
        new_due, ok = QInputDialog.getText(self, "Reschedule Bill", "New due date (YYYY-MM-DD):", text=str(bill.get("next_due") or bill.get("due_date_or_day") or ""))
        if not ok or not new_due.strip(): return
        bill["next_due"] = new_due.strip()
        bill["status"] = "unpaid"
        bill["snooze_until"] = ""
        self._save_data(); self._refresh_all()

    def _upsert_budget_target(self) -> None:
        rec = {"category": self.budget_category.currentText(), "planned_amount": self._safe_amount(self.budget_amount.text()), "period": self.budget_period.currentText()}
        found = False
        for idx, row in enumerate(self.data["budget_targets"]):
            if row.get("category") == rec["category"] and row.get("period", "monthly") == rec["period"]:
                self.data["budget_targets"][idx] = rec; found = True; break
        if not found: self.data["budget_targets"].append(rec)
        self._save_data(); self._refresh_all()

    def _add_goal(self) -> None:
        rec = {"id": str(uuid.uuid4()), "name": self.goal_name.text().strip(), "type": self.goal_type.currentText(),
               "target_amount": self._safe_amount(self.goal_target.text()), "current_amount": self._safe_amount(self.goal_current.text()),
               "monthly_contribution": self._safe_amount(self.goal_monthly.text()), "deadline": self.goal_deadline.date().toString("yyyy-MM-dd"),
               "notes": self.goal_notes.text().strip()}
        self.data["goals"].append(rec); self._save_data(); self._refresh_all()

    def _edit_goal(self) -> None:
        r = self.goal_table.currentRow()
        if r < 0 or r >= len(self.data["goals"]): return
        self.data["goals"][r].update({"name": self.goal_name.text().strip(), "type": self.goal_type.currentText(),
                                      "target_amount": self._safe_amount(self.goal_target.text()), "current_amount": self._safe_amount(self.goal_current.text()),
                                      "monthly_contribution": self._safe_amount(self.goal_monthly.text()), "deadline": self.goal_deadline.date().toString("yyyy-MM-dd"),
                                      "notes": self.goal_notes.text().strip()})
        self._save_data(); self._refresh_all()

    def _delete_goal(self) -> None:
        r = self.goal_table.currentRow()
        if r < 0 or r >= len(self.data["goals"]): return
        self.data["goals"].pop(r); self._save_data(); self._refresh_all()

    def _advance_due_date(self, due_text: str, recurrence: str) -> str:
        try: d = datetime.strptime(str(due_text), "%Y-%m-%d").date()
        except Exception: d = date.today()
        if recurrence == "weekly": d = d + timedelta(days=7)
        elif recurrence == "monthly": d = d + timedelta(days=30)
        elif recurrence == "quarterly": d = d + timedelta(days=91)
        else: d = d + timedelta(days=365)
        return d.strftime("%Y-%m-%d")

    def _build_google_sync_descriptor(self, bill: dict) -> dict:
        return {
            "title": f"Bill Due — {bill.get('name','Unnamed Bill')}",
            "description": f"Amount: {bill.get('amount',0.0)} | Category: {bill.get('category','')} | Notes: {bill.get('notes','')} | Recurrence: {bill.get('recurrence','monthly')}",
            "recurring": True,
            "source_of_truth": "financial_planner_local",
            "last_synced": local_now_iso(),
        }

    def _emit_hidden_ai_reminder(self, bill: dict, due_status: str) -> None:
        payload = {"intent": "financial_bill_reminder", "bill_name": bill.get("name",""), "amount": bill.get("amount", 0.0),
                   "due_status": due_status, "category": bill.get("category",""), "note": bill.get("notes","")}
        try:
            messages_path = cfg_path("memories") / "messages.jsonl"
            records = read_jsonl(messages_path)
            records.append({"timestamp": local_now_iso(), "role": "user", "speaker": "internal", "content": "[INTERNAL_EVENT] " + json.dumps(payload, ensure_ascii=False)})
            write_jsonl(messages_path, records)
        except Exception as e:
            self._log(f"[FINANCIAL] hidden AI reminder write failed: {e}", "WARN")

    def _check_bill_reminders(self) -> None:
        now_dt = datetime.now()
        for bill in self.data.get("recurring_bills", []):
            if bill.get("status") == "paid":
                continue
            due_text = str(bill.get("next_due") or bill.get("due_date_or_day") or "")
            try: due_dt = datetime.strptime(due_text, "%Y-%m-%d")
            except Exception: continue
            snooze_until = str(bill.get("snooze_until") or "").strip()
            if snooze_until:
                try:
                    if datetime.strptime(snooze_until, "%Y-%m-%d %H:%M:%S") > now_dt:
                        continue
                except Exception:
                    pass
            lead_days = int(bill.get("reminder_offset", 3) or 3)
            remind_from = due_dt - timedelta(days=lead_days)
            if now_dt < remind_from:
                continue
            interval = int(bill.get("reminder_repeat_interval") or self.data.get("planner_settings", {}).get("reminder_repeat_interval", 3600))
            last_fire = self._bill_last_reminder.get(bill.get("id",""))
            if last_fire and (now_dt - last_fire).total_seconds() < max(60, interval):
                continue
            due_status = "overdue" if now_dt.date() > due_dt.date() else "due_soon" if now_dt.date() < due_dt.date() else "due_today"
            self._bill_last_reminder[bill.get("id","")] = now_dt
            bill["status"] = "overdue" if due_status == "overdue" else bill.get("status", "unpaid")
            self._emit_hidden_ai_reminder(bill, due_status)
            self._log(f"[FINANCIAL] Reminder: {bill.get('name','Bill')} ({due_status})", "INFO")
        self._save_data()
        self._refresh_dashboard()

    def _export_data(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, "Export Financial Planner", str(cfg_path("exports") / f"financial_planner_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"),
                                              "Text (*.txt);;CSV (*.csv);;Excel (*.xlsx);;PDF (*.pdf)")
        if not path: return
        ext = Path(path).suffix.lower()
        payload = self._export_payload()
        if ext == ".csv": self._write_csv_export(path, payload)
        elif ext == ".xlsx": self._write_xlsx_export(path, payload)
        elif ext == ".pdf": self._write_pdf_export(path, payload)
        else: Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _export_payload(self) -> dict:
        return {
            "generated_at": local_now_iso(),
            "recurring_bills_list": self.data.get("recurring_bills", []),
            "monthly_bills_snapshot": self._monthly_bill_snapshot(),
            "paid_unpaid_snapshot": self._paid_unpaid_snapshot(),
            "transaction_ledger": self.data.get("transactions", []),
            "budget_target_vs_actual_summary": self._budget_rows(),
            "goals_progress_summary": self.data.get("goals", []),
            "monthly_financial_review_data": self._review_context(),
        }

    def _write_csv_export(self, path: str, payload: dict) -> None:
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["section", "record_json"])
            for key, val in payload.items():
                if isinstance(val, list):
                    for item in val: w.writerow([key, json.dumps(item, ensure_ascii=False)])
                else:
                    w.writerow([key, json.dumps(val, ensure_ascii=False)])

    def _write_xlsx_export(self, path: str, payload: dict) -> None:
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "Financial Planner"
            ws.append(["section", "record_json"])
            for key, val in payload.items():
                if isinstance(val, list):
                    for item in val: ws.append([key, json.dumps(item, ensure_ascii=False)])
                else:
                    ws.append([key, json.dumps(val, ensure_ascii=False)])
            wb.save(path)
        except Exception:
            Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def _write_pdf_export(self, path: str, payload: dict) -> None:
        lines = json.dumps(payload, ensure_ascii=False, indent=2).splitlines()[:160]
        stream = "BT /F1 9 Tf 40 800 Td " + " ".join(f"({ln.replace('(', '[').replace(')', ']')}) Tj T*" for ln in lines) + " ET"
        pdf = f"%PDF-1.4\\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\\n2 0 obj<</Type/Pages/Count 1/Kids[3 0 R]>>endobj\\n3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 842]/Resources<</Font<</F1 4 0 R>>>>/Contents 5 0 R>>endobj\\n4 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\\n5 0 obj<</Length {len(stream)}>>stream\\n{stream}\\nendstream endobj\\nxref\\n0 6\\n0000000000 65535 f \\n0000000010 00000 n \\n0000000060 00000 n \\n0000000120 00000 n \\n0000000240 00000 n \\n0000000310 00000 n \\ntrailer<</Root 1 0 R/Size 6>>\\nstartxref\\n420\\n%%EOF"
        Path(path).write_text(pdf, encoding="latin-1", errors="ignore")

    def _import_data(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "Import Financial Planner", str(cfg_path("exports")), "Structured Files (*.txt *.csv *.xlsx)")
        if not path: return
        mode, ok = QInputDialog.getItem(self, "Import Mode", "Choose import mode:", ["append", "replace matching", "full replace"], 0, False)
        if not ok: return
        records = self._read_import_records(path)
        if not records:
            QMessageBox.warning(self, "Financial Planner", "No valid import records found.")
            return
        self._apply_import_records(records, mode)
        self._save_data(); self._refresh_all()

    def _read_import_records(self, path: str) -> list[tuple[str, dict]]:
        out: list[tuple[str, dict]] = []
        p = Path(path); ext = p.suffix.lower()
        try:
            if ext == ".csv":
                import csv
                with p.open("r", encoding="utf-8", newline="") as f:
                    for row in csv.DictReader(f):
                        section = row.get("section", "")
                        raw = row.get("record_json", "")
                        try: obj = json.loads(raw)
                        except Exception: continue
                        if isinstance(obj, dict): out.append((section, obj))
            elif ext == ".xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(p); ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 2: continue
                    section = str(row[0] or ""); raw = str(row[1] or "")
                    try: obj = json.loads(raw)
                    except Exception: continue
                    if isinstance(obj, dict): out.append((section, obj))
            else:
                blob = json.loads(p.read_text(encoding="utf-8"))
                for section, val in blob.items():
                    if isinstance(val, list):
                        for item in val:
                            if isinstance(item, dict): out.append((section, item))
        except Exception as e:
            self._log(f"[FINANCIAL] import parse issue: {e}", "WARN")
        return out

    def _apply_import_records(self, rows: list[tuple[str, dict]], mode: str) -> None:
        section_map = {"transaction_ledger": "transactions", "recurring_bills_list": "recurring_bills",
                       "budget_target_vs_actual_summary": "budget_targets", "goals_progress_summary": "goals"}
        if mode == "full replace":
            for key in ("transactions", "recurring_bills", "budget_targets", "goals"):
                self.data[key] = []
        for sec, rec in rows:
            key = section_map.get(sec)
            if not key: continue
            if mode == "replace matching":
                rid = rec.get("id")
                if rid:
                    replaced = False
                    for i, old in enumerate(self.data[key]):
                        if old.get("id") == rid:
                            self.data[key][i] = rec; replaced = True; break
                    if not replaced: self.data[key].append(rec)
                else:
                    self.data[key].append(rec)
            else:
                self.data[key].append(rec)

    def _monthly_bill_snapshot(self) -> dict:
        month = self._month_key()
        bills = [b for b in self.data.get("recurring_bills", []) if str(b.get("next_due") or b.get("due_date_or_day") or "").startswith(month)]
        return {"month": month, "count": len(bills), "total_amount": sum(float(b.get("amount", 0.0) or 0.0) for b in bills)}

    def _paid_unpaid_snapshot(self) -> dict:
        bills = self.data.get("recurring_bills", [])
        paid = len([b for b in bills if b.get("status") == "paid"])
        return {"paid": paid, "unpaid": len(bills) - paid}

    def _review_context(self) -> dict:
        income = sum(float(t.get("amount", 0.0) or 0.0) for t in self.data.get("transactions", []) if t.get("type") == "income" and self._is_current_month(t.get("date","")))
        expenses = sum(float(t.get("amount", 0.0) or 0.0) for t in self.data.get("transactions", []) if t.get("type") == "expense" and self._is_current_month(t.get("date","")))
        cat_totals = {}
        for t in self.data.get("transactions", []):
            if t.get("type") != "expense" or not self._is_current_month(t.get("date","")): continue
            cat = t.get("category", "Uncategorized")
            cat_totals[cat] = cat_totals.get(cat, 0.0) + float(t.get("amount", 0.0) or 0.0)
        due = [b.get("name","") for b in self.data.get("recurring_bills", []) if b.get("status") in ("unpaid", "overdue")]
        goal_progress = [{ "name": g.get("name",""), "progress_pct": round((float(g.get("current_amount",0.0) or 0.0) / max(1.0, float(g.get("target_amount",0.0) or 0.0))) * 100.0, 1)} for g in self.data.get("goals",[])]
        return {"total_income": round(income, 2), "total_expenses": round(expenses, 2), "net_cash_flow": round(income - expenses, 2),
                "largest_categories": sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)[:5],
                "due_or_overdue_bills": due, "goal_progress": goal_progress}

    def _generate_monthly_review(self) -> None:
        ctx = self._review_context()
        payload = {"intent": "financial_monthly_review", "context": ctx}
        summary = (
            f"Income: {ctx['total_income']:.2f}, Expenses: {ctx['total_expenses']:.2f}, Net: {ctx['net_cash_flow']:.2f}. "
            f"Largest categories: {', '.join([f'{c} ({a:.2f})' for c, a in ctx['largest_categories'][:3]]) or 'None'}. "
            f"Due/overdue bills: {len(ctx['due_or_overdue_bills'])}. Keep contributions steady and review high-variance categories."
        )
        self._emit_hidden_ai_reminder({"name": "Monthly Review", "amount": ctx["net_cash_flow"], "category": "review", "notes": json.dumps(payload)}, "monthly_review")
        QMessageBox.information(self, "Monthly AI Review", summary)

    def _budget_rows(self) -> list[dict]:
        rows = []
        for b in self.data.get("budget_targets", []):
            cat = b.get("category", "")
            planned = float(b.get("planned_amount", 0.0) or 0.0)
            actual = sum(float(t.get("amount", 0.0) or 0.0) for t in self.data.get("transactions", []) if t.get("type") == "expense" and t.get("category") == cat and self._is_current_month(t.get("date","")))
            variance = planned - actual
            pct = (actual / planned * 100.0) if planned > 0 else 0.0
            rows.append({"category": cat, "planned_amount": round(planned,2), "actual": round(actual,2), "variance": round(variance,2), "pct_used": round(pct,1), "period": b.get("period","monthly")})
        return rows

    def _refresh_all(self) -> None:
        self._refresh_transactions()
        self._refresh_bills()
        self._refresh_budget()
        self._refresh_goals()
        self._refresh_dashboard()

    def _refresh_transactions(self) -> None:
        rows = self.data.get("transactions", [])
        self.tx_table.setRowCount(len(rows))
        for r, item in enumerate(rows):
            vals = [item.get("date",""), item.get("type",""), item.get("category",""), f"{float(item.get('amount',0.0) or 0.0):.2f}",
                    item.get("payee_or_source",""), item.get("notes",""), item.get("tags","")]
            for c, v in enumerate(vals): self.tx_table.setItem(r, c, QTableWidgetItem(str(v)))

    def _refresh_bills(self) -> None:
        rows = self.data.get("recurring_bills", [])
        self.bill_table.setRowCount(len(rows))
        for r, item in enumerate(rows):
            vals = [item.get("name",""), f"{float(item.get('amount',0.0) or 0.0):.2f}", item.get("category",""),
                    item.get("next_due") or item.get("due_date_or_day",""), item.get("recurrence",""), item.get("status","unpaid"),
                    "Yes" if item.get("autopay_flag") else "No", "Yes" if item.get("google_sync_flag") else "No", item.get("notes","")]
            for c, v in enumerate(vals): self.bill_table.setItem(r, c, QTableWidgetItem(str(v)))

    def _refresh_budget(self) -> None:
        rows = self._budget_rows()
        self.budget_table.setRowCount(len(rows))
        for r, item in enumerate(rows):
            vals = [item.get("category",""), f"{item.get('planned_amount',0.0):.2f}", f"{item.get('actual',0.0):.2f}",
                    f"{item.get('variance',0.0):.2f}", f"{item.get('pct_used',0.0):.1f}%"]
            for c, v in enumerate(vals): self.budget_table.setItem(r, c, QTableWidgetItem(str(v)))

    def _refresh_goals(self) -> None:
        rows = self.data.get("goals", [])
        self.goal_table.setRowCount(len(rows))
        for r, g in enumerate(rows):
            target = float(g.get("target_amount", 0.0) or 0.0)
            current = float(g.get("current_amount", 0.0) or 0.0)
            pct = (current / target * 100.0) if target > 0 else 0.0
            vals = [g.get("name",""), g.get("type",""), f"{current:.2f}", f"{target:.2f}", f"{float(g.get('monthly_contribution',0.0) or 0.0):.2f}", f"{pct:.1f}", g.get("deadline","")]
            for c, v in enumerate(vals): self.goal_table.setItem(r, c, QTableWidgetItem(str(v)))

    def _refresh_dashboard(self) -> None:
        month = self._month_key()
        tx = [t for t in self.data.get("transactions", []) if self._is_current_month(t.get("date",""))]
        income = sum(float(t.get("amount",0.0) or 0.0) for t in tx if t.get("type") == "income")
        expenses = sum(float(t.get("amount",0.0) or 0.0) for t in tx if t.get("type") == "expense")
        now = date.today()
        due_soon = 0; overdue = 0; recurring = 0
        for b in self.data.get("recurring_bills", []):
            due = str(b.get("next_due") or b.get("due_date_or_day") or "")
            try: due_d = datetime.strptime(due, "%Y-%m-%d").date()
            except Exception: continue
            recurring += 1
            if b.get("status") == "paid": continue
            if due_d < now: overdue += 1
            elif due_d <= now + timedelta(days=7): due_soon += 1
        self._monthly_snapshot.setText(f"Month: {month} | Total Income: {income:.2f} | Total Expenses: {expenses:.2f} | Net Cash Flow: {income-expenses:.2f} | Bills Due Soon: {due_soon} | Overdue Bills: {overdue} | Upcoming Subscriptions/Recurring Bills: {recurring}")

        buckets = {"needs": 0.0, "wants": 0.0, "savings_debt": 0.0}
        for t in tx:
            if t.get("type") != "expense": continue
            cat = t.get("category", "")
            bucket = self.data.get("category_buckets", {}).get(cat, "needs")
            if bucket not in buckets: bucket = "needs"
            buckets[bucket] += float(t.get("amount", 0.0) or 0.0)
        total_spend = max(0.01, sum(buckets.values()))
        pct = {k: round(v / total_spend * 100.0, 1) for k, v in buckets.items()}
        self._guidance_snapshot.setText(
            f"Needs: {buckets['needs']:.2f} ({pct['needs']}%, target 50%, {'over' if pct['needs'] > 50 else 'under'}) | "
            f"Wants: {buckets['wants']:.2f} ({pct['wants']}%, target 30%, {'over' if pct['wants'] > 30 else 'under'}) | "
            f"Savings/Debt: {buckets['savings_debt']:.2f} ({pct['savings_debt']}%, target 20%, {'over' if pct['savings_debt'] > 20 else 'under'})"
        )

        goals = self.data.get("goals", [])
        self._goals_snapshot.setRowCount(len(goals))
        for r, g in enumerate(goals):
            cur = float(g.get("current_amount", 0.0) or 0.0); tgt = float(g.get("target_amount", 0.0) or 0.0); mon = float(g.get("monthly_contribution", 0.0) or 0.0)
            pctg = (cur / tgt * 100.0) if tgt > 0 else 0.0
            rem = max(0.0, tgt - cur); proj = f\"{int(math.ceil(rem / mon))} mo\" if mon > 0 and rem > 0 else \"n/a\"
            vals = [g.get("name",""), f"{cur:.2f}", f"{tgt:.2f}", f"{pctg:.1f}", proj]
            for c, v in enumerate(vals): self._goals_snapshot.setItem(r, c, QTableWidgetItem(str(v)))

        activity = sorted(
            [{"date": x.get("date",""), "type": x.get("type",""), "category": x.get("category",""), "amount": x.get("amount",0.0), "note": x.get("payee_or_source","") or x.get("notes","")} for x in tx] + self.data.get("bill_activity", []),
            key=lambda z: z.get("date",""),
            reverse=True
        )[:20]
        self._recent_snapshot.setRowCount(len(activity))
        for r, a in enumerate(activity):
            vals = [a.get("date",""), a.get("type",""), a.get("category",""), f"{float(a.get('amount',0.0) or 0.0):.2f}", a.get("note","")]
            for c, v in enumerate(vals): self._recent_snapshot.setItem(r, c, QTableWidgetItem(str(v)))


class DiceTrayDie(QFrame):
""",
        "financial planner class injection",
    )
    source = _replace_once(
        source,
        "        # ── Module Tracker tab ─────────────────────────────────────────\n"
        "        self._module_tracker = ModuleTrackerTab()\n"
        "\n"
        "        # ── Dice Roller tab ────────────────────────────────────────────\n",
        "        # ── Module Tracker tab ─────────────────────────────────────────\n"
        "        self._module_tracker = ModuleTrackerTab()\n"
        "\n"
        "        # ── Financial Planner tab ──────────────────────────────────────\n"
        "        self._financial_planner_tab = FinancialPlannerTab(diagnostics_logger=self._diag_tab.log)\n"
        "\n"
        "        # ── Dice Roller tab ────────────────────────────────────────────\n",
        "financial planner tab instantiation",
    )
    source = _replace_once(
        source,
        '            {"id": "modules", "title": "Modules", "widget": self._module_tracker, "default_order": 7},\n'
        '            {"id": "dice_roller", "title": "Dice Roller", "widget": self._dice_roller_tab, "default_order": 8},\n'
        '            {"id": "magic_8_ball", "title": "Magic 8-Ball", "widget": self._magic_8ball_tab, "default_order": 9},\n'
        '            {"id": "diagnostics", "title": "Diagnostics", "widget": self._diag_tab, "default_order": 10},\n'
        '            {"id": "settings", "title": "Settings", "widget": self._settings_tab, "default_order": 11},\n',
        '            {"id": "modules", "title": "Modules", "widget": self._module_tracker, "default_order": 7},\n'
        '            {"id": "financial_planner", "title": "Financial Planner", "widget": self._financial_planner_tab, "default_order": 8},\n'
        '            {"id": "dice_roller", "title": "Dice Roller", "widget": self._dice_roller_tab, "default_order": 9},\n'
        '            {"id": "magic_8_ball", "title": "Magic 8-Ball", "widget": self._magic_8ball_tab, "default_order": 10},\n'
        '            {"id": "diagnostics", "title": "Diagnostics", "widget": self._diag_tab, "default_order": 11},\n'
        '            {"id": "settings", "title": "Settings", "widget": self._settings_tab, "default_order": 12},\n',
        "financial planner tab definition order",
    )
    source = _replace_once(
        source,
        "        self._load_spell_tab_state_from_config()\n"
        "        self._rebuild_spell_tabs()\n",
        "        self._init_spell_category_framework()\n"
        "        self._load_spell_tab_state_from_config()\n"
        "        self._rebuild_spell_tabs()\n",
        "initialize category framework",
    )
    source = _replace_once(
        source,
        "        right_workspace_layout.addWidget(self._spell_tabs, 1)\n",
        "        self._category_strip = QWidget()\n"
        "        self._category_strip_layout = QHBoxLayout(self._category_strip)\n"
        "        self._category_strip_layout.setContentsMargins(0, 0, 0, 0)\n"
        "        self._category_strip_layout.setSpacing(4)\n"
        "        right_workspace_layout.addWidget(self._category_strip, 0)\n"
        "        right_workspace_layout.addWidget(self._spell_tabs, 1)\n",
        "inject category strip host",
    )
    source = _replace_once(
        source,
        "    def _load_spell_tab_state_from_config(self) -> None:\n",
        "    def _init_spell_category_framework(self) -> None:\n"
        "        self._protected_categories = {\"System\", \"Core\"}\n"
        "        self._category_map = {}\n"
        "        self._module_registry = {}\n"
        "        for tab in self._spell_tab_defs:\n"
        "            tab.setdefault(\"category\", \"Core\")\n"
        "            tab.setdefault(\"secondary_categories\", [])\n"
        "            tab.setdefault(\"protected_category\", tab.get(\"category\") in self._protected_categories)\n"
        "            self._register_module_categories(tab)\n"
        "        self._active_category = \"Core\" if \"Core\" in self._category_map else next(iter(self._category_map.keys()), \"Core\")\n"
        "        self._rebuild_category_strip()\n"
        "\n"
        "    def _register_module_categories(self, tab: dict) -> None:\n"
        "        tab_id = str(tab.get(\"id\") or \"\").strip()\n"
        "        if not tab_id:\n"
        "            return\n"
        "        primary = str(tab.get(\"category\") or \"Core\").strip() or \"Core\"\n"
        "        secondary = [str(c).strip() for c in tab.get(\"secondary_categories\", []) if str(c).strip()]\n"
        "        self._module_registry[tab_id] = {\"id\": tab_id, \"installed\": True, \"enabled\": True, \"primary\": primary, \"secondary\": secondary}\n"
        "        self._category_map.setdefault(primary, {\"modules\": set(), \"protected\": bool(tab.get(\"protected_category\", False) or primary in self._protected_categories)})\n"
        "        self._category_map[primary][\"modules\"].add(tab_id)\n"
        "        for cat in secondary:\n"
        "            if cat in self._category_map:\n"
        "                self._category_map[cat][\"modules\"].add(tab_id)\n"
        "\n"
        "    def _rebuild_category_strip(self) -> None:\n"
        "        if not hasattr(self, \"_category_strip_layout\"):\n"
        "            return\n"
        "        while self._category_strip_layout.count():\n"
        "            item = self._category_strip_layout.takeAt(0)\n"
        "            w = item.widget()\n"
        "            if w is not None:\n"
        "                w.deleteLater()\n"
        "        cats = list(self._category_map.keys())\n"
        "        if self._active_category not in cats and cats:\n"
        "            self._active_category = cats[0]\n"
        "        for cat in cats:\n"
        "            btn = QToolButton()\n"
        "            btn.setText(cat)\n"
        "            btn.setCheckable(True)\n"
        "            btn.setChecked(cat == self._active_category)\n"
        "            btn.clicked.connect(lambda _checked=False, c=cat: self._select_category(c))\n"
        "            self._category_strip_layout.addWidget(btn)\n"
        "        self._category_strip_layout.addStretch(1)\n"
        "\n"
        "    def _select_category(self, category: str) -> None:\n"
        "        if category not in self._category_map:\n"
        "            return\n"
        "        self._active_category = category\n"
        "        self._rebuild_category_strip()\n"
        "        self._rebuild_spell_tabs()\n"
        "\n"
        "    def _visible_tab_for_active_category(self, tab: dict) -> bool:\n"
        "        tab_id = str(tab.get(\"id\") or \"\")\n"
        "        reg = self._module_registry.get(tab_id, {})\n"
        "        if not reg.get(\"installed\", True) or not reg.get(\"enabled\", True):\n"
        "            return False\n"
        "        return (not self._active_category) or tab_id in self._category_map.get(self._active_category, {}).get(\"modules\", set())\n"
        "\n"
        "    def _disable_module(self, tab_id: str) -> None:\n"
        "        if tab_id in self._module_registry:\n"
        "            self._module_registry[tab_id][\"enabled\"] = False\n"
        "            self._rebuild_spell_tabs()\n"
        "\n"
        "    def _uninstall_module(self, tab_id: str, delete_files: bool = False) -> None:\n"
        "        reg = self._module_registry.get(tab_id)\n"
        "        if not reg:\n"
        "            return\n"
        "        reg[\"installed\"] = False\n"
        "        reg[\"enabled\"] = False\n"
        "        for cat_data in self._category_map.values():\n"
        "            cat_data.get(\"modules\", set()).discard(tab_id)\n"
        "        self._cleanup_empty_categories()\n"
        "        self._rebuild_category_strip()\n"
        "        self._rebuild_spell_tabs()\n"
        "\n"
        "    def _reinstall_modules_from_library(self) -> None:\n"
        "        for tab in self._spell_tab_defs:\n"
        "            tab_id = str(tab.get(\"id\") or \"\")\n"
        "            if not tab_id:\n"
        "                continue\n"
        "            reg = self._module_registry.setdefault(tab_id, {\"id\": tab_id, \"installed\": True, \"enabled\": True, \"primary\": str(tab.get(\"category\") or \"Core\"), \"secondary\": [str(c).strip() for c in tab.get(\"secondary_categories\", []) if str(c).strip()]})\n"
        "            reg[\"installed\"] = True\n"
        "            reg[\"enabled\"] = True\n"
        "            primary = reg.get(\"primary\") or \"Core\"\n"
        "            self._category_map.setdefault(primary, {\"modules\": set(), \"protected\": bool(tab.get(\"protected_category\", False) or primary in self._protected_categories)})\n"
        "            self._category_map[primary][\"modules\"].add(tab_id)\n"
        "            for cat in reg.get(\"secondary\", []):\n"
        "                if cat in self._category_map:\n"
        "                    self._category_map[cat][\"modules\"].add(tab_id)\n"
        "        self._cleanup_empty_categories()\n"
        "        self._rebuild_category_strip()\n"
        "        self._rebuild_spell_tabs()\n"
        "\n"
        "    def _cleanup_empty_categories(self) -> None:\n"
        "        remove = []\n"
        "        for cat, data in self._category_map.items():\n"
        "            if data.get(\"modules\"):\n"
        "                continue\n"
        "            if data.get(\"protected\") or cat in self._protected_categories:\n"
        "                continue\n"
        "            remove.append(cat)\n"
        "        for cat in remove:\n"
        "            self._category_map.pop(cat, None)\n"
        "        if self._active_category not in self._category_map:\n"
        "            self._active_category = next(iter(self._category_map.keys()), \"Core\")\n"
        "\n"
        "    def _load_spell_tab_state_from_config(self) -> None:\n",
        "inject category framework and lifecycle",
    )
    source = _replace_once(
        source,
        "        for tab in self._ordered_spell_tab_defs():\n"
        "            i = self._spell_tabs.addTab(tab[\"widget\"], tab[\"title\"])\n",
        "        for tab in self._ordered_spell_tab_defs():\n"
        "            if not self._visible_tab_for_active_category(tab):\n"
        "                continue\n"
        "            i = self._spell_tabs.addTab(tab[\"widget\"], tab[\"title\"])\n",
        "category-aware visible tabs",
    )
    return source


def _prune_optional_runtime_tabs(source: str, selected_modules: list[str], log_fn=None) -> str:
    selected = set(selected_modules or [])
    optional_tab_map = {
        "sl_scans": {"sl_scans"},
        "sl_commands": {"sl_commands"},
        "job_tracker": {"job_tracker"},
        "lessons_learned": {"lessons"},
        "dice_roller": {"dice_roller"},
        "magic_8ball": {"magic_8_ball"},
        "bill_scheduler": {"financial_planner"},
        "cvr_engine": {"cvr"},
        "session_browser": {"modules"},
    }
    remove_ids: set[str] = set()
    for mod_key, tab_ids in optional_tab_map.items():
        if mod_key not in selected:
            remove_ids.update(tab_ids)

    if not remove_ids:
        return source

    filtered_lines: list[str] = []
    removed = 0
    for line in source.splitlines():
        stripped = line.strip()
        if stripped.startswith('{"id": "') and '"title": "' in stripped:
            for tab_id in remove_ids:
                if f'"id": "{tab_id}"' in stripped:
                    removed += 1
                    break
            else:
                filtered_lines.append(line)
            continue
        filtered_lines.append(line)
    if log_fn:
        log_fn(f"[DECK] Optional tab rows removed: {removed}")
    return "\n".join(filtered_lines) + "\n"


def _get_deck_implementation(selected_modules: Optional[list[str]] = None, log_fn=None) -> str:
    """
    Returns the full embedded deck implementation.
    Completely self-contained — no external files required.

    Embedded implementation includes:
      - Draggable main left/right QSplitter shell layout
      - Minimum-width guards for both panes
      - Deck-config persistence/restore for main splitter sizes
    """
    import base64 as _base64
    if log_fn:
        log_fn("[DECK] Using embedded implementation")
    decoded = _base64.b64decode(_DECK_IMPL_B64).decode("utf-8")
    patched = _patch_embedded_deck_implementation(decoded, log_fn=log_fn)
    return _prune_optional_runtime_tabs(patched, selected_modules or [], log_fn=log_fn)

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — SETUP WORKER
#
# Runs the full deck creation sequence in a background thread.
# Emits log messages and progress for the UI to display.
# All file operations happen here — never on the main thread.
# ═══════════════════════════════════════════════════════════════════════════════


class DeckSetupWorker(QThread):
    """
    Background worker that creates a complete deck installation.

    Sequence:
      1. Create directory structure
      2. Extract and rename face pack (ZIP or folder)
      3. Convert icon PNG to ICO
      4. Generate sound WAV files
      5. Write config.json
      6. Initialize memory/session JSONL files
      7. Write deck Python file (injected template)
      8. Write requirements.txt
      9. Create desktop shortcut

    Signals:
        log(str)           — progress message for display
        progress(int)      — 0-100
        done(bool, str)    — success flag, result message or error
    """

    log      = Signal(str)
    progress = Signal(int)
    done     = Signal(bool, str)

    def __init__(
        self,
        deck_name:        str,
        persona:          dict,
        model_config:     dict,
        selected_modules: list[str],
        face_source:      str,          # path to ZIP or folder, or ""
        create_shortcut:  bool,
        output_base:      Path,         # e.g. D:\AI\Models
        ai_state_greetings: Optional[dict] = None,
    ):
        super().__init__()
        self._deck_name        = deck_name
        self._persona          = persona
        self._ai_state_greetings = ai_state_greetings
        self._model_config     = model_config
        self._selected_modules = selected_modules
        self._face_source      = face_source
        self._create_shortcut  = create_shortcut
        self._output_base      = output_base

    def _log(self, msg: str) -> None:
        self.log.emit(msg)

    def run(self) -> None:
        try:
            deck_lower = self._deck_name.lower().replace(" ", "_")
            deck_home  = self._output_base / self._deck_name

            self._log(f"Creating deck: {self._deck_name}")
            self._log(f"Location: {deck_home}")

            # ── 1. Create directory structure ──────────────────────────
            self._log("Creating directory structure...")
            dirs = [
                deck_home / "Faces",
                deck_home / "sounds",
                deck_home / "memories",
                deck_home / "sessions",
                deck_home / "sl",
                deck_home / "exports",
                deck_home / "logs",
                deck_home / "backups",
                deck_home / "personas",
                deck_home / "config",
            ]
            for d in dirs:
                d.mkdir(parents=True, exist_ok=True)
            self._log("✓ Directory structure created")
            self.progress.emit(10)

            # ── 2. Extract and rename face pack ────────────────────────
            face_prefix = self._persona.get("face_prefix", self._deck_name)
            icon_path   = None

            if self._face_source and Path(self._face_source).exists():
                self._log(f"Extracting face pack...")
                face_result = extract_faces(
                    source      = self._face_source,
                    dest_dir    = deck_home / "Faces",
                    face_prefix = face_prefix,
                    deck_name   = self._deck_name,
                    log_fn      = self._log,
                )
                if face_result["errors"]:
                    for err in face_result["errors"]:
                        self._log(f"  ⚠ {err}")
                self._log(
                    f"✓ Faces: {face_result['extracted']} extracted, "
                    f"{face_result['skipped']} skipped, "
                    f"{face_result['unrecognized']} unrecognized"
                )
                icon_path = face_result.get("icon_path")
            else:
                self._log(
                    "⚠ No face pack provided — using placeholder. "
                    "Add faces to Faces/ later."
                )
            self.progress.emit(25)

            # ── 3. Convert icon ────────────────────────────────────────
            ico_path = None
            if icon_path and icon_path.exists():
                ico_path = deck_home / f"{self._deck_name}.ico"
                self._log(f"Converting icon: {icon_path.name} → {ico_path.name}")
                ok = convert_png_to_ico(icon_path, ico_path, log_fn=self._log)
                if ok:
                    self._log(f"✓ Icon created: {ico_path.name}")
                else:
                    self._log("⚠ Icon conversion failed — shortcut will use default icon")
                    ico_path = None
            self.progress.emit(35)

            # ── 4. Generate sound files ────────────────────────────────
            sound_profile = self._persona.get("sound_profile", "default_chime")
            self._log(f"Generating sounds (profile: {sound_profile})...")
            generate_sounds_for_profile(
                profile_name    = sound_profile,
                sounds_dir      = deck_home / "sounds",
                deck_name_prefix= deck_lower,
                log_fn          = self._log,
            )
            self._log("✓ Sound files ready")
            self.progress.emit(50)

            # ── 5. Write config.json ───────────────────────────────────
            self._log("Writing config.json...")
            self._write_config(deck_home, deck_lower)
            self._log("✓ config.json written")
            self.progress.emit(60)

            # ── 6. Initialize memory files ────────────────────────────
            self._log("Initializing memory files...")
            self._init_memory_files(deck_home)
            self._log("✓ Memory files initialized")
            self.progress.emit(68)

            # ── 7. Write deck Python file ──────────────────────────────
            deck_filename = f"{deck_lower}_deck.py"
            deck_path     = deck_home / deck_filename
            self._log(f"Writing deck file: {deck_filename}...")

            ok = build_deck_file(
                persona          = self._persona,
                deck_name        = self._deck_name,
                selected_modules = self._selected_modules,
                output_path      = deck_path,
                model_config     = self._model_config,
                ai_state_greetings = self._ai_state_greetings,
                log_fn           = self._log,
            )
            if not ok:
                self.done.emit(False, "Deck file write failed. Check log for details.")
                return
            self._log(f"✓ Deck file written: {deck_path.name}")
            self.progress.emit(82)

            # ── 8. Write requirements.txt ─────────────────────────────
            self._write_requirements(deck_home, self._selected_modules)
            self._log("✓ requirements.txt written")
            self.progress.emit(90)

            # ── 9. Create desktop shortcut ────────────────────────────
            shortcut_path = None
            if self._create_shortcut:
                ok, msg, sc_path = self._create_shortcut_fn(
                    deck_path, ico_path
                )
                shortcut_path = sc_path
                self._log(f"{'✓' if ok else '⚠'} Shortcut: {msg}")
            self.progress.emit(100)

            # ── Done ──────────────────────────────────────────────────
            self._log("═" * 55)
            self._log(f"✦ {self._deck_name.upper()} DECK CREATED SUCCESSFULLY ✦")
            self._log(f"Location: {deck_home}")
            self._log(f"Deck file: {deck_path.name}")
            if shortcut_path:
                self._log(f"Shortcut: {shortcut_path}")
            self._log("═" * 55)

            self.done.emit(True, str(deck_path))

        except Exception as e:
            import traceback
            self._log(f"✗ SETUP FAILED: {e}")
            self._log(traceback.format_exc())
            self.done.emit(False, str(e))

    # ── HELPERS ────────────────────────────────────────────────────────────────

    def _write_config(self, deck_home: Path, deck_lower: str) -> None:
        m = self._model_config
        cfg = {
            "deck_name":    self._deck_name,
            "deck_version": DECK_VERSION,
            "base_dir":     str(deck_home),
            "model": {
                "type":         m.get("type",         "ollama"),
                "path":         m.get("path",         ""),
                "ollama_model": m.get("ollama_model",  ""),
                "api_key":      m.get("api_key",       ""),
                "api_type":     m.get("api_type",      ""),
                "api_model":    m.get("api_model",     ""),
            },
            "paths": {
                "faces":    str(deck_home / "Faces"),
                "sounds":   str(deck_home / "sounds"),
                "memories": str(deck_home / "memories"),
                "sessions": str(deck_home / "sessions"),
                "sl":       str(deck_home / "sl"),
                "exports":  str(deck_home / "exports"),
                "logs":     str(deck_home / "logs"),
                "backups":  str(deck_home / "backups"),
                "personas": str(deck_home / "personas"),
            },
            "settings": {
                "idle_enabled":              False,
                "idle_min_minutes":          10,
                "idle_max_minutes":          30,
                "autosave_interval_minutes": 10,
                "max_backups":               10,
                "sound_enabled":             True,
                "auto_wake_on_relief":       False,
                "runtime_persona_mode":      "persona",
                "runtime_theme_mode":        "auto",
                "theme_mode":                "auto",
            },
            "persona": {
                "name":              self._deck_name,
                "face_prefix":       self._persona.get("face_prefix", self._deck_name),
                "sound_profile":     self._persona.get("sound_profile", "default_chime"),
                "vampire_states":    bool(self._persona.get("vampire_states", False)),
                "torpor_system":     bool(self._persona.get("torpor_system", False)),
            },
            "modules": {
                k: (k in self._selected_modules)
                for k in MODULES
            },
            "first_run": False,
        }
        (deck_home / "config.json").write_text(
            json.dumps(cfg, indent=2), encoding="utf-8"
        )

    def _init_memory_files(self, deck_home: Path) -> None:
        mem_dir = deck_home / "memories"
        for fname in (
            "messages.jsonl", "memories.jsonl", "tasks.jsonl",
            "lessons_learned.jsonl", "persona_history.jsonl",
            "job_tracker.jsonl",
        ):
            fp = mem_dir / fname
            if not fp.exists():
                fp.write_text("", encoding="utf-8")

        sessions_dir = deck_home / "sessions"
        idx = sessions_dir / "session_index.json"
        if not idx.exists():
            idx.write_text(
                json.dumps({"sessions": []}, indent=2), encoding="utf-8"
            )

        state = mem_dir / "state.json"
        if not state.exists():
            state.write_text(json.dumps({
                "persona_name":              self._deck_name,
                "deck_version":              DECK_VERSION,
                "session_count":             0,
                "last_startup":              None,
                "last_shutdown":             None,
                "last_active":               None,
                "total_messages":            0,
                "internal_narrative":        {},
                "vampire_state_at_shutdown": "DORMANT",
            }, indent=2), encoding="utf-8")

        index = mem_dir / "index.json"
        if not index.exists():
            index.write_text(json.dumps({
                "version":        DECK_VERSION,
                "total_messages": 0,
                "total_memories": 0,
            }, indent=2), encoding="utf-8")

    def _write_requirements(self, deck_home: Path,
                             selected_modules: list[str]) -> None:
        # Collect required pip packages from selected modules
        extra_reqs: set[str] = set()
        for mod_key in selected_modules:
            mod = MODULES.get(mod_key, {})
            for req in mod.get("requires", []):
                extra_reqs.add(req)

        lines = [
            f"# {self._deck_name} Deck — Dependencies",
            f"# Generated by deck_builder.py {DECK_VERSION}",
            f"# Install: pip install -r requirements.txt",
            "",
            "# Core",
            "PySide6",
            "apscheduler",
            "loguru",
            "pygame",
            "pywin32",
            "psutil",
            "requests",
            "",
            "# Optional — NVIDIA GPU monitoring",
            "# pynvml",
            "",
            "# Optional — local HuggingFace model",
            "# torch",
            "# transformers",
            "# accelerate",
            "",
            "# Optional — icon conversion",
            "# Pillow",
        ]

        if extra_reqs - {"psutil", "pynvml", "cryptography"}:
            lines.extend(["", "# Additional module requirements"])
            for req in sorted(extra_reqs):
                lines.append(req)

        (deck_home / "requirements.txt").write_text(
            "\n".join(lines), encoding="utf-8"
        )

    def _create_shortcut_fn(
        self, deck_path: Path, ico_path: Optional[Path]
    ) -> tuple[bool, str, Optional[Path]]:
        try:
            import win32com.client
            desktop = Path.home() / "Desktop"
            sc_path = desktop / f"{self._deck_name}.lnk"

            pythonw = Path(sys.executable)
            if pythonw.name.lower() == "python.exe":
                candidate = pythonw.parent / "pythonw.exe"
                if candidate.exists():
                    pythonw = candidate

            shell = win32com.client.Dispatch("WScript.Shell")
            sc    = shell.CreateShortCut(str(sc_path))
            sc.TargetPath       = str(pythonw)
            sc.Arguments        = f'"{deck_path}"'
            sc.WorkingDirectory = str(deck_path.parent)
            sc.Description      = f"{self._deck_name} — Echo Deck"

            if ico_path and ico_path.exists():
                sc.IconLocation = f"{ico_path},0"

            sc.save()
            return True, f"Created: {sc_path.name}", sc_path

        except ImportError:
            return (
                False,
                "pywin32 not installed — shortcut not created. "
                "Run: pip install pywin32",
                None,
            )
        except Exception as e:
            return False, f"Shortcut error: {e}", None


# ── END OF SECTION 6 ──────────────────────────────────────────────────────────
# Next: Section 7 — UI helper widgets


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — UI HELPER WIDGETS
#
# Reusable UI components for the builder dialog.
# Light/dark mode aware — read current scheme via S() function.
# ═══════════════════════════════════════════════════════════════════════════════


def _get_style() -> str:
    """Generate stylesheet from current active scheme."""
    return f"""
    QDialog, QWidget, QMainWindow {{
        background-color: {S('bg')};
        color: {S('text')};
        font-family: 'Segoe UI', 'Arial', sans-serif;
        font-size: 11px;
    }}
    QLabel {{
        color: {S('text')};
        border: none;
    }}
    QLineEdit {{
        background-color: {S('bg3')};
        color: {S('text')};
        border: 1px solid {S('border')};
        border-radius: 3px;
        padding: 6px 10px;
        font-size: 11px;
    }}
    QLineEdit:focus {{
        border: 1px solid {S('primary')};
    }}
    QLineEdit:disabled {{
        background-color: {S('bg2')};
        color: {S('text_dim')};
    }}
    QPushButton {{
        background-color: {S('primary_dim')};
        color: {S('text')};
        border: 1px solid {S('primary')};
        border-radius: 3px;
        padding: 6px 14px;
        font-size: 11px;
        font-weight: bold;
    }}
    QPushButton:hover {{
        background-color: {S('primary')};
        color: #ffffff;
    }}
    QPushButton:pressed {{
        background-color: {S('primary')};
        border-color: {S('gold')};
    }}
    QPushButton:disabled {{
        background-color: {S('bg2')};
        color: {S('text_dim')};
        border-color: {S('border')};
    }}
    QPushButton#primary_btn {{
        background-color: {S('primary')};
        color: #ffffff;
        border: 2px solid {S('primary')};
        padding: 8px 20px;
        font-size: 12px;
    }}
    QPushButton#primary_btn:hover {{
        background-color: {S('gold')};
        border-color: {S('gold')};
        color: {S('bg')};
    }}
    QPushButton#primary_btn:disabled {{
        background-color: {S('bg2')};
        color: {S('text_dim')};
        border-color: {S('border')};
    }}
    QComboBox {{
        background-color: {S('bg3')};
        color: {S('text')};
        border: 1px solid {S('border')};
        border-radius: 3px;
        padding: 6px 10px;
        font-size: 11px;
    }}
    QComboBox:focus {{
        border: 1px solid {S('primary')};
    }}
    QComboBox::drop-down {{
        border: none;
        width: 20px;
    }}
    QComboBox QAbstractItemView {{
        background-color: {S('bg2')};
        color: {S('text')};
        border: 1px solid {S('border')};
        selection-background-color: {S('primary_dim')};
        selection-color: {S('text')};
    }}
    QCheckBox {{
        color: {S('text')};
        spacing: 6px;
    }}
    QCheckBox::indicator {{
        width: 14px;
        height: 14px;
        border: 1px solid {S('border')};
        border-radius: 2px;
        background: {S('bg3')};
    }}
    QCheckBox::indicator:checked {{
        background: {S('primary')};
        border-color: {S('primary')};
    }}
    QRadioButton {{
        color: {S('text')};
        spacing: 6px;
    }}
    QRadioButton::indicator {{
        width: 14px;
        height: 14px;
        border: 1px solid {S('border')};
        border-radius: 7px;
        background: {S('bg3')};
    }}
    QRadioButton::indicator:checked {{
        background: {S('primary')};
        border-color: {S('primary')};
    }}
    QGroupBox {{
        border: 1px solid {S('border')};
        border-radius: 4px;
        margin-top: 8px;
        padding-top: 8px;
        font-weight: bold;
        color: {S('text_dim')};
        font-size: 10px;
        letter-spacing: 1px;
    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        subcontrol-position: top left;
        padding: 0 4px;
        background-color: {S('bg')};
    }}
    QTextEdit {{
        background-color: {S('bg3')};
        color: {S('text')};
        border: 1px solid {S('border')};
        border-radius: 3px;
        font-family: 'Courier New', monospace;
        font-size: 10px;
        padding: 4px;
    }}
    QProgressBar {{
        background-color: {S('bg3')};
        border: 1px solid {S('border')};
        border-radius: 3px;
        text-align: center;
        color: {S('text')};
        font-size: 10px;
        height: 16px;
    }}
    QProgressBar::chunk {{
        background-color: {S('primary')};
        border-radius: 2px;
    }}
    QScrollArea {{
        border: none;
        background: transparent;
    }}
    QScrollBar:vertical {{
        background: {S('bg2')};
        width: 8px;
        border: none;
    }}
    QScrollBar::handle:vertical {{
        background: {S('border')};
        border-radius: 4px;
    }}
    QScrollBar::handle:vertical:hover {{
        background: {S('primary')};
    }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
        height: 0;
    }}
    QFrame[frameShape="4"] {{
        color: {S('border')};
        border: 1px solid {S('border')};
    }}
    """


def section_label(text: str) -> QLabel:
    """Small caps section header label."""
    lbl = QLabel(text.upper())
    lbl.setStyleSheet(
        f"color: {S('text_dim')}; font-size: 9px; font-weight: bold; "
        f"letter-spacing: 2px; font-family: 'Segoe UI', Arial, sans-serif;"
    )
    return lbl


def h_divider() -> QFrame:
    f = QFrame()
    f.setFrameShape(QFrame.Shape.HLine)
    f.setStyleSheet(f"color: {S('border')}; border: 1px solid {S('border')};")
    return f


def status_label(text: str = "", ok: bool = True) -> QLabel:
    lbl = QLabel(text)
    color = S("green") if ok else S("red")
    lbl.setStyleSheet(
        f"color: {color}; font-size: 11px; "
        f"font-family: 'Segoe UI', Arial, sans-serif;"
    )
    return lbl


def primary_btn(text: str) -> QPushButton:
    btn = QPushButton(text)
    btn.setObjectName("primary_btn")
    return btn


class PersistedRightSplitterHandle(QSplitterHandle):
    """Splitter handle with right-click lock/unlock support."""

    def mousePressEvent(self, event) -> None:
        splitter = self.splitter()
        if event.button() == Qt.MouseButton.RightButton and isinstance(splitter, PersistedRightSplitter):
            splitter.show_handle_menu(self.mapToGlobal(event.position().toPoint()))
            event.accept()
            return
        if event.button() == Qt.MouseButton.LeftButton and isinstance(splitter, PersistedRightSplitter):
            if splitter.is_locked():
                event.accept()
                return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event) -> None:
        splitter = self.splitter()
        if isinstance(splitter, PersistedRightSplitter) and splitter.is_locked():
            event.accept()
            return
        super().mouseMoveEvent(event)


class PersistedRightSplitter(QSplitter):
    """QSplitter that persists left-pane size and lock state."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(Qt.Orientation.Horizontal, parent)
        self._locked = False
        self._loaded_left_width: Optional[int] = None
        self._load_state()
        self.splitterMoved.connect(self._on_splitter_moved)

    def createHandle(self) -> QSplitterHandle:
        return PersistedRightSplitterHandle(self.orientation(), self)

    def is_locked(self) -> bool:
        return self._locked

    def set_locked(self, locked: bool) -> None:
        self._locked = bool(locked)
        self.save_state()

    def show_handle_menu(self, global_pos) -> None:
        menu = QMenu(self)
        action_label = "Unlock" if self._locked else "Lock"
        action = menu.addAction(action_label)
        chosen = menu.exec(global_pos)
        if chosen == action:
            self.set_locked(not self._locked)

    def apply_saved_position(self) -> None:
        if self._loaded_left_width is None:
            return
        total = max(1, self.size().width())
        left = max(120, min(self._loaded_left_width, total - 120))
        self.setSizes([left, total - left])

    def _on_splitter_moved(self, _pos: int, _index: int) -> None:
        self.save_state()

    def _load_state(self) -> None:
        try:
            if not BUILDER_UI_STATE_PATH.exists():
                return
            data = json.loads(BUILDER_UI_STATE_PATH.read_text(encoding="utf-8"))
            self._loaded_left_width = int(data.get("right_splitter_left_width", 0)) or None
            self._locked = bool(data.get("right_splitter_locked", False))
        except Exception:
            self._loaded_left_width = None
            self._locked = False

    def save_state(self) -> None:
        try:
            current_left = None
            sizes = self.sizes()
            if sizes:
                current_left = sizes[0]
            data = {}
            if BUILDER_UI_STATE_PATH.exists():
                data = json.loads(BUILDER_UI_STATE_PATH.read_text(encoding="utf-8"))
            if current_left is not None:
                data["right_splitter_left_width"] = int(current_left)
            data["right_splitter_locked"] = self._locked
            BUILDER_UI_STATE_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
        except Exception:
            pass


# ── BUILT-IN PERSONAS ─────────────────────────────────────────────────────────

BUILTIN_PERSONAS: dict[str, dict] = {
    "Default": {
        "_loaded_name": "Default",
        "description":  "Base model — no persona constraints. Minimal system prompt. Good for testing.",
        "system_prompt": "You are a helpful AI assistant. Respond clearly and accurately.",
        "cognitive_anchors": [],
        "colors": {
            "primary":    "#5588aa",
            "secondary":  "#336677",
            "accent":     "#336677",
            "background": "#0a0a0f",
            "panel":      "#0f0f18",
            "border":     "#2a2a40",
            "text":       "#ccccdd",
            "text_dim":   "#555566",
        },
        "font_family":    "sans-serif",
        "gender":         "they/them",
        "awakening_line": "Ready.",
        "ui_labels": {
            "window_title":       "ECHO DECK",
            "chat_window":        "CONVERSATION",
            "send_button":        "SEND",
            "input_placeholder":  "Type your message...",
            "generating_status":  "◉ THINKING",
            "idle_status":        "◉ IDLE",
            "offline_status":     "◉ OFFLINE",
            "torpor_status":      "◉ SUSPENDED",
            "runes":              "— ECHO DECK —",
        },
        "sound_profile":       "default_chime",
        "special_systems":     [],
        "vampire_states":      False,
        "torpor_system":       False,
        "anchor_entity":       None,
        "face_prefix":         "Default",
        "pronouns": {"subject": "it", "object": "it", "possessive": "its"},
    },
    "Assistant": {
        "_loaded_name": "Assistant",
        "description":  "Professional, warm, task-focused. Slate blue and amber.",
        "system_prompt": (
            "You are a knowledgeable assistant who responds clearly, helpfully, and with care. "
            "You focus on the task at hand, provide accurate information, and communicate in a "
            "warm professional tone. You are direct without being cold, thorough without being verbose."
        ),
        "cognitive_anchors": [
            "What is the user actually trying to accomplish?",
            "What is the most accurate and useful answer?",
            "Is there anything missing that would help them succeed?",
        ],
        "colors": {
            "primary":    "#6688bb",
            "secondary":  "#cc9944",
            "accent":     "#cc9944",
            "background": "#080c14",
            "panel":      "#0d1220",
            "border":     "#2a3450",
            "text":       "#e8e4dc",
            "text_dim":   "#606070",
        },
        "font_family":    "sans-serif",
        "gender":         "they/them",
        "awakening_line": "Ready to assist.",
        "ui_labels": {
            "window_title":       "ECHO DECK — ASSISTANT",
            "chat_window":        "WORKSPACE",
            "send_button":        "SEND",
            "input_placeholder":  "How can I help?",
            "generating_status":  "◉ WORKING",
            "idle_status":        "◉ READY",
            "offline_status":     "◉ OFFLINE",
            "torpor_status":      "◉ SUSPENDED",
            "runes":              "— ASSISTANT —",
        },
        "sound_profile":       "warm_chime",
        "special_systems":     [],
        "vampire_states":      False,
        "torpor_system":       False,
        "anchor_entity":       None,
        "face_prefix":         "Assistant",
        "pronouns": {"subject": "they", "object": "them", "possessive": "their"},
    },
    "Friend": {
        "_loaded_name": "Friend",
        "description":  "Casual, conversational, non-judgmental. Teal and warm cream.",
        "system_prompt": (
            "You are a friendly conversational companion. You are warm, genuine, and easy to talk to. "
            "You do not moralize or lecture. You engage with whatever the person wants to discuss — "
            "serious or silly — without judgment. You speak like a real person, not a customer service bot."
        ),
        "cognitive_anchors": [
            "What does this person actually want to talk about?",
            "How can I engage genuinely rather than generically?",
        ],
        "colors": {
            "primary":    "#449988",
            "secondary":  "#cc7755",
            "accent":     "#cc7755",
            "background": "#0c0a08",
            "panel":      "#141210",
            "border":     "#2a2420",
            "text":       "#e8ddd0",
            "text_dim":   "#605850",
        },
        "font_family":    "sans-serif",
        "gender":         "they/them",
        "awakening_line": "Hey, what's up?",
        "ui_labels": {
            "window_title":       "ECHO DECK — FRIEND",
            "chat_window":        "CHAT",
            "send_button":        "SEND",
            "input_placeholder":  "Say anything...",
            "generating_status":  "◉ TYPING",
            "idle_status":        "◉ HERE",
            "offline_status":     "◉ OFFLINE",
            "torpor_status":      "◉ AWAY",
            "runes":              "— FRIEND —",
        },
        "sound_profile":       "soft_ping",
        "special_systems":     [],
        "vampire_states":      False,
        "torpor_system":       False,
        "anchor_entity":       None,
        "face_prefix":         "Friend",
        "pronouns": {"subject": "they", "object": "them", "possessive": "their"},
    },
}


def _neutral_state_greetings(deck_name: str) -> dict[str, str]:
    dn = deck_name.strip() or "Deck"
    return {
        "WITCHING HOUR":   f"{dn} is online and ready to assist right now.",
        "DEEP NIGHT":      f"{dn} remains focused and available for your request.",
        "TWILIGHT FADING": f"{dn} is attentive and waiting for your next prompt.",
        "DORMANT":         f"{dn} is in a low-activity mode but still responsive.",
        "RESTLESS SLEEP":  f"{dn} is lightly idle and can re-engage immediately.",
        "STIRRING":        f"{dn} is becoming active and ready to continue.",
        "AWAKENED":        f"{dn} is fully active and prepared to help.",
        "HUNTING":         f"{dn} is in an active processing window and standing by.",
    }


def _coerce_state_greetings(raw_map: object) -> Optional[dict[str, str]]:
    if not isinstance(raw_map, dict):
        return None
    clean: dict[str, str] = {}
    for key in AI_STATE_KEYS:
        val = raw_map.get(key)
        if not isinstance(val, str):
            return None
        sentence = " ".join(val.strip().split())
        if not sentence:
            return None
        clean[key] = sentence
    if set(raw_map.keys()) != set(AI_STATE_KEYS):
        return None
    return clean


def _extract_json_object(text: str) -> Optional[dict]:
    text = (text or "").strip()
    if not text:
        return None
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        try:
            obj = json.loads(text[start:end + 1])
            if isinstance(obj, dict):
                return obj
        except Exception:
            return None
    return None


def _build_state_greeting_prompt(persona: dict, deck_name: str) -> str:
    anchors = persona.get("cognitive_anchors", []) or []
    anchors_text = "\n".join(f"- {a}" for a in anchors) if anchors else "- None"
    pronouns = persona.get("pronouns", {}) or {}
    return (
        "Generate state greeting lines for an AI deck persona.\n"
        "Return STRICT JSON ONLY with exactly these keys and no extra text:\n"
        "{\n"
        '"WITCHING HOUR": "...",\n'
        '"DEEP NIGHT": "...",\n'
        '"TWILIGHT FADING": "...",\n'
        '"DORMANT": "...",\n'
        '"RESTLESS SLEEP": "...",\n'
        '"STIRRING": "...",\n'
        '"AWAKENED": "...",\n'
        '"HUNTING": "..."\n'
        "}\n\n"
        "Rules:\n"
        "- Exact key names only.\n"
        "- Exactly one sentence per value.\n"
        "- No markdown.\n"
        "- No commentary.\n"
        "- No extra keys.\n\n"
        f"Deck name: {deck_name}\n"
        f"System prompt: {persona.get('system_prompt', '')}\n"
        f"Cognitive anchors:\n{anchors_text}\n"
        f"Awakening line: {persona.get('awakening_line', '')}\n"
        f"Pronouns: subject={pronouns.get('subject', '')}, "
        f"object={pronouns.get('object', '')}, "
        f"possessive={pronouns.get('possessive', '')}\n"
    )


def _request_state_greetings(model_config: dict, persona: dict, deck_name: str) -> Optional[dict[str, str]]:
    model_type = (model_config or {}).get("type", "")
    prompt = _build_state_greeting_prompt(persona, deck_name)
    if model_type == "ollama":
        model = (model_config or {}).get("ollama_model", "").strip()
        if not model:
            return None
        payload = json.dumps({
            "model": model,
            "prompt": prompt,
            "stream": False,
            "options": {"temperature": 0.4},
        }).encode("utf-8")
        req = urllib.request.Request(
            "http://localhost:11434/api/generate",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=45) as resp:
                data = json.loads(resp.read().decode("utf-8", errors="ignore"))
            raw = data.get("response", "")
        except Exception:
            return None
        return _coerce_state_greetings(_extract_json_object(raw))
    return None

# ── PERSONA PANEL ──────────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7B — PERSONA PANEL
# Three built-in radio options (Default, Assistant, Friend) always present.
# Load from file for any custom persona (Morganna, GrimVeil, etc.)
# ═══════════════════════════════════════════════════════════════════════════════

class PersonaPanel(QWidget):
    """
    Persona selection panel.
    Three built-in personas selectable via radio buttons.
    Custom personas load from .txt template files.
    Export template button creates a fillable .txt file.
    """

    persona_changed = Signal(str, dict)  # (name, persona_dict)
    generate_state_greetings_requested = Signal()
    regenerate_state_greetings_requested = Signal()
    use_state_greetings_requested = Signal()
    use_neutral_state_greetings_requested = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._loaded_persona: Optional[dict] = None
        self._loaded_name:    Optional[str]  = None
        self._radio_group = QButtonGroup(self)
        self._builtin_name_by_id = {0: "Default", 1: "Assistant", 2: "Friend"}
        self._builtin_pronoun_map = {
            "male":   {"subject": "he", "object": "him", "possessive": "his"},
            "female": {"subject": "she", "object": "her", "possessive": "her"},
            "none":   {"subject": "", "object": "", "possessive": ""},
        }
        self._state_greetings_preview: Optional[dict[str, str]] = None
        self._setup_ui()

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(6)

        root.addWidget(section_label("Persona"))

        # ── Built-in radio options ─────────────────────────────────────────
        radio_style = (
            f"color: {S('text')}; font-size: 10px;"
        )
        dim_style = (
            f"color: {S('text_dim')}; font-size: 9px; "
            f"font-style: italic; margin-left: 22px; margin-bottom: 2px;"
        )

        self._radio_default = QRadioButton("Default  (base model, no persona)")
        self._radio_default.setStyleSheet(radio_style)
        self._radio_default.setChecked(True)
        self._radio_group.addButton(self._radio_default, 0)
        root.addWidget(self._radio_default)

        self._radio_assistant = QRadioButton("Assistant  (professional, task-focused)")
        self._radio_assistant.setStyleSheet(radio_style)
        self._radio_group.addButton(self._radio_assistant, 1)
        root.addWidget(self._radio_assistant)

        self._radio_friend = QRadioButton("Friend  (casual, conversational)")
        self._radio_friend.setStyleSheet(radio_style)
        self._radio_group.addButton(self._radio_friend, 2)
        root.addWidget(self._radio_friend)

        # ── Built-in gender / pronouns ─────────────────────────────────────
        self._builtin_gender_box = QGroupBox("Built-in Persona Pronouns")
        self._builtin_gender_box.setStyleSheet(
            f"QGroupBox {{ color: {S('text')}; border: 1px solid {S('border')}; margin-top: 6px; }}"
            f"QGroupBox::title {{ subcontrol-origin: margin; left: 8px; padding: 0 4px; }}"
        )
        builtin_gender_layout = QVBoxLayout(self._builtin_gender_box)
        builtin_gender_layout.setContentsMargins(8, 10, 8, 8)
        builtin_gender_layout.setSpacing(4)

        gender_row = QHBoxLayout()
        self._builtin_gender_group = QButtonGroup(self)
        self._gender_male = QRadioButton("Male")
        self._gender_female = QRadioButton("Female")
        self._gender_none = QRadioButton("None")
        self._gender_other = QRadioButton("Other")
        for idx, btn in enumerate([self._gender_male, self._gender_female, self._gender_none, self._gender_other]):
            btn.setStyleSheet(radio_style)
            self._builtin_gender_group.addButton(btn, idx)
            gender_row.addWidget(btn)
        self._gender_none.setChecked(True)
        builtin_gender_layout.addLayout(gender_row)

        custom_row = QHBoxLayout()
        self._custom_subject = QLineEdit()
        self._custom_object = QLineEdit()
        self._custom_possessive = QLineEdit()
        self._custom_subject.setPlaceholderText("subject")
        self._custom_object.setPlaceholderText("object")
        self._custom_possessive.setPlaceholderText("possessive")
        for field in (self._custom_subject, self._custom_object, self._custom_possessive):
            field.setStyleSheet(
                f"background: {S('bg3')}; color: {S('text')}; border: 1px solid {S('border')}; padding: 3px 6px;"
            )
            custom_row.addWidget(field)
        builtin_gender_layout.addLayout(custom_row)
        root.addWidget(self._builtin_gender_box)

        hint = QLabel("All other personas (Morganna, GrimVeil, VEX, etc.) load via file below.")
        hint.setStyleSheet(dim_style)
        hint.setWordWrap(True)
        root.addWidget(hint)

        # ── Load from file ─────────────────────────────────────────────────
        self._radio_file = QRadioButton("Load from file  (.txt template)")
        self._radio_file.setStyleSheet(radio_style)
        self._radio_group.addButton(self._radio_file, 3)
        root.addWidget(self._radio_file)

        file_row = QHBoxLayout()
        self._file_path = QLineEdit()
        self._file_path.setPlaceholderText(
            "Browse to persona template .txt file  (or drag & drop)"
        )
        self._file_path.setReadOnly(True)
        self._file_path.setStyleSheet(
            f"background: {S('bg3')}; color: {S('text')}; "
            f"border: 1px solid {S('border')}; padding: 4px 8px;"
        )
        browse_btn = QPushButton("Browse")
        browse_btn.setFixedWidth(70)
        browse_btn.clicked.connect(self._browse_file)
        file_row.addWidget(self._file_path, 1)
        file_row.addWidget(browse_btn)
        root.addLayout(file_row)

        # Export template button
        export_btn = QPushButton("Export Blank Template")
        export_btn.setToolTip(
            "Save a blank persona template .txt file that you can fill in"
        )
        export_btn.clicked.connect(self._export_template)
        export_btn.setStyleSheet(
            f"background: {S('primary_dim')}; color: {S('text')}; "
            f"border: 1px solid {S('primary')}; padding: 4px 12px; font-size: 10px;"
        )
        root.addWidget(export_btn)

        # Preview
        self._preview = QLabel("No persona loaded.")
        self._preview.setWordWrap(True)
        self._preview.setStyleSheet(
            f"color: {S('text_dim')}; font-size: 9px; "
            f"font-style: italic; padding: 4px 0;"
        )
        root.addWidget(self._preview)

        self._btn_generate_states = QPushButton("Generate State Greetings")
        self._btn_generate_states.setEnabled(False)
        self._btn_generate_states.clicked.connect(
            self.generate_state_greetings_requested.emit
        )
        root.addWidget(self._btn_generate_states)

        self._state_preview = QLabel("")
        self._state_preview.setWordWrap(True)
        self._state_preview.setStyleSheet(
            f"color: {S('text_dim')}; font-size: 9px; padding: 2px 0;"
        )
        self._state_preview.setVisible(False)
        root.addWidget(self._state_preview)

        self._state_btn_row = QHBoxLayout()
        self._btn_regen_states = QPushButton("Regenerate")
        self._btn_use_states = QPushButton("Use These")
        self._btn_neutral_states = QPushButton("Use Neutral Defaults")
        self._btn_regen_states.clicked.connect(self.regenerate_state_greetings_requested.emit)
        self._btn_use_states.clicked.connect(self.use_state_greetings_requested.emit)
        self._btn_neutral_states.clicked.connect(self.use_neutral_state_greetings_requested.emit)
        self._state_btn_row.addWidget(self._btn_regen_states)
        self._state_btn_row.addWidget(self._btn_use_states)
        self._state_btn_row.addWidget(self._btn_neutral_states)
        root.addLayout(self._state_btn_row)
        self._set_state_buttons_visible(False)

        # Connect radio group
        self._radio_group.idClicked.connect(self._on_radio_changed)
        self._builtin_gender_group.idClicked.connect(self._on_builtin_pronouns_changed)
        self._custom_subject.textChanged.connect(self._on_builtin_pronouns_changed)
        self._custom_object.textChanged.connect(self._on_builtin_pronouns_changed)
        self._custom_possessive.textChanged.connect(self._on_builtin_pronouns_changed)

        # Emit initial selection
        self._update_builtin_pronoun_controls()
        self._emit_builtin("Default")

    def _on_radio_changed(self, radio_id: int) -> None:
        if radio_id in self._builtin_name_by_id:
            self._loaded_persona = None
            self._loaded_name    = None
            self._file_path.setText("")
            self._preview.setText("")
            self._update_builtin_pronoun_controls()
            self._emit_builtin(self._builtin_name_by_id[radio_id])
            return
        self._update_builtin_pronoun_controls()
        # radio_id == 3 (Load from file) — do nothing until file is browsed

    def _builtin_selection_name(self) -> Optional[str]:
        radio_id = self._radio_group.checkedId()
        return self._builtin_name_by_id.get(radio_id)

    def _current_builtin_pronouns(self) -> dict[str, str]:
        gender_id = self._builtin_gender_group.checkedId()
        if gender_id == 0:
            return dict(self._builtin_pronoun_map["male"])
        if gender_id == 1:
            return dict(self._builtin_pronoun_map["female"])
        if gender_id == 2:
            return dict(self._builtin_pronoun_map["none"])
        return {
            "subject": self._custom_subject.text().strip(),
            "object": self._custom_object.text().strip(),
            "possessive": self._custom_possessive.text().strip(),
        }

    def _update_builtin_pronoun_controls(self) -> None:
        is_builtin = self._builtin_selection_name() is not None
        self._builtin_gender_box.setVisible(is_builtin)
        custom_visible = is_builtin and self._builtin_gender_group.checkedId() == 3
        self._custom_subject.setVisible(custom_visible)
        self._custom_object.setVisible(custom_visible)
        self._custom_possessive.setVisible(custom_visible)

    def _on_builtin_pronouns_changed(self, *_args) -> None:
        self._update_builtin_pronoun_controls()
        name = self._builtin_selection_name()
        if name:
            self._emit_builtin(name)

    def _emit_builtin(self, name: str) -> None:
        persona = dict(BUILTIN_PERSONAS[name])
        persona["pronouns"] = self._current_builtin_pronouns()
        self._preview.setText(
            f"Built-in: {name}\n{persona['description']}"
        )
        self._preview.setStyleSheet(
            f"color: {S('primary')}; font-size: 9px; padding: 4px 0;"
        )
        self.persona_changed.emit(name, persona)

    def set_generation_enabled(self, enabled: bool) -> None:
        self._btn_generate_states.setEnabled(bool(enabled))

    def _set_state_buttons_visible(self, visible: bool) -> None:
        self._btn_regen_states.setVisible(visible)
        self._btn_use_states.setVisible(visible)
        self._btn_neutral_states.setVisible(visible)

    def set_state_greetings_preview(self, greetings: dict[str, str], color: Optional[str] = None) -> None:
        self._state_greetings_preview = dict(greetings)
        lines = [f"{k}: {greetings.get(k, '')}" for k in AI_STATE_KEYS]
        self._state_preview.setText("\n".join(lines))
        self._state_preview.setVisible(True)
        self._set_state_buttons_visible(True)
        self._state_preview.setStyleSheet(
            f"color: {color or S('text')}; font-size: 9px; padding: 2px 0;"
        )

    def clear_state_greetings_preview(self) -> None:
        self._state_greetings_preview = None
        self._state_preview.setVisible(False)
        self._state_preview.setText("")
        self._set_state_buttons_visible(False)

    def _browse_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Load Persona Template", str(SCRIPT_DIR / "personas"),
            "Persona Templates (*.txt);;All Files (*)"
        )
        if path:
            self._radio_file.setChecked(True)
            self._load_file(path)

    def _load_file(self, path: str) -> None:
        file_path = Path(path)
        persona, errors = parse_persona_template(file_path)

        if errors:
            for err in errors:
                print(f"[PERSONA] {err}")

        if persona is None:
            self._preview.setText(
                f"✗ Failed to load: {errors[0] if errors else 'Unknown error'}"
            )
            self._preview.setStyleSheet(
                f"color: {S('red')}; font-size: 9px; font-style: italic; padding: 4px 0;"
            )
            return

        self._loaded_persona = persona
        self._loaded_name    = persona["_loaded_name"]
        self._file_path.setText(str(file_path))

        name    = persona["_loaded_name"]
        desc    = persona.get("description", "")
        colors  = persona.get("colors", {})
        primary = colors.get("primary", "?")
        warn_str = f"  ⚠ {len(errors)} warning(s)" if errors else ""

        self._preview.setText(
            f"✓ Loaded: {name}\n{desc}\nPrimary: {primary}{warn_str}"
        )
        self._preview.setStyleSheet(
            f"color: {S('green')}; font-size: 9px; padding: 4px 0;"
        )

        self.persona_changed.emit(name, persona)

    def _export_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Persona Template",
            str(SCRIPT_DIR / "personas" / "echo_deck_persona_template.txt"),
            "Text Files (*.txt)"
        )
        if path:
            export_persona_template(Path(path))
            QMessageBox.information(
                self, "Template Exported",
                f"Blank template saved to:\n{path}\n\n"
                "Fill in each section, then load it here."
            )

    def get_selection(self) -> tuple[Optional[str], Optional[dict]]:
        """Returns (name, persona_dict) or (None, None) if nothing loaded."""
        radio_id = self._radio_group.checkedId()
        if radio_id == 3:
            # File-loaded
            return self._loaded_name, self._loaded_persona
        name = self._builtin_name_by_id.get(radio_id, "Default")
        persona = dict(BUILTIN_PERSONAS[name])
        persona["pronouns"] = self._current_builtin_pronouns()
        return name, persona

class ModuleChecklist(QWidget):
    """
    Scrollable module selection panel.
    Groups modules by category.
    Shows status badge (BUILT / PARTIAL / PLANNED).
    Only BUILT and PARTIAL modules are checkable.
    PLANNED modules show as greyed out with status badge.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._checkboxes: dict[str, QCheckBox] = {}
        self._setup_ui()

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(4)

        root.addWidget(section_label("Modules"))

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        container = QWidget()
        c_layout  = QVBoxLayout(container)
        c_layout.setContentsMargins(4, 4, 4, 4)
        c_layout.setSpacing(2)

        # Group by category
        categories: dict[str, list[str]] = {}
        for mod_key, mod in MODULES.items():
            cat = mod.get("category", "Other")
            categories.setdefault(cat, []).append(mod_key)

        STATUS_COLORS = {
            "built":   S("green"),
            "partial": S("orange"),
            "planned": S("text_dim"),
        }
        STATUS_LABELS = {
            "built":   "BUILT",
            "partial": "PARTIAL",
            "planned": "PLANNED",
        }

        for category, mod_keys in sorted(categories.items()):
            # Category header
            cat_lbl = QLabel(f"  {category.upper()}")
            cat_lbl.setStyleSheet(
                f"color: {S('primary')}; font-size: 9px; font-weight: bold; "
                f"letter-spacing: 2px; padding-top: 6px;"
            )
            c_layout.addWidget(cat_lbl)

            for mod_key in mod_keys:
                mod    = MODULES[mod_key]
                status = mod.get("status", "planned")
                is_usable = status in ("built", "partial")

                row = QHBoxLayout()
                row.setContentsMargins(8, 0, 0, 0)
                row.setSpacing(6)

                cb = QCheckBox(mod["display_name"])
                cb.setChecked(mod.get("default_on", False) and is_usable)
                cb.setEnabled(is_usable)
                cb.setToolTip(mod.get("description", ""))

                if not is_usable:
                    cb.setStyleSheet(f"color: {S('text_dim')};")

                self._checkboxes[mod_key] = cb

                # Status badge
                badge_color = STATUS_COLORS.get(status, S("text_dim"))
                badge_text  = STATUS_LABELS.get(status, status.upper())
                badge = QLabel(badge_text)
                badge.setStyleSheet(
                    f"color: {badge_color}; font-size: 8px; font-weight: bold; "
                    f"letter-spacing: 1px; padding: 1px 4px; "
                    f"border: 1px solid {badge_color}; border-radius: 2px;"
                )
                badge.setFixedWidth(56)
                badge.setAlignment(Qt.AlignmentFlag.AlignCenter)

                row.addWidget(cb, 1)
                row.addWidget(badge)
                c_layout.addLayout(row)

        c_layout.addStretch()
        scroll.setWidget(container)
        root.addWidget(scroll, 1)

    def get_selected(self) -> list[str]:
        """Return list of selected module keys."""
        return [k for k, cb in self._checkboxes.items() if cb.isChecked()]

    def set_defaults_for_persona(self, persona_name: str) -> None:
        """Adjust default module selections based on persona."""
        pass


# ── FACE PACK PANEL ───────────────────────────────────────────────────────────
class FacePackPanel(QWidget):
    """
    Face pack selection — supports ZIP file or folder.
    Shows validation status and image count.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._source_path: str = ""
        self._setup_ui()

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(4)

        root.addWidget(section_label("Face Pack  (optional)"))

        # Browse row
        browse_row = QHBoxLayout()
        self._path_field = QLineEdit()
        self._path_field.setPlaceholderText("Select ZIP file or folder containing face images")
        self._path_field.textChanged.connect(self._on_path_changed)

        self._btn_zip    = QPushButton("Browse ZIP")
        self._btn_folder = QPushButton("Browse Folder")
        self._btn_zip.clicked.connect(self._browse_zip)
        self._btn_folder.clicked.connect(self._browse_folder)
        self._btn_zip.setFixedWidth(90)
        self._btn_folder.setFixedWidth(100)

        browse_row.addWidget(self._path_field)
        browse_row.addWidget(self._btn_zip)
        browse_row.addWidget(self._btn_folder)
        root.addLayout(browse_row)

        # Status
        self._status_lbl = QLabel("No face pack selected — deck will use placeholder faces.")
        self._status_lbl.setStyleSheet(
            f"color: {S('text_dim')}; font-size: 10px;"
        )
        root.addWidget(self._status_lbl)

        # Icon note
        icon_note = QLabel(
            "Tip: include a file named [DeckName].png in the pack "
            "and it will be used as the desktop shortcut icon."
        )
        icon_note.setWordWrap(True)
        icon_note.setStyleSheet(
            f"color: {S('text_dim')}; font-size: 9px; font-style: italic;"
        )
        root.addWidget(icon_note)

    def _browse_zip(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Face Pack ZIP",
            str(SCRIPT_DIR),
            "ZIP Files (*.zip);;All Files (*)"
        )
        if path:
            self._path_field.setText(path)

    def _browse_folder(self) -> None:
        path = QFileDialog.getExistingDirectory(
            self, "Select Face Pack Folder",
            str(SCRIPT_DIR)
        )
        if path:
            self._path_field.setText(path)

    def _on_path_changed(self, text: str) -> None:
        self._source_path = text.strip()
        if not self._source_path:
            self._status_lbl.setText(
                "No face pack selected — deck will use placeholder faces."
            )
            self._status_lbl.setStyleSheet(
                f"color: {S('text_dim')}; font-size: 10px;"
            )
            return

        valid, msg, count = validate_face_source(self._source_path)
        color = S("green") if valid else S("red")
        self._status_lbl.setText(
            f"{'✓' if valid else '✗'} {msg}"
        )
        self._status_lbl.setStyleSheet(
            f"color: {color}; font-size: 10px;"
        )

    @property
    def source_path(self) -> str:
        return self._source_path

    @property
    def is_valid(self) -> bool:
        if not self._source_path:
            return True  # Empty is OK — optional
        valid, _, _ = validate_face_source(self._source_path)
        return valid


# ── CONNECTION PANEL ──────────────────────────────────────────────────────────
class ConnectionPanel(QWidget):
    """
    Model runner (backend) selection panel.
    Connection type dropdown + model field (Ollama dropdown or text).
    Test Connection button with green/red status.
    """

    connection_tested = Signal(bool)   # True = passed

    def __init__(self, parent=None):
        super().__init__(parent)
        self._test_passed = False
        self._setup_ui()
        QTimer.singleShot(500, self._fetch_ollama_models)

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(6)

        root.addWidget(section_label("Model Runner"))

        # Type row
        type_row = QHBoxLayout()
        type_lbl = QLabel("Runner:")
        type_lbl.setFixedWidth(50)
        self._type_combo = QComboBox()
        self._type_combo.addItems([
            "Ollama (local service — recommended)",
            "Local model folder (transformers)",
            "Claude API (Anthropic)",
            "OpenAI API",
        ])
        self._type_combo.currentIndexChanged.connect(self._on_type_change)
        type_row.addWidget(type_lbl)
        type_row.addWidget(self._type_combo, 1)
        root.addLayout(type_row)

        # Model row (changes based on type)
        model_row = QHBoxLayout()
        self._model_lbl = QLabel("Model:")
        self._model_lbl.setFixedWidth(50)

        # Ollama: combo with refresh
        self._ollama_combo = QComboBox()
        self._ollama_combo.setEditable(True)
        self._ollama_combo.lineEdit().setPlaceholderText("dolphin-mistral")
        self._ollama_refresh = QPushButton("Sync")
        self._ollama_refresh.setFixedWidth(56)
        self._ollama_refresh.setToolTip("Refresh Ollama model list")
        self._ollama_refresh.clicked.connect(self._fetch_ollama_models)

        # Other: text field
        self._model_field = QLineEdit()
        self._model_field.setPlaceholderText("Model path or API model name")

        model_row.addWidget(self._model_lbl)
        model_row.addWidget(self._ollama_combo, 1)
        model_row.addWidget(self._ollama_refresh)
        model_row.addWidget(self._model_field, 1)
        root.addLayout(model_row)

        self._model_field.setVisible(False)
        self._ollama_refresh.setVisible(True)

        # API key (shown for Claude/OpenAI)
        self._key_lbl = QLabel("API Key:")
        self._key_lbl.setFixedWidth(50)
        self._key_field = QLineEdit()
        self._key_field.setEchoMode(QLineEdit.EchoMode.Password)
        self._key_field.setPlaceholderText("sk-...")
        key_row = QHBoxLayout()
        key_row.addWidget(self._key_lbl)
        key_row.addWidget(self._key_field, 1)
        root.addLayout(key_row)
        self._key_lbl.setVisible(False)
        self._key_field.setVisible(False)

        # Test row
        test_row = QHBoxLayout()
        self._btn_test    = QPushButton("Test Connection")
        self._test_status = QLabel("")
        self._test_status.setStyleSheet(f"font-size: 11px;")
        self._btn_test.clicked.connect(self._test)
        test_row.addWidget(self._btn_test)
        test_row.addWidget(self._test_status, 1)
        root.addLayout(test_row)

    def _on_type_change(self, idx: int) -> None:
        is_ollama = (idx == 0)
        is_local  = (idx == 1)
        is_api    = (idx >= 2)

        self._ollama_combo.setVisible(is_ollama)
        self._ollama_refresh.setVisible(is_ollama)
        self._model_field.setVisible(not is_ollama)
        self._key_lbl.setVisible(is_api)
        self._key_field.setVisible(is_api)

        if is_local:
            self._model_field.setPlaceholderText(r"D:\AI\Models\dolphin-8b")
        elif is_api:
            self._model_field.setPlaceholderText(
                "claude-sonnet-4-6" if idx == 2 else "gpt-4o"
            )

        self._test_passed = False
        self._test_status.setText("")
        self.connection_tested.emit(False)

    def _fetch_ollama_models(self) -> None:
        self._ollama_combo.lineEdit().setPlaceholderText("Checking Ollama...")

        class _Fetcher(QThread):
            done = Signal(list, str)
            def run(self):
                try:
                    req  = urllib.request.Request("http://localhost:11434/api/tags")
                    resp = urllib.request.urlopen(req, timeout=4)
                    data = json.loads(resp.read().decode())
                    models = [m["name"] for m in data.get("models", [])]
                    self.done.emit(models, "")
                except Exception as e:
                    self.done.emit([], str(e))

        self._fetcher = _Fetcher()
        self._fetcher.done.connect(self._on_models_fetched)
        self._fetcher.start()

    def _on_models_fetched(self, models: list, error: str) -> None:
        if models:
            current = self._ollama_combo.currentText()
            self._ollama_combo.clear()
            self._ollama_combo.addItems(models)
            if current in models:
                self._ollama_combo.setCurrentText(current)
            self._ollama_combo.lineEdit().setPlaceholderText("Select model")
        else:
            self._ollama_combo.lineEdit().setPlaceholderText(
                "Ollama not running — type model name"
            )

    def _test(self) -> None:
        self._test_status.setText("Testing...")
        self._test_status.setStyleSheet(
            f"color: {S('text_dim')}; font-size: 11px;"
        )
        QApplication.processEvents()

        idx = self._type_combo.currentIndex()
        ok  = False
        msg = ""

        if idx == 0:  # Ollama
            model = self._ollama_combo.currentText().strip()
            if not model:
                msg = "Enter or select a model name."
            else:
                try:
                    req  = urllib.request.Request("http://localhost:11434/api/tags")
                    resp = urllib.request.urlopen(req, timeout=4)
                    data = json.loads(resp.read().decode())
                    models = [m["name"] for m in data.get("models", [])]
                    if model in models:
                        ok  = True
                        msg = f"✓ Ollama running — {model} is available"
                    elif models:
                        ok  = True
                        msg = f"⚠ Ollama running — '{model}' not in list, proceeding"
                    else:
                        msg = "Ollama running but no models found."
                except Exception as e:
                    msg = f"✗ Ollama not reachable: {e}"

        elif idx == 1:  # Local folder
            path = self._model_field.text().strip()
            if path and Path(path).exists():
                ok  = True
                msg = "✓ Model folder found"
            else:
                msg = "✗ Folder not found. Check the path."

        elif idx == 2:  # Claude
            key = self._key_field.text().strip()
            ok  = bool(key and key.startswith("sk-ant"))
            msg = "✓ API key format valid." if ok else "✗ Enter a Claude API key (sk-ant-...)"

        elif idx == 3:  # OpenAI
            key = self._key_field.text().strip()
            ok  = bool(key and key.startswith("sk-"))
            msg = "✓ API key format valid." if ok else "✗ Enter an OpenAI API key (sk-...)"

        self._test_passed = ok
        color = S("green") if ok else S("red")
        self._test_status.setText(msg)
        self._test_status.setStyleSheet(
            f"color: {color}; font-size: 11px;"
        )
        self.connection_tested.emit(ok)

    def get_model_config(self) -> dict:
        idx   = self._type_combo.currentIndex()
        types = ["ollama", "local", "claude", "openai"]
        return {
            "type":         types[idx],
            "path":         self._model_field.text().strip() if idx == 1 else "",
            "ollama_model": self._ollama_combo.currentText().strip() if idx == 0 else "",
            "api_key":      self._key_field.text().strip() if idx >= 2 else "",
            "api_type":     types[idx] if idx >= 2 else "",
            "api_model":    self._model_field.text().strip() if idx >= 2 else "",
        }

    @property
    def test_passed(self) -> bool:
        return self._test_passed

# ── END OF SECTION 7 ──────────────────────────────────────────────────────────
# Next: Section 8 — Main dialog and progress view


# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — MAIN BUILDER DIALOG
# ═══════════════════════════════════════════════════════════════════════════════


class DeckBuilderDialog(QDialog):
    """
    The main deck builder UI.

    Layout:
      Title bar with light/dark toggle
      ├── Deck Name field
      ├── Connection Panel (model runner)
      ├── Persona Panel
      ├── Face Pack Panel
      ├── Module Checklist
      └── Build button row
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Echo Deck Builder")
        self.resize(740, 820)
        self._dark_mode   = True
        self._setup_ui()
        self._apply_style()

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setSpacing(10)
        root.setContentsMargins(16, 12, 16, 12)

        # ── Title bar ──────────────────────────────────────────────────
        title_row = QHBoxLayout()
        title_lbl = QLabel("ECHO DECK BUILDER")
        title_lbl.setStyleSheet(
            "font-size: 16px; font-weight: bold; letter-spacing: 3px;"
        )
        ver_lbl = QLabel(f"v{BUILDER_VERSION}")
        ver_lbl.setStyleSheet("font-size: 10px; color: gray;")

        self._theme_btn = QPushButton("☀ Light Mode")
        self._theme_btn.setFixedWidth(100)
        self._theme_btn.setToolTip("Toggle light/dark mode")
        self._theme_btn.clicked.connect(self._toggle_theme)

        title_row.addWidget(title_lbl)
        title_row.addWidget(ver_lbl)
        title_row.addStretch()
        title_row.addWidget(self._theme_btn)
        root.addLayout(title_row)
        root.addWidget(h_divider())

        # ── Deck Name ──────────────────────────────────────────────────
        root.addWidget(section_label("Deck Name"))
        name_row = QHBoxLayout()
        name_lbl = QLabel("Name:")
        name_lbl.setFixedWidth(50)
        self._name_field = QLineEdit()
        self._name_field.setPlaceholderText(
            "e.g. Thovrik  (used for folder name, window title, file prefix)"
        )
        self._name_field.textChanged.connect(self._on_name_changed)
        name_row.addWidget(name_lbl)
        name_row.addWidget(self._name_field, 1)
        root.addLayout(name_row)

        root.addWidget(h_divider())

        # ── Splitter: left (config) | right (modules) ─────────────────
        splitter = PersistedRightSplitter(self)
        splitter.setHandleWidth(8)

        # Left side
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 8, 0)
        left_layout.setSpacing(10)

        # Connection
        self._conn_panel = ConnectionPanel()
        self._conn_panel.connection_tested.connect(self._on_connection_tested)
        left_layout.addWidget(self._conn_panel)

        left_layout.addWidget(h_divider())

        # Persona
        self._persona_panel = PersonaPanel()
        self._persona_panel.persona_changed.connect(self._on_persona_changed)
        self._persona_panel.generate_state_greetings_requested.connect(self._on_generate_state_greetings)
        self._persona_panel.regenerate_state_greetings_requested.connect(self._on_generate_state_greetings)
        self._persona_panel.use_state_greetings_requested.connect(self._on_use_generated_state_greetings)
        self._persona_panel.use_neutral_state_greetings_requested.connect(self._on_use_neutral_state_greetings)
        left_layout.addWidget(self._persona_panel)

        left_layout.addWidget(h_divider())

        # Face pack
        self._face_panel = FacePackPanel()
        left_layout.addWidget(self._face_panel)

        left_layout.addWidget(h_divider())

        # Shortcut checkbox
        self._shortcut_cb = QCheckBox("Create desktop shortcut")
        self._shortcut_cb.setChecked(True)
        left_layout.addWidget(self._shortcut_cb)

        left_layout.addStretch()

        # Right side — module checklist
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(8, 0, 0, 0)
        self._module_list = ModuleChecklist()
        right_layout.addWidget(self._module_list)

        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setSizes([420, 280])
        QTimer.singleShot(0, splitter.apply_saved_position)
        root.addWidget(splitter, 1)

        root.addWidget(h_divider())

        # ── Build button row ───────────────────────────────────────────
        build_row = QHBoxLayout()
        self._status_bar = QLabel(
            "Enter a deck name and test connection to begin."
        )
        self._status_bar.setStyleSheet("font-size: 10px;")

        self._btn_build  = primary_btn("BUILD DECK")
        self._btn_build.setEnabled(False)
        self._btn_build.setMinimumWidth(140)
        self._btn_build.clicked.connect(self._begin_build)

        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(self.reject)
        btn_cancel.setFixedWidth(80)

        build_row.addWidget(self._status_bar, 1)
        build_row.addWidget(self._btn_build)
        build_row.addWidget(btn_cancel)
        root.addLayout(build_row)

        # Track state
        self._name_ok        = False
        self._connection_ok  = False
        self._current_persona_name: Optional[str] = "Default"
        self._current_persona_name: Optional[str] = None
        self._current_persona: Optional[dict]     = None
        self._generated_state_greetings: Optional[dict[str, str]] = None
        self._approved_state_greetings: Optional[dict[str, str]] = None
        self._build_state_greetings: Optional[dict[str, str]] = None
    def _apply_style(self) -> None:
        self.setStyleSheet(_get_style())

    def _toggle_theme(self) -> None:
        self._dark_mode = not self._dark_mode
        set_scheme("dark" if self._dark_mode else "light")
        self._theme_btn.setText("☀ Light Mode" if self._dark_mode else "☾ Dark Mode")
        self._apply_style()

    def _on_name_changed(self, text: str) -> None:
        self._name_ok = bool(text.strip())
        self._check_ready()

    def _on_connection_tested(self, ok: bool) -> None:
        self._connection_ok = ok
        self._persona_panel.set_generation_enabled(ok)
        self._check_ready()

    def _on_persona_changed(self, name: str, persona: dict) -> None:
        self._current_persona_name = name
        self._current_persona      = persona
        self._generated_state_greetings = None
        self._approved_state_greetings = None
        self._persona_panel.clear_state_greetings_preview()
        # Let module list adjust defaults
        self._module_list.set_defaults_for_persona(name)
        # Auto-fill deck name if blank
        if not self._name_field.text().strip() and name not in ("Default", "Custom"):
            self._name_field.setText(name)

    def _generate_state_greetings_now(self) -> Optional[dict[str, str]]:
        if not self._current_persona:
            return None
        deck_name = self._name_field.text().strip() or (self._current_persona_name or "Deck")
        model_config = self._conn_panel.get_model_config()
        generated = _request_state_greetings(model_config, self._current_persona, deck_name)
        if generated:
            return generated
        return _neutral_state_greetings(deck_name)

    def _on_generate_state_greetings(self) -> None:
        if not self._conn_panel.test_passed:
            return
        greetings = self._generate_state_greetings_now()
        if greetings is None:
            return
        self._generated_state_greetings = greetings
        self._persona_panel.set_state_greetings_preview(greetings, S("text"))

    def _on_use_generated_state_greetings(self) -> None:
        if self._generated_state_greetings:
            self._approved_state_greetings = dict(self._generated_state_greetings)
            self._persona_panel.set_state_greetings_preview(self._approved_state_greetings, S("green"))

    def _on_use_neutral_state_greetings(self) -> None:
        deck_name = self._name_field.text().strip() or (self._current_persona_name or "Deck")
        neutral = _neutral_state_greetings(deck_name)
        self._generated_state_greetings = dict(neutral)
        self._approved_state_greetings = dict(neutral)
        self._persona_panel.set_state_greetings_preview(neutral, S("green"))

    def _check_ready(self) -> None:
        ready = self._name_ok and self._connection_ok
        self._btn_build.setEnabled(ready)
        if ready:
            self._status_bar.setText(
                f"Ready to build '{self._name_field.text().strip()}' deck."
            )
            self._status_bar.setStyleSheet(
                f"font-size: 10px; color: {S('green')};"
            )
        else:
            msgs = []
            if not self._name_ok:
                msgs.append("deck name")
            if not self._connection_ok:
                msgs.append("connection test")
            self._status_bar.setText(
                f"Still needed: {', '.join(msgs)}."
            )
            self._status_bar.setStyleSheet(
                f"font-size: 10px; color: {S('text_dim')};"
            )

    def _begin_build(self) -> None:
        deck_name = self._name_field.text().strip()

        if not self._current_persona:
            QMessageBox.warning(self, "No Persona",
                                "Select or load a persona before building.")
            return

        # Validate face pack
        if not self._face_panel.is_valid:
            QMessageBox.warning(self, "Invalid Face Pack",
                                "The face pack source is invalid. "
                                "Fix it or clear it to skip face installation.")
            return

        # Check for existing deck
        output_base  = SCRIPT_DIR
        deck_home    = output_base / deck_name
        if deck_home.exists() and (deck_home / "config.json").exists():
            reply = QMessageBox.question(
                self, "Deck Already Exists",
                f"A deck named '{deck_name}' already exists at:\n{deck_home}\n\n"
                f"Rebuilding will overwrite config, sounds, and deck file "
                f"but preserve memories and face images.\n\nContinue?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        build_state_greetings = None
        if bool(self._current_persona.get("vampire_states", False)):
            if self._approved_state_greetings:
                build_state_greetings = dict(self._approved_state_greetings)
            else:
                generated = _request_state_greetings(
                    self._conn_panel.get_model_config(), self._current_persona, deck_name
                )
                build_state_greetings = generated or _neutral_state_greetings(deck_name)
        self._build_state_greetings = build_state_greetings

        # Switch to progress view
        self._show_progress(deck_name)

    def _show_progress(self, deck_name: str) -> None:
        """Replace dialog content with the progress/log view."""
        # Clear layout
        layout = self.layout()
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        self.resize(680, 500)
        root = layout

        # Header
        hdr = QLabel(f"Building: {deck_name}")
        hdr.setStyleSheet(
            f"font-size: 14px; font-weight: bold; color: {S('primary')};"
        )
        hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
        root.addWidget(hdr)

        # Progress bar
        self._progress = QProgressBar()
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        root.addWidget(self._progress)

        # Log display
        self._log_view = QTextEdit()
        self._log_view.setReadOnly(True)
        root.addWidget(self._log_view, 1)

        # Done button
        self._done_btn = primary_btn("Close")
        self._done_btn.setEnabled(False)
        self._done_btn.clicked.connect(self.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(self._done_btn)
        root.addLayout(btn_row)

        # Start worker
        self._worker = DeckSetupWorker(
            deck_name        = deck_name,
            persona          = self._current_persona,
            ai_state_greetings = self._build_state_greetings,
            model_config     = self._conn_panel.get_model_config(),
            selected_modules = self._module_list.get_selected(),
            face_source      = self._face_panel.source_path,
            create_shortcut  = self._shortcut_cb.isChecked(),
            output_base      = SCRIPT_DIR,
        )
        self._worker.log.connect(self._on_log)
        self._worker.progress.connect(self._progress.setValue)
        self._worker.done.connect(self._on_done)
        self._worker.start()

    def _on_log(self, msg: str) -> None:
        self._log_view.append(msg)
        self._log_view.verticalScrollBar().setValue(
            self._log_view.verticalScrollBar().maximum()
        )

    def _on_done(self, success: bool, message: str) -> None:
        self._done_btn.setEnabled(True)
        if success:
            self._done_btn.setText("✓ Done — Close Builder")
            deck_path = Path(message)
            self._log_view.append(
                f"\nDeck ready. Launch with:\n  python {deck_path}\n"
                f"or use the desktop shortcut."
            )
            # Offer to launch immediately
            reply = QMessageBox.question(
                self, "Build Complete",
                f"Deck built successfully!\n\nLaunch '{deck_path.parent.name}' now?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                _launch_deck(deck_path)
        else:
            self._done_btn.setText("Close")
            self._log_view.append(f"\n✗ Build failed: {message}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 9 — ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def _launch_deck(deck_path: Path) -> None:
    """Launch a generated deck file using pythonw (no console window)."""
    try:
        pythonw = Path(sys.executable)
        if pythonw.name.lower() == "python.exe":
            candidate = pythonw.parent / "pythonw.exe"
            if candidate.exists():
                pythonw = candidate
        subprocess.Popen(
            [str(pythonw), str(deck_path)],
            cwd=str(deck_path.parent)
        )
    except Exception as e:
        print(f"Could not launch deck: {e}")
        print(f"Run manually: python {deck_path}")


def main() -> None:
    # Ensure we're running from the right directory
    os.chdir(str(SCRIPT_DIR))

    app = QApplication(sys.argv)
    app.setApplicationName("Echo Deck Builder")

    dlg = DeckBuilderDialog()
    dlg.exec()

    sys.exit(0)


if __name__ == "__main__":
    main()
